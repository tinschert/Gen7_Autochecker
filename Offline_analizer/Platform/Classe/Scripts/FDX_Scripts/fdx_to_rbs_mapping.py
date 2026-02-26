# -*- coding: utf-8 -*-
# @file fdx_to_rbs_mapping.py
# @author ADAS_HIL_TEAM
# @date 08-21-2023

##################################################################
# C O P Y R I G H T S
# ----------------------------------------------------------------
# Copyright (c) 2023 by Robert Bosch GmbH. All rights reserved.

# The reproduction, distribution and utilization of this file as
# well as the communication of its contents to others without express
# authorization is prohibited. Offenders will be held liable for the
# payment of damages. All rights reserved in the event of the grant
# of a patent, utility model or design.

##################################################################


from openpyxl import load_workbook
import pandas as pd
try:
    pd.set_option('future.no_silent_downcasting', True)
except:
    pass
import numpy as np
import re
import os, sys
from collections import Counter
try:
    from Control.logging_config import logger
except ImportError:
    sys.path.append(os.path.dirname(os.path.abspath(__file__)) + r"\..\Control")
    from logging_config import logger
sys.path.append(os.path.dirname(os.path.abspath(__file__)) + r"\..\Rbs_Scripts")
from create_sysvar import generate_sysvar, save_to_file

#allow longer names up to 100 characters
pd.options.display.max_colwidth = 100
pd.options.display.colheader_justify = 'left'

column_names = ["Name", "Message", "Value Type", "Initial Value", "Factor", "Offset", "Minimum",  "Maximum", "Unit", "texttable", "texttable values", "max_value", "variant", "sender", "direction", "texttable mapping", "conversion", "array name", "data_type"]

node_head = '''/**
 * @file FDX_map.can
 * @author Rafael Herrera
 * @date 12.07.2021
 * @brief handles the conversion from FDX received variables to restbus variables
 */
'''

convert_formula_in = '''/**
 * @fn: convert_signal()
 * @brief: Calculates the conversion from input a to output b
 */
double convert_signal(double sig_a, double sig_a_offset, double sig_a_factor, double sig_b_offset, double sig_b_factor, double sig_b_conversion)
{
   
  double sig_a_phys;
  double sig_b_raw_offset,sig_b_phys_offset;
  double sig_b;
  
  sig_a_phys = ((((sig_a * sig_a_factor) + (sig_a_offset * sig_a_factor)) * sig_b_conversion));
  sig_b_raw_offset = (sig_b_offset/sig_b_factor);
  sig_b = sig_a_phys;// - sig_b_raw_offset;
  return sig_b;
   
}
'''

convert_formula_out = '''/**
 * @fn: convert_signal()
 * @brief: Calculates the conversion from input a to output b
 */
double convert_signal(double sig_a, double sig_a_offset, double sig_a_factor, double sig_b_offset, double sig_b_factor, double sig_b_conversion)
{
  
  double sig_a_phys;
  double sig_b_raw_offset;
  double sig_b;
  
  sig_a_phys = ((((sig_a * sig_a_factor) + (sig_a_offset)) * sig_b_conversion) / sig_b_factor);
  sig_b_raw_offset = (sig_b_offset/sig_b_factor);
  sig_b = sig_a_phys - sig_b_raw_offset;
  return sig_b;
}
'''

# read worksheet 'sheet_name' from workbook 'wb'
def generate_dataframe(df_FDX,sheet_patt):
    """
    generate dataframe for sheet and put into global dataframe

    Args:
      df_FDX(dataframe): dataframe of excel sheet
      sheet_patt(string): for which sheet dataframe to generate

    Returns:

    """
    global node, node_df, num_of_rows, sheet
    global signal_cloe
    global col_name, col_namespace, col_initial_value, col_factor, col_offset, col_minimum, col_maximum, col_unit, col_texttable, col_texttable_values, col_max_value, col_variant, col_sender, col_direction, col_texttable_mapping, col_conversion, col_array_name, col_index
    sheet = sheet_patt
    
    node = ['/*@!Encoding:1252*/']
    node.append(node_head)
    #ws1 = wb[sheet_name]
    #ws2 = wb_classe[sheet_name]
    #print(node_name, bus_name, bus_type)

    # create the different dataframes
    #node_fdx = pd.DataFrame(ws1.values)
    #node_classe = pd.DataFrame(ws2.values)
    #node_classe = node_classe.drop(labels=0, axis=0)
    #node_df = pd.concat([node_fdx, node_classe])
    df_FDX.loc[-1] = list(df_FDX.columns)  # adding a row
    df_FDX.index = df_FDX.index + 1  # shifting index
    df_FDX = df_FDX.sort_index()
    df_FDX = df_FDX.replace(np.nan, '', regex=True)
    node_df = df_FDX.rename(columns={x: y for x, y in zip(df_FDX.columns, range(0, len(df_FDX.columns)))})
    #node_df = node_df.replace(np.nan, '', regex=True)

    # get column names
    col_names = list(node_df.iloc[0])
    # get column index
    try:
        col_index = [col_names.index(col) for col in column_names]
    except Exception as e:
        logger.error(f"{e} -> Column names in Excel are different from the Script:")
        logger.error(';'.join(column_names))
        return False
    [col_name, col_namespace, col_value_type, col_initial_value, col_factor, col_offset, col_minimum, col_maximum, col_unit, col_texttable, col_texttable_values, col_max_value, col_variant, col_sender, col_direction, col_texttable_mapping, col_conversion, col_array_name, col_data_type] = col_index

    signal_cloe.extend(node_df[node_df[col_sender] == "Cloe"][col_name].tolist())
    num_of_rows = node_df.shape[0]
    logger.info(f"Signal count: {num_of_rows-1}")

def generate_includes():
    """ will add include section in .can file with adding into node list"""
    # Case create node includes
    node.append('includes {')
    if (sheet == "fdx_in"):
        node.append('  #include "fdx_in_cus.cin"')
    node.append('}')

def generate_variables():
    """will add variable section in .can file with adding into node list """
    # Case create node variables
    node.append('variables {')
    node.append('}')

def generate_onstart():
    """ will add onstart section in .can file with adding into node list """
    node.append('')
    node.append('on start {')
    node.append('}')

def generate_onstop():
    """will add onstop section in .can file with adding into node list """
    node.append('')
    node.append('on stopMeasurement {')
    node.append('}')

# def generate_doublearray():
#     global sysvar_df
#     message_count = 0
#     array_df = node_df
#     array_df = array_df[(array_df[col_array_name] != "")]
#     num_of_rows_array = array_df.shape[0]
#     for i in range(num_of_rows_array):
#         if (i > 0)  and (array_df[col_array_name].values[i-1] != array_df[col_array_name].values[i]):
#             array_name = array_df[col_array_name].values[i]
#             signal_array_df = array_df
#             signal_array_df = signal_array_df[(signal_array_df[col_array_name] == array_name)]
#             num_of_rows_signal_array = signal_array_df.shape[0]
#             node.append('')   
#             node.append('on sysvar_update {0}::{0}'.format(array_name))
#             node.append('{')
#             for j in range(num_of_rows_signal_array):
#                 sender = signal_array_df[col_sender].values[j]
#                 if (j == 0) and (sender == "Cloe"):
#                     # generate sysvar
#                     add_sysvar_df = signal_array_df[col_index].iloc[[row for row in range(j, j+num_of_rows_signal_array, 2)]]
#                     add_sysvar_df.columns = [i for i in range(len(col_index))]
#                     sysvar_df = pd.concat([sysvar_df, add_sysvar_df])
#                     if (j == num_of_rows_signal_array-1) or (signal_array_df[col_sender].values[j+1] == "Cloe"):
#                         continue
#                     namespace = signal_array_df[col_namespace].values[j]
#                     signal_name = signal_array_df[col_name].values[j]
#                     direction = signal_array_df[col_direction].values[j]
#                     texttable = signal_array_df[col_texttable].values[j]
#                     sig_a_offset = signal_array_df[col_offset].values[j]
#                     sig_a_factor = signal_array_df[col_factor].values[j]
#                     sig_b_offset = signal_array_df[col_offset].values[j+1]
#                     sig_b_factor = signal_array_df[col_factor].values[j+1]
#                     sig_b_conversion = signal_array_df[col_conversion].values[j]
#                     sig_a_min = signal_array_df[col_minimum].values[j]
#                     sig_b_min = signal_array_df[col_minimum].values[j+1]
#                     array_length = int(num_of_rows_signal_array/2)
#                     node.append('')  
#                     node.append('  int i = 0;')
#                     node.append('  int sig_a_array_len = {0};'.format(array_length))
#                     node.append('  double sig_a_offset = {0};'.format(sig_a_offset))
#                     node.append('  double sig_a_factor = {0};'.format(sig_a_factor))
#                     node.append('  double sig_b_offset = {0};'.format(sig_b_offset))
#                     node.append('  double sig_b_factor = {0};'.format(sig_b_factor))
#                     node.append('  double sig_b_conversion = {0};'.format(sig_b_conversion))
#                     for k in range (array_length):
#                         node.append('  double sig_a_{0};'.format(k))
#                     node.append('  double sig_a_min = {0};'.format(sig_a_min))
#                     node.append('  double sig_b_min = {0};'.format(sig_b_min))
#                     node.append('')
#                     node.append('  for (i=0; i<sig_a_array_len;i++)')
#                     node.append('  {')
#                     for k in range (array_length):
#                         node.append('    sig_a_{0} = @this[{0}];'.format(k))
#                     node.append('  }')
#                     node.append('')
#                     node.append('  if ((sig_a_min < 0) && (sig_b_min < 0))')
#                     node.append('  {')
#                     for k in range (num_of_rows_signal_array-1):
#                         if (signal_array_df[col_direction].values[k] == "source"):
#                             sig_a_name = signal_array_df[col_name].values[k]
#                             sig_a_id = int(re.search(r'\d+', sig_a_name).group())
#                             sig_b_name = signal_array_df[col_name].values[k+1]
#                             sig_b_namespace = signal_array_df[col_namespace].values[k+1]
#                             node.append('    @{0}::{1}_Rv = convert_signal(sig_a_{2}, sig_a_offset, sig_a_factor, sig_b_offset, sig_b_factor, sig_b_conversion);'.format(sig_b_namespace, sig_b_name , sig_a_id))
#                     node.append('  }')
#                     node.append('  else if ((sig_a_min < 0) && (sig_b_min >= 0))')
#                     node.append('  {')
#                     for k in range (num_of_rows_signal_array-1):
#                         if (signal_array_df[col_direction].values[k] == "source"):
#                             sig_a_name = signal_array_df[col_name].values[k]
#                             sig_a_id = int(re.search(r'\d+', sig_a_name).group())
#                             sig_b_name = signal_array_df[col_name].values[k+1]
#                             sig_b_namespace = signal_array_df[col_namespace].values[k+1]
#                             node.append('    @{0}::{1}_Rv = abs(convert_signal(sig_a_{2}, sig_a_offset, sig_a_factor, sig_b_offset, sig_b_factor, sig_b_conversion));'.format(sig_b_namespace, sig_b_name , sig_a_id))
#                     node.append('  }')
#                     node.append('}')
#                     message_count += 1
#     print("\tdouble array: generated with " + str(message_count) + " messages.")

def generate_onsysvarenum():
    """for each sheet it will add into global node data frame  as per sheet and inside data will put data into
    list for generating can file """
    global sysvar_df
    message_count = 0
    enum_df = node_df
    enum_df = enum_df[(enum_df[col_texttable] != "")]
    num_of_rows_enum = enum_df.shape[0]
    variants_row = set(enum_df[col_variant].values)
    to_delete = {'', "multi", "variant"}
    variants_row.difference_update(to_delete)

    for i in range(num_of_rows_enum):
        if (enum_df[col_sender].values[i] == "Cloe"):
            # generate sysvar
            add_sysvar_df = enum_df[col_index].iloc[[i]]
            add_sysvar_df.columns = [i for i in range(len(col_index))]
            sysvar_df = pd.concat([sysvar_df, add_sysvar_df])
            if (i == num_of_rows_enum-1) or (enum_df[col_sender].values[i+1] == "Cloe") or (enum_df[col_name].values[i+1] == "na"):
                continue

            if (sheet == "fdx_in"):
                namespace = enum_df[col_namespace].values[i]
                signal_name =  enum_df[col_name].values[i]
                namespace_b = []
                signal_name_b = []
                initial_value_b = []
                enum_mapping = []
                if enum_df[col_variant].values[i] == "multi":
                    idx = 1
                    while "variant" in enum_df[col_variant].values[i+idx]:
                        namespace_b.append(enum_df[col_namespace].values[i+idx])
                        signal_name_b.append(enum_df[col_name].values[i+idx])
                        init_val = enum_df[col_initial_value].values[i + idx]
                        if init_val is not None:
                            initial_value_b.append(init_val)
                        else:
                            logger.error(f"Missing initial value for {enum_df[col_name].values[i + idx]} signal")
                            raise Exception("Please check the database")
                        mapping = enum_df[col_texttable_mapping].values[i + idx]
                        if mapping != "":
                            enum_mapping.append(mapping)
                        else:
                            logger.error(f"Missing texttable mapping for {enum_df[col_name].values[i + idx]} signal")
                            raise Exception("Please check the database")
                        if i+idx == num_of_rows_enum:
                            break
                        idx+=1
                    combined_data =  list(zip(namespace_b, signal_name_b, initial_value_b, enum_mapping))
                else:
                    idx = 1
                    while "classe" in enum_df[col_sender].values[i+idx] or "Rbs" in enum_df[col_sender].values[i+idx] or "XCP" in enum_df[col_sender].values[i+idx]:
                        namespace_b.append(enum_df[col_namespace].values[i+idx])
                        signal_name_b.append(enum_df[col_name].values[i+idx])
                        init_val = enum_df[col_initial_value].values[i+idx]
                        if init_val != "":
                            initial_value_b.append(init_val)
                        else:
                            logger.error(f"Missing initial value for {enum_df[col_name].values[i+idx]} signal")
                            raise Exception("Please check the database")
                        mapping = enum_df[col_texttable_mapping].values[i+idx]
                        if mapping != "":
                            enum_mapping.append(mapping)
                        else:
                            logger.error(f"Missing texttable mapping for {enum_df[col_name].values[i+idx]} signal")
                            raise Exception("Please check the database")
                        idx+=1
                        if i+idx == num_of_rows_enum:
                            break
                    combined_data_common = list(zip(namespace_b, signal_name_b, initial_value_b, enum_mapping))
 
            elif (sheet == "fdx_out"):
                namespace_b = []
                signal_name_b = []
                initial_value_b = []
                enum_mapping = []
                namespace = enum_df[col_namespace].values[i+1]
                signal_name =  enum_df[col_name].values[i+1]
                idx = 1
                while "classe" in enum_df[col_sender].values[i+idx] or "Rbs" in enum_df[col_sender].values[i+idx] or "XCP" in enum_df[col_sender].values[i+idx]:
                    namespace_b.append(enum_df[col_namespace].values[i])
                    signal_name_b.append(enum_df[col_name].values[i])
                    init_val = enum_df[col_initial_value].values[i]
                    if init_val != "":
                        initial_value_b.append(init_val)
                    else:
                        logger.error(f"Missing initial value for {enum_df[col_name].values[i]} signal")
                        raise Exception("Please check the database")
                    mapping = enum_df[col_texttable_mapping].values[i+1]
                    if mapping != "":
                        enum_mapping.append(mapping)
                    else:
                        logger.error(f"Missing texttable mapping for {enum_df[col_name].values[i]} signal")
                        raise Exception("Please check the database")
                    idx+=1
                    if i+idx == num_of_rows_enum:
                        break
                combined_data_common = list(zip(namespace_b, signal_name_b, initial_value_b, enum_mapping))


            sender = enum_df[col_sender].values[i+1]
            node.append('')  
            if (sheet == "fdx_in"):
                node.append(f'on sysvar_update {namespace}::{signal_name}')
            elif (sheet == "fdx_out" and sender != "classe"):
                node.append(f'on signal_update {namespace}::{signal_name}')
            else:
                node.append(f'on sysvar_update {namespace}::{signal_name}')
            node.append('{')
            node.append('')
            if (sheet == "fdx_in"):
                node.append(f'  switch (@{namespace}::{signal_name})')
            elif (sheet == "fdx_out" and sender != "classe" and "XCP" not in sender ):
                node.append(f'  switch (${namespace}::{signal_name})')
            else:
                node.append(f'  switch (@{namespace}::{signal_name})')
            node.append('  {')

            if  enum_df[col_variant].values[i] != "multi":
                enum_mapping = enum_mapping[0].split()
                enum_length = len(enum_mapping)
                for k in range (enum_length):
                    enum_member = str(enum_mapping[k])
                    enum_member = enum_member.split(':')
                    sig_a_map = enum_member[0]
                    node.append(f'    case {sig_a_map}:')
                    for variant in range(0,len(combined_data_common)):
                        val_temp = combined_data_common[variant][3].split("\n")[k]
                        val = val_temp[2]
                        if "XCP" in sender:
                            node.append(f'      @{sender}::{combined_data_common[variant][0][:-1]}::{combined_data_common[variant][1]} = {val};')
                        elif (sheet == "fdx_in" and sheet != "fdx_out" and sender != "classe" ):
                            node.append(f'      ${combined_data_common[variant][0]}::{combined_data_common[variant][1]} = {val};')
                        else:
                            node.append(f'      @{combined_data_common[variant][0]}::{combined_data_common[variant][1]} = {val};')
                    node.append('      break;')

                node.append('    default:'.format(sig_a_map))
                for variant in range(0,len(combined_data_common)):
                    if "XCP" in sender:
                        node.append(f'      @{sender}::{combined_data_common[variant][0][:-1]}::{combined_data_common[variant][1]} = {combined_data_common[variant][2]};')
                    elif (sheet == "fdx_in" and sheet != "fdx_out" and sender != "classe" ):
                        node.append(f'      ${combined_data_common[variant][0]}::{combined_data_common[variant][1]} = {combined_data_common[variant][2]};')
                    else:
                        node.append(f'      @{combined_data_common[variant][0]}::{combined_data_common[variant][1]} = {combined_data_common[variant][2]};')
                node.append('      break;')
                node.append('  }')
                node.append('}')
                message_count += 1
            else:
                enum_mapping = enum_mapping[0].split()
                enum_length = len(enum_mapping)
                for k in range (enum_length):
                    enum_member = str(enum_mapping[k])
                    enum_member = enum_member.split(':')
                    sig_a_map = enum_member[0]
                    node.append(f'    case {sig_a_map}:')
                    for variant in range(0,len(combined_data)):
                        val_temp = combined_data[variant][3].split("\n")[k]
                        val = val_temp[2]
                        if (sheet == "fdx_in" and sheet != "fdx_out" and sender != "classe" ):
                            node.append(f'      ${combined_data[variant][0]}::{combined_data[variant][1]} = {val};')
                        else:
                            node.append(f'      @{combined_data[variant][0]}::{combined_data[variant][1]} = {val};')
                    node.append('      break;')

                node.append('    default:'.format(sig_a_map))
                for variant in range(0,len(combined_data)):
                    if "XCP" in sender:
                        node.append(f'      @{sender}::{combined_data[variant][0][:-1]}::{combined_data[variant][1]} = {combined_data[variant][2]};')
                    elif (sheet == "fdx_in" and sheet != "fdx_out" and sender != "classe" ):
                        node.append(f'      ${combined_data[variant][0]}::{combined_data[variant][1]} = {combined_data[variant][2]};')
                    else:
                        node.append(f'      @{combined_data[variant][0]}::{combined_data[variant][1]} = {combined_data[variant][2]};')
                node.append('      break;')
                node.append('  }')
                node.append('}')
                message_count += 1

    logger.info(f"enum int: generated with {str(message_count)} messages.")



def generate_onsysvardouble():
    """for each sheet it will take data check data integrity and add into global node data frame
    for generate can file  """
    global sysvar_df
    message_count = 0
    sysvardouble_df = node_df
    sysvardouble_df = sysvardouble_df[(sysvardouble_df[col_texttable] != "texttable")]
    sysvardouble_df = sysvardouble_df[(sysvardouble_df[col_array_name] == "")]
    num_of_rows_double = sysvardouble_df.shape[0]
    variants_row = set(sysvardouble_df[col_variant].values)
    to_delete = {'', "multi", "variant"}
    variants_row.difference_update(to_delete)
    number_of_variants = len(variants_row)
    variants_row = list(variants_row)
    variants_row.sort()

    for i in range(num_of_rows_double):
        if (sysvardouble_df[col_sender].values[i] == "Cloe"):
            # generate sysvar
            add_sysvar_df = sysvardouble_df[col_index].iloc[[i]]
            add_sysvar_df.columns = [i for i in range(len(col_index))]
            sysvar_df = pd.concat([sysvar_df, add_sysvar_df])
            if (i == num_of_rows_double-1) or (sysvardouble_df[col_sender].values[i+1] == "Cloe") or (sysvardouble_df[col_name].values[i+1] == "na"):
                continue
            if (sheet == "fdx_in"):
                if sysvardouble_df[col_variant].values[i] == "multi":
                    namespace = sysvardouble_df[col_namespace].values[i]
                    signal_name = sysvardouble_df[col_name].values[i]
                    # direction = sysvardouble_df[col_direction].values[i]
                    # texttable = sysvardouble_df[col_texttable].values[i]
                    sig_a_offset = sysvardouble_df[col_offset].values[i]
                    sig_a_factor = sysvardouble_df[col_factor].values[i]
                    sig_a_min = sysvardouble_df[col_minimum].values[i]
                    sig_b_conversion = sysvardouble_df[col_conversion].values[i]
                    sig_b_offset = []
                    sig_b_factor = []
                    sig_b_min = []
                    sig_b_name = []
                    sig_b_namespace = []
                    for variant in range(1,number_of_variants+1):
                        sig_b_offset.append(sysvardouble_df[col_offset].values[i+variant])
                        sig_b_factor.append(sysvardouble_df[col_factor].values[i+variant])
                        sig_b_min.append(sysvardouble_df[col_minimum].values[i+variant])
                        sig_b_name.append(sysvardouble_df[col_name].values[i+variant])
                        sig_b_namespace.append(sysvardouble_df[col_namespace].values[i+variant])
                else:
                    namespace = sysvardouble_df[col_namespace].values[i]
                    signal_name = sysvardouble_df[col_name].values[i]
                    sig_b_name = sysvardouble_df[col_name].values[i + 1]
                    sig_b_namespace = sysvardouble_df[col_namespace].values[i + 1]
                    # direction = sysvardouble_df[col_direction].values[i]
                    # texttable = sysvardouble_df[col_texttable].values[i]
                    sig_a_offset = sysvardouble_df[col_offset].values[i]
                    check_int_double_integrity("offset",sig_a_offset,signal_name)
                    sig_a_factor = sysvardouble_df[col_factor].values[i]
                    check_int_double_integrity("factor", sig_a_factor, signal_name)
                    sig_b_offset = sysvardouble_df[col_offset].values[i+1]
                    check_int_double_integrity("offset", sig_b_offset, sig_b_name)
                    sig_b_factor = sysvardouble_df[col_factor].values[i+1]
                    check_int_double_integrity("factor", sig_b_factor, sig_b_name)
                    sig_b_conversion = sysvardouble_df[col_conversion].values[i]
                    check_int_double_integrity("conversion", sig_b_conversion, sig_b_name)
                    sig_a_min = sysvardouble_df[col_minimum].values[i]
                    check_int_double_integrity("minimum", sig_a_min, signal_name)
                    sig_b_min = sysvardouble_df[col_minimum].values[i+1]
                    check_int_double_integrity("minimum", sig_b_min, sig_b_name)


            elif (sheet == "fdx_out"):
                namespace = sysvardouble_df[col_namespace].values[i+1]
                signal_name = sysvardouble_df[col_name].values[i+1]
                sig_b_name = sysvardouble_df[col_name].values[i]
                sig_b_namespace = sysvardouble_df[col_namespace].values[i]
                # direction = sysvardouble_df[col_direction].values[i+1]
                # texttable = sysvardouble_df[col_texttable].values[i+1]
                sig_a_offset = sysvardouble_df[col_offset].values[i+1]
                check_int_double_integrity("offset", sig_a_offset, signal_name)
                sig_a_factor = sysvardouble_df[col_factor].values[i+1]
                check_int_double_integrity("factor", sig_a_factor, signal_name)
                sig_b_offset = sysvardouble_df[col_offset].values[i]
                check_int_double_integrity("offset", sig_b_offset, sig_b_name)
                sig_b_factor = sysvardouble_df[col_factor].values[i]
                check_int_double_integrity("factor", sig_b_factor, sig_b_name)
                sig_b_conversion = sysvardouble_df[col_conversion].values[i]
                check_int_double_integrity("conversion", sig_b_conversion, sig_b_name)
                sig_a_min = sysvardouble_df[col_minimum].values[i+1]
                check_int_double_integrity("minimum", sig_a_min, signal_name)
                sig_b_min = sysvardouble_df[col_minimum].values[i]
                check_int_double_integrity("minimum", sig_b_min, sig_b_name)


            node.append('')
            sender = sysvardouble_df[col_sender].values[i+1]
            if (sheet == "fdx_in"):
                node.append(f'on sysvar_update {namespace}::{signal_name}')
            elif (sheet == "fdx_out" and sender != "classe"):
                node.append(f'on signal_update {namespace}::{signal_name}')
            else:
                node.append(f'on sysvar_update {namespace}::{signal_name}')
            node.append('{')
            node.append(f'  double sig_a_offset = {sig_a_offset};')
            node.append(f'  double sig_a_factor = {sig_a_factor};')
            node.append(f'  double sig_a;')
            node.append(f'  double sig_a_min = {sig_a_min};')
            if  sysvardouble_df[col_variant].values[i] != "multi":
                node.append(f'  double sig_b_offset = {sig_b_offset};')
                node.append(f'  double sig_b_factor = {sig_b_factor};')
                node.append(f'  double sig_b_conversion = {sig_b_conversion};')
                node.append(f'  double sig_b_min = {sig_b_min};')
                node.append('')
            else:
                node.append(f'  double sig_b_offset;')
                node.append(f'  double sig_b_factor;')
                node.append(f'  double sig_b_conversion;')
                node.append(f'  double sig_b_min;')
                node.append('')
                for variant in range(0,number_of_variants):
                    statement = lambda var : "if" if var == 0 else "else if"
                    node.append(f'  {statement(variant)} (@Vehicle_Model::variant == {variants_row[variant]})')
                    node.append(f'      {{ sig_b_offset = {sig_b_offset[variant]}; sig_b_factor = {sig_b_factor[variant]}; sig_b_conversion = {sig_b_conversion}; sig_b_min = {sig_b_min[variant]}; }} ')
                node.append('')
            if (sheet == "fdx_in"):
                node.append('  sig_a = @this;')
            elif (sheet == "fdx_out" and sender == "classe"):
                node.append('  sig_a = @this;')
            else:
                node.append(f'  sig_a = getRawSignal({namespace}::{signal_name});')
            node.append('')
            node.append('  if ((((sig_a_min < 0) && (sig_b_min < 0))) || ((sig_a_min >= 0) && (sig_b_min >= 0)) )')
            node.append('  {')
            if  sysvardouble_df[col_variant].values[i] != "multi":
                if "XCP" in sender:
                    node.append(f'    @{sender}::{sig_b_namespace[:-1]}::{sig_b_name} = convert_signal(sig_a, sig_a_offset, sig_a_factor, sig_b_offset, sig_b_factor, sig_b_conversion);')
                elif (sheet == "fdx_in" and sender != "classe"):
                    node.append(f'    ${sig_b_namespace}::{sig_b_name} = convert_signal(sig_a, sig_a_offset, sig_a_factor, sig_b_offset, sig_b_factor, sig_b_conversion);')
                elif (sheet == "fdx_out" or sender == 'classe'):
                    node.append(f'    @{sig_b_namespace}::{sig_b_name} = convert_signal(sig_a, sig_a_offset, sig_a_factor, sig_b_offset, sig_b_factor, sig_b_conversion);')
            else:
                for variant in range(0,number_of_variants):
                    if (sheet == "fdx_in" and sender != "classe"):
                        node.append(f'    ${sig_b_namespace[variant]}::{sig_b_name[variant]} = convert_signal(sig_a, sig_a_offset, sig_a_factor, sig_b_offset, sig_b_factor, sig_b_conversion);')
                    elif (sheet == "fdx_out" or sender == 'classe'):
                        node.append(f'    @{sig_b_namespace[variant]}::{sig_b_name[variant]} = convert_signal(sig_a, sig_a_offset, sig_a_factor, sig_b_offset, sig_b_factor, sig_b_conversion);')
            node.append('  }')
            node.append('  else if (((sig_a_min < 0) && (sig_b_min >= 0)) || ((sig_a_min >= 0) && (sig_b_min < 0)))')
            node.append('  {')
            if  sysvardouble_df[col_variant].values[i] != "multi":
                if "XCP" in sender:
                    node.append(f'    @{sender}::{sig_b_namespace[:-1]}::{sig_b_name} = abs(convert_signal(sig_a, sig_a_offset, sig_a_factor, sig_b_offset, sig_b_factor, sig_b_conversion));')
                elif (sheet == "fdx_in" and sender != "classe"):
                    node.append(f'    ${sig_b_namespace}::{sig_b_name} = abs(convert_signal(sig_a, sig_a_offset, sig_a_factor, sig_b_offset, sig_b_factor, sig_b_conversion));')
                elif (sheet == "fdx_out" or sender == "classe"):
                    node.append(f'    @{sig_b_namespace}::{sig_b_name} = abs(convert_signal(sig_a, sig_a_offset, sig_a_factor, sig_b_offset, sig_b_factor, sig_b_conversion));')
            else:
                for variant in range(0,number_of_variants):
                    if (sheet == "fdx_in" and sender != "classe"):
                        node.append(f'    ${sig_b_namespace[variant]}::{sig_b_name[variant]} = abs(convert_signal(sig_a, sig_a_offset, sig_a_factor, sig_b_offset, sig_b_factor, sig_b_conversion));')
                    elif (sheet == "fdx_out" or sender == "classe"):
                        node.append(f'    @{sig_b_namespace[variant]}::{sig_b_name[variant]} = abs(convert_signal(sig_a, sig_a_offset, sig_a_factor, sig_b_offset, sig_b_factor, sig_b_conversion));')
            node.append('  }')
            node.append('}')
            message_count += 1
    if (sheet == "fdx_in"):
        node.append(convert_formula_in)
    elif (sheet == "fdx_out"):
        node.append(convert_formula_out)
    logger.info(f"double signal : generated with {str(message_count)} messages.")


def check_int_double_integrity(value_type, data, signal_name):
    """
    check signal value integrity
    

    Args:
      value_type(str): column value type of signal
      data(str): data of particular signal for column value type
      signal_name(str):signal name for which data integrity has to checked

    Returns:

    """
    if data == "":
        logger.error(f"Missing {value_type} value for {signal_name} signal")
        raise Exception("Please check the database")

# def generate_mapping(sysvar_df):
#     # get column index
#     try:
#         col_name = column_names.index("Name")
#         col_namespace = column_names.index("Message")
#         col_value_type = column_names.index("Value Type")
#         col_unit = column_names.index("Unit")
#         col_texttable = column_names.index("texttable")
#         col_direction = column_names.index("direction")
#         col_data_type = column_names.index("data_type")
#     except:
#         print("\nColumn names in Excel are different from the Script:")
#         print(';'.join(column_names))
#         return False
#
#     xml = ['<?xml version="1.0" encoding="iso-8859-1"?>']
#     xml.append('<canoefdxdescription version="1.0">')
#     #print('\n'.join(xml))
#     direction_list = ["out", "source"]
#     group_id = 1
#     for item in direction_list:
#         item_xml = []
#         if item == "out":
#             #print('    <identifier>EasyDataRead</identifier>')
#             item_xml.append('    <identifier>EasyDataRead</identifier>')
#         elif item == "source":
#             #print('    <identifier>EasyDataWrite</identifier>')
#             item_xml.append('    <identifier>EasyDataWrite</identifier>')
#         #print(sysvar_df[col_direction])
#         sub_sysvar_df = sysvar_df[sysvar_df[col_direction] == item]
#         size_sum = 0
#         offset = 0
#         for i in range(sub_sysvar_df.shape[0]):
#             signal_name = sub_sysvar_df[col_name].iloc[i]
#             namespace = sub_sysvar_df[col_namespace].iloc[i]
#             value = "raw"
#             size = 8
#             unit = sub_sysvar_df[col_unit].iloc[i]
#             texttable = sub_sysvar_df[col_texttable].iloc[i]
#             data_type = sub_sysvar_df[col_data_type].iloc[i]
#             value_type = "int64"
#             if data_type == "int" or data_type == "int enum":
#                 value_type = "uint64" if sub_sysvar_df[col_value_type].iloc[i] == "Unsigned" else "int64"
#             else:
#                 value_type = "double"
#             #print(signal_name, namespace, value, value_type, size, unit)
#             item_xml.append('    <item offset="{0}" size="{1}" type="{2}">'.format(str(offset), str(size), value_type))
#             item_xml.append('      <identifier>{}</identifier>'.format(signal_name))
#             item_xml.append('      <sysvar name="{0}" namespace="{1}" unit="{2}" value="{3}" />'.format(signal_name, namespace, unit, value))
#             item_xml.append('    </item>')
#             offset += size
#             size_sum += size
#         xml.append('  <datagroup groupID="{0}" size="{1}">'.format(str(group_id), str(size_sum)))
#         xml.extend(item_xml)
#         xml.append('  </datagroup>')
#         #print('  <datagroup groupID="{0}" size="{1}">'.format(str(group_id), str(size_sum)))
#         #print('\n'.join(item_xml))
#         #print('  </datagroup>')
#         group_id += 1
#     xml.append('</canoefdxdescription>')
#     return xml

def save_file(file_path):
    """
    Open and save node .can file
    

    Args:
      file_path(str):file path name where can file to save

    Returns:

    """
    # Open and save node .can file
    file = ('{0}/{1}.can'.format(file_path, sheet))
    with open(file, 'w') as outfile:
        outfile.write("\n".join(str(item) for item in node))
        outfile.close()
    logger.info(f"{file} updated successfully")


# main function to create nodes
def create_mapping_main(df_FDX, file_path):
    """
    from Fdx excel sheet it will create .can file as per fdx sheet

    Args:
      df_FDX(dataframe): dataframe of Fdx_sheet
      file_path(str):file path name where can file to save

    Returns:

    """
    global sysvar_df, signal_cloe
    sheet_list =list(df_FDX.keys())
    sysvar_df = pd.DataFrame([column_names])
    signal_cloe = ["Name"]
    dict_fdx_data={"fdx_in":"","fdx_out":""}
    for sheet_pattern in dict_fdx_data.keys():
        sheet_order = [sheet for sheet in sheet_list if sheet_pattern in sheet]
        sheet_order = sorted(sheet_order)
        dict_fdx_data[sheet_pattern] = pd.concat([df_FDX[sheet] for sheet in sheet_order], axis=0, ignore_index=True)

    for sheet_pat, df_value in dict_fdx_data.items():
        logger.info(f"++++++++++Creating mapping for: {sheet_pat} ++++++++++")
        try:
            generate_dataframe(df_value,sheet_pat)
            generate_includes()
            generate_variables()
            generate_onstart()
            generate_onstop()
            generate_onsysvarenum()
            generate_onsysvardouble()
            save_file(file_path+"/Nodes")
        except Exception as exp:
            logger.error(f"Failed to create node for: {sheet_pat}")
            logger.error(f"Error: {exp}")
            return False
    # try:
    #     index_list = []
    #     for i in range(sysvar_df.shape[0]):
    #         try:
    #             index_list.append(signal_cloe.index(sysvar_df.iloc[i][col_name]))
    #         except:
    #             print(sysvar_df.iloc[i][col_name], " signal is not found in sysvar_df.")
    #             raise Exception("Please check that 3 generate functions.")
    #     sysvar_df["Index"] = index_list
    #     sysvar_df = sysvar_df.sort_values("Index")
    #     sysvar_xml = generate_sysvar(sysvar_df, FDX=True)
    #     fdx_xml = generate_mapping(sysvar_df)
    #     save_to_file(sysvar_xml, file_path+"/Databases", "FDX_Database")
    #     save_to_file(fdx_xml, file_path+"/Databases", "FDX_Mapping")
    # except Exception as exp:
    #     print("\nFailed to create sysvar&mapping:", exp)
    #     return False
    return True


def external_call():
    """ """
    try:
        logger.info("###### START 'Update FDX Nodes' DEBUG INFORMATION ######")
        script_path = os.path.dirname(os.path.abspath(__file__))
        wb = load_workbook(script_path + r'/../../../../CustomerPrj/FDX/FDX_Database.xlsx', data_only=True)
        #wb_classe = load_workbook(script_path + '/../../Classe_Database.xlsx', data_only=True)
        #df_FDX=pd.ExcelFile(script_path + r'/../../../../CustomerPrj/FDX/FDX_Database.xlsx')
        df_FDX=pd.read_excel(script_path + r'/../../../../CustomerPrj/FDX/FDX_Database.xlsx',sheet_name=None,dtype=object)
        create_mapping_main(df_FDX, script_path + r'/../../../../CustomerPrj/FDX/')
        #df_FDX.close()
    except Exception as e:
        logger.error(f"Update Update FDX Nodes failed. Exception --> {e}")
    logger.info("###### END 'Update FDX Nodes' DEBUG INFORMATION ######")
    logger.info('-' * 80)


if __name__ == "__main__":
    external_call()
