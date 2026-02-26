# -*- coding: utf-8 -*-
# @file test_database.py
# @author ADAS_HIL_TEAM
# @date 08-21-2023

##################################################################
# C O P Y R I G H T S
# ----------------------------------------------------------------
# Copyright (c) 2023 by Robert Bosch GmbH. All rights reserved.

# The reproduction, distribution and utilization of this file as
# well as the communication of its contents to others without express
# authorization is prohibited. Offenders will be held liable for the
# payment of damages. All rights reserved in the event of the grant
# of a patent, utility model or design.

##################################################################


import os, sys
import math
import pandas as pd
from openpyxl import load_workbook

try:
    from Control.logging_config import logger
except ImportError:
    sys.path.append(os.path.dirname(os.path.abspath(__file__)) + r"\..\Control")
    from logging_config import logger

empty_types = ["", "nan"]


def cell_empty(cell_val, column_name):
    """
    Description: Function compares column value is empty or "nan". logs error if condition is not satisfied.

    Args:
        cell_val (str): Column value
        column_name (str): Column Name

    Returns: True or False

    """
    global status, LoggedError
    if (str(cell_val).strip() in empty_types):
        status = status and False
        LoggedError += column_name + " column is empty \n"
        return True
    else:
        status = status and True
        return False


def empty_cell(cell_val):
    """
    Description: Function compares column value is empty or "nan"

    Args:
        cell_val (str): Column value:

    Returns: True or False

    """
    if (str(cell_val).strip() in empty_types):
        return True
    else:
        return False


def tc_1_signal_name_test(row):
    """
    Description: Function checks signal name is empty or not.

    Args:
        row (Pandas Series): Series contains all column values of individual row

    Returns:

    """
    # Case 0: column "Name" i.e. signal name should not be empty
    if cell_empty(row["Signal"], "Signal"):
        return


def tc_2_message_name_test(row):
    """
    Description: Function checks Message name is empty or not.

    Args:
        row (Pandas Series): Series contains all column values of individual row

    Returns:

    """
    global status, LoggedError
    # Case 0: "Message" name should not be empty
    if cell_empty(row["Message"], "Message"):
        return


def tc_3_startbit_test(row):
    """
    Description: Function checks startbit of signal is unique for given message/pdu, and it is not empty.
                 logs error if condition is not satisfied.

    Args:
        row (Pandas Series): Series contains all column values of individual row

    Returns:

    """
    global status, LoggedError
    # Case 0: "Startbit" should be integer
    if cell_empty(row["Startbit"], "Startbit"):
        return
    if type(row["Startbit"]) != int:
        status = status and False
        LoggedError += "Startbit should be integer \n"
        return
    else:
        status = status and True

    # Case 1: In one message two signals shall not have same "start bit" if input_file is dbc
    #         In one pdu two signals shall not have same "start bit" if input_file is arxml
    # if row["input_file"] == "dbc":
    #     if empty_cell(row["Message"]) is False:
    #         if row["Message"] in dict_message_startbit.keys():
    #             if row["Startbit"] in dict_message_startbit[row["Message"]]:
    #                 status = status and False
    #                 LoggedError += "Signal has repeated start bit \n"
    #             else:
    #                 dict_message_startbit[row["Message"]].append(row["Startbit"])
    #                 status = status and True
    #         else:
    #             dict_message_startbit[row["Message"]] = [row["Startbit"]]
    #             status = status and True
    # elif row["input_file"] == "arxml":
    #     if empty_cell(row["PDU"]) is False:
    #         if row["PDU"] in dict_pdu_startbit.keys():
    #             if row["Startbit"] in dict_pdu_startbit[row["PDU"]]:
    #                 status = status and False
    #                 LoggedError += "Signal has repeated start bit \n"
    #             else:
    #                 dict_pdu_startbit[row["PDU"]].append(row["Startbit"])
    #                 status = status and True
    #         else:
    #             dict_pdu_startbit[row["PDU"]] = [row["Startbit"]]
    #             status = status and True


def tc_4_length_bit_test(row):
    """
    Description: Function checks length of signal is not empty and type is int.
                 logs error if condition is not satisfied.

    Args:
        row (Pandas Series): Series contains all column values of individual row

    Returns:

    """
    global status, LoggedError
    # Case 0: "Length [Bit]" should be integer
    if cell_empty(row["Signal_Length [Bit]"], "Signal_Length [Bit]"):
        return
    if type(row["Signal_Length [Bit]"]) != int:
        status = status and False
        LoggedError += "Signal_Length [Bit] should be integer \n"
        return
    else:
        status = status and True


def tc_5_byte_order_test(row):
    """
    Description: Function checks bye order of signal is not empty and type is string.
                 logs error if condition is not satisfied.

    Args:
        row (Pandas Series): Series contains all column values of individual row

    Returns:

    """
    global status, LoggedError
    # Case 0: "Byte Order" should not be empty
    if cell_empty(row["Byte Order"], "Byte Order"):
        return
    if type(row["Byte Order"]) != str:
        status = status and False
        LoggedError += "Byte Order should be string \n"
        return
    else:
        status = status and True


def tc_6_value_type_test(row):
    """
    Description: Function checks value type is not empty and type is string. logs error if condition is not satisfied.

    Args:
        row (Pandas Series): Series contains all column values of individual row

    Returns:

    """
    global status, LoggedError
    # Case 0: "Value Type" should not be empty
    if cell_empty(row["Value Type"], "Value Type"):
        return
    if type(row["Value Type"]) != str:
        status = status and False
        LoggedError += "Value Type should be string \n"
        return
    else:
        status = status and True


def tc_7_initial_value_test(row):
    """
    Description: Function checks initial value of signal is not empty and type is int or float,
                 initial value >= minimum value and <= maximum value
                 logs error if condition is not satisfied.

    Args:
        row (Pandas Series): Series contains all column values of individual row

    Returns:

    """
    global status, LoggedError
    # Case 0: "Initial Value" column must have integer or float values
    if cell_empty(row["Initial Value"], "Initial Value"):
        return
    if (type(row["Initial Value"]) != int) and (type(row["Initial Value"]) != float):
        status = status and False
        LoggedError += "Initial Value is neither Integer nor Float \n"
        return
    else:
        status = status and True

    # Case 1: "Initial Value" should be +ve if "Value Type" is Unsigned
    # if empty_cell(row["Value Type"]) is False:
    #     if row["Value Type"] == "Unsigned":
    #         if row["Initial Value"] < 0:
    #             status = status and False
    #             LoggedError += "Initial Value is -ve as Value Type is Unsigned \n"
    #         else:
    #             status = status and True

    # Case 2: "Initial Value" value shall not be less than "Minimum" value
    if empty_cell(row["Minimum"]) is False:
        if row["Initial Value"] < row["Minimum"]:
            status = status and False
            LoggedError += "Initial Value is less than Minimum value \n"
        else:
            status = status and True

    # Case 3: "Initial Value" value shall not be greater than "Maximum" value
    if empty_cell(row["Maximum"]) is False:
        if row["Initial Value"] > row["Maximum"]:
            status = status and False
            LoggedError += "Initial Value is greater than Maximum value \n"
        else:
            status = status and True


def tc_8_factor_test(row):
    """
    Description: Function checks factor of signal is not empty and type is int or float, factor is nonzero and +ve
                 if value type is unsigned. logs error if condition is not satisfied.

    Args:
        row (Pandas Series): Series contains all column values of individual row

    Returns:

    """
    global status, LoggedError
    # Case 0: "Factor" column must have integer or float values
    if cell_empty(row["Factor"], "Factor"):
        return
    if (type(row["Factor"]) != int) and (type(row["Factor"]) != float):
        status = status and False
        LoggedError += "Factor is neither Integer nor Float \n"
        return
    else:
        status = status and True

    # Case 1: "Factor" should be +ve if "Value Type" is Unsigned, and it should be nonzero
    if empty_cell(row["Value Type"]) is False:
        if row["Value Type"] == "Unsigned":
            if row["Factor"] <= 0:
                status = status and False
                if row["Factor"] == 0:
                    LoggedError += "Factor is 0 \n"
                else:
                    LoggedError += "Factor is -ve as Value Type is Unsigned \n"
            else:
                status = status and True
        else:  # Signed Or IEEE Float
            if row["Factor"] == 0:
                status = status and False
                LoggedError += "Factor is 0 \n"
            else:
                status = status and True


def tc_9_offset_test(row):
    """
    Description: Function checks offset of signal is not empty and type is int or float. Offset is <= Minimum value
                 if value type is unsigned. logs error if condition is not satisfied.

    Args:
        row (Pandas Series): Series contains all column values of individual row

    Returns:

    """
    global status, LoggedError
    # Case 0: "Offset" column must have integer or float values
    if cell_empty(row["Offset"], "Offset"):
        return
    if (type(row["Offset"]) != int) and (type(row["Offset"]) != float):
        status = status and False
        LoggedError += "Offset is neither Integer nor Float \n"
        return
    else:
        status = status and True

    # Case 1: "Offset" should be <= "Minimum" value if "Value Type" is Unsigned
    if empty_cell(row["Value Type"]) is False and empty_cell(row["Minimum"]) is False:
        if row["Value Type"] == "Unsigned":
            if row["Offset"] > row["Minimum"]:
                status = status and False
                LoggedError += "Offset should be <= minimum value \n"
            else:
                status = status and True


def tc_10_minimum_value_test(row):
    """
    Description: Function checks Minimum value of signal is not empty and type is int or float.
                 logs error if condition is not satisfied.

    Args:
        row (Pandas Series): Series contains all column values of individual row

    Returns:

    """
    global status, LoggedError
    # Case 0: "Minimum" column must have integer or float values
    if cell_empty(row["Minimum"], "Minimum"):
        return
    if (type(row["Minimum"]) != int) and (type(row["Minimum"]) != float):
        status = status and False
        LoggedError += "Minimum value is neither Integer nor Float \n"
        return
    else:
        status = status and True

    # Case 1: "Minimum" value should be +ve if "Value Type" is Unsigned
    # if empty_cell(row["Value Type"]) is False:
    #     if row["Value Type"] == "Unsigned":
    #         if row["Minimum"] < 0:
    #             status = status and False
    #             LoggedError += "Minimum value is -ve as Value Type is Unsigned \n"
    #         else:
    #             status = status and True


def tc_11_maximum_value_test(row):
    """
    Description: Function checks maximum value of signal is not empty, type is int or float.
                 Maximum value is positive if value type is unsigned, Maximum value is <= "max_value"
                 or <= ((2 ^ Length [Bit]) - 1). logs error if condition is not satisfied.

    Args:
        row (Pandas Series): Series contains all column values of individual row

    Returns:

    """
    global status, LoggedError
    # Case 0: "Maximum" column must have integer or float values
    if cell_empty(row["Maximum"], "Maximum"):
        return
    if (type(row["Maximum"]) != int) and (type(row["Maximum"]) != float):
        status = status and False
        LoggedError += "Maximum value is neither Integer nor Float \n"
        return
    else:
        status = status and True

    # Case 1: "Maximum" value should be +ve if "Value Type" is Unsigned
    if empty_cell(row["Value Type"]) is False:
        if row["Value Type"] == "Unsigned":
            if row["Maximum"] < 0:
                status = status and False
                LoggedError = LoggedError + "Maximum value is -ve as Value Type is Unsigned \n"
            else:
                status = status and True

    # Case 2: "Maximum" value shall not be less than or equal to "Minimum" value
    if empty_cell(row["Minimum"]) is False:
        if row["Maximum"] < row["Minimum"]:
            status = status and False
            LoggedError += "Maximum value is less than to Minimum value \n"
        else:
            status = status and True

    # Case 3: "Maximum" raw value shall be <= "max_value" column or <= ((2 ^ Length [Bit]) - 1)
    # if empty_cell(row["Factor"]) is False and empty_cell(row["Offset"]) is False and empty_cell(
    #         row["max_value"]) is False and empty_cell(row["Length [Bit]"]) is False:
    #     if ((type(row["Factor"]) is int) or (type(row["Factor"]) is float)) and row["Factor"] != 0:
    #         Max_raw = (row["Maximum"] - row["Offset"]) // row["Factor"]  # Raw value should be integer
    #         if Max_raw > row["max_value"] or Max_raw > (math.pow(2, row["Length [Bit]"]) - 1):
    #             status = status and False
    #             LoggedError += "\"Maximum"" raw value is greater than \"max_value"" or greater than ((2 ^ Length [Bit]) - 1) \n"
    #         else:
    #             status = status and True


def tc_12_value_table_test(row):
    """
    Description: Function checks if value table is not empty. logs error if condition is not satisfied.

    Args:
        row (Pandas Series): Series contains all column values of individual row

    Returns:

    """
    # Case 0: “Value Table” column shall have data if “texttable” column filled with texttable or not empty
    if empty_cell(row["texttable"]) is False:
        if row["texttable"] == "texttable":
            if cell_empty(row["Value Table"], "Value Table"):
                return


def tc_13_message_id_test(row):
    """
    Description: Function checks message id is not empty and is in hex format. Each signal mapped to message has
                 same message id. logs error if condition is not satisfied.

    Args:
        row (Pandas Series): Series contains all column values of individual row

    Returns:

    """
    global status, LoggedError
    # Case 0: "Message ID" should not contain special character or empty
    if cell_empty(row["Message ID"], "Message ID"):
        return
    try:
        hex(int(row["Message ID"], 16))
        status = status and True
    except ValueError as Argument:
        status = status and False
        LoggedError += "Message Id should not contain special character \n"
        return

    # Case 1: In one message each signal shall have same "Message ID"
    if empty_cell(row["Message"]) is False:
        if row["Message"] in dict_message_messageid.keys():
            if row["Message ID"] not in dict_message_messageid[row["Message"]]:
                status = status and False
                LoggedError += "Message ID is not same with other signals \n"
            else:
                dict_message_messageid[row["Message"]].append(row["Message ID"])
                status = status and True
        else:
            dict_message_messageid[row["Message"]] = [row["Message ID"]]
            status = status and True


def tc_14_cycle_time_test(row):
    """
    Description: Function checks cycle time of signal is not empty, is > 0 and type is int. Each signal mapped to one
                 message/pdu has same cycle time. logs error if condition is not satisfied.

    Args:
        row (Pandas Series): Series contains all column values of individual row

    Returns:

    """
    global status, LoggedError
    # Case 0: "Cycle Time [ms]" should be integer
    if cell_empty(row["Cycle Time [ms]"], "Cycle Time [ms]"):
        return
    if type(row["Cycle Time [ms]"]) != int:
        status = status and False
        LoggedError += "Cycle Time [ms] should be integer \n"
        return
    else:
        status = status and True

    # Case 1: "Cycle Time [ms]" should be >= 0
    if row["Cycle Time [ms]"] < 0:
        status = status and False
        LoggedError += "Cycle Time [ms] should be >= 0 \n"
    else:
        status = status and True

    # Case 2: In one message each signal shall have same "Cycle Time [ms]" if input_file is dbc
    #         In one pdu each signal shall have same "Cycle Time [ms]" if input_file is arxml
    if row["input_file"] == "dbc":
        if empty_cell(row["Message"]) is False:
            if row["Message"] in dict_message_cycletime.keys():
                if row["Cycle Time [ms]"] not in dict_message_cycletime[row["Message"]]:
                    status = status and False
                    LoggedError += "Cycle time is not same with other signals \n"
                else:
                    dict_message_cycletime[row["Message"]].append(row["Cycle Time [ms]"])
                    status = status and True
            else:
                dict_message_cycletime[row["Message"]] = [row["Cycle Time [ms]"]]
                status = status and True
    elif row["input_file"] == "arxml":
        if empty_cell(row["PDU"]) is False:
            if row["PDU"] in dict_pdu_cycletime.keys():
                if row["Cycle Time [ms]"] not in dict_pdu_cycletime[row["PDU"]]:
                    status = status and False
                    LoggedError += "Cycle time is not same with other signals \n"
                else:
                    dict_pdu_cycletime[row["PDU"]].append(row["Cycle Time [ms]"])
                    status = status and True
            else:
                dict_pdu_cycletime[row["PDU"]] = [row["Cycle Time [ms]"]]
                status = status and True


def tc_15_texttable_test(row):
    """
    Description: Function checks texttable column is not empty if texttable values column is not empty.
                 logs error if condition is not satisfied.

    Args:
        row (Pandas Series): Series contains all column values of individual row

    Returns:

    """
    # Case 0: “texttable” column shall not be empty if “texttable values” column filled with data
    if empty_cell(row["texttable values"]) is False:
        if cell_empty(row["texttable"], "texttable"):
            return


def tc_16_texttable_values_test(row):
    """
    Description: Function checks texttable values column is not empty. logs error if condition is not satisfied.

    Args:
        row (Pandas Series): Series contains all column values of individual row

    Returns:

    """
    # Case 0: “texttable values” column shall have data if “texttable” column filled with texttable or not empty
    if empty_cell(row["texttable"]) is False:
        if row["texttable"] == "texttable":
            if cell_empty(row["texttable values"], "texttable values"):
                return


def tc_17_max_value_test(row):
    """
    Description: Function checks max_value of signal is not empty and type is int or float. max_value is
                 <= ((2 ^ Length [Bit]) - 1). logs error if condition is not satisfied.

    Args:
        row (Pandas Series): Series contains all column values of individual row

    Returns:

    """
    global status, LoggedError
    # Case 0: "max_value" must have integer or float values
    if cell_empty(row["max_value"], "max_value"):
        return
    if (type(row["max_value"]) != int) and (type(row["max_value"]) != float):
        status = status and False
        LoggedError += "max_value is neither Integer nor Float \n"
        return
    else:
        status = status and True

    # # Case 1: "max_value" shall be <= ((2 ^ Length [Bit]) - 1)
    # if empty_cell(row["Length [Bit]"]) is False:
    #     if row["max_value"] > (math.pow(2, row["Length [Bit]"]) - 1):
    #         status = status and False
    #         LoggedError += "max_value is greater than ((2 ^ Length [Bit]) - 1) \n"
    #     else:
    #         status = status and True


def tc_18_dlc_test(row):
    """
    Description: Function checks DLC of signal is not empty, type is int and DLC is > 0.
                 logs error if condition is not satisfied.

    Args:
        row (Pandas Series): Series contains all column values of individual row

    Returns:

    """
    global status, LoggedError
    # Case 0: "dlc" should be integer
    if cell_empty(row["dlc"], "dlc"):
        return
    if type(row["dlc"]) != int:
        status = status and False
        LoggedError += "dlc should be integer \n"
        return
    else:
        status = status and True

    # Case 1: "dlc" should be > 0
    if row["dlc"] < 0:
        status = status and False
        LoggedError += "dlc should be > 0 \n"
    else:
        status = status and True


def tc_19_sender_test(row):
    """
    Description: Function checks sender of signal is not empty. Sender of each signal mapped to message/pdu is same.
                 logs error if condition is not satisfied.

    Args:
        row (Pandas Series): Series contains all column values of individual row

    Returns:

    """
    global status, LoggedError
    # Case 0: column "sender" should not be empty
    if cell_empty(row["sender"], "sender"):
        return

    # Case 1: In one message each signal shall have same "sender" if input_file is dbc
    #         In one pdu each signal shall have same "sender" if input_file is arxml
    if row["input_file"] == "dbc":
        if empty_cell(row["Message"]) is False:
            if row["Message"] in dict_message_sender.keys():
                if row["sender"] not in dict_message_sender[row["Message"]]:
                    status = status and False
                    LoggedError += "Sender is not same with other signals \n"
                else:
                    dict_message_sender[row["Message"]].append(row["sender"])
                    status = status and True
            else:
                dict_message_sender[row["Message"]] = [row["sender"]]
                status = status and True
    elif row["input_file"] == "arxml":
        if empty_cell(row["PDU"]) is False:
            if row["PDU"] in dict_pdu_sender.keys():
                if row["sender"] not in dict_pdu_sender[row["PDU"]]:
                    status = status and False
                    LoggedError += "Sender is not same with other signals \n"
                else:
                    dict_pdu_sender[row["PDU"]].append(row["sender"])
                    status = status and True
            else:
                dict_pdu_sender[row["PDU"]] = [row["sender"]]
                status = status and True

def tc_19_eth_sender_test(row):
    """
    Description: Function checks sender of signal is not empty. Sender of each signal mapped to message/pdu is same.
                 logs error if condition is not satisfied.

    Args:
        row (Pandas Series): Series contains all column values of individual row

    Returns:

    """
    global status, LoggedError
    # Case 0: column "sender" should not be empty
    if cell_empty(row["Node Name"], "Node Name"):
        return


def tc_20_pdu_test(row):
    """
    Description: Function checks PDU is not empty. logs error if condition is not satisfied.

    Args:
        row (Pandas Series): Series contains all column values of individual row

    Returns:

    """
    global status, LoggedError
    # Case 0: "PDU" should not be empty
    if cell_empty(row["PDU"], "PDU"):
        return
    
def tc_21_profile_test(row):
    """
    Description: Function checks siganl is of type CRC or not if siganl is pf type CRC checks Profile is 
    defined or not for the siganal.

    Args:
        row (Pandas Series): Series contains all column values of individual row

    Returns:

    """
    global status, LoggedError
    if empty_cell(row["Signal"]) is False:
        if "CRC" in row["Signal"] and "profile" not in str(row["EndToEndProtection"]).lower():
            status = status and False
            LoggedError += "CRC pattern matched but no Profile defined \n"
        else:
            status = status and True

def tc_23_payload_pdu_type_test(row):
    global status, LoggedError
    if (str(row["Payload_PDU_Type"]) in empty_types) and (str(row["Payload_PDU"]) not in empty_types):
        status = status and False
        LoggedError += "Payload_PDU_Type is empty but Payload_PDU is not empty. \n"
    elif str(row["Payload_PDU_Type"]) not in empty_types and str(row["Payload_PDU"]) in empty_types:
        status = status and False
        LoggedError += "Payload_PDU is empty but Payload_PDU_Type is not empty. \n"
    else:
        status = status and True

def findEndBit(startbit,bitlength):   # ONLY FOR MOTOROLA
    count = 1
    while count!=bitlength:
        if (startbit+1) % 8 == 0:
            startbit -= 16
        count += 1
        startbit += 1
    return int(startbit)

# def find_bit_range(end_bit):
#     bit_start_point = end_bit
#     while bit_start_point % 8 != 0:
#         bit_start_point -= 1
#     bit_end_point = bit_start_point + 7
#     return bit_start_point, bit_end_point

def check_startbit_overlapped(number_list, startbit, length): #only for motorola byte order
    """
    Remove elements from number_list starting from startbit for the given length.
    Update the occupied list and adjust startbit accordingly.
    """
    counter = 0
    occupied = []
    while counter < length and startbit >= 0:
        if (startbit + 1) % 8 == 0:
            if startbit in number_list:
                #number_list.remove(startbit)
                occupied.append(startbit)
            startbit = (startbit + 1) - 16
        else:
            if startbit in number_list:
                #number_list.remove(startbit)
                occupied.append(startbit)
            startbit = startbit + 1
        counter += 1
    return occupied

def next_number_to_be_divisible_by_8(input_number): #only for motorola
    # if input_number > 0 and input_number % 8 == 0:
    #     return input_number
    # else:
    next_number = input_number + 1
    while (next_number) % 8 != 0:
        next_number += 1
    return next_number

def findStartBit(startbit, length):# ONLY FOR MOTOROLA arxml
    """
    

    Args:
      startbit: 
      length: 

    Returns:

    """
    if length<1:
        return int(startbit)
    while length != 1:
        if startbit % 8 == 0:
            startbit += 16
        length -= 1
        startbit -= 1
    return int(startbit)

def generate_occupied(startbit, length):
    occupied = []
    while len(occupied) != length:
        if startbit % 8 == 0:
            occupied.append(startbit)
            startbit+=15
        else:
            occupied.append(startbit)
            startbit-=1
    return occupied

def tc_22_test_startbit(df, input_file_type):
    messages = df['Message'].drop_duplicates().tolist()
    for message in messages:
        message_df = df[(df["Message"] == message)]
        byte_order_types = message_df['Byte Order'].dropna().drop_duplicates().tolist()
        # print(byte_order_types)
        if len(byte_order_types) > 1:
            logger.warning(f"more then one byte_order_types in network : {message_df['network_name'].iloc[0]} sender : {message_df['sender'].iloc[0]} message : {message_df['Message'].iloc[0]}")
        elif len(byte_order_types) == 1 and "Intel" in byte_order_types:
            message_df.loc[:, "Payload_PDU_Type"] = message_df["Payload_PDU_Type"].fillna("NaN_placeholder")
            for payload_pdu_type, payload_pdu_type_df in message_df.groupby(message_df["Payload_PDU_Type"]):
                if payload_pdu_type == "CONTAINER-I-PDU" or payload_pdu_type == "N-PDU":
                    for pdu_type, pdu_type_df in payload_pdu_type_df.groupby(payload_pdu_type_df["PDU_Type"]):
                        if pdu_type == "I-SIGNAL-I-PDU":
                            for pdu, pdu_df in pdu_type_df.groupby(pdu_type_df["PDU"]):
                                sorted_df = pdu_df.sort_values(by="Startbit")
                                for idx in range(len(sorted_df) - 1):
                                    current_row  = sorted_df.iloc[idx]
                                    if empty_cell(current_row["Startbit"]) is False and empty_cell(current_row["Signal_Length [Bit]"]) is False:
                                        print("Hello")
                                        next_row = sorted_df.iloc[idx + 1]
                                        end_bit = (current_row["Startbit"] + current_row["Signal_Length [Bit]"]) - 1
                                        next_start_bit = next_row["Startbit"]
                                        if next_start_bit <= end_bit:
                                            current_row_df = pd.DataFrame([next_row], dtype=object)
                                            row = current_row_df.iloc[0]
                                            condition = (df["network_name"] == row["network_name"]) & \
                                                        (df["sender"] == row["sender"]) & \
                                                        (df["Message"] == row["Message"]) & \
                                                        (df["Byte Order"] == row["Byte Order"]) & \
                                                        (df["Signal"] == row["Signal"]) & \
                                                        (df["Startbit"] == row["Startbit"])
                                            df.loc[condition, "Remarks"] = "startbit is <= endbit of previous signal \n"
                                            df.loc[condition, "Status"] = "Fail"
                elif payload_pdu_type == "NaN_placeholder":
                    for pdu_type, pdu_type_df in payload_pdu_type_df.groupby(payload_pdu_type_df["PDU_Type"]):
                        if pdu_type == "I-SIGNAL-I-PDU":
                            for pdu, pdu_df in pdu_type_df.groupby(pdu_type_df["PDU"]):
                                sorted_df = pdu_df.sort_values(by="Startbit")
                                for idx in range(len(sorted_df) - 1):
                                    current_row  = sorted_df.iloc[idx]
                                    if empty_cell(current_row["Startbit"]) is False and empty_cell(current_row["Signal_Length [Bit]"]) is False:
                                        next_row = sorted_df.iloc[idx + 1]
                                        end_bit = (current_row["Startbit"] + current_row["Signal_Length [Bit]"]) - 1
                                        next_start_bit = next_row["Startbit"]
                                        if next_start_bit <= end_bit:
                                            current_row_df = pd.DataFrame([next_row], dtype=object)
                                            row = current_row_df.iloc[0]
                                            condition = (df["network_name"] == row["network_name"]) & \
                                                        (df["sender"] == row["sender"]) & \
                                                        (df["Message"] == row["Message"]) & \
                                                        (df["Byte Order"] == row["Byte Order"]) & \
                                                        (df["Signal"] == row["Signal"]) & \
                                                        (df["Startbit"] == row["Startbit"])
                                            df.loc[condition, "Remarks"] = "startbit is <= endbit of previous signal \n"
                                            df.loc[condition, "Status"] = "Fail"
        elif len(byte_order_types) == 1 and "Motorola" in byte_order_types and input_file_type == "dbc":
            sorted_df = message_df.sort_values(by="Startbit")
            start_bits = sorted_df['Startbit'].tolist()
            max_bit = max(start_bits)
            max_bit = next_number_to_be_divisible_by_8(max_bit)
            total_bits = list(range(max_bit))
            occupied_bits = []
            for idx in range(len(sorted_df)):
                current_row  = sorted_df.iloc[idx]
                if empty_cell(current_row["Startbit"]) is False and empty_cell(current_row["Signal_Length [Bit]"]) is False:
                    #end_bit = findEndBit(current_row["Startbit"], current_row["Length [Bit]"])
                    # if current_row["network_name"] == "CAN_MPC3" and current_row["sender"] == "MPC3" and current_row["Message"] == "ABS_UN_PR_4":
                    #     print("total bits : ", total_bits)
                    #     print("occupied bits : ", occupied_bits)
                    if current_row["Startbit"] not in occupied_bits:
                        occupied = check_startbit_overlapped(total_bits, current_row["Startbit"], current_row["Signal_Length [Bit]"])
                        # if current_row["network_name"] == "CAN_MPC3" and current_row["sender"] == "MPC3" and current_row["Message"] == "ABS_UN_PR_4":
                        #     print("occupied : ", occupied)
                        if len(occupied) == current_row["Signal_Length [Bit]"]:
                            for bit in occupied:
                                if bit not in occupied_bits:
                                    occupied_bits.append(bit)
                                else:
                                    current_row_df = pd.DataFrame([current_row], dtype=object)
                                    row = current_row_df.iloc[0]
                                    condition = (df["network_name"] == row["network_name"]) & \
                                                (df["sender"] == row["sender"]) & \
                                                (df["Message"] == row["Message"]) & \
                                                (df["Byte Order"] == row["Byte Order"]) & \
                                                (df["Signal"] == row["Signal"]) & \
                                                (df["Startbit"] == row["Startbit"])
                                    df.loc[condition, "Remarks"] = "startbit is overlapped in previous signal layout \n"
                                    df.loc[condition, "Status"] = "Fail"
                        else:
                            current_row_df = pd.DataFrame([current_row], dtype=object)
                            row = current_row_df.iloc[0]
                            condition = (df["network_name"] == row["network_name"]) & \
                                        (df["sender"] == row["sender"]) & \
                                        (df["Message"] == row["Message"]) & \
                                        (df["Byte Order"] == row["Byte Order"]) & \
                                        (df["Signal"] == row["Signal"]) & \
                                        (df["Startbit"] == row["Startbit"])
                            df.loc[condition, "Remarks"] = "bit range contains negative range \n"
                            df.loc[condition, "Status"] = "Fail"
                    else:
                        current_row_df = pd.DataFrame([current_row], dtype=object)
                        row = current_row_df.iloc[0]
                        condition = (df["network_name"] == row["network_name"]) & \
                                    (df["sender"] == row["sender"]) & \
                                    (df["Message"] == row["Message"]) & \
                                    (df["Byte Order"] == row["Byte Order"]) & \
                                    (df["Signal"] == row["Signal"]) & \
                                    (df["Startbit"] == row["Startbit"])
                        df.loc[condition, "Remarks"] = "startbit is overlapped in previous signal layout \n"
                        df.loc[condition, "Status"] = "Fail"
        elif len(byte_order_types) == 1 and "Motorola" in byte_order_types and input_file_type == "arxml":
            # print(message_df)
            for payload_pdu_type, payload_pdu_type_df in message_df.groupby(message_df["Payload_PDU_Type"]):
                if payload_pdu_type == "CONTAINER-I-PDU" or payload_pdu_type == "N-PDU":
                    for pdu_type, pdu_type_df in payload_pdu_type_df.groupby(payload_pdu_type_df["PDU_Type"]):
                        if pdu_type == "I-SIGNAL-I-PDU":
                            for pdu, pdu_df in pdu_type_df.groupby(pdu_type_df["PDU"]):
                                # print("pdu is : ", pdu)
                                sorted_df = pdu_df.sort_values(by="Startbit")
                                occupied_bits = []
                                for idx in range(len(sorted_df)):
                                    current_row  = sorted_df.iloc[idx]
                                    if empty_cell(current_row["Startbit"]) is False and empty_cell(current_row["Signal_Length [Bit]"]) is False:
                                        occupied = generate_occupied(current_row["Startbit"], current_row["Signal_Length [Bit]"])
                                        # print(current_row["Signal"])
                                        # print(occupied)
                                        for bit in occupied:
                                            if bit not in occupied_bits:
                                                occupied_bits.append(bit)
                                            else:
                                                current_row_df = pd.DataFrame([current_row], dtype=object)
                                                row = current_row_df.iloc[0]
                                                condition = (df["network_name"] == row["network_name"]) & \
                                                            (df["sender"] == row["sender"]) & \
                                                            (df["Message"] == row["Message"]) & \
                                                            (df["Byte Order"] == row["Byte Order"]) & \
                                                            (df["Signal"] == row["Signal"]) & \
                                                            (df["Startbit"] == row["Startbit"])
                                                df.loc[condition, "Remarks"] = "startbit is overlapped in previous signal layout \n"
                                                df.loc[condition, "Status"] = "Fail"
                elif payload_pdu_type == "NaN_placeholder":
                    for pdu_type, pdu_type_df in payload_pdu_type_df.groupby(payload_pdu_type_df["PDU_Type"]):
                        if pdu_type == "I-SIGNAL-I-PDU":
                            for pdu, pdu_df in pdu_type_df.groupby(pdu_type_df["PDU"]):
                                print("pdu is : ", pdu)
                                sorted_df = pdu_df.sort_values(by="Startbit")
                                occupied_bits = []
                                for idx in range(len(sorted_df)):
                                    current_row  = sorted_df.iloc[idx]
                                    if empty_cell(current_row["Startbit"]) is False and empty_cell(current_row["Signal_Length [Bit]"]) is False:
                                        occupied = generate_occupied(current_row["Startbit"], current_row["Signal_Length [Bit]"])
                                        print(current_row["Signal"])
                                        print(occupied)
                                        for bit in occupied:
                                            if bit not in occupied_bits:
                                                occupied_bits.append(bit)
                                            else:
                                                current_row_df = pd.DataFrame([current_row], dtype=object)
                                                row = current_row_df.iloc[0]
                                                condition = (df["network_name"] == row["network_name"]) & \
                                                            (df["sender"] == row["sender"]) & \
                                                            (df["Message"] == row["Message"]) & \
                                                            (df["Byte Order"] == row["Byte Order"]) & \
                                                            (df["Signal"] == row["Signal"]) & \
                                                            (df["Startbit"] == row["Startbit"])
                                                df.loc[condition, "Remarks"] = "startbit is overlapped in previous signal layout \n"
                                                df.loc[condition, "Status"] = "Fail"


    return df

def test_varient_cols(df, log_file_path, db_type):
    if db_type == "CAN":
        columns_to_check = df.columns[df.columns.get_loc('network_name') + 1:]
    if db_type == "ETHERNET":
        columns_to_check = df.columns[df.columns.get_loc('Network Name') + 1:]
    with open(log_file_path, 'a') as log_file:
        log_file.write(f"################{db_type}################\n")
        for i, row in df.iterrows():
            sub_row = row[columns_to_check]
            empty_columns = sub_row[sub_row.isna()].index.tolist()
            if empty_columns:
                log_file.write(f"[ERROR] : signal : {row.iloc[0]} has empty values in variant columns : {empty_columns}\n")
    return

def test_sysvar_database(df, input_file_type):
    """
    Description: Function takes reference of SysVarDatabase sheet and evaluates each row based on certain conditions.
                 Generates report Test_Database.xlsx that consist of pass/fail status of each signal and statistics i.e.
                 % passed signal etc.

    Args:
        df (DataFrame): Dataframe of SysVarDatabase sheet of the Autosar_Gen_Database.xlsx

    Returns:

    """
    global status, LoggedError, dict_message_startbit, dict_pdu_startbit, dict_pdu_cycletime, dict_message_cycletime, dict_message_messageid, dict_message_sender, dict_pdu_sender, signal_pass, signal_fail
    dict_message_startbit = {}
    dict_pdu_startbit = {}
    dict_message_cycletime = {}
    dict_pdu_cycletime = {}
    dict_message_messageid = {}
    dict_message_sender = {}
    dict_pdu_sender = {}
    signal_pass, signal_fail = 0, 0
    rows = df.shape[0]  # No of rows
    col = df.shape[1]  # No of columns
    df["Status"] = None  # Add new column to excel
    df["Remarks"] = None  # Add new column to excel
    
    
    df = tc_22_test_startbit(df, input_file_type)
    for i, row in df.iterrows():
        status = True
        LoggedError = ""

        # Test scenario for column "Name"
        tc_1_signal_name_test(row)

        # Test scenario for column "Message"
        tc_2_message_name_test(row)

        # Test scenario for column "Startbit"
        tc_3_startbit_test(row)

        # Test scenario for column "Length [Bit]"
        tc_4_length_bit_test(row)

        # Test scenario for column "Byte Order"
        tc_5_byte_order_test(row)

        # Test scenario for column "Value Type"
        if row["input_file"] == "dbc":
            tc_6_value_type_test(row)

        # Test scenario for column "Initial Value"
        tc_7_initial_value_test(row)

        # Test scenario for column "Factor"
        tc_8_factor_test(row)

        # Test scenario for column "Offset"
        tc_9_offset_test(row)

        # Test scenario for column "Minimum"
        tc_10_minimum_value_test(row)

        # Test scenario for column "Maximum"
        tc_11_maximum_value_test(row)

        # Test scenario for column "Value Table"
        if row["input_file"] == "dbc":
            tc_12_value_table_test(row)

        # Test scenario for column "Message ID"
        tc_13_message_id_test(row)

        # Test scenario for column "Cycle Time [ms]"
        tc_14_cycle_time_test(row)

        # Test scenario for column "texttable"
        tc_15_texttable_test(row)

        # Test scenario for column "texttable values"
        tc_16_texttable_values_test(row)

        # Test scenario for column "max_value"
        tc_17_max_value_test(row)

        # Test scenario for column "dlc"
        if row["input_file"] == "dbc":
            tc_18_dlc_test(row)

        # Test scenario for column "sender"
        tc_19_sender_test(row)

        # Test scenario for column "PDU"
        if row["input_file"] == "arxml":
            tc_20_pdu_test(row)

        # Test scenario for CRC Signal
        tc_21_profile_test(row)

        # Test scenario for Payload_PDU_Type
        tc_23_payload_pdu_type_test(row)

        # Updating "Status" and "Remarks" columns
        # df.loc[row_indexer, "col"]
        if status and df.loc[i, "Status"] == None:
            df.loc[i, "Status"] = "Pass"
            df.loc[i, "Remarks"] = ""
        else:
            df.loc[i, "Status"] = "Fail"
            if df.loc[i, "Remarks"] != None:
                df.loc[i, "Remarks"] += LoggedError
            else:
                df.loc[i, "Remarks"] = LoggedError


    # Deleting columns
    # del df["variant"]
    # del df["gw"]
    # del df["a_variant_veh_0"]
    # del df["a_variant_veh_1"]
    # del df["b_variant_veh_0"]
    # del df["b_variant_veh_1"]

    return df

#ETHERNET functions
def tc_1_eth_service_name_test(row):
    """
    Description: Function checks service name is empty or not.

    Args:
        row (Pandas Series): Series contains all column values of individual row

    Returns:

    """
    # Case 0: column "Service" i.e. service name should not be empty
    if cell_empty(row["Service"], "Service"):
        return
    
def tc_2_eth_service_test(row):
    """
    Description: Function checks for each service Service ID, Instance ID
    Major version, Minor version exists or not.

    Args:
        row (Pandas Series): Series contains all column values of individual row

    Returns:

    """
    cell_empty(row["Service ID"], "Service ID")
    cell_empty(row["Major version"], "Major version")
    cell_empty(row["Minor version"], "Minor version")
    cell_empty(row["Instance ID"], "Instance ID")

def tc_3_eth_method_test(row):
    """
    Description: Function checks for each row Member Type column is empty or not.
    If Member Type == method it checks Member Id, Member, Parameter and Parametr Type columns are not empty.

    Args:
        row (Pandas Series): Series contains all column values of individual row

    Returns:

    """
    global status, LoggedError
    cell_empty(row["Member Type"], "Member Type")
    if (row["Member Type"] == "method"):
        cell_empty(row["Member ID"], "Member ID")
        cell_empty(row["Member"], "Member")
        cell_empty(row["Parameter"], "Parameter")
        cell_empty(row["Parameter Type"], "Parameter Type")
        if empty_cell(row["Member ID"]) is False:
            if type(row["Member ID"]) != int:
                status = status and False
                LoggedError += "Member ID should be integer \n"
            else:
                status = status and True

def tc_4_eth_event_test(row):
    """
    Description: Function checks for Member Type column If Member Type = event,
    It checks Event Group and Event Id columns are not empty.

    Args:
        row (Pandas Series): Series contains all column values of individual row

    Returns:

    """
    global status, LoggedError
    if (row["Member Type"] == "event"):
        cell_empty(row["Event Group"], "Event Group")
        cell_empty(row["Member ID"], "Member ID")
        member_id = row['Member ID']
        member = row['Member']
        if member_id not in member_dict:
            member_dict[member_id] = set()
        if member in member_dict[member_id]:
            status = status and False
            LoggedError += "Event Group has repeated Member \n"
        else:
            status = status and True
            member_dict[member_id].add(member)


def tc_5_eth_minimum_maximum_test(row):
    """
    Description: Function checks Maximum and Minimum columns are not empty.
    and also checks Maximum is > Minimum

    Args:
        row (Pandas Series): Series contains all column values of individual row

    Returns:

    """
    global status, LoggedError
    
    if (row["Member Type"] == "method"):
        cell_empty(row["Minimum"], "Minimum")
        cell_empty(row["Maximum"], "Maximum")
        if ((empty_cell(row["Minimum"]) is False) and (empty_cell(row["Minimum"]) is False)):
            if (row["Maximum"] < row["Minimum"]):
                status = status and False
                LoggedError += "Maximum value is less than or equal to Minimum value \n"
            else:
                status = status and True

def tc_6_eth_sd_type_test(row):
    """
    Description: Function checks SD Type columns is not empty.

    Args:
        row (Pandas Series): Series contains all column values of individual row

    Returns:

    """
    cell_empty(row["SD Type"], "SD Type")

def test_sysvar_database_eth_someip(df):
    """
    Description: Function takes reference of SysVarDatabaseETH sheet and evaluates each row based on certain conditions.
                 Generates report Test_Database.xlsx that consist of pass/fail status of each signal and statistics i.e.
                 % passed signal etc.

    Args:
        df (DataFrame): Dataframe of SysVarDatabase sheet of the Autosar_Gen_Database.xlsx

    Returns:

    """

    global status, LoggedError, member_dict
    member_dict = {}
    # signal_pass, signal_fail = 0, 0
    # rows = df.shape[0]  # No of rows
    # col = df.shape[1]  # No of columns
    df["Status"] = None  # Add new column to excel
    df["Remarks"] = None  # Add new column to excel
    for i, row in df.iterrows():
        status = True
        LoggedError = ""

        # Test scenario for column "Service"
        tc_1_eth_service_name_test(row)

        # Test scenario for "Service" should has service_id, instance_id, minor_version, major_version
        tc_2_eth_service_test(row)

        #Test Scenario to check member type is method if true it should has Method Id and Method Name
        tc_3_eth_method_test(row)

        #Test Scenario to check If member type is event then it shold has event group
        tc_4_eth_event_test(row)

        #Test Scenario to check Minimum and Maximum columns are empty and Maximum is > Minimum
        tc_5_eth_minimum_maximum_test(row)

        #Test Scenario for SD Type column
        tc_6_eth_sd_type_test(row)

        # if status:
        #     df["Status"][i] = "Pass"
        #     df["Remarks"][i] = ""
        # else:
        #     df["Status"][i] = "Fail"
        #     df["Remarks"][i] = LoggedError

        if status and df.loc[i, "Status"] == None:
            df.loc[i, "Status"] = "Pass"
            df.loc[i, "Remarks"] = ""
        else:
            df.loc[i, "Status"] = "Fail"
            if df.loc[i, "Remarks"] != None:
                df.loc[i, "Remarks"] += LoggedError
            else:
                df.loc[i, "Remarks"] = LoggedError

    return df

def test_sysvar_database_eth_eth_pdu(df):
    """
    Description: Function takes reference of SysVarDatabaseETH sheet and evaluates each row based on certain conditions.
                 Generates report Test_Database.xlsx that consist of pass/fail status of each signal and statistics i.e.
                 % passed signal etc.

    Args:
        df (DataFrame): Dataframe of SysVarDatabase sheet of the Autosar_Gen_Database.xlsx

    Returns:

    """
    global status, LoggedError, member_dict
    member_dict = {}
    # signal_pass, signal_fail = 0, 0
    # rows = df.shape[0]  # No of rows
    # col = df.shape[1]  # No of columns
    df["Status"] = None  # Add new column to excel
    df["Remarks"] = None  # Add new column to excel
    
    for i, row in df.iterrows():
        status = True
        LoggedError = ""

        # Test scenario for column "Startbit"
        tc_3_startbit_test(row)

        # Test scenario for column "Length [Bit]"
        tc_4_length_bit_test(row)

        # Test scenario for column "Byte Order"
        # tc_5_byte_order_test(row)

        # Test scenario for column "Initial Value"
        tc_7_initial_value_test(row)

        # Test scenario for column "Factor"
        tc_8_factor_test(row)

        # Test scenario for column "Offset"
        tc_9_offset_test(row)

        # Test scenario for column "Minimum"
        tc_10_minimum_value_test(row)

        # Test scenario for column "Maximum"
        tc_11_maximum_value_test(row)

        # Test scenario for column "texttable"
        tc_15_texttable_test(row)

        # Test scenario for column "texttable values"
        tc_16_texttable_values_test(row)

        # Test scenario for column "max_value"
        tc_17_max_value_test(row)

        # Test scenario for column "sender"
        tc_19_eth_sender_test(row)

        # Test scenario for Payload_PDU_Type
        tc_23_payload_pdu_type_test(row)

        if status and df.loc[i, "Status"] == None:
            df.loc[i, "Status"] = "Pass"
            df.loc[i, "Remarks"] = ""
        else:
            df.loc[i, "Status"] = "Fail"
            if df.loc[i, "Remarks"] != None:
                df.loc[i, "Remarks"] += LoggedError
            else:
                df.loc[i, "Remarks"] = LoggedError
    return df

def write_to_excel(script_path, df, test_database_info_file_path, database_type):
    total_signals = len(df)
    pass_df = df[(df['Status'] == "Pass")]
    fail_df = df[(df['Status'] == "Fail")]
    signal_pass = len(pass_df)
    signal_fail = len(fail_df)
    
    pass_in_per = str(round((signal_pass / (signal_pass + signal_fail) * 100), 2)) + "%"
    script_path = os.path.dirname(os.path.abspath(__file__))
    test_database_path = script_path + f"/../../../../CustomerPrj/Restbus/Test_Database_{database_type}.xlsx"
    #df.to_excel(test_database_path, index=False)
    with pd.ExcelWriter(test_database_path, engine='openpyxl', mode='w') as writer:
        df.to_excel(writer, sheet_name='Test_SysVarDatabase', index=False)

    wb = load_workbook(test_database_path)
    wb.create_sheet("SysVarDatabase Statistics")
    ws1 = wb["SysVarDatabase Statistics"]
    ws1["A1"] = "Total signals"
    ws1["B1"] = total_signals
    ws1["A2"] = "Signals passed"
    ws1["B2"] = signal_pass
    ws1["A3"] = "Signals failed"
    ws1["B3"] = signal_fail
    ws1["A4"] = "% Passed"
    pass_in_per = str(round((signal_pass / (signal_pass + signal_fail) * 100), 2)) + "%"
    ws1["B4"] = pass_in_per

    wb.save(test_database_path)
    with open(test_database_info_file_path, 'a') as database_log:
        logger.info(f"Total number of signals in {database_type} Database = {signal_pass + signal_fail}")
        database_log.write(f"Total number of signals in {database_type} Database = {signal_pass + signal_fail}\n")
        logger.info(f"Number of signals passed in {database_type} Database = {signal_pass}")
        database_log.write(f"Number of signals passed in {database_type} Database = {signal_pass}\n")
        logger.info(f"Number of signals Failed in {database_type} Database= {signal_fail}")
        database_log.write(f"Number of signals Failed in {database_type} Database = {signal_fail}\n")
        logger.info(f"Passed % = {pass_in_per}")
        database_log.write(f"Passed % in {database_type} Database = {pass_in_per}\n")
    return

def main():
    try:
        sysvar_database_sheet = "SysVarDatabase"
        sysvar_database_eth_sheet = "SysVarDatabaseETH"
        script_path = os.path.dirname(os.path.abspath(__file__))
        autosar_path = script_path + r'\..\..\..\..\CustomerPrj\Restbus\Autosar_Gen_Database.xlsx'
        wb = load_workbook(autosar_path, read_only=True)
        if sysvar_database_sheet in wb.sheetnames:
            df_sys = pd.read_excel(autosar_path, sheet_name=sysvar_database_sheet, dtype=object)
            input_file_type = df_sys['input_file'].dropna().drop_duplicates().tolist()
            input_file_type = input_file_type[0]
            # if input_file_type == "arxml":
            #     sys_df = df_sys[df_sys['PDU_Type'] == 'I-SIGNAL-I-PDU']
            # else:
            #     sys_df = df_sys

        if sysvar_database_eth_sheet in wb.sheetnames:
            df_sys_eth = pd.read_excel(autosar_path, sheet_name=sysvar_database_eth_sheet, dtype=object)
        
        can_dataframes = []
        eth_dataframes = []
        eth_pdu_dataframes = []
        log_file_path = os.path.join(os.path.dirname(autosar_path), 'test_database_variant_columns_log.txt')
        test_database_info_file_path = os.path.join(os.path.dirname(autosar_path), 'test_database_info.txt')
        if os.path.exists(log_file_path):
            os.remove(log_file_path)
        if os.path.exists(test_database_info_file_path):
            os.remove(test_database_info_file_path)
        if (sysvar_database_sheet in wb.sheetnames):
            logger.info(f"###### STARTED 'Test Database' for CAN ######")
            test_varient_cols(df_sys, log_file_path, db_type = "CAN")
            network_names = df_sys['network_name'].drop_duplicates().tolist()
            for network in network_names:
                network_df = df_sys[(df_sys['network_name'] == network)]
                grouped_data = network_df.groupby('sender')
                
                for sender, sender_df in grouped_data:
                    network_sender_df = f"{network}_{sender}_df"
                    network_sender_df = test_sysvar_database(sender_df, input_file_type)
                    can_dataframes.append(network_sender_df)
            merged_df_can = pd.concat(can_dataframes, ignore_index=True)
            write_to_excel(script_path, merged_df_can, test_database_info_file_path, database_type = "CAN")
            logger.info(f"###### FINISHED 'Test Database' for CAN ######")

        # Ethernet_Someip
        if (sysvar_database_eth_sheet in wb.sheetnames) and not(df_sys_eth.shape[0] <= 1):
            # logger.info(f"###### STARTED 'Test Database' for ETHERNET ######")
            #test_varient_cols(df_sys_eth, log_file_path, db_type = "ETHERNET")
            logger.info(f"###### STARTED 'Test Database' for ETHERNET_SOMEIP ######")
            some_ip_df = df_sys_eth[(df_sys_eth['Autosar Type'] == "ETH_SOMEIP")]
            network_names = some_ip_df['Network Name'].drop_duplicates().tolist()
            for network in network_names:
                network_df = some_ip_df[(some_ip_df['Network Name'] == network)]
                grouped_data = network_df.groupby('Node Name')
                for sender, sender_df in grouped_data:
                    network_sender_df = f"{network}_{sender}_df"
                    network_sender_df = test_sysvar_database_eth_someip(sender_df)
                    eth_dataframes.append(network_sender_df)
            merged_df_eth = pd.concat(eth_dataframes, ignore_index=True)
            write_to_excel(script_path, merged_df_eth, test_database_info_file_path, database_type = "ETH_SOMEIP")
            logger.info(f"###### FINISHED 'Test Database' for ETHERNET_SOMEIP ######")

        # Ethernet_ETH_PDU
        if (sysvar_database_eth_sheet in wb.sheetnames) and not(df_sys_eth.shape[0] <= 1):
            # logger.info(f"###### STARTED 'Test Database' for ETHERNET ######")
            #test_varient_cols(df_sys_eth, log_file_path, db_type = "ETHERNET")
            logger.info(f"###### STARTED 'Test Database' for ETHERNET_ETH_PDU ######")
            eth_pdu_df = df_sys_eth[(df_sys_eth['Autosar Type'] == "ETH_PDU")]
            network_names = eth_pdu_df['Network Name'].drop_duplicates().tolist()
            for network in network_names:
                network_df = eth_pdu_df[(eth_pdu_df['Network Name'] == network)]
                grouped_data = network_df.groupby('Node Name')
                for sender, sender_df in grouped_data:
                    network_sender_df = f"{network}_{sender}_df"
                    
                    network_sender_df = test_sysvar_database_eth_eth_pdu(sender_df)
                    eth_pdu_dataframes.append(network_sender_df)
            merged_df_eth_pdu = pd.concat(eth_pdu_dataframes, ignore_index=True)
            write_to_excel(script_path, merged_df_eth_pdu, test_database_info_file_path, database_type = "ETH_PDU")
            logger.info(f"###### FINISHED 'Test Database' for ETHERNET_ETH_PDU ######")
    except Exception as e:
        logger.warning(f"Error occured during testing Database : {e}")
        raise Exception(e)

if __name__ == "__main__":
    main()
             