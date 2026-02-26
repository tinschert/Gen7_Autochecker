# -*- coding: utf-8 -*-
# @file create_nodes.py
# @author ADAS_HIL_TEAM
# @date 08-21-2023

##################################################################
# C O P Y R I G H T S
# ----------------------------------------------------------------
# Copyright (c) 2023 by Robert Bosch GmbH. All rights reserved.

# The reproduction, distribution and utilization of this file as
# well as the communication of its contents to others without express
# authorization is prohibited. Offenders will be held liable for the
# payment of damages. All rights reserved in the event of the grant
# of a patent, utility model or design.

##################################################################


import os, sys, traceback
import re
import pandas as pd
try:
    pd.set_option('future.no_silent_downcasting', True)
except:
    pass
import numpy as np
from math import ceil
from openpyxl import load_workbook

try:
    from Rbs_Scripts.create_OsekTp import * #OSEK_TP
except ImportError:
    sys.path.append(os.path.dirname(os.path.abspath(__file__)))
    from create_OsekTp import * #OSEK_TP

try:
    from Rbs_Scripts.create_Opaque_byte_order import * #Opaque_Byte_Order
except ImportError:
    sys.path.append(os.path.dirname(os.path.abspath(__file__)))
    from create_Opaque_byte_order import * #Opaque_Byte_Order
    
try:
    from Rbs_Scripts.update_sysvartab import getMappingData
except ImportError:
    sys.path.append(os.path.dirname(os.path.abspath(__file__)))
    from update_sysvartab import getMappingData
try:
    sys.path.append(os.getcwd() + r"\..\..\CustomerPrj\Restbus\Scripts")
    from pattern_matching import *
except ImportError:
    sys.path.append(os.getcwd() + r"\..\..\..\..\CustomerPrj\Restbus\Scripts")
    from pattern_matching import *

try:
    from Control.logging_config import logger
except ImportError:
    sys.path.append(os.path.dirname(os.path.abspath(__file__)) + r"\..\Control")
    from logging_config import logger

#END IMPORTS

# allow longer names up to 100 characters
pd.options.display.max_colwidth = 100
pd.options.display.colheader_justify = 'left'

# can_column_names = ("Signal",
#                     "Signal_Group",
#                     "PDU",
#                     "PDU_Type",
#                     "Payload_PDU_Type",
#                     "Payload_PDU",
#                     "Selector_Field_Signal",
#                     "Message",
#                     "Message ID",
#                     "Startbit",
#                     "PDU_Length [Byte]",
#                     "Signal_Length [Bit]",
#                     "Initial Value",
#                     "max_value",
#                     "Cycle Time [ms]",
#                     "texttable",
#                     "texttable values",
#                     "Value Type",
#                     "Comment",
#                     "dlc",
#                     "variant",
#                     "Value Table",
#                     "EndToEndProtection",
#                     "Byte Order",
#                     "Factor",
#                     "Offset",
#                     "Minimum",
#                     "Maximum",
#                     "Unit",
#                     'Block_size', 
#                     'Address_formate', 
#                     'Padding_active', 
#                     'STMin',
#                     'MAXFC_wait'
#                     )


skip_pdu_type_list = ['NM-PDU', 'GENERAL-PURPOSE-PDU', 'GENERAL-PURPOSE-I-PDU', 'USER-DEFINED-PDU', 'USER-DEFINED-I-PDU']

flt_evt_node = ['']
flt_evt_cnt = 0  # FLT_EVT

def check_df_with_value(df, col_index, value_to_check):
    if df.empty:
        return False
    return value_to_check in df.iloc[:, col_index].values
    

def preprocess_node_df(df):
    """
    Preprocess node_df and returns updated node_df for Arxml

    Args:
      df (dataframe): dataframe to be converted

    Returns:
        converted df

    """
    #securedpdu_copy_col_index = [col_msg, col_msg_id, col_cycle_time, col_endtoendprotection]
    if input_file!="dbc":
        df = df.fillna('')
        for i, row in df.iterrows():
            if "NM-PDU" == row[col_pdu_type]:
                df = df.drop(i)
                continue
            if row[col_pdu_type] in skip_pdu_type_list:
                continue
            pdu = row[col_pdu]
            if row[col_payload_pdu_type] == 'N-PDU':
                continue
            # if check_PDU_in_SecuredPdu(pdu, securedPdu_dict):  # Checking I-SIGNAL-I-PDU -> PDU mapped to SECURED-I-PDU - > PDU
            #     for msg, values in securedPdu_dict.items():
            #         if values[1] == pdu:
            #             for index in securedpdu_copy_col_index:
            #                 df.loc[i,index] = values[2][index]
            #                 #df[index][i] = values[2][index]
            #                 # print(f'preprocess_node_df::check_PDU_in_SecuredPdu::i = {i}')

            if not(is_empty_or_spaces(row[col_byte_order])) and ("moto" in row[col_byte_order].lower()):  # converts Motorola StartBit
                #print(f"{row[col_pdu]}:")
                #print(f"{row[col_name]}: Length-> {row[col_signal_length_bit]} ,Before -> {row[col_startbit]}")
                df.loc[i, col_startbit] = findStartBit(row[col_startbit], row[col_signal_length_bit])
                #df[col_startbit][i] = findStartBit(row[col_startbit], row[col_length])
                #print(f"After -> {df[col_startbit][i]}")
                #print("-"*20)
    return df

def extractHilCtrl(wb):
    """
    extract hil_ctrl mapping to node

    Args:
      wb (workbook): excel workbook

    Returns:
        dict_hilctrl

    """
    mapping_sheets = ["DBCmapping","ArxmlMapping","ETHArxmlMapping"]
    dict_hilctrl = {}
    try:
        for mapping_sheet in mapping_sheets:
            if mapping_sheet not in wb.sheetnames:
                continue
            df = pd.DataFrame(wb[mapping_sheet].values)
            df.columns = df.iloc[0]
            df = df.drop(index=0)
            df = df.replace(np.nan, '', regex=True)
            key1 = mapping_sheet.lower().replace("mapping", "")   # dbc/arxml
            key1 = "dbc" if "dbc" in key1.lower() else "arxml"
            if key1 not in dict_hilctrl.keys():
                dict_hilctrl[key1] = {}
            for i, row in df.iterrows():
                key2 = row["canoe_ecu_node"].strip()
                if key1 not in dict_hilctrl.keys():
                    dict_hilctrl[key1] = {key2:row["hil_ctrl"].strip()}
                else:
                    if key2 not in dict_hilctrl[key1].keys():
                        dict_hilctrl[key1][key2] = row["hil_ctrl"].strip()
        return dict_hilctrl
    except Exception as e:
        logger.error(f"Error while extracting hil_ctrl node names from dbcmapping or arxml mapping -> {e}")

def getColNamespace(entity):
    """
    categarise node based on entity (message ,pdu or signal group)

    Args:
      entity: 

    Returns:column  based on node entity

    """
    if "message" in entity:
        return col_msg
    elif "pdu" in entity:
        return col_pdu
    elif "group" in entity:
        return col_siggrp
    else:
        return col_msg


# read worksheet 'sheet_name' from workbook 'wb'
def generate_dataframe(wb, sheet_name):
    """
    create pandas dataframe from workbook and declare all column index

    Args:
      wb (workbook): excel workbook
      sheet_name (str): sheet name

    """
    global node, node_name, bus_name, bus_type_from_sheet_name, bus_type, input_file, node_df, num_of_rows, hil_ctrl_node_dict, is_public_can, network_name, node_entity
    global col_name,col_siggrp,col_pdu,col_pdu_type,col_payload_pdu_type,col_payload_pdu,col_selector_field_signal,col_msg,col_msg_id,\
     col_startbit,col_pdu_length_byte,col_signal_length_bit,col_initial_value,col_max_value,col_cycle_time,col_texttable,col_texttable_values,\
     col_value_type,col_comment,col_dlc,col_variant,col_value_table,col_endtoendprotection,col_byte_order,col_factor,col_offset,col_minimum,\
     col_maximum,col_unit,col_cantp_pdu_type,col_block_size,col_address_formate,col_padding_active,col_stmin,col_maxfc_wait, \
     col_header_id, col_payload_pdu_length, col_signal_base_type, col_transmission_mode,col_payload_pdu_dlc

    global col_namespace
    global current_sheet, block_message_list
    global var_IL

    current_sheet = sheet_name
    block_message_list = []
    script_path = "\\".join(os.path.dirname(os.path.abspath(__file__)).split("\\")[-4:])
    script_path = "// Autogenerated by -> " + script_path + r"\create_nodes.py"
    node = ['/*@!Encoding:1252*/', script_path]

    is_public_can = True if "public" in dict_network_info[sheet_name].lower() else False
    node_entity = dict_entity[sheet_name]
    network_name = dict_network_name[sheet_name]
    ws1 = wb[sheet_name]
    name_list = sheet_name.split('_')
    input_file = name_list[-1]
    node_name = name_list[-2]
    bus_name = name_list[-3]
    bus_type_from_sheet_name = name_list[-4] #bus type for node name(.can) & for cus_bus
    bus_type = dict_network_type[sheet_name] #bus type from network category column: CAN/CANFD/ETH
    # print(node_name, bus_name, bus_type)
    try:
        hil_ctrl_node_dict = dict_hilctrl["arxml"] if "arxml" in input_file.lower() else dict_hilctrl["dbc"]
    except Exception as e:
        logger.error(f"Error while reading hil ctrl data in dbc mapping -> {e}")

    var_IL = get_canoe_IL(bus_type, input_file)

    # create the different dataframes
    node_df = pd.DataFrame(ws1.values)
    node_df = node_df.replace(np.nan, '', regex=True)

    # get column names
    col_names = list(node_df.iloc[0])
    # get column index
    try:
        col_index = [col_names.index(col) for col in col_names]
    except:
        logger.error("Column names in Excel are different from the Script:")
        logger.error(';'.join(col_names))
        return False
    try:
        [col_name,col_siggrp,col_pdu,col_pdu_type,col_payload_pdu_type,col_payload_pdu,col_selector_field_signal,col_msg,col_msg_id,col_header_id,
        col_startbit,col_pdu_length_byte,col_signal_length_bit, col_payload_pdu_length, col_signal_base_type, col_initial_value,col_max_value,col_transmission_mode, col_cycle_time,col_texttable,col_texttable_values,
        col_value_type,col_comment,col_dlc,col_payload_pdu_dlc,col_variant,col_value_table,col_endtoendprotection,col_byte_order,col_factor,col_offset,col_minimum,
        col_maximum,col_unit,col_cantp_pdu_type,col_block_size,col_address_formate,col_padding_active,col_stmin,col_maxfc_wait] = col_index
    except Exception as e:
        logger.error(f"New column is added or removed from excel, please check and update thw col index in script -> {e}")
        return False

    node_df_Heading = node_df.head(1)  # Sort colum Heading(Example ::'Name','group','pdu','Message','Multiplexing/Group','Startbit','Length [Bit]',...)
    node_df = node_df.iloc[1:]

    col_namespace = getColNamespace(node_entity)
    
    if sheet_name in dict_block_messages.keys():
        blk_list_ids = dict_block_messages[sheet_name]
        block_msg_df = node_df[node_df[col_msg_id].isin(blk_list_ids)]
        block_message_list = list(block_msg_df[col_namespace].unique())
        node_df = node_df[~node_df[col_msg_id].isin(blk_list_ids)]
    elif sheet_name in dict_pass_messages.keys():
        pass_list_ids = dict_pass_messages[sheet_name]
        block_msg_df = node_df[~node_df[col_msg_id].isin(pass_list_ids)]
        block_message_list = list(block_msg_df[col_namespace].unique())
        node_df = node_df[node_df[col_msg_id].isin(pass_list_ids)]

    # diag filer
    node_df = node_df[node_df[col_namespace].str.contains(diag_msg_pattern, flags=re.IGNORECASE ,regex=True) == False]  # Filter out Diag
    #node_df = node_df[node_df[col_namespace].str.contains(stbm_msg_pattern, regex=True) == False]  # Filter out STBM

    #check_df_integrity(node_df)

      # Remove colum Heading(Example ::'Name','group','pdu','Message','Multiplexing/Group','Startbit','Length [Bit]',...)
    sort_by_column_names = sort_column_by(bus_type, input_file)  # get sort by column names based on bus:Eth or CAN

    node_df = df_sort_asc(node_df, col_names, sort_by_column_names)  # Sort in ascending df by column names
    node_df = pd.concat([node_df_Heading, node_df], ignore_index=True)  # Restore Heading
    num_of_rows = node_df.shape[0]
    
    # npdu_dict = map_npdu_to_ipdu_dict(node_df, col_msg, col_pdu, col_pdu_type, col_payload_pdu)
    # securedPdu_dict = map_securedpdu_to_ipdu_dict(node_df, col_msg, col_pdu, col_pdu_type, col_payload_pdu)

    node_df = preprocess_node_df(node_df)
    num_of_rows = node_df.shape[0]
    logger.info(f"Signal count: {num_of_rows - 1}")


def generate_includes():
    """
    generate includes section based on IL usage
    """
    # includes Section
    node.append('includes')
    node.append('{')
    node.extend(get_includes(var_IL))

    filtered_df = node_df.iloc[1:]
    for pdu_type, pdu_type_df in filtered_df.groupby(col_pdu_type):
        if ('SEC' in pdu_type.upper()):
            sec_pdu_list = list(pdu_type_df[col_pdu].unique())
            node.extend(secoc_includes(var_IL,sec_pdu_list))
            break

    osektp_list = []
    for pdu_type, pdu_type_df in filtered_df.groupby(col_pdu_type):
        if ('N-PDU' in pdu_type.upper()):
            for pdu, pdu_df in pdu_type_df.groupby(col_pdu):
                if not(is_empty_or_spaces(pdu)):
                    if check_df_with_value(pdu_df, col_cantp_pdu_type, 'Data_pdu'):
                        osektp_list.extend(generate_osektp_includes_can(network_name, node_name, pdu))   # defined in create_OsekTp.py
                    elif check_df_with_value(pdu_df, col_cantp_pdu_type, 'Flowcontrol_pdu'):
                        osektp_list.extend(generate_osektp_includes_can(network_name, node_name, pdu))   # defined in create_OsekTp.py
    if osektp_list:
        node.append(r'  #include "OSEKTP\OSEKTP.cin"  //common definations for osektp')
        #Specific implementation for radars needed in ADAS HIL
        #Restbus sending must not happen from sysvar but directly from radar model dlls
        #Conditionnal compilation should prevent any significant change on RBS only
        #Including the radar dlls
        if ("Radar" in node_name):
            node.append("  #if ADAS_HIL")
            node.append(r'    #pragma library("..\..\..\Platform\Classe\DLL\star_model\capldll.dll")')
            node.append(r'    #pragma library("..\..\..\Platform\Classe\DLL\clara_model\Release64\clara_model.dll")')
            node.append(r'    #pragma library("..\..\..\Platform\Classe\DLL\Release64\RoadObj.dll")')
            node.append("  #endif")
        node.extend(osektp_list)

    for byte_order, byte_order_df in filtered_df.groupby(col_byte_order):
        if 'OPAQUE' in byte_order.upper():
            for pdu_byte_order, pdu_byte_order_df in byte_order_df.groupby(col_pdu):
                for signal_byte_order, signal_byte_order_df in pdu_byte_order_df.groupby(col_name):
                    node.append(f'  #include "Opaque_Byte_Order\\Opaque_{network_name}_{node_name}_{pdu_byte_order}_{signal_byte_order}.cin" /*Signal Opaque byte order handling*/')

    node.append('}')
    node.append('')
    logger.info('includes: generated for {0} '.format(node_name))


def generate_variables():
    """
    create node variables ::
        create crc ,timestamp and  counters(Alive and Block) variables  for each message/pdu 
        create npdu and secoc variables for each message/pdu based on messahe type
    """
    # Case create node variables
    global flt_evt_cnt
    message_count = 0
    Npdu_found_flag = True
    evt_cnt = 0
    node.append('variables {')
    part_msg = []  # Pdu or Msg
    part_ac = []
    part_bc = []
    part_ts = []
    part_crc = []
    part_secoc = []
    part_Npdu = []
    for i in range(num_of_rows):
        # print('we are on message i = ' + str(i))
        namespace = str(node_df[col_namespace].values[i])
        msg_name = str(node_df[col_msg].values[i])
        pdu_name = str(node_df[col_pdu].values[i])
        pdu_type= str(node_df[col_pdu_type].values[i])
        Payload_PDU_Type = str(node_df[col_payload_pdu_type].values[i])
        payload_pdu_name = str(node_df[col_payload_pdu].values[i])

        if ("I-SIGNAL-I-PDU" in pdu_type) and Payload_PDU_Type == "MULTIPLEXED-I-PDU":
            continue
        if "MULTIPLEXED-I-PDU" in pdu_type:
            part_msg.extend(generate_MsgType(bus_type, input_file, msg_name, pdu_name))
            continue
        if "SECURED-I-PDU" in pdu_type or pdu_type in skip_pdu_type_list:
            continue
        elif "N-PDU" in pdu_type:
            if str(node_df[col_cantp_pdu_type].values[i]) == 'Flowcontrol_pdu':
                if Npdu_found_flag:
                    node.append('  int i;')
                    node.append('  char sysvarElementName[256];')
                    Npdu_found_flag = False
                node.append('  char mSysVarNameSpace_{0}_{1}[50] = "{0}::{1}";'.format(network_name, payload_pdu_name))
            continue
            
##            if check_PDU_in_NPDU(namespace, npdu_dict):
##                if "I-SIGNAL" in pdu_type.upper():# Skip,If I-Signal pdu is mappded to N-pdu
##                    node.append('  int Token_FC_{0} = 0;'.format(namespace))
##                    continue

        cycletime = node_df[col_cycle_time].values[i]
        pdu_type = node_df[col_pdu_type].values[i]
        if match_pattern(namespace,stbm_msg_pattern):
            continue
        try:
            cycletime = float(cycletime)
        except:
            cycletime = 0
        if str(cycletime).endswith(r'.0'):
            cycletime = int(cycletime)

        if (i > 0) and (node_df[col_namespace].values[i - 1] != node_df[col_namespace].values[i]):
            # node.append('  message {0} Msg_{0};'.format(namespace))
            if str(node_df[col_payload_pdu_type].values[i]) == 'N-PDU':
                if "I-SIGNAL" in pdu_type.upper():# Skip,If I-Signal pdu is mappded to N-pdu
                    node.append('  int Token_FC_{0} = 0;'.format(namespace))
                    if namespace in pdu_skip:
                        node.append('  msTimer {0}_Timer;'.format(namespace))
                        node.append('  int count;')
                        node.append('  int i;')
                    continue
            flt_evt_pattern = match_flt_evt_pattern(namespace)
            if (flt_evt_pattern != '0x00'):  # Filter in all FLT_EVT message
                if (flt_evt_pattern != '0xFF'):  # Filter out FLT_EVT message from 'FD3'
                    flt_evt_cnt += 1
            else:
                if check_PDU_in_SecuredPdu(namespace, securedPdu_dict):
                    retrieved_msg = get_message_from_SecuredPdu(namespace, securedPdu_dict)
                    part_msg.extend(generate_MsgType(bus_type, input_file, retrieved_msg, pdu_name))  # Import from pattern_matching
                    # if ('SEC' in pdu_type.upper()):
                    # msg_name =
                    part_secoc.extend(generate_SecocVar(bus_type, input_file, namespace))  # Import from pattern_matching
                else:
                    if Payload_PDU_Type == "CONTAINER-I-PDU" and "I-SIGNAL-I-PDU" in pdu_type:
                        if (node_df[col_payload_pdu].values[i - 1] != node_df[col_payload_pdu].values[i]):
                            part_msg.extend(generate_MsgType(bus_type, input_file, msg_name, pdu_name))  # Import from pattern_matching
                    else:
                        part_msg.extend(generate_MsgType(bus_type, input_file, msg_name, pdu_name))  # Import from pattern_matching

                message_df = node_df
                message_df = message_df[(message_df[col_namespace] == namespace)]
                num_of_signals = message_df.shape[0]
                ac_cnt = 0  # Alive counter
                bc_cnt = 0  # Block counter
                ts_cnt = 0  # Time stamp
                crc_cnt = 0  # CRC

                for j in range(num_of_signals):
                    signal_name = str(message_df[col_name].values[j])
                    sig_length = int(message_df[col_signal_length_bit].values[j])
                    para_dict = {
                        "namespace": namespace,
                        "signal_name": signal_name,
                        "sig_length": sig_length,
                        "bus_type": bus_type,
                        "input_file": input_file
                    }
                    break_out = 0
                    if (break_out != 1):
                        for sig_pattern, sig_fun in ac_var_dict.items():
                            if (None != re.search(sig_pattern, signal_name, re.M | re.I)):  # Pattern :: Alive counter
                                part_ac = part_ac + sig_fun(para_dict)
                                ac_cnt += 1
                                break_out = 1
                                break
                    if (break_out != 1):
                        for sig_pattern, sig_fun in bc_var_dict.items():
                            if (None != re.search(sig_pattern, signal_name, re.M | re.I)):  # Pattern :: Block counter
                                part_bc = part_bc + sig_fun(para_dict)
                                bc_cnt += 1
                                break_out = 1
                                break
                    if (break_out != 1):
                        for sig_pattern, sig_fun in ts_var_dict.items():
                            if (None != re.search(sig_pattern, signal_name, re.M | re.I)):  # Pattern :: Time stamp
                                part_ts = part_ts + sig_fun(para_dict)
                                ts_cnt += 1
                                break_out = 1
                                break
                    if (break_out != 1):
                        for sig_pattern, sig_fun in crc_var_dict.items():
                            if (None != re.search(sig_pattern, signal_name, re.M | re.I)):  # Pattern :: CRC
                                part_crc = part_crc + sig_fun(para_dict)
                                crc_cnt += 1
                                break_out = 1
                                break

    node.extend(part_msg)
    node.append('')
    
    filtered_df = node_df.iloc[1:]
    for pdu_type, pdu_type_df in filtered_df.groupby(col_pdu_type):
        if ('SEC' in pdu_type.upper()):
            node.extend(part_secoc)
            node.append('')
            break

    if (current_sheet == rbs_sheet_name):  # Old:PSA_CAN_FD3_Rbs
        if (flt_evt_cnt > 0):
            node.extend(rbs_node_ack(bus_type, input_file))

    if (len(part_ac) > 0):
        node.extend(part_ac)
        node.append('')
    if (len(part_bc) > 0):
        node.extend(part_bc)
        node.append('')
    if (len(part_ts) > 0):
        node.extend(part_ts)
        node.append('')
    if (len(part_crc) > 0):
        node.extend(part_crc)
        node.append('')

    node.append('}')
    logger.info(f"Variables: generated with {str(message_count)} messages.")

def generate_onprestart():
    """
    on prestart:initialize and handle all IL control
    """
    node.append('')
    node.append('on preStart {')
    node.append('  long result;')
    if(var_IL=='AsrPDUIL'):
        node.append('  result=ARILControlInit ();//Initialization of AUTOSAR PDU IL,to prevent the IL autostart function.')
        node.append('  switch(result)')
        node.append('  {')
        node.append('    case  0    : write("ARILControlInit :: {0} :: No error."); break;'.format(node_name))
        node.append('    case -1    : write("ARILControlInit :: {0} :: IL is in a state that ignores the given request: State graph of the IL (Link)."); break;'.format(node_name))
        node.append('    case -2    : write("ARILControlInit :: {0} :: Allowed value range of a signal exceeded. (The exceeded value is transmitted to the bus nevertheless).");  break;'.format(node_name))
        node.append('    case -50   : write("ARILControlInit :: {0} :: Node layer is not active. Presumably it is deactivated in the node’s configuration dialog."); break;'.format(node_name))
        node.append('    case -51   : write("ARILControlInit :: {0} :: The node is connected to another bus like the CAN Bus."); break;'.format(node_name))
        node.append('    case -100  : write("ARILControlInit :: {0} :: Signal/message not found in DB, or not mapped to the node."); break;'.format(node_name))
        node.append('    case -101  : write("ARILControlInit :: {0} :: Maximum possible value range exceeded. (There is no transmission to the bus)."); break;'.format(node_name))
        node.append('    case -102  : write("ARILControlInit :: {0} :: A network management-, diagnostic, or transport protocol signal was set. IL declines such types of signals, because it is not intended to connect these layers to the IL."); break;'.format(node_name))
        node.append('    case -103  : write("ARILControlInit :: {0} :: Invalid value supplied (for all types of requests)."); break;'.format(node_name))
        node.append('    case -104  : write("ARILControlInit :: {0} :: General error for invalid requests that are not described furthermore."); break;'.format(node_name))
        node.append('   }')
    elif(var_IL=='CANoeIL'):
        node.append('  result=ILControlInit ();//Initialization of CANoe IL,to prevent the IL autostart function.')
        node.append('  switch(result)')
        node.append('  {')
        node.append('    case  0   : write("ILControlInit :: {0} :: No error."); break;'.format(node_name))
        node.append('    case -1   : write("ILControlInit :: {0} :: Momentary state of the IL does not permit this query."); break;'.format(node_name))
        node.append('    case -50  : write("ILControlInit :: {0} :: Nodelayer is inactive - possibly deactivated in the nodes configuration dialog.");  break;'.format(node_name))
        node.append('   }')
    else:
        node.append('')
    node.append('')
    node.append('  {0}_sim_Disable();'.format(node_name))
    node.append('  {0}_sim_SetCycleTimeOffset();'.format(node_name))
    node.append('}')
    
def generate_onstart(): 
    """
    on start:start IL simultion and configure all IL parameters
    """
    node.append('')
    node.append('on start {')
    node.append('')
    if(var_IL=='AsrPDUIL'):
        node.append('  ARILControlSimulationOn();//Starts the simulation of the IL. The IL is operational and started. The IL will send PDUs.')
        node.append('  {0}_sim_DisableMsg();'.format(node_name))
        node.append('  {0}_sim_Disable();'.format(node_name))
        node.append('  ARILControlStart ();//When the IL is started, then it is fully operating and sending.')
    elif(var_IL=='CANoeIL'):
        node.append('  ILControlSimulationOn();//Starts the simulation of the IL.')
        if (bus_type.upper() == 'CAN'):
            node.append('  {0}_sim_SetCANParameter();'.format(node_name))
        elif (bus_type.upper() == 'CANFD'):
            node.append('  {0}_sim_SetCANFDParameter();'.format(node_name))
        node.append('  ILControlStart ();//Cyclical sending starts; setting signals is now possible.')
    else:
        node.append('  //ILControlSimulationOn();//Starts the simulation of the IL.')
        if (bus_type.upper() == 'CAN'):
            node.append('  //{0}_sim_SetCANParameter();'.format(node_name))
        elif (bus_type.upper() == 'CANFD'):
            node.append('  //{0}_sim_SetCANFDParameter();'.format(node_name))
        node.append('  //No ILControlStart()')
    node.append('  {0}_sim_Reset();'.format(node_name))
    generate_OnStart_callTx()
    node.append('}')


def generate_onstart_ini_CANTP():
    filtered_df = node_df.iloc[1:]
    pdu_count = 0
    for pdu_type, pdu_type_df in filtered_df.groupby(col_pdu_type):
        if ('N-PDU' in pdu_type.upper()):
            node.append('')
            node.append('OnStart_init_CANTP_{0}()'.format(node_name))
            node.append('{')
            node.append('  long dbHandle; \n  dbHandle = CanTpGetDBConnection();')
            node.append('  if (dbHandle > 0) \n  { \n    CanTpCloseConnection(dbHandle); \n  }')
            node.append('}')
            break


def generate_onstop():
    """
    on stopMeasurement:Stop IL simultion
    """
    node.append('')
    node.append('on stopMeasurement {')
    node.append('')
    node.append('  {0}_sim_DisableMsg();'.format(node_name))
    node.append('  {0}_sim_Disable();'.format(node_name))
    if(var_IL=='AsrPDUIL'):
        node.append('  ARILControlStop();//Stops sending of PDUs.')
        node.append('  ARILControlSimulationOff();//Stops the simulation of the IL. The IL is not operational. All API function except function ARILControlSimulationOn will not work.')
    elif(var_IL=='CANoeIL'):
        node.append('  ILControlStop();//Cyclical sending is stopped; setting signals is now no longer possible.')
        node.append('  ILControlSimulationOff();//Stops the simulation of the IL. After that no other function to control the IL has an effect to the IL.')
    else:
        node.append('//No ILControlStop() and ILControlSimulationOff()')
    node.append('')
    node.append('}')


def generate_onsim():
    """ generate onsim"""
    if node_name in hil_ctrl_node_dict.keys() and hil_ctrl_node_dict[node_name] != "":
        nod_name = hil_ctrl_node_dict[node_name]
    else:
        nod_name = node_name
    filtered_df = node_df.iloc[1:]
    node.append('')
    node.append('on sysvar_update hil_ctrl::{0}'.format(nod_name))
    node.append('{')

    for pdu_type, pdu_type_df in filtered_df.groupby(col_pdu_type):
        if ('N-PDU' in pdu_type.upper()):
            for payload_pdu, payload_pdu_df in pdu_type_df.groupby(col_payload_pdu):
                if not (is_empty_or_spaces(payload_pdu)):
                    if check_df_with_value(payload_pdu_df, col_cantp_pdu_type, 'Flowcontrol_pdu') and payload_pdu in pdu_skip:
                            node.append('  CallTx_{0}_{1}_hil_ctrl_{2}();'.format(node_name, payload_pdu_df[col_pdu].values[0], nod_name))

    node.append('  {0}_sim_Reset();'.format(node_name))
    node.append('}')


def generate_onbus():
    """ generate onbus"""
    node.append('')
    node.append('on sysvar_update Cus_bus::bus_{0}_{1}_ON_OFF'.format(bus_type_from_sheet_name, bus_name))
    node.append('{')
    node.append('  {0}_sim_Reset();'.format(node_name))
    node.append('}')

def generate_customer_functions():
    """gets customer specific functions"""
    try:
        node.extend(get_customer_functions(network_name, node_name))
    except:
        return


def generate_onvariant():
    """ generate onvariant"""
    node.append('')
    node.append('on sysvar_update hil_ctrl::variant')
    node.append('{')
    node.append('  {0}_sim_Reset();'.format(node_name))
    node.append('}')


def generate_void_sim_Reset():
    """ generate void sim Reset"""
    if node_name in hil_ctrl_node_dict.keys() and hil_ctrl_node_dict[node_name] != "":
        nod_name = hil_ctrl_node_dict[node_name]
    else:
        nod_name = node_name
    node.append('')
    node.append('void {0}_sim_Reset()'.format(node_name))
    node.append('{')
    node.append('')
    node.append('  {0}_sim_DisableMsg();'.format(node_name))
    node.append('  {0}_sim_Disable();'.format(node_name))
    node.append('')

    #try to get customer specific conditions
    cus_condition = ''
    try:
        cus_condition = get_customer_reset_condition(network_name,node_name)
        if cus_condition!='':
            cus_condition = ' && ' + cus_condition
    except:
        pass

    node.append('  if ((@Cus_bus::bus_{0}_{1}_ON_OFF == 1) && (@hil_ctrl::{2} == 1){3})'.format(bus_type_from_sheet_name, bus_name, nod_name, cus_condition))
    node.append('  {')
    node.append('    switch(@hil_ctrl::variant)')
    node.append('    {')
    node.append('      case a_variant:{0}_sim_start_a_variant();break;'.format(node_name))
    node.append('      case b_variant:{0}_sim_start_b_variant();break;'.format(node_name))
    node.append('    }')
    node.append('    if (@hil_ctrl::hil_mode == @hil_ctrl::hil_mode::Restbus) ') 
    node.append('    {')
    node.append('       {0}_sim_EnableEventMsg();'.format(node_name))
    node.append('    }')
    node.append('    else')
    node.append('    {')
    node.append('       {0}_sim_DisableEventMsg();'.format(node_name))
    node.append('    }')
    node.append('  }')
    node.append('}')


def generate_void_sim_start():
    """
    based on variant, enable(set system variable to one) all cyclic message/pdu
    """
    variant_list =['a_variant','b_variant']
    for variant in variant_list:
        node.append('')
        node.append('void {0}_sim_start_{1}()'.format(node_name,variant))
        node.append('{')
        # Create initialization of timers and node variables
        message_count = 0
        for i in range(num_of_rows):
            # print('we are on voidstart i = ' + str(i))
            namespace = str(node_df[col_namespace].values[i])
            pdu_type = str(node_df[col_pdu_type].values[i])
            payload_pdu_type = str(node_df[col_payload_pdu_type].values[i])
            container_i_pdu_namespace = str(node_df[col_payload_pdu].values[i])
            # print(f'generate_void_sim_start::namespace = {namespace}, pdu_type = {pdu_type}')
            if "N-PDU" in pdu_type:
                if str(node_df[col_cantp_pdu_type].values[i]) == 'Flowcontrol_pdu':
                    generate_OnStart_callTx_FC(node_name,namespace)

            if match_pattern(namespace,stbm_msg_pattern) or "SECURED-I-PDU" in pdu_type or pdu_type in skip_pdu_type_list:
                continue
            cycletime = node_df[col_cycle_time].values[i]
            try:
                cycletime = float(cycletime)
            except:
                cycletime = 0
            if str(cycletime).endswith(r'.0'):
                cycletime = int(cycletime)

            if (node_df[col_variant].values[i] == variant or node_df[col_variant].values[i] == 'Common'):
                if (i > 0) and (node_df[col_namespace].values[i - 1] != node_df[col_namespace].values[i]):
                    if (cycletime not in [0, 0.0]):
                        if (match_flt_evt_pattern(namespace) == '0x00'):  # Filter out all FLT_EVT message
                            # node.appen-d('  settimer(Tmr_{0}, {1}.0);'.format(namespace, cycletime))
                            # print(f'generate_void_sim_start::match_flt_evt_pattern::namespace = {namespace}, pdu_type = {pdu_type}')
                            # if "SECURED-I-PDU" in pdu_type:
                            #     node.append('  @{1}::{0}::{0}_ON_OFF = 1;'.format(str(node_df[col_payload_pdu].values[i]), network_name))
                            # else:
                            node.append('  @{1}::{0}::{0}_ON_OFF = 1;'.format(namespace, network_name))
                        if (("CONTAINER-I-PDU" in payload_pdu_type) and (node_df[col_payload_pdu].values[i - 1] != node_df[col_payload_pdu].values[i])):
                            node.append('  @{1}::{0}::{0}_ON_OFF = 1;'.format(container_i_pdu_namespace, network_name))   # CONTAINER-I-PDU
                    else:
                        pass
                    message_count += 1
                    if payload_pdu_type == 'N-PDU':
                        node.append('  //@{1}::{0}::{0}_ON_OFF = 1;'.format(namespace,network_name))
                        message_count += 1

        node.append('}')
        logger.info(f"void :: vgenerate_sim_start_{str(variant)}: with {str(message_count)} messages.")

def generate_void_sim_Disable():
    """
    Disable(set system variable to zero) all cyclic message/pdu
    """
    node.append('')
    node.append('void {0}_sim_Disable()'.format(node_name))
    node.append('{')
    # Create turning off of timers and node variables
    message_count = 0
    for i in range(num_of_rows):
        # print('we are on voidstop i = ' + str(i))
        namespace = str(node_df[col_namespace].values[i])
        pdu_type = str(node_df[col_pdu_type].values[i])
        payload_pdu_type = str(node_df[col_payload_pdu_type].values[i])
        container_i_pdu_namespace = str(node_df[col_payload_pdu].values[i])
        #if namespace in pdu_list:
        if "N-PDU" in pdu_type or "SECURED-I-PDU" in pdu_type or pdu_type in skip_pdu_type_list:
            continue

        if match_pattern(namespace,stbm_msg_pattern):
            continue

        if ("I-SIGNAL-I-PDU" in pdu_type) and payload_pdu_type == "MULTIPLEXED-I-PDU":
            continue

        if (i > 0) and (node_df[col_namespace].values[i - 1] != node_df[col_namespace].values[i]):
            if (match_flt_evt_pattern(namespace) == '0x00'):  # Filter out all FLT_EVT message
                node.append('  @{1}::{0}::{0}_ON_OFF = 0;'.format(namespace,network_name))
                message_count += 1
            if (("CONTAINER-I-PDU" in payload_pdu_type) and (node_df[col_payload_pdu].values[i - 1] != node_df[col_payload_pdu].values[i])):
                node.append('  @{1}::{0}::{0}_ON_OFF = 0;'.format(container_i_pdu_namespace, network_name))

    node.append('}')
    logger.info(f"generated :: void_{str(node_name)}_sim_Disable with {str(message_count)} messages.")

def generate_void_sim_DisableMsg():
    """
    Disable all message/pdu
    """
    node.append('')
    node.append('void {0}_sim_DisableMsg()'.format(node_name))
    node.append('{')
    if(var_IL=='AsrPDUIL'):
        DisableMsg = 'ARILFaultInjectionDisablePDU'
    elif(var_IL=='CANoeIL'):
        DisableMsg = 'ILFaultInjectionDisableMsg'
    else:
        DisableMsg = 'ILFaultInjectionDisableMsg' #Default
    # Create turning off of timers and node variables
    message_count = 0
    for i in range(num_of_rows):
        # print('we are on voidstop i = ' + str(i))
        namespace = str(node_df[col_namespace].values[i])
        pdu_type= str(node_df[col_pdu_type].values[i])
        payload_pdu_type = str(node_df[col_payload_pdu_type].values[i])
##        if namespace in pdu_list:
        if "N-PDU" in pdu_type and str(node_df[col_cantp_pdu_type].values[i]) == 'Flowcontrol_pdu':
            node.append('//   if (gHandle_{0}_{1} > 0)'.format(node_name,namespace))
            node.append('//    {')
            node.append('//       CanTpCloseConnection(gHandle_{0}_{1});'.format(node_name,namespace))
            node.append('//    }')
            continue
        elif "N-PDU" in pdu_type or "SECURED-I-PDU" in pdu_type or pdu_type in skip_pdu_type_list:
            continue
        elif ("I-SIGNAL-I-PDU" in pdu_type) and payload_pdu_type == "MULTIPLEXED-I-PDU":
            continue

        if payload_pdu_type == 'N-PDU':
            continue

        if (i > 0) and (node_df[col_namespace].values[i - 1] != node_df[col_namespace].values[i]):
            if (match_flt_evt_pattern(namespace) == '0x00'):  # Filter out all FLT_EVT message
                node.append('   {0}({1}::{2});'.format(DisableMsg,network_name,namespace))
                message_count += 1

    for namesp in block_message_list:
        node.append('   {0}({1}::{2});'.format(DisableMsg, network_name, namesp))
    node.append('}')
    logger.info(f"void: generate_void_sim_DisableMsg with {str(message_count)} messages.")
    
def generate_void_sim_EventMsg():
    """
    Disable and Enable all event based message/pdu
    """
    if node_name in hil_ctrl_node_dict.keys() and hil_ctrl_node_dict[node_name] != "":
        nod_name = hil_ctrl_node_dict[node_name]
    else:
        nod_name = node_name
    node.append('on sysvar hil_ctrl::hil_mode')
    node.append('{')
    node.append('  {0}_sim_Reset();'.format(node_name))
    node.append('}')
    node.append('')
    node.append('void {0}_sim_DisableEventMsg()'.format(node_name))
    node.append('{')
    if(var_IL=='AsrPDUIL'):
        DisableMsg = 'ARILFaultInjectionDisablePDU'
        EnableMsg  = 'ARILFaultInjectionEnablePDU'
    elif(var_IL=='CANoeIL'):
        DisableMsg = 'ILFaultInjectionDisableMsg'
        EnableMsg  = 'ILFaultInjectionEnableMsg'
    else:#Default
        DisableMsg = 'ILFaultInjectionDisableMsg' 
        EnableMsg  = 'ILFaultInjectionEnableMsg'
    # Create turning off of timers and node variables
    message_count = 0
    for i in range(num_of_rows):
        # print('we are on voidstop i = ' + str(i))
        namespace = str(node_df[col_namespace].values[i])
        pdu_type= str(node_df[col_pdu_type].values[i])
        payload_pdu_type = str(node_df[col_payload_pdu_type].values[i])
##        if namespace in pdu_list:
        if "N-PDU" in pdu_type or "SECURED-I-PDU" in pdu_type or pdu_type in skip_pdu_type_list:
            continue
        elif ("I-SIGNAL-I-PDU" in pdu_type) and payload_pdu_type == "MULTIPLEXED-I-PDU":
            continue
##            continue
        if payload_pdu_type in ['N-PDU', 'SECURED-I-PDU']:
            # Skip,If I-Signal pdu is mappded to N-pdu or SECURED-I-PDU
            continue
        if (i > 0) and (node_df[col_namespace].values[i - 1] != node_df[col_namespace].values[i]):
            if (match_flt_evt_pattern(namespace) == '0x00'):  # Filter out all FLT_EVT message
                cycletime = node_df[col_cycle_time].values[i]
                try:
                    cycletime = float(cycletime)
                except:
                    cycletime = 0
                if str(cycletime).endswith(r'.0'):
                    cycletime = int(cycletime)
                if cycletime in [0, 0.0]:
                    node.append('   {0}({1}::{2});'.format(DisableMsg,network_name,namespace))
                    message_count += 1
    node.append('')
    node.append('}')
    logger.info(f"void: generate_void_sim_DisableEventMsg with {str(message_count)} messages.")
    node.append('')
    node.append('void {0}_sim_EnableEventMsg()'.format(node_name))
    node.append('{')
    message_count = 0
    for i in range(num_of_rows):
        # print('we are on voidstop i = ' + str(i))
        namespace = str(node_df[col_namespace].values[i])
        pdu_type= str(node_df[col_pdu_type].values[i])
        payload_pdu_type = str(node_df[col_payload_pdu_type].values[i])
##        if namespace in pdu_list:
        ##            continue

        if "N-PDU" in pdu_type or "SECURED-I-PDU" in pdu_type or pdu_type in skip_pdu_type_list:
            continue
        elif ("I-SIGNAL-I-PDU" in pdu_type) and payload_pdu_type == "MULTIPLEXED-I-PDU":
            continue

        if payload_pdu_type in ['N-PDU', 'SECURED-I-PDU']:
            # Skip,If I-Signal pdu is mappded to N-pdu or SECURED-I-PDU
            continue
        if (i > 0) and (node_df[col_namespace].values[i - 1] != node_df[col_namespace].values[i]):
            if (match_flt_evt_pattern(namespace) == '0x00'):  # Filter out all FLT_EVT message
                cycletime = node_df[col_cycle_time].values[i]
                try:
                    cycletime = float(cycletime)
                except:
                    cycletime = 0
                if str(cycletime).endswith(r'.0'):
                    cycletime = int(cycletime)
                if cycletime in [0, 0.0]:
                    node.append('   {0}({1}::{2});'.format(EnableMsg,network_name,namespace))
                    message_count += 1
    node.append('')
    node.append('}')
    logger.info(f"void: generate_void_sim_EventMsg with {str(message_count)} messages.")

def generate_void_sim_SetCANParameter():#for future purpose
    """
    Configuare all CAN Parameter
    """
    if(var_IL=='CANoeIL'):
        node.append('void {0}_sim_SetCANParameter()'.format(node_name))
        node.append('{')
        node.append('  //For future purpose')
        node.append('  //ILFaultInjectionSetMsgDlc(dbMsg msg, dword dlc)')
        node.append('  //ILFaultInjectionSetMsgLength(dbMsg msg, dword msgLength)')
        node.append('  //ILFaultInjectionResetMsgDlc(dbMsg msg)')
        node.append('  //ILFaultInjectionResetMsgLength (dbMsg msg)')
        node.append('}')
        logger.info(f"generate_void_sim_SetCANParameter")
    

def generate_void_sim_SetCANFDParameter():
    """
    Configuare all CAN-FD Parameter
    """
    if(IsNodeCANFdEnabled(node_name,bus_type) and var_IL=='CANoeIL'):
        node.append('')
        node.append('void {0}_sim_SetCANFDParameter()'.format(node_name))
        node.append('{')
        node.append('   //CAN FD Parameter Setting')
        node.append('   int FDF =1;')
        node.append('   int BRS =1;')
        # Setting CAN-FD Parameters
        message_count = 0
        for i in range(num_of_rows):
            # print('we are on voidstop i = ' + str(i))
            namespace = str(node_df[col_namespace].values[i])
            pdu_type=str(node_df[col_pdu_type].values[i])
            if "N-PDU" in pdu_type or "SECURED-I-PDU" in pdu_type or pdu_type in skip_pdu_type_list:
                continue
            
            if str(node_df[col_payload_pdu_type].values[i]) =='N-PDU':
                # Skip,If I-Signal pdu is mappded to N-pdu
                continue

            if match_pattern(namespace,stbm_msg_pattern):
                continue
            if (i > 0) and (node_df[col_namespace].values[i - 1] != node_df[col_namespace].values[i]):
                if (match_flt_evt_pattern(namespace) == '0x00'):  # Filter out all FLT_EVT message
                    node.append('   ILSetCANFDParam({0}::{1},FDF,BRS,{0}::{1}.dlc);'.format(network_name,namespace))
                    message_count += 1
        node.append('}')
        logger.info(f"void: generate_sim_SetCANFDParameter with {str(message_count)} messages.")
        
def generate_on_CycleTime():
    """
    Configuare cycle time for cyclic message/pdu
    """
    # Create on CycleTime for node variables
    message_count = 0
    if(var_IL=='AsrPDUIL'):
        node.append('//  ARILSetPDUTimingCyclic (dbMsg dbMessage, long TrueOrFalse, long offset, long period, long disturbanceCount, long flags);')
        node.append('//  TrueOrFalse ; 0: denotes the False timing; 1: denotes the True timing; 3: denotes the True and the False timing.')
        node.append('//  offset :Defines the delay in [ms] from now when the first cyclic transmission will start.')
        node.append('//  period :Defines the period in [ms] for the cyclic transmission.')
        node.append('//  disturbanceCount :Reserved/unused; should be set to -1 (infinite).')
        node.append('//  flags :Reserved; should be set to 0.')
    for i in range(num_of_rows):
        # print('we are on voidstop i = ' + str(i))
        namespace = str(node_df[col_namespace].values[i])
        pdu_type=str(node_df[col_pdu_type].values[i])
        pdu_name = node_df[col_pdu].values[i]
        #if namespace in pdu_list:
        if "N-PDU" in pdu_type or "SECURED-I-PDU" in pdu_type or "MULTIPLEXED-I-PDU" in pdu_type or pdu_type in skip_pdu_type_list:
            continue
        
        if str(node_df[col_payload_pdu_type].values[i]) == 'N-PDU':
            # Skip,If I-Signal pdu is mappded to N-pdu
            continue

        if match_pattern(namespace,stbm_msg_pattern):
            continue
        if (i > 0) and (node_df[col_namespace].values[i - 1] != node_df[col_namespace].values[i]):
            if (match_flt_evt_pattern(namespace) == '0x00'):  # Filter out all FLT_EVT message
                cycletime = node_df[col_cycle_time].values[i]
                try:
                    cycletime = float(cycletime)
                except:
                    cycletime = 0
                if str(cycletime).endswith(r'.0'):
                    cycletime = int(cycletime)
                if cycletime not in [0, 0.0]:
                    node.append('')
                    node.append('on sysvar {1}::{0}::{0}_CycleTime'.format(namespace,network_name))
                    node.append('{')
                    if(var_IL=='AsrPDUIL'):
                        node.append('  ARILSetPDUTimingCyclic ({0}::{1}, 3, {2}, @this, -1, 0);//TrueOrFalse=3,offset={2};disturbanceCount =-1;flags=0'.format(network_name,namespace,message_count))
                        node.append('  //ARILResetPDUTimingCyclic({0},3);//TrueOrFalse=3'.format(namespace))
                    elif(var_IL=='CANoeIL'):
                        if not is_empty_or_spaces(pdu_name):
                            node.append('  ILNodeSetPDUTimingCyclic  ({0}, 3, {1}, {2}, -1, 0);//TrueOrFalse=3;offset={1};cycletime={2};disturbanceCount =-1;flags=0'.format(pdu_name,message_count,cycletime))
                            node.append('  //ILNodeResetPDUTimingCyclic ({0},3);//TrueOrFalse=3'.format(pdu_name))
                        else:  # Default
                            node.append('  ILFaultInjectionSetMsgCycleTime({0}::{1},@this);'.format(network_name,namespace))
                            node.append('  //ILFaultInjectionResetMsgCycleTime({0});'.format(namespace))
                    else:#Default
                        node.append('  ILFaultInjectionSetMsgCycleTime({0}::{1},@this);'.format(network_name,namespace))
                        node.append('  //ILFaultInjectionResetMsgCycleTime({0});'.format(namespace))
                    node.append('}')
                    if((message_count+1)%10!=0):#Should not collide with 10ms pdu/msg
                        message_count += 1
                    else:
                        message_count += 2
    logger.info(f"void: generate_on_CycleTime with {str(message_count)} messages.")

def generate_void_sim_SetCycleTimeOffset():
    """
    Configuare cycle time and offset for cyclic message/pdu
    """
    # Create on CycleTime for node variables
    message_count = 0
    node.append('')
    node.append('void {0}_sim_SetCycleTimeOffset()'.format(node_name))
    node.append('{')
    node.append('')
    if(var_IL=='AsrPDUIL'):
        node.append('//  ARILSetPDUTimingCyclic (dbMsg dbMessage, long TrueOrFalse, long offset, long period, long disturbanceCount, long flags);')
        node.append('//  TrueOrFalse ; 0: denotes the False timing; 1: denotes the True timing; 3: denotes the True and the False timing.')
        node.append('//  offset :Defines the delay in [ms] from now when the first cyclic transmission will start.')
        node.append('//  period :Defines the period in [ms] for the cyclic transmission.')
        node.append('//  disturbanceCount :Reserved/unused; should be set to -1 (infinite).')
        node.append('//  flags :Reserved; should be set to 0.')
    elif(var_IL=='CANoeIL'):
        node.append('//  ILNodeSetPDUTimingCyclic (dbMsg dbMessage, long TrueOrFalse, long offset, long period, long disturbanceCount, long flags);')
        node.append('//  TrueOrFalse ; 0: denotes the False timing; 1: denotes the True timing; 3: denotes the True and the False timing.')
        node.append('//  offset :Defines the delay in [ms] from now when the first cyclic transmission will start.')
        node.append('//  period :Defines the period in [ms] for the cyclic transmission.')
        node.append('//  disturbanceCount :Reserved/unused; should be set to -1 (infinite).')
        node.append('//  flags :Reserved; should be set to 0.')
    for i in range(num_of_rows):
        # print('we are on voidstop i = ' + str(i))
        namespace = str(node_df[col_namespace].values[i])
        pdu_type=str(node_df[col_pdu_type].values[i])
        
        msg_name=str(node_df[col_name].values[i])
        pdu_name=str(node_df[col_pdu].values[i])
        #if namespace in pdu_list:
        if "N-PDU" in pdu_type or "SECURED-I-PDU" in pdu_type or "MULTIPLEXED-I-PDU" in pdu_type or pdu_type in skip_pdu_type_list:
            continue
        
        if str(node_df[col_payload_pdu_type].values[i]) == 'N-PDU':
            # Skip,If I-Signal pdu is mappded to N-pdu
            continue

        if match_pattern(namespace,stbm_msg_pattern):
            continue
        if (i > 0) and (node_df[col_namespace].values[i - 1] != node_df[col_namespace].values[i]):
            if (match_flt_evt_pattern(namespace) == '0x00'):  # Filter out all FLT_EVT message
                cycletime = node_df[col_cycle_time].values[i]
                try:
                    cycletime = float(cycletime)
                except:
                    cycletime = 0
                if str(cycletime).endswith(r'.0'):
                    cycletime = int(cycletime)
                if cycletime not in [0, 0.0]:
                    if(var_IL=='AsrPDUIL'):
                        node.append('  ARILSetPDUTimingCyclic ({0}::{1}, 3, {2}, {3}, -1, 0);//TrueOrFalse=3;offset={2};cycletime={3};disturbanceCount =-1;flags=0'.format(network_name,namespace,message_count,cycletime))
                        node.append('  //ARILResetPDUTimingCyclic({0},3);//TrueOrFalse=3'.format(namespace))
                    elif(var_IL=='CANoeIL'):
                        if not is_empty_or_spaces(pdu_name):
                            node.append('  ILNodeSetPDUTimingCyclic  ({0}, 3, {1}, {2}, -1, 0);//TrueOrFalse=3;offset={1};cycletime={2};disturbanceCount =-1;flags=0'.format(pdu_name,message_count,cycletime))
                            node.append('  //ILNodeResetPDUTimingCyclic ({0},3);//TrueOrFalse=3'.format(pdu_name))
                    '''
                    else:#Default
                        node.append('  ILFaultInjectionSetMsgCycleTime({0}::{1},@this);'.format(network_name,namespace))
                        node.append('  //ILFaultInjectionResetMsgCycleTime({0});'.format(namespace))
                    '''
                    if((message_count+1)%10!=0):#Should not collide with 10ms pdu/msg
                        message_count += 1
                    else:
                        message_count += 2
    node.append('')
    node.append('}')      
    logger.info(f"void: generate_void_sim_SetCycleTimeOffset with {str(message_count)} messages.")

    
def generate_dword_applILTx():
    """ """
    if(var_IL=='AsrPDUIL'):
        generate_dword_applPDUILTxPending()
    elif(var_IL=='CANoeIL'):
        generate_dword_applILTxPending()
    else:
        generate_dword_applILTxPending()
        
def generate_dword_applPDUILTxPending():
    """
    applPDUILTxPending : Callback is being called before the IL sends a PDU to the bus
    """
    node.append('')
    node.append('dword applPDUILTxPending (dword busContext, dword shortID, dword longID, char name[], dword & aPDULength, byte data[])//This callback is being called before the IL sends a PDU to the bus'.format(node_name))
    node.append('{')
    # Create turning off of timers and node variables
    message_count = 0
    for i in range(num_of_rows):
        # print('we are on voidstop i = ' + str(i))
        namespace = str(node_df[col_namespace].values[i])
        pdu_type = str(node_df[col_pdu_type].values[i])
        payload_pdu_type = str(node_df[col_payload_pdu_type].values[i])
    
        if "N-PDU" in pdu_type or "SECURED-I-PDU" in pdu_type or pdu_type in skip_pdu_type_list:
            continue

        if payload_pdu_type == 'N-PDU':
            # Skip,If I-Signal pdu is mappded to N-pdu 
            continue

        if ("I-SIGNAL-I-PDU" in pdu_type) and payload_pdu_type == "MULTIPLEXED-I-PDU":
            continue

##        if namespace in pdu_list:
##            continue
        if match_pattern(namespace,stbm_msg_pattern):
            continue
        if (i > 0) and (node_df[col_namespace].values[i - 1] != node_df[col_namespace].values[i]):
            if (match_flt_evt_pattern(namespace) == '0x00'):  # Filter out all FLT_EVT message
                node.append('  if(strncmp(name, "{0}", elcount(name)) == 0)'.format(namespace))
                node.append('  {')
                node.append('    cfg_Pdu_{0}(aPDULength, data);'.format(namespace))
                node.append("    return 1; // don't prevent sending of the pdu")
                node.append('  }')
                message_count += 1
    node.append("  return 1; // don't prevent sending of the pdu")
    node.append('}')
    logger.info(f"void: generate_dword_applILTxPending with {str(message_count)} messages.")

def generate_dword_applILTxPending():
    """
    applPDUILTxPending : Callback is being called before the IL sends a message to the bus
    """
    node.append('')
    node.append('dword applILTxPending (long aId, dword aDlc, byte data[])//This callback is being called before the IL sends a message to the bus'.format(node_name))
    node.append('{')
    node.append('  switch(aId)')
    node.append('  {')
    # Create turning off of timers and node variables
    message_count = 0
    for i in range(num_of_rows):
        # print('we are on voidstop i = ' + str(i))
        namespace = str(node_df[col_namespace].values[i])
        if match_pattern(namespace,stbm_msg_pattern):
            continue
        if (i > 0) and (node_df[col_namespace].values[i - 1] != node_df[col_namespace].values[i]):
            if (match_flt_evt_pattern(namespace) == '0x00'):  # Filter out all FLT_EVT message
                node.append('    case {0}::{1}.id:cfg_Msg_{1}(aDlc, data);break;'.format(network_name,namespace))
                message_count += 1
    node.append('    default :;//write("Node:{0} ::No Cyclic event");'.format(node_name))
    node.append('           break;'.format(node_name))
    node.append('  }')
    node.append("  return 1; // don't prevent sending of the message")
    node.append('}')
    logger.info(f"void: generate_dword_applILTxPending with {str(message_count)} messages.")

def generate_void_OnSecurityLocalPDUPreTx():
    """
    OnSecurityLocalPDUPreTx : Callback handler is called, after all data updates and the automatic Authenticator (CMAC) calculation has be done
    """
    if(var_IL=='AsrPDUIL'):
        filtered_df = node_df.iloc[1:]
        pdu_count = 0
        sec_fun = []  #Fault Injection function defination for pdu 
        node.append('')
        node.append('void OnSecurityLocalPDUPreTx(char pduName[], dword dataId, byte payload[], dword payloadLength, qword& authInfoHigh, qword& authInfo, dword authInfoBitLength, qword& freshness, dword freshnessBitLength, dword freshnessValueId)//This callback handler is called, after all data updates and the automatic Authenticator (CMAC) calculation has be done'.format(node_name))
        node.append('{')
        for pdu_type, pdu_type_df in filtered_df.groupby(col_pdu_type):
            if ('SEC' in pdu_type.upper()):
                for payload_pdu, payload_pdu_df in pdu_type_df.groupby(col_payload_pdu):
                    if is_empty_or_spaces(payload_pdu):
                        continue
                    pdu_count +=1
                    node.append('  if(strncmp(pduName, "{0}", elcount(pduName)) == 0)'.format(payload_pdu))
                    node.append('  {')
                    node.append('    cfg_Secoc_{0}(dataId ,payload, payloadLength, authInfoHigh, authInfo, authInfoBitLength, freshness, freshnessBitLength, freshnessValueId);'.format(payload_pdu))
                    node.append('  }')
                    sec_fun.extend(generate_Secoc_Fault(payload_pdu, network_name))
        node.append('}')
        node.extend(sec_fun)
        logger.info(f"void: generate_void_OnSecurityLocalPDUPreTx with {str(pdu_count)} pdu.")

def generate_ontmr():
    """
    Calculates and fault injection for all counters(alive and block),timestamps and crc based on message/pdu (and also signal group if applicable)
    """
    # Create initialization of timers and node variables.
    tmr_count = 0
    count_normal = 0  # Normal
    count_ac = 0  # Alive Counter
    count_bc = 0  # Block Counter
    count_ts = 0  # Time Stamp
    count_crc = 0  # CRC or CHKSUM
    signal_sum = 0
    filtered_df = node_df.iloc[1:]

    # Controlling "CONTAINER-I-PDU" mapped "I-SIGNAL-I-PDU"
    for payload_pdu_type, payload_pdu_type_df in filtered_df.groupby(col_payload_pdu_type):
        if payload_pdu_type == "CONTAINER-I-PDU":
            for payload_pdu, payload_pdu_df in payload_pdu_type_df.groupby(col_payload_pdu):
                pdu_type = str(payload_pdu_df[col_pdu_type].values[0])
                if "I-SIGNAL-I-PDU" in pdu_type:
                    sysvar_on = []
                    sysvar_off = []
                    node.append('')
                    node.append('on sysvar {1}::{0}::{0}_ON_OFF'.format(payload_pdu, network_name))
                    node.append('{')
                    node.append('  if (@this == 1)')
                    node.append('  {')
                    for pdu, pdu_df in payload_pdu_df.groupby(col_pdu):
                        sysvar_on.append('     @{1}::{0}::{0}_ON_OFF = 1;'.format(pdu, network_name))
                        sysvar_off.append('     @{1}::{0}::{0}_ON_OFF = 0;'.format(pdu, network_name))
                    node.extend(sysvar_on)
                    node.append('  }')
                    node.append('  else')
                    node.append('  {')
                    node.extend(sysvar_off)
                    node.append('  }')
                    node.append('}')

    for payload_pdu_type, payload_pdu_type_df in filtered_df.groupby(col_payload_pdu_type):
        for message, message_df in payload_pdu_type_df.groupby(col_msg):
            #print("message:", message)
            for pdu, pdu_df in message_df.groupby(col_pdu):
                #print("PDU:", pdu)
                if is_empty_or_spaces(pdu):
                    namespace=message
                    EnableMsg  = 'ILFaultInjectionEnableMsg'
                    DisableMsg = 'ILFaultInjectionDisableMsg'
                    MsgEvent   = 'ILSetMsgEvent'
                    function_def = 'void cfg_Msg_{0}(dword aDlc, byte data[])'
                    SigGrp_def   = 'void cfg_SigGrp_{0}(dword aDlc, byte data[])'
                    SigGrp_cal   = '   cfg_SigGrp_{0}(aDlc, data);'
                else:
                    namespace=pdu
                    EnableMsg   = 'ARILFaultInjectionEnablePDU'
                    DisableMsg  = 'ARILFaultInjectionDisablePDU'
                    MsgEvent    = 'ARILSetPDUEvent'
                    function_def = 'void cfg_Pdu_{0}(dword & aPDULength, byte data[])'
                    SigGrp_def   = 'void cfg_SigGrp_{0}(dword & aPDULength, byte data[])'
                    SigGrp_cal   = '   cfg_SigGrp_{0}(aPDULength, data);'

                pdu_type = str(pdu_df[col_pdu_type].values[0])
                pdu_length_byte = pdu_df[col_pdu_length_byte].values[0]
                header_id = pdu_df[col_header_id].values[0]

                if "N-PDU" in pdu_type or "SECURED-I-PDU" in pdu_type or pdu_type in skip_pdu_type_list:
                    continue

                if match_pattern(namespace,stbm_msg_pattern):
                    continue

                if match_flt_evt_pattern(namespace) != '0x00':
                    continue

                cycletime = pdu_df[col_cycle_time].values[0]

                try:
                    cycletime = float(cycletime)
                except:
                    cycletime = 0
                if str(cycletime).endswith(r'.0'):
                    cycletime = int(cycletime)

                temp_df=pdu_df
                # N-PDU read and write void function
                if not (is_empty_or_spaces(pdu)):
                    if payload_pdu_type == 'N-PDU':
                        pdu_type = list(pdu_df[col_pdu_type].loc[1:].unique())[0].upper()

                        byte_order = str(pdu_df[col_byte_order].values[0])

                        pdu_startbit = min(list(pdu_df[col_startbit].values))
                        pdu_startbit_length = pdu_df[col_signal_length_bit].values[list(pdu_df[col_startbit].values).index(pdu_startbit)]

                        pdu_endbit = max(list(pdu_df[col_startbit].values))
                        pdu_endbit_length = pdu_df[col_signal_length_bit].values[list(pdu_df[col_startbit].values).index(pdu_endbit)]

                        DataLength=Get_DataLength(pdu_startbit,pdu_startbit_length,pdu_endbit,pdu_endbit_length,byte_order)

                        if "I-SIGNAL" in pdu_type:
                            node.extend(generate_npdu_onsyvar(pdu, pdu_df))
                            if pdu in pdu_skip:
                                node.extend(generate_npdu_write_fielddata(node_name, pdu, pdu_df, byte_order))
                            else:
                                node.extend(generate_npdu_write(network_name,pdu, pdu_df))
                            if pdu in pdu_skip:
                                node.extend(generate_npdu_timer(network_name, node_name, pdu, pdu_df))
                            node.extend(generate_npdu_Transmit(network_name, pdu, pdu_df))
                        continue
                    if ("I-SIGNAL-I-PDU" in pdu_type) and payload_pdu_type == "MULTIPLEXED-I-PDU":
                        for row_index, row in temp_df.iterrows():
                            signal_name = str(row[col_name])
                            sfd_signal_type = str(row[col_selector_field_signal])
                            payload_pdu = str(row[col_payload_pdu])
                            if (sfd_signal_type == "yes"):
                                node.append('')
                                node.append(
                                    'on sysvar {1}::{0}::{2}_Selection_Field_Code'.format(payload_pdu, network_name,
                                                                                       signal_name))
                                node.append('{')
                                node.append('${0}::{1}=@this;'.format(payload_pdu, signal_name))
                                node.append(' }')
                                node.append(' ')
                        continue

                node.append('')
                node.append('on sysvar {1}::{0}::{0}_ON_OFF'.format(namespace,network_name))
                node.append('{')
                node.append('  if (@this == 1)')
                node.append('  {')
                if cycletime in [0, 0.0]:#Event
                    node.append('     {0}({1}::{2});'.format(MsgEvent,network_name,namespace))
                    node.append('     @{1}::{0}::{0}_ON_OFF = 0;'.format(namespace,network_name))
                    node.append('  }')
                else :#Cyclic
                    node.append('    {0}({1}::{2});'.format(EnableMsg,network_name,namespace))
                    node.append('  }')
                    node.append('  else{')
                    node.append('         {0}({1}::{2});'.format(DisableMsg,network_name,namespace))
                    node.append('      }')
                node.append('}')
                node.append('')
                node.append(function_def.format(namespace))
                node.append('{')


                Sig_grp_fun           = []  # Signal belongs to signal group
                without_Sig_grp_fun   = []  # Signal without signal group

                siggrp_counts = pdu_df[pdu_df[col_siggrp] != ''][col_siggrp].nunique()
                empty_siggrp_counts = pdu_df[pdu_df[col_siggrp] == ''][col_siggrp].nunique()

                e2e_def=False
                if siggrp_counts == 1 :
                    if (input_file.lower()!='dbc') and (pdu_df[pdu_df[col_endtoendprotection] != ''][col_endtoendprotection].nunique()!=0):#If pdu group has E2E profile
                        e2e_def=True

                # if check_PDU_in_SecuredPdu(namespace, securedPdu_dict):
                #     retrieved_msg = get_message_from_SecuredPdu(namespace, securedPdu_dict)
                #     print(f'generate_ontmr::ontmr_variables::if ::retrieved_msg = {retrieved_msg}, pdu = {pdu}, e2e_def = {e2e_def}')
                #     node.extend(ontmr_variables(bus_type, input_file, retrieved_msg, pdu, network_name,e2e_def))  # Import from pattern_matching
                # else:
                #     print(f'generate_ontmr::ontmr_variables::else ::message = {message}, pdu = {pdu}, e2e_def = {e2e_def}')
                signal_dict = {}
                for signal_name, signal_df in pdu_df.groupby(col_name):
                    signal_byte_order = signal_df[col_byte_order].values[0]
                    signal_start_bit = signal_df[col_startbit].values[0]
                    signal_length = signal_df[col_signal_length_bit].values[0]
                    signal_para_list = [signal_byte_order, signal_start_bit, signal_length]
                    if 'OPAQUE' in signal_byte_order.upper():
                        signal_dict[signal_name] = signal_para_list

                # signal_para_list = [signal_name, signal_byte_order, signal_start_bit, signal_length]
                # print(f'pdu_df = \n{pdu_df}')
                para_dict = {
                    "bus_type": bus_type,
                    "input_file": input_file,
                    "message": message,
                    "pdu": pdu,
                    "network_name": network_name,
                    "e2e_def": e2e_def,
                    "signal_dict": signal_dict,
                    "Payload_PDU_Type": payload_pdu_type,
                    "PDU_Type": pdu_type,
                    "PDU_Length [Byte]": pdu_length_byte,
                    "Header ID": header_id
                }

                node.extend(ontmr_variables(para_dict))  # Import from pattern_matching

                if ("MULTIPLEXED-I-PDU" in pdu_type):
                    node.append('}')
                    node.append('')
                    continue

                for siggrp, siggrp_df in pdu_df.groupby(col_siggrp):
                    #print("Signal group:", siggrp)
                    if siggrp_counts > 1 :
                        if not is_empty_or_spaces(siggrp):
                            node.append(SigGrp_cal.format(siggrp))#call signal group config

                    num_of_signals = siggrp_df.shape[0] #print("num_of_signals",num_of_signals)

                    byte_order = str(siggrp_df[col_byte_order].values[0])

                    SigGrp_startbit = min(list(siggrp_df[col_startbit].values)) #print("SigGrp_startbit",SigGrp_startbit)
                    SigGrp_startbit_length = siggrp_df[col_signal_length_bit].values[list(siggrp_df[col_startbit].values).index(SigGrp_startbit)] #print("SigGrp_startbit_length",SigGrp_startbit_length)

                    SigGrp_endbit = max(list(siggrp_df[col_startbit].values)) #print("SigGrp_endbit",SigGrp_endbit)
                    SigGrp_endbit_length = siggrp_df[col_signal_length_bit].values[list(siggrp_df[col_startbit].values).index(SigGrp_endbit)] #print("SigGrp_endbit_length",SigGrp_endbit_length)


                    if not is_empty_or_spaces(siggrp):
                        DataLength=Get_DataLength(SigGrp_startbit,SigGrp_startbit_length,SigGrp_endbit,SigGrp_endbit_length,byte_order)

                    siggrp_info = {
                        "siggrp": siggrp,
                        "SigGrp_startbit": SigGrp_startbit,
                        "SigGrp_startbit_length": SigGrp_startbit_length,
                        "SigGrp_endbit": SigGrp_endbit,
                        "SigGrp_endbit_length": SigGrp_endbit_length,
                        "siggrp_counts": siggrp_counts,
                        "empty_siggrp_counts": empty_siggrp_counts
                    }

                    total_used_msg_bytes = ceil((SigGrp_endbit + SigGrp_endbit_length)/8)-1 #print("total_used_msg_bytes",total_used_msg_bytes)

                    part_msg_null   = []  # Normal
                    part_msg_ac     = []  # Alive Counter
                    part_msg_bc     = []  # Block Counter
                    part_msg_ts     = []  # Time Stamp
                    part_msg_crc    = []  # CRC or CHKSUM
                    tmr_count += 1
                    signal_sum += num_of_signals
                    msg_length = 0  # in bits
                    # for j in range(num_of_signals):
                    # msg_length = msg_length + byteOrder_to_8byte(int(siggrp_df[col_signal_length_bit].values[j]))

                    for j in range(num_of_signals):
                        break_out = 0
                        message_name = str(siggrp_df[col_namespace].values[j])
                        signal_name = str(siggrp_df[col_name].values[j])
                        byte_order = str(siggrp_df[col_byte_order].values[j])
                        endtoendprotection = str(siggrp_df[col_endtoendprotection].values[j])
                        comment = str(siggrp_df[col_comment].values[j])
                        startbit = int(siggrp_df[col_startbit].values[j]) if (
                                    str(siggrp_df[col_startbit].values[j]) != "") else 0
                        msg_length = byteOrder_to_8byte(int(siggrp_df[col_signal_length_bit].values[j])) if (
                                    str(siggrp_df[col_signal_length_bit].values[j]) != "") else 0

                        sig_factor = float(siggrp_df[col_factor].values[j]) if (
                                    str(siggrp_df[col_factor].values[j]) != "") else float(1)

                        sig_offset = float(siggrp_df[col_offset].values[j]) if (
                                    str(siggrp_df[col_offset].values[j]) != "") else float(0)

                        dlc = int(siggrp_df[col_dlc].values[j]) if (str(siggrp_df[col_dlc].values[j]) != "") else 0

                        sig_length = int(siggrp_df[col_signal_length_bit].values[j]) if (str(siggrp_df[col_signal_length_bit].values[j]) != "") else 0

                        para_dict = {
                            "namespace": message_name,
                            "signal_name": signal_name,
                            "siggrp_info": siggrp_info,
                            "startbit": startbit,
                            "msg_length": msg_length,
                            "sig_length": sig_length,
                            "byte_order": byte_order,
                            "sig_factor": sig_factor,
                            "sig_offset": sig_offset,
                            "comment": comment,
                            "cycletime": cycletime,
                            "dlc": dlc,
                            "bus_type": bus_type,
                            "input_file": input_file,
                            "network_name": network_name,
                            "is_public_can": is_public_can,
                            "total_used_msg_bytes": total_used_msg_bytes,
                            "node_name": node_name,
                            "endtoendprotection": endtoendprotection,
                            "PDU_Length [Byte]": pdu_length_byte
                        }
                        
                        if (break_out != 1):
                            for sig_pattern, sig_fun in ac_dict.items():
                                if (None != re.search(sig_pattern, signal_name, re.M | re.I)):  # Pattern :: Alive counter
                                    #print(siggrp,signal_name)
                                    if (input_file=='dbc') or is_empty_or_spaces(endtoendprotection):
                                        part_msg_ac = part_msg_ac + sig_fun(para_dict)
                                    else:
                                        part_msg_ac = part_msg_ac + SigGrp_ARILFaultInjectionDisturbSequenceCounter(para_dict)
                                    count_ac += 1
                                    break_out = 1
                                    break
                        if (break_out != 1):
                            for sig_pattern, sig_fun in bc_dict.items():
                                if (None != re.search(sig_pattern, signal_name, re.M | re.I)):  # Pattern :: Block counter
                                    part_msg_bc = part_msg_bc + sig_fun(para_dict)
                                    count_bc += 1
                                    break_out = 1
                                    break
                        if (break_out != 1):
                            for sig_pattern, sig_fun in ts_dict.items():
                                if (None != re.search(sig_pattern, signal_name, re.M | re.I)):  # Pattern :: Time Stamp
                                    part_msg_ts = part_msg_ts + sig_fun(para_dict)
                                    count_ts += 1
                                    break_out = 1
                                    break
                        if (break_out != 1):
                            for sig_pattern, sig_fun in crc_dict.items():
                                if (None != re.search(sig_pattern, signal_name, re.M | re.I)):  # Pattern ::CRC / CHK_SUM
                                    if (input_file=='dbc') or is_empty_or_spaces(endtoendprotection):
                                        part_msg_crc = part_msg_crc + sig_fun(para_dict)
                                    else:
                                        part_msg_crc = part_msg_crc + SigGrp_ARILFaultInjectionDisturbChecksum(para_dict)
                                    count_crc += 1
                                    break_out = 1
                                    break
                        if (break_out != 1):
                            count_normal += 1
                            #part_msg_null.append('   Msg_{0}.{1} = f_Phys2Raw(@{0}::{1}_Rv,{2},{3});'.format(message_name, signal_name,sig_factor,sig_offset))#Msg_{0}.{1}.phys = @{0}::{1}_Rv

                    if is_empty_or_spaces(siggrp):
                        without_Sig_grp_fun.extend(part_msg_null)
                        without_Sig_grp_fun.extend(part_msg_ac)
                        without_Sig_grp_fun.extend(part_msg_bc)
                        without_Sig_grp_fun.extend(part_msg_ts)
                        without_Sig_grp_fun.extend(part_msg_crc)
                    else:
                        if siggrp_counts > 1:
                            Sig_grp_fun.append(SigGrp_def.format(siggrp))
                            Sig_grp_fun.append('{')
                            Sig_grp_fun.append('   int64 startbit,bitlength;')
                            Sig_grp_fun.append('   int64 i,k;//For looping')
                            Sig_grp_fun.append('   byte DataLength = {0};//Signal group length in byte'.format(DataLength))
                            Sig_grp_fun.append('   byte SigGrp_Data[{0}];'.format(DataLength))
                            Sig_grp_fun.append('   long result;')
                            if (input_file.lower()!='dbc') and (siggrp_df[siggrp_df[col_endtoendprotection] != ''][col_endtoendprotection].nunique()!=0):#If signal group has E2E profile
                                Sig_grp_fun.append('   long type=0;//reserved (should be set to 0).')
                                Sig_grp_fun.append('   long disturbanceMode  =0; //0 :: set to disturbanceValue //2 :: set to Random value //3 :: Offset Signal value is increased by disturbanceValue')
                                Sig_grp_fun.append('   long disturbanceCount =0;//-1 ::Infinite Disturbance is continuously applied //0 ::Stop An active disturbance is stopped and the SequenceCounter/crc will be calculated again appropriately //n > 0 ::Count Do exactly n Repetition/disturbances')
                                Sig_grp_fun.append('   long disturbanceValue =0;//According to the disturbance mode the SequenceCounter/crc will optionally be set to this value')
                                Sig_grp_fun.append('   dword crc[1];')

                            Sig_grp_fun.extend(part_msg_null)
                            Sig_grp_fun.extend(part_msg_ac)
                            Sig_grp_fun.extend(part_msg_bc)
                            Sig_grp_fun.extend(part_msg_ts)
                            Sig_grp_fun.extend(part_msg_crc)
                            Sig_grp_fun.append('')
                            Sig_grp_fun.append('}')
                            Sig_grp_fun.append('')
                        else:
                            without_Sig_grp_fun.extend(part_msg_null)
                            without_Sig_grp_fun.extend(part_msg_ac)
                            without_Sig_grp_fun.extend(part_msg_bc)
                            without_Sig_grp_fun.extend(part_msg_ts)
                            without_Sig_grp_fun.extend(part_msg_crc)

                node.extend(without_Sig_grp_fun)
                node.append('}')
                node.append('')
                node.extend(Sig_grp_fun)

    logger.info("Timer count: {0} on timer functions were generated with {1} signals in total.".format(str(tmr_count),str(signal_sum)))
    logger.info("\tNormal: {0}; Alive Counter: {1}; Time Stamp: {2}; CRC or CHKSUM: {3}".format(count_normal, count_ac,count_ts, count_crc))
                
def generate_onmsg():
    """ """
    # Case create node variables
    global flt_evt_cnt
    global flt_evt_node
    message_count = 0
    part_msg = []
    for i in range(num_of_rows):
        # print('we are on message i = ' + str(i))
        namespace = node_df[col_namespace].values[i]
        if match_pattern(namespace,stbm_msg_pattern):
            continue
        flt_evt_pattern = match_flt_evt_pattern(namespace)
        if ((flt_evt_pattern != '0x00') and (
                flt_evt_pattern != '0xFF')):  # Filter in all FLT_EVT message and Filter out FLT_EVT message from 'FD3'
            if (i > 0) and (node_df[col_namespace].values[i - 1] != node_df[col_namespace].values[i]):
                part_msg = part_msg + FLT_EVENT(namespace)
    flt_evt_node.extend(part_msg)
    if (current_sheet == rbs_sheet_name):  # Old:PSA_CAN_FD3_Rbs
        if (flt_evt_cnt > 0):
            temp_node, flt_evt_node = rbs_node_ack_onmsg(flt_evt_node)
            node.extend(temp_node)
            logger.info(f"on message: generated with {str(flt_evt_cnt)} messages.")
        flt_evt_cnt = 0  # clear count


def getCopyRights(file):
    """
    checks if the given file has copyrights header,
    if present extracts it as a string and returns else returns empty strs

    Args:
      file: str
      Returns: 

    Returns:
      

    """
    try:
        return_string = ""
        if os.path.exists(file):
            # check if copyrights is present
            copyrights_str = "Copyright (c)"
            copyrights_lines = []

            found = False
            with open(file) as existing_file:
                existing_file_lines = existing_file.readlines()
                for i, lin in enumerate(existing_file_lines):
                    if copyrights_str in lin:
                        found = True
                    if "############" in lin and r"*/" in existing_file_lines[i + 1]:
                        copyrights_lines.append(lin)
                        copyrights_lines.append(existing_file_lines[i + 1])
                        if found:
                            return_string = ""
                            for l in copyrights_lines:
                                return_string += l
                            break
                    else:
                        copyrights_lines.append(lin)

                    if i > 40:
                        break
                existing_file.close()
        return return_string
    except Exception as e:
        logger.warning(f"Error while finding copyrights header -> {e}")
        return ""

def save_file(file_path):
    """
    

    Args:
      file_path: 

    Returns:

    """
    # Open and save node .can file
    try:
        file = ('{3}/{0}_{1}_{2}.can'.format(node_name, bus_type_from_sheet_name, bus_name, file_path))
        file_content = node
        copyrights = getCopyRights(file)

        if copyrights:
            file_content[0] = copyrights

        with open(file, 'w') as outfile:
            outfile.write("\n".join(str(item) for item in file_content))
        logger.info(f"{file} updated successfully")
    except:
        logger.warning(f"{node_name} was NOT updated successfully", exc_info=True)



def check_df_integrity(node_df):
    """
    

    Args:
      node_df: 

    Returns:

    """
    # Replacing Header with Top Row
    new_header = node_df.iloc[0]  # grab the first row for the header
    df = node_df[1:]  # take the data less the header row
    df.columns = new_header  # set the header row as the df header

    # Check integrity of Message ID column
    mess_id_status = df['Message ID'].str.contains("^(?:0[xX])[A-F0-9]+$", na=False, regex=True).values.tolist()
    indices = [i for i, x in enumerate(mess_id_status) if x is False]
    for row in indices:
        logger.warning(f"Message ID format on row {row + 2} is not correct")

    # Check for not allowed symbols or date and time
    df_initial = df['Initial Value'].astype(str)
    df_factor = df['Factor'].astype(str)
    df_offset = df['Offset'].astype(str)
    df_mimimum = df['Minimum'].astype(str)
    df_maximum = df['Maximum'].astype(str)

    initial_value_status = df_initial.str.contains(r'[!@#$%^&*(),?":{}|<>\/]', na=False, regex=True).values.tolist()
    indices = [i for i, x in enumerate(initial_value_status) if x is True]
    for row in indices:
        logger.warning(f"Initial value format on row {row + 2} is not correct")

    status = df_factor.str.contains(r'[!@#$%^&*(),?":{}|<>\/]', na=False, regex=True).values.tolist()
    indices = [i for i, x in enumerate(status) if x is True]
    for row in indices:
        logger.warning(f"Initial value format on row {row + 2} is not correct")

    status = df_offset.str.contains(r'[!@#$%^&*(),?":{}|<>\/]', na=False, regex=True).values.tolist()
    indices = [i for i, x in enumerate(status) if x is True]
    for row in indices:
        logger.warning(f"Offset format on row {row + 2} is not correct")

    status = df_mimimum.str.contains(r'[!@#$%^&*(),?":{}|<>\/]', na=False, regex=True).values.tolist()
    indices = [i for i, x in enumerate(status) if x is True]
    for row in indices:
        logger.warning(f"Minimum format on row {row + 2} is not correct")

    status = df_maximum.str.contains(r'[!@#$%^&*(),?":{}|<>\/]', na=False, regex=True).values.tolist()
    indices = [i for i, x in enumerate(status) if x is True]
    for row in indices:
        logger.warning(f"Maximum format on row {row + 2} is not correct")


def read_values_from_row(row):
    """
    

    Args:
      row: 

    Returns:

    """
    #factor, offset, message_name, signal_name, startbit, bitlength
    row_data = {}

    try:
        #add all must needed values from excel(cannot be empty) which we cant assign
        pdu_name = row[col_pdu]
        signal_name = row[col_name]
        start_bit = int(row[col_startbit])
        bit_length = int(row[col_signal_length_bit])
        byte_order = str(row[col_byte_order])
    except:
        raise Exception(f"EMPTY CELL: Error while reading for signal -> {row[col_name]}")
    

    #read values which we can assgn if empty (give user defined values if empty or not correct)
    try:
        factor = np.format_float_positional(row[col_factor], min_digits=1)    #format_float_positional is to convert exponential values to float eg: 3.046e-05 to float:: min_digits is to add .0 if int 
        if not(factor):
            factor = float(1)
    except:
        logger.warning(f"EMPTY CELL: Error occured will reading factor of signal -> {row[col_name]}")
        factor = float(1)
    try:
        offset = np.format_float_positional(row[col_offset], min_digits=1)
        if not(offset):
            offset = float(1)
    except:
        logger.warning(f"EMPTY CELL: Error occured will reading offset of signal -> {row[col_name]}")
        offset = float(0)
    
    row_data = {'pdu_name':pdu_name, 'signal_name':signal_name, 'start_bit':start_bit, 'bit_length':bit_length, 'factor':factor, 'offset':offset ,'byte_order':byte_order}

    return row_data
    
def generateNpduDatalength(sysvardatabase_df, col_sender, col_msg,col_pdu, col_pdu_type, col_payload_pdu_type,col_pdu_length, col_byte_order, col_startbit, col_signal_length_bit):
    # Create *.cin for each N-pdu ->For Osek TP handling.
    npdu_datalength_dict = {}
    filtered_df = sysvardatabase_df.iloc[1:]
    for sender_node, sender_node_df in filtered_df.groupby(col_sender):
        for message, message_df in sender_node_df.groupby(col_msg):
            for pdu, pdu_df in message_df.groupby(col_pdu):
                if is_empty_or_spaces(pdu):
                    namespace = message
                else:
                    namespace = pdu
                pdu_type = str(pdu_df[col_pdu_type].values[0])

                if "N-PDU" in pdu_type or "SECURED-I-PDU" in pdu_type or "MULTIPLEXED-I-PDU" in pdu_type or pdu_type in skip_pdu_type_list:
                    continue
                if match_pattern(namespace, stbm_msg_pattern):
                    continue
                if match_flt_evt_pattern(namespace) != '0x00':
                    continue
                if not is_empty_or_spaces(pdu):
                    if str(pdu_df[col_payload_pdu_type].values[0]) == 'N-PDU':
                        pdu_type = list(pdu_df[col_pdu_type].loc[1:].unique())[0].upper()
                        byte_order = str(pdu_df[col_byte_order].values[0])
                        if 'moto' in byte_order.lower():
                            for sig_index, signal_row in pdu_df.iterrows():
                                pdu_df.loc[sig_index, col_startbit] = findStartBit(signal_row[col_startbit], signal_row[col_signal_length_bit])
                        pdu_startbit = min(list(pdu_df[col_startbit].values))
                        pdu_startbit_length = pdu_df[col_signal_length_bit].values[
                            list(pdu_df[col_startbit].values).index(pdu_startbit)]
                        pdu_endbit = max(list(pdu_df[col_startbit].values))
                        pdu_endbit_length = pdu_df[col_signal_length_bit].values[
                            list(pdu_df[col_startbit].values).index(pdu_endbit)]
                        DataLength = Get_DataLength(pdu_startbit, pdu_startbit_length, pdu_endbit, pdu_endbit_length,byte_order)
                        #DataLength = pdu_df[col_pdu_length].values[0]
                        npdu_datalength_dict[pdu] = DataLength
    return npdu_datalength_dict
def generate_OsekTP_cin(path):
    """
    As per NPDU data it will call osekTp main function

    Args:
      path: location of folder where cin files are generated

    Returns:cin file for osekTp

    """
    filtered_df = node_df.iloc[1:]
    for pdu_type, pdu_type_df in filtered_df.groupby(col_pdu_type):
        if ('N-PDU' in pdu_type.upper()):
            for pdu, pdu_df in pdu_type_df.groupby(col_pdu):
                if not is_empty_or_spaces(pdu):
                    message_name = pdu_df[col_msg].values[0]
                    payload_pdu = pdu_df[col_payload_pdu].values[0]
                    cantp_pdu_type = pdu_df[col_cantp_pdu_type].values[0]
                    para_dict = {
                        "network_name": network_name,
                        "node_name": node_name,
                        "message_name": message_name,
                        "pdu_name": pdu,
                        "payload_pdu": payload_pdu,
                        "file_path": path,
                        "cantp_pdu_type":cantp_pdu_type
                    }

                    PDU_Data_Length = npdu_datalength_dict.get(payload_pdu, None)

                    if PDU_Data_Length is not None:
                        create_OsekTP_node_main(para_dict, PDU_Data_Length, npdu_dict, hil_ctrl_node_dict)


def generate_opaque_byte_order_cin(path):
    filtered_df = node_df.iloc[1:]
    for message, message_df in filtered_df.groupby(col_msg):
        for pdu, pdu_df in message_df.groupby(col_pdu):
            pdu_type = str(pdu_df[col_pdu_type].values[0])
            if "N-PDU" in pdu_type or "SECURED-I-PDU" in pdu_type:
                continue
            if match_pattern(pdu, stbm_msg_pattern):
                continue
            if match_flt_evt_pattern(pdu) != '0x00':
                continue
            for signal_name, signal_df in pdu_df.groupby(col_name):
                signal_byte_order = signal_df[col_byte_order].values[0]
                signal_start_bit = signal_df[col_startbit].values[0]
                signal_length = signal_df[col_signal_length_bit].values[0]
                signal_para_list = [network_name, node_name, pdu, signal_name, signal_byte_order, signal_start_bit, signal_length, path]
                if 'OPAQUE' in signal_byte_order.upper():
                    create_Opaque_main(signal_para_list, path)


def generate_npdu_onsyvar(pdu_name, pdu_df):
    """
    

    Args:
      pdu_name: 
      pdu_df: 

    Returns:

    """
    npdu_list =[]
    Npdu_name = ""
    for key, value in npdu_dict.items():
        if value[1] == pdu_name:#Value[1]-->I-signal Pdu
            Npdu_name = value[0]#Value[0]-->NPdu
            break
    cycletime = pdu_df[col_cycle_time].values[0]
    if node_name in hil_ctrl_node_dict.keys() and hil_ctrl_node_dict[node_name] != "":
        nod_name = hil_ctrl_node_dict[node_name]
    else:
        nod_name = node_name
    try:
        cycletime = float(cycletime)
    except:
        cycletime = 0
    if str(cycletime).endswith(r'.0'):
        cycletime = int(cycletime)

    npdu_list.append('')
    npdu_list.append('on sysvar {0}::{1}::{1}_ON_OFF'.format(network_name,pdu_name))
    npdu_list.append('{')

    # try to get customer specific on_off condition conditions
    cus_condition = ''
    try:
        cus_condition = get_customer_onoff_condition(network_name, node_name, pdu_name)
        if cus_condition != '':
            cus_condition = ' && ' + cus_condition
    except:
        pass

    npdu_list.append('  if ((@this == 1)&& (@hil_ctrl::{0} == 1){1})'.format(nod_name, cus_condition))
    npdu_list.append('   {')
##    npdu_list.append('    if(Token_FC_{0}==0)'.format(pdu_name))
##    npdu_list.append('     {')
    if pdu_name in pdu_skip:
        npdu_list.append('      cancelTimer({0}_Timer); '.format(pdu_name))
        npdu_list.append('      count = 1; ')
    npdu_list.append('      Transmit_Pdu_{0}_{1}();'.format(node_name, pdu_name))
##    npdu_list.append('     }')
    if pdu_name in pdu_skip:
        npdu_list.append('      if (@sysvar::{0}::{1}::txBuffermode == 1 || @sysvar::{0}::{1}::txBuffermode == 2)'.format(network_name, pdu_name))
        npdu_list.append('      {')
        npdu_list.append('         setTimer({0}_Timer,50); '.format(pdu_name))
        npdu_list.append('      }')
    npdu_list.append('      @sysvar::{0}::{1}::{1}_ON_OFF= 0;'.format(network_name, pdu_name))
    npdu_list.append('   }')
    npdu_list.append('}')
    npdu_list.append(' ')
    npdu_list.append('')
    return npdu_list

def generate_npdu_write_bufferwise(pdu_name,buffer_fullname, buffer_df):
    """
    generate write npdu data bufferwise

    Args:
      pdu_name: 
      buffer_fullname: 
      buffer_df: 

    Returns:

    """
    write_list = []
    write_list.append("")
    write_list.append(f"void {buffer_fullname}_write(byte byteArray[])")
    write_list.append("{")
    write_list.append("    qword raw_value;")
    
    for i, row in buffer_df.iterrows():
        row_data = read_values_from_row(row)
        signal_name_split = row_data["signal_name"].split("_")
        write_list.append("")
        if "MOTO" in row_data['byte_order'].upper():
            write_list.append(f"    raw_value = convertPhysicalToRaw(@{network_name}::{row_data['pdu_name']}::{'_'.join(signal_name_split[:-1])}.{signal_name_split[-1]}, {row_data['factor']}, {row_data['offset']});")
            write_list.append(f"    copyBitsToByteArrayBE(raw_value, byteArray, {row_data['start_bit']}, {row_data['bit_length']});")
        elif "INTEL" in row_data['byte_order'].upper():
            write_list.append(f"    raw_value = convertPhysicalToRaw(@{network_name}::{row_data['pdu_name']}::{'_'.join(signal_name_split[:-1])}.{signal_name_split[-1]}, {row_data['factor']}, {row_data['offset']});")
            write_list.append(f"    copyBitsToByteArrayLE(raw_value, byteArray, {row_data['start_bit']}, {row_data['bit_length']});")
    write_list.append("}")
    write_list.append("")
    node.extend(write_list)

def generate_npdu_write(network_name,pdu_name, pdu_df):
    """
    

    Args:
      network_name: 
      pdu_name: 
      pdu_df: 

    Returns:

    """
    write_list = []
    write_list.append("")
    write_list.append(f"void Write_OsekTP_{node_name}_{pdu_name}(byte byteArray[])")
    write_list.append("{")
    write_list.append("    qword raw_value;")
    #ADAS HIL specific part
    #Adding the part to get the byte array from the radar model with conditionnal compiling
    #As it is needed only for locations/objects, we check for the location/object pdu
    if ("Radar" in node_name) and (pdu_name_pattern in pdu_name):
        if node_name[-2:].lower() == "fc":
            radar_id = "1"
        elif node_name[-2:].lower() == "fl":
            radar_id = "2"
        elif node_name[-2:].lower() == "fr":
            radar_id = "3"
        elif node_name[-2:].lower() == "rl":
            radar_id = "4"
        elif node_name[-2:].lower() == "rr":
            radar_id = "5"
        else:
            radar_id = "0"
        write_list.extend(generate_target_radar_sysvar(node_name, radar_id))#moved to pattern_matching_canTP py file
    pdu_df["BUFFER"] = pdu_df[col_name].apply(lambda x: int(x.split("_")[2].replace(pdu_shortname_pattern,"")) if re.match(Npdu_sig_patterns, x, flags=re.I) else int(-1))
    Switch_flag=False
    Break_flag=False
    # print(f'pdu_df["BUFFER"] = \n{pdu_df["BUFFER"]}')
    # print(f'len(pdu_df["BUFFER"]) = {len(pdu_df["BUFFER"])}')
    #print(pdu_df.sort_values("BUFFER", ascending=False).groupby("BUFFER"))
    pdu_df=pdu_df.sort_values("BUFFER", ascending=False)
    for buffer, buffer_df in pdu_df.groupby("BUFFER",sort=False):  #.groupby("BUFFER")
        if buffer==-1:
            if (len(pdu_df["BUFFER"]) > 1) and (Break_flag==False):
                #Break must be at the end of the switch case 
                write_list.append("    break;")
                write_list.append("    }")
                #Adding parsing for ADAS HIL
                write_list.append("    #if ADAS_HIL")
                write_list.append("    }")
                write_list.append("    #endif")
                Break_flag=True
            for i, row in buffer_df.iterrows():
                row_data = read_values_from_row(row)
                write_list.append("")
                if re.search(TimeStamp_pattern,row_data['signal_name']):       #To select STBM time to timestamp variable
                    write_list.append(f"    @{network_name}::{row_data['pdu_name']}::{row_data['signal_name']} = (timeNowInt64()-@CAN_Radar::TimeSync_CAN_Radar::SYNC_LastTimeReceivedNs)/100000+@CAN_Radar::TimeSync_CAN_Radar::stbmTime_100us;  //Compute real time according to Autosar specification")

                if "MOTO" in row_data['byte_order'].upper():
                    write_list.append(f"    raw_value = convertPhysicalToRaw(@{network_name}::{row_data['pdu_name']}::{row_data['signal_name']}, {row_data['factor']}, {row_data['offset']});")
                    write_list.append(f"    copyBitsToByteArrayBE(raw_value, byteArray, {row_data['start_bit']}, {row_data['bit_length']});")
                elif "INTEL" in row_data['byte_order'].upper():
                    write_list.append(f"    raw_value = convertPhysicalToRaw(@{network_name}::{row_data['pdu_name']}::{row_data['signal_name']}, {row_data['factor']}, {row_data['offset']});")
                    write_list.append(f"    copyBitsToByteArrayLE(raw_value, byteArray, {row_data['start_bit']}, {row_data['bit_length']});")

        else:
            if Switch_flag==False:
                write_list.append(f'    switch(@{network_name}_{pdu_name}::Write_No_signals)')
                write_list.append("    {")
                Switch_flag=True
            buffer_fullname = "_".join(buffer_df[col_name].values[0].split("_")[:-1])
            buffer_number =re.findall(r'\d+', buffer_fullname)
            generate_npdu_write_bufferwise(pdu_name, buffer_fullname, buffer_df)
            write_list.append(f'    case {buffer_number[0]}:')
            write_list.append(f"    {buffer_fullname}_write(byteArray);")

    write_list.append("}")
    write_list.append("")
    return write_list


def generate_npdu_Transmit(network_name, pdu_name, pdu_df):
    """
    

    Args:
      pdu_name: 
      pdu_df: 

    Returns:

    """
    npdu_list =[]
    Npdu_name = ""
    for key, value in npdu_dict.items():
        if value[1] == pdu_name:#Value[1]-->I-signal Pdu
            Npdu_name = value[0]#Value[0]-->NPdu
            break
    signal_name = pdu_df[col_name].values[0]
    npdu_list.append('')
    npdu_list.append('void Transmit_Pdu_{0}_{1}()'.format(node_name,pdu_name))
    npdu_list.append('{')
    npdu_list.append('   long txCount ;')
    npdu_list.append('   pdu {0} pdu_{0};'.format(pdu_name))
    if pdu_name in pdu_skip:
        #pdu skip is specific to field data PDU
        npdu_list.append('   if (@sysvar::{0}::{1}::txBuffermode == 2) // Big_Packet'.format(network_name, pdu_name))
        npdu_list.append('   {')
        npdu_list.append('     byte txDataBuffer[kBufferSizeJumbo_{0}_{1}_large];'.format(node_name, Npdu_name))
        npdu_list.append('     byte txDataBuffer_null[kBufferSizeJumbo_{0}_{1}_large];'.format(node_name, Npdu_name))
        npdu_list.append('     txCount = kBufferSizeJumbo_{0}_{1}_large;'.format(node_name, Npdu_name))
        npdu_list.append('     Write_OsekTP_{0}_{1}_Big_Packet(txDataBuffer);'.format(node_name, pdu_name))
        npdu_list.append('     CanTpSendData(gHandle_{0}_{1}, txDataBuffer, txCount);'.format(node_name, Npdu_name))
        npdu_list.append('     memcpy(txDataBuffer, txDataBuffer_null, txCount);')
        npdu_list.append('   }')
        npdu_list.append('   else if(@sysvar::{0}::{1}::txBuffermode == 1) //  Small_Packet'.format(network_name, pdu_name))
        npdu_list.append('   {')
        npdu_list.append('     byte txDataBuffer[kBufferSizeJumbo_{0}_{1}_short];'.format(node_name, Npdu_name))
        npdu_list.append('     byte txDataBuffer_null[kBufferSizeJumbo_{0}_{1}_short];'.format(node_name, Npdu_name))
        npdu_list.append('     txCount = kBufferSizeJumbo_{0}_{1}_short;'.format(node_name, Npdu_name))
        npdu_list.append('     Write_OsekTP_{0}_{1}_Small_Packet(txDataBuffer);'.format(node_name, pdu_name))
        npdu_list.append('     CanTpSendData(gHandle_{0}_{1}, txDataBuffer, txCount);'.format(node_name, Npdu_name))
        npdu_list.append('     memcpy(txDataBuffer, txDataBuffer_null, txCount); \n   }')
        npdu_list.append('   else if(@sysvar::{0}::{1}::txBuffermode == 0) //  User_defined'.format(network_name, pdu_name))
        npdu_list.append('   {')
        npdu_list.append('     byte txDataBuffer[kBufferSizeJumbo_{0}_{1}];'.format(node_name, Npdu_name))
        npdu_list.append('     byte sendString[kBufferSizeJumbo_{0}_{1}];'.format(node_name, Npdu_name))
        npdu_list.append('     byte txDataBuffer_null[kBufferSizeJumbo_{0}_{1}];'.format(node_name, Npdu_name))
        npdu_list.append('     byte emptyString[kBufferSizeJumbo_{0}_{1}];'.format(node_name, Npdu_name))
        npdu_list.append('     long copiedBytes; \n     long TxLength; \n')
        npdu_list.append('     // Preparation same as sysSendData \n     TxLength = 0;')
        npdu_list.append('     if (sysGetVariableData(sysvar::{0}::{1}::{2}::SignalValue, sendString, copiedBytes) == 0)'.format(network_name, pdu_name, signal_name))
        npdu_list.append('     { \n       TxLength=copiedBytes; \n       memcpy( txDataBuffer, sendString, TxLength); \n     } \n')
        npdu_list.append('     // Start of TransmitTxBuffer()')
        npdu_list.append('     CanTpSendData(gHandle_{0}_{1}, txDataBuffer, TxLength);'.format(node_name, Npdu_name))
        npdu_list.append('     sysSetVariableData(sysvar::{0}::{1}::{2}::SignalValue,txDataBuffer,TxLength);'.format(network_name, pdu_name, signal_name))
        npdu_list.append('     // End of TransmitTxBuffer() \n\n     memcpy(txDataBuffer, txDataBuffer_null, txCount); ')
        npdu_list.append('   }')

    elif 'OPAQUE' in pdu_df[col_byte_order].values[0]:
        #if PDU is of type NPDU and not in PDU skip
        npdu_list.append('   byte txDataBuffer[kBufferSizeJumbo_{0}_{1}];'.format(node_name, Npdu_name))
        npdu_list.append('   byte sendString[kBufferSizeJumbo_{0}_{1}];'.format(node_name, Npdu_name))
        npdu_list.append('   byte txDataBuffer_null[kBufferSizeJumbo_{0}_{1}];'.format(node_name, Npdu_name))
        npdu_list.append('   byte emptyString[kBufferSizeJumbo_{0}_{1}];'.format(node_name, Npdu_name))
        npdu_list.append('   long copiedBytes; \n     long TxLength; \n')
        npdu_list.append('   // Preparation same as sysSendData \n     TxLength = 0;')
        npdu_list.append('   if (sysGetVariableData(sysvar::{0}::{1}::{2}::SignalValue, sendString, copiedBytes) == 0)'.format(network_name, pdu_name, signal_name))
        npdu_list.append('   { \n       TxLength=copiedBytes; \n       memcpy( txDataBuffer, sendString, TxLength); \n     } \n')
        npdu_list.append('   // Start of TransmitTxBuffer()')
        npdu_list.append('   CanTpSendData(gHandle_{0}_{1}, txDataBuffer, TxLength);'.format(node_name, Npdu_name))
        npdu_list.append('   sysSetVariableData(sysvar::{0}::{1}::{2}::SignalValue,txDataBuffer,TxLength);'.format(network_name, pdu_name, signal_name))
        npdu_list.append('   // End of TransmitTxBuffer() \n\n     memcpy(txDataBuffer, txDataBuffer_null, txCount); ')
    else:
        #not OPAQUE byte order but NPDU
        npdu_list.append('   byte txDataBuffer[kBufferSizeJumbo_{0}_{1}];'.format(node_name,Npdu_name))
        npdu_list.append('   byte txDataBuffer_null[kBufferSizeJumbo_{0}_{1}];'.format(node_name,Npdu_name))
        npdu_list.append('   txCount = kBufferSizeJumbo_{0}_{1};'.format(node_name,Npdu_name))
        npdu_list.append('   Write_OsekTP_{0}_{1}(txDataBuffer);'.format(node_name,pdu_name))
        npdu_list.append('   // OSEK TP supports up to 4 GiB with Jumbo frames, but limit that to a much lower value for this demo')
        npdu_list.append('   if (txCount > kBufferSizeJumbo_{0}_{1})'.format(node_name,Npdu_name))
        npdu_list.append('   {')
        npdu_list.append('      txCount = kBufferSizeJumbo_{0}_{1};'.format(node_name,Npdu_name))
        npdu_list.append('   }')
        npdu_list.append('   CanTpSendData(gHandle_{0}_{1}, txDataBuffer, txCount);'.format(node_name,Npdu_name))
        npdu_list.append('   memcpy(txDataBuffer, txDataBuffer_null, txCount);')
    npdu_list.append('}')
    return npdu_list
       
def generate_CanTp_ReceptionInd():
    """ """
    filtered_df = node_df.iloc[1:]
    pdu_count = 0
    
    for pdu_type, pdu_type_df in filtered_df.groupby(col_pdu_type):
        if "N-PDU" in pdu_type:
            for pdu, pdu_df in pdu_type_df.groupby(col_pdu):
                if check_df_with_value(pdu_df, col_cantp_pdu_type, 'Flowcontrol_pdu'):
                    Npdu_name = pdu_df[col_payload_pdu].values[0]
                    # try:
                    #     Npdu_name = pdu.rsplit("_", maxsplit=2)[0]
                    # except:
                    #     logger.warning("Error while spliting npdu flow control name")
                    #     continue
                    #filter syvardf by networkname
                    sysvar_nw_df = sysvar_df[sysvar_df[col_network_name]==network_name]
                    osek_fun_cal = []
                    osek_fun_def = []
                    for pdu_type_sysvar, pdu_type_sysvar_df in sysvar_nw_df.groupby(col_pdu_type):
                        if "I-SIGNAL" in pdu_type_sysvar.upper():
                            for pdu_sysvar, pdu_sysvar_df in pdu_type_sysvar_df.groupby(col_pdu):
                                if re.search(Npdu_name, pdu_sysvar, flags=re.I):
                                    pdu_count+=1
                                    npdu_data_node_name = pdu_sysvar_df[col_sender].values[0]
                                    # node.extend(generate_npdu_read(pdu_sysvar, pdu_sysvar_df)) #read function defination
                                    if pdu_count == 1:
                                        node.append('')
                                        node.append('CanTp_ReceptionInd(long connHandle, byte data[])//This function is called when data has been received on the connection with the given handle. The number of bytes received is available as length of the byte array.'.format(node_name))
                                        node.append('{')
                                        node.append('  // Output of Source Address')
                                        node.append('  long rxCount;')
                                        node.append('  byte connId;')
                                        node.append('  ')
                                        node.append('  rxCount = elcount(data);')
                                        node.append('  ')

                                    #reception fuction call
                                    osek_fun_cal.append("")
                                    osek_fun_cal.append('  if (connHandle == gHandle_{0}_{1})'.format(node_name,pdu))
                                    osek_fun_cal.append('  {')
                                    #osek_fun_cal.append('   @sysvar::{0}::{1}::{2}::sysNoOfBytesReceived = rxCount;'.format(network_name,node_name,pdu))
                                    osek_fun_cal.append('   connId = kId_{0}_{1};'.format(node_name,pdu))
                                    osek_fun_cal.append(f'   if ((@hil_ctrl::{dict_hilctrl[input_file][node_name]} == 1) & (@hil_ctrl::{dict_hilctrl[input_file][npdu_data_node_name]} == 1))')
                                    osek_fun_cal.append('   {')
                                    #enable loop back reading of what is sent on bus for debugging
                                    if pdu_name_pattern in pdu:
                                        osek_fun_cal.append(f'   if(@{network_name}_{Npdu_name}::TP_enable_loopback_reading)')
                                        osek_fun_cal.append('   {')
                                        osek_fun_cal.append('       SetDataReceived_{0}_{1}(data, rxCount);'.format(node_name,Npdu_name))
                                        osek_fun_cal.append('   }')
                                    osek_fun_cal.append('   }')
                                    osek_fun_cal.append(f'   else if ((@hil_ctrl::{dict_hilctrl[input_file][node_name]} == 2) & (@hil_ctrl::{dict_hilctrl[input_file][npdu_data_node_name]} == 1))')
                                    osek_fun_cal.append('   {')
                                    #enable loop back reading of what is sent on bus for debugging
                                    if pdu_name_pattern in pdu:
                                        osek_fun_cal.append(f'   if(@{network_name}_{Npdu_name}::TP_enable_loopback_reading)')
                                        osek_fun_cal.append('   {')
                                        osek_fun_cal.append('       SetDataReceived_{0}_{1}(data, rxCount);'.format(node_name,Npdu_name))
                                        osek_fun_cal.append('   }')
                                    osek_fun_cal.append('   }')
                                    osek_fun_cal.append(f'   else if ((@hil_ctrl::{dict_hilctrl[input_file][node_name]} == 0) & (@hil_ctrl::{dict_hilctrl[input_file][npdu_data_node_name]} == 1))')
                                    osek_fun_cal.append('   {')
                                    #enable loop back reading of what is sent on bus for debugging
                                    if pdu_name_pattern in pdu:
                                        osek_fun_cal.append(f'   if(@{network_name}_{Npdu_name}::TP_enable_loopback_reading)')
                                        osek_fun_cal.append('   {')
                                        osek_fun_cal.append('       SetDataReceived_{0}_{1}(data, rxCount);'.format(node_name,Npdu_name))
                                        osek_fun_cal.append('   }')
                                    osek_fun_cal.append('   }')
                                    osek_fun_cal.append(f'   else if ((@hil_ctrl::{dict_hilctrl[input_file][node_name]} == 2) & (@hil_ctrl::{dict_hilctrl[input_file][npdu_data_node_name]} == 2))')
                                    osek_fun_cal.append('   {')
                                    # try:
                                    osek_fun_cal.append('   SetDataReceived_{0}_{1}(data, rxCount);'.format(node_name,Npdu_name))
                                    # except:
                                    #     logger.warning(' No data to split for NPDU')
                                    osek_fun_cal.append('   }')
                                    osek_fun_cal.append(f'   else if ((@hil_ctrl::{dict_hilctrl[input_file][node_name]} == 1) & (@hil_ctrl::{dict_hilctrl[input_file][npdu_data_node_name]} == 2))')
                                    osek_fun_cal.append('   {')
                                    # try:
                                    osek_fun_cal.append('   SetDataReceived_{0}_{1}(data, rxCount);'.format(node_name,Npdu_name))
                                    # except:
                                    #     logger.warning(' No data to split for NPDU')
                                    osek_fun_cal.append('   }')
                                    osek_fun_cal.append('  }')
                                    osek_fun_cal.append("")
                    
                    if(pdu_count>0):
                        # generate_Set_Read_Number_of_signals(network_name,pdu_df[col_payload_pdu].values[0])
                        # node.append('')
                        # node.append('CanTp_ReceptionInd(long connHandle, byte data[])//This function is called when data has been received on the connection with the given handle. The number of bytes received is available as length of the byte array.'.format(node_name))
                        # node.append('{')
                        # node.append('  // Output of Source Address')
                        # node.append('  long rxCount;')
                        # node.append('  byte connId;')
                        # node.append('  ')
                        # node.append('  rxCount = elcount(data);')
                        # node.append('  ')
                        node.extend(osek_fun_cal)
                        # node.append('   //write("%s, connection %d: Reception indication for %d bytes", "{0}", connId, rxCount);'.format(network_name))
                        # node.append('}')
                        node.extend(osek_fun_def)
                        logger.info(f"void: generate_CanTp_ReceptionInd with {str(pdu_count)} pdu.")
            if (pdu_count > 0):
                node.append('   //write("%s, connection %d: Reception indication for %d bytes", "{0}", connId, rxCount);'.format(network_name))
                node.append('}')

            for pdu, pdu_df in pdu_type_df.groupby(col_pdu):
                if check_df_with_value(pdu_df, col_cantp_pdu_type, 'Flowcontrol_pdu'):
                    Npdu_name = pdu_df[col_payload_pdu].values[0]
                    # try:
                    #     Npdu_name = pdu.rsplit("_", maxsplit=2)[0]
                    # except:
                    #     logger.warning("Error while spliting npdu flow control name")
                    #     continue
                    #filter syvardf by networkname
                    sysvar_nw_df = sysvar_df[sysvar_df[col_network_name]==network_name]
                    for pdu_type_sysvar, pdu_type_sysvar_df in sysvar_nw_df.groupby(col_pdu_type):
                        if "I-SIGNAL" in pdu_type_sysvar.upper():
                            for pdu_sysvar, pdu_sysvar_df in pdu_type_sysvar_df.groupby(col_pdu):
                                if re.search(Npdu_name, pdu_sysvar, flags=re.I):
                                    pdu_count+=1
                                    npdu_data_node_name = pdu_sysvar_df[col_sender].values[0]
                                    node.extend(generate_npdu_read(pdu_sysvar, pdu_sysvar_df)) #read function defination
                    if (pdu_count > 0):
                        read_sig_para_dict = {'network_name':network_name,
                                              'pdu':pdu_df[col_payload_pdu].values[0],
                                              'node_name':node_name}
                        node.extend(generate_Set_Read_Number_of_signals(read_sig_para_dict))

       
def generate_OnStart_callTx():
    """ """
    filtered_df = node_df.iloc[1:]
    pdu_count = 0
    for pdu_type, pdu_type_df in filtered_df.groupby(col_pdu_type):
        if ('N-PDU' in pdu_type.upper()):
            node.append('  OnStart_init_CANTP_{0}();'.format(node_name))
            for payload_pdu, payload_pdu_df in pdu_type_df.groupby(col_payload_pdu):
                if not(is_empty_or_spaces(payload_pdu)):
                    node.append('  OnStart_callTx_{0}_{1}();'.format(node_name, payload_pdu_df[col_pdu].values[0]))
                    if check_df_with_value(payload_pdu_df, col_cantp_pdu_type, 'Flowcontrol_pdu') and payload_pdu in pdu_skip:
                        node.append('  CallTx_{0}_{1}_hil_ctrl_{2}();'.format(node_name, payload_pdu_df[col_pdu].values[0], hil_ctrl_node_dict[node_name]))
                    pdu_count += 1
                    
    if(pdu_count>0):
        logger.info(f"void: generate_OnStart_callTx with {str(pdu_count)} pdu.")

def generate_OnStart_callTx_FC(node_name,pdu):
    """
    

    Args:
      node_name: 
      pdu: 

    Returns:

    """
    node.append('//  OnStart_callTx_{0}_{1}();'.format(node_name,pdu))

def generate_npdu_read_bufferwise(pdu_name,buffer_fullname, buffer_df):
    """
    generate read received npdu data bufferwise

    Args:
      pdu_name: 
      buffer_fullname: 
      buffer_df: 

    Returns:

    """
    read_list = []
    read_list.append("")
    read_list.append(f"void {sender_setDataReceived}_{buffer_fullname}_SetDataReceived(byte byteArray[])")
    read_list.append("{")
    read_list.append("    qword raw_value;")
    
    for i, row in buffer_df.iterrows():
        row_data = read_values_from_row(row)
        signal_name_split = row_data["signal_name"].split("_")
        read_list.append("")
        if "MOTO" in row_data['byte_order'].upper():
            read_list.append(f"    raw_value=copyBitsFromByteArrayBE( byteArray, {row_data['start_bit']}, {row_data['bit_length']});")
            read_list.append(f"    @{network_name}::{row_data['pdu_name']}::{'_'.join(signal_name_split[:-1])}.{signal_name_split[-1]} = convertRawToPhysical(raw_value, {row_data['factor']}, {row_data['offset']});")
        elif "INTEL" in row_data['byte_order'].upper():
            read_list.append(f"    raw_value=copyBitsFromByteArrayLE( byteArray, {row_data['start_bit']}, {row_data['bit_length']});")
            read_list.append(f"    @{network_name}::{row_data['pdu_name']}::{'_'.join(signal_name_split[:-1])}.{signal_name_split[-1]} = convertRawToPhysical(raw_value, {row_data['factor']}, {row_data['offset']});")
    read_list.append("}")
    read_list.append("")
    node.extend(read_list)

def generate_npdu_read(pdu_name, pdu_df):
    """
    

    Args:
      pdu_name: 
      pdu_df: 

    Returns:

    """
    global sender_setDataReceived
    Npdu_name = ""
    for key, value in npdu_dict.items():
        if value[1] == pdu_name:#Value[1]-->I-signal Pdu #Value[0]-->NPdu
            Npdu_name = pdu_name
            break
    read_list = []
    read_list.append("")
    read_list.append(f"void SetDataReceived_{node_name}_{Npdu_name}(byte byteArray[], long count)//I-signal-i pdu : {pdu_name}->Mapped to ->N-pdu : {Npdu_name}")
    read_list.append("{")
    read_list.append("    qword raw_value;")

    #creates a new column to know which buffer it is, later we can groupby and use it
    pdu_df["BUFFER"] = pdu_df[col_name].apply(lambda x: x.split("_")[2] if re.match(Npdu_sig_patterns, x, flags=re.I) else "-")

    for buffer, buffer_df in pdu_df.groupby("BUFFER"):
        if buffer=="-":
            for i, row in buffer_df.iterrows():
                row_data = read_values_from_row(row)
                read_list.append("")

                if "MOTO" in row_data['byte_order'].upper():
                    read_list.append(f"    raw_value=copyBitsFromByteArrayBE( byteArray, {row_data['start_bit']}, {row_data['bit_length']});")
                    read_list.append(f"    @{network_name}::{row_data['pdu_name']}::{row_data['signal_name']} = convertRawToPhysical(raw_value, {row_data['factor']}, {row_data['offset']});")
                elif "INTEL" in row_data['byte_order'].upper():
                    read_list.append(f"    raw_value=copyBitsFromByteArrayLE( byteArray, {row_data['start_bit']}, {row_data['bit_length']});")
                    read_list.append(f"    @{network_name}::{row_data['pdu_name']}::{row_data['signal_name']} = convertRawToPhysical(raw_value, {row_data['factor']}, {row_data['offset']});")
                elif "OPAQUE" in row_data['byte_order'].upper():
                    read_list.append(f"    @{network_name}::{row_data['pdu_name']}::rxBuffercount = count;")
                    read_list.append(f"    sysSetVariableArraySize(sysvar::{network_name}::{row_data['pdu_name']}::{row_data['signal_name']}::SignalValue, count);")
                    read_list.append(f"    sysSetVariableData(sysvar::{network_name}::{row_data['pdu_name']}::{row_data['signal_name']}::SignalValue, byteArray, count);")
        else:
            buffer_fullname = "_".join(buffer_df[col_name].values[0].split("_")[:-1])
            sender_setDataReceived = buffer_df[col_sender].values[0]
            generate_npdu_read_bufferwise(pdu_name, buffer_fullname, buffer_df)

            read_list.append("")
            read_list.append(f"    if(@{network_name}::{row_data['pdu_name']}::{buffer_fullname}.Read == 1)")
            read_list.append("    {")
            read_list.append(f"        {sender_setDataReceived}_{buffer_fullname}_SetDataReceived(byteArray);")
            read_list.append("    }")

    read_list.append("}")
    read_list.append("")
    return read_list


# main function to create nodes
def create_node_main(workbook, sheet, path):
    """
    

    Args:
      workbook: 
      sheet: 
      path: 

    Returns:

    """
    logger.info("++++++++++Creating node for: {0} ++++++++++".format(sheet))
    # global node_osektp, cinfile_list, pdu_list
    # cinfile_list = []
##    global pdu_list
##    pdu_list = []

    try:
        generate_dataframe(workbook, sheet)
        '''
        for ind in node_df.index:
            # if "NPDU" in node_df[col_pdu][ind].upper().split('_'):
            # if osektp_pdu_pattern in node_df[col_pdu][ind].split('_'):
            pdu_pattern_check = [True if re.search(osektp_pdu_pattern, i) else False for i in
                                    node_df[col_pdu][ind].split('_')]
            if (True in pdu_pattern_check):
                pdu_list.append(node_df[col_pdu][ind])
        #generate_includes(cinfile_list)
        '''
        generate_OsekTP_cin(path)
        generate_opaque_byte_order_cin(path)
        generate_includes()
        generate_variables()
        generate_onprestart()
        generate_onstart()
        generate_onstart_ini_CANTP()
        generate_onstop()
        generate_onsim()
        generate_onvariant()
        generate_customer_functions()
        generate_onbus()
        generate_void_sim_Reset()
        generate_void_sim_start()
        generate_void_sim_Disable()
        generate_void_sim_DisableMsg()
        generate_void_sim_EventMsg()
        generate_void_sim_SetCANParameter()
        generate_void_sim_SetCANFDParameter()
        generate_dword_applILTx()
        generate_ontmr()
        generate_on_CycleTime()
        generate_void_sim_SetCycleTimeOffset()
        generate_onmsg()
        generate_void_OnSecurityLocalPDUPreTx()
        generate_CanTp_ReceptionInd()
        save_file(path)
        return True
    except Exception as exp:
        logger.error(f"Failed to create node for: {sheet}")
        traceback.print_exc()
        logger.error(f"Error: {exp}")
        raise exp

def generate_sysvardatabase_dataframe(wb, sht_name):
    """
    reads sysvarsheet and store as global sheet

    Args:
      wb: 
      sht_name: 

    Returns:

    """
    global sysvar_df, col_sender, col_network_name, npdu_dict, securedPdu_dict, npdu_datalength_dict

    ws_sysvar = wb[sht_name]

    # create the different dataframes
    sysvar_df = pd.DataFrame(ws_sysvar.values)
    sysvar_df = sysvar_df.replace(np.nan, '', regex=True)
    # get column names
    col_names = list(sysvar_df.iloc[0])
    col_sender = col_names.index("sender")
    col_network_name = col_names.index("network_name")
    col_msg = col_names.index("Message")
    col_pdu = col_names.index("PDU")
    col_pdu_type = col_names.index("PDU_Type")
    col_payload_pdu = col_names.index("Payload_PDU")
    col_payload_pdu_type = col_names.index("Payload_PDU_Type")
    col_start_bit = col_names.index("Startbit")
    col_signal_length_bit = col_names.index("Signal_Length [Bit]")
    col_byte_order = col_names.index("Byte Order")
    col_pdu_length = col_names.index("PDU_Length [Byte]")
    npdu_dict = map_npdu_to_ipdu_dict_sysvar(sysvar_df, col_sender, col_msg, col_pdu, col_pdu_type, col_payload_pdu)
    securedPdu_dict = map_securedpdu_to_ipdu_dict_sysvar(sysvar_df, col_sender, col_msg, col_pdu, col_pdu_type, col_payload_pdu)
    npdu_datalength_dict = generateNpduDatalength(sysvar_df, col_sender, col_msg, col_pdu, col_pdu_type, col_payload_pdu_type,col_pdu_length, col_byte_order, col_start_bit, col_signal_length_bit)


def initialize_dict(wb):
    """
    

    Args:
      wb: 

    Returns:

    """
    global dict_network_info, dict_entity, dict_hilctrl, dict_network_name, dict_network_type, dict_block_messages, dict_pass_messages
    dict_network_info = getMappingData(wb, "network_category")
    dict_network_name = getMappingData(wb, "network_name")
    dict_network_type = getMappingData(wb, "network_type")
    dict_entity = getMappingData(wb, "node_entity")
    dict_hilctrl = extractHilCtrl(wb)
    dict_blk_msg = getMappingData(wb, "block_messages")
    dict_block_messages = {}
    for st, value in dict_blk_msg.items():
        if value!='na':
            value = value.strip().split(";")
            value = [i.strip().upper() for i in value]
            value = ['0X' + i if not (i.upper().startswith('0X')) else i for i in value]
            dict_block_messages[st] = value
    

    dict_pass_msg = getMappingData(wb, "pass_messages")
    dict_pass_messages = {}
    for st, value in dict_pass_msg.items():
        if value!='na':
            value = value.strip().split(";")
            value = [i.strip().upper() for i in value]
            value = ['0X'+i if not(i.upper().startswith('0X')) else i for i in value]
            dict_pass_messages[st] = value
    


def external_call():
    """ """
    try:
        logger.info(f"###### START 'Create Nodes(.can)' DEBUG INFORMATION ######")
        script_path = os.path.dirname(os.path.abspath(__file__))
        autosar_path = script_path + r'/../../../../CustomerPrj/Restbus/Autosar_Gen_Database.xlsx'
        workbook = load_workbook(autosar_path, data_only=True)
        initialize_dict(workbook)
        generate_sysvardatabase_dataframe(workbook,"SysVarDatabase")
        sheet_list = [name for name in workbook.sheetnames if len(name.split('_')) > 2]

        for sheet in sheet_list:
            if (len(sheet.split('_')) > 2) and ("can_" in sheet.lower()) and ("sysvardatabase" not in sheet.lower()):  # if it is an ECU sheet
                create_node_main(workbook, sheet, script_path + r'/../../../../CustomerPrj/Restbus/Nodes')
    except Exception as e:
        logger.error(f"Create Nodes(.can) failed. Exception --> {e}")
        raise Exception(e)
    logger.info("###### END 'Create Nodes(.can)' DEBUG INFORMATION ######")
    logger.info('-' * 80)


if __name__ == "__main__":
    external_call()
