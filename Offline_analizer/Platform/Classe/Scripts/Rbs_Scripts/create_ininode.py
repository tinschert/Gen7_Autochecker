# -*- coding: utf-8 -*-
# @file create_ininode.py
# @author ADAS_HIL_TEAM
# @date 08-21-2023

##################################################################
# C O P Y R I G H T S
# ----------------------------------------------------------------
# Copyright (c) 2023 by Robert Bosch GmbH. All rights reserved.

# The reproduction, distribution and utilization of this file as
# well as the communication of its contents to others without express
# authorization is prohibited. Offenders will be held liable for the
# payment of damages. All rights reserved in the event of the grant
# of a patent, utility model or design.

##################################################################

"""used to create initialization code for the variables and signals"""

import re

from openpyxl import load_workbook
import pandas as pd
try:
    pd.set_option('future.no_silent_downcasting', True)
except:
    pass
import numpy as np
import os, sys
import traceback
try:
    from Rbs_Scripts.create_nodes import extractHilCtrl, getCopyRights
except ImportError:
    sys.path.append(os.path.dirname(os.path.abspath(__file__)))
    from create_nodes import extractHilCtrl, getCopyRights

try:
    sys.path.append(os.getcwd() + r"\..\..\CustomerPrj\Restbus\Scripts")
    from pattern_matching import *
except ImportError:
    sys.path.append(os.path.dirname(os.path.abspath(__file__)) + r"\..\..\..\..\CustomerPrj\Restbus\Scripts")
    from pattern_matching import *

try:
    from Control.logging_config import logger
except ImportError:
    sys.path.append(os.path.dirname(os.path.abspath(__file__)) + r"\..\Control")
    from logging_config import logger

try:
    from create_autosar import extract_vehicle_variants
except:
    from Rbs_Scripts.create_autosar import extract_vehicle_variants

# allow longer names up to 100 characters
pd.options.display.max_colwidth = 100
pd.options.display.colheader_justify = 'left'

# Case create simulated RBS & Bosch ECUs - Gateway is excluded as the initialization of the variables is done as per sending node.
# node_list = ['Rbs', 'FRadar', 'FVideo', 'Map', 'CRadarFL', 'CRadarFR', 'CRadarRL', 'CRadarRR']
node_list = []
ignore_nodes = ["Pass"]

# for updating initialization.can. If sheets are empty then nodes are ignored here
rbs_init_node_list = []
classe_init_node_list = ["classe", "customer"]
mapping_sheets = ["DBCmapping", "ArxmlMapping"]
skip_pdu_type_list = ['NM-PDU', 'MULTIPLEXED-I-PDU', 'GENERAL-PURPOSE-PDU', 'GENERAL-PURPOSE-I-PDU', 'N-PDU', 'SECURED-I-PDU']

def define_colIndex_ini_rbs(col_names, variant_columns):
    """
    declare index of the rbs columns as global

    Args:
      col_names (list): column names
      variant_columns (list): variant column list

    """
    global col_name, col_group, col_pdu, col_pdu_type, col_payload_pdu, col_payload_pdu_type,col_msg, col_signal_length, col_byte_order, col_init_value, col_sender, col_entity, col_nwt_name, col_multicanoe, col_texttable_values, variant_columns_index_dict
    variant_columns_index_dict = {}
    try:
        col_name = col_names.index("Signal")  # Signal
        col_group = col_names.index("Signal_Group")  # group
        col_pdu = col_names.index("PDU")  # pdu
        col_pdu_type = col_names.index("PDU_Type")  # pdu type
        col_payload_pdu = col_names.index("Payload_PDU")  # pdu type
        col_payload_pdu_type = col_names.index("Payload_PDU_Type")  # payload pdu type
        col_msg = col_names.index("Message")  # Message
        col_signal_length = col_names.index("Signal_Length [Bit]")  # Length [Bit]
        col_byte_order = col_names.index("Byte Order")  # Byte Order
        col_init_value = col_names.index("Initial Value")
        col_sender = col_names.index("sender")
        col_entity = col_names.index("node_entity")
        col_nwt_name = col_names.index("network_name")
        col_multicanoe = col_names.index("multicanoe")
        col_texttable_values = col_names.index("texttable values")
        for varinat in variant_columns:
            variant_columns_index_dict[varinat] = col_names.index(varinat)
       
    except Exception as e:
        logger.error(
            f"{e} SysVarDatabase::Column 'Name' or 'sender' or 'Message' or bus_type, or input_file is not found in Excel. Please check it in Autosar_Gen_Database.xlsx")
        raise e
    
def define_colIndex_ini_rbs_ethernet(col_names, variant_columns):
    """
    declare index of the rbs columns as global

    Args:
      col_names (list): column names
      variant_columns (list): variant column list

    """
    global col_name, col_group, col_pdu, col_pdu_type, col_payload_pdu, col_payload_pdu_type,col_signal_length, col_byte_order, col_init_value, col_sender, col_nwt_name, col_nwt_protocol, variant_columns_index_dict
    variant_columns_index_dict = {}
    try:
        col_name = col_names.index("Signal")  # Signal
        col_group = col_names.index("Signal_Group")  # group
        col_pdu = col_names.index("PDU")  # pdu
        col_pdu_type = col_names.index("PDU_Type")  # pdu type
        col_payload_pdu = col_names.index("Payload_PDU")  # pdu type
        col_payload_pdu_type = col_names.index("Payload_PDU_Type")  # payload pdu type
        col_signal_length = col_names.index("Signal_Length [Bit]")  # Length [Bit]
        col_byte_order = col_names.index("Byte Order")  # Byte Order
        col_init_value = col_names.index("Initial Value")
        col_sender = col_names.index("Node Name")
        col_nwt_name = col_names.index("Network Name")
        col_nwt_protocol  = col_names.index("Network Protocol")
        for varinat in variant_columns:
            variant_columns_index_dict[varinat] = col_names.index(varinat)
       
    except Exception as e:
        logger.error(
            f"{e} SysVarDatabaseETH::Column 'Name' or 'sender' or 'Message' or bus_type, or input_file is not found in Excel. Please check it in Autosar_Gen_Database.xlsx")
        raise e
    


def generate_dataframe_rbs(wb, sysVarSheet, variant_columns):
    """
    generates pandas dataframe of sysvar databse sheet for RBS

    Args:
      wb (workbook): excel workbook
      sysVarSheet (str): sysvar sheet name
      variant_columns (list): list of variant column names

    Returns:
        dataframe of sysvar sheet

    """

    ini_df = pd.DataFrame(wb[sysVarSheet].values)
    sysvar_cols = list(ini_df.iloc[0])
    define_colIndex_ini_rbs(sysvar_cols, variant_columns)

    node_list.clear()
    # generate node list
    node_list.extend(list(ini_df[col_sender].loc[1:].unique()))
    for i in node_list:
        if i in ignore_nodes:
            node_list.remove(i)

    #filter df based on entity
    unique_entity = list(ini_df[col_entity].loc[1:].unique())
    df_list = []
    for entity in unique_entity:
        col_namespace = getColNamespace(entity)
        temp_df = ini_df[ini_df[col_entity] == entity]
        temp_df = filter_column_df(temp_df, col_namespace, col_name)
        df_list.append(temp_df)
    ini_df = pd.concat(df_list, axis=0)
    ini_df.loc[-1] = list(sysvar_cols)  # adding a column row
    ini_df.index = ini_df.index + 1  # shifting index
    ini_df = ini_df.sort_index()
    return ini_df

def generate_dataframe_rbs_ethernet(wb, sysVarSheet, variant_columns):
    """
    generates pandas dataframe of sysvar databse sheet for RBS ETHERNET

    Args:
      wb (workbook): excel workbook
      sysVarSheet (str): sysvar sheet name
      variant_columns (list): list of variant column names

    Returns:
        dataframe of sysvar sheet

    """
    try:
        try:
            ini_df = pd.DataFrame(wb[sysVarSheet].values)
        except KeyError:
            return None
        if len(ini_df) <= 1:
            return None
        
        sysvar_cols = list(ini_df.iloc[0])
        define_colIndex_ini_rbs_ethernet(sysvar_cols, variant_columns)

        node_list.clear()
        # generate node list
        node_list.extend(list(ini_df[col_sender].loc[1:].unique()))
        for i in node_list:
            if i in ignore_nodes:
                node_list.remove(i)

        #filter df based on entity
        ini_df = ini_df[ini_df[col_nwt_protocol] == 'ETH_PDU']
        ini_df = filter_column_df(ini_df, col_pdu, col_name)
        ini_df.loc[-1] = list(sysvar_cols)  # adding a column row
        ini_df.index = ini_df.index + 1  # shifting index
        ini_df = ini_df.sort_index()
        return ini_df
    except Exception as e:
        logger.error(f"Error in reading sysvar database sheet {sysVarSheet} for creating dataframe in RBS ETHERNET")
        raise e
    
def generate_ininode_senderwise_ethernet(ini_df, variant):
    """
    generate ininode senderwise

    Args:
      ini_df (dataframe): sysvar dataframe
      variant (str): variant name

    Returns:
        list: file content for the given node

    """
    global signal_count, opaque_byte_order_sig_dict
    flag_count=1
    ininode = []
    opaque_byte_order_sig_dict = {}
    signal_count = [0] * len(node_list)
    check_duplicates_columns = [col_name, col_pdu, col_nwt_name]  # in a network
    df_duplicates = ini_df[ini_df.duplicated(subset=check_duplicates_columns)]
    if not (df_duplicates.empty):
        for i, row in df_duplicates.iterrows():
            logger.warning(f"{row[col_name]} signal DUPLICATED in {col_nwt_name} network")

    ini_df = ini_df.drop_duplicates(subset=check_duplicates_columns, keep="first")

    for node in node_list:
        ininode_main = []
        ininode_sender=[]
        ininode_main.append('void ' + 'init_{0}_ETH_{1}()'.format(node,variant))
        ininode_main.append('{')
        ini_df_filter_sender = ini_df[ini_df[col_sender] == node]
        signal_count[node_list.index(node)] = ini_df_filter_sender.shape[0]
        ininode_sender += generate_ininode_ethernet(ini_df_filter_sender, variant_columns_index_dict[variant], node)
        ini_count = 0
        line_count = 0
        for line in ininode_sender:
            if line_count%25==0:
                ini_count += 1
                if ini_count!=1:
                    ininode_main.append('  }')
                ininode_main.append('  if(init_count == {0})'.format(ini_count))
                ininode_main.append('  {')
            ininode_main.append(line)
            line_count += 1
        if len(ininode_sender)!=0:
            ininode_main.append('  if (prestart_mode==0) init_count = 0;')
            flag_count += 1
            ininode_main.append('  initFlag = {0};'.format(flag_count))
            ininode_main.append('  }')
            ininode_main.append('  init_count++;')
        ininode_main.append('}')
        ininode += ininode_main
    logger.info("=============== Summary ===============")
    logger.info("The number of signals of each node:")
    logger.info("---------------")
    logger.info("{:<10} {:<5}".format("Node", "Count"))
    for node, count in zip(node_list, signal_count):
        logger.info("{:<10} {:<5}".format(node, str(count)))
    return ininode

def generate_variables_ethernet(ini_node, flag_opaque):
    variables = []
    variables.append('variables \n{')
    if opaque_byte_order_sig_dict != {} and flag_opaque:
        variables.append('  int i;')
        for key, val in opaque_byte_order_sig_dict.items():
            variables.append('  byte ' + key + '_byte_Array[' + val + '];')
    variables.append('} \n')
    ini_node = variables + ini_node
    return ini_node



def generate_ininode_ethernet(ini_df, varinat_column_index, node=""):
    """
    here ini node logic is created

    Args:
      ini_df (dataframe): sysvar dataframe
      varinat_column_index (int): column index
      node (str):  (Default value = "") node name

    Returns:
        list: generated logic

    """
    # count the number of signals
    # flag count for the node number
    # set of 25 signals batch
    signal_batch = 25

    # Case create restbus ini
    ininode = []
    # get node name and dataframe
    #node_df = ini_df[ini_df[col_init_value] != 'na']
    num_of_rows_df = ini_df.shape[0]

    # node check: in case the node name is different from SysVarDatabase
    if ini_df.empty:
        logger.warning(f"Node '{node}' is Empty in Excel, please check it in Autosar_Gen_Database.xlsx.")
        return []
        # input("press any key to continue:")
    if node+'_ETH' not in rbs_init_node_list:
        rbs_init_node_list.append(node+'_ETH')
    # print('we are on node: ' + str(node) + ', and we have the num of rows: ' + str(num_of_rows_df))

    for i in range(num_of_rows_df):
        # if (i*100//num_of_rows_df) % 10 == 0:       # restrain output
        #    print('\tprocess completed: ' + "%.1f" % (i*100/num_of_rows_df) + '%')
        signal_name = ini_df[col_name].values[i]
        signal_byte_order = ini_df[col_byte_order].values[i]
        signal_length = ini_df[col_signal_length].values[i]
        
        #filter to skip null signal initialization
        if (signal_name=='-') or(signal_name and "null_" in str(signal_name).lower()):
            continue
        
        namespace = ini_df[col_pdu].values[i]
        pdu_type = str(ini_df[col_pdu_type].values[i]).upper()

        if (pdu_type in skip_pdu_type_list):  # skip all special pdu initialization
            continue


        network_name = ini_df[col_nwt_name].values[i]
        try:
            network_name = dict_network_excelNaming_canoeNaming.get(network_name, network_name)
        except:
            pass
        variant_value = ini_df[varinat_column_index].values[i]
        if variant_value == None:
            variant_value = 0

        # case new set of 25 signals
        if i % signal_batch == 0:
            # except for the first signal
            if i != 0:
                pass
                #ininode.append('  }')
            ini_count = i // signal_batch + 1
            #ininode.append('  if(init_count == {0})'.format(ini_count))
            #ininode.append('  {')
        
        if "OPAQUE" in signal_byte_order.upper():
            pass
            # ininode.append('  sysSetVariableData(sysvar::{0}::{1}::{2}::SignalValue, {2}_byte_Array, {3});'.format(network_name, namespace, signal_name, signal_length // 8))
            # if signal_name not in opaque_byte_order_sig_dict.keys():
            #     opaque_byte_order_sig_dict[signal_name] = str(signal_length // 8)
        else:
            ininode.append('  ${0}::{1}::{2} = {3};'.format(network_name, namespace, signal_name, variant_value))

    return ininode


def getColNamespace(entity):
    """
    takes entity name as input and return column index

    Args:
      entity (str): entity name msg/pdu/grp

    Returns:
        int: index of column
    """
    if "message" in entity:
        return col_msg
    elif "pdu" in entity:
        return col_pdu
    elif "group" in entity:
        return col_group
    else:
        return col_msg

def generate_dataframe_classe(wb, node):
    """
    generates pandas dataframe of CLASSE excel sheets

    Args:
      wb (workbook): classe database excel
      node (str): node name CLASSE

    Returns:
        combined dataframe

    """
    # add classe to node_list
    node_list.clear()
    node_list.append(node)
    ini_df_combined = pd.DataFrame()
    num_of_sheets = len(wb.worksheets)
    for sheet_index in range(num_of_sheets):
        sheet = wb.worksheets[sheet_index]
        ws2 = wb[sheet.title]
        if sheet_index != 0:
            ws2.delete_rows(1)
        ini_df = pd.DataFrame(ws2.values)
        ini_df_combined = pd.concat([ini_df_combined, ini_df])
    return ini_df_combined


def get_namespace_col(bus_type="", input_file=""):
    """
    not used in this file

    Args:
      bus_type:  (Default value = "")
      input_file:  (Default value = "")

    Returns:

    """
    if (grp_or_pdu_or_msg(bus_type, input_file) == 'grp'):
        return col_group  # group
    elif (grp_or_pdu_or_msg(bus_type, input_file) == 'pdu'):
        return col_pdu  # pdu
    elif (grp_or_pdu_or_msg(bus_type, input_file) == 'msg'):
        return col_msg  # Message
    else:
        return col_msg  # default



def generate_ininode_multicanoe_senderwise(ini_df, variant, multicanoe_category, multicanoe_node_list):
    """
    generate ininode multicanoe senderwise

    Args:
      ini_df (dataframe): sysvar dataframe
      variant (str): variant name
      multicanoe_category (str): master/slave
      multicanoe_node_list (list): node list

    Returns:
        list: file content for the given node
    """
    global signal_count_multicanoe
    flag_count=1
    ininode = []
    opaque_byte_order_sig_dict = {}
    signal_count_multicanoe = [0] * len(multicanoe_node_list)
    check_duplicates_columns = [col_name, col_msg, col_nwt_name]  # in a network
    df_duplicates = ini_df[ini_df.duplicated(subset=check_duplicates_columns)]
    if not (df_duplicates.empty):
        for i, row in df_duplicates.iterrows():
            logger.warning(f"{row[col_name]} signal DUPLICATED in {col_nwt_name} network")

    ini_df = ini_df.drop_duplicates(subset=check_duplicates_columns, keep="first")
    ini_df = ini_df[ini_df[col_multicanoe] == multicanoe_category]
    for node in multicanoe_node_list:
        ininode_main = []
        ininode_sender=[]
        ininode_main.append('void ' + 'init_{0}_{1}()'.format(node,variant))
        ininode_main.append('{')
        ini_df_filter_sender = ini_df[ini_df[col_sender] == node]
        signal_count_multicanoe[multicanoe_node_list.index(node)] = ini_df_filter_sender.shape[0]
        ininode_sender += generate_ininode(ini_df_filter_sender, variant_columns_index_dict[variant], node)
        ini_count = 0
        line_count = 0
        for line in ininode_sender:
            if line_count%25==0:
                ini_count += 1
                if ini_count!=1:
                    ininode_main.append('  }')
                ininode_main.append('  if(init_count == {0})'.format(ini_count))
                ininode_main.append('  {')
            ininode_main.append(line)
            line_count += 1
        if len(ininode_sender)!=0:
            ininode_main.append('  if (prestart_mode==0) init_count = 0;')
            flag_count += 1
            ininode_main.append('  initFlag = {0};'.format(flag_count))
            ininode_main.append('  }')
            ininode_main.append('  init_count++;')
        ininode_main.append('}')
        ininode += ininode_main
    logger.info("=============== Summary ===============")
    logger.info("The number of signals of each node:")
    logger.info("---------------")
    logger.info("{:<10} {:<5}".format("Node", "Count"))
    for node, count in zip(node_list, signal_count_multicanoe):
        logger.info("{:<10} {:<5}".format(node, str(count)))
    return ininode



def generate_ininode_senderwise(ini_df, variant):
    """
    generate ininode senderwise

    Args:
      ini_df (dataframe): sysvar dataframe
      variant (str): variant name

    Returns:
        list: file content for the given node

    """
    global signal_count, opaque_byte_order_sig_dict
    flag_count=1
    ininode = []
    opaque_byte_order_sig_dict = {}
    signal_count = [0] * len(node_list)
    check_duplicates_columns = [col_name, col_msg, col_nwt_name]  # in a network
    df_duplicates = ini_df[ini_df.duplicated(subset=check_duplicates_columns)]
    if not (df_duplicates.empty):
        for i, row in df_duplicates.iterrows():
            logger.warning(f"{row[col_name]} signal DUPLICATED in {col_nwt_name} network")

    ini_df = ini_df.drop_duplicates(subset=check_duplicates_columns, keep="first")

    for node in node_list:
        ininode_main = []
        ininode_sender=[]
        ininode_main.append('void ' + 'init_{0}_{1}()'.format(node,variant))
        ininode_main.append('{')
        ini_df_filter_sender = ini_df[ini_df[col_sender] == node]
        signal_count[node_list.index(node)] = ini_df_filter_sender.shape[0]
        ininode_sender += generate_ininode(ini_df_filter_sender, variant_columns_index_dict[variant], node)
        ini_count = 0
        line_count = 0
        for line in ininode_sender:
            if line_count%25==0:
                ini_count += 1
                if ini_count!=1:
                    ininode_main.append('  }')
                ininode_main.append('  if(init_count == {0})'.format(ini_count))
                ininode_main.append('  {')
            ininode_main.append(line)
            line_count += 1
        if len(ininode_sender)!=0:
            ininode_main.append('  if (prestart_mode==0) init_count = 0;')
            flag_count += 1
            ininode_main.append('  initFlag = {0};'.format(flag_count))
            ininode_main.append('  }')
            ininode_main.append('  init_count++;')
        ininode_main.append('}')
        ininode += ininode_main
    logger.info("=============== Summary ===============")
    logger.info("The number of signals of each node:")
    logger.info("---------------")
    logger.info("{:<10} {:<5}".format("Node", "Count"))
    for node, count in zip(node_list, signal_count):
        logger.info("{:<10} {:<5}".format(node, str(count)))
    return ininode


def generate_ininode(ini_df, varinat_column_index, node=""):
    """
    here ini node logic is created

    Args:
      ini_df (dataframe): sysvar dataframe
      varinat_column_index (int): column index
      node (str):  (Default value = "") node name

    Returns:
        list: generated logic

    """
    unique_entity = list(ini_df[col_entity].unique())
    #col_namespace = getColNamespace(unique_entity[0]) if len(unique_entity)==1 else False
    # count the number of signals
    # flag count for the node number
    # set of 25 signals batch
    signal_batch = 25

    # Case create restbus ini
    ininode = []
    # get node name and dataframe
    #node_df = ini_df[ini_df[col_init_value] != 'na']
    num_of_rows_df = ini_df.shape[0]

    # node check: in case the node name is different from SysVarDatabase
    if ini_df.empty:
        logger.warning(f"Node '{node}' is Empty in Excel, please check it in Autosar_Gen_Database.xlsx.")
        return []
        # input("press any key to continue:")
    if node not in rbs_init_node_list:
        rbs_init_node_list.append(node)
    # print('we are on node: ' + str(node) + ', and we have the num of rows: ' + str(num_of_rows_df))
    #npdu_dict = map_npdu_to_ipdu_dict_sysvar(ini_df, col_sender, col_msg, col_pdu ,col_pdu_type, col_payload_pdu, column_indexed=False)

    for i in range(num_of_rows_df):
        #if col_namespace == False:
        col_namespace = getColNamespace(ini_df[col_entity].values[i])
        # if (i*100//num_of_rows_df) % 10 == 0:       # restrain output
        #    print('\tprocess completed: ' + "%.1f" % (i*100/num_of_rows_df) + '%')
        signal_name = ini_df[col_name].values[i]
        signal_byte_order = ini_df[col_byte_order].values[i]
        signal_length = ini_df[col_signal_length].values[i]
        
        #filter to skip null signal initialization
        if (signal_name=='-') or(signal_name and "null_" in str(signal_name).lower()):
            continue
        
        namespace = ini_df[col_namespace].values[i]

        pdu_name = str(ini_df[col_pdu].values[i]).strip()
        pdu_type = str(ini_df[col_pdu_type].values[i]).upper()
        payload_pdu_type = str(ini_df[col_payload_pdu_type].values[i]).upper()

        if (pdu_type in skip_pdu_type_list):  # skip all special pdu initialization
            continue


        network_name = ini_df[col_nwt_name].values[i]
        variant_value = ini_df[varinat_column_index].values[i]
        if variant_value == None:
            variant_value = 0

        # case new set of 25 signals
        if i % signal_batch == 0:
            # except for the first signal
            if i != 0:
                pass
                #ininode.append('  }')
            ini_count = i // signal_batch + 1
            #ininode.append('  if(init_count == {0})'.format(ini_count))
            #ininode.append('  {')
        if(("I-SIGNAL" in pdu_type) and (payload_pdu_type == 'N-PDU')):
            if re.match(Npdu_sig_patterns, signal_name, flags=re.I):
                signal_split = signal_name.split("_")
                loc_buffer_name = '_'.join(signal_split[:-1])
                if "OPAQUE" in signal_byte_order.upper():
                    ininode.append("  sysSetVariableData(sysvar::{0}::{1}::{2}::{3}::SignalValue, {3}_byte_Array, {4});".format(network_name, namespace, loc_buffer_name, signal_split[-1], signal_length // 8))
                    if signal_split[-1] not in opaque_byte_order_sig_dict.keys():
                        opaque_byte_order_sig_dict[signal_split[-1]] = str(signal_length // 8)
                else:
                    ininode.append("  @{0}::{1}::{2}.{3} = {4};".format(network_name, namespace, loc_buffer_name, signal_split[-1], variant_value))
            else:
                if "OPAQUE" in signal_byte_order.upper():
                    ininode.append('  sysSetVariableData(sysvar::{0}::{1}::{2}::SignalValue, {2}_byte_Array, {3});'.format(network_name, namespace, signal_name, signal_length // 8))
                    if signal_name not in opaque_byte_order_sig_dict.keys():
                        opaque_byte_order_sig_dict[signal_name] = str(signal_length // 8)
                else:
                    ininode.append('  @{0}::{1}::{2} = {3};'.format(network_name, namespace, signal_name, variant_value))

        else:
            if "OPAQUE" in signal_byte_order.upper():
                ininode.append('  sysSetVariableData(sysvar::{0}::{1}::{2}::SignalValue, {2}_byte_Array, {3});'.format(network_name, namespace, signal_name, signal_length // 8))
                if signal_name not in opaque_byte_order_sig_dict.keys():
                    opaque_byte_order_sig_dict[signal_name] = str(signal_length // 8)
            else:
                ininode.append('  ${0}::{1}::{2} = {3};'.format(network_name, namespace, signal_name, variant_value))
        # close function
    # print(f"generate_ininode::opaque_byte_order_sig_dict = {opaque_byte_order_sig_dict}")
    return ininode


def generate_ininode_classe(ini_df, variant):
    """
    here ini node logic is created for CLASSE

    Args:
      ini_df (dataframe): sysvar dataframe
      variant (str): vrainat name

    Returns:
        list: generated logic

    """
    global col_namespace, col_group, col_pdu, col_msg
    # get columns name & sender
    col_names = list(ini_df.iloc[0])
    try:
        col_name = col_names.index("Name")  # Signal
        col_group = col_names.index("group")  # group
        col_pdu = col_names.index("pdu")  # pdu
        col_msg = col_names.index("Message")  # Message
        col_init_value = col_names.index("Initial Value")
        col_sender = col_names.index("sender")
        col_datatype = col_names.index("texttable")
        col_texttable_values = col_names.index("texttable values")
        col_namespace = col_msg

        variant_column_index = col_names.index(variant)
    except Exception as e:
        logger.error(
            "{e} ::Column 'Name' or 'sender' or 'Message' is not found in Excel. Please check it in Autosar_Gen_Database.xlsx")
        raise e

    # diag filer
    #ini_df = filter_column_df(ini_df, col_namespace, col_name)

    num_of_nodes = len(node_list)
    # count the number of signals
    signal_count = [0] * num_of_nodes
    # flag count for the node number
    flag_count = 1
    # set of 25 signals batch
    signal_batch = 25
    datatype_list = ['floatarray', 'intarray', 'texttable_array']
    struct_datatype_list = ['float_struct', 'int_struct', 'texttable_struct', 'struct_struct']

    # Case create restbus ini
    ininode = []
    for j in range(num_of_nodes):
        # get node name and dataframe
        node = node_list[j]
        node_df_temp = ini_df[ini_df[col_init_value] != 'na']
        node_df = node_df_temp[node_df_temp[col_sender] == node]
        num_of_rows_df = node_df.shape[0]
        signal_count[j] = num_of_rows_df
        # node check: in case the node name is different from SysVarDatabase
        if node_df.empty:
            logger.warning(f"Node '{node}' is not found in Excel, please check it in Autosar_Gen_Database.xlsx.")
            # input("press any key to continue:")
            continue
        # print('we are on node: ' + str(node) + ', and we have the num of rows: ' + str(num_of_rows_df))
        # save in ininode
        ininode.append('void ' + 'init_{0}_{1}()'.format(node,variant))
        ininode.append('{')
        ininode.append("  int object_id;")
        ininode.append("  int array_len;")
        
        # variables for handling array initialization
        batch_count = 0
        is_previous_row_array = False

        # variables for struct variable initialization


        ini_count = 0
        for i in range(num_of_rows_df):
            # if (i*100//num_of_rows_df) % 10 == 0:       # restrain output
            #    print('\tprocess completed: ' + "%.1f" % (i*100/num_of_rows_df) + '%')
            signal_name = node_df[col_name].values[i]
            namespace = node_df[col_namespace].values[i]
            datatype = node_df[col_datatype].values[i]
            texttable_value = node_df[col_texttable_values].values[i]
            variable_name = signal_name
            varinat_value = node_df[variant_column_index].values[i]

            # case new set of 25 signals
            if batch_count % signal_batch == 0:
                # except for the first signal
                if i != 0:
                    ininode.append('  }')
                ini_count +=1 
                ininode.append('  if(init_count == {0})'.format(ini_count))
                ininode.append('  {')
            if datatype in struct_datatype_list:
                pattern = r"(\_t\::|\_st\::)"
                if texttable_value == 'Abstract' or datatype == 'struct_struct':
                    continue
                signal_name = re.sub(pattern, ".", signal_name)
                ininode.append(f'    @{namespace}::{signal_name} = {varinat_value};')
                batch_count += 1
                is_previous_row_array = False
            elif datatype in datatype_list:
                #array initialization should be done in seperate init count
                if not(is_previous_row_array):
                    if i!=0:
                        ininode.append('  }')
                    ini_count +=1 
                    ininode.append('  if(init_count == {0})'.format(ini_count))
                    ininode.append('  {')
                ininode.append("  object_id = 0;")
                # Extract array length
                pattern = r"\[(.*?)\]"
                matches = re.findall(pattern, signal_name)
                for match in matches:
                    ininode.append(f"  array_len = {match};")

                ininode.append(f"  for (object_id; object_id < array_len; object_id++)")
                ininode.append("  {")
                signal = variable_name.replace("::", ".").split("[")[0].rsplit(".",1)[0].rsplit("_",1)[0]
                sig_name = variable_name.split("::")[1].split("[")[0]
                sig_to_generate = signal + "." + sig_name
                ininode.append(f'    @{namespace}::{sig_to_generate}[object_id] = {varinat_value};')
                ininode.append("  }")
                batch_count = 0
                is_previous_row_array = True
            elif datatype == 'struct':
                # TODO: Implement recursive initialization
                struct_member_names = node_df[node_df[col_name].str.contains(texttable_value.split('::')[-1])][col_name].tolist()
                struct_member_types = node_df[node_df[col_name].str.contains(texttable_value.split('::')[-1])][col_datatype].tolist()
                struct_member_typevalue = node_df[node_df[col_name].str.contains(texttable_value.split('::')[-1])][col_texttable_values].tolist()
                struct_member_variantvalue = node_df[node_df[col_name].str.contains(texttable_value.split('::')[-1])][col_init_value].tolist()
                struct_member_parent = [variable_name] * len(struct_member_names)
                struct_stack = list(map(list, zip(struct_member_names, struct_member_types, struct_member_typevalue, struct_member_variantvalue, struct_member_parent)))
                while struct_stack:
                    if struct_stack[0][1] == 'struct_struct':
                        struct_name = struct_stack[0][0].split('::')[-1]
                        struct_type = struct_stack[0][2].split('::')[-1]
                        struct_member_names = node_df[node_df[col_name].str.contains(struct_type)][col_name].tolist()
                        struct_member_types = node_df[node_df[col_name].str.contains(struct_type)][col_datatype].tolist()
                        struct_member_typevalue = node_df[node_df[col_name].str.contains(struct_type)][col_texttable_values].tolist()
                        struct_member_variantvalue = node_df[node_df[col_name].str.contains(struct_type)][col_init_value].tolist()
                        struct_member_parent = [struct_stack[0][4] + '.' + struct_name] * len(struct_member_names)
                        struct_stack = list(map(list, zip(struct_member_names, struct_member_types, struct_member_typevalue, struct_member_variantvalue, struct_member_parent))) + struct_stack[1:]
                    else:
                        ininode.append('  @{0}::{1} = {2};'.format(namespace, struct_stack[0][4] + '.' + struct_stack[0][0].split('::')[-1], struct_stack[0][3]))
                        struct_stack.pop(0)
                        batch_count += 1
                        if batch_count % signal_batch == 0:
                            ini_count +=1 
                            ininode.append('  }')
                            ininode.append('  if(init_count == {0})'.format(ini_count))
                            ininode.append('  {')        
                is_previous_row_array = False
            else:
                ininode.append('  @{0}::{1} = {2};'.format(namespace, variable_name, varinat_value))
                batch_count += 1
                is_previous_row_array = False
            # close function
        ininode.append('  if (prestart_mode==0) init_count = 0;')
        if node == "classe":
            flag_count += 1
        else:
            flag_count += 2
        ininode.append('  initFlag = {0};'.format(flag_count))
        ininode.append('  }')
        ininode.append('  init_count++;')
        ininode.append('}')

    logger.info("=============== Summary ===============")
    logger.info("The number of signals of each node:")
    logger.info("---------------")
    logger.info("{:<10} {:<5}".format("Node", "Count"))
    for node, count in zip(node_list, signal_count):
        logger.info("{:<10} {:<5}".format(node, str(count)))

    return ininode

def append_file(ininode, file_name):
    """
    adds the headers and writes content to file

    Args:
      ininode (list): file content to write
      file_name (str): file path

    """
    # Open and save ini node signal.can file
    line_separator = "/***** Automated code generation starts here *****/"
    # add comments here
    comments = r"""
// Autogenerated by -> Platform\Classe\Scripts\Rbs_Scripts\create_ininode.py 
includes
{
}

variables
{
}
/***** Automated code generation starts here *****/
"""
    # read signal.can file and save content
    #original_text = []
    # with open(file_name, 'r') as outfile:
    #     line = outfile.readline()
    #     while line:
    #         if line.strip('\n') == line_separator:
    #             original_text.append(line.strip('\n'))
    #             break
    #         original_text.append(line.strip('\n'))
    #         line = outfile.readline()
    #     outfile.close()
    # rewrite signal.can file with ini nodes
    copyrights = getCopyRights(file_name)
    with open(file_name, 'w') as outfile:
        #outfile.write("\n".join(item for item in original_text))
        #outfile.write('\n' + line_separator + '\n')
        outfile.write(copyrights + comments + '\n')
        outfile.write("\n".join(str(item) for item in ininode))
        outfile.close()
    logger.info(f"{file_name} updated successfully")


def create_file(ininode, file_name):
    """
    adds the headers and writes content to file

    Args:
      ininode (list): file content to write
      file_name (str): file path

    """
    # Open and save ini node init_rbs.cin file
    # add comments here
    comments = '''
includes
{
}
'''
    # rewrite signal.cin file with ini nodes
    script_path = "\\".join(os.path.dirname(os.path.abspath(__file__)).split("\\")[-4:])
    copyrights = getCopyRights(file_name)
    with open(file_name, 'w') as outfile:
        if copyrights:
            outfile.write(copyrights + "// Autogenerated by -> " + script_path + r"\create_ininode.py "+ '\n' + comments + '\n')
        else:
            outfile.write("/*@!Encoding:1252*/" '\n' + "// Autogenerated by -> " + script_path + r"\create_ininode.py " + '\n' + comments + '\n')
        outfile.write("\n".join(str(item) for item in ininode))
    logger.info(file_name + " updated successfully")


def generate_variables(ini_node, flag_opaque):
    variables = []
    variables.append('variables \n{')
    if opaque_byte_order_sig_dict != {} and flag_opaque:
        variables.append('  int i;')
        for key, val in opaque_byte_order_sig_dict.items():
            variables.append('  byte ' + key + '_byte_Array[' + val + '];')
    variables.append('} \n')
    ini_node = variables + ini_node
    return ini_node


def generate_Opaque_array_initialize(ini_node):
    byte_array_initialize = []
    byte_array_initialize.append('')
    byte_array_initialize.append('void Initialize_opaque_byte_order_signal_array()')
    byte_array_initialize.append('{')
    if opaque_byte_order_sig_dict != {}:
        for key, val in opaque_byte_order_sig_dict.items():
            byte_array_initialize.append('  for(i = 0; i < ' + val + '; i++)')
            byte_array_initialize.append('  {')
            byte_array_initialize.append('    ' + key + '_byte_Array[i] = 0;')
            byte_array_initialize.append('  }')
    byte_array_initialize.append('}')
    ini_node = ini_node + byte_array_initialize
    return ini_node


def update_initialization_file(file_path, variant_columns, hit_ctrl_mapping_dict, can_variants_present, eth_variants_present, multicanoe_category="",
                               multicanoe_node_list=[]):
    """
    upadtes initialization.can file

    Args:
      file_path (str): file pathof initialization.can
      variant_columns (list):list of variant column names
      hit_ctrl_mapping_dict (dict): hil_ctrl namespace to ecu node mapping dict
      multicanoe_category:  (Default value = "")
      multicanoe_node_list:  (Default value = [])

    """
    try:
        #check if multicanoe is configured
        if multicanoe_node_list:
            rbs_init_node_list.clear()
            rbs_init_node_list.extend(multicanoe_node_list)
    
        logger.info(f"Start updating {file_path} ...")
        #--------------init_sequence_without_timers function ----------------------------
        default_init_count = '125'

        init_for_loop_templete = "for (init_count=0;init_count<<init-count>;) init_<node>_a_variant_veh_0();"

        init_sequence_without_timers = []
        temp = """/**
  * @brief Initializes the signals using the signal.ini. Order: 1. FDX, 2. Classe, 3. RBS.
  * In those functions the signals are split into groups of 25. So max number of signals is 25*125=3125. This is for CANoe so it does not get a big load peak. (RW)
  */
void init_sequence_without_timers()
{"""
        init_sequence_without_timers.append(temp)
        init_sequence_without_timers.append('\twrite("Start of Prestart...");')
        if ("master" in multicanoe_category.lower()):
            copyManualChangesFromInitToInitMaster(source_path, destination_path)
        if ("slave" not in multicanoe_category.lower()):
            # hardcoded fdx init
            fdx_init = """  /**
    * Initialize FDX
    */
  //for (init_count=0;init_count<125;) init_fdx(); 
  //write("Initializing FDX done.");"""
            init_sequence_without_timers.append(fdx_init)

            init_sequence_without_timers.append("")
            init_sequence_without_timers.append("\t@hil_ctrl::hil_mode = 1;")
            init_sequence_without_timers.append("")

            #Classe
            temp = """  /**
    * Initialize Classe
    */
  //125 groups - should be enough for most cases :)
  //only autogenerated classe variables initialized(considering variant columns list index 0 as default"""
            init_sequence_without_timers.append(temp)
            init_sequence_without_timers.append(
                f"\tfor (init_count=0;init_count<{default_init_count};) init_customer_{variant_columns[0]}();")
            init_sequence_without_timers.append(
                f"\tfor (init_count=0;init_count<{default_init_count};) init_classe_init_value();")
            init_sequence_without_timers.append('\twrite("Initializing Classe done.");\n')

        #Rbs
        temp = """  /**
    * Initialize RBS
    */
  //Initializing to default as a_variant_veh_0
  //125 groups - should be enough for most cases :)"""
        init_sequence_without_timers.append(temp)
        for node in rbs_init_node_list:
            try:
                init_count = str(dict_sigGroupCount_nodeName_mapping[node])
            except:
                init_count = default_init_count
            init_sequence_without_timers.append(
                f"\tfor (init_count=0;init_count < {init_count};) init_{node}_{variant_columns[0]}();")

        init_sequence_without_timers.append('\twrite("Initializing RBS done.");\n')

        #end
        temp = """  /**
    * End of preStart initialization.
    */
  initcounter = 0;
  initFlag = 0;
  init_count = 1;
  hand_over_to_user();
  write("End of Prestart.");
}\n"""
        
        init_sequence_without_timers.append(temp)
        #---------------------------------------------------------------------------

        #--------------classe_init_with_timers function ----------------------------

        classe_init_with_timers = []
        if("slave" not in multicanoe_category.lower()):

            temp = """/**
  * @brief Initialization of Classe using timers. This can be called on the run.
  */
void classe_init_with_timers()
{ """
            classe_init_with_timers.append(temp)
            temp = """  if (initFlag==0)
  {
    write("Start initializing Classe.");   
    initFlag ++;
  }"""
            classe_init_with_timers.append(temp)
            classe_node_count = len(classe_init_node_list)
            
            for i, node in enumerate(classe_init_node_list):
                i=i+1
                if i==1:
                    temp = "\tif (initFlag=={0}) {{".format(str(i))
                else:
                    temp = "\t}}else if (initFlag=={0}) {{".format(str(i))
                classe_init_with_timers.append(temp)
                if node == "classe":
                    classe_init_with_timers.append(f"\t\t\t\tinit_classe_init_value();")
                else:
                    classe_init_with_timers.append("\t\tswitch(@hil_ctrl::vehicle)\n\t\t{")
                    for variant in variant_columns:
                        classe_init_with_timers.append(f"\t\t\tcase {variant}:")
                        classe_init_with_timers.append(f"\t\t\t\tinit_{node}_{variant}();\n\t\t\t\tbreak;")
                    classe_init_with_timers.append("\t\t}")
            classe_init_with_timers.append("\t}")

            # end part
            temp = f"""  if (initFlag=={classe_node_count + 1})
  {{
    initcounter = 0;
    initFlag = 0;
    init_count = 1;	
    cancelTimer(Classe_Update_Timer);
    setTimerCyclic(Bus_Update_Timer,Update_Timer_Cycle);
    write("Initializing Classe done."); 
  }}"""
            
            classe_init_with_timers.append(temp)
            classe_init_with_timers.append("}\n")
            #---------------------------------------------------------------------------


        #--------------rbs_init_with_timers function -------------------------------
        rbs_init_with_timers = []
        temp = """/**
  * @brief Initialization of RBS using timers. This can be called on the run.
  */
void rbs_init_with_timers()
{"""
        rbs_init_with_timers.append(temp)
        temp = """  if(initFlag==0)
  {
    initFlag=1;
    write("Start of initializing RBS variables...");
  }"""
        rbs_init_with_timers.append(temp)

        rbs_node_count = len(rbs_init_node_list)

        rbs_node_dict_for_with_timers = {}
        for nd in rbs_init_node_list:
            key = nd.replace('_ETH', '') if nd.endswith('_ETH') else nd
            rbs_node_dict_for_with_timers.setdefault(key, []).append(nd)
                    
        for i, node in enumerate(rbs_node_dict_for_with_timers.keys()):
            i=i+1

            hil_ctrl_name = ""
            if node in hit_ctrl_mapping_dict["dbc"].keys():
                hil_ctrl_name = hit_ctrl_mapping_dict["dbc"][node]
            elif node in hit_ctrl_mapping_dict["arxml"].keys():
                hil_ctrl_name = hit_ctrl_mapping_dict["arxml"][node]
            else:
                raise Exception(
                    f"Node {node} ETH not found in mapping sheets, error while trying to get hil_ctrl mapping for {node} for rbs_init_with_timers function")

            if i == 1:
                temp = "\tif ((@hil_ctrl::{0} == 1) && (initFlag == {1})) {{".format(hil_ctrl_name, str(i))
            else:
                temp = "\t}}else if ((@hil_ctrl::{0} == 1) && (initFlag == {1})) {{".format(hil_ctrl_name, str(i))
            rbs_init_with_timers.append(temp)
            rbs_init_with_timers.append("\t\tswitch(@hil_ctrl::vehicle)\n\t\t{")
            for variant in variant_columns:
                rbs_init_with_timers.append(f"\t\t\tcase {variant}:")
                for nd in rbs_node_dict_for_with_timers[node]:
                    rbs_init_with_timers.append(f"\t\t\t\tinit_{nd}_{variant}();")
                rbs_init_with_timers.append(f"\t\t\t\tbreak;")
            rbs_init_with_timers.append("\t\t}")
        
        #skip flag part
        temp = [f"initFlag=={str(i)}" for i in range(1,rbs_node_count+1)]
        temp = " || ".join(temp)
        rbs_init_with_timers.append("\t}}else if({0}) {{".format(temp))
        temp="""    // In the case that one ECU has not to be initialized anymore increase the flag to skip that one
    write("%d skipped",initFlag);
    initFlag++;
    init_count = 1;
  }"""
        rbs_init_with_timers.append(temp)

        # end part
        temp = f"""  if(initFlag=={rbs_node_count + 1}) {{
    // End of the RBS initialization is reached
    initcounter = 0;
    initFlag = 0;
    init_count = 1;	
    hand_over_to_user();
    write("End of initializing RBS variables.");
  }}"""
        
        rbs_init_with_timers.append(temp)
        rbs_init_with_timers.append("}\n")

        #-----------------------------------------------------------------------------------
        #start updating initialization.can file
        autogen_line_separator = r"*** Automated code generation starts here ***"

        init_sequence_without_timers = "\n".join(init_sequence_without_timers)
        classe_init_with_timers = "\n".join(classe_init_with_timers)
        rbs_init_with_timers = "\n".join(rbs_init_with_timers)


        manual_includes_parsed = False
        includes_start_index = None

        file_data = []
        with open(file_path, 'r') as file:
            for index, line in enumerate(file.readlines()):
                if autogen_line_separator in line.strip('\n'):
                    file_data.append(line.strip('\n'))
                    file_data.append("\n")
                    break
                elif ("#include" in line) and (manual_includes_parsed):  # Remove old includes
                    continue
                else:
                    if ('automated_includes_starts' in line.lower()):
                        includes_start_index = index + 1  # Index where new includes shall be added
                        manual_includes_parsed = True #added all manual includes
                    file_data.append(line.strip('\n'))
            file.close()

        # Generate includes
        rbs_cin = []
        customer_cin = []
        for variant in variant_columns:
            rbs_multicanoe_category = variant + '_' + multicanoe_category if multicanoe_category else variant
            if can_variants_present:
                rbs_cin.append(f'\t#include "init_rbs_can_{rbs_multicanoe_category}.cin"\n')
            if eth_variants_present:
                rbs_cin.append(f'\t#include "init_rbs_eth_{rbs_multicanoe_category}.cin"\n')
            customer_cin.append(
                f'\t#include "..\\..\\..\\..\\CustomerPrj\\Classe\\Nodes\\Initialization\\customer_init_{variant}.cin"\n')

        includes_data = ['\t#include "..\\..\\..\\..\\Platform\\Classe\\Nodes\\Initialization\\classe_init_value.cin"\n',
                         "".join(rbs_cin), "".join(customer_cin), '}']

        if (includes_start_index is not None):
            file_data[includes_start_index] = "".join(includes_data)
        file_data.append(init_sequence_without_timers)
        file_data.append(classe_init_with_timers)
        file_data.append(rbs_init_with_timers)
        #copyrights = getCopyRights(file_path)
        with open(file_path, 'w') as file:
            #file.write(copyrights)
            file.write("\n".join(file_data))
            file.close()
        
        logger.info("Successfully updated Initialization.can")

    except Exception as e:
        logger.error(f"Error while updating initialization.can file -> {e}")
        traceback.print_exc()
        raise Exception(f"Error while updating initialization.can file -> {e}")


def extractMultiCanoeMapping(wb, skip_empty=False):
    """
    gets multicanoe info from mapping sheet

    Args:
      wb (workbook): excel worbook
      skip_empty (bool):  (Default value = False)

    Returns:
        dict_multicanoe (dict): node and category mapping

    """
    dict_multicanoe = {}
    try:
        for mapping_sheet in mapping_sheets:
            df = pd.DataFrame(wb[mapping_sheet].values)
            df.columns = df.iloc[0]
            df = df.drop(index=0)
            df = df.replace(np.nan, '', regex=True)
            for i, row in df.iterrows():
                nwt_name = row["network_name"].strip()
                canoe_ecu_name = row["canoe_ecu_node"].strip()
                multicanoe_category = row["multicanoe"].strip()
                if skip_empty:
                    #skip if data is empty
                    try:
                        #sheet_name = nwt_name + "_" + canoe_ecu_name + "_" + mapping_sheet.lower().replace("mapping", "")
                        temp_df = pd.DataFrame(wb["SysVarDatabase"].values)
                        temp_df.columns = temp_df.iloc[0]
                        temp_df = temp_df.drop(index=0)
                        temp_df = temp_df.replace(np.nan, '', regex=True)
                        if (temp_df[(temp_df["multicanoe"] == multicanoe_category) & (
                                temp_df["sender"] == canoe_ecu_name)].empty):
                            continue
                    except Exception as e:
                        logger.warning(
                            f"Error while checking if sheet is empty for -> {nwt_name}_{canoe_ecu_name}  --> {e}")

                if multicanoe_category != "":
                    if multicanoe_category not in dict_multicanoe.keys():
                        dict_multicanoe[multicanoe_category] = [canoe_ecu_name]
                    else:
                        if canoe_ecu_name not in dict_multicanoe[multicanoe_category]:
                            dict_multicanoe[multicanoe_category].append(canoe_ecu_name)
        return dict_multicanoe

    except Exception as e:
        logger.error(f"Error while extracting multicanoe mapping names from dbcmapping or arxml mapping -> {e}")


def copyManualChangesFromInitToInitMaster(source_path, destination_path):
    try:

        with open(source_path, 'r') as source_file:

            content = source_file.readlines()

        #
        modified_content = []

        # Iterate through each line in the content
        for i, line in enumerate(content):
            if line.strip().startswith('#include "init_rbs_'):
                # Replace .cin with _master.cin
                line = line.replace('.cin', '_master.cin')
            modified_content.append(line)

            if line.startswith("/************ Platform functions ************/"):
                startindex = i + 1  # Insert after this line

                # print(startindex)
        if startindex is not None:
            modified_content = modified_content[startindex:]
            # print(modified_content)

        # Open the destination file for reading and appending
        with open(destination_path, 'r+') as destination_file:

            existing_content = destination_file.readlines()

            # Find the index to insert after
            insert_index = None
            for i, line in enumerate(existing_content):
                # print(line)
                if line.startswith("/************ Platform functions ************/"):
                    insert_index = i + 1  # Insert after this line +1
                    # print(insert_index)

                    break

            # Insert the modified content into the existing content
            if insert_index is not None:
                existing_content[insert_index:] = modified_content
            else:
                # If the insert line is not found, append the modified content at the end
                existing_content.extend(modified_content)

            # Set the file pointer to the beginning to overwrite the existing content
            destination_file.seek(0)

            destination_file.writelines(existing_content)
            logger.info("Successfully copied manual changes from Initialization.can to Initialization_master.can")

        # print(f"Content modified in '{source_path}' appended to '{destination_path}' successfully.")

    except FileNotFoundError:
        print(f"One of the files does not exist.")
    except Exception as e:
        print(f"An error occurred: {e}")


script_path = os.path.dirname(os.path.abspath(__file__))
source_path = os.path.join(script_path, r'../../../../CustomerPrj/Classe/Nodes/Initialization/Initialization.can')
destination_path = os.path.join(script_path,
                                r'../../../../CustomerPrj/Classe/Nodes/Initialization/Initialization_master.can')


def external_call():
    """ """
    try:
        script_path = os.path.dirname(os.path.abspath(__file__))
        logger.info("###### START 'Create IniNodes(.can)' DEBUG INFORMATION ######")
        autosar_path = script_path + r'/../../../../CustomerPrj/Restbus/Autosar_Gen_Database.xlsx'
        classe_database = script_path + r'\..\..\..\..\Platform\Classe\Classe_Database.xlsx'
        customer_database = script_path + r'\..\..\..\..\CustomerPrj\Classe\Customer_Database.xlsx'
        wb = load_workbook(autosar_path, data_only=True)
        wb_classe = load_workbook(classe_database, data_only=True)
        wb_classe_customer = load_workbook(customer_database, data_only=True)
        variant_columns = extract_vehicle_variants(autosar_path)

        hit_ctrl_mapping_dict = extractHilCtrl(wb)
        dict_multicanoe = extractMultiCanoeMapping(wb,skip_empty=True)

        can_variants_present = False
        eth_variants_present = False

        can_sysvarSheet = "SysVarDatabase"
        # variant_columns = ["a_variant_veh_0", "a_variant_veh_1", "b_variant_veh_0", "b_variant_veh_1"]

        ini_df_main = generate_dataframe_rbs(wb, can_sysvarSheet, variant_columns)
        flag_opaque = True
        for variant in variant_columns:
            ini_node = generate_ininode_senderwise(ini_df_main, variant)
            if flag_opaque:
                ini_node = generate_variables(ini_node, flag_opaque)
                ini_node = generate_Opaque_array_initialize(ini_node)
                flag_opaque = False
            else:
                ini_node = generate_variables(ini_node, flag_opaque)
            create_file(ini_node,
                        script_path + r'/../../../../CustomerPrj/Classe/Nodes/Initialization/init_rbs_can_' + variant + '.cin')
            can_variants_present = True

        eth_sysvarSheet = "SysVarDatabaseETH"
        ini_df_main_eth = generate_dataframe_rbs_ethernet(wb, eth_sysvarSheet, variant_columns)
        if ini_df_main_eth is not None:
            flag_opaque = True
            for variant in variant_columns:
                ini_node = generate_ininode_senderwise_ethernet(ini_df_main_eth, variant)
                if flag_opaque:
                    #ini_node = generate_variables_ethernet(ini_node, flag_opaque)
                    #ini_node = generate_Opaque_array_initialize_ethernet(ini_node)
                    flag_opaque = False
                ini_node = generate_variables_ethernet(ini_node, flag_opaque)
                create_file(ini_node,
                            script_path + r'/../../../../CustomerPrj/Classe/Nodes/Initialization/init_rbs_eth_' + variant + '.cin')
                eth_variants_present = True


        # update initialization.can file
        update_initialization_file(
            script_path + r'/../../../../CustomerPrj/Classe/Nodes/Initialization/Initialization.can',
            variant_columns, hit_ctrl_mapping_dict, can_variants_present, eth_variants_present)

        # if multi canoe is configured
        if dict_multicanoe:
            logger.info("MultiCanoe configured in mapping sheets, Starting .cin generation ...")
            flag_opaque = True
            for variant in variant_columns:
                for multicanoe_category, node_list in dict_multicanoe.items():
                    ini_node = generate_ininode_multicanoe_senderwise(ini_df_main, variant, multicanoe_category,
                                                                      node_list)
                    if flag_opaque:
                        ini_node = generate_variables(ini_node, flag_opaque)
                        ini_node = generate_Opaque_array_initialize(ini_node)
                        flag_opaque = False
                    else:
                        ini_node = generate_variables(ini_node, flag_opaque)
                    create_file(ini_node,
                                script_path + r'/../../../../CustomerPrj/Classe/Nodes/Initialization/init_rbs_can_' + variant + '_' + multicanoe_category + '.cin')
                    logger.info(f"Generated .cin for {variant}  ->  {multicanoe_category}")

            for multicanoe_category, node_list in dict_multicanoe.items():
                update_initialization_file(
                    script_path + r'/../../../../CustomerPrj/Classe/Nodes/Initialization/Initialization_' + multicanoe_category + '.can',
                    variant_columns, hit_ctrl_mapping_dict,can_variants_present, eth_variants_present, multicanoe_category=multicanoe_category,
                    multicanoe_node_list=node_list)
                logger.info(f"Updated Initialization_{multicanoe_category}.can for multicanoe")


        # for classe
        ini_df = generate_dataframe_classe(wb_classe, "classe")
        ini_node = generate_ininode_classe(ini_df, "init_value")
        append_file(ini_node, script_path + r'/../../../../Platform/Classe/Nodes/Initialization/classe_init_value.cin')
        logger.info("###### END 'Create IniNodes(.can)' DEBUG INFORMATION ######")

        ini_df = generate_dataframe_classe(wb_classe_customer, "customer")
        for variant in variant_columns:
            ini_node = generate_ininode_classe(ini_df, variant)
            append_file(ini_node,
                        script_path + r'/../../../../CustomerPrj/Classe/Nodes/Initialization/customer_init_' + variant + '.cin')

        logger.info('-' * 80)
    except Exception as e:
        logger.error(f"'Create IniNodes(.can) failed. Exception --> {e}")
        raise Exception(e)
    logger.info("###### END 'Create IniNodes(.can)' DEBUG INFORMATION ######")

    logger.info('-' * 80)


if __name__ == "__main__":
    external_call()

