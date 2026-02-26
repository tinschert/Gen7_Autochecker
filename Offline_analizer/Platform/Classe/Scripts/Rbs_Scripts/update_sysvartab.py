# -*- coding: utf-8 -*-
# @file update_sysvartab.py
# @author ADAS_HIL_TEAM
# @date 08-21-2023

##################################################################
# C O P Y R I G H T S
# ----------------------------------------------------------------
# Copyright (c) 2023 by Robert Bosch GmbH. All rights reserved.

# The reproduction, distribution and utilization of this file as
# well as the communication of its contents to others without express
# authorization is prohibited. Offenders will be held liable for the
# payment of damages. All rights reserved in the event of the grant
# of a patent, utility model or design.

##################################################################

"""
merges all node sheets into on sysvar database sheet
"""

import os, sys
import numpy as np
import pandas as pd
try:
    pd.set_option('future.no_silent_downcasting', True)
except:
    pass
import pythoncom

try:
    from Control.logging_config import logger
except ImportError:
    sys.path.append(os.path.dirname(os.path.abspath(__file__)) + r"\..\Control")
    from logging_config import logger
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from win32com.client import Dispatch

try:
    from create_autosar import extract_vehicle_variants
except:
    from Rbs_Scripts.create_autosar import extract_vehicle_variants

try:
    sys.path.append(os.getcwd() + r"\..\..\CustomerPrj\Restbus\Scripts")
    from pattern_matching import *
    from update_gw_ruleset import update_gw
except ImportError:
    sys.path.append(os.path.dirname(os.path.abspath(__file__)) + r"\..\..\..\..\CustomerPrj\Restbus\Scripts")
    from update_gw_ruleset import update_gw
    from pattern_matching import *

column_names = ["Signal",
                    "Signal_Group",
                    "PDU",
                    "PDU_Type",
                    "Payload_PDU_Type",
                    "Payload_PDU",
                    "Selector_Field_Signal",
                    "Message",
                    "Message ID",
                    "Header ID",
                    "Startbit",
                    "PDU_Length [Byte]",
                    "Signal_Length [Bit]",
                    "Payload_PDU_Length [Byte]",
                    "Signal Base Type",
                    "Initial Value",
                    "max_value",
                    "Transmission Mode",
                    "Cycle Time [ms]",
                    "texttable",
                    "texttable values",
                    "Value Type",
                    "Comment",
                    "dlc",
                    "Payload PDU DLC",
                    "variant",
                    "Value Table",
                    "EndToEndProtection",
                    "Byte Order",
                    "Factor",
                    "Offset",
                    "Minimum",
                    "Maximum",
                    "Unit",
                    'Cantp_Pdu_Type',
                    'Block_size', 
                    'Address_formate', 
                    'Padding_active', 
                    'STMin',
                    'MAXFC_wait',
                    "sender", "gw", "multicanoe", 'node_entity', 'network_type','input_file', "dbc_file_index","network_name"]

eth_column_names = ['Service',
                    'Service Instance',
                    'Service ID',
                    'Major version',
                    'Minor version',
                    'Instance ID',
                    'Member Type',
                    'Event Group',
                    'Event_GroupId',
                    'Field Name',
                    'Member',
                    'Member ID',
                    'Field Type',
                    'Parameter',
                    'Parameter Type',
                    'No of Parameters',
                    'parameter_dtype_l1',
                    'parameter_dtype_l2',
                    'parameter_dtype_l3',
                    'parameter_dtype_l4',
                    'parameter_dtype_l5',
                    'Signal',
                    'No of Elements',
                    'Fire_and_forget',
                    "Signal_Group",
                    "PDU",
                    "PDU_Type",
                    "Payload_PDU_Type",
                    "Payload_PDU",
                    "Selector_Field_Signal",
                    "Startbit",
                    "PDU_Length [Byte]",
                    "Signal_Length [Bit]",
                    "Initial Value",
                    "max_value",
                    "Cycle Time [ms]",
                    "texttable",
                    "texttable values",
                    "Value Type",
                    "Comment",
                    "dlc",
                    "variant",
                    "Value Table",
                    "EndToEndProtection",
                    'Signal Data Type',
                    'Byte Order',
                    'Factor',
                    'Offset',
                    'Minimum',
                    'Maximum',
                    'Unit',
                    'IP_Address',
                    'UDP_Port',
                    'TCP_Port',
                    'TP_Protocol',
                    'VLAN_ID',
                    'VLAN_Name',
                    'SD Type',
                    "Network Protocol",
                    'Autosar Type',
                    'Network Name',
                    'Node Name']

col_name = column_names.index("Signal")
col_message = column_names.index("Message")
col_message_id = column_names.index("Message ID")
col_sender = column_names.index("sender")
col_gw = column_names.index("gw")
col_variant = column_names.index("variant")
col_pdu = column_names.index("PDU")
col_pdu_type = column_names.index("PDU_Type")
col_payload_pdu = column_names.index("Payload_PDU")
col_network_name = column_names.index("network_name")

variant_columns_added = False

# allow longer names up to 100 characters
pd.options.display.max_colwidth = 100
pd.options.display.colheader_justify = 'left'

dict_mappingsheet_data = {}

dict_Column_Datatype = [{'Startbit': int}, {'Signal_Length [Bit]': int}, {'Initial Value': float}, {'Factor': float},
                        {'Offset': float},
                        {'Minimum': float}, {'Maximum': float}, {'Cycle Time [ms]': int}, {'max_value': float},
                        {'dlc': int}, {'dbc_file_index':int}]


def filterSomeipData(df):
    """
    Filters out services which are provided and consumed by same node
    
    Args:
        df (DataFrame): Input dataframe
    
    Returns:
        DataFrame: Filtered df
    """
    # Filter rows based on 'SD Type'
    pdu_data = df[df["Network Protocol"] == 'ETH_PDU']
    someip_data = df[df["Network Protocol"] == 'ETH_SOMEIP']
    df_provided = someip_data[someip_data['SD Type'] == 'provide']
    df_consumed = someip_data[someip_data['SD Type'] == 'consume']


    # Create a unique identifier for services
    df_provided['temp'] = 'sif_' + df_provided['Service ID'].astype(str) + '_' + \
                              df_provided['Major version'].astype(str) + '_' + \
                              df_provided['Minor version'].astype(str) + '_' + \
                              df_provided['Instance ID'].astype(str)

    df_consumed['temp'] = 'sif_' + df_consumed['Service ID'].astype(str) + '_' + \
                              df_consumed['Major version'].astype(str) + '_' + \
                              df_consumed['Minor version'].astype(str) + '_' + \
                              df_consumed['Instance ID'].astype(str)

    # Filter consumed services not in provided services
    provided_rbs_service_list = df_provided['temp'].unique()
    rbs_df_consumed_filtered = df_consumed[~df_consumed['temp'].isin(provided_rbs_service_list)]

    # Concatenate provided and filtered consumed dataframes, drop the 'temp' column

    concat_list = []
    if not (df_provided.empty):
        concat_list.append(df_provided)
    if not (rbs_df_consumed_filtered.empty):
        concat_list.append(rbs_df_consumed_filtered)
    if not (pdu_data.empty):
        concat_list.append(pdu_data)

    # Concatenate provided and filtered consumed dataframes, drop the 'temp' column
    if len(concat_list)>1:
        result_df = pd.concat(concat_list)
        if 'temp' in result_df.columns:
            result_df = result_df.drop(columns=['temp'])
    elif len(concat_list) == 1:
        result_df = concat_list[0]
    else:
        result_df = df
    
    logger.info("Filtered out services which are provided and consumed by same node")
    
    return result_df


def datatype_formatter(np_array):  # to make sure all the data types of columns are as per 'dict_Column_Datatype' to avoid creating duplicates in sysvar sheet
    """
    converts datatype of the pandas dataframe and returns as numpy array
    Args:
      np_array (array): numpy array to process
    Returns:
        formatted np array

    """
    df = pd.DataFrame(np_array, columns=column_names)
    for pair in dict_Column_Datatype:
        try:
            df = df.astype(pair)
        except ValueError as vr:
            logger.warning(f"Empty cell found in the column {pair}, hence couldn't convert the datatype for this column")
    #df = df.drop_duplicates(subset=duplicate_columnlist_sysvar, keep="first")
    return df.to_numpy()

def getMappingData(wb, mapping_column_name):
    """
    returns a dict with sheetname as key and required column as value

    Args:
      wb (workbook): excel workbook
      mapping_column_name (str): required column

    Returns:
        dict: maping data

    """

    mapping_sheets = {"DBCmapping": "dbc", "ArxmlMapping": "arxml", "ETHArxmlMapping": "arxml"}
    dict_mapping = {}
    try:
        if dict_mappingsheet_data == {}:
            for mapping_sheet in mapping_sheets.keys():
                if mapping_sheet not in wb.sheetnames:
                    continue
                df = pd.DataFrame(wb[mapping_sheet].values)
                df.columns = df.iloc[0]
                df = df.drop(index=0)
                df = df.replace(np.nan, '', regex=True)
                dict_mappingsheet_data[mapping_sheet] = df

        for mapping_sheet, df in dict_mappingsheet_data.items():
            file_type = mapping_sheets.get(mapping_sheet, "")
            for i, row in df.iterrows():
                if mapping_column_name not in df.columns:
                    continue
                key = row["network_name"].strip() + "_" + row["canoe_ecu_node"].strip() + "_" + file_type
                dict_mapping[key] = str(row[mapping_column_name]).strip()
        return dict_mapping
    except Exception as e:
        logger.error(f"Error while extracting data from mapping sheets -> {e}")
        raise Exception(e)

def generate_database(wb, car_variants: list):
    """
    main function, takes all node sheets and merges into signle sysvar database sheet for both can and eth sheets
    at the end takes varaint values from previous excel and retains the column without any change
    so if there is no varaint values filled in avriant columna then it should be filled manually

    Args:
      wb (workbook): excel workbook

    Returns:
        dict: merged data for CAN and ETH

    """
    global variant_columns_added
    if not(variant_columns_added):
        column_names.extend(car_variants)
        eth_column_names.extend(car_variants)
        variant_columns_added=True

    column_size = len(column_names)
    dict_file_index = getMappingData(wb,"mapping_canoe_ecu_to_db")
    dict_entity = getMappingData(wb, "node_entity")
    dict_nwt_name = getMappingData(wb, "network_name")
    dict_nwt_type = getMappingData(wb, "network_type")
    dict_multicanoe = getMappingData(wb, "multicanoe")
    dict_blk_msg = getMappingData(wb, "block_messages")
    dict_block_messages = {}
    for st, value in dict_blk_msg.items():
        if (value not in ['na', '', None]):
            value = value.strip().split(";")
            value = [i.strip() for i in value]
            if st.startswith('CAN'):
                value = ['0X' + i.upper() if not (i.upper().startswith('0X')) else i.upper() for i in value]
            dict_block_messages[st] = value

    dict_pass_msg = getMappingData(wb, "pass_messages")
    dict_pass_messages = {}
    for st, value in dict_pass_msg.items():
        if (value not in ['na', '', None]):
            value = value.strip().split(";")
            value = [i.strip() for i in value]
            if st.startswith('CAN'):
                value = ['0X' + i.upper() if not (i.upper().startswith('0X')) else i.upper() for i in value]
            dict_pass_messages[st] = value

    sheet_list = wb.sheetnames
    copy_col_index = [col_message, col_message_id]
    database_list = []
    ethernet_database_list = []
    for sheet in sheet_list:
        if "CAN_" in sheet.upper():  # if it is an ECU sheet
            ws = wb[sheet]
            sheetname_split = sheet.split("_")
            sheet_data = ws.values
            cols = next(sheet_data)
            sheet_df = pd.DataFrame(sheet_data, columns=cols)
            sheet_df = sheet_df.replace(np.nan, '', regex=True)
            npdu_dict = map_npdu_to_ipdu_dict(sheet_df, col_message, col_pdu, col_pdu_type, col_payload_pdu, column_indexed=False)
            securedPdu_dict = map_securedpdu_to_ipdu_dict(sheet_df, col_message, col_pdu, col_pdu_type, col_payload_pdu, column_indexed=False)
            for pdu_type, pdu_type_df in sheet_df.groupby(col_pdu_type):
                if pdu_type == 'I-SIGNAL-I-PDU':
                    for i, row in pdu_type_df.iterrows():
                        if check_PDU_in_NPDU(row[col_pdu], npdu_dict):
                            for msg, values in npdu_dict.items():
                                if values[1] == row[col_pdu]:
                                    for index in copy_col_index:
                                        sheet_df.loc[i, index] = values[2][index]
                                        #sheet_df[index][i] = values[2][index]
                        elif check_PDU_in_SecuredPdu(row[col_pdu], securedPdu_dict):
                            for msg, values in securedPdu_dict.items():
                                if values[1] == row[col_pdu]:
                                    for index in copy_col_index:
                                        sheet_df.loc[i, index] = values[2][index]
                                        #sheet_df[index][i] = values[2][index]
            if sheet in dict_block_messages.keys():
                blk_list_ids = dict_block_messages[sheet]
                sheet_df = sheet_df[~sheet_df[col_message_id].isin(blk_list_ids)]
            elif sheet in dict_pass_messages.keys():
                pass_list_ids = dict_pass_messages[sheet]
                sheet_df = sheet_df[sheet_df[col_message_id].isin(pass_list_ids)]
            sheet_array = sheet_df.to_numpy()
            logger.info(f"{sheet} size: {sheet_array.shape}")
            if sheet_array.shape[0] != 0:
                # add columns: sender, gw, a_variant, b_variant
                add_array = np.array([[''] * (column_size - col_sender) for _ in range(sheet_array.shape[0])],
                                     dtype=sheet_array.dtype)
                add_array[:, 0] = sheet.split('_')[-2]  # sender
                add_array[:, 2] = dict_multicanoe[sheet]  # multicanoe master/slave
                add_array[:, 3] = dict_entity[sheet]  # entity message/pdu
                add_array[:, 4] = dict_nwt_type[sheet]  # bus_type  (can/ canfd etc)
                add_array[:, 5] = sheetname_split[-1]  # file type   (dbc/arxml)
                add_array[:, 6] = dict_file_index[sheet]  # file_index in mapping sheet
                add_array[:, 7] = dict_nwt_name[sheet]  # network_name
                database_list.append(np.concatenate([sheet_array[:, :col_sender], add_array], axis=1))
            else:
                logger.info(f"{sheet} is empty!")
        elif "ETH_" in sheet.upper():  # for ethernet sheets
            ws = wb[sheet]
            sheet_data = ws.values
            cols = next(sheet_data)
            sheet_df = pd.DataFrame(sheet_data, columns=cols)
            sheet_df = sheet_df.replace(np.nan, '', regex=True)
            sheet_df = filterSomeipData(sheet_df)
            if sheet in dict_block_messages.keys():
                blk_pdu_list = dict_block_messages[sheet]
                sheet_df = sheet_df[~sheet_df[col_pdu].isin(blk_pdu_list)]
            elif sheet in dict_pass_messages.keys():
                pass_pdu_list = dict_pass_messages[sheet]
                sheet_df = sheet_df[sheet_df[col_pdu].isin(pass_pdu_list)]
            sheet_array = sheet_df.to_numpy()
            logger.info(f"{sheet} size: {sheet_array.shape}")
            if sheet_array.shape[0] != 0:
                network_name_column = np.full((sheet_array.shape[0], 1), "_".join(sheet.split('_')[:2]))
                sheet_array = np.hstack((sheet_array, network_name_column))
                node_name_column = np.full((sheet_array.shape[0], 1), sheet.split('_')[-2])
                sheet_array = np.hstack((sheet_array, node_name_column))
                for var in car_variants:
                    var_col = np.full((sheet_array.shape[0], 1), '')
                    sheet_array = np.hstack((sheet_array, var_col))

                ethernet_database_list.append(sheet_array)
            else:
                logger.info(f"{sheet} is empty!")

    try:
        if database_list:
            database_array = np.concatenate(database_list, axis=0)
        else:
            database_array = []
        if ethernet_database_list:
            eth_database_array = np.concatenate(ethernet_database_list, axis=0)
        else:
            eth_database_array = []
    except ValueError as e:
        logger.error(f"Error while sheet concatenation -> {e}")
        raise Exception(f"Run create node sheets before running update sysvar -> {e}")
    database_array = datatype_formatter(database_array)
    # sort array by column Message & Name
    database_array = database_array[np.lexsort((database_array[:, col_name], database_array[:, col_message]))]
    database_array_filtered = []

    for element in database_array.tolist():
        database_array_filtered.append(element)

    database_array = np.array(database_array_filtered, dtype=object)
    if len(database_array)!=0:
        logger.info(f"SysVarDatabase size: {database_array.shape}")
    if len(eth_database_array)!=0:
        logger.info(f"SysVarDatabaseETH size: {eth_database_array.shape}")
    return {"CAN": database_array, "ETH": eth_database_array}


def update_variants(wb, old_sysvar, database_array, variant_list):
    """
    Updates the 'database_array' with values of variant columns (using cut_index) 
    from the 'old_sysvar' sheet in the workbook 'wb'.

    Args:
        wb (Workbook): The workbook containing the 'old_sysvar' sheet.
        old_sysvar (str): The name of the sheet in the workbook.
        database_array (ndarray): The array to be updated.

    Returns:
        ndarray: The updated 'database_array' with values from the 'old_sysvar' sheet.

    """
    try:
        if old_sysvar not in wb.sheetnames:
            logger.warning("While updating variant values, Old SysvarSheet is empty or not found")
            return database_array
        sysvar_ws = wb[old_sysvar]

        old_columns = [cell.value for cell in sysvar_ws[1]]

        cut_index_old = len(old_columns) - len(variant_list)
        cut_index_new = len(column_names) - len(variant_list)

        sysvar_array = pd.DataFrame(sysvar_ws.values).fillna('').iloc[1:].to_numpy()

        # Create a dictionary from sysvar_array for efficient lookups
        sysvar_dict = {(row[0], row[old_columns.index('Message')], row[old_columns.index('network_name')], row[old_columns.index('sender')]): row for row in sysvar_array}

        for i in range(database_array.shape[0]):
            name = database_array[i, col_name]
            message = database_array[i, col_message]
            nwt_name = database_array[i, col_network_name]
            sender_name = database_array[i, col_sender]

            # Look up the matching row in sysvar_dict
            matched_sysvar = sysvar_dict.get((name, message, nwt_name, sender_name), None)

            if matched_sysvar is not None:
                database_array[i, cut_index_new:] = matched_sysvar[cut_index_old:]
            else:
                database_array[i, cut_index_new:] = ''

        logger.info("Variant columns were updated for CAN according to old SySVarDatabase sheet.")
        return database_array
    except Exception as e:
        logger.error(f"An error occurred while updating variants: {str(e)}")
        raise Exception(f"An error occurred while updating variants: {str(e)}")



def update_variants_eth(wb, old_sysvar, database_array, variant_list):
    """
    Updates the variant columns in the provided database array based on the old SysVar sheet in the workbook for ethernet.
    Args:
        wb (openpyxl.Workbook): The workbook containing the old SysVar sheet.
        old_sysvar (str): The name of the old SysVar sheet.
        database_array (numpy.ndarray): The array representing the current database.
        variant_list (list): The list of variant names.
    Returns:
        numpy.ndarray: The updated database array with variant columns updated.
    Raises:
        Exception: If an error occurs during the update process.
    Notes:
        - The function checks if the old SysVar sheet exists in the workbook.
        - It creates a dictionary from the old SysVar sheet for efficient lookups.
        - It updates the variant columns in the database array based on the old SysVar sheet.
        - If a matching row is not found in the old SysVar sheet, the variant columns are set to empty.
    """

    try:
        if old_sysvar not in wb.sheetnames:
            logger.warning("While updating variant values, Old SysvarSheet is empty or not found")
            return database_array
        sysvar_ws = wb[old_sysvar]

        old_columns = [cell.value for cell in sysvar_ws[1]]
        cut_index_old = len(old_columns) - len(variant_list)
        cut_index_new = len(eth_column_names) - len(variant_list)

        sysvar_array = pd.DataFrame(sysvar_ws.values).fillna('').iloc[1:].to_numpy()
        if sysvar_array.size == 0:
            for variant in variant_list:
                database_array = np.column_stack((database_array, np.full((database_array.shape[0], 1), '', dtype=object)))
            logger.info("Variant columns were added to Ethernet database array as old SysVar sheet is empty or not found.")
            return database_array

        # Create a dictionary from sysvar_array for efficient lookups
        sysvar_dict = {(row[eth_column_names.index('Signal')], row[eth_column_names.index('PDU')], row[eth_column_names.index('Network Name')], row[eth_column_names.index('Node Name')]): row for row in sysvar_array}

        for i in range(database_array.shape[0]):
            name = database_array[i, eth_column_names.index('Signal')]
            pdu = database_array[i, eth_column_names.index('PDU')]
            nwt_name = database_array[i, eth_column_names.index('Network Name')]
            sender_name = database_array[i, eth_column_names.index('Node Name')]

            # Look up the matching row in sysvar_dict
            matched_sysvar = sysvar_dict.get((name, pdu, nwt_name, sender_name), None)

            if matched_sysvar is not None:
                database_array[i, cut_index_new:] = matched_sysvar[cut_index_old:]
            else:
                database_array[i, cut_index_new:] = ''

        logger.info("Variant columns were updated for Ethernet according to old SySVarDatabase sheet.")
        return database_array
    except Exception as e:
        logger.error(f"An error occurred while updating variants: {str(e)}")
        raise Exception(f"An error occurred while updating variants: {str(e)}")


def save_database(wb, dict_sheet_names, dict_arrays, file_name):
    """
    save the generated excel

    Args:
      wb (workbook): workbook of excel
      sysvar_sheet (str): sheet name
      database_array (array): generated array
      file_name (str): filenme

    """
    sheet_list = wb.sheetnames
    # not exist, create sysvar sheet; otherwise remove it and create a new one
    dict_column_names = {"CAN": column_names, "ETH": eth_column_names}
    for nwt, sysvar_sheet in dict_sheet_names.items():
        if isinstance(dict_arrays[nwt], np.ndarray) and dict_arrays[nwt].size == 0:
            continue
        elif isinstance(dict_arrays[nwt], list) and len(dict_arrays[nwt]) == 0:
            continue

        if sysvar_sheet not in sheet_list:
            sysvar_ws = wb.create_sheet(sysvar_sheet)
        else:
            sysvar_ws = wb[sysvar_sheet]
            wb.remove(sysvar_ws)
            sysvar_ws = wb.create_sheet(sysvar_sheet)
        # numpy->pandas->worksheet
        database_df = pd.DataFrame(dict_arrays[nwt], columns=dict_column_names[nwt])
        for row in dataframe_to_rows(database_df, index=False):
            sysvar_ws.append(row)
    wb.save(file_name)
    try:
        pythoncom.CoInitialize()
        xlApp = Dispatch("Excel.Application")
        xlApp.Visible = False
        xlBook = xlApp.Workbooks.Open(os.path.abspath(os.path.join(os.getcwd(), file_name)))
        xlBook.Save()
        xlBook.Close()
        xlApp.Quit()
        logger.info("Changes have been successfully saved into {}.".format(file_name))
    except Exception as exp:
        logger.error("Failed to save Excel:", exp)
        logger.error("Please open {} with Excel and save it manually.".format(file_name))


def check_sysvardatabase_integrity(database):
    """
    checks integrity after merging all node sheets into single, prints warning if any issues

    Args:
      database (array): merged database array

    """
    columns_to_read = ['Signal', 'PDU', 'Minimum', 'Maximum', 'Message ID', 'sender', 'network_name']
    df = pd.read_excel(database, sheet_name='SysVarDatabase', usecols=columns_to_read)#Column names : Name,pdu,Message,Multiplexing/Group,Minimum,Maximum,Message ID,sender,network_name
    # Check duplicated Signal/Message in SysVarDatabase sheet
    duplicated = df.duplicated().values.tolist()
    indices = [i for i, x in enumerate(duplicated) if x is True]
    for row in indices:
        logger.warning(f"Signal/Message on row {row + 1} duplicated in SysVarDatabase sheet")

    # Check if value in column Minimum is > then value in column Maximum
    comparison_column = np.where(df["Minimum"] >= df["Maximum"], True, False).tolist()
    indices = [i for i, x in enumerate(comparison_column) if x is True]
    for row in indices:
        logger.warning(f"Minimum >= Maximum on row {row + 2} in SysVarDatabase sheet")

    logger.info("###### END 'Update SysVarDatabase' DEBUG INFORMATION ######")
    logger.info('-' * 80)


def external_call():
    """ external call for jenkins"""
    try:
        logger.info(f"###### START 'Update SysVarDatabase' DEBUG INFORMATION ######")
        script_path = os.path.dirname(os.path.abspath(__file__))
        autosar_path = script_path + r'/../../../../CustomerPrj/Restbus/Autosar_Gen_Database.xlsx'
        variant_list = extract_vehicle_variants(autosar_path)
        wb_r = load_workbook(autosar_path, data_only=True)
        
        dict_sysvar_sheets = {"CAN": "SysVarDatabase", "ETH": "SysVarDatabaseETH"}
        dict_arrays = generate_database(wb_r, variant_list)

        #process CAN sheet by filling variants values etc
        database_array = dict_arrays["CAN"]
        database_array = update_gw(database_array)
        if variant_list:
            database_array = update_variants(wb_r, dict_sysvar_sheets["CAN"], database_array, variant_list)
        dict_arrays["CAN"] = database_array

        #process ETH sheet by filling variants
        eth_database_array = dict_arrays["ETH"]
        if variant_list:
            eth_database_array = update_variants_eth(wb_r, dict_sysvar_sheets["ETH"], eth_database_array, variant_list)
        dict_arrays["ETH"] = eth_database_array
        wb_r.close()
        
        # workbook for writing
        wb_w = load_workbook(autosar_path)
        save_database(wb_w, dict_sysvar_sheets, dict_arrays, autosar_path)
        #check_sysvardatabase_integrity(autosar_path)
    except Exception as e:
        logger.error(f"Update SysVarDatabase failed. Exception --> {e}")
        raise Exception(e)
    logger.info("###### END 'Update SysVarDatabase' DEBUG INFORMATION ######")


if __name__ == "__main__":
    external_call()

