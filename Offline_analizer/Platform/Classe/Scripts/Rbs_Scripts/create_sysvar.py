# -*- coding: utf-8 -*-
# @file create_sysvar.py
# @author ADAS_HIL_TEAM
# @date 08-21-2023

##################################################################
# C O P Y R I G H T S
# ----------------------------------------------------------------
# Copyright (c) 2023 by Robert Bosch GmbH. All rights reserved.

# The reproduction, distribution and utilization of this file as
# well as the communication of its contents to others without express
# authorization is prohibited. Offenders will be held liable for the
# payment of damages. All rights reserved in the event of the grant
# of a patent, utility model or design.

##################################################################


from openpyxl import load_workbook
import pandas as pd
try:
    pd.set_option('future.no_silent_downcasting', True)
except:
    pass
import numpy as np
import unidecode
import re
import os, sys
from tkinter import messagebox
import traceback
from collections import defaultdict
import ast

try:
    sys.path.append(os.path.dirname(os.path.abspath(__file__)) + r"\..\..\CustomerPrj\Restbus\Scripts")
    from pattern_matching import *
except ImportError:
    sys.path.append(os.path.dirname(os.path.abspath(__file__)) + r"\..\..\..\..\CustomerPrj\Restbus\Scripts")
    from pattern_matching import *

try:
    from Control.logging_config import logger
except ImportError:
    sys.path.append(os.path.dirname(os.path.abspath(__file__)) + r"\..\Control")
    from logging_config import logger

try:
    from create_autosar import extract_vehicle_variants
except:
    from Rbs_Scripts.create_autosar import extract_vehicle_variants

# allow longer names up to 100 characters
pd.options.display.max_colwidth = 100
pd.options.display.colheader_justify = 'left'

bus_xml = []

required_restbus_can_column_names = ("Signal",
                    "Signal_Group",
                    "PDU",
                    "PDU_Type",
                    "Payload_PDU_Type",
                    "Payload_PDU",
                    "Selector_Field_Signal",
                    "Message",
                    "Message ID",
                    "Startbit",
                    "PDU_Length [Byte]",
                    "Signal_Length [Bit]",
                    "Initial Value",
                    "max_value",
                    "Cycle Time [ms]",
                    "texttable",
                    "texttable values",
                    "Value Type",
                    "Comment",
                    "dlc",
                    "variant",
                    "Value Table",
                    "EndToEndProtection",
                    "Byte Order",
                    "Factor",
                    "Offset",
                    "Minimum",
                    "Maximum",
                    "Unit",
                    'Block_size', 
                    'Address_formate', 
                    'Padding_active', 
                    'STMin',
                    'MAXFC_wait',
                    "sender", "gw", "multicanoe", 'node_entity', 'network_type','input_file', "dbc_file_index","network_name")




skip_pdu_type_list = ['NM-PDU', 'MULTIPLEXED-I-PDU', 'GENERAL-PURPOSE-PDU', 'GENERAL-PURPOSE-I-PDU']

column_names_classe = ["Name", "group", "pdu", "Message", "Value Type", "Initial Value", "Unit", "texttable", "texttable values", "Offset", "Minimum" , "Factor", "Maximum", "max_value"]
column_names_TP=["Block_size", "Address_formate","Padding_active", "STMin","MAXFC_wait","Message ID"]

osektp_param_dict = {"Block_size": "sysBlockSize", "Address_formate": "Address_formate", "Padding_active": "Padding_active", 
                     "STMin": "sysSTmin", "MAXFC_wait": "MAXFC_wait", "Message ID":"sysTxIdentifier"}

number_pattern = re.compile(r'(0[xX][0-9a-fA-F]+)|(\d+)')


def getColNamespace(entity):
    """
    takes entity name as input and return column index

    Args:
      entity (str): entity name msg/pdu/grp

    Returns:
        int: index of column
    """
    if "message" in entity:
        return col_msg
    elif "pdu" in entity:
        return col_pdu
    elif "group" in entity:
        return col_group
    else:
        return col_msg


def create_bus_sysvar(wb_sheet_list):
    """
    create one sysvar for each network on/off

    Args:
      wb_sheet_list (list): sheet list

    Returns:
        list of xml content for on/off sysvars
    """
    # open working sheet
    bus_xml = []
    bus_list = []
    for sheet in wb_sheet_list:
        sheet_split = sheet.split('_')
        if len(sheet_split) > 2 and ("sysvardatabase" not in sheet.lower()):  # if it is an ECU sheet
            bus_list.append(sheet_split[-4] + '_' + sheet_split[-3])
    bus_list = list(dict.fromkeys(bus_list))
    bus_xml.append('  <namespace name="Cus_bus" comment="" interface="">')
    for bus in bus_list:
        bus_xml.append(
            '    <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="" name="bus_{0}_ON_OFF" comment="" bitcount="64" isSigned="false" encoding="65001" type="int" startValue="1" minValue="0" maxValue="1" />'.format(
                bus))
    bus_xml.append('  </namespace>')
    return bus_xml


def generate_sysvar_header():
    """generate sysvar header """
    xml = ['<?xml version="1.0" encoding="utf-8"?>']
    xml.append('<systemvariables version="4">')
    xml.append('  <namespace name="" comment="" interface="">')
    return xml


def generate_sysvar_footer(bus_xml=[]):
    """
    generate sysvar footer

    Args:
      bus_xml (list):  (Default value = []) bus sysvars to append at the end

    Returns:
        upadted list of xml content
    """
    xml = ['    </namespace>']
    xml.extend(bus_xml)
    xml.append('  </namespace>')
    xml.append('</systemvariables>')
    return xml


def generate_sysvar_footer_classe():
    """ generate sysvar footer classe"""
    xml = ['    </namespace>', '</systemvariables>']
    return xml


def define_classe_colIndex(classe_columns):
    """
    define classe colIndex makes global

    Args:
      classe_columns (list): column list
    """
    global col_name, col_group, col_pdu, col_msg, col_value_type, col_ini_value, col_unit, col_texttable, col_enum, col_offset, col_min, col_factor, col_max, col_max_value
    try:
        col_index = [classe_columns.index(col) for col in column_names_classe]
        [col_name, col_group, col_pdu, col_msg, col_value_type, col_ini_value, col_unit, col_texttable, col_enum,
         col_offset, col_min, col_factor, col_max, col_max_value] = col_index
    except Exception as e:
        logger.error(f"{e} - Column names in Excel are different from the Script:")
        logger.error(';'.join(column_names_classe))
        return False


def filter_dataframe(df, col_index, value):
    """
    filter_dataframe for given column and value

    Args:
      df (dataframe): dataframe to filter
      col_index (int): index of the column to filter
      value (str/number): value to filter

    Returns:
        filtered dataframe

    """
    cols = list(df.iloc[0])
    df = df[df[col_index] == value]
    # df = df[df[col_input_file].str.contains(input_file, regex=False) != False]
    df.loc[-1] = list(cols)  # adding a column row
    df.index = df.index + 1  # shifting index
    df = df.sort_index()
    return df


def generate_Sig_sysvar(sysvar_name):
    """
    generates xml variables for CRC handling for signal

    Args:
      sysvar_name (str): crc_variable name

    Returns:
        updated list of xml

    """
    sysvar_name_fault_inject = sysvar_name + "_FaultInject"
    sysvar_name_cnt_cycle = sysvar_name + "_FaultInject_CntCycle"
    
    temp_xml = ['        <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="" name="{0}" comment="" bitcount="32" isSigned="false" encoding="65001" type="int" startValue="0" minValue="0" minValuePhys="0" maxValue="3" maxValuePhys="3">'.format(sysvar_name_fault_inject),
           '			<valuetable name="{0}" definesMinMax="true">'.format(sysvar_name_fault_inject),
           '				<valuetableentry value="0" lowerBound="0" upperBound="0" description="Disable" displayString="Disable" />',
           '				<valuetableentry value="1" lowerBound="1" upperBound="1" description="Enabled:Set_to_zero" displayString="Enabled:Set_to_zero" />',
           '				<valuetableentry value="2" lowerBound="2" upperBound="2" description="Enabled:Freeze_previous_value" displayString="Enabled:Freeze_previous_value" />',
           '				<valuetableentry value="3" lowerBound="3" upperBound="3" description="User_Defined" displayString="User_Defined" />',
           '			</valuetable>',
           '        </variable>',
           '        <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="" name="{0}" comment="" bitcount="32" isSigned="false" encoding="65001" type="int" startValue="0" minValue="0" minValuePhys="0"/>'.format(sysvar_name_cnt_cycle)]

    return temp_xml

def generate_SigGrp_sysvar(sysvar_name):
    """
    generates xml variables for CRC handling for signal group

    Args:
      sysvar_name (str): crc_variable name

    Returns:
        updated list of xml

    """
    sysvar_name_fault_inject = sysvar_name + "_FaultInject"
    sysvar_name_cnt_cycle = sysvar_name + "_FaultInject_CntCycle"
    
    temp_xml = ['        <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="" name="{0}" comment="" bitcount="32" isSigned="false" encoding="65001" type="int" startValue="0" minValue="0" minValuePhys="0" maxValue="3" maxValuePhys="3">'.format(sysvar_name_fault_inject),
           '			<valuetable name="{0}" definesMinMax="true">'.format(sysvar_name_fault_inject),
           '				<valuetableentry value="0" lowerBound="0" upperBound="0" description="Disable" displayString="Disable" />',
           '				<valuetableentry value="1" lowerBound="1" upperBound="1" description="Enabled:Set_to_zero" displayString="Enabled:Set_to_zero" />',
           '				<valuetableentry value="2" lowerBound="2" upperBound="2" description="Enabled:Freeze_previous_value" displayString="Enabled:Freeze_previous_value" />',
           '			</valuetable>',
           '        </variable>',
           '        <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="" name="{0}" comment="" bitcount="32" isSigned="false" encoding="65001" type="int" startValue="0" minValue="0" minValuePhys="0"/>'.format(sysvar_name_cnt_cycle)]
    return temp_xml

def create_rb_sysvar(wb, sysvarsheet, variant_list):
    """
    generate sysvars for rbs signals

    Args:
      wb (workbook): workbook of excel
      sysvarsheet (str): sysvarshett

    Returns:
        updated xml content

    """
    global col_name,col_siggrp,col_pdu,col_pdu_type,col_payload_pdu_type,col_payload_pdu,col_selector_field_signal,col_msg,col_msgID,\
        col_startbit,col_pdu_length_byte,col_signal_length_bit,col_initial_value,col_max_value,col_cycletime,col_texttable,col_enum,\
        col_value_type,col_comment,col_dlc,col_variant,col_value_table,col_endtoendprotection,col_byte_order,col_factor,col_offset,col_minimum,\
        col_maximum,col_unit,col_cantp_pdu_type,col_BS,col_Addformate,col_padding,col_STmin,col_MaxFC,col_sender, col_gw, col_multicanoe, col_entity, col_network_type,\
            col_input_file, col_dbc_file_index,col_network_name, col_header_id, col_payload_pdu_length, col_signal_base_type, col_transmission_mode,col_payload_pdu_dlc
    logger.info("++++++++++++Start RBS xml generation+++++++++++++++")
    bus_xml = create_bus_sysvar(wb.sheetnames)
    ws = wb[sysvarsheet]
    del wb
    sysvar_df = pd.DataFrame(ws.values)
    sysvar_df = sysvar_df.replace(np.nan, '', regex=True)
    col_names = list(sysvar_df.iloc[0])
    logger_info = {}
    try:
        #col_index = [col_names.index(col) for col in required_restbus_can_column_names]
        col_index = [col_names.index(col) for col in col_names if col not in variant_list]
        col_index_TP = [col_names.index(col) for col in column_names_TP]
        [col_name,col_siggrp,col_pdu,col_pdu_type,col_payload_pdu_type,col_payload_pdu,col_selector_field_signal,col_msg,col_msgID,col_header_id,
        col_startbit,col_pdu_length_byte,col_signal_length_bit,col_payload_pdu_length,col_signal_base_type,col_initial_value,col_max_value,col_transmission_mode,col_cycletime,col_texttable,col_enum,
        col_value_type,col_comment,col_dlc,col_payload_pdu_dlc,col_variant,col_value_table,col_endtoendprotection,col_byte_order,col_factor,col_offset,col_minimum,
        col_maximum,col_unit,col_cantp_pdu_type,col_BS,col_Addformate,col_padding,col_STmin,col_MaxFC,col_sender, col_gw, col_multicanoe, col_entity, col_network_type,col_input_file, col_dbc_file_index,col_network_name] = col_index
        # [col_BS, col_Addformate, col_padding, col_STmin, col_MaxFC, col_msgID] = col_index_TP
        Col_TP = [col_BS, col_Addformate, col_padding, col_STmin, col_MaxFC, col_msgID]
    except Exception as e:
        logger.error(f"New column might be added or removed from excel CAN sysvar sheet, please check and update the col index in script -> {e}")
        raise Exception("check if new columns are added in excel, check the order of columns as well")
        #logger.error(f"Column names in Excel are different from the Script OR -->{e}")

    # sysvar_df = sysvar_df.drop(sysvar_df[sysvar_df[col_pdu_type] == 'NM-PDU'].index)
    xml = ['<?xml version="1.0" encoding="utf-8"?>']
    xml.append('<systemvariables version="4">')
    network_names = list(sysvar_df[col_network_name].loc[1:].unique())
    for nwt_name in network_names:
        container_pdu_list = []
        logger_info[nwt_name] = 0
        xml.append('  <namespace name="{0}" comment="" interface="">'.format(nwt_name))
        df_nwt = sysvar_df[sysvar_df[col_network_name] == nwt_name]

        npdu_dict = map_npdu_to_ipdu_dict(df_nwt, col_msg, col_pdu, col_pdu_type, col_payload_pdu, column_indexed = False)
        securedPdu_dict = map_securedpdu_to_ipdu_dict(df_nwt, col_msg, col_pdu, col_pdu_type, col_payload_pdu, column_indexed = False)
        NPDU_Identifier_dict, NPDU_RX_Identifier_dict = get_NPDU_Identifier_dict(df_nwt, col_msg, col_pdu, col_pdu_type,col_msgID, col_payload_pdu, column_indexed = False)
        #filter_nw=df_nwt.iloc[1:]
        for payload_pdu_type, payload_pdu_type_df in df_nwt.groupby(col_payload_pdu_type):
            #print(payload_pdu_type)
            if (payload_pdu_type == "MULTIPLEXED-I-PDU"):
                #print(payload_pdu_type)
                for msg,msg_df in payload_pdu_type_df.groupby(col_msg):
                    msg_name = str(list(msg_df[col_msg].unique())[0])
                    main_payload_pdu =str(list(msg_df[col_payload_pdu].unique())[0])
                    try:
                        cycle_time = float(list(msg_df[col_payload_pdu].unique())[0])
                    except ValueError as e:
                        cycle_time = 0
                    xml.append('    <namespace name="{0}" comment="" interface="">'.format(main_payload_pdu))
                    xml.append(
                        '        <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="" name="{0}_ON_OFF" comment="" bitcount="64" isSigned="false" encoding="65001" type="int" startValue="0" minValue="0" maxValue="1" />'.
                        format(main_payload_pdu))
                    if (cycle_time != 0):
                        xml.append('        <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="" name="{0}_CycleTime" comment="" bitcount="64" isSigned="true" encoding="65001" type="int" startValue="{1}" minValue="0"/>'.
                            format(main_payload_pdu, cycle_time))
                    
                    for row_index, row in msg_df.iterrows():
                        signal_name = str(row[col_name])
                        pdu_name = str(row[col_pdu])
                        selector_field_code = str(row[col_selector_field_signal])
                        min_value = int(row[col_minimum])
                        max_value=int(row[col_maximum])
                        if(selector_field_code == "yes"):
                            xml.append('        <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="" name="{0}_Selection_Field_Code" comment="" bitcount="64" isSigned="true" encoding="65001" type="float" startValue="0" minValue="{1}" maxValue="{2}" />'.format(
                                    signal_name,min_value,max_value))
                    xml.append('   </namespace>')

            else:
                entity_namespaces = list(payload_pdu_type_df[col_entity].unique())
                for entity in entity_namespaces:
                    #df_ent = filter_dataframe(payload_pdu_type_df, col_entity, entity)
                    df_ent = payload_pdu_type_df[payload_pdu_type_df[col_entity] == entity]
                    col_namespace = getColNamespace(entity.lower())

                    df_ent = df_ent[
                        df_ent[col_namespace].str.contains(diag_msg_pattern, flags=re.IGNORECASE, regex=True) == False]  # Filter out Diag
                    df_ent = df_ent[
                        df_ent[col_namespace].str.contains(flt_evt_msg_pattern, flags=re.IGNORECASE, regex=True) == False]  # Filter out FLT_EVT
                    df_ent = df_ent[
                        df_ent[col_namespace].str.contains(stbm_msg_pattern, flags=re.IGNORECASE, regex=True) == False]  # Filter out STBM

                    check_duplicates_columns = [col_name, col_namespace, col_payload_pdu_type, col_sender,col_network_name]  # in a network
                    df_duplicates = df_ent[df_ent.duplicated(subset=check_duplicates_columns)]
                    if not (df_duplicates.empty):
                        for i, row in df_duplicates.iterrows():
                            logger.warning(f"{row[0]} signal DUPLICATED in {nwt_name} network")

                    df_ent = df_ent.drop_duplicates(subset=check_duplicates_columns, keep="first")
                    all_namespace = list(df_ent[col_namespace].unique())
                    logger_info[nwt_name] += len(all_namespace)
                    i = 0
                    for namespace in all_namespace:
                        df_namespace = df_ent[df_ent[col_namespace] == namespace]
                        #df_namespace = df_ent[df_ent[col_namespace]==namespace]

                        pdu_type = df_namespace[col_pdu_type].values[0]
                        main_payload_pdu_type =str (list(df_namespace[col_payload_pdu_type].unique())[0])


                        xml_enum = []
                        if "N-PDU" in pdu_type:
                            xml.append('    <namespace name="{0}" comment="" interface="">'.format(str(df_namespace[col_sender].values[0])))
                            xml.append('        <namespace name="{0}" comment="" interface="">'.format(namespace))
                            for j in range(0, len(col_index_TP)):
                                min_value = df_namespace[Col_TP[j]].values[0]
                                if type(min_value) != "str":
                                    enum_signal = sysvar_df[col_texttable].values[i]
                                else:
                                    enum_signal = "texttable"
                                initial_value = df_namespace[Col_TP[j]].values[0]
                                signal_name = osektp_param_dict[column_names_TP[j]]
                                msg_name = df_namespace[col_namespace].values[0]  # to check duplicate signal with same msg ID

                        # if "_Data" in msg_name:
                        #     if re.sub("Data$","FC",msg_name,flags=re.M|re.I) in NPDU_Identifier_dict.keys():
                        #         RX_Identifier=int(NPDU_Identifier_dict[re.sub("Data$","FC",msg_name,flags=re.M|re.I)],16)
                        # elif "_FC" in msg_name:
                        #     if re.sub("FC$","Data",msg_name,flags=re.M|re.I) in NPDU_Identifier_dict.keys():
                        #         RX_Identifier=int(NPDU_Identifier_dict[re.sub("FC$","Data",msg_name,flags=re.M|re.I)],16)
                        # else:
                        #     RX_Identifier=None

                                RX_Identifier = NPDU_RX_Identifier_dict.get(msg_name, None)
                                if RX_Identifier is not None:
                                    RX_Identifier = int(RX_Identifier, 16)

                                if "Address_formate" in signal_name or "Padding_active" in signal_name:
                                    data_type = "string"
                                elif "sysTxIdentifier" in signal_name:
                                    data_type = "longlong"
                                    initial_value=int(initial_value,16)

                                else:
                                    data_type = "longlong" if enum_signal == "texttable" else "float"

                                ##                if type(signal_name)==str:
                                ##                    data_type="string"
                                value_type = df_namespace[col_value_type].values[0]
                                signed = "true" if value_type == "Signed" else "false"

                                # correct unit format
                                try:
                                    unit = str(df_namespace[col_unit].values[0]).encode("latin-1").decode("utf-8")
                                except:
                                    unit = str(df_namespace[col_unit].values[0])

                                # remove all the accents
                                unit = unidecode.unidecode(unit)

                                factor = float(1)

                                offset = float(0)

                                min_value = df_namespace[Col_TP[j]].values[0]
                                if data_type == "string":
                                    min_value = str(min_value)
                                else:
                                    try:
                                        min_value = int(min_value)
                                    except:
                                        min_value = 0

                                max_value = df_namespace[Col_TP[j]].values[0]
                                if data_type == "string":
                                    max_value = str(max_value)
                                else:
                                    try:
                                        max_value = int(max_value)
                                    except:
                                        max_value = 0

                                if data_type == "string":
                                    initial_value = str(initial_value)
                                else:
                                    try:
                                        initial_value = int(initial_value)
                                    except:
                                        initial_value = 0

                                # initial_value =float(0)#Raw Adoptation::Need to be removed while considering phyiscal

                                ##                if (initial_value < min_value) or (initial_value > max_value):
                                ##                    #logger.warning(f"{signal_name} - > initial value not in range of min & max, TAKEN: initial_value = min_value")
                                ##                    initial_value = min_value
                                ##
                                ##                if (min_value >= max_value):
                                ##                    logger.error(f"{signal_name} - > min value is more or equal to max value, TAKEN: min_value=max_value")
                                ##                    raise Exception(f"{signal_name} - > min value is more or equal to max value, TAKEN: min_value=max_value")
                                if data_type != "string":
                                    if (min_value < 0):
                                        signed = "true"

                                # generate xml for enum
                                ##                if data_type != "float":
                                ##                    factor=int(factor)
                                ##                    offset = int(offset)
                                ##                    min_value=int(min_value)
                                ##                    max_value=int(max_value)
                                ##                    initial_value=int(initial_value)

                                # xml_enum = []
                                # print('enum signal loop different namespace is ' + str(enum_signal))
                                # if enum_signal == '':
                                # write signal without enums
                                # print("appended "+str(signal_name))
                                # print("min_value "+str(min_value))
                                xml_enum.append('                <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="{0}" name="{1}" comment="" bitcount="64" isSigned="{2}" encoding="65001" type="{3}" startValue="{4}" minValue="{5}" maxValue="{6}" />'.format(unit, signal_name, signed, data_type, initial_value, min_value, max_value))
                            xml_enum.append('                <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="{0}" name="{1}" comment="" bitcount="64" isSigned="{2}" encoding="65001" type="{3}" startValue="{4}" minValue="{5}" maxValue="{6}" />'.format("", "sysECUAddress", "false", "int", "0", "0", "0"))
                            xml_enum.append('                <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="{0}" name="{1}" comment="" bitcount="64" isSigned="{2}" encoding="65001" type="{3}" startValue="{4}" minValue="{5}" maxValue="{6}" />'.format("", "sysTargetAddress", "false", "int", "0", "0", "0"))
                            xml_enum.append('                <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="{0}" name="{1}" comment="" bitcount="64" isSigned="{2}" encoding="65001" type="{3}" startValue="{4}" minValue="{5}" maxValue="{6}" />'.format("", "sysAddrExtension", "false", "int", "0", "0", "0"))
                            xml_enum.append('                <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="{0}" name="{1}" comment="" bitcount="64" isSigned="{2}" encoding="65001" type="{3}" startValue="{4}" minValue="{5}" maxValue="{6}" />'.format("", "sysBaseAddress", "false", "int", "0", "0", "0"))
                            xml_enum.append('                <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="{0}" name="{1}" comment="" bitcount="64" isSigned="{2}" encoding="65001" type="{3}" startValue="{4}" minValue="{5}" maxValue="{6}" />'.format("", "sysUseExtId", "false", "int", "0", "0", "0"))
                            xml_enum.append('                <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="{0}" name="{1}" comment="" bitcount="64" isSigned="{2}" encoding="65001" type="{3}" startValue="{4}" minValue="{5}" maxValue="{6}" />'.format("", "sysMaxFDFrameLen", "false", "int", "64", "0", "0"))
                            if str(df_namespace[col_payload_pdu].values[0]) in pdu_skip:
                                xml_enum.append('                <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="{0}" name="{1}" comment="" bitcount="64" isSigned="{2}" encoding="65001" type="{3}" startValue="{4}" minValue="{5}" maxValue="{6}" />'.format("", "sysUseFC", "false", "int", "1", "0", "1"))
                            else:
                                xml_enum.append('                <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="{0}" name="{1}" comment="" bitcount="64" isSigned="{2}" encoding="65001" type="{3}" startValue="{4}" minValue="{5}" maxValue="{6}" />'.format("", "sysUseFC", "false", "int", "0", "0", "1"))
                            xml_enum.append('                <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="{0}" name="{1}" comment="" bitcount="64" isSigned="{2}" encoding="65001" type="{3}" startValue="{4}" minValue="{5}" maxValue="{6}" />'.format("", "sysUseFCBlockSize", "false", "int", "0", "0", "0"))
                            xml_enum.append('                <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="{0}" name="{1}" comment="" bitcount="64" isSigned="{2}" encoding="65001" type="{3}" startValue="{4}" minValue="{5}" maxValue="{6}" />'.format("", "sysStartSN", "false", "int", "0", "0", "0"))
                            xml_enum.append('                <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="{0}" name="{1}" comment="" bitcount="64" isSigned="{2}" encoding="65001" type="{3}" startValue="{4}" minValue="{5}" maxValue="{6}" />'.format("", "sysMaxReceiveLength", "false", "int", "424749", "0", "424749"))
                            xml_enum.append('                <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="{0}" name="{1}" comment="" bitcount="64" isSigned="{2}" encoding="65001" type="{3}" startValue="{4}" minValue="{5}" maxValue="{6}" />'.format("", "sysBitRateSwitch", "false", "int", "1", "0", "0"))
                            xml_enum.append('                <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="{0}" name="{1}" comment="" bitcount="64" isSigned="{2}" encoding="65001" type="{3}" startValue="{4}" minValue="{5}" maxValue="{6}" />'.format("", "sysDataTransmittedConf", "false", "int", "0", "0", "0"))
                            xml_enum.append('                <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="{0}" name="{1}" comment="" bitcount="64" isSigned="{2}" encoding="65001" type="{3}" startValue="{4}" minValue="{5}" maxValue="{6}" />'.format("", "sysDataReceivedInd", "false", "int", "0", "0", "0"))
                            xml_enum.append('                <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="{0}" name="{1}" comment="" bitcount="64" isSigned="{2}" encoding="65001" type="{3}" startValue="{4}" minValue="{5}" maxValue="{6}" />'.format("", "sysErrorInd", "false", "int", "0", "0", "0"))
                            xml_enum.append('                <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="{0}" name="{1}" comment="" bitcount="64" isSigned="{2}" encoding="65001" type="{3}" startValue="{4}" minValue="{5}" maxValue="{6}" />'.format("", "sysRxIdentifier", "false", "longlong", RX_Identifier, "0", "0"))
                            xml_enum.append('                <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="{0}" name="{1}" comment="" bitcount="64" isSigned="{2}" encoding="65001" type="{3}" startValue="{4}" minValue="{5}" maxValue="{6}" />'.format("", "sysSendData", "false", "int", "0", "0", "0"))
                            xml_enum.append('                <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="{0}" name="{1}" comment="" bitcount="64" isSigned="{2}" encoding="65001" type="{3}" startValue="{4}" minValue="{5}" maxValue="{6}" />'.format("", "sysPaddingValue", "false", "int", "0", "0", "0"))
                            xml_enum.append('                <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="{0}" name="{1}" comment="" bitcount="64" isSigned="{2}" encoding="65001" type="{3}" startValue="{4}" minValue="{5}" maxValue="{6}" />'.format("", "sysAddrMode", "false", "int", "0", "0", "0"))
                            xml_enum.append('                <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="{0}" name="{1}" comment="" bitcount="64" isSigned="{2}" encoding="65001" type="{3}" startValue="{4}" minValue="{5}" maxValue="{6}" />'.format("", "sysFlowControlDelay", "false", "int", "0", "0", "0"))


                            # non_enum_count += 12
                            df_namespace = filter_sysvarDF_IL(df_namespace, col_name)

                            xml.append('            <namespace name="{0}" comment="" interface="">'.format("DefaultSettings"))
                            xml.extend(xml_enum)
                            xml.append('            </namespace>')
                            xml.append('            <namespace name="{0}" comment="" interface="">'.format("GeneralSettings"))
                            xml.extend(xml_enum)
                            xml.append('                <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="{0}" name="{1}" comment="" bitcount="64" isSigned="{2}" encoding="65001" type="{3}" startValue="{4}" minValue="{5}" maxValue="{6}" />'.format("", "sysResetToDefaultSettings", "false", "float", "0", "0", "0"))
                            xml.append('            </namespace>')
                            xml.append('            <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="{0}" name="{1}" comment="" bitcount="64" isSigned="{2}" encoding="65001" type="{3}" startValue="{4}" minValue="{5}" maxValue="{6}" />'.format("", "sysNoOfBytesToFill", "false", "int", "0", "0", "0"))
                            xml.append('            <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="{0}" name="{1}" comment="" bitcount="64" isSigned="{2}" encoding="65001" type="{3}" minValue="{4}" maxValue="{5}" />'.format("", "sysDataToTransmit", "false", "data", "0", "0"))
                            xml.append('            <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="{0}" name="{1}" comment="" bitcount="64" isSigned="{2}" encoding="65001" type="{3}" startValue="{4}" minValue="{5}" maxValue="{6}" />'.format("", "sysFillData", "false", "int", "0", "0", "0"))
                            xml.append('            <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="{0}" name="{1}" comment="" bitcount="64" isSigned="{2}" encoding="65001" type="{3}" startValue="{4}" minValue="{5}" maxValue="{6}" />'.format("", "sysClearData", "false", "int", "0", "0", "0"))
                            xml.append('            <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="{0}" name="{1}" comment="" bitcount="64" isSigned="{2}" encoding="65001" type="{3}" startValue="{4}" minValue="{5}" maxValue="{6}" />'.format("", "sysNoOfBytesReceived", "false", "int", "0", "0", "0"))
                            xml.append('            <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="{0}" name="{1}" comment="" bitcount="64" isSigned="{2}" encoding="65001" type="{3}" startValue="{4}" minValue="{5}" maxValue="{6}" />'.format("", "sysNoOfBytesToSend", "false", "int", "0", "0", "0"))
                            xml.append('            <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="{0}" name="{1}" comment="" bitcount="64" isSigned="{2}" encoding="65001" type="{3}" minValue="{4}" maxValue="{5}" />'.format("", "sysDataReceived", "false", "data", "0", "0"))
                            xml.append('            <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="{0}" name="{1}" comment="" bitcount="64" isSigned="{2}" encoding="65001" type="{3}" startValue="{4}" minValue="{5}" maxValue="{6}" />'.format("", "sysDataToTransmit_Status", "false", "string", "0", "0", "0"))

                            xml.append('        </namespace>')
                        elif "SECURED-I-PDU" in pdu_type or "MULTIPLEXED-I-PDU" in pdu_type:
                            continue
                        elif ("I-SIGNAL-I-PDU" in pdu_type) and (main_payload_pdu_type == "MULTIPLEXED-I-PDU"):
                            continue
                        else:
                            xml.append('    <namespace name="{0}" comment="" interface="">'.format(namespace))
                            try:
                                cycle_time = int(list(df_namespace[col_cycletime].unique())[0])
                            except ValueError as e:
                                cycle_time = 0
                            pdu_type = list(df_namespace[col_pdu_type].unique())[0]
                            payload_pdu_type = list(df_namespace[col_payload_pdu_type].unique())[0]

                            if ("I-SIGNAL-I-PDU" in pdu_type) and payload_pdu_type == 'N-PDU': # for i-signal-i-pdu which are npdu
                                temp_df = df_namespace
                                temp_df["BUFFER"] = temp_df[col_name].apply(lambda x: x.split("_")[2] if re.match(Npdu_sig_patterns, x, flags=re.I) else "-")
        ##                        xml.append('        <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="{0}" name="{1}" comment="" bitcount="64" isSigned="{2}" encoding="65001" type="{3}" startValue="{4}" minValue="{5}" maxValue="{6}" />'.format("", "Read_No_signals", "false", "int", "0", "0", "500"))
        ##                        xml.append('        <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="{0}" name="{1}" comment="" bitcount="64" isSigned="{2}" encoding="65001" type="{3}" startValue="{4}" minValue="{5}" maxValue="{6}" />'.format("", "Write_No_signals", "false", "int", "0", "0", "500"))
        ##                        #temp_df = temp_df.sort_values(by='BUFFER')
                                for buffer, buffer_df in temp_df.groupby("BUFFER"):
                                    if buffer=="-":
                                        for i, row in buffer_df.iterrows():
                                            row_data = read_values_from_row_can(row)
                                            data_type = "float"
                                            if row_data['max_value'] == "" or row_data['min_value'] == "":
                                                is_signed = "true"
                                                min_max_section = ""
                                            else:
                                                min_max_section = f'minValue="{row_data["min_value"]}" maxValue="{row_data["max_value"]}"'
                                            # if 'OPAQUE' in row_data['byte_order'].upper():
                                                # xml.append('        <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="{0}" name="{1}" comment="" bitcount="64" isSigned="{2}" encoding="65001" type="{3}" arrayLength="{4}" {5} />'.format(
                                                #         row_data['unit'], row_data['signal_name'], row_data['signed'], 'data', '0', min_max_section))
                                            if 'OPAQUE' in row_data['byte_order'].upper():
                                                xml.append('        <namespace name="{0}" comment="" interface="">'.format(df_namespace[col_name].values[0]))
                                                xml.append('            <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="{0}" name="{1}" comment="" bitcount="64" isSigned="{2}" encoding="65001" type="{3}" arrayLength="{4}" {5} />'.format(row_data['unit'], 'SignalValue', row_data['signed'], 'data', '0', min_max_section))
                                                xml.append('            <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="{0}" name="{1}" comment="" bitcount="64" isSigned="{2}" encoding="65001" type="{3}" startValue="{4}" minValue="{5}" maxValue="{6}" />'.format("", "sysFillData", "false", "int", "0", "0", "0"))
                                                xml.append('            <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="{0}" name="{1}" comment="" bitcount="64" isSigned="{2}" encoding="65001" type="{3}" startValue="{4}" minValue="{5}" maxValue="{6}" />'.format("", "sysNoOfBytesToFill", "false", "int", "0", "0", "0"))
                                                xml.append('            <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="{0}" name="{1}" comment="" bitcount="64" isSigned="{2}" encoding="65001" type="{3}" startValue="{4}" minValue="{5}" maxValue="{6}" />'.format("", "sysClearData", "false", "int", "0", "0", "0"))
                                                xml.append('            <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="{0}" name="{1}" comment="" bitcount="64" isSigned="{2}" encoding="65001" type="{3}" startValue="{4}" minValue="{5}" maxValue="{6}" />'.format("", "sysNoOfBytesToSend", "false", "int", "0", "0", "0"))                                                
                                                xml.append('        </namespace>')
                                                xml.append('        <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="" name="rxBuffercount" comment="" bitcount="64" isSigned="true" encoding="65001" type="longlong" startValue="{0}" minValue="{1}" minValuePhys="{2}" maxValue="{3}" maxValuePhys="{4}" />'.format("0", "0", "0", "0", "0"))
                                            else:
                                                xml.append('        <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="{0}" name="{1}" comment="" bitcount="64" isSigned="{2}" encoding="65001" type="{3}" startValue="{4}" {5} />'.format(
                                                    row_data['unit'], row_data['signal_name'], row_data['signed'], data_type, row_data['initial_value'], min_max_section))

                                    else:
                                        buffer_fullname = "_".join(buffer_df[col_name].values[0].split("_")[:-1])
                                        pdu_name = buffer_df[col_pdu].values[0]
                                        vsysvar_namespace = re.sub(r"\d+$", "", buffer_fullname)
                                        xml.append(f'        <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="" name="{buffer_fullname}" comment="" bitcount="32" isSigned="true" encoding="65001" type="struct" structDefinition="{pdu_name}::{vsysvar_namespace}" />')
                            elif ("I-SIGNAL-I-PDU" in pdu_type) and ('OPAQUE' in df_namespace[col_byte_order].values[0].upper()):
                                if df_namespace[col_minimum].values[0] == "" or df_namespace[col_maximum].values[0] == "":
                                    min_max_section = ""
                                else:
                                    min_max_section = f'minValue="{df_namespace[col_minimum].values[0]}" maxValue="{df_namespace[col_maximum].values[0]}"'
                                # xml.append('        <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="{0}" name="{1}" comment="" bitcount="64" isSigned="{2}" encoding="65001" type="{3}" arrayLength="{4}" {5} />'.format(
                                #         df_namespace[col_unit].values[0], df_namespace[col_name].values[0], "false", 'data', '0', min_max_section))
                                xml.append('        <namespace name="{0}" comment="" interface="">'.format(df_namespace[col_name].values[0]))
                                xml.append('            <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="{0}" name="{1}" comment="" bitcount="64" isSigned="{2}" encoding="65001" type="{3}" arrayLength="{4}" {5} />'.format(df_namespace[col_unit].values[0], 'SignalValue', "false", 'data', '0', min_max_section))
                                xml.append('            <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="{0}" name="{1}" comment="" bitcount="64" isSigned="{2}" encoding="65001" type="{3}" startValue="{4}" minValue="{5}" maxValue="{6}" />'.format("", "sysFillData", "false", "int", "0", "0", "0"))
                                xml.append('            <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="{0}" name="{1}" comment="" bitcount="64" isSigned="{2}" encoding="65001" type="{3}" startValue="{4}" minValue="{5}" maxValue="{6}" />'.format("", "sysNoOfBytesToFill", "false", "int", "0", "0", "0"))
                                xml.append('            <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="{0}" name="{1}" comment="" bitcount="64" isSigned="{2}" encoding="65001" type="{3}" startValue="{4}" minValue="{5}" maxValue="{6}" />'.format("", "sysClearData", "false", "int", "0", "0", "0"))
                                xml.append('            <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="{0}" name="{1}" comment="" bitcount="64" isSigned="{2}" encoding="65001" type="{3}" startValue="{4}" minValue="{5}" maxValue="{6}" />'.format("", "sysNoOfBytesToSend", "false", "int", "0", "0", "0"))
                                xml.append('        </namespace>')

                            xml.append('        <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="" name="{0}_ON_OFF" comment="" bitcount="64" isSigned="false" encoding="65001" type="int" startValue="0" minValue="0" maxValue="1" />'.format(namespace))
                            if cycle_time>0:
                                xml.append('        <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="" name="{0}_CycleTime" comment="" bitcount="64" isSigned="true" encoding="65001" type="int" startValue="{1}" minValue="0"/>'.format(namespace, cycle_time))
                            if namespace in pdu_skip:
                                xml.append('        <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="" name="txBuffermode" comment="" bitcount="32" isSigned="true" encoding="65001" type="int" startValue="{0}" minValue="{1}" minValuePhys="{2}" maxValue="{3}" maxValuePhys="{4}" />'.format("0", "0", "0", "2", "2"))
                                

                            #if secured-i-pdu
                            if main_payload_pdu_type=='SECURED-I-PDU':
                            # if "SEC" in pdu_type.upper():
                                xml.extend(generate_SigGrp_sysvar("{0}_Freshness".format(namespace)))
                                xml.extend(generate_SigGrp_sysvar("{0}_AuthInfo".format(namespace)))


                        df_namespace = filter_sysvarDF_IL(df_namespace, col_name)

                        if not (df_namespace.empty):
                            added_signal_list = []  # to avoid duplicates
                            for i, row in df_namespace.iterrows():
                                row_data = read_values_from_row_can(row)

                                if row_data['signal_name'] in added_signal_list:
                                    continue
                                else:
                                    added_signal_list.append(row_data['signal_name'])
                                data_type = "float"  # no longlong bcz we are adding only counters and CRC which are not texttable

                                if ("N-PDU" in row_data['pdu_type']) or row_data['payload_pdu_type'] == 'N-PDU':
                                    continue

                                if filterSignalForSysvar(row_data['signal_name']):
                                    if row_data['max_value'] == "" or row_data['min_value'] == "":
                                        is_signed = "true"
                                        min_max_section = ""
                                    else:
                                        min_max_section = f'minValue="{row_data["min_value"]}" maxValue="{row_data["max_value"]}"'
                                    xml.append('        <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="{0}" name="{1}_FaultInject" comment="" bitcount="64" isSigned="{2}" encoding="65001" type="{3}" startValue="{4}" {5} />'.format(
                                            row_data['unit'], row_data['signal_name'], row_data['signed'], data_type, row_data['initial_value'], min_max_section))
                                else:
                                    #default fault injection sysvar
                                    if (row_data['input_file']=='dbc') or (is_empty_or_spaces(row_data['endtoendprotection'])):
                                        xml.extend(generate_Sig_sysvar(row_data['signal_name']))
                                    else:
                                        xml.extend(generate_SigGrp_sysvar(row_data['signal_name']))

                        xml.append('    </namespace>')
                        df_namespace_payload_pdu = df_ent[df_ent[col_namespace] == namespace]
                        payload_pdu_name = df_namespace_payload_pdu[col_payload_pdu].values[0]
                        if payload_pdu_type == "CONTAINER-I-PDU" and (payload_pdu_name not in container_pdu_list):
                            xml.append('    <namespace name="{0}" comment="" interface="">'.format(payload_pdu_name))
                            xml.append('        <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="" name="{0}_ON_OFF" comment="CONTAINER-I-PDU" bitcount="64" isSigned="false" encoding="65001" type="int" startValue="0" minValue="0" maxValue="1" />'.format(payload_pdu_name))
                            xml.append('    </namespace>')
                            container_pdu_list.append(payload_pdu_name)
                        i=i+1
        xml.append('  </namespace>')
    xml.extend(bus_xml)
    #xml.append('</systemvariables>')

    # log summary
    logger.info(f"Generated sysvars for = {network_names} networks")
    for nwt, num_msg in logger_info.items():
        logger.info(f"{num_msg} messages in {nwt}")

    return xml

def rbs_sysvar_footer():
    """
    adds footer and returns
    """
    footer = []
    footer.append('</systemvariables>')
    return footer

def create_rb_sysvar_ethernet(wb, sysvarsheet, variant_list):
    """
    create sysvars for ethernet signals

    Args:
        wb (workbook): workbook of the excel
        sysvarsheet (str): sysvar sheet name

    Returns:
        list of generated sysvars

    """

    #someip arguement namespae generation

    def get_array_info(arr):
        temp = arr.split('[')
        arr_name = temp[0]
        size = int(temp[-1].split(']')[0])
        return (arr_name, size)


    def get_signal_info(row, namespace=""):
        signal_list = []
        signal = row[col_signal]
        no_elements = row[col_no_of_elements]
        if not is_empty_or_spaces(signal):
            if signal=='2D_ARRAY_':
                signal = 'index'
            if not is_empty_or_spaces(no_elements):
                for i in range(int(no_elements)):
                    if namespace:
                        signal_list.append(namespace + '::' + signal + f'_{i}')
                    else:
                        signal_list.append(f'{signal}_{i}')
            else:
                if namespace:
                    signal_list.append(namespace + '::' + signal)
                else:
                    signal_list.append(signal)
        else:
            if namespace:
                signal_list.append(namespace)
            else:
                signal_list.append("")
        return signal_list


    def generate_namespace_for_a_column(value):
        res = []
        if value:
            if '[' in value:
                arr_name, arr_size = get_array_info(value)
                if arr_name == '2D_ARRAY_':
                    arr_name = 'index'
                for i in range(arr_size):
                    res.append(arr_name+f'_{i}')
            else:
                res.append(value)
        return res
    def generate_parameter_namespaces(df):
        col_parameter_mapping = {1: col_para_level1,
                                2:col_para_level2,
                                3:col_para_level3,
                                4:col_para_level4,
                                5:col_para_level5}

        df = df.fillna("")
        return_list = []
        
        for i, row in df.iterrows():
            namespace = ""
            result_namespaces = []
            paramerter_level_dict = {}
            
            parameter_list = []
        
            if not(is_empty_or_spaces(row[col_parameter])):
                if not(is_empty_or_spaces(row[col_no_of_parameters])):
                    for i in range(int(row[col_no_of_parameters])):
                        parameter_list.append(row[col_parameter] + f'_{i}')
                else:
                    parameter_list.append(row[col_parameter])
                    
            for i in range(1,6):
                col_name = col_parameter_mapping.get(i)
                key = f'l{i}'
                paramerter_level_dict[key] = generate_namespace_for_a_column(row[col_name])
            
            for parameter in parameter_list:
                if paramerter_level_dict['l1']:
                    for l1 in paramerter_level_dict['l1']:
                        if paramerter_level_dict['l2']:
                            for l2 in paramerter_level_dict['l2']:
                                if paramerter_level_dict['l3']:
                                    for l3 in paramerter_level_dict['l3']:
                                        if paramerter_level_dict['l4']:
                                            for l4 in paramerter_level_dict['l4']:
                                                if paramerter_level_dict['l5']:
                                                    for l5 in paramerter_level_dict['l5']:
                                                        result_namespaces.append(f'{parameter}::{l1}::{l2}::{l3}::{l4}::{l5}')
                                                else:
                                                    result_namespaces.append(f'{parameter}::{l1}::{l2}::{l3}::{l4}')
                                        
                                        else:
                                            result_namespaces.append(f'{parameter}::{l1}::{l2}::{l3}')
                                    
                                else:
                                    result_namespaces.append(f'{parameter}::{l1}::{l2}')
                        else:
                            result_namespaces.append(f'{parameter}::{l1}')
                else:
                    result_namespaces.append(f'{parameter}')
            
            
            mini = row[col_minimum]
            maxi = row[col_maximum]
            if not(is_empty_or_spaces(maxi)) and 'e' in str(maxi).lower():
                try:
                    maxi = int(maxi)
                except:
                    maxi = ''

            if not(is_empty_or_spaces(mini)) and str(mini).endswith(r'.0'):
                mini = int(mini)
            if not(is_empty_or_spaces(maxi)) and str(maxi).endswith(r'.0'):
                maxi = int(maxi)

            unit = row[col_unit]
            datatype = get_sysvar_type(row[col_signal_data_type])
            if 'nounit' in unit.lower():
                unit = ''
            is_signed = 'true'
            if not(is_empty_or_spaces(mini)) and (mini >= 0):
                is_signed = 'false'
            signal_details = {'maxValue':maxi, 'minValue':mini, 'unit':unit, 'isSigned': is_signed, 'type':datatype}
            signal_details = str(signal_details)

            final_namespace = []
            signal_info = get_signal_info(row)
            if signal_info:
                if result_namespaces:
                    for nmps in result_namespaces:
                        for signal_nm in signal_info:
                            if signal_nm:
                                final_namespace.append(f'{nmps}::{signal_nm};{signal_details}')
                            else:
                                final_namespace.append(f'{nmps};{signal_details}')
                else:
                    for signal_nm in signal_info:
                        final_namespace.append(signal_nm)
            else:
                final_namespace.append('')
                
            temp = final_namespace.copy()
            return_list.extend(temp)            


        #print(return_list)                        
        return return_list
    
    def get_parameter_xml_data(df):
        def nested_dict():
            return defaultdict(nested_dict)

        def add_to_dict(paths, nested):
            for path in paths:
                current = nested
                for node in path[:-1]:
                    current = current[node]
                current[path[-1]] = None

        def generate_xml(nested, depth=0):
            xml = []
            for key, value in nested.items():
                if value is None:
                    signal, attrs = key.split(';')
                    attrs_dict = ast.literal_eval(attrs)
                    sig_details = ""
                    for k, v in attrs_dict.items():
                        if not(is_empty_or_spaces(v)):
                            sig_details += f' {k}="{v}"'
                    xml.append(f'{"    " * depth}<variable anlyzLocal="2" readOnly="false" valueSequence="false" name="{signal}" comment="" bitcount="64" encoding="65001" {sig_details}/>')
                else:
                    xml.append(f'{"    " * depth}<namespace name="{key}" comment="" interface="">')
                    xml.extend(generate_xml(value, depth + 1))
                    xml.append(f'{"    " * depth}</namespace>')
            return xml
        
        name_space_data = generate_parameter_namespaces(df)
        input_data = []
        for nmps in name_space_data:
            input_data.append(nmps.split('::'))

        nested = nested_dict()
        add_to_dict(input_data, nested)
        eth_parameter_xml = generate_xml(nested, depth=5)
        return eth_parameter_xml

    def get_sysvar_type(datatype):
        """

        Args:
            datatype:

        Returns:

        """
        if is_empty_or_spaces(datatype):
            return 'float'

        if 'string' in datatype:
            return 'string'
        elif 'bool' in datatype:
            return 'int'
        elif ('float' in datatype) or ('double' in datatype):
            return 'float'
        elif 'int' in datatype:
            if 'uint' in datatype:
                if '64' in datatype:
                    return 'longlong'
                else:
                    return 'int'
            else:
                if '64' in datatype:
                    return 'longlong'
                else:
                    return 'int'
        else:
            return 'float'

    ## END someip arguement namespace generation

    def getFieldValueNamespace(field_type_df, isempty_list):
        field_value_namespace_list = []
        if all(isempty_list):
            for i, row in field_type_df.iterrows():
                signal = row[col_signal]
                if is_empty_or_spaces(signal):
                    signal = row[col_field_name]
                min_value = row[col_minimum]
                max_value = row[col_maximum]
                unit = row[col_unit]
                if 'nounit' in unit.lower():
                    unit = ''
                is_signed = 'true'
                if not(is_empty_or_spaces(min_value)) and (min_value >= 0):
                    is_signed = 'false'

                sig_details = f''
                if str(min_value):
                    sig_details += f'minValue = "{str(min_value)}" '
                if str(max_value):
                    sig_details += f'maxValue = "{str(max_value)}" '
                if str(unit):
                    sig_details += f'unit = "{str(unit)}" '

                sig_details += f'type = "{get_sysvar_type(row[col_signal_data_type])}" '


                #if column col_no_of_elements is not empty
                if not is_empty_or_spaces(row[col_no_of_elements]):
                    for j in range(int(row[col_no_of_elements])):
                        field_value_namespace_list.append(f'                  <variable anlyzLocal="2" readOnly="false" valueSequence="false" name="{signal}_{j}" comment="signal" bitcount="64" isSigned="{is_signed}" encoding="65001" {sig_details}/>')
                else:
                    field_value_namespace_list.append(f'                  <variable anlyzLocal="2" readOnly="false" valueSequence="false" name="{signal}" comment="signal" bitcount="64" isSigned="{is_signed}" encoding="65001" {sig_details}/>')

        else:
            for parameter, parameter_df in field_type_df.groupby(col_parameter):
                field_value_namespace_list.extend(get_parameter_xml_data(parameter_df))
        return field_value_namespace_list

    def getEventValueNamespace(event_df, isempty_list):
        event_value_namespace_list = []
        if all(isempty_list):
            for i, row in event_df.iterrows():
                signal = row[col_signal]
                if is_empty_or_spaces(signal):
                    signal = row[col_member]
                min_value = row[col_minimum]
                max_value = row[col_maximum]
                unit = row[col_unit]
                if 'nounit' in unit.lower():
                    unit = ''
                is_signed = 'true'
                if not (is_empty_or_spaces(min_value)) and (min_value >= 0):
                    is_signed = 'false'

                sig_details = f''
                if str(min_value):
                    sig_details += f'minValue = "{str(min_value)}" '
                if str(max_value):
                    sig_details += f'maxValue = "{str(max_value)}" '
                if str(unit):
                    sig_details += f'unit = "{str(unit)}" '

                sig_details += f'type = "{get_sysvar_type(row[col_signal_data_type])}" '

                # if column col_no_of_elements is not empty
                if not is_empty_or_spaces(row[col_no_of_elements]):
                    for j in range(int(row[col_no_of_elements])):
                        event_value_namespace_list.append(f'                  <variable anlyzLocal="2" readOnly="false" valueSequence="false" name="{signal}_{j}" comment="signal" bitcount="64" isSigned="{is_signed}" encoding="65001" {sig_details}/>')
                else:
                    event_value_namespace_list.append(f'                  <variable anlyzLocal="2" readOnly="false" valueSequence="false" name="{signal}" comment="signal" bitcount="64" isSigned="{is_signed}" encoding="65001" {sig_details}/>')

        else:
            for parameter, parameter_df in event_df.groupby(col_parameter):
                event_value_namespace_list.extend(get_parameter_xml_data(parameter_df))
        return event_value_namespace_list


    logger.info(f"#---Start sysvar generation for Ethernet -> {sysvarsheet}---#")
    eth_xml = []
    if sysvarsheet not in wb.sheetnames:
        logger.info(f"#---Ethernet {sysvarsheet} not found ---#")
        return []

    ws = wb[sysvarsheet]


    del wb
    sysvar_df = pd.DataFrame(ws.values)
    if sysvar_df.shape[0] <= 1:
        logger.info(f"#---Ethernet {sysvarsheet} is empty ---#")
        return []

    sysvar_df = sysvar_df.replace(np.nan, '', regex=True)

    col_names = list(sysvar_df.iloc[0])
    sysvar_df = sysvar_df.iloc[1:]
    try:
        col_index = [col_names.index(col) for col in col_names if col not in variant_list]
        [col_service, col_service_instance_name,col_service_id, col_major_version, col_minor_version, col_instance_id, col_member_type, col_event_group,col_EventGroup_ID,col_field_name,
         col_member, col_member_id, col_field_type, col_parameter, col_parameter_type, col_no_of_parameters, col_para_level1, col_para_level2, 
         col_para_level3, col_para_level4, col_para_level5,col_signal, col_no_of_elements, col_fire_and_forget,col_signal_group,col_pdu,col_pdu_type,
         col_payload_pdu_type,col_payload_pdu,col_selector_field_signal,col_startbit,col_pdu_length_byte,col_signal_length_bit,col_initial_value,col_max_value,
         col_cycletime,col_texttable,col_enum,col_value_type,col_comment,col_dlc,col_variant,col_value_table,col_endtoendprotection, col_signal_data_type,
         col_signal_byte_order, col_factor, col_offset, col_minimum, col_maximum, col_unit, col_ip_address, col_udp_port, col_tcp_port, col_tp_protocol, col_vlan_id, col_vlan_name, col_sd_type, col_nwt_protocol, col_autosar_type,
         col_network_name, col_node_name] = col_index

    except Exception as e:
        logger.error(f"Column names in Excel are different from the Script OR -->{e}")
        logger.error(';'.join(required_columns_rb))

    # remove services of given ip address
    block_nwt_endpoint_dict = {}
    try:
        block_nwt_endpoint_dict = block_network_endpoint_dict  # list of tuple [(ipaddress, id),]
    except Exception as e:
        pass
    if block_nwt_endpoint_dict != {}:
        for key, block_nwt_endpoint_list in block_nwt_endpoint_dict.items():
            nwt_nm, node_nm = key.split("::")
            sysvar_df = sysvar_df[~((sysvar_df[col_ip_address].isin([ip for ip, id in block_nwt_endpoint_list])) &
                                    (sysvar_df[col_node_name] == node_nm) &
                                    (sysvar_df[col_network_name] == nwt_nm))]
            #sysvar_df = sysvar_df[~sysvar_df[col_ip_address].isin([ip for ip, id in block_nwt_endpoint_list])]

    for network_name, network_df in sysvar_df.groupby(col_network_name):
        eth_xml.append(f'  <namespace name="{network_name}" comment="" interface="">')

        for node_name, node_df in network_df.groupby(col_node_name):
            eth_xml.append(f'    <namespace name="{node_name}" comment="" interface="">')

            for nwt_protocol, nwt_protocol_df in node_df.groupby(col_nwt_protocol):
                if nwt_protocol.upper() == 'ETH_SOMEIP':
                    for SDType, SDType_df in nwt_protocol_df.groupby(col_sd_type):
                        servive_type,control_namespace = ("CONSUMED_SERVICES","Subscribe") if "consume" in SDType.lower() else ("PROVIDED_SERVICES","Provide")
                        eth_xml.append(f'      <namespace name="{servive_type}" comment="" interface="">')

                        for srvID, serviceID_df in SDType_df.groupby(col_service_id):
                            for MajVer, MajVer_df in serviceID_df.groupby(col_major_version):
                                for MinVer, MinVer_df in MajVer_df.groupby(col_minor_version):
                                    for InstID, InstID_df in MinVer_df.groupby(col_instance_id):
                                        unique_udp_ports = InstID_df[col_udp_port].unique()
                                        for udp_port, udp_port_df in InstID_df.groupby(col_udp_port):
                                            service_name = udp_port_df[col_service].iloc[0]
                                            autosar_type = udp_port_df[col_autosar_type].iloc[0]
                                            if 'adaptive' in autosar_type.lower():
                                                service_instance_name = udp_port_df[col_service_instance_name].iloc[0]
                                                service_comment = service_instance_name + ' | ' + service_name
                                            else:
                                                service_comment = service_name
                                            port_namespace = f'_{udp_port}' if ('consume' in SDType.lower() and len(unique_udp_ports)>1) else ''
                                            eth_xml.append(f'        <namespace name="sif_{srvID}_{MajVer}_{MinVer}_{InstID}{port_namespace}" comment="{service_comment}" interface="">')
                                            if 'provide' in SDType.lower():
                                                eth_xml.append(f'          <namespace name="CONTROLS" comment="" interface="">')
                                                eth_xml.append(f'           <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="" name="{control_namespace}" comment="" bitcount="32" isSigned="false" encoding="65001" type="int" startValue="1" minValue="0" minValuePhys="0" maxValue="1" maxValuePhys="1">')
                                                eth_xml.append(f'             <valuetable name="{control_namespace}" definesMinMax="true">')
                                                eth_xml.append(f'               <valuetableentry value="0" lowerBound="0" upperBound="0" description="No" displayString="No" />')
                                                eth_xml.append(f'               <valuetableentry value="1" lowerBound="1" upperBound="1" description="Yes" displayString="Yes" />')
                                                eth_xml.append(f'             </valuetable>')
                                                eth_xml.append(f'           </variable>')
                                                eth_xml.append(f'          </namespace>') #service CONTROLS

                                            for MemberType, MemberType_df in udp_port_df.groupby(col_member_type):
                                                if "method" in MemberType.lower():  # if method
                                                    eth_xml.append(f'          <namespace name="METHODS" comment="" interface="">')

                                                    for Method, Method_df in MemberType_df.groupby(col_member):

                                                        eth_xml.append(f'            <namespace name="{Method}" comment="" interface="">')
                                                        eth_xml.append(f'              <namespace name="CONTROLS" comment="" interface="">')
                                                        # if "PROVIDE" in servive_type.upper():
                                                        #     eth_xml.append(f'                <variable anlyzLocal="2" readOnly="false" valueSequence="false" name="Enable" comment="" bitcount="64" isSigned="false" encoding="65001" type="longlong"  startValue="0" minValue="0" maxValue="1" />')
                                                        if "CONSUME" in servive_type.upper():
                                                            eth_xml.append(f'                <variable anlyzLocal="2" readOnly="false" valueSequence="false" name="SendOnce" comment="" bitcount="64" isSigned="false" encoding="65001" type="longlong"  startValue="0" minValue="0" maxValue="1" />')
                                                        eth_xml.append(f'                <variable anlyzLocal="2" readOnly="false" valueSequence="false" name="RequestID" comment="" bitcount="64" isSigned="false" encoding="65001" type="longlong"  startValue="0" minValue="0" maxValue="1" />')

                                                        if not('true' in list(Method_df[col_fire_and_forget]) and 'PROVIDE' in servive_type.upper()):
                                                            eth_xml.append(f'                <variable anlyzLocal="2" readOnly="false" valueSequence="false" name="ReturnCode" comment="" bitcount="64" isSigned="false" encoding="65001" type="longlong"  startValue="0" minValue="0" maxValue="1" />')

                                                        eth_xml.append(f'              </namespace>')  # CONTROLS

                                                        for parameter_type, parameter_type_df in Method_df.groupby(col_parameter_type):
                                                            if is_empty_or_spaces(parameter_type):
                                                                continue
                                                            eth_xml.append(f'              <namespace name="{parameter_type}" comment="" interface="">')
                                                            for parameter, parameter_df in parameter_type_df.groupby(col_parameter):
                                                                eth_xml.extend(get_parameter_xml_data(parameter_df))
                                                            eth_xml.append(f'              </namespace>')  #parameter_type

                                                        eth_xml.append(f'            </namespace>')  #namespace method_name

                                                    eth_xml.append(f'          </namespace>')  #METHODS
                                                elif "event" in MemberType.lower():  # if method
                                                    eth_xml.append(f'          <namespace name="EVENTGROUPS" comment="" interface="">')
                                                    for evggrp_id, evnt_grp_df in MemberType_df.groupby(col_event_group):
                                                        eth_xml.append(f'            <namespace name="{evggrp_id}" comment="" interface="">')

                                                        if control_namespace=="Subscribe":  # only needed for consume event group
                                                            eth_xml.append(f'              <namespace name="CONTROLS" comment="" interface="">')
                                                            eth_xml.append(f'                <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="" name="{control_namespace}" comment="" bitcount="32" isSigned="false" encoding="65001" type="int" startValue="1" minValue="0" minValuePhys="0" maxValue="1" maxValuePhys="1">')
                                                            eth_xml.append(f'                  <valuetable name="{control_namespace}" definesMinMax="true">')
                                                            eth_xml.append(f'                    <valuetableentry value="0" lowerBound="0" upperBound="0" description="No" displayString="No" />')
                                                            eth_xml.append(f'                    <valuetableentry value="1" lowerBound="1" upperBound="1" description="Yes" displayString="Yes" />')
                                                            eth_xml.append(f'                  </valuetable>')
                                                            eth_xml.append(f'                </variable>')
                                                            eth_xml.append(f'              </namespace>')  # evnt_gropu CONTROLS

                                                        eth_xml.append(f'              <namespace name="EVENTS" comment="" interface="">')
                                                        for evnt, evntdf in evnt_grp_df.groupby(col_member):
                                                            # check_column = [col_no_of_parameters, col_para_level1,
                                                            #                 col_para_level2, col_para_level3,
                                                            #                 col_para_level4, col_para_level5]
                                                            # isempty_list = []
                                                            # for col in check_column:
                                                            #     is_empty_col = []
                                                            #     for cell in evntdf[col]:
                                                            #         if is_empty_or_spaces(cell):
                                                            #             is_empty_col.append(True)
                                                            #         else:
                                                            #             is_empty_col.append(False)
                                                            #     if all(is_empty_col):
                                                            #         isempty_list.append(True)
                                                            #     else:
                                                            #         isempty_list.append(False)

                                                            eth_xml.append(f'                <namespace name="{evnt}" comment="" interface="">')
                                                            eth_xml.append(f'                  <namespace name="CONTROLS" comment="" interface="">')
                                                            if "PROVIDE" in servive_type.upper():
                                                                eth_xml.append(f'                <variable anlyzLocal="2" readOnly="false" valueSequence="false" name="SendOnce" comment="" bitcount="64" isSigned="false" encoding="65001" type="longlong"  startValue="0" minValue="0" maxValue="1" />')

                                                            eth_xml.append(f'                  </namespace>') #event controls

                                                            # eth_xml.append(f'                  <namespace name="Value" comment="" interface="">')
                                                            # # ---#TEMP_FIX
                                                            # # Check if any event_name has multiple event_ids
                                                            # has_duplicates = evntdf.groupby(col_member)[col_member_id].nunique().gt(1).any()
                                                            # if not(has_duplicates):
                                                            #     temp_ip = evntdf[col_ip_address].iloc[0]
                                                            #     evntdf = evntdf[evntdf[col_ip_address] == temp_ip]
                                                            #     evntdf[col_parameter] = evnt
                                                            #     eth_xml.extend(getEventValueNamespace(evntdf, isempty_list))
                                                            # # ---
                                                            # eth_xml.append(f'                  </namespace>')  # event Value

                                                            eth_xml.append(f'                </namespace>') #evnt

                                                        eth_xml.append(f'              </namespace>')  # EVENTS namepce

                                                        eth_xml.append(f'            </namespace>') #evnt_grp name
                                                    eth_xml.append(f'          </namespace>') #EVENTGROUPS namespace

                                                elif "field" in MemberType.lower():
                                                    eth_xml.append(f'          <namespace name="FIELDS" comment="" interface="">')
                                                    for field, field_df in MemberType_df.groupby(col_field_name):
                                                        field_df[col_parameter] = field
                                                        field_df.loc[field_df[col_field_type] == 'getter', col_parameter_type] = 'RETURN_PARAMETER'
                                                        field_df.loc[field_df[col_field_type] == 'setter', col_parameter_type] = 'IN_PARAMETER'


                                                        check_column = [col_no_of_parameters, col_para_level1, col_para_level2,col_para_level3, col_para_level4, col_para_level5]
                                                        isempty_list = []
                                                        for col in check_column:
                                                            is_empty_col = []
                                                            for cell in field_df[col]:
                                                                if is_empty_or_spaces(cell):
                                                                    is_empty_col.append(True)
                                                                else:
                                                                    is_empty_col.append(False)
                                                            if all(is_empty_col):
                                                                isempty_list.append(True)
                                                            else:
                                                                isempty_list.append(False)


                                                        eth_xml.append(f'            <namespace name="{field}" comment="" interface="">')
                                                        eth_xml.append(f'              <namespace name="Value" comment="" interface="">') # main Value

                                                        temp_filed_type = field_df[col_field_type].unique()[0]
                                                        main_value_df = field_df[field_df[col_field_type] == temp_filed_type]
                                                        #---
                                                        eth_xml.extend(getFieldValueNamespace(main_value_df, isempty_list))
                                                        #---
                                                        eth_xml.append(f'              </namespace>') # main Value close

                                                        for field_type, field_type_df in field_df.groupby(col_field_type):
                                                            fld_map = {'getter':'GET', 'setter':'SET', 'notifier':'NOTIFY'}
                                                            field_type_namespace = fld_map.get(field_type.lower(), '')
                                                            field_type_name = field_type_df[col_member].iloc[0]

                                                            if field_type == 'notifier' and "PROVIDE" in servive_type.upper():
                                                                eth_xml.append(f'              <namespace name="{field_type_namespace}" comment="{field_type_name}" interface="">')
                                                                eth_xml.append(f'                <namespace name="CONTROLS" comment="" interface="">')
                                                                eth_xml.append(f'                  <variable anlyzLocal="2" readOnly="false" valueSequence="false" name="SendOnce" comment="" bitcount="64" isSigned="false" encoding="65001" type="longlong"  startValue="0" minValue="0" maxValue="1" />')
                                                                eth_xml.append(f'                </namespace>') #controls
                                                                eth_xml.append(f'              </namespace>')  #field type
                                                            elif field_type in ['getter','setter']:
                                                                eth_xml.append(f'              <namespace name="{field_type_namespace}" comment="{field_type_name}" interface="">')
                                                                if "CONSUME" in servive_type.upper():
                                                                    eth_xml.append(f'                <namespace name="CONTROLS" comment="" interface="">')
                                                                    eth_xml.append(f'                  <variable anlyzLocal="2" readOnly="false" valueSequence="false" name="SendOnce" comment="" bitcount="64" isSigned="false" encoding="65001" type="longlong"  startValue="0" minValue="0" maxValue="1" />')
                                                                    eth_xml.append(f'                </namespace>')  # CONTROLS

                                                                    eth_xml.append(f'                <namespace name="RETURN_PARAMETERS" comment="" interface="">')
                                                                    eth_xml.append(f'                  <variable anlyzLocal="2" readOnly="false" valueSequence="false" name="RequestID" comment="" bitcount="64" isSigned="false" encoding="65001" type="longlong"  startValue="0" minValue="0" maxValue="1" />')
                                                                    eth_xml.append(f'                  <variable anlyzLocal="2" readOnly="false" valueSequence="false" name="ReturnCode" comment="" bitcount="64" isSigned="false" encoding="65001" type="longlong"  startValue="0" minValue="0" maxValue="1" />')
                                                                    eth_xml.append(f'                </namespace>')  # RETURN_PARAMETER
                                                                if "PROVIDE" in servive_type.upper():
                                                                    eth_xml.append(f'                <namespace name="RETURN_PARAMETERS" comment="" interface="">')
                                                                    eth_xml.append(f'                  <variable anlyzLocal="2" readOnly="false" valueSequence="false" name="ReturnCode" comment="" bitcount="64" isSigned="false" encoding="65001" type="longlong"  startValue="0" minValue="0" maxValue="1" />')
                                                                    eth_xml.append(f'                </namespace>')  # RETURN_PARAMETER

                                                                if field_type == 'setter':
                                                                    eth_xml.append(f'                <namespace name="IN_PARAMETERS" comment="" interface="">')
                                                                    eth_xml.append(f'                <namespace name="Value" comment="Signal values" interface="">')
                                                                    eth_xml.extend(getFieldValueNamespace(field_type_df, isempty_list))
                                                                    eth_xml.append(f'                </namespace>')  # Value
                                                                    if "PROVIDE" in servive_type.upper():
                                                                        eth_xml.append(f'                  <variable anlyzLocal="2" readOnly="false" valueSequence="false" name="RequestID" comment="" bitcount="64" isSigned="false" encoding="65001" type="longlong"  startValue="0" minValue="0" maxValue="1" />')
                                                                    eth_xml.append(f'                </namespace>')  # IN_PARAMETER
                                                                if 'getter' in field_type.lower() and "PROVIDE" in servive_type.upper():
                                                                    eth_xml.append(f'                <namespace name="IN_PARAMETERS" comment="" interface="">')
                                                                    eth_xml.append(f'                  <variable anlyzLocal="2" readOnly="false" valueSequence="false" name="RequestID" comment="" bitcount="64" isSigned="false" encoding="65001" type="longlong"  startValue="0" minValue="0" maxValue="1" />')
                                                                    eth_xml.append(f'                </namespace>')


                                                                eth_xml.append(f'              </namespace>')  #field type


                                                                # if "CONSUME" in servive_type.upper() and 'SETTER' in field_type.upper():
                                                                #     eth_xml.append(f'                <namespace name="Value" comment="" interface="">')
                                                                #     eth_xml.extend(getFieldValueNamespace(field_type_df, isempty_list))
                                                                #     eth_xml.append(f'                </namespace>')  # Value

                                                            # elif field_type == 'setter':
                                                            #     eth_xml.append(f'                <namespace name="CONTROLS" comment="" interface="">')
                                                            #     if "CONSUME" in servive_type.upper():
                                                            #         eth_xml.append(f'                  <variable anlyzLocal="2" readOnly="false" valueSequence="false" name="SendOnce" comment="" bitcount="64" isSigned="false" encoding="65001" type="longlong"  startValue="0" minValue="0" maxValue="1" />')
                                                            #     eth_xml.append(f'                  <variable anlyzLocal="2" readOnly="false" valueSequence="false" name="RequestID" comment="" bitcount="64" isSigned="false" encoding="65001" type="longlong"  startValue="0" minValue="0" maxValue="1" />')
                                                            #     eth_xml.append(f'                  <variable anlyzLocal="2" readOnly="false" valueSequence="false" name="ReturnCode" comment="" bitcount="64" isSigned="false" encoding="65001" type="longlong"  startValue="0" minValue="0" maxValue="1" />')
                                                            #     eth_xml.append(f'                </namespace>')  # CONTROLS
                                                            #     eth_xml.append(f'                <namespace name="Value" comment="" interface="">')
                                                            #     for parameter, parameter_df in field_type_df.groupby(col_parameter):
                                                            #         eth_xml.extend(get_parameter_xml_data(parameter_df))
                                                            #     eth_xml.append(f'                </namespace>')  # IN para

                                                        eth_xml.append(f'            </namespace>')  #field

                                                    eth_xml.append(f'          </namespace>')  #FIELDS

                                            eth_xml.append(f'        </namespace>')  #service

                        eth_xml.append(f'      </namespace>')  #type of service provide/consume namespace

                elif nwt_protocol.upper() == 'ETH_PDU':
                    eth_xml.append(f'      <namespace name="PDUs" comment="" interface="">')
                    for payload_pdu_type, payload_pdu_type_df in nwt_protocol_df.groupby(col_payload_pdu_type):
                        if is_empty_or_spaces(payload_pdu_type):  #normal eth pdu
                            for normal_pdu, normal_pdu_df in  payload_pdu_type_df.groupby(col_pdu):
                                eth_xml.append(f'        <namespace name="{normal_pdu}" comment="" interface="">')
                                eth_xml.append(f'            <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="" name="{normal_pdu}_ON_OFF" comment="" bitcount="64" isSigned="false" encoding="65001" type="int" startValue="0" minValue="0" maxValue="1" />')
                                try:
                                    cycle_time = float(normal_pdu_df.iloc[0][col_cycletime])
                                except:
                                    cycle_time = 0
                                if cycle_time>0:
                                    eth_xml.append(f'            <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="" name="{normal_pdu}_CycleTime" comment="" bitcount="64" isSigned="true" encoding="65001" type="float" startValue="{cycle_time}" minValue="0"/>')
                                counter_data_df =  filter_sysvarDF_IL(normal_pdu_df, col_signal)  #filter to check if counters/crc are present
                                if not(counter_data_df.empty):
                                    for i, row in counter_data_df.iterrows():
                                        eth_xml.extend(generate_SigGrp_sysvar(row[col_signal]))
                                eth_xml.append(f'        </namespace>')  #pdu name
                        elif "MULTIPLEXED" in payload_pdu_type:
                            for main_multiplex_pdu, main_multiplex_pdu_df in payload_pdu_type_df.groupby(col_payload_pdu):
                                multiplexed_pdu_count = main_multiplex_pdu_df[col_pdu].nunique()
                                eth_xml.append(f'        <namespace name="{main_multiplex_pdu}" comment="" interface="">')
                                eth_xml.append(f'            <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="" name="{main_multiplex_pdu}_ON_OFF" comment="" bitcount="64" isSigned="false" encoding="65001" type="int" startValue="0" minValue="0" maxValue="1" />')
                                selction_field_df = main_multiplex_pdu_df[main_multiplex_pdu_df[col_selector_field_signal]=='yes']
                                selectior_field_signal = selction_field_df.iloc[0][col_signal]
                                eth_xml.append(f'            <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="" name="{selectior_field_signal}_Selection_Field_Code" comment="" bitcount="32" isSigned="true" encoding="65001" type="int" startValue="0" minValue="0" minValuePhys="0" maxValue="{multiplexed_pdu_count}" maxValuePhys="{multiplexed_pdu_count}" />')
                                eth_xml.append(f'        </namespace>')  #pdu name
                    eth_xml.append(f'      </namespace>')  #PDUs namespace

            eth_xml.append(f'    </namespace>')  #node name

        eth_xml.append(f'  </namespace>')   #network name

    # this is temporary fix for CAPL keyword handling need to be revisited-----------------#TEMP_FIX
    eth_xml = [line.replace('name="message"', 'name="message_"') if 'name="message"' in line else line for line in eth_xml]
    # ----------------------------------------------
    return eth_xml

def read_values_from_row_can(row):
    """
    reads values from the dataframe row , checks integrity if fails assigns default

    Args:
      row (dataframe row object): dataframe row

    Returns:
        dict: column value mapping

    """
    #takes row as input and send read values from row as dict, this is needed because we sometimes we cant take vales from excel directly
    row_data = {}
    try:
        #add all must needed values from excel(cannot be empty) which we cant assign
        pdu_name = row[col_pdu]
        pdu_type = row[col_pdu_type]
        payload_pdu_type = row[col_payload_pdu_type]
        signal_name = row[col_name]
        message_name = row[col_msg]
        byte_order = row[col_byte_order]
        value_type = row[col_value_type].lower()
        endtoendprotection = row[col_endtoendprotection]
        input_file_type = row[col_input_file]
        
    except:
        logger.error(f"EMPTY CELL: Error while reading startbit/bitlength for signal -> {row[col_name]}")
        raise Exception(f"EMPTY CELL: Error while reading startbit/bitlength for signal -> {row[col_name]}")
    

    #read values which we can assgn if empty (give user defined values if empty or not correct)
    try:
        factor = float(row[col_factor])    #format_float_positional is to convert exponential values to float eg: 3.046e-05 to float:: min_digits is to add .0 if int 
        if not(factor):
            factor = float(1)
    except:
        logger.warning(f"EMPTY CELL: Error occured will reading factor of signal -> {row[col_name]}")
        factor = float(1)
    try:
        offset = float(row[col_offset])
        if not(offset):
            offset = float(1)
    except:
        logger.warning(f"EMPTY CELL: Error occured will reading offset of signal -> {row[col_name]}")
        offset = float(0)
    
    try:
        initial_value = float(row[col_ini_value])
    except:
        initial_value = float(0)

    try:
        unit = str(row[col_unit]).encode("latin-1").decode("utf-8")
    except:
        unit = str(row[col_unit])
    unit = unidecode.unidecode(unit)

    try:
        min_value = float(row[col_minimum])
        max_value = float(row[col_maximum])
    except Exception as e:
        logger.warning(f"Error while reading MIN/MAX value msg: {row[col_name]} -->{e} , taking 2 power length as max")
        
        #need to be reviewed *******************
        min_value = float(0)
        try:
            max_value = float(row[col_max_value])
        except:
            max_value = ""
        #raise Exception(f"Error while reading MIN/MAX value -->{e}")


    #checks and changes if needed
   
    if max_value != "":
        if (initial_value < min_value) or (initial_value > max_value):
            initial_value = min_value

    signed = "true" if value_type == "signed" else "false"
    if (min_value < 0):
        signed = "true"

    if max_value != "":
        if (min_value >= max_value):
            logger.warning(f"{signal_name} - > min value is more or equal to max value, TAKEN: min_value=0 max_value= (2^length) -1")
            #commented exception because too many signals have this error and need to proceed
            #raise Exception(f"{signal_name} - > min value is more or equal to max value, TAKEN: min_value=max_value")

    if max_value != "":
        row_data = {'pdu_name':pdu_name, 'pdu_type':pdu_type, 'payload_pdu_type':payload_pdu_type, 'signal_name':signal_name, 'message_name':message_name, 'byte_order': byte_order, 'factor':np.format_float_positional(factor, min_digits=1),
                    'offset':np.format_float_positional(offset, min_digits=1), 'initial_value':np.format_float_positional(initial_value, min_digits=1), 'value_type':value_type,
                    'unit':unit, 'min_value':np.format_float_positional(min_value, min_digits=1), 'max_value':np.format_float_positional(max_value, min_digits=1),
                    'endtoendprotection':endtoendprotection,'signed':signed, 'input_file':input_file_type}
    else:
        row_data = {'pdu_name': pdu_name, 'pdu_type': pdu_type, 'payload_pdu_type':payload_pdu_type, 'signal_name': signal_name,
                    'message_name': message_name, 'byte_order': byte_order, 'factor': np.format_float_positional(factor, min_digits=1),
                    'offset': np.format_float_positional(offset, min_digits=1),
                    'initial_value': np.format_float_positional(initial_value, min_digits=1), 'value_type': value_type,
                    'unit': unit, 'min_value': np.format_float_positional(min_value, min_digits=1),
                    'max_value': "",
                    'endtoendprotection': endtoendprotection, 'signed': signed, 'input_file':input_file_type}
    return row_data


def read_values_from_row_eth(row):
    """
    reads values from the dataframe row , checks integrity if fails assigns default

    Args:
      row (dataframe row object): dataframe row

    Returns:
        dict: column value mapping

    """
    #takes row as input and send read values from row as dict, this is needed because we sometimes we cant take vales from excel directly
    row_data = {}
    try:
        #add all must needed values from excel(cannot be empty) which we cant assign
        pdu_name = row[col_pdu]
        pdu_type = row[col_pdu_type]
        signal_name = row[col_name]
        message_name = row[col_msg]
        value_type = row[col_value_type].lower()
        endtoendprotection = row[col_endtoendprotection]
        
    except:
        logger.error(f"EMPTY CELL: Error while reading startbit/bitlength for signal -> {row[col_name]}")
        raise Exception(f"EMPTY CELL: Error while reading startbit/bitlength for signal -> {row[col_name]}")
    

    #read values which we can assgn if empty (give user defined values if empty or not correct)
    try:
        factor = float(row[col_factor])    #format_float_positional is to convert exponential values to float eg: 3.046e-05 to float:: min_digits is to add .0 if int 
        if not(factor):
            factor = float(1)
    except:
        logger.warning(f"EMPTY CELL: Error occured will reading factor of signal -> {row[col_name]}")
        factor = float(1)
    try:
        offset = float(row[col_offset])
        if not(offset):
            offset = float(1)
    except:
        logger.warning(f"EMPTY CELL: Error occured will reading offset of signal -> {row[col_name]}")
        offset = float(0)
    
    try:
        initial_value = float(row[col_ini_value])
    except:
        initial_value = float(0)

    try:
        unit = str(row[col_unit]).encode("latin-1").decode("utf-8")
    except:
        unit = str(row[col_unit])
    unit = unidecode.unidecode(unit)

    try:
        min_value = float(row[col_minimum])
        max_value = float(row[col_maximum])
    except Exception as e:
        logger.warning(f"Error while reading MIN/MAX value msg: {row[col_name]} -->{e} , taking 2 power length as max")
        
        #need to be reviewed *******************
        min_value = float(0)
        try:
            max_value = float(row[col_max_value])
        except:
            max_value = ""
        #raise Exception(f"Error while reading MIN/MAX value -->{e}")


    #checks and changes if needed
   
    if max_value != "":
        if (initial_value < min_value) or (initial_value > max_value):
            initial_value = min_value

    signed = "true" if value_type == "signed" else "false"
    if (min_value < 0):
        signed = "true"

    if max_value != "":
        if (min_value >= max_value):
            logger.warning(f"{signal_name} - > min value is more or equal to max value, TAKEN: min_value=0 max_value= (2^length) -1")
            #commented exception because too many signals have this error and need to proceed
            #raise Exception(f"{signal_name} - > min value is more or equal to max value, TAKEN: min_value=max_value")

    if max_value != "":
        row_data = {'pdu_name':pdu_name, 'pdu_type':pdu_type, 'signal_name':signal_name, 'message_name':message_name, 'factor':np.format_float_positional(factor, min_digits=1),
                    'offset':np.format_float_positional(offset, min_digits=1), 'initial_value':np.format_float_positional(initial_value, min_digits=1), 'value_type':value_type,
                    'unit':unit, 'min_value':np.format_float_positional(min_value, min_digits=1), 'max_value':np.format_float_positional(max_value, min_digits=1),
                    'endtoendprotection':endtoendprotection,'signed':signed}
    else:
        row_data = {'pdu_name': pdu_name, 'pdu_type': pdu_type, 'signal_name': signal_name,
                    'message_name': message_name, 'factor': np.format_float_positional(factor, min_digits=1),
                    'offset': np.format_float_positional(offset, min_digits=1),
                    'initial_value': np.format_float_positional(initial_value, min_digits=1), 'value_type': value_type,
                    'unit': unit, 'min_value': np.format_float_positional(min_value, min_digits=1),
                    'max_value': "",
                    'endtoendprotection': endtoendprotection, 'signed': signed}
    return row_data



def generate_sysvar(sysvar_df, FDX=False, Classe=False):
    """
    generate sysvar main for rbs

    Args:
      sysvar_df (dataframe): sysvar dataframe
      FDX (bool):  (Default value = False)
      Classe (bool):  (Default value = False) - deprecated

    Returns:
        generated list of xml

    """
    global dict_duplicateCheck

    dict_duplicateCheck = {"": {}}
    xml = []
    col_namespace = col_msg

    # get column index of FDX
    if FDX:
        try:
            col_names = list(sysvar_df.iloc[0])
            col_data_type = col_names.index("data_type")
        except:
            logger.error("Column data_type is not found in FDX Excel.")
            return False

    # sysvar_df = sysvar_df[sysvar_df[col_namespace].str.contains(diag_msg_pattern, regex=True)==False]#Filter out Diag
    # sysvar_df = sysvar_df[sysvar_df[col_namespace].str.contains(flt_evt_msg_pattern, regex=True)==False]#Filter out FLT_EVT
    # sysvar_df = sysvar_df[sysvar_df[col_namespace].str.contains(stbm_msg_pattern, regex=True)==False]#Filter out STBM

    num_of_rows = sysvar_df.shape[0]
    signal_count = num_of_rows - 1

    # count signal/message
    enum_count = 0
    non_enum_count = 0
    message_count = 0
    # case normal signal
    for i in range(1, num_of_rows):
        # print('we are on i = ' + str(i))
        enum_signal = sysvar_df[col_texttable].values[i]
        signal_name = sysvar_df[col_name].values[i]

        if FDX:
            data_type = "longlong" if "int" in sysvar_df[col_data_type].values[i] else "float"
        else:
            data_type = "longlong" if enum_signal == "texttable" else "float"

        value_type = sysvar_df[col_value_type].values[i]
        signed = "true" if value_type == "Signed" else "false"

        # correct unit format
        try:
            unit = str(sysvar_df[col_unit].values[i]).encode("latin-1").decode("utf-8")
        except:
            unit = str(sysvar_df[col_unit].values[i])

        # remove all the accents
        unit = unidecode.unidecode(unit)

        factor = sysvar_df[col_factor].values[i]
        try:
            factor = float(factor)
        except:
            factor = float(1)

        offset = sysvar_df[col_offset].values[i]
        try:
            offset = float(offset)
        except:
            offset = float(0)

        min_value = sysvar_df[col_min].values[i]
        try:
            min_value = float(min_value)
        except:
            min_value = float(0)

        max_value = sysvar_df[col_max].values[i]
        try:
            max_value = float(max_value)
        except:
            max_value = float(0)

        if FDX:
            try:
                if offset != 0:
                    min_value = int((float(sysvar_df[col_min].values[i]) // float(sysvar_df[col_factor].values[i])) - (
                            offset // float(sysvar_df[col_factor].values[i])))
                else:
                    min_value = int(float(sysvar_df[col_min].values[i]) // float(sysvar_df[col_factor].values[i]))
                    if signed == "true" and min_value == 0:
                        min_value = -65535
                        min_value = int(min_value)
            except:
                min_value = 0
                logger.error(f"max_value is not an integer: {signal_name}")
        else:
            min_value = sysvar_df[col_min].values[i]
            try:
                min_value = float(min_value)
            except:
                min_value = float(0)

        max_value = sysvar_df[col_max].values[i]

        if FDX:
            try:
                max_value = int(max_value)
            except:
                max_value = 0
                logger.error(f"max_value is not an integer: {signal_name}")
        else:
            try:
                max_value = float(max_value)
            except:
                max_value = float(0)
        initial_value = sysvar_df[col_ini_value].values[i]

        try:
            initial_value = float(initial_value)
        except:
            initial_value = float(0)

        # initial_value =float(0)#Raw Adoptation::Need to be removed while considering phyiscal

        if FDX:
            initial_value = int(initial_value)
        else:
            if (initial_value < min_value) or (initial_value > max_value):
                # logger.warning(f"{signal_name} - > initial value not in range of min & max, TAKEN: initial_value = min_value")
                initial_value = min_value

            if (min_value >= max_value):
                logger.error(f"{signal_name} - > min value is more or equal to max value")
                raise Exception(f"{signal_name} - > min value is more or equal to max value")

            if (min_value < 0):
                signed = "true"

        '''
        try
            max_value = sysvar_df[col_max].values[i]
        except:
            max_value = 0
            logger.error(f"max_value is not an integer: {signal_name}")

        try:
            min_value = sysvar_df[col_min].values[i]
        except:
            min_value = 0
            logger.error(f"min_value is not an integer: {signal_name}")

        if offset != 0:
            min_value = float((float(sysvar_df[col_min].values[i]) // float(sysvar_df[col_factor].values[i])) - (offset // float(sysvar_df[col_factor].values[i])))
        else:
            min_value = float(float(sysvar_df[col_min].values[i]) // float(sysvar_df[col_factor].values[i]))
            
            if signed == "true" and min_value == 0:
                min_value = -65535
        '''

        # generate xml for enum
        if data_type != "float":
            factor = int(factor)
            offset = int(offset)
            min_value = int(min_value)
            max_value = int(max_value)
            initial_value = int(initial_value)

        xml_enum = []
        '''
        # verify if signal is array
        if '[' in signal_name and ']' in signal_name:
            signal_name = signal_name.split('[')[0]
            if signal_name not in signal_array.keys():
                signal_array[signal_name] = {}
                signal_array[signal_name]["count"] = 1
                signal_array[signal_name]["ini_value"] = [str(initial_value)]
                signal_array[signal_name]["xml_value"] = [unit, signal_name, signed, min_value, max_value]
            else:
                signal_array[signal_name]["count"] += 1
                signal_array[signal_name]["ini_value"].append(str(initial_value))
        '''

        if (enum_signal == ''):
            # write signal without enums
            xml_enum.append(
                '      <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="{0}" name="{1}" comment="" bitcount="64" isSigned="{2}" encoding="65001" type="{3}" startValue="{4}" minValue="{5}" maxValue="{6}" />'.format(
                    unit, signal_name, signed, data_type, initial_value, min_value, max_value))
            non_enum_count += 1
        elif (enum_signal == 'texttable'):
            # write signal with enums
            enum_list = sysvar_df[col_enum].values[i]
            if not enum_list:
                logger.warning(f"texttable values missing: {signal_name}")
                continue
            # split the string into a list
            parse_enum = enum_list.split('\n')
            # remove the Logical value prefix and spaces
            # parse_enum = [str(elem).strip('LogicalValue :') for elem in parse_enum]
            parse_enum = [str(elem)[14:] for elem in parse_enum]
            enum_length = len(parse_enum)
            results = []
            for j in range(enum_length):
                enum_elem = parse_enum[j]
                # split the elements, first element is the enum number and the next can be concatenated
                # correct the format of latin letters
                try:
                    enum_string = str(enum_elem).encode("latin-1").decode("utf-8")
                except:
                    enum_string = str(enum_elem)
                # remove all the accents
                enum_string = unidecode.unidecode(enum_string)
                first_elem = number_pattern.match(enum_string)
                if not first_elem:
                    logger.warning(f"texttable values not correct: {signal_name}")
                    continue
                else:
                    first_elem = first_elem.group()
                second_elem = enum_string[len(first_elem):]
                # filter illegal characters
                second_elem = second_elem.replace(' ', '')
                second_elem = second_elem.replace('&', '')
                second_elem = second_elem.replace('"', '')
                second_elem = second_elem.replace("'", '')
                second_elem = second_elem.replace('(', '')
                second_elem = second_elem.replace(')', '')
                results.append([first_elem, second_elem])
            xml_enum.append(
                '      <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="{0}" name="{1}" comment="" bitcount="64" isSigned="{2}" encoding="65001" type="{3}" startValue="{4}" minValue="{5}" minValuePhys="{5}" maxValue="{6}" maxValuePhys="{6}">'.format(
                    unit, signal_name, signed, data_type, initial_value, min_value, max_value))
            xml_enum.append('        <valuetable name="{0}" definesMinMax="true">'.format(signal_name))
            for enum_nr, enum_name in results:
                if ('0x' in enum_nr) or ('0X' in enum_nr):
                    enum_nr = int(enum_nr, 16)
                xml_enum.append(
                    '          <valuetableentry value="{0}" lowerBound="{0}" upperBound="{0}" description="{1}" displayString="{1}" />'.format(
                        enum_nr, enum_name.replace('<<', 'left shift').replace('>>',
                                                                               'right shift')))  # added replace function to avoid error in xml file when less/greater than symbol present
            xml_enum.append('        </valuetable>')
            xml_enum.append('      </variable>')
            enum_count += 1
        # first row or case where namespace is new
        if sysvar_df[col_namespace].values[i - 1] != sysvar_df[col_namespace].values[i]:
            # close <namespace>
            if (i > 1):
                '''
                # append array type of signals
                if signal_array.keys():
                    for key in signal_array.keys():
                        xml.append('      <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="{0}" name="{1}" comment="" bitcount="64" isSigned="{2}" encoding="65001" type="floatarray" arrayLength="{5}" startValue="{6}" minValue="{3}" maxValue="{4}" />'.format(*signal_array[key]["xml_value"], signal_array[key]["count"], ';'.join(signal_array[key]["ini_value"])))
                    signal_array = {}
                '''
                xml.append('    </namespace>')
            # write first namespace Column 6
            namespace = sysvar_df[col_namespace].values[i]

            xml.append('    <namespace name="{0}" comment="" interface="">'.format(namespace))
            # create the frame ON_OFF variable, if not FDX nor Classe
            message_count += 1
        # extend xml_enum
        xml.extend(xml_enum)
    '''
    # append array type of signals
    if signal_array.keys():
        for key in signal_array.keys():
            xml.append('      <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="{0}" name="{1}" comment="" bitcount="64" isSigned="{2}" encoding="65001" type="floatarray" arrayLength="{5}" startValue="{6}" minValue="{3}" maxValue="{4}" />'.format(*signal_array[key]["xml_value"], signal_array[key]["count"], ';'.join(signal_array[key]["ini_value"])))
    '''
    # xml.append('    </namespace>')
    logger.info("=============== Summary ===============")
    logger.info(f"The number of total signals: {signal_count}")
    logger.info(f"The number of signals with texttable: {enum_count}")
    logger.info(f"The number of signals without texttable: {non_enum_count}")
    logger.info(f"The number of total messages: {message_count}")
    return xml


def generate_sysvar_classe(sysvar_df):
    """
    generate sysvar main Classe

    Args:
      sysvar_df (dataframe): classe database dataframe

    Returns:
        generated list of sysvars

    """
    xml = []
    
    try:
        col_namespace = col_msg
        namespaces = sysvar_df[col_namespace].unique().tolist()[1:]
        for namespace in namespaces:
            df = sysvar_df[sysvar_df[col_namespace] == namespace]
            num_of_rows = df.shape[0]
            signals_with_struct = len(df[df[col_name].str.contains('::')])
            xml.append(f'    <namespace name="{namespace}" comment="" interface="">')
            namespace_index = len(xml) - 1
            # count signal/message
            enum_count = 0
            non_enum_count = 0
            message_count = 1
            # case normal signal
            struct_signals = 0
            signals_in_namespace = 0
            for i in range(0, num_of_rows):
                enum_signal = df[col_texttable].values[i]
                signal_name = df[col_name].values[i]
                msg_name = df[col_namespace].values[i]
    
                match enum_signal:
                    case "texttable":
                        data_type = "longlong"
                    case "int":
                        data_type = "int"
                    case "string":
                        data_type = "string"
                    case "float":
                        data_type = "float"
                    case "intarray":
                        data_type = "intarray"
                    case "floatarray":
                        data_type = "floatarray"
                    case "int_struct":
                        data_type = "int"
                    case "int64":
                        data_type = "longlong"
                    case "float_struct":
                        data_type = "float"
                    case "texttable_struct":
                        data_type = "int"
                    case "texttable_array":
                        data_type = "intarray"
                    case "struct_struct":
                        # Set the data type to the string in 'texttable values' column
                        data_type = df[col_enum].values[i]
                    case "struct":
                        data_type = df[col_enum].values[i]
    
                value_type = df[col_value_type].values[i]
                signed = "true" if value_type == "Signed" else "false"
    
                # correct unit format
                try:
                    unit = str(df[col_unit].values[i]).encode("latin-1").decode("utf-8")
                except:
                    unit = str(df[col_unit].values[i])
    
                # remove all the accents
                unit = unidecode.unidecode(unit)
                factor = df[col_factor].values[i]
    
                try:
                    factor = float(factor)
                except:
                    factor = float(1)
    
                offset = df[col_offset].values[i]
    
                try:
                    offset = float(offset)
                except:
                    offset = float(0)
    
                min_value = df[col_min].values[i]
    
                try:
                    min_value = float(min_value)
                except:
                    min_value = float(0)
    
                max_value = df[col_max].values[i]
    
                try:
                    max_value = float(max_value)
                except:
                    max_value = float(0)
    
                col_length = 6  # Column Length [Bit] index
                bit_count = int(df[col_length].values[i])
    
                if max_value < min_value:
                    max_value = min_value
    
                initial_value = df[col_ini_value].values[i]
                res = isinstance(initial_value, str)
    
                if res is False or initial_value == "na" or (str(initial_value).strip() == ""):
                    try:
                        initial_value = float(initial_value)
                    except:
                        initial_value = float(0)
    
                    if (initial_value < min_value) or (initial_value > max_value):
                        initial_value = min_value
    
                if min_value >= max_value:
                    logger.error(f"{signal_name} - > min value is more or equal to max value, TAKEN: min_value=max_value")
                    raise Exception(
                        f"{signal_name} - > min value is more or equal to max value, TAKEN: min_value=max_value")
    
                if min_value < 0:
                    signed = "true"
    
                # generate xml for enum
                if data_type != "float" and data_type != "string":
                    factor = int(factor)
                    offset = int(offset)
                    min_value = int(min_value)
                    max_value = int(max_value)
                    if initial_value != "na":
                        initial_value = int(initial_value)
    
                xml_enum = []
    
                # Create signals with longlong data type #
                if enum_signal == 'int' or enum_signal == 'int64' or enum_signal == 'float':
                    xml_enum.append(
                        f'      <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="{unit}" context="CANoe_master" name="{signal_name}" comment="" bitcount="64" isSigned="{signed}" encoding="65001" type="{data_type}" startValue="{initial_value}" minValue="{min_value}" maxValue="{max_value}" />')
                    non_enum_count += 1
    
                # Create signals with string data type
                elif enum_signal == "string":
                    xml_enum.append(
                        f'      <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="" context="CANoe_master" name="{signal_name}" comment="" bitcount="64" isSigned="{signed}" encoding="65001" type="{data_type}" startValue="{initial_value}" />')
                    non_enum_count += 1
    
                elif enum_signal == "struct":
                    xml_enum.append(
                        f'      <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="" name="{signal_name}" comment="" bitcount="{bit_count}" isSigned="true" encoding="65001" type="struct" structDefinition="{data_type}" />'
                    )
                    non_enum_count += 1
    
                elif enum_signal == "int_struct" or enum_signal == "float_struct" or enum_signal == 'struct_struct' or enum_signal == "floatarray" or enum_signal == "intarray" or enum_signal == "texttable_struct" or enum_signal == "texttable_array":
                    struct_name = signal_name.split("::")[0]
                    #if struct_name == 'ts_debug_output_data_t':
                        #print("xxx")
                    if struct_signals > 0:
                        index = -abs(struct_signals) - 1
                        if struct_name not in xml[index]:
                            xml.append("      </struct>")
                            struct_var = xml[index - 1].split('"')[1].rsplit("_", 1)[0]
                            struct_var_variable = xml[index - 1].split('"')[1]
                            xml.append(
                                f'      <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="" context="CANoe_master" name="{struct_var}" comment="" bitcount="32000" isSigned="true" encoding="65001" type="struct" structDefinition="{namespace + "::" + struct_var_variable}" />')
                            struct_signals = 0
    
                    struct_header = f'      <struct name="{struct_name}" isUnion="false" definedBinaryLayout="True" comment="">'
                    reduced_xml = xml[namespace_index:]
                    if struct_header not in reduced_xml:
                        xml.append(struct_header)
    
                    if "[" in signal_name:
                        search_results = re.finditer(r'\[.*?\]', signal_name)
                        for item in search_results:
                            arr_length = item.group(0).lstrip("[").rstrip("]")
                    else:
                        arr_length = None
                    # print(signal_name.split("::")[1])
    
                    if arr_length is not None and enum_signal != "texttable_array":
                        struct_signal = f'        <structMember relativeOffset="0" byteOrder="0" isOptional="False" isHidden="False" name="{signal_name.split("::")[1].split("[")[0]}" comment="" bitcount="{bit_count}" isSigned="true" encoding="65001" type="{data_type}" arrayLength="{arr_length}" minValue="{min_value}" minValuePhys="{min_value}" maxValue="{max_value}" maxValuePhys="{max_value}" />'
                        xml.append(struct_signal)
                    else:
                        if enum_signal == "texttable_struct":
                            struct_signal = f'        <structMember relativeOffset="0" byteOrder="0" isOptional="False" isHidden="False" name="{signal_name.split("::")[1]}" comment="" bitcount="{bit_count}" isSigned="true" encoding="65001" type="{data_type}" startValue="{initial_value}" minValue="{min_value}" minValuePhys="{min_value}" maxValue="{max_value}" maxValuePhys="{max_value}" >'
                        elif enum_signal == "texttable_array":
                            struct_signal = f'        <structMember relativeOffset="0" byteOrder="0" isOptional="False" isHidden="False" name="{signal_name.split("::")[1].split("[")[0]}" comment="" bitcount="{bit_count}" isSigned="true" encoding="65001" type="{data_type}" arrayLength="{arr_length}" minValue="{min_value}" minValuePhys="{min_value}" maxValue="{max_value}" maxValuePhys="{max_value}" >'
                        elif enum_signal == "struct_struct":
                            struct_signal = f'        <structMember relativeOffset="0" byteOrder="0" isOptional="False" isHidden="False" name="{signal_name.split("::")[1]}" comment="" bitcount="{bit_count}" isSigned="true" encoding="65001" type="struct" structDefinition="{data_type}" />'
                        else:
                            if value_type:
                                struct_signal = f'        <structMember relativeOffset="0" byteOrder="0" isOptional="False" isHidden="False" name="{signal_name.split("::")[1]}" comment="" bitcount="{bit_count}" isSigned="{signed}" encoding="65001" type="{data_type}" startValue="{initial_value}" minValue="{min_value}" minValuePhys="{min_value}" maxValue="{max_value}" maxValuePhys="{max_value}" />'
                            else:
                                struct_signal = f'        <structMember relativeOffset="0" byteOrder="0" isOptional="False" isHidden="False" name="{signal_name.split("::")[1]}" comment="" bitcount="{bit_count}" isSigned="true" encoding="65001" type="{data_type}" startValue="{initial_value}" minValue="{min_value}" minValuePhys="{min_value}" maxValue="{max_value}" maxValuePhys="{max_value}" />'
                        xml.append(struct_signal)
                        if enum_signal == "texttable_struct" or enum_signal == "texttable_array":
                            xml_texttable_struct = []
                            xml_texttable_struct = generate_texttable_struct(df, i, signal_name)
                            xml.extend(xml_texttable_struct)
                    if enum_signal != "texttable_struct" and enum_signal != "texttable_array":
                        struct_signals += 1
                    else:
                        struct_signals += len(xml_texttable_struct) + 1
                    signals_in_namespace += 1
                    if signals_in_namespace == df.shape[0] or signals_in_namespace == signals_with_struct:
                        xml.append("      </struct>")
                        xml.append(
                            f'      <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="" context="CANoe_master" name="{signal_name.split("::")[0].rsplit("_", 1)[0]}" comment="" bitcount="32000" isSigned="true" encoding="65001" type="struct" structDefinition="{namespace + "::" + struct_name}" />')
    
                # Create signals with enum data type #
                elif enum_signal == 'texttable':
                    xml_enum = generate_texttable(df, xml_enum, i, unit, signal_name, signed, data_type, initial_value,
                                                  min_value, max_value)
    
                # extend xml_enum
                xml.extend(xml_enum)
            xml.append('    </namespace>')
    
            logger.info("=============== Summary ===============")
            logger.info(f"The number of total signals: {num_of_rows}")
            logger.info(f"The number of total messages: {message_count}")
        return xml
    except Exception as e:
        logger.error(traceback.format_exc())
        raise e
        


def generate_texttable(df, xml_enum, i, unit, signal_name, signed, data_type, initial_value, min_value, max_value):
    """
    generates sysvar for classe texttable type

    Args:
      df: 
      xml_enum: 
      i: 
      unit: 
      signal_name: 
      signed: 
      data_type: 
      initial_value: 
      min_value: 
      max_value: 

    Returns:

    """
    # write signal with enums
    enum_count = 0
    enum_list = df[col_enum].values[i]
    if not enum_list:
        logger.warning(f"texttable values missing: {signal_name}")
    # split the string into a list
    parse_enum = enum_list.split('\n')
    # remove the Logical value prefix and spaces
    # parse_enum = [str(elem).strip('LogicalValue :') for elem in parse_enum]
    parse_enum = [str(elem)[14:] for elem in parse_enum]
    enum_length = len(parse_enum)
    results = []
    for j in range(enum_length):
        enum_elem = parse_enum[j]
        # split the elements, first element is the enum number and the next can be concatenated
        # correct the format of latin letters
        try:
            enum_string = str(enum_elem).encode("latin-1").decode("utf-8")
        except:
            enum_string = str(enum_elem)
        enum_string = unidecode.unidecode(enum_string)
        first_elem = number_pattern.match(enum_string)
        if not first_elem:
            logger.warning(f"texttable values not correct: {signal_name}")
            continue
        else:
            first_elem = first_elem.group()
        second_elem = enum_string[len(first_elem):]
        # filter illegal characters
        second_elem = second_elem.replace(' ', '')
        second_elem = second_elem.replace('&', '')
        second_elem = second_elem.replace('"', '')
        second_elem = second_elem.replace("'", '')
        second_elem = second_elem.replace('(', '')
        second_elem = second_elem.replace(')', '')
        results.append([first_elem, second_elem])
    xml_enum.append(
        f'      <variable anlyzLocal="2" readOnly="false" valueSequence="false" unit="{unit}" context="CANoe_master" name="{signal_name}" comment="" bitcount="64" isSigned="{signed}" encoding="65001" type="{data_type}" startValue="{initial_value}" minValue="{min_value}" minValuePhys="{min_value}" maxValue="{max_value}" maxValuePhys="{max_value}">')
    xml_enum.append(f'        <valuetable name="{signal_name}" definesMinMax="true">')
    for enum_nr, enum_name in results:
        if ('0x' in enum_nr) or ('0X' in enum_nr):
            enum_nr = int(enum_nr, 0)
        xml_enum.append(
            '          <valuetableentry value="{0}" lowerBound="{0}" upperBound="{0}" description="{1}" displayString="{1}" />'.format(
                enum_nr, enum_name.replace('<<', 'left shift').replace('>>',
                                                                       'right shift')))  # added replace function to avoid error in xml file when less/greater than symbol present
    xml_enum.append('        </valuetable>')
    xml_enum.append('      </variable>')
    enum_count += 1
    return xml_enum


def generate_texttable_struct(df, i, signal_name):
    """
    generates sysvar for classe struct type

    Args:
      df: 
      i: 
      signal_name: 

    Returns:

    """
    # write signal with enums
    enum_count = 0
    xml_struct = []
    enum_list = df[col_enum].values[i]
    if not enum_list:
        logger.warning(f"texttable values missing: {signal_name}")
    # split the string into a list
    parse_enum = enum_list.split('\n')
    # remove the Logical value prefix and spaces
    # parse_enum = [str(elem).strip('LogicalValue :') for elem in parse_enum]
    parse_enum = [str(elem)[14:] for elem in parse_enum]
    enum_length = len(parse_enum)
    results = []
    for j in range(enum_length):
        enum_elem = parse_enum[j]
        # split the elements, first element is the enum number and the next can be concatenated
        # correct the format of latin letters
        try:
            enum_string = str(enum_elem).encode("latin-1").decode("utf-8")
        except:
            enum_string = str(enum_elem)
        enum_string = unidecode.unidecode(enum_string)
        first_elem = number_pattern.match(enum_string)
        if not first_elem:
            logger.warning(f"texttable values not correct: {signal_name}")
            continue
        else:
            first_elem = first_elem.group()
        second_elem = enum_string[len(first_elem):]
        # filter illegal characters
        second_elem = second_elem.replace(' ', '')
        second_elem = second_elem.replace('&', '')
        second_elem = second_elem.replace('"', '')
        second_elem = second_elem.replace("'", '')
        second_elem = second_elem.replace('(', '')
        second_elem = second_elem.replace(')', '')
        results.append([first_elem, second_elem])
    if "[" in signal_name:
        xml_struct.append(f'        <valuetable name="{signal_name.split("::")[1].split("[")[0]}_vt" definesMinMax="true">')
    else:
        xml_struct.append(f'        <valuetable name="{signal_name.split("::")[1]}_vt" definesMinMax="true">')
    for enum_nr, enum_name in results:
        if ('0x' in enum_nr) or ('0X' in enum_nr):
            enum_nr = int(enum_nr, 0)
        xml_struct.append(
            '          <valuetableentry value="{0}" lowerBound="{0}" upperBound="{0}" description="{1}" displayString="{1}" />'.format(
                enum_nr, enum_name.replace('<<', 'left shift').replace('>>',
                                                                       'right shift')))  # added replace function to avoid error in xml file when less/greater than symbol present
    xml_struct.append('        </valuetable>')
    xml_struct.append('      </structMember>')
    enum_count += 1
    return xml_struct

def save_to_file(xml, path, name):
    """
    save generated content to xml file

    Args:
      xml (list): list of generated content
      path (str): path of the file
      name (str): name of the file

    """
    # Open and save xml file
    if "fdx_init" in name:
        file = ('{0}/{1}.cin'.format(path, name))
    else:
        file = ('{0}/{1}_gen.xml'.format(path, name))
    with open(file, 'w') as outfile:
        outfile.write("\n".join(str(item) for item in xml))
        outfile.close()
    logger.info(f"{file} updated successfully")


def external_call():
    """ external call used by jenkins"""
    try:
        logger.info(f"###### START 'Create SysVar(.xml)' DEBUG INFORMATION ######")
        script_path = os.path.dirname(os.path.abspath(__file__))
        autosar_path = script_path + r'\..\..\..\..\CustomerPrj\Restbus\Autosar_Gen_Database.xlsx'
        classe_database = script_path + r'\..\..\..\..\Platform\Classe\Classe_Database.xlsx'
        customer_database = script_path + r'\..\..\..\..\CustomerPrj\Classe\Customer_Database.xlsx'
        wb = load_workbook(autosar_path, data_only=True)
        sysvarSheet = "SysVarDatabase"
        eth_sysvar_sheet = "SysVarDatabaseETH"
        variant_list = extract_vehicle_variants(autosar_path)

        xml = create_rb_sysvar(wb, sysvarSheet, variant_list)

        eth_xml = create_rb_sysvar_ethernet(wb, eth_sysvar_sheet, variant_list)
        xml.extend(eth_xml)

        xml.extend(rbs_sysvar_footer())

        save_to_file(xml, script_path + r"\..\..\..\..\CustomerPrj\Restbus\Database", sysvarSheet)
        logger.info(f"-----------------------------------END XML creation for RBS-----------------------------------------")

        databases = {"Classe Platform": [classe_database, r"\..\..\..\..\Platform\Classe\Database", "ClasseDatabase"],
                     "Classe Customer": [customer_database, r"\..\..\..\..\CustomerPrj\Classe\Database",
                                         "CustomerDatabase"]}

        for tag, database in databases.items():
            logger.info(f"+++++++++++Start {tag} xml generation++++++++++++")

            #  Update car variants in CustomerDatabase.xlsx
            if database[2] == "CustomerDatabase":
                # Read Excel file sheet by sheet
                excel_data = pd.read_excel(database[0], sheet_name=None)

                # Access each DataFrame corresponding to a sheet
                for sheet_name, df in excel_data.items():
                    # change the enumeration of vehicle variable
                    if sheet_name == "hil_ctrl":
                        filtered_index = df.index[df["Name"] == "vehicle"].tolist()
                        vehicle_texttable = ""
                        for i, vehicle in enumerate(variant_list):
                            vehicle_texttable = vehicle_texttable + f"LogicalValue: {i} " + vehicle + "\n"
                        df.at[filtered_index[0], "texttable values"] = vehicle_texttable

                    # Determine the index of column "gw"
                    column_gw_index = df.columns.get_loc('gw')
                    # Rename columns based on DBCMapping variant list
                    for index, new_column in enumerate(variant_list):
                        if len(df.columns) > column_gw_index + index + 1:
                            df.columns.values[column_gw_index + index + 1] = new_column
                        else:
                            df.insert(column_gw_index + index + 1, new_column, None)

                    # Determine the number of columns to keep based on the length of variant list
                    num_columns_to_keep = column_gw_index + len(variant_list) + 1

                    # Slice the DataFrame to keep only the required number of columns
                    df = df.iloc[:, :num_columns_to_keep]

                    # Write back to the Excel file
                    with pd.ExcelWriter(database[0], engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                        df.to_excel(writer, sheet_name=sheet_name, index=False)

            df = pd.concat(pd.read_excel(database[0], sheet_name=None), ignore_index=True)
            df.loc[-1] = list(df.columns)  # adding a row
            df.index = df.index + 1  # shifting index
            df = df.sort_index()
            classe_df = df.replace(np.nan, '', regex=True)
            define_classe_colIndex(list(classe_df.columns))
            classe_df_with_indexes = classe_df.rename(
                columns={x: y for x, y in zip(classe_df.columns, range(0, len(classe_df.columns)))})
            xml = generate_sysvar_header()
            xml.extend(generate_sysvar_classe(classe_df_with_indexes))
            xml.extend(generate_sysvar_footer_classe())
            save_to_file(xml, script_path + database[1], database[2])

    except Exception as exp:
        logger.error(f"Create Sysvar(.xlm) execution failed --> {exp}")
        traceback.print_exc()
        raise Exception(exp)
    logger.info(f"###### END 'Create SysVar(.xml)' DEBUG INFORMATION ######")
    logger.info('-' * 80)


if __name__ == "__main__":
    external_call()
