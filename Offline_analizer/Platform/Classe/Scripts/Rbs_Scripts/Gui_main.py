# -*- coding: utf-8 -*-
# @file Gui_main.py
# @author ADAS_HIL_TEAM
# @date 08-21-2023

##################################################################
# C O P Y R I G H T S
# ----------------------------------------------------------------
# Copyright (c) 2023-2024 by Robert Bosch GmbH. All rights reserved.

# The reproduction, distribution and utilization of this file as
# well as the communication of its contents to others without express
# authorization is prohibited. Offenders will be held liable for the
# payment of damages. All rights reserved in the event of the grant
# of a patent, utility model or design.

##################################################################
import traceback

import os, sys
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from openpyxl import Workbook, load_workbook
import easygui
import re
import subprocess as sp
import pandas as pd
try:
    pd.set_option('future.no_silent_downcasting', True)
except:
    pass

if "Scripts" in os.getcwd():
    os.chdir(os.path.abspath(os.path.join(os.getcwd(), "..")))
sys.path.append(os.path.dirname(os.path.abspath(__file__)) + r"\..")
from create_autosar import generate_dbc_map, update_autosar_main, extract_vehicle_variants
from update_sysvartab import generate_database, update_variants, update_variants_eth, save_database, check_sysvardatabase_integrity
from create_nodes import create_node_main, initialize_dict, extractHilCtrl,generate_sysvardatabase_dataframe
from create_nodes_Eth import external_call as createNodesEthernetMain
from create_sysvar import generate_sysvar_header, generate_sysvar_footer, create_rb_sysvar, rbs_sysvar_footer,create_rb_sysvar_ethernet,define_classe_colIndex, generate_sysvar, save_to_file, generate_sysvar_footer_classe, generate_sysvar_classe
from create_ininode import generate_dataframe_rbs, generate_variables, generate_Opaque_array_initialize, generate_ininode_senderwise, generate_ininode_classe, append_file, create_file, generate_dataframe_classe, update_initialization_file,extractMultiCanoeMapping,\
    generate_ininode_multicanoe_senderwise, generate_dataframe_rbs_ethernet,generate_ininode_senderwise_ethernet,generate_variables_ethernet
#from update_ini import update_iniout, create_ini
from create_gw import generate_gw
#from Rbs_Scripts import create_arxml_v5
from classicarxml_parser import external_call as create_arxml_can
from create_arxml_eth import external_call as create_arxml_eth
from test_database import main as test_sysvar_can_eth
try:
    from Sanity_check_scripts.arxml_sanity_check.arxml_sanity_check import main as arxml_sanity_check_main
except Exception as e:
    print(f"Failed to import arxml_sanity_check: {e}")

sys.path.append(os.path.dirname(os.path.abspath(__file__)) + r"\..")
from FDX_Scripts.fdx_to_rbs_mapping import create_mapping_main
from FDX_Scripts.fdx_database_gen import create_mapping_main_database
import Triggers_Checks_gen.Generate_Triggers_XLS_table as triggers_Gen
#from CarMaker_scripts.cm_to_rbs_mapping import create_mapping_main_cm
from Control import XCP_to_A2L_diff
import HIL_Specific_triggers.fdx_to_hil_specific_json as hil_specific
from Control.logging_config import logger
from Control.file_search import update_database_cycle_times
from Multicanoe_Scripts.create_multicanoe_cfgs import create_multicanoe_main, createCfg3
from Documentation_scripts.copyrights_generator import generate_copy_rights
from Documentation_scripts.autogen_release_note import releaseNoteAutoGeneration
try:
    from Documentation_scripts.offline_scripts_docu import generate_offline_docu
except Exception as e:
    print(f"Failed to import offline_scripts_docu: {e}")
import Create_delivery_zip.create_delivery_zip as delivery


sys.path.append(os.getcwd() + r"\..\..\..\CustomerPrj\Restbus\Scripts")
from update_gw_ruleset import update_gw
from threading import Thread
from tkinter.messagebox import askyesno
import numpy as np


# import Triggers_Checks_gen.Extract_Used_Triggers_From_YML_Testcases
# import Triggers_Checks_gen.Update_FDX_in_Sheet
# import Triggers_Checks_gen.Update_EVALD_signal_list_YML__from_A2L_file


class WorkflowGUI:
    """ MAIN GUI for all automation tasks"""
    def __init__(self):
        # determine the main window
        self.customer = None
        self.window = tk.Tk()
        self.window.title("Classe Generation Tool")
        self.window.geometry("1500x650")
        self.window.resizable(False, False)
        self.window.iconbitmap(self.__icon__())

        self.excel_file = os.path.abspath(
            os.path.join(os.getcwd(), r"../../../CustomerPrj/Restbus/Autosar_Gen_Database.xlsx"))
        self.excel_file_fdx = os.path.abspath(os.path.join(os.getcwd(), r"../../../CustomerPrj/FDX/FDX_Database.xlsx"))
        # self.excel_file_carmaker = os.path.abspath(
        #     os.path.join(os.getcwd(), r"../../../Platform/Carmaker/CM_Database.xlsx"))
        self.excel_file_classe = os.path.abspath(
            os.path.join(os.getcwd(), r"../../../Platform/Classe/Classe_Database.xlsx"))
        self.excel_file_classe_customer = os.path.abspath(
            os.path.join(os.getcwd(), r"../../../CustomerPrj/Classe/Customer_Database.xlsx"))
        self.excel_file_xcp = os.path.abspath(os.path.join(os.getcwd(),r"../../../CustomerPrj/XCP/XCP_Database.xlsx"))
        self.json_database = os.path.abspath(
            os.path.join(os.getcwd(), r"../../../CustomerPrj/FDX/Databases/data_dump.json"))
        self.specific_hil_json = os.path.abspath(
            os.path.join(os.getcwd(), r"../../../CustomerPrj/Classe/Database/hil_specific_triggers.json"))
        self.fdx_directory = os.path.abspath(os.path.join(os.getcwd(),r"../../../CustomerPrj/FDX/"))
        # self.cm_directory = os.path.abspath(os.path.join(os.getcwd(),r"../../../Platform/Carmaker/"))
        self.dbc_directory = os.path.abspath(os.path.join(os.getcwd(),r"../../../CustomerPrj/Restbus/Database/DBC"))
        self.main_directory = os.path.abspath(os.path.join(os.getcwd(),r"../../../"))
        self.script_names = ["Update Node Sheet DBC", "Update SysVarDatabase", "Create Nodes(.can)",
                             "Create SysVar(.xml)", "Create IniNodes(.cin)", "Create SysVarMapping(.can)",
                             "Update Nodes", "Update Databases"]
        self.script_callback = [self.update_sheet, self.update_sysvar, self.create_nodes, self.create_sysvar,
                                self.create_ininode, self.create_gw]
        self.num_steps = len(self.script_names)

        # self.__dbc_map = ((0, 0), (1, 0), (2, 3), (3, 0), (4, (2,3)), (5, 2), (6, (2,3)), (7, 4), (8, 5), (9, 6), (10, 7), (11, 0))
        self.__dbc_map = generate_dbc_map(self.excel_file, "DBCmapping")
        #self.__arxml_map = generate_dbc_map(self.excel_file, "ArxmlMapping")
        self.variant_list = extract_vehicle_variants(self.excel_file)
        self.sysvar_sheet_name = "SysVarDatabase"
        self.classe_columns = ['Name', 'Message', 'Multiplexing/Group', 'Startbit', 'Length [Bit]', 'Byte Order',
                               'Value Type', 'Initial Value', 'Factor',
                               'Offset', 'Minimum', 'Maximum', 'Unit', 'Value Table', 'Comment', 'Message ID',
                               'Cycle Time [ms]', 'texttable', 'texttable values',
                               'max_value', 'dlc', 'variant', 'sender', 'gw', 'a_variant_veh_0', 'a_variant_veh_1',
                               'b_variant_veh_0', 'b_variant_veh_1']

        self.classe_sheets = ["hil_ctrl", "hil_vm", "hil_drv", "hil_adas", "classe_fvideo", "classe_radarfc",
                              "classe_radarfr",
                              "classe_radarfl", "classe_radarrr", "classe_radarrl", "classe_fdx"]

    # add an icon if it exist
    def __icon__(self):
        """icon for gui"""
        return "classe.ico" if os.path.exists("classe.ico") else ""

    def run_subprocess(self):
        """ open debug info in notepad if error occurs will be help full if console closes"""
        sp.Popen(["notepad.exe", "Debug.log"])

    # GUI main content
    def main_content(self):
        """ main content of gui, creating frames and buttons"""
        frame_list = [None] * self.num_steps
        canvas_list = [None] * self.num_steps
        label_list = [None] * self.num_steps
        button_list = [None] * self.num_steps

        ################# section create node sheet DBC#########################
        frame_list[0] = tk.Frame(self.window, highlightbackground="grey", highlightthickness=2)
        frame_list[0].grid(column=0, row=1, padx=50, pady=2, sticky=tk.EW)
        ttk.Label(frame_list[0], text="Create Node Sheet DBC", anchor=tk.CENTER, width=25).grid(column=0, row=0,
                                                                                                padx=10, pady=5)
        self.create_node_sheet_btn = ttk.Button(frame_list[0], text="CAN",
                                                command=lambda: Thread(target=self.create_autosar).start(), width=10)
        self.create_node_sheet_btn.grid(column=0, row=1, padx=10, pady=5)

        ################# section create node sheet ARXML #########################
        frame_list[0] = tk.Frame(self.window, highlightbackground="grey", highlightthickness=2)
        frame_list[0].grid(column=0, row=3, padx=50, pady=2, sticky=tk.EW)
        ttk.Label(frame_list[0], text="Create Node Sheet ARXML", anchor=tk.CENTER, width=25).grid(column=0, row=0,
                                                                                                  padx=10, pady=5)
        self.create_node_sheet_arxml_btn = ttk.Button(frame_list[0], text="CAN",
                                                      command=lambda: Thread(target=self.create_autosar_arxml).start(),
                                                      width=10)
        self.create_node_sheet_arxml_btn.grid(column=0, row=1, padx=10, pady=9, sticky='w')

        self.create_node_sheet_arxml_eth_btn = ttk.Button(frame_list[0], text="ETH",
                                                      command=lambda: Thread(target=self.createAutosarArxmlEth).start(),
                                                      width=10)
        self.create_node_sheet_arxml_eth_btn.grid(column=0, row=1, padx=10, pady=9, sticky='e')

        # ################# section for autosar #################################
        ttk.Label(self.window, text="Update Autosar", font=("Arial Bold", 12), anchor=tk.CENTER).grid(column=0, row=0, padx=50, pady=10)
        # frame_list[0] = tk.Frame(self.window, highlightbackground="goldenrod", highlightthickness=2)
        # frame_list[0].grid(column=0, row=3, padx=50, pady=2, sticky=tk.EW)
        # ttk.Label(frame_list[0], text=self.script_names[0], anchor=tk.CENTER, width=25).grid(column=0, row=0, padx=10, pady=5)
        # self.update_node_sheet_btn = ttk.Button(frame_list[0], text="Execute", command=lambda :Thread(target=self.script_callback[0]).start(), width=10)
        # self.update_node_sheet_btn.grid(column=0, row=1, padx=10, pady=5)
        # ttk.Label(self.window, text="Please check them manually.", foreground="goldenrod", anchor=tk.CENTER).grid(column=0, row=4, padx=50)

        ############### section for Update HIL specific JSON #################
        frame_list[0] = tk.Frame(self.window, highlightbackground="grey", highlightthickness=2)
        frame_list[0].grid(column=3, row=1, padx=10, pady=2, sticky=tk.EW)
        ttk.Label(frame_list[0], text="Update HIL specific JSON", anchor=tk.CENTER, width=25).grid(column=0, row=0,
                                                                                                   padx=10, pady=5)
        self.json_btn = ttk.Button(frame_list[0], text="Execute",
                                   command=lambda: Thread(target=self.update_hil_json).start(), width=10)
        self.json_btn.grid(column=0, row=1, padx=10, pady=5)

        # section for FDX_Database Excel all triggers
        ttk.Label(self.window, text="Update FDX", font=("Arial Bold", 12), anchor=tk.CENTER).grid(column=3,
                                                                                                           row=0,
                                                                                                           padx=50,
                                                                                                           pady=0)
        frame_list[0] = tk.Frame(self.window, highlightbackground="grey", highlightthickness=2)
        frame_list[0].grid(column=3, row=3, padx=10, pady=2, sticky=tk.EW)
        ttk.Label(frame_list[0], text="Update TriggersDatabase", anchor=tk.CENTER, width=25).grid(column=0, row=0,
                                                                                                  padx=10, pady=5)
        self.update_fdx_database_btn = ttk.Button(frame_list[0], text="Execute",
                                                  command=lambda: Thread(target=self.update_fdx_database_excel).start(),
                                                  width=10)
        self.update_fdx_database_btn.grid(column=0, row=1, padx=10, pady=5)

        # section for FDX_Database Excel only test triggers

        # frame_list[0] = tk.Frame(self.window, highlightbackground="grey", highlightthickness=1)
        # frame_list[0].grid(column=1, row=3, padx=50, pady=0, sticky=tk.EW)
        #
        # ttk.Label(frame_list[0], text="Update TriggersTestCatalog", anchor=tk.CENTER, width=25).grid(column=0, row=0, padx=10, pady=5)
        # ttk.Button(frame_list[0], text="Execute", command=self.update_test_triggers, width=10).grid(column=0, row=1, padx=10, pady=5)
        #
        # # section for FDX_in update
        #
        # frame_list[0] = tk.Frame(self.window, highlightbackground="grey", highlightthickness=1)
        # frame_list[0].grid(column=1, row=5, padx=50, pady=0, sticky=tk.EW)
        #
        # ttk.Label(frame_list[0], text="Update FDX_in sheet", anchor=tk.CENTER, width=25).grid(column=0, row=0, padx=10, pady=5)
        # ttk.Button(frame_list[0], text="Execute", command=self.update_fdx_in_sheet, width=10).grid(column=0, row=1, padx=10, pady=5)
        #
        # # section for Update xcp_checksdatabase sheet
        #
        # frame_list[0] = tk.Frame(self.window, highlightbackground="grey", highlightthickness=1)
        # frame_list[0].grid(column=1, row=7, padx=50, pady=0, sticky=tk.EW)
        #
        # ttk.Label(frame_list[0], text="Update Checks Database", anchor=tk.CENTER, width=25).grid(column=0, row=0, padx=10, pady=5)
        # ttk.Button(frame_list[0], text="Execute", command=self.update_xcpcheckdatabase, width=10).grid(column=0, row=1, padx=10, pady=5)


        ########### section for Multicanoe #######################
        ttk.Label(self.window, text="Update Multicanoe", font=("Arial Bold", 12), anchor=tk.CENTER).grid(column=4, row=5,
                                                                                                  columnspan=3,
                                                                                                  padx=10, pady=10,sticky="S")
        frame_list[0] = tk.Frame(self.window, highlightbackground="grey", highlightthickness=2)
        frame_list[0].grid(column=4, row=6, padx=60, pady=2, sticky=tk.N, rowspan=4)

        ttk.Label(frame_list[0], text="Update Multicanoe .cfg", anchor=tk.CENTER, width=25).grid(column=0, row=0, padx=10,
                                                                                    pady=5)
        self.upadte_multicanoe_btn = ttk.Button(frame_list[0], text="Execute",
                                         command=lambda: Thread(target=self.createMulticanoe).start(), width=10)
        self.upadte_multicanoe_btn.grid(column=0, row=1, padx=10, pady=5)
        self.create_maincfg_3_state = tk.IntVar()
        self.create_maincfg_3_checkbox = tk.Checkbutton(frame_list[0], text='Main_Cfg_pf3', variable=self.create_maincfg_3_state, onvalue=1, offvalue=0,
                                            anchor="w")
        self.create_maincfg_3_checkbox.grid(column=0, row=2, padx=10, sticky="WE")

        self.create_master_slave_3_state = tk.IntVar()
        self.create_master_slave_checkbox = tk.Checkbutton(frame_list[0], text='Master_Slave_pf3', variable=self.create_master_slave_3_state,
                                                 onvalue=1, offvalue=0, anchor="w")
        self.create_master_slave_checkbox.grid(column=0, row=4, padx=10, sticky="WE")

        self.create_master_slave_2_state = tk.IntVar()
        self.create_master_slave_2_checkbox = tk.Checkbutton(frame_list[0], text='Master_Salve_pf2', variable=self.create_master_slave_2_state, onvalue=1,
                                              offvalue=0, anchor="w")
        self.create_master_slave_2_checkbox.grid(column=0, row=5, padx=10, sticky="WE")



        ########### section for Update xcpconfig #######################
        ttk.Label(self.window, text="Update XCP", font=("Arial Bold", 12), anchor=tk.CENTER).grid(column=4, row=0,
                                                                                                      columnspan=3,
                                                                                                      padx=10, pady=10)
        frame_list[0] = tk.Frame(self.window, highlightbackground="grey", highlightthickness=2)
        frame_list[0].grid(column=4, row=1, padx=60, pady=2, sticky=tk.EW, rowspan=4)

        ttk.Label(frame_list[0], text="Update xcpconfig", anchor=tk.CENTER, width=25).grid(column=0, row=0, padx=10,
                                                                                           pady=5)
        self.update_xcp_btn = ttk.Button(frame_list[0], text="Execute",
                                         command=lambda: Thread(target=self.update_xcpconfig).start(), width=10)
        self.update_xcp_btn.grid(column=0, row=1, padx=10, pady=5)

        self.customer = (self.get_customer(self.main_directory)).upper()
        if self.customer == "FORD":
            self.dasy_state = tk.IntVar()
            self.dasy_checkbox = tk.Checkbutton(frame_list[0], text='DASy', variable=self.dasy_state, onvalue=1, offvalue=0, anchor="w")
            self.dasy_checkbox.grid(column=0, row=2, padx=10, sticky="WE")

            self.front_obj_state = tk.IntVar()
            self.front_obj_checkbox = tk.Checkbutton(frame_list[0], text='RadarFC', variable=self.front_obj_state, onvalue=1, offvalue=0, anchor="w")
            self.front_obj_checkbox.grid(column=0, row=4, padx=10, sticky="WE")

            self.corner_state = tk.IntVar()
            self.corner_checkbox = tk.Checkbutton(frame_list[0], text='Radar Corner', variable=self.corner_state, onvalue=1, offvalue=0, anchor="w")
            self.corner_checkbox.grid(column=0, row=5, padx=10, sticky="WE")
        else:
            self.dasy_state_delta1 = tk.IntVar()
            self.dasy_checkbox = tk.Checkbutton(frame_list[0], text='DASy Delta1', variable=self.dasy_state_delta1, onvalue=1,
                                                offvalue=0, anchor="w")
            self.dasy_checkbox.grid(column=0, row=2, padx=10, sticky="WE")

            self.dasy_state_delta5 = tk.IntVar()
            self.dasy_checkbox = tk.Checkbutton(frame_list[0], text='DASy Delta5', variable=self.dasy_state_delta5, onvalue=1,
                                                offvalue=0, anchor="w")
            self.dasy_checkbox.grid(column=0, row=3, padx=10, sticky="WE")

            self.front_loc_state = tk.IntVar()
            self.front_loc_checkbox = tk.Checkbutton(frame_list[0], text='RadarFC Locations',
                                                     variable=self.front_loc_state, onvalue=1, offvalue=0, anchor="w")
            self.front_loc_checkbox.grid(column=0, row=4, padx=10, sticky="WE")

            self.front_obj_state = tk.IntVar()
            self.front_obj_checkbox = tk.Checkbutton(frame_list[0], text='RadarFC Objects',
                                                     variable=self.front_obj_state, onvalue=1, offvalue=0, anchor="w")
            self.front_obj_checkbox.grid(column=0, row=5, padx=10, sticky="WE")

            self.corner_state = tk.IntVar()
            self.corner_checkbox = tk.Checkbutton(frame_list[0], text='Radar Corner', variable=self.corner_state,
                                                  onvalue=1, offvalue=0, anchor="w")
            self.corner_checkbox.grid(column=0, row=6, padx=10, sticky="WE")



        ######################Delivery tasks automation section###############
        ttk.Label(self.window, text="Delivery Automation", font=("Arial Bold", 12), anchor=tk.CENTER).grid(column=7, row=0,
                                                                                                      columnspan=3,
                                                                                                      padx=10, pady=10)
        frame_list[0] = tk.Frame(self.window, highlightbackground="grey", highlightthickness=2)
        frame_list[0].grid(column=7, row=1, padx=30, pady=2, sticky=tk.N, rowspan=4)

        self.update_copyrights_state = tk.IntVar()
        self.update_copyrights_state_checkbox = tk.Checkbutton(frame_list[0], text='Update Copyrights', variable=self.update_copyrights_state, onvalue=1, offvalue=0, anchor="w")
        self.update_copyrights_state_checkbox.grid(column=0, row=2, padx=10, pady=(10,0),sticky="WE")

        self.update_doxygen_state = tk.IntVar()
        self.update_doxygen_checkbox = tk.Checkbutton(frame_list[0], text='Update Doxygen', variable=self.update_doxygen_state, onvalue=1, offvalue=0, anchor="w")
        self.update_doxygen_checkbox.grid(column=0, row=4, padx=10, sticky="WE")

        #handled in new GUI
        # self.update_releasenote_state = tk.IntVar()
        # self.update_releasenote_checkbox = tk.Checkbutton(frame_list[0], text='Update Release Note', variable=self.update_releasenote_state, onvalue=1, offvalue=0, anchor="w")
        # self.update_releasenote_checkbox.grid(column=0, row=6, padx=10, sticky="WE")

        self.zip_files_state = tk.IntVar()
        self.zip_files_checkbox = tk.Checkbutton(frame_list[0], text='ZIP/Encrypt Files',variable=self.zip_files_state, onvalue=1, offvalue=0,anchor="w")
        self.zip_files_checkbox.grid(column=0, row=6, padx=10, sticky="WE")

        self.offline_docu_state = tk.IntVar()
        self.offline_docu_checkbox = tk.Checkbutton(frame_list[0], text='Offline Analyzer Docu', variable=self.offline_docu_state,
                                                 onvalue=1, offvalue=0, anchor="w")
        self.offline_docu_checkbox.grid(column=0, row=8, padx=10, sticky="WE")

        self.delivery_auto_btn = ttk.Button(frame_list[0], text="Execute",command=lambda: Thread(target=self.deliveryAutomation).start(), width=10)
        self.delivery_auto_btn.grid(column=0, row=10, padx=10, pady=15)

        #release note gui button
        # self.relese_note_gui_btn = ttk.Button(self.window, text="Auto Gen Release Note",
        #                                         command=lambda: Thread(target=self.releaseNoteAutoGen).start(), width=25)
        # self.relese_note_gui_btn.grid(column=7, row=6, padx=60, pady=5, ipady=9)






        ################## section for restbus ########################
        ttk.Label(self.window, text="Update Restbus", font=("Arial Bold", 12), anchor=tk.CENTER).grid(column=1, row=0,
                                                                                                      columnspan=1,
                                                                                                      padx=10, pady=10)

        self.buttons = {}
        for i, pos in enumerate(range(1, (self.num_steps - 3) * 2, 2), 1):
            frame_list[i] = tk.Frame(self.window, highlightbackground="slate grey", highlightthickness=2)
            frame_list[i].grid(column=1, row=pos, padx=10, pady=2, sticky=tk.EW)

            if i < self.num_steps - 3:
                canvas_list[i] = tk.Canvas(self.window, height=20, width=30)
                canvas_list[i].grid(column=1, row=pos + 1, sticky=tk.EW)
                canvas_list[i].create_line(100, 0, 100, 20, arrow=tk.LAST)
                # canvas_list[i].create_line(100, 0, 100, 20, fill="slate grey", arrow=tk.LAST)

        ttk.Label(frame_list[1], text=self.script_names[1], anchor=tk.CENTER, width=25).grid(column=0, row=0, padx=10,
                                                                                             pady=5)
        self.buttons[f"button_1"] = ttk.Button(frame_list[1], text="Execute",
                                               command=lambda: Thread(target=self.script_callback[1]).start(), width=10)
        self.buttons[f"button_1"].grid(column=0, row=1, padx=10, pady=5)
        ttk.Label(frame_list[2], text=self.script_names[2], anchor=tk.CENTER, width=25).grid(column=0, row=0, padx=10,
                                                                                             pady=5)
        self.buttons[f"button_2"] = ttk.Button(frame_list[2], text="CAN",
                                               command=lambda: Thread(target=self.script_callback[2]).start(), width=10)
        self.buttons[f"button_2"].grid(column=0, row=1, padx=10, pady=9, sticky='w')
        self.buttons[f"button_2_eth"] = ttk.Button(frame_list[2], text="ETH",
                                               command=lambda: Thread(target=self.createNodesETH).start(), width=10)
        self.buttons[f"button_2_eth"].grid(column=0, row=1, padx=10, pady=9, sticky='e')

        ttk.Label(frame_list[3], text=self.script_names[3], anchor=tk.CENTER, width=25).grid(column=0, row=0, padx=10,
                                                                                             pady=5)
        self.buttons[f"button_3"] = ttk.Button(frame_list[3], text="Execute",
                                               command=lambda: Thread(target=self.script_callback[3]).start(), width=10)
        self.buttons[f"button_3"].grid(column=0, row=1, padx=10, pady=5)
        ttk.Label(frame_list[4], text=self.script_names[4], anchor=tk.CENTER, width=25).grid(column=0, row=0, padx=10,
                                                                                             pady=5)
        self.buttons[f"button_4"] = ttk.Button(frame_list[4], text="Execute",
                                               command=lambda: Thread(target=self.script_callback[4]).start(), width=10)
        self.buttons[f"button_4"].grid(column=0, row=1, padx=10, pady=5)

        #ttk.Label(frame_list[5], text=self.script_names[5], anchor=tk.CENTER, width=25).grid(column=0, row=0, padx=10,
        #                                                                                     pady=5)
        #self.buttons[f"button_5"] = ttk.Button(frame_list[5], text="Execute",
        #                                       command=lambda: Thread(target=self.script_callback[5]).start(), width=10)
        #self.buttons[f"button_5"].grid(column=0, row=1, padx=10, pady=5)

        ttk.Label(frame_list[5], text=self.script_names[5], anchor=tk.CENTER, width=25).grid(column=0, row=0, padx=10,
                                                                                             pady=5)
        self.buttons[f"button_5"] = ttk.Button(frame_list[5], text="Execute",
                                               command=lambda: Thread(target=self.script_callback[5]).start(), width=10)
        self.buttons[f"button_5"].grid(column=0, row=1, padx=10, pady=5)

        frame_list[0] = tk.Frame(self.window, highlightbackground="grey", highlightthickness=2)
        frame_list[0].grid(column=2, row=1, padx=60, pady=2, sticky=tk.EW)
        ttk.Label(frame_list[0], text="Test Database", anchor=tk.CENTER, width=22).grid(column=0, row=0,padx=10, pady=5)
        self.create_test_database_btn = ttk.Button(frame_list[0], text="Execute", command=lambda: Thread(target=self.test_database).start(), width=10)
        self.create_test_database_btn.grid(column=0, row=1, padx=10, pady=5)

        self.sanity_check_btn = ttk.Button(self.window, text="ARXML Sanity Check",
                                          command=lambda: Thread(target=self.arxml_sanity_check).start(), width=25)
        self.sanity_check_btn.grid(column=2, row=3, padx=60, pady=5, ipady=15)

        self.execute_all_btn = ttk.Button(self.window, text="Execute RBS",
                                          command=lambda: Thread(target=self.update_restbus).start(), width=25)
        self.execute_all_btn.grid(column=2, row=5, padx=60, pady=5, ipady=15)
        #
        # #copyrights button
        # self.copyrights_update_btn = ttk.Button(self.window, text="Update Copyrights\n     *.(py,can,cin)",
        #                                   command=lambda: Thread(target=self.updateCopyRights).start(), width=25)
        # self.copyrights_update_btn.grid(column=2, row=5, padx=60, pady=5, ipady=9)
        #
        # #Doxygen_button
        # self.doxygen_update_btn = ttk.Button(self.window, text="Update Doxygen Doc",
        #                                         command=lambda: Thread(target=self.updateDoxygenDocumentation).start(), width=25)
        # self.doxygen_update_btn.grid(column=2, row=7, padx=60, pady=5, ipady=15)


        ################# section for fdx ############################
        # ttk.Label(self.window, text="Update FDX", font=("Arial Bold", 12), anchor=tk.CENTER).grid(column=4, row=0, padx=20, pady=10)

        frame_list[self.num_steps - 2] = tk.Frame(self.window, highlightbackground="slate grey", highlightthickness=2)
        frame_list[self.num_steps - 2].grid(column=3, row=5, padx=10, pady=2, sticky=tk.EW)

        label_list[self.num_steps - 2] = ttk.Label(frame_list[self.num_steps - 2],
                                                   text=self.script_names[self.num_steps - 2], anchor=tk.CENTER,
                                                   width=25)
        label_list[self.num_steps - 2].grid(column=0, row=0, padx=10, pady=2, sticky=tk.EW)
        self.update_nodes_btn = ttk.Button(frame_list[self.num_steps - 2], text="Execute",
                                           command=lambda: Thread(target=self.update_fdx).start(), width=10)
        self.update_nodes_btn.grid(column=0, row=1, padx=10, pady=5)

        frame_list[self.num_steps - 1] = tk.Frame(self.window, highlightbackground="slate grey", highlightthickness=2)
        frame_list[self.num_steps - 1].grid(column=3, row=7, padx=10, pady=2, sticky=tk.EW)

        label_list[self.num_steps - 1] = ttk.Label(frame_list[self.num_steps - 1],
                                                   text=self.script_names[self.num_steps - 1], anchor=tk.CENTER,
                                                   width=25)
        label_list[self.num_steps - 1].grid(column=0, row=1, padx=10, pady=2, sticky=tk.EW)
        self.update_databases_btn = ttk.Button(frame_list[self.num_steps - 1], text="Execute",
                                               command=lambda: Thread(target=self.update_database).start(),
                                               width=10)
        self.update_databases_btn.grid(column=0, row=2, padx=10, pady=5)

        ############### section for Update CarMaker nodes #################
        # ttk.Label(self.window, text="Update CarMaker", font=("Arial Bold", 12), anchor=tk.CENTER).grid(column=6, row=0,
        #                                                                                                padx=20, pady=10)
        # frame_list[0] = tk.Frame(self.window, highlightbackground="grey", highlightthickness=2)
        # frame_list[0].grid(column=6, row=1, padx=20, pady=2, sticky=tk.EW)
        # ttk.Label(frame_list[0], text="Update CM Nodes", anchor=tk.CENTER, width=25).grid(column=0, row=0,
        #                                                                                   padx=10, pady=5)
        # self.cm_btn = ttk.Button(frame_list[0], text="Execute",
        #                          command=lambda: Thread(target=self.update_carmaker).start(), width=10)
        # self.cm_btn.grid(column=0, row=1, padx=10, pady=5)

    def get_customer(self, path):
        """
        Get the customer name from the cfg name

        Args:
          path: path to main git folder

        Returns:
            str: customer name
        """
        for (dirpath, dirnames, filenames) in os.walk(path):
            for file in filenames:
                if file.endswith(".cfg"):
                    return file.split(".")[0].split("_")[1]

    def create_autosar(self):
        """ Extract data from DBC files and create node sheets"""
        logger.info("###### START 'Create Autosar' DEBUG INFORMATION ######")
        self.disable_bnts()
        try:
            update_autosar_main(self.__dbc_map)
            logger.info("###### END 'Create Autosar' DEBUG INFORMATION ######")
            logger.info('-' * 80)
            messagebox.showinfo(title="Create Autosar", message="Successfully created Autosar from DBC database.")
            self.enable_btns()
        except Exception as exp:
            logger.error(f"Failed to create autosar: {exp}")
            messagebox.showerror(title="Error", message="Failed to create autosar.\nPlease check the Console output.")
            self.run_subprocess()

    def create_autosar_arxml(self):
        """ Extract CAN data from ARXML files and create node sheets"""
        logger.info("###### START 'Create Autosar ARXML' DEBUG INFORMATION ######")
        self.disable_bnts()
        try:
            result = create_arxml_can('CAN')
            if not(result):
                raise Exception("Error occured, check above errors")
            logger.info("###### END 'Create Autosar ARXML CAN' DEBUG INFORMATION ######")
            logger.info('-' * 80)
            messagebox.showinfo(title="Create Autosar ARXML CAN",
                                message="Successfully created Autosar from ARXML CAN database.")
            self.enable_btns()
        except Exception as exp:
            logger.error(f"Failed to extract ARXML CAN{exp}")
            messagebox.showerror(title="Error", message="Failed to create autosar.\nPlease check the Console output.")
            self.enable_btns()
            self.run_subprocess()

    def createAutosarArxmlEth(self):
        """ extract eth data from arxml and create node sheets """
        logger.info("########## START 'Create Node Sheets ARXML ETH' DEBUG INFORMATION ##############")
        self.disable_bnts()
        try:
            result = create_arxml_eth()
            if not(result):
                raise Exception("Error occured, check above errors")
            logger.info("###### END 'Create Autosar ARXML ETH' DEBUG INFORMATION ######")
            logger.info('-' * 80)
            messagebox.showinfo(title="Create Autosar ARXML ETH",
                                message="Successfully created Autosar from ARXML ETH database.")
            self.enable_btns()
        except Exception as exp:
            logger.error(f"Failed to extract arxml eth: {exp}")
            messagebox.showerror(title="Error", message="Failed to extract arxml eth.\nPlease check the Console output.")
            self.enable_btns()
            self.run_subprocess()


    # call update_autosar
    # NOTE : Function update_sheet shall be removed in future as we merged in create autosar
    def update_sheet(self):
        """ """
        logger.info(f"###### START 'Update Node Sheet' DEBUG INFORMATION ######")
        self.disable_bnts()
        messagebox.showwarning(title="Warning",
                               message="Please make sure that the DBC file names and the order of sheets in Excel are not modified.")
        dbc_list = find_DBC_file("../../../CustomerPrj/Restbus/Database/DBC", select=False)

        wb = load_workbook(self.excel_file)
        sheet_list = wb.sheetnames

        counter = 0
        for sheet_index, dbc_index in self.__dbc_map:
            try:
                if type(dbc_index) is int:
                    logger.info(f"DBC index is {dbc_index}")
                    logger.info(f"DBC is {dbc_list[dbc_index]}")
                    parsed_tuple = parse_DBC_file(dbc_list[dbc_index])

                elif type(dbc_index) is tuple:
                    dbcs = []
                    for dbc in dbc_index:
                        logger.info(f"DBC index is {dbc}")
                        logger.info(f"DBC is {dbc_list[dbc]}")
                        dbcs.append(dbc_list[dbc])
                    parsed_tuple = parse_DBC_file_multiple_dbcs(dbcs)

                else:
                    parsed_tuple = 0
                    logger.error("Format of DBCs files to be parsed is not valid")

                if not parsed_tuple:
                    logger.error(f"Failed to update sheet: {sheet_list[sheet_index]}")
                    logger.error(f"The selected DBC file is not correct: {dbc_list[dbc_index[0]]}")
                    continue
                new_wb = update_sheet(wb, sheet_index, parsed_tuple)
                if new_wb is not False:
                    wb = new_wb
                    counter += 1
                else:
                    logger.error(f"Failed to update sheet: {sheet_list[sheet_index]}")


            except Exception as exp:
                logger.error(f"Failed to update sheet: {sheet_list[sheet_index]} {exp}")
                self.run_subprocess()
        wb.save(self.excel_file)
        save_excel(self.excel_file)

        # Code to organize data of all worksheets alphabetically
        # logger.info(
        #    f"###### Organising the worksheets data alphabetically in progress - This may take longer time ######")
        # book = load_workbook(self.excel_file)
        # writer = pd.ExcelWriter(self.excel_file, engine='openpyxl')
        # writer.book = book
        # writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        # sheet_names = wb.sheetnames
        # for sheet_index, dbc_index in self.__dbc_map:
        #    original_signals = pd.read_excel(self.excel_file, sheet_name=sheet_index)
        #    sorted_signals = original_signals.sort_values(by=['Message', 'Name'], ignore_index=False)
        #    sorted_signals.to_excel(writer, sheet_name=sheet_names[sheet_index], index=False)
        # writer.save()

        messagebox.showinfo(title="Update Sheets", message="{0} of {1} node sheets were updated.".format(str(counter),
                                                                                                         str(len(
                                                                                                             self.__dbc_map))))
        logger.info(f"###### END 'Update Node Sheet' DEBUG INFORMATION ######")
        logger.info('-' * 80)
        self.enable_btns()
        return counter == len(self.__dbc_map)

    # call update_sysvartab
    def update_sysvar(self):
        """ merges all nodes into single sysvardatabase sheet"""
        logger.info(f"###### START 'Update SysVarDatabase' DEBUG INFORMATION ######")
        self.disable_bnts()
        wb_r = load_workbook(self.excel_file, data_only=True)
        # workbook for writing
        
        dict_sysvar_sheets = {"CAN": "SysVarDatabase", "ETH": "SysVarDatabaseETH"}

        try:
            dict_arrays = generate_database(wb_r, self.variant_list)
            database_array = dict_arrays["CAN"]
            database_array = update_gw(database_array)
            if self.variant_list:
                database_array = update_variants(wb_r, dict_sysvar_sheets["CAN"], database_array, self.variant_list)
            dict_arrays["CAN"] = database_array
            #process ETH sheet by filling variants
            eth_database_array = dict_arrays["ETH"]
            if self.variant_list:
                eth_database_array = update_variants_eth(wb_r, dict_sysvar_sheets["ETH"], eth_database_array, self.variant_list)
            dict_arrays["ETH"] = eth_database_array
            wb_r.close()
            wb_w = load_workbook(self.excel_file)
            save_database(wb_w, dict_sysvar_sheets, dict_arrays, self.excel_file)
            #check_sysvardatabase_integrity(self.excel_file)
        except Exception as exp:
            logger.error("Failed to update SysVarDatabase:", exp)
            logger.info("###### END 'Update SysVarDatabase' DEBUG INFORMATION ######")
            messagebox.showerror(title="Error",
                                 message="SysVarDatabase sheet was NOT updated.\nPlease check the Console output.")
            self.run_subprocess()
            return False
        messagebox.showinfo(title="Update SysVarDatabase", message="Successfully updated SysVarDatabase sheet.")
        self.enable_btns()
        return True

    # call test_database
    def test_database(self):
        """ tests the sysvardatabase sheet for data integrity """
        logger.info(f"###### START 'Test Database' DEBUG INFORMATION ######")
        self.disable_bnts()
        sysvar_sheet = "SysVarDatabase"
        df_sys = pd.read_excel(self.excel_file, sheet_name=sysvar_sheet, dtype=object)
        try:
            test_sysvar_can_eth()
            logger.info("###### END 'Test Database' DEBUG INFORMATION ######")
            logger.info('-' * 80)
        except Exception as exp:
            logger.error("Failed to Test Database:", exp)
            messagebox.showerror(title="Error", message="Test_Database.xlsx was NOT updated.\nPlease check the Console output.")
            self.run_subprocess()
            return False
        messagebox.showinfo(title="Test Database", message="Successfully updated Test_Database.xlsx.")
        self.enable_btns()
        return True

    # call create_nodes
    def create_nodes(self):
        """ creates capl files for each node sheet"""
        logger.info(f"###### START 'Create Nodes(.can)' DEBUG INFORMATION ######")
        self.disable_bnts()
        workbook = load_workbook(self.excel_file, data_only=True)
        initialize_dict(workbook)
        generate_sysvardatabase_dataframe(workbook,"SysVarDatabase")
        # get node sheets
        sheet_list = [name for name in workbook.sheetnames if len(name.split('_')) > 2 and ("can_" in name.lower())]

        counter = 0
        for sheet in sheet_list:
            try:
                create_node_main(workbook, sheet,r'../../../CustomerPrj/Restbus/Nodes')
                counter += 1
            except Exception as exp:
                logger.error(f"'Create Nodes(.can)' execution failed: {exp}")
                messagebox.showerror(title="Error", message="Failed to create nodes.\nPlease check the Console output.")
                self.run_subprocess()
                return False
        logger.info("###### END 'Create Nodes(.can)' DEBUG INFORMATION ######")
        logger.info('-' * 80)
        messagebox.showinfo(title="Create Nodes",
                            message="{0} node flies were created from {1} node sheets.".format(str(counter),
                                                                                               str(len(sheet_list))))
        self.enable_btns()
        return counter == len(sheet_list)

    def createNodesETH(self):
        """create .can files for ethernet"""
        logger.info(f"###### START 'Create Nodes(.can)' DEBUG INFORMATION ######")
        self.disable_bnts()
        try:
            createNodesEthernetMain()
        except Exception as exp:
            traceback.print_exc()
            logger.error(f"'Create Nodes(.can)' execution failed: {exp}")
            messagebox.showerror(title="Error", message="Failed to create nodes.\nPlease check the Console output.")
            self.enable_btns()
            self.run_subprocess()
            return False
        logger.info("###### END 'Create Nodes(.can)' DEBUG INFORMATION ######")
        logger.info('-' * 80)
        messagebox.showinfo(title="Create Nodes ETH", message="Successfully executed")
        self.enable_btns()
        return


    # call create_sysvar
    def create_sysvar(self):
        """ updates the xml file for system variables"""
        logger.info(f"###### START 'Create SysVar(.xml)' DEBUG INFORMATION ######")
        self.disable_bnts()
        wb = load_workbook(self.excel_file, data_only=True)
        try:
            sysvarSheet = "SysVarDatabase"
            eth_sysvar_sheet = "SysVarDatabaseETH"
            #RBS
            try:
                xml = create_rb_sysvar(wb, sysvarSheet, self.variant_list)
                eth_xml = create_rb_sysvar_ethernet(wb, eth_sysvar_sheet, self.variant_list)
                xml.extend(eth_xml)
                xml.extend(rbs_sysvar_footer())
                save_to_file(xml,r"../../../CustomerPrj/Restbus/Database", self.sysvar_sheet_name)
            except Exception as rbs_exp:
                logger.error(f"Error occured while generating RBS xml file, new columns might be added, please run from bigining and try else -> {rbs_exp}")
                logger.info(f"-------------------------------------Proceeding with CLASSE xml generation---------------------------------------------")


            #CLASSE
            databases = {"Classe Platform": [self.excel_file_classe,r"../../Classe/Database",r"ClasseDatabase"],
                         "Classe Customer": [self.excel_file_classe_customer,r"../../../CustomerPrj/Classe/Database",r"CustomerDatabase"]}

            for tag, database in databases.items():
                logger.info(f"+++++++++++Start {tag} xml generation++++++++++++")

                #  Update car variants in CustomerDatabase.xlsx
                if database[2] == "CustomerDatabase":
                    # Read Excel file sheet by sheet
                    excel_data = pd.read_excel(database[0], sheet_name=None)

                    # Access each DataFrame corresponding to a sheet
                    for sheet_name, df in excel_data.items():
                        # change the enumeration of vehicle variable
                        if sheet_name == "hil_ctrl":
                            filtered_index = df.index[df["Name"] == "vehicle"].tolist()
                            vehicle_texttable = ""
                            for i, vehicle in enumerate(self.variant_list):
                                vehicle_texttable = vehicle_texttable + f"LogicalValue: {i} " + vehicle + "\n"
                            df.at[filtered_index[0], "texttable values"] = vehicle_texttable

                        # Determine the index of column "gw"
                        column_gw_index = df.columns.get_loc('gw')
                        # Rename columns based on DBCMapping variant list
                        for index, new_column in enumerate(self.variant_list):
                            if len(df.columns) > column_gw_index + index + 1:
                                df.columns.values[column_gw_index + index + 1] = new_column
                            else:
                                df.insert(column_gw_index + index + 1, new_column, None)

                        # Determine the number of columns to keep based on the length of variant list
                        num_columns_to_keep = column_gw_index + len(self.variant_list) + 1

                        # Slice the DataFrame to keep only the required number of columns
                        df = df.iloc[:, :num_columns_to_keep]

                        # Write back to the Excel file
                        with pd.ExcelWriter(database[0], engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                            df.to_excel(writer, sheet_name=sheet_name, index=False)

                df = pd.concat(pd.read_excel(database[0], sheet_name=None), ignore_index=True)
                df.loc[-1] = list(df.columns)  # adding a row
                df.index = df.index + 1  # shifting index
                df = df.sort_index()
                classe_df = df.replace(np.nan, '', regex=True)
                define_classe_colIndex(list(classe_df.columns))
                classe_df_with_indexes = classe_df.rename(
                    columns={x: y for x, y in zip(classe_df.columns, range(0, len(classe_df.columns)))})
                xml = generate_sysvar_header()
                xml.extend(generate_sysvar_classe(classe_df_with_indexes))
                xml.extend(generate_sysvar_footer_classe())
                save_to_file(xml, database[1], database[2])

            logger.info(f"###### END 'Create SysVar(.xml)' DEBUG INFORMATION ######")
            logger.info('-' * 80)
        except Exception as exp:
            logger.error(f"{'Create SysVar(.xml)'} execution failed: {exp}")
            traceback.print_exc()
            messagebox.showerror(title="Error", message="Failed to create System Variables.\nPlease check the Console output.")
            self.run_subprocess()
            return False
        messagebox.showinfo(title="Create SysVar", message='''Successfully created System Variables
                                                              Autosar_Gen_Database.xlsx --> SysVarDatabase_gen.xml 
        Classe_Database.xlsx --> ClasseDatabase_gen.xml
        Customer_Database.xlsx --> CustomerDatabase_gen.xml''')
        self.enable_btns()
        return True

    # call create_ininode
    def create_ininode(self):
        """ craete/upadte .cin files for initialization"""
        logger.info("###### START 'Create IniNodes(.can)' DEBUG INFORMATION ######")
        self.disable_bnts()
        wb = load_workbook(self.excel_file, data_only=True)
        wb_classe = load_workbook(self.excel_file_classe, data_only=True)
        wb_classe_customer = load_workbook(self.excel_file_classe_customer, data_only=True)

        hit_ctrl_mapping_dict = extractHilCtrl(wb)
        dict_multicanoe = extractMultiCanoeMapping(wb,skip_empty=True)

        sysvarSheet = "SysVarDatabase"
        can_variants_present = False
        eth_variants_present = False

        try:
            can_sysvarSheet = "SysVarDatabase"
            ini_df_main = generate_dataframe_rbs(wb, can_sysvarSheet, self.variant_list)
            flag_opaque = True
            for variant in self.variant_list:
                ini_node = generate_ininode_senderwise(ini_df_main, variant)
                if flag_opaque:
                    ini_node = generate_variables(ini_node, flag_opaque)
                    ini_node = generate_Opaque_array_initialize(ini_node)
                    flag_opaque = False
                else:
                    ini_node = generate_variables(ini_node, flag_opaque)
                create_file(ini_node,r'../../../CustomerPrj/Classe/Nodes/Initialization/init_rbs_can_'+variant+'.cin')
                can_variants_present = True

            eth_sysvarSheet = "SysVarDatabaseETH"
            ini_df_main_eth = generate_dataframe_rbs_ethernet(wb, eth_sysvarSheet, self.variant_list)
            if ini_df_main_eth is not None:
                flag_opaque = True
                for variant in self.variant_list:
                    ini_node = generate_ininode_senderwise_ethernet(ini_df_main_eth, variant)
                    if flag_opaque:
                        #ini_node = generate_variables_ethernet(ini_node, flag_opaque)
                        #ini_node = generate_Opaque_array_initialize_ethernet(ini_node)
                        flag_opaque = False
                    ini_node = generate_variables_ethernet(ini_node, flag_opaque)
                    create_file(ini_node, r'../../../CustomerPrj/Classe/Nodes/Initialization/init_rbs_eth_' + variant + '.cin')
                    eth_variants_present = True
            
            #update initialization.can file
            update_initialization_file('../../../CustomerPrj/Classe/Nodes/Initialization/Initialization.can',
                                        self.variant_list, hit_ctrl_mapping_dict, can_variants_present, eth_variants_present)

            # if multi canoe is configured
            if dict_multicanoe:
                logger.info("MultiCanoe configured in mapping sheets, Starting .cin generation ...")
                flag_opaque = True
                for variant in self.variant_list:
                    for multicanoe_category, node_list in dict_multicanoe.items():
                        ini_node = generate_ininode_multicanoe_senderwise(ini_df_main, variant, multicanoe_category, node_list)
                        if flag_opaque:
                            ini_node = generate_variables(ini_node, flag_opaque)
                            ini_node = generate_Opaque_array_initialize(ini_node)
                            flag_opaque = False
                        else:
                            ini_node = generate_variables(ini_node, flag_opaque)
                        create_file(ini_node, r'../../../CustomerPrj/Classe/Nodes/Initialization/init_rbs_can_' + variant + '_' + multicanoe_category + '.cin')
                        logger.info(f"Generated .cin for {variant}  ->  {multicanoe_category}")

                for multicanoe_category, node_list in dict_multicanoe.items():
                    update_initialization_file('../../../CustomerPrj/Classe/Nodes/Initialization/Initialization_'+ multicanoe_category+'.can',
                        self.variant_list, hit_ctrl_mapping_dict, can_variants_present, eth_variants_present, multicanoe_category = multicanoe_category, multicanoe_node_list=node_list)
                    logger.info(f"Updated Initialization_{multicanoe_category}.can for multicanoe")  
            #for classe
            ini_df = generate_dataframe_classe(wb_classe, "classe")
            ini_node = generate_ininode_classe(ini_df, "init_value")
            append_file(ini_node,r'../../../Platform/Classe/Nodes/Initialization/classe_init_value.cin')
            logger.info("###### END 'Create IniNodes(.can)' DEBUG INFORMATION ######")

            ini_df = generate_dataframe_classe(wb_classe_customer, "customer")
            for variant in self.variant_list:
                ini_node = generate_ininode_classe(ini_df, variant)
                append_file(ini_node,r'../../../CustomerPrj/Classe/Nodes/Initialization/customer_init_' + variant +r'.cin')
            logger.info("###### END 'Create IniNodes(.can)' DEBUG INFORMATION ######")
            logger.info('-' * 80)
        except Exception as exp:
            traceback.print_exc()
            logger.error(f"Failed to created ininode: {exp}")
            messagebox.showerror(title="Error",
                                 message="Failed to create IniNode from SysVarDatabase & ClasseDatabase sheets.\nPlease check the Console output.")
            self.run_subprocess()
            return False
        messagebox.showinfo(title="Create IniNode",
                            message="Successfully created IniNode from SysVarDatabase & ClasseDatabase sheets.")
        self.enable_btns()
        return True

    # call update_ini
    #update_ini shall be removed from gui in future as ini files are removed from initialization.can 
    def update_ini(self):
        """ """
        logger.info("###### START 'Update Ini(.ini)' DEBUG INFORMATION ######")
        self.disable_bnts()
        wb = load_workbook(self.excel_file)
        wb_classe = load_workbook(self.excel_file_classe, data_only=True)
        inisheet = "ini_out"
        sheet_list = ["SysVarDatabase", "Classe"]
        try:
            wb = update_iniout(wb, wb_classe, inisheet, sheet_list, self.excel_file)
            create_ini(wb, inisheet,r"../../../CustomerPrj/Restbus/Variants")
            logger.info("###### END 'Update Ini(.ini)' DEBUG INFORMATION ######")
            logger.info('-' * 80)
        except Exception as exp:
            logger.error(f"Failed to update ini: {exp}")
            messagebox.showerror(title="Error", message="Failed to update Ini sheet.\nPlease check the Console output.")
            self.run_subprocess()
            return False
        messagebox.showinfo(title="Update Ini", message="Successfully updated Ini sheet and Ini files.")
        self.enable_btns()
        return True

    # call create_gw
    def create_gw(self):
        """ creates caplt for gateway if needed"""
        logger.info("###### START 'Create SysVarMapping' DEBUG INFORMATION ######")
        self.disable_bnts()
        logger.info("Execution skipped, since gateway is not present")
        logger.info("###### END 'Create SysVarMapping' DEBUG INFORMATION ######")
        messagebox.showinfo(title="Create GW", message="Execution skipped, since gateway is not present.")
        self.enable_btns()
        return True
        wb = load_workbook(self.excel_file, data_only=True)
        sys_sheet = "SysVarDatabase"
        try:
            generate_gw(wb[sys_sheet], r"../../../CustomerPrj/Restbus/Nodes")
            logger.info("###### END 'Create SysVarMapping' DEBUG INFORMATION ######")
            logger.info('-' * 80)
        except Exception as exp:
            logger.error(f"Failed to create sysvarmapping: {exp}")
            messagebox.showerror(title="Error",
                                 message="Failed to create SysVar Mapping file.\nPlease check the Console output.")
            self.run_subprocess()
            return False
        messagebox.showinfo(title="Create GW", message="Successfully created SysVar Mapping file.")
        self.enable_btns()
        return True

    def arxml_sanity_check(self):
        self.disable_bnts()
        
        # Ask for directory
        folder_path = filedialog.askdirectory(title="Select Folder")
        
        if folder_path:
            logger.info('Start Analyzing files in the selected directory.')
            get_canoe_warnings = messagebox.askyesno("Canoe Warnings", "Do you want to get Canoe write window warnings?")
            arxml_sanity_check_main(folder_path, get_canoe_warnings)
        else:
            # Ask for ARXML files if no directory is selected
            file_paths = filedialog.askopenfilenames(title="Select One or More ARXML Files", filetypes=[("ARXML files", "*.arxml"), ("All files", "*.*")])
            
            if file_paths:
                logger.info('Start Analyzing selected files.')
                get_canoe_warnings = messagebox.askyesno("Canoe Warnings", "Do you want to get Canoe write window warnings?")
                arxml_sanity_check_main(list(file_paths), get_canoe_warnings)
            else:
                logger.info('No files or directory selected.')
                messagebox.showwarning(title="No Selection", message="No files or directory selected.")
        
        logger.info('Finished checking files.')
        messagebox.showinfo(title="Arxml Checks", message="Finished analyzing files,\nPlease check error_log.txt for report")
        
        self.enable_btns()


    # update Restbus
    def update_restbus(self):
        """ runs all restbus scripts in order"""
        messagebox.showinfo(title="Update Sheets", message="All RBS scripts except will be executed.")
        self.disable_bnts()
        for i in range(1, self.num_steps):
            if i <= 5:
                success = self.script_callback[i]()
                if not success:
                    messagebox.showinfo(title="Info",
                                        message="Please find the error and fix it, then re-run the scripts.")
                    break
        messagebox.showinfo(title="Execute RBS", message="FINISHED Successfully.")
        self.enable_btns()

    # update FDX
    def update_fdx(self):
        """ updates fdx nodes capl"""
        logger.info("###### START 'Update FDX Nodes' DEBUG INFORMATION ######")
        self.disable_bnts()
        df_FDX = pd.read_excel(self.excel_file_fdx, sheet_name=None,dtype=object)
        success = create_mapping_main(df_FDX, self.fdx_directory)
        #success = create_mapping_main(wb, wb_classe, self.fdx_directory)
        logger.info("###### END 'Update FDX Nodes' DEBUG INFORMATION ######")
        logger.info('-' * 80)
        if not success:
            messagebox.showerror(title="Error",
                                 message="Failed to create Mapping for FDX.\nPlease check the Console output.")
            self.run_subprocess()
            return False
        messagebox.showinfo(title="Create Nodes Mapping", message="Successfully updated Nodes for FDX.")
        self.enable_btns()
        return True

    def update_database(self):
        """ updates fdx xml """
        logger.info("###### START 'Update FDX Database' DEBUG INFORMATION ######")
        self.disable_bnts()
        df_FDX = pd.read_excel(self.excel_file_fdx, sheet_name=None, dtype=object)
        success = create_mapping_main_database(df_FDX, self.fdx_directory)
        logger.info("###### END 'Update FDX Database' DEBUG INFORMATION ######")
        logger.info('-' * 80)
        if not success:
            messagebox.showerror(title="Error",
                                 message="Failed to create Database mapping for FDX.\nPlease check the Console output.")
            self.run_subprocess()
            return False
        messagebox.showinfo(title="Create Database Mapping", message="Successfully created Database mapping for FDX.")
        self.enable_btns()
        return True

    def update_fdx_database_excel(self):
        """ """
        logger.info("###### START 'Update TriggersDatabase' DEBUG INFORMATION ######")
        self.disable_bnts()
        success = triggers_Gen.main_sequence(self.excel_file_fdx, self.json_database)
        logger.info("###### START 'Update TriggersDatabase' DEBUG INFORMATION ######")
        logger.info('-' * 80)
        if not success:
            messagebox.showerror(title="Error",
                                 message="Failed to create fdxTriggersDatabase triggers database.\nPlease check the Console output.")
            self.run_subprocess()
            return False
        messagebox.showinfo(title="Create fdxTriggersDatabase triggers",
                            message="Successfully created FDX Database excel triggers.")
        self.enable_btns()
        return True

    # update CarMaker
    # def update_carmaker(self):
    #     logger.info("###### START 'Update CarMaker Nodes' DEBUG INFORMATION ######")
    #     self.disable_bnts()
    #     wb = load_workbook(self.excel_file_carmaker, data_only=True)
    #     success = create_mapping_main_cm(wb, self.cm_directory)
    #     logger.info("###### END 'Update CarMaker Nodes' DEBUG INFORMATION ######")
    #     logger.info('-' * 80)
    #     if not success:
    #         messagebox.showerror(title="Error",
    #                              message="Failed to create Mapping for CarMaker.\nPlease check the Console output.")
    #         self.run_subprocess()
    #         return False
    #     messagebox.showinfo(title="Create CarMaker Nodes Mapping", message="Successfully updated Nodes for CarMaker.")
    #     self.enable_btns()
    #     return True


    def update_hil_json(self):
        """ """
        logger.info("###### START 'Update HIL specific JSON' DEBUG INFORMATION ######")
        self.disable_bnts()
        success = hil_specific.main_sequence(self.excel_file_fdx, self.specific_hil_json)
        logger.info("###### END 'Update HIL specific JSON' DEBUG INFORMATION ######")
        logger.info('-' * 80)
        if not success:
            messagebox.showerror(title="Error",
                                 message="Failed to create HIL_specific_triggers.json .\nPlease check the Console output.")
            self.run_subprocess()
            return False
        messagebox.showinfo(title="Create HIL_specific_triggers.json",
                            message="Successfully created HIL_specific_triggers.json.")
        self.enable_btns()
        return True

    def update_xcpconfig(self):
        """ this is main function for updating XCP related files, reduces a2l, updates XCP_config as per XCP_database sheet"""

        def generate_dasy(xcp, a2l_dasy, dasy_sheet, ecu, a2l_type):
            """
            generate and update xcpconfig and reduce a2l for dasy as per XCP_database sheet

            Args:
              xcp (object): XCP_to_A2L_diff instance
              a2l_dasy (str):
              dasy_sheet (str):
              ecu (str):
              a2l_type (str):

            """
            logger.info(f"###### Start procedure for {a2l_type} #######")
            xcp.delete_signals(xcpconfig, a2l_dasy, ecu, self.customer, lean_choise)
            ecu_index_dasy = xcp.find_ECU(xcpconfig, ecu)
            logger.info("Updating XCP_Database.xlsx cycle times for DASy")
            update_database_cycle_times(dasy_sheet, a2l_dasy, xcpconfig, ecu_index_dasy, ecu)
            logger.info("Extract Dasy data from XCP_Database.xlsx")
            xcp_dasy = xcp.parseXCP(dasy_sheet, self.excel_file_xcp)
            logger.info(f"Start parsing full XCP list from  {a2l_dasy}")
            a2l_dasy_list = xcp.extract_A2l(a2l_dasy)
            logger.info("Compare Dasy XCP_Database.xlsx XCP entries against A2L file")
            xcp.parseA2l(a2l_dasy, xcp_dasy, a2l_dasy_list)
            if xcp.a2lfile_missvar != {}:
                raise Exception("Missing XCP variables")

            if xcp_dasy:
                logger.info("Updating XCP configuration file for DASy")
                """" Get the index for DASy ECU"""
                for missing_xcp in reversed(xcp_dasy):
                    xcp.add_XCP(xcpconfig, missing_xcp, ecu_index_dasy, "Dasy")
                logger.info("Finished updating XCP configuration file for DASy")
                complete_list_dasy = xcp.parseXCPconfig2(xcpconfig, ecu_index_dasy)
                if lean_choise and a2l_dasy is not None:
                    xcp.reduce_A2l(a2l_dasy, complete_list_dasy, a2l_type, self.customer)

        self.disable_bnts()
        try:
            success = True
            filepath = os.path.abspath(os.path.join(os.getcwd(),r"Control/Missing_XCP_variables_in_A2L.txt"))

            with open(filepath, 'w') as file:
                file.close()


            xcpconfig = r'..\\..\\..\\CustomerPrj\\XCP\\XCP_config_gen.xcpcfg'

            dict_rear = {"xcp_RadarFL": "RadarFL", "xcp_RadarFR": "RadarFR", "xcp_RadarRL": "RadarRL", "xcp_RadarRR": "RadarRR"}
            al2_list = []
            a2l_dasy = None
            a2l_dasy_delta1 = None
            a2l_dasy_delta5 = None
            ################# Checkboxes logic #############################################

            if self.customer == "FORD":
                if self.dasy_state.get() == 0 and self.front_obj_state.get() == 0 and self.corner_state.get() == 0:
                    messagebox.showwarning(title="Warning", message="No checkbox selected.")
                    self.enable_btns()
                    return

                if self.dasy_state.get():
                    a2l_dasy = self.checkbox_logic("Select the DASy <<AL2 file>> file : ")
                    al2_list.append(a2l_dasy)
                    if not a2l_dasy:
                        return
                else:
                    a2l_dasy = None
            else:
                if self.dasy_state_delta1.get() == 0 and self.dasy_state_delta5.get() == 0 and self.front_loc_state.get() == 0 and self.front_obj_state.get() == 0 and self.corner_state.get() == 0:
                    messagebox.showwarning(title="Warning", message="No checkbox selected.")
                    self.enable_btns()
                    return

                if self.dasy_state_delta1.get():
                    a2l_dasy_delta1 = self.checkbox_logic("Select DASy Delta1 <<AL2 file>> file : ")
                    al2_list.append(a2l_dasy_delta1)
                    if not a2l_dasy_delta1:
                        return

                if self.dasy_state_delta5.get():
                    a2l_dasy_delta5 = self.checkbox_logic("Select DASy Delta5 <<AL2 file>> file : ")
                    al2_list.append(a2l_dasy_delta5)
                    if not a2l_dasy_delta5:
                        return

            if self.customer != "FORD":
                if self.front_loc_state.get():
                    a2l_front_loc = self.checkbox_logic("Select the RadarFC (Locations) <<AL2 file>> file : ")
                    al2_list.append(a2l_front_loc)
                    if not a2l_front_loc:
                        return
                else:
                    a2l_front_loc = None
            else:
                a2l_front_loc = None

            if self.front_obj_state.get():
                a2l_front_obj = self.checkbox_logic("Select the RadarFC (Objects) <<AL2 file>> file : ")
                al2_list.append(a2l_front_obj)
                if not a2l_front_obj:
                    return
            else:
                a2l_front_obj = None

            if self.corner_state.get():
                a2l_rear = self.checkbox_logic("Select the Radar Rear <<AL2 file>> file : ")
                al2_list.append(a2l_rear)
                if not a2l_rear:
                    return
            else:
                a2l_rear = None

            lean_choise = askyesno(title="A2L lean", message="Do you want to lean the selected A2Ls?")
        ################# Checkboxes logic END #############################################
            xcp = XCP_to_A2L_diff.XCPConfigurator(al2_list[0])
            ############### DaSy Section ##################################
            if a2l_dasy:
                dasy_sheet = "xcp_Dasy"
                ecu = "ACP_FD3"
                a2l_type = "ACP_FD3"
                generate_dasy(xcp, a2l_dasy, dasy_sheet, ecu, a2l_type)

            if a2l_dasy_delta1:
                dasy_sheet = "xcp_DPCdelta1"
                ecu = "DPCdelta1"
                a2l_type = "DPCDelta1"
                generate_dasy(xcp, a2l_dasy_delta1, dasy_sheet, ecu, a2l_type)

            if a2l_dasy_delta5:
                dasy_sheet = "xcp_DPCdelta5"
                ecu = "DPCdelta5"
                a2l_type = "DPCDelta5"
                generate_dasy(xcp, a2l_dasy_delta5, dasy_sheet, ecu, a2l_type)

            ############### End DaSy Section ##################################

            ############### RadarFC Locations Section ##################################

            if a2l_front_loc:
                tag = "RadarFC_Loc"
                a2l_type = "RadarFC_Locations_Gamma"
                logger.info(f"###### Start procedure for {a2l_type} #######")
                xcp.delete_signals(xcpconfig, a2l_front_loc, tag, self.customer, lean_choise)
                """" Get the index for Front Radar ECU"""
                ecu_index_front = xcp.find_ECU(xcpconfig, "RadarFC_Loc")
                logger.info("Updating XCP_Database.xlsx cycle times for RadarFC Locations")
                update_database_cycle_times("xcp_RPCgamma1", a2l_front_loc, xcpconfig, ecu_index_front, tag)
                logger.info("Extract RadarFC Locations data from XCP_Database.xlsx")
                xcp_fc_loc = xcp.parseXCP("xcp_RPCgamma1", self.excel_file_xcp)
                logger.info(f"Start parsing full XCP list from  {a2l_front_loc}")
                a2l_front_loc_list = xcp.extract_A2l(a2l_front_loc)
                logger.info("Compare Dasy XCP_Database.xlsx XCP entries against A2L file")
                xcp.parseA2l(a2l_front_loc, xcp_fc_loc, a2l_front_loc_list)
                if xcp.a2lfile_missvar != {}:
                    raise Exception("Missing XCP variables")

                if xcp_fc_loc:
                    logger.info("Updating XCP configuration file for RadarFC Locations")
                    for missing_xcp in reversed(xcp_fc_loc):
                        xcp.add_XCP(xcpconfig, missing_xcp, ecu_index_front, "Radar")
                    logger.info("Finished updating XCP configuration file for RadarFC Locations")
                    complete_list_front = xcp.parseXCPconfig2(xcpconfig, ecu_index_front)
                    if lean_choise and a2l_front_loc is not None:
                        xcp.reduce_A2l(a2l_front_loc, complete_list_front, a2l_type, self.customer)
                ############### End RadarFC Locations Section ##################################

                ############### RadarFC Objects Section ##################################

            if a2l_front_obj:
                if self.customer != "FORD":
                    tag = "RadarFC_Obj"
                    a2l_type = "RadarFC_Objects_Alpha"
                    sheet_name = "xcp_RPCalpha2"
                else:
                    tag = "RadarFC"
                    a2l_type = "RadarFC_Objects"
                    sheet_name = "xcp_RadarFC"
                logger.info(f"###### Start procedure for {a2l_type} #######")
                xcp.delete_signals(xcpconfig, a2l_front_obj, tag, self.customer, lean_choise)
                ecu_index_front = xcp.find_ECU(xcpconfig, tag)
                """ Update excel database cycle times """
                logger.info("Updating XCP_Database.xlsx cycle times for RadarFC Objects")
                update_database_cycle_times(sheet_name, a2l_front_obj, xcpconfig, ecu_index_front, tag)
                logger.info("Extract RadarFC Locations data from XCP_Database.xlsx")
                xcp_fc_obj = xcp.parseXCP(sheet_name, self.excel_file_xcp)
                logger.info(f"Start parsing full XCP list from  {a2l_front_obj}")
                a2l_front_obj_list = xcp.extract_A2l(a2l_front_obj)
                logger.info("Compare Dasy XCP_Database.xlsx XCP entries against A2L file")
                xcp.parseA2l(a2l_front_obj, xcp_fc_obj, a2l_front_obj_list)
                if xcp.a2lfile_missvar != {}:
                    raise Exception("Missing XCP variables")

                for missing_xcp in reversed(xcp_fc_obj):
                    xcp.add_XCP(xcpconfig, missing_xcp, ecu_index_front, "Radar")
                logger.info("Finished updating XCP configuration file for RadarFC")
                complete_list_front = xcp.parseXCPconfig2(xcpconfig, ecu_index_front)
                if lean_choise and a2l_front_obj is not None:
                    xcp.reduce_A2l(a2l_front_obj, complete_list_front, a2l_type, self.customer)

            ############### END RadarFC Objects Section ##################################

            ############### Radar Rear Section ##################################

            if a2l_rear:
                logger.info(f"###### Start procedure for RearRadar(s) #######")
                logger.info(f"Start parsing full XCP list from  {a2l_rear}")
                a2l_rear_list = xcp.extract_A2l(a2l_rear)
                reduced = False
                for sheet, rear_ecu in dict_rear.items():
                    logger.info(f"###### Updating for {rear_ecu} #######")
                    logger.info("Extract Rear Radar data from XCP_Database.xlsx")
                    xcp_rr = xcp.parseXCP(sheet, self.excel_file_xcp)
                    if xcp_rr:
                        xcp.delete_signals(xcpconfig, a2l_rear, rear_ecu, self.customer, lean_choise)
                        ecu_index_rear = xcp.find_ECU(xcpconfig, rear_ecu)
                        logger.info("Updating XCP_Database.xlsx cycle times for Radar Rear")
                        update_database_cycle_times(rear_ecu, a2l_rear, xcpconfig, ecu_index_rear, tag)
                        logger.info("Compare RearRadar XCP_Database.xlsx XCP entries against A2L file")
                        xcp.parseA2l(a2l_rear, xcp_rr, a2l_rear_list)
                        if xcp.a2lfile_missvar != {}:
                            raise Exception("Missing XCP variables")
                        if ecu_index_rear is not None:
                            ecu_index_rear = xcp.find_ECU(xcpconfig, rear_ecu)
                            logger.info(f"Updating XCP configuration file for {rear_ecu}")
                            for missing_xcp in reversed(xcp_rr):
                                xcp.add_XCP(xcpconfig, missing_xcp, ecu_index_rear, "Radar")
                            logger.info(f"Finished updating XCP configuration file for {rear_ecu}")
                            complete_list_rear = xcp.parseXCPconfig2(xcpconfig, ecu_index_rear)
                            if complete_list_rear is not None and reduced is False:
                                if lean_choise and a2l_rear is not None:
                                    xcp.reduce_A2l(a2l_rear, complete_list_rear, "RadarCorner", self.customer)
                                    reduced = True
                    else:
                        logger.info(f"No XCP entries in XCP_Database.xlsx")

            if lean_choise is True:
                messagebox.showinfo(title="Update xcpcfg file",
                                    message="Successfully updated xcpcfg file and reduced A2L.")
            else:
                messagebox.showinfo(title="Update xcpcfg file", message="Successfully updated xcpcfg file.")
        except Exception as e:
            success = False
            logger.error(f"Error occured-> {e}")
        if not success:
            messagebox.showerror(title="Error",
                                 message="Failed to create xcpcfg file.\nPlease check the Console output.")
            self.run_subprocess()
        self.enable_btns()

    def checkbox_logic(self, text):
        """
        takes a2l file input from user, checks if the given input is valid or not, gives warning

        Args:
          text: title for dialog window

        Returns:
             str: path of selected file by user

        """
        a2l = easygui.fileopenbox(text)
        if a2l is None:
            messagebox.showwarning(title="Warning", message="Please select the a2l from browse window.")
            self.enable_btns()
            return
        if "reduced" in a2l:
            messagebox.showwarning(title="Warning", message="Please select the original a2l.")
            self.enable_btns()
            return
        return a2l

    def createMulticanoe(self):
        """ for creating multicanoe cfgs from main cfg"""
        logger.info("------------------------Start Creating Multicanoe ----------------------")
        self.disable_bnts()
        try:
            if not(self.create_maincfg_3_state.get() or self.create_master_slave_2_state.get() or self.create_master_slave_3_state.get()):
                messagebox.showwarning(title="Warning", message="No checkbox selected")
                self.enable_btns()
                return

            if (self.create_maincfg_3_state.get()):
                if not(self.create_master_slave_2_state.get() or self.create_master_slave_3_state.get()):
                    createCfg3(quit_canoe=True)
                else:
                    createCfg3(quit_canoe=False)
            if self.create_master_slave_3_state.get():
                create_multicanoe_main(main3_needed=True)
            if self.create_master_slave_2_state.get():
                create_multicanoe_main(main3_needed=False)

            messagebox.showinfo(title="Create Multicanoe", message="Created Successfully.")
            self.enable_btns()
        except Exception as e:
            logger.error(f"Error while creating multicanoe -> {e}")
            messagebox.showerror(title="Error",
                                 message="Error while creating multicanoe.\nPlease check the Console output.")
            self.run_subprocess()
            self.enable_btns()
            return False


    def updateCopyRights(self):
        """upadtes the copyrights header for complete repo"""
        self.disable_bnts()
        try:
            logger.info(f"-----------------------START COPYRIGHTS UPDATE--------------------------")
            generate_copy_rights()
            logger.info(f"-----------------------END COPYRIGHTS UPDATE--------------------------")
            self.enable_btns()
            # messagebox.showinfo(title="Copyrights Update",
            #                     message="Successfully updated copyrights\nfor *.(py,can,cin)")
        except Exception as e:
            logger.error(f"Error -> {e}")
            messagebox.showerror(title="Error",
                                 message="Failed to update Copyrights\nCheck console for info")
            self.enable_btns()

    def generateOfflineDocu(self):
        """updates the offline analyzer documentation"""
        self.disable_bnts()
        try:
            logger.info(f"-----------------------START OFFLINE ANALYZER DOCU UPDATE--------------------------")
            generate_offline_docu()
            logger.info(f"-----------------------END OFFLINE ANALYZER DOCU UPDATE--------------------------")
            self.enable_btns()
        except Exception as e:
            logger.error(f"Error -> {e}")
            messagebox.showerror(title="Error",
                                 message="Failed to update Offline Analyzer Documentation\nCheck console for info")
            self.enable_btns()

    def updateDoxygenDocumentation(self):
        """upadtes doxygen documentation for both cus and platform folders excluding blacklists"""
        def updateExcludeInDoxygenConfig(config_path, blacklist_file_path):
            """
            updates doxygen configs to exclude black list
            Args:
                config_path: doxygen gen config path
                blacklist_file_path: blacklist txt file path
            """
            source_directory = r"..\..\..\.."
            try:
                with open(blacklist_file_path, "r") as file:
                    # Read each line and store them in a list
                    entries = [line.strip() for line in file]
                    entries_modified = [source_directory + line if "\\" in line else line for line in entries]
                    logger.info(f"Filtered files and folders -- {entries_modified}")

                exclude_patterns = ""
                for entry in entries_modified[:-1]:
                    exclude_patterns += entry + " \\\n" + "\t"*6 + " "

                exclude_patterns += entries_modified[-1]
                exclude_line_found = False

                config_file_content = []
                with open(config_path, "r") as cfg_file:
                    for line in cfg_file:
                        if exclude_line_found:
                            if "#END_EXCLUDE" in line:
                                exclude_line_found = False
                                config_file_content.append("\n#END_EXCLUDE\n")
                                continue
                            else:
                                continue

                        split = line.split("=")
                        if len(split) >= 1:
                            if split[0].strip() == "EXCLUDE":
                                append_exclude = "EXCLUDE" + " "*16 + "= " + exclude_patterns
                                config_file_content.append(append_exclude)
                                exclude_line_found=True
                            else:
                                config_file_content.append(line)
                        else:
                            config_file_content.append(line)
                with open(config_path, "w") as cfg_file:
                    cfg_file.writelines(config_file_content)
                logger.info(f"Updated Doxygen Config file according to {blacklist_file_path}")

            except FileNotFoundError:
                logger.error(f"File '{blacklist_file_path}' not found.")
            except Exception as e:
                logger.info(f"An error occurred: {str(e)}")



        script_path = os.path.dirname(os.path.abspath(__file__))

        custprj_bat_path = script_path + r"\..\..\..\..\CustomerPrj\Docs\Doxygen\Run_doxygen.bat"
        cus_config_path = script_path + r"\..\..\..\..\CustomerPrj\Docs\Doxygen\Config\Doxyfile"
        platform_bat_path = script_path + r"\..\..\..\Docs\Doxygen\Run_doxygen.bat"
        platform_config_path = script_path + r"\..\..\..\Docs\Doxygen\Config\Doxyfile"

        platform_black_list_path = script_path + r"\..\..\..\..\Platform\Classe\Scripts\Create_delivery_zip\black_list.txt"
        customer_black_list_path = script_path + r"\..\..\..\..\Deployment\cus_black_list.txt"

        self.disable_bnts()
        try:
            logger.info(f"-----------------------START UPDATING DOC FOR CustomerPrj folder--------------------------")
            updateExcludeInDoxygenConfig(cus_config_path, customer_black_list_path)
            sp.run([custprj_bat_path], cwd=os.path.dirname(custprj_bat_path), stdin=sp.PIPE, text=True)
            logger.info(f"-----------------------END UPDATING DOC FOR CustomerPrj folder--------------------------")

            logger.info(f"-----------------------START UPDATING DOC FOR Platform folder--------------------------")
            updateExcludeInDoxygenConfig(platform_config_path, platform_black_list_path)
            sp.run([platform_bat_path], cwd=os.path.dirname(platform_bat_path), stdin=sp.PIPE, text=True)
            logger.info(f"-----------------------END UPDATING DOC FOR Platform folder--------------------------")
            logger.info(f"-----------------------Updated Doc for Both CustPrj and Platform folders --------------------------")
            # messagebox.showinfo(title="Documentation Update",
            #                     message="Successfully updated Doxygen documentation\nfor both cus and platform")
            self.enable_btns()

        except Exception as e:
            logger.error(f"Error -> {e}")
            messagebox.showerror(title="Error", message="Failed to update Doc\nCheck console for info")
            self.enable_btns()

    def deliveryZipFiles(self):
        """
        Zips files excluding blacklist
        """
        try:
            logger.info("--------Start zipping files----------")
            zipping = delivery.Deliver()
            zipping.createDeliveryZipMain()
            logger.info("--------End zipping files----------")
            messagebox.showinfo(title="Execution completed",
                                message="Successfully completed given tasks")
        except Exception as e:
            logger.error(f"Error occured while creating zip -> {e}")
            messagebox.showerror(title="Error",message="Error occured while creating zip\nCheck console for info")

    def deliveryAutomation(self):
        """
        main function for delivery automation, checks if any checkbox is enabled if yes triggers that task

        """
        self.disable_bnts()
        logger.info("-----------Start Delivery Automation---------------")
        try:
            if not(self.update_copyrights_state.get() or self.update_doxygen_state.get() or self.zip_files_state.get() or self.offline_docu_state.get()):
                messagebox.showwarning(title="Warning", message="No checkbox selected")
                self.enable_btns()
                return
            if self.update_copyrights_state.get():
                self.updateCopyRights()
            if self.update_doxygen_state.get():
                self.updateDoxygenDocumentation()
            if self.zip_files_state.get():
                self.deliveryZipFiles()
            if self.offline_docu_state.get():
                self.generateOfflineDocu()
        except Exception as e:
            logger.error(f"Error -> {e}")
            messagebox.showerror(title="Error",
                                 message="Error occured\nCheck console for info")
            self.enable_btns()

    def releaseNoteAutoGen(self):
        """
        generate release note using pull request title/description

        """
        self.disable_bnts()
        try:
            logger.info(f"-----------------------START RELEASE NOTE GUI--------------------------")
            rel_note_gui = releaseNoteAutoGeneration()
            rel_note_gui.mainloop()
            logger.info(f"-----------------------END RELEASE NOTE GUI--------------------------")
            self.enable_btns()
        except Exception as e:
            logger.error(f"Error -> {e}")
            messagebox.showerror(title="Error",
                                 message="Failed to open Release note GUI\nCheck console for info")
            logger.info(f"-----------------------END RELEASE NOTE GUI--------------------------")
            self.enable_btns()

    def disable_bnts(self):
        """ disables all button"""
        self.create_node_sheet_arxml_btn.config(state=tk.DISABLED)
        self.create_node_sheet_arxml_eth_btn.config(state=tk.DISABLED)
        self.create_node_sheet_btn.config(state=tk.DISABLED)
        # self.update_node_sheet_btn.config(state=tk.DISABLED)
        self.json_btn.config(state=tk.DISABLED)
        self.update_fdx_database_btn.config(state=tk.DISABLED)
        self.update_xcp_btn.config(state=tk.DISABLED)
        self.buttons["button_1"].config(state=tk.DISABLED)
        self.buttons["button_2"].config(state=tk.DISABLED)
        self.buttons["button_2_eth"].config(state=tk.DISABLED)
        self.buttons["button_3"].config(state=tk.DISABLED)
        self.buttons["button_4"].config(state=tk.DISABLED)
        self.buttons["button_5"].config(state=tk.DISABLED)
        #self.buttons["button_6"].config(state=tk.DISABLED)
        self.execute_all_btn.config(state=tk.DISABLED)
        self.update_nodes_btn.config(state=tk.DISABLED)
        self.update_databases_btn.config(state=tk.DISABLED)
        # self.cm_btn.config(state=tk.DISABLED)
        self.create_test_database_btn.config(state=tk.DISABLED)
        self.upadte_multicanoe_btn.config(state=tk.DISABLED)
        self.delivery_auto_btn.config(state=tk.DISABLED)
        self.sanity_check_btn.config(state=tk.DISABLED)

    def enable_btns(self):
        """ Enables all buttons """
        self.create_node_sheet_arxml_btn.config(state=tk.NORMAL)
        self.create_node_sheet_arxml_eth_btn.config(state=tk.NORMAL)
        self.create_node_sheet_btn.config(state=tk.NORMAL)
        # self.update_node_sheet_btn.config(state=tk.NORMAL)
        self.json_btn.config(state=tk.NORMAL)
        self.update_fdx_database_btn.config(state=tk.NORMAL)
        self.update_xcp_btn.config(state=tk.NORMAL)
        self.buttons["button_1"].config(state=tk.NORMAL)
        self.buttons["button_2"].config(state=tk.NORMAL)
        self.buttons["button_2_eth"].config(state=tk.NORMAL)
        self.buttons["button_3"].config(state=tk.NORMAL)
        self.buttons["button_4"].config(state=tk.NORMAL)
        self.buttons["button_5"].config(state=tk.NORMAL)
        #self.buttons["button_6"].config(state=tk.NORMAL)
        self.execute_all_btn.config(state=tk.NORMAL)
        self.update_nodes_btn.config(state=tk.NORMAL)
        self.update_databases_btn.config(state=tk.NORMAL)
        # self.cm_btn.config(state=tk.NORMAL)
        self.create_test_database_btn.config(state=tk.NORMAL)
        self.upadte_multicanoe_btn.config(state=tk.NORMAL)
        self.delivery_auto_btn.config(state=tk.NORMAL)
        self.sanity_check_btn.config(state=tk.NORMAL)

    def __call__(self):
        self.main_content()
        self.window.mainloop()


if __name__ == "__main__":
    APP = WorkflowGUI()
    APP()
