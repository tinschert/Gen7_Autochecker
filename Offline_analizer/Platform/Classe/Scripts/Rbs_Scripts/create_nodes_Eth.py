# -*- coding: utf-8 -*-
# @file create_nodes_Eth.py
# @author ADAS_HIL_TEAM
# @date 11-02-2023

##################################################################
# C O P Y R I G H T S
# ----------------------------------------------------------------
# Copyright (c) 2023 by Robert Bosch GmbH. All rights reserved.

# The reproduction, distribution and utilization of this file as
# well as the communication of its contents to others without express
# authorization is prohibited. Offenders will be held liable for the
# payment of damages. All rights reserved in the event of the grant
# of a patent, utility model or design.

##################################################################
import traceback

import os, sys
import re
import pandas as pd
try:
    pd.set_option('future.no_silent_downcasting', True)
except:
    pass
import numpy as np
from math import ceil
from openpyxl import load_workbook

try:
    from Rbs_Scripts.update_sysvartab import getMappingData
    from Rbs_Scripts.create_nodes import extractHilCtrl
except ImportError:
    sys.path.append(os.path.dirname(os.path.abspath(__file__)))
    from update_sysvartab import getMappingData
    from create_nodes import extractHilCtrl
try:
    sys.path.append(os.getcwd() + r"\..\..\CustomerPrj\Restbus\Scripts")
    from pattern_matching import *
except ImportError:
    sys.path.append(os.getcwd() + r"\..\..\..\..\CustomerPrj\Restbus\Scripts")
    from pattern_matching import *

try:
    from Control.logging_config import logger
except ImportError:
    sys.path.append(os.path.dirname(os.path.abspath(__file__)) + r"\..\Control")
    from logging_config import logger

#END IMPORTS

# allow longer names up to 100 characters
pd.options.display.max_colwidth = 100
pd.options.display.colheader_justify = 'left'

flt_evt_node = ['']
flt_evt_cnt = 0  # FLT_EVT

def getColNamespace(entity):
    if "message" in entity:
        return col_msg
    elif "pdu" in entity:
        return col_pdu
    elif "group" in entity:
        return col_group
    else:
        return col_msg
    


#someip arguement namespae generation

def get_array_info(arr):
    temp = arr.split('[')
    arr_name = temp[0]
    size = int(temp[-1].split(']')[0])
    return (arr_name, size)


def get_signal_info(row, namespace=""):
    signal_list = []
    signal = row[col_Signal_name]
    no_elements = row[col_No_of_Elements]
    if not is_empty_or_spaces(signal):
        if not is_empty_or_spaces(no_elements):
            for i in range(int(no_elements)):
                if namespace:
                    signal_list.append(namespace + '::' + signal + f'_{i}')
                else:
                    signal_list.append(f'{signal}_{i}')
        else:
            if namespace:
                signal_list.append(namespace + '::' + signal)
            else:
                signal_list.append(signal)
    else:
        if namespace:
            signal_list.append(namespace)
        else:
            signal_list.append("")
    return signal_list


def generate_namespace_for_a_column(value):
    res = []
    if value:
        if '[' in value:
            arr_name, arr_size = get_array_info(value)
            for i in range(arr_size):
                res.append(arr_name+f'_{i}')
        else:
            res.append(value)
    return res
def generate_parameter_namespaces(df):
    col_parameter_mapping = {1: col_para_level1,
                             2:col_para_level2,
                             3:col_para_level3,
                             4:col_para_level4,
                             5:col_para_level5}

    df = df.fillna("")
    return_list = []
    
    for i, row in df.iterrows():
        namespace = ""
        result_namespaces = []
        paramerter_level_dict = {}
        
        for i in range(1,6):
            col_name = col_parameter_mapping.get(i)
            key = f'l{i}'
            paramerter_level_dict[key] = generate_namespace_for_a_column(row[col_name])
        
        if paramerter_level_dict['l1']:
            for l1 in paramerter_level_dict['l1']:
                if paramerter_level_dict['l2']:
                    for l2 in paramerter_level_dict['l2']:
                        if paramerter_level_dict['l3']:
                            for l3 in paramerter_level_dict['l3']:
                                if paramerter_level_dict['l4']:
                                    for l4 in paramerter_level_dict['l4']:
                                        if paramerter_level_dict['l5']:
                                            for l5 in paramerter_level_dict['l5']:
                                                result_namespaces.append(f'{l1}::{l2}::{l3}::{l4}::{l5}')
                                        else:
                                            result_namespaces.append(f'{l1}::{l2}::{l3}::{l4}')
                                   
                                else:
                                    result_namespaces.append(f'{l1}::{l2}::{l3}')
                            
                        else:
                            result_namespaces.append(f'{l1}::{l2}')
                else:
                    result_namespaces.append(f'{l1}')
        
        
        final_namespace = []
        signal_info = get_signal_info(row)
        if signal_info:
            if result_namespaces:
                for nmps in result_namespaces:
                    for signal_nm in signal_info:
                        final_namespace.append(f'{nmps}::{signal_nm}')
            else:
                for signal_nm in signal_info:
                    final_namespace.append(signal_nm)
        else:
            final_namespace.append('')
        datatype = row[col_Signal_DataType]
        final_namespace = [z + f'#{datatype}' for z in final_namespace]
        temp = final_namespace.copy()
        return_list.extend(temp)            

    return return_list

def replace_para_namespace_pattern(input_string):
    # If the string is just _number or _number::, do nothing
    if re.fullmatch(r'_\d+::?$', input_string):
        return input_string.replace('::', '.')
    # Otherwise, replace _number:: in middle of string with [number]::
    input_string = re.sub(r'(?<=\w)_(\d+)::', r'[\1]::', input_string)
    # And replace _number at the end of the string with [number]
    input_string = re.sub(r'(?<=\w)_(\d+)$', r'[\1]', input_string)
    input_string = input_string.replace('::', '.')

    if '2D_ARRAY_' in input_string:
        input_string = input_string.replace('.2D_ARRAY_', '')
        input_string = input_string.replace('2D_ARRAY_', '')

    input_string = r'.' + input_string

    return input_string
def replace_for_direct_namespace(input_string):
    if input_string.startswith(r'.'):
        return input_string[1:]
    return input_string


## END someip arguement namespace generation

def filterSomeipData_1(df):
    """
    Filters out services which are provided and consumed by same node
    
    Args:
        df (DataFrame): Input dataframe
    
    Returns:
        DataFrame: Filtered df
    """
    # Filter rows based on 'SD Type'
    pdu_data = df[df[col_network_protocol]=='ETH_PDU']
    someip_data = df[df[col_network_protocol]=='ETH_SOMEIP']
    df_provided = someip_data[someip_data[col_SDType] == 'provide']
    df_consumed = someip_data[someip_data[col_SDType] == 'consume']

    # Create a unique identifier for services
    df_provided['temp'] = 'sif_' + df_provided[col_Service_ID].astype(str) + '_' + \
                              df_provided[col_Major_version].astype(str) + '_' + \
                              df_provided[col_Minor_version].astype(str) + '_' + \
                              df_provided[col_Instance_ID].astype(str)

    df_consumed['temp'] = 'sif_' + df_consumed[col_Service_ID].astype(str) + '_' + \
                              df_consumed[col_Major_version].astype(str) + '_' + \
                              df_consumed[col_Minor_version].astype(str) + '_' + \
                              df_consumed[col_Instance_ID].astype(str)

    # Filter consumed services not in provided services
    provided_rbs_service_list = df_provided['temp'].unique()
    rbs_df_consumed_filtered = df_consumed[~df_consumed['temp'].isin(provided_rbs_service_list)]

    columns_row0_df = df.iloc[[0]]
    # Concatenate provided and filtered consumed dataframes, drop the 'temp' column
    concat_list = [columns_row0_df]
    if not(df_provided.empty):
        concat_list.append(df_provided)
    if not(rbs_df_consumed_filtered.empty):
        concat_list.append(rbs_df_consumed_filtered)
    if not (pdu_data.empty):
        concat_list.append(pdu_data)

    if not(df.empty) and len(concat_list)>1:
        result_df = pd.concat(concat_list)
        if 'temp' in result_df.columns:
            result_df = result_df.drop(columns=['temp'])
    elif len(concat_list) == 1:
        result_df = concat_list[0]
    else:
        result_df = df
    
    logger.info("Filtered out services which are provided and consumed by same node")
    
    return result_df


def filterSomeipData_2(df):
    """
     if in VectorSimulation Node consuming same servic with diff IP take 1 occurrence
        DataFrame: Filtered df
    """
    # Filter rows based on 'SD Type'
    pdu_data = df[df[col_network_protocol] == 'ETH_PDU']
    someip_data = df[df[col_network_protocol] == 'ETH_SOMEIP']
    df_provided = someip_data[someip_data[col_SDType] == 'provide']
    df_consumed = someip_data[someip_data[col_SDType] == 'consume']

    # Create a unique identifier for services
    df_provided['temp'] = 'sif_' + df_provided[col_Service_ID].astype(str) + '_' + \
                          df_provided[col_Major_version].astype(str) + '_' + \
                          df_provided[col_Minor_version].astype(str) + '_' + \
                          df_provided[col_Instance_ID].astype(str)

    df_consumed['temp'] = 'sif_' + df_consumed[col_Service_ID].astype(str) + '_' + \
                          df_consumed[col_Major_version].astype(str) + '_' + \
                          df_consumed[col_Minor_version].astype(str) + '_' + \
                          df_consumed[col_Instance_ID].astype(str)

    # if in VectorSimulation Node consuming same servic with diff IP take 1 occurrence
    df_consumed = df_consumed.drop_duplicates(subset=[col for col in df_consumed.columns if col not in [col_ip_address, col_Service_name, col_service_instance_name]], keep='first')

    columns_row0_df = df.iloc[[0]]
    # Concatenate provided and filtered consumed dataframes, drop the 'temp' column
    concat_list = [columns_row0_df]
    if not (df_provided.empty):
        concat_list.append(df_provided)
    if not (df_consumed.empty):
        concat_list.append(df_consumed)
    if not (pdu_data.empty):
        concat_list.append(pdu_data)

    if not (df.empty) and len(concat_list) > 1:
        result_df = pd.concat(concat_list)
        if 'temp' in result_df.columns:
            result_df = result_df.drop(columns=['temp'])
    elif len(concat_list) == 1:
        result_df = concat_list[0]
    else:
        result_df = df
    return result_df



# read worksheet 'sheet_name' from workbook 'wb'
def generate_dataframe(wb, sheet_name):
    global node, node_name, bus_name, bus_type_from_sheet_name, bus_type, input_file, node_df, num_of_rows, hil_ctrl_node_dict, is_public_can, network_name, node_entity
    global col_Service_name,col_service_instance_name,col_Service_ID,col_Major_version,col_Minor_version,col_Instance_ID,col_MemberType,\
        col_EventGroup,col_field_name,col_Member_name,col_Member_ID,col_Parameter,col_ParameterType,col_No_of_Parameters,\
        col_para_level1, col_para_level2, col_para_level3, col_para_level4, col_para_level5,\
        col_Signal_name,col_No_of_Elements,col_fire_and_forget,col_signal_group,col_pdu,col_pdu_type, \
        col_payload_pdu_type,col_payload_pdu,col_selector_field_signal,col_startbit,col_pdu_length_byte, \
        col_signal_length_bit,col_initial_value,col_max_value,col_cycle_time,col_texttable,col_enum,\
        col_value_type,col_comment,col_dlc,col_variant,col_value_table,col_endtoendprotection,col_Signal_ByteOrder, \
        col_Signal_DataType,col_Factor,col_Offset,col_Minimum,col_Maximum,col_Unit, col_ip_address, col_udp_port, col_tcp_port, col_tp_protocol, \
        col_SDType, col_vlan_id, col_vlan_name, col_network_protocol, col_autosar_type, col_field_type, col_EventGroup_ID
    global block_nwt_endpoint_list, blocked_df, generated_app_endpoint_initialize_list
    global col_namespace
    global current_sheet, blk_pdu_list

    current_sheet = sheet_name
    script_path = "\\".join(os.path.dirname(os.path.abspath(__file__)).split("\\")[-4:])
    script_path = "// Autogenerated by -> " + script_path + r"\create_nodes.py"
    node = ['/*@!Encoding:1252*/', script_path]
    blk_pdu_list = []
    #is_public_can = True if "public" in dict_network_info[sheet_name].lower() else False
    #node_entity = dict_entity[sheet_name]
    network_name = dict_network_name[sheet_name]
    ws1 = wb[sheet_name]
    name_list = sheet_name.split('_')
    input_file = name_list[-1]
    node_name = name_list[-2]
    bus_name = name_list[-3]
    bus_type_from_sheet_name = name_list[-4] #bus type for node name(.can) & for cus_bus
    #bus_type = dict_network_type[sheet_name] #bus type from network category column: CAN/CANFD/ETH
    bus_type="ETH"
    #print(input_file, node_name, bus_name,bus_type_from_sheet_name,bus_type)
    

    try:
        hil_ctrl_node_dict = dict_hilctrl["arxml"] if "arxml" in input_file.lower() else dict_hilctrl["dbc"]
    except Exception as e:
        logger.error(f"Error while reading hil ctrl data in dbc mapping -> {e}")

    # create the different dataframes
    node_df = pd.DataFrame(ws1.values)
    node_df = node_df.replace(np.nan, '', regex=True)

    if sheet_name in dict_block_messages.keys():
        blk_pdu_list = dict_block_messages[sheet_name]
        node_df = node_df[~node_df[col_pdu].isin(blk_pdu_list)]
    elif sheet_name in dict_pass_messages.keys():
        pass_pdu_list = dict_pass_messages[sheet_name]
        block_msg_df = node_df[~node_df[col_pdu].isin(pass_pdu_list)]
        blk_pdu_list = list(block_msg_df[col_pdu].unique())
        node_df = node_df[node_df[col_pdu].isin(pass_pdu_list)]

    # get column names
    col_names = list(node_df.iloc[0])
    
    # get column index
    try:
        col_index = [col_names.index(col) for col in col_names]

        [col_Service_name,col_service_instance_name,col_Service_ID,col_Major_version,col_Minor_version,col_Instance_ID,col_MemberType,col_EventGroup,col_EventGroup_ID,col_field_name,col_Member_name,col_Member_ID,col_field_type,col_Parameter,col_ParameterType,col_No_of_Parameters,
        col_para_level1, col_para_level2, col_para_level3, col_para_level4, col_para_level5,
        col_Signal_name,col_No_of_Elements,col_fire_and_forget,col_signal_group,col_pdu,col_pdu_type,
         col_payload_pdu_type,col_payload_pdu,col_selector_field_signal,col_startbit,col_pdu_length_byte,col_signal_length_bit,col_initial_value,col_max_value,
         col_cycle_time,col_texttable,col_enum,col_value_type,col_comment,col_dlc,col_variant,col_value_table,col_endtoendprotection,
         col_Signal_DataType,col_Signal_ByteOrder,col_Factor,col_Offset,col_Minimum,col_Maximum,col_Unit,
         col_ip_address, col_udp_port, col_tcp_port, col_tp_protocol,col_vlan_id, col_vlan_name,col_SDType, col_network_protocol, col_autosar_type] = col_index

    except Exception as e:
        logger.error("Column names in Excel are different from the Script: update the col_index list")
        raise Exception(f"Column names in Excel are different from the Script: -> {e}")
        
    #col_namespace = getColNamespace(node_entity)
        
    # diag filer
    #node_df = node_df[node_df[col_namespace].str.contains(diag_msg_pattern, regex=True) == False]  # Filter out Diag
    #node_df = node_df[node_df[col_namespace].str.contains(stbm_msg_pattern, regex=True) == False]  # Filter out STBM

    #check_df_integrity(node_df)

    #node_df_Heading = node_df.head(1)  # Sort colum Heading(Example ::'Name','group','pdu','Message','Multiplexing/Group','Startbit','Length [Bit]',...)
    #node_df = node_df.iloc[1:]  # Remove colum Heading(Example ::'Name','group','pdu','Message','Multiplexing/Group','Startbit','Length [Bit]',...)
    #sort_by_column_names = sort_column_by(bus_type, input_file)  # get sort by column names based on bus:Eth or CAN
    #node_df = df_sort_asc(node_df, column_names, sort_by_column_names)  # Sort in ascending df by column names
    #node_df = pd.concat([node_df_Heading, node_df], ignore_index=True)  # Restore Heading

    #filter out services provided and consumed by same node
    node_df = filterSomeipData_1(node_df)

    #for removing ipaddress from tcpip stack 
    block_nwt_endpoint_list = []
    blocked_df = pd.DataFrame([])
    try:
        block_nwt_endpoint_list = block_network_endpoint_dict.get(f"{network_name}::{node_name}", []) #list of tuple [(ipaddress, id),]
    except:
        pass
    
    if len(block_nwt_endpoint_list)!=0:
        blocked_df = node_df[node_df[col_ip_address].isin([ip for ip, id in block_nwt_endpoint_list])]
        node_df = node_df[~node_df[col_ip_address].isin([ip for ip, id in block_nwt_endpoint_list])]
        node_df = node_df.reset_index(drop=True)

    #generated_app_endpoint_initialize_list = add_remove_csi_endpoints()  # generate this list before filtering ndoe_df
    generated_app_endpoint_initialize_list = [] #commented canoe ports are fixed in canoe 18

    #filter someip where multiple IP are consuming same service in same VectorSimulationNode
    node_df = filterSomeipData_2(node_df)
    
    num_of_rows = node_df.shape[0]
    
    logger.info(f"Signal count: {num_of_rows - 1}")

def AREthGetValueFunction(datatype):
    """

    Args:
        datatype:

    Returns:

    """
    if is_empty_or_spaces(datatype):
        return 'AREthGetValueDWord'

    if 'string' in datatype:
        return 'AREthGetValueString'
    elif ('float' in datatype) or ('double' in datatype):
        return 'AREthGetValueFloat'
    elif 'int' in datatype:
        if 'uint' in datatype:
            if '64' in datatype:
                return 'AREthGetValueQWord'
            else:
                return 'AREthGetValueDWord'
        else:
            if '64' in datatype:
                return 'AREthGetValueInt64'
            else:
                return 'AREthGetValueLong'
    else:
        return 'AREthGetValueDWord'

def AREthSetValueFunction(datatype):
    """

    Args:
        datatype:

    Returns:

    """
    if is_empty_or_spaces(datatype):
        return 'AREthSetValueDWord'
    if 'string' in datatype:
        return 'AREthSetValueString'
    elif ('float' in datatype) or ('double' in datatype):
        return 'AREthSetValueFloat'
    elif 'int' in datatype:
        if 'uint' in datatype:
            if '64' in datatype:
                return 'AREthSetValueQWord'
            else:
                return 'AREthSetValueDWord'
        else:
            if '64' in datatype:
                return 'AREthSetValueInt64'
            else:
                return 'AREthSetValueLong'
    else:
        return 'AREthSetValueDWord'

def generate_includes():
    """
    generate includes section based on IL usage
    """
    # includes Section
    global var_IL
    node.append('includes')
    node.append('{')
    var_IL=get_canoe_IL(bus_type, input_file)
    node.extend(get_includes(var_IL))
    node.extend(get_eth_rbs_includes(network_name ,node_name))
    node.append('}')
    node.append('')
    logger.info('includes: generated for {0} '.format(node_name))


def generate_variables():
    """
    create node variables ::
        create variables for each Provide Method 
        create variables for each Consume Service
        create variables for each Consumed Method
    """
    pdu_count = 0

    node.append('variables {')
    node.append('    // Define global variables for this CAPL file')
    node.append('    dword gIL_BusContext = 0;')
    node.append('    dword gAfterPreStart = 0;')
    node.append('    char gECU[256] = "{0}";'.format(get_EthNodeName(bus_type,input_file,bus_name,node_name)))
    node.append('    char gETH1BusName[256] = "{0}";'.format(get_EthBusName(bus_type,input_file,bus_name,node_name)))
    part_srv = []  # Service

    ms_timers_notifiers = []  # Timer

    filtered_df = node_df.iloc[1:]
    for network_protocol, network_protocol_df in filtered_df.groupby(col_network_protocol):
        if 'ETH_PDU' in network_protocol.upper():
            part_ac = []
            part_bc = []
            part_ts = []
            part_crc = []
            for payload_pdu_type, payload_pdu_type_df in network_protocol_df.groupby(col_payload_pdu_type):
                if is_empty_or_spaces(payload_pdu_type):
                    for pdu_type, pdu_type_df in filtered_df.groupby(col_pdu_type):
                        if "I-SIGNAL" in pdu_type.upper():
                            for pdu, pdu_df in pdu_type_df.groupby(col_pdu):
                                for signal_name, signal_df in pdu_df.groupby(col_Signal_name):
                                    sig_length = signal_df[col_signal_length_bit].values[0]
                                    ac_cnt = 0  # Alive counter
                                    bc_cnt = 0  # Block counter
                                    ts_cnt = 0  # Time stamp
                                    crc_cnt = 0  # CRC

                                    para_dict = {
                                        "namespace": pdu,
                                        "signal_name": signal_name,
                                        "sig_length": sig_length,
                                        "bus_type": bus_type,
                                        "input_file": input_file
                                    }

                                    break_out = 0
                                    if (break_out != 1):
                                        for sig_pattern, sig_fun in ac_var_dict.items():
                                            if (None != re.search(sig_pattern, signal_name,
                                                                  re.M | re.I)):  # Pattern :: Alive counter
                                                part_ac = part_ac + sig_fun(para_dict)
                                                ac_cnt += 1
                                                break_out = 1
                                                break
                                    if (break_out != 1):
                                        for sig_pattern, sig_fun in bc_var_dict.items():
                                            if (None != re.search(sig_pattern, signal_name,
                                                                  re.M | re.I)):  # Pattern :: Block counter
                                                part_bc = part_bc + sig_fun(para_dict)
                                                bc_cnt += 1
                                                break_out = 1
                                                break
                                    if (break_out != 1):
                                        for sig_pattern, sig_fun in ts_var_dict.items():
                                            if (None != re.search(sig_pattern, signal_name,
                                                                  re.M | re.I)):  # Pattern :: Time stamp
                                                part_ts = part_ts + sig_fun(para_dict)
                                                ts_cnt += 1
                                                break_out = 1
                                                break
                                    if (break_out != 1):
                                        for sig_pattern, sig_fun in crc_var_dict.items():
                                            if (None != re.search(sig_pattern, signal_name, re.M | re.I)):  # Pattern :: CRC
                                                part_crc = part_crc + sig_fun(para_dict)
                                                crc_cnt += 1
                                                break_out = 1
                                                break
            if (len(part_ac) > 0):
                node.extend(part_ac)
                node.append('')
            if (len(part_bc) > 0):
                node.extend(part_bc)
                node.append('')
            if (len(part_ts) > 0):
                node.extend(part_ts)
                node.append('')
            if (len(part_crc) > 0):
                node.extend(part_crc)
                node.append('')
        elif 'ETH_SOMEIP' in network_protocol.upper():
            for srvID, srvID_df in filtered_df.groupby(col_Service_ID):
                #print("Service ID:", srvID)
                for MajVer, MajVer_df in srvID_df.groupby(col_Major_version):
                    #print("MajorVersion:", MajVer)
                    for MinVer, MinVer_df in MajVer_df.groupby(col_Minor_version):
                        #print("MinorVersion:", MinVer)
                        for InstID, InstID_df in MinVer_df.groupby(col_Instance_ID):
                            #print("Instance_ID:", InstID)
                            for SDType, SDType_df in InstID_df.groupby(col_SDType):
                                if "provide"  in SDType.lower():
                                    for MemberType,MemberType_df in SDType_df.groupby(col_MemberType):
                                        if "method" in MemberType.lower():
                                            for Method, Method_df in MemberType_df.groupby(col_Member_name):
                                                node.append("    dword gPm_sif_{0}_{1}_{2}_{3}_{4}; // Provide Service 'sif_{0}_{1}_{2}_{3}' Method '{4}'".format(srvID,MajVer,MinVer,InstID,Method))
                                        elif "field" in MemberType.lower():
                                            tem_field_map = {'setter': 'set_', 'getter': 'get_', 'notifier': ''}
                                            if 'notifier' in MemberType_df[col_field_type].values:
                                                ms_timers_notifiers.append(f'    msTimer Timer_sif_{srvID}_{MajVer}_{MinVer}_{InstID};')
                                            for field, field_df in MemberType_df.groupby(col_field_name):
                                                for field_type, field_type_df in field_df.groupby(col_field_type):
                                                    field_suffix = tem_field_map.get(field_type.lower()) + field
                                                    node.append("    dword gP{5}_sif_{0}_{1}_{2}_{3}_{4}; // Provide Service 'sif_{0}_{1}_{2}_{3}' Field '{4}'".format(srvID, MajVer, MinVer, InstID, field_suffix, field_type.lower()[0]))

                                                    if field_type.lower() == 'notifier':
                                                        node.append("    dword gP{5}_sif_{0}_{1}_{2}_{3}_{4}_Flag = 0; // FLAG for Provide Service 'sif_{0}_{1}_{2}_{3}' Field '{4}'".format(srvID, MajVer, MinVer, InstID, field_suffix, field_type.lower()[0]))
                                        elif 'event' in MemberType.lower():
                                            for event, event_df in MemberType_df.groupby(col_Member_name):
                                                node.append("    dword gPe_sif_{0}_{1}_{2}_{3}_{4}; // Provide Service 'sif_{0}_{1}_{2}_{3}' event '{4}'".format(srvID,MajVer,MinVer,InstID,event))



                                elif "consume"  in SDType.lower():
                                    unique_udp_ports = SDType_df[col_udp_port].unique()
                                    for udp_port in unique_udp_ports:
                                        port_namespace = f'_{udp_port}' if len(unique_udp_ports)>1 else ''
                                        node.append("    dword gCService_sif_{0}_{1}_{2}_{3}{4}; // Consume Service 'sif_{0}_{1}_{2}_{3}{4}'".format(srvID,MajVer,MinVer,InstID,port_namespace))
                                        for MemberType,MemberType_df in SDType_df.groupby(col_MemberType):
                                            if "method" in MemberType.lower():
                                                for Method, Method_df in MemberType_df.groupby(col_Member_name):
                                                    node.append("    dword gCm_sif_{0}_{1}_{2}_{3}_{4}{5}; // Consume Service 'sif_{0}_{1}_{2}_{3}{5}' Method '{4}'".format(srvID,MajVer,MinVer,InstID,Method, port_namespace))
                                            elif "field" in MemberType.lower():
                                                for field, field_df in MemberType_df.groupby(col_field_name):
                                                    for field_type, field_type_df in field_df.groupby(col_field_type):
                                                        if field_type.lower() in ['getter', 'setter']:
                                                            field_suffix = 'set_' + field if 'setter' in field_type.lower() else 'get_' + field
                                                            node.append("    dword gC{5}_sif_{0}_{1}_{2}_{3}_{4}{6}; // Consume Service 'sif_{0}_{1}_{2}_{3}{6}' Field '{4}'".format(srvID, MajVer, MinVer, InstID, field_suffix, field_type.lower()[0], port_namespace))
                                                        elif field_type.lower() == 'notifier':
                                                            node.append("    dword gCn_sif_{0}_{1}_{2}_{3}_{4}{6}; // Consume Service 'sif_{0}_{1}_{2}_{3}{6}' Field '{4}'".format(srvID, MajVer, MinVer, InstID, field, field, port_namespace))
                                            elif 'event' in MemberType.lower():
                                                for event, event_df in MemberType_df.groupby(col_Member_name):
                                                    node.append("    dword gCe_sif_{0}_{1}_{2}_{3}_{4}{5}; // Consume Service 'sif_{0}_{1}_{2}_{3}{5}' event '{4}'".format(srvID,MajVer,MinVer,InstID,event, port_namespace))

    node.extend(ms_timers_notifiers)
    node.append('}')
    #logger.info(f"Variables: generated with {str(message_count)} messages.")
    logger.info(f"Generated Variables for Services and Pdus.")

def generate_InitProvidedServices():
    # Initialse Provided Services
    """
    Initialse Provided Services ::
        Enables Service for each Service provided using "EnableService"
        create object for each Provide Method using "AREthGetProvidedObjectHandle"
        create OnMethod_Request Callback for each provided Method using "AREthRegisterCallback"
    """
    service_count = 0
    method_count  = 0
    node.append('void {0}_InitProvidedServices()'.format(node_name))
    node.append('{')
    node.append('  if (gAfterPreStart == 0) return;')
    node.append('  #if (TOOL_MAJOR_VERSION >= 13)')
    node.append('  if (AREthILControlGetStatus() != 4) return;//4->active[Returns current status of AUTOSAR Eth IL]')
    node.append('  #endif')
    node.append('  write("{0} InitProvidedServices called [%.6f].", TimeNowNS()/1e9);'.format(node_name))

    filtered_df = node_df.iloc[1:]
    for srvID, srvID_df in filtered_df.groupby(col_Service_ID):
        #print("Service ID:", srvID)
        for MajVer, MajVer_df in srvID_df.groupby(col_Major_version):
            #print("MajorVersion:", MajVer)
            for MinVer, MinVer_df in MajVer_df.groupby(col_Minor_version):
                #print("MinorVersion:", MinVer)
                for InstID, InstID_df in MinVer_df.groupby(col_Instance_ID):
                    #print("Instance_ID:", InstID)
                    for SDType, SDType_df in InstID_df.groupby(col_SDType):
                        #print("SD Type:", SDType)
                        autosar_type = SDType_df[col_autosar_type].iloc[0]
                        
                        if 'classic' in autosar_type.lower():
                            service_namespace = 'sif_{0}'.format(srvID)
                        else:
                            service_namespace = SDType_df[col_Service_name].unique()[0]
                        
                        if "provide" in SDType.lower():
                            service_count += 1
                            node.append('  EnableService("{6}::{1}::{2}::{3}", @sysvar::ETH_{4}::{5}::PROVIDED_SERVICES::sif_{0}_{1}_{2}_{3}::CONTROLS::Provide); // Enable Service "sif_{0}_{1}_{2}_{3}" '.format(srvID,MajVer,MinVer,InstID,bus_name,node_name,service_namespace))
                            
                            for MemberType,MemberType_df in SDType_df.groupby(col_MemberType):
                                if "method" in MemberType.lower():
                                    for Method, Method_df in MemberType_df.groupby(col_Member_name):
                                        #print("Method_Name:", Method)
                                        method_count += 1
                                        node.append('  gPm_sif_{0}_{1}_{2}_{3}_{4} = AREthGetProvidedObjectHandle("{5}::{4}::{1}::{2}::{3}");// create provide object of Service "sif_{0}_{1}_{2}_{3}" Method "{4}"'.format(srvID,MajVer,MinVer,InstID,Method,service_namespace))
                                        node.append('  AREthRegisterCallback(gPm_sif_{0}_{1}_{2}_{3}_{4}, "OnMethod_sif_{0}_{1}_{2}_{3}_{4}_Request");// Callback for Service "sif_{0}_{1}_{2}_{3}" Method "{4}"'.format(srvID,MajVer,MinVer,InstID,Method))

                                elif "field" in MemberType.lower():
                                    for field, field_df in MemberType_df.groupby(col_field_name):
                                        for field_type, field_type_df in field_df.groupby(col_field_type):
                                            field_type_name = field_type_df[col_Member_name].iloc[0]
                                            if 'getter' in field_type.lower():
                                                field_suffix = 'get_' + field
                                                node.append('  gPg_sif_{0}_{1}_{2}_{3}_{4} = AREthGetProvidedObjectHandle("{5}::{4}::{1}::{2}::{3}");// create provide object of Service "sif_{0}_{1}_{2}_{3}" Field getter "{4}"'.format(srvID, MajVer, MinVer, InstID, field_suffix, service_namespace))
                                                node.append('  AREthRegisterCallback(gPg_sif_{0}_{1}_{2}_{3}_{4}, "OnGetter_sif_{0}_{1}_{2}_{3}_{4}_Request");// Callback for Service "sif_{0}_{1}_{2}_{3}" Field getter "{4}"'.format(srvID, MajVer, MinVer, InstID, field_suffix))
                                            elif 'setter' in field_type.lower():
                                                field_suffix = 'set_' + field
                                                node.append('  gPs_sif_{0}_{1}_{2}_{3}_{4} = AREthGetProvidedObjectHandle("{5}::{4}::{1}::{2}::{3}");// create provide object of Service "sif_{0}_{1}_{2}_{3}" Field setter "{4}"'.format(srvID, MajVer, MinVer, InstID, field_suffix, service_namespace))
                                                node.append('  AREthRegisterCallback(gPs_sif_{0}_{1}_{2}_{3}_{4}, "OnSetter_sif_{0}_{1}_{2}_{3}_{4}_Request");// Callback for Service "sif_{0}_{1}_{2}_{3}" Field setter "{4}"'.format(srvID, MajVer, MinVer, InstID, field_suffix))

                                            elif 'notifier' in field_type.lower():
                                                node.append('  gPn_sif_{0}_{1}_{2}_{3}_{4} = AREthGetProvidedObjectHandle("{5}::{4}::{1}::{2}::{3}");// create provide object of Service "sif_{0}_{1}_{2}_{3}" Field notifier "{4}"'.format(srvID, MajVer, MinVer, InstID, field, service_namespace))
                                                node.append('  AREthRegisterCallback(gPn_sif_{0}_{1}_{2}_{3}_{4}, "OnPrepareNotification_sif_{0}_{1}_{2}_{3}_{4}");// preprocess notification before sending for Service "sif_{0}_{1}_{2}_{3}" field "{4}"'.format(srvID, MajVer, MinVer, InstID, field))
                                                
                                elif "event" in MemberType.lower():
                                    for event, event_df in MemberType_df.groupby(col_Member_name):
                                        node.append('  gPe_sif_{0}_{1}_{2}_{3}_{4} = AREthGetProvidedObjectHandle("{5}::{4}::{1}::{2}::{3}");// create provide object of Service "sif_{0}_{1}_{2}_{3}" Event "{4}"'.format(srvID,MajVer,MinVer,InstID,event,service_namespace))
                                        node.append('  AREthRegisterCallback(gPe_sif_{0}_{1}_{2}_{3}_{4}, "OnPrepareEvent_sif_{0}_{1}_{2}_{3}_{4}");// preprocess event before sending for Service "sif_{0}_{1}_{2}_{3}" Event "{4}"'.format(srvID,MajVer,MinVer,InstID,event))


                                                
    node.append('}')
    node.append('')
    logger.info(f" InitProvidedServices: generated with {str(service_count)} Services and {str(method_count)} Methods.")
    
def generate_FinishProvidedServices():
    """
    Finish Provided Services:: Removes Provide-object for each method Provided
    """
    service_count = 0
    method_count  = 0
    node.append('void {0}_FinishProvidedServices()'.format(node_name))
    node.append('{')
    node.append('  if (gAfterPreStart == 0) return;')
    node.append('  #if (TOOL_MAJOR_VERSION >= 13)')
    node.append('  if (AREthILControlGetStatus() != 4) return;//4->active[Returns current status of AUTOSAR Eth IL]')
    node.append('  #endif')
    node.append('  write("{0} FinishProvidedServices called [%.6f].", TimeNowNS()/1e9);'.format(node_name))

    filtered_df = node_df.iloc[1:]
    for srvID, srvID_df in filtered_df.groupby(col_Service_ID):
        #print("Service ID:", srvID)
        for MajVer, MajVer_df in srvID_df.groupby(col_Major_version):
            #print("MajorVersion:", MajVer)
            for MinVer, MinVer_df in MajVer_df.groupby(col_Minor_version):
                #print("MinorVersion:", MinVer)
                for InstID, InstID_df in MinVer_df.groupby(col_Instance_ID):
                    #print("Instance_ID:", InstID)
                    for SDType, SDType_df in InstID_df.groupby(col_SDType):
                        #print("SD Type:", SDType)
                        if "provide"  in SDType.lower():
                            service_count += 1
                            for MemberType,MemberType_df in SDType_df.groupby(col_MemberType):
                                if "method" in MemberType.lower():
                                    for Method, Method_df in MemberType_df.groupby(col_Member_name):
                                        #print("Method_Name:", Method)
                                        method_count += 1
                                        node.append('  gPm_sif_{0}_{1}_{2}_{3}_{4} = 0;// remove provide object of Service "sif_{0}_{1}_{2}_{3}" Method "{4}"'.format(srvID,MajVer,MinVer,InstID,Method))

                                elif "field" in MemberType.lower():
                                    for field, field_df in MemberType_df.groupby(col_field_name):
                                        for field_type, field_type_df in field_df.groupby(col_field_type):
                                            field_type_name = field_type_df[col_Member_name].iloc[0]
                                            tem_field_map = {'setter': 'set_', 'getter': 'get_', 'notifier': ''}
                                            field_suffix = tem_field_map.get(field_type.lower()) + field
                                            node.append('  gP{5}_sif_{0}_{1}_{2}_{3}_{4} = 0;// remove provide object of Service "sif_{0}_{1}_{2}_{3}" Field "{4}"'.format(srvID, MajVer, MinVer, InstID, field_suffix, field_type.lower()[0]))
                                            if 'notifier' in field_type.lower():
                                                node.append('  gP{5}_sif_{0}_{1}_{2}_{3}_{4}_Flag = 0;// remove provide object of Service "sif_{0}_{1}_{2}_{3}" Field "{4}"'.format(srvID, MajVer, MinVer, InstID, field_suffix, field_type.lower()[0]))
                                elif 'event' in MemberType.lower():
                                    for event, event_df in MemberType_df.groupby(col_Member_name):
                                        node.append('  gPe_sif_{0}_{1}_{2}_{3}_{4} = 0;// remove provide object of Service "sif_{0}_{1}_{2}_{3}" Event "{4}"'.format(srvID,MajVer,MinVer,InstID,event))

    node.append('}')
    node.append('')
    logger.info(f" FinishProvidedServices: generated with {str(method_count)} Methods.")
    
def generate_InitConsumedServices():
    """
    Initialse Consumed Services
        create object for each Consumed Service using "AREthGetConsumedObjectHandle"
        create OnMethod_Response Callback for each consumed Method using "AREthCreateMethodCall"
        subscribes for each consumed EventGroup using "SubscribeEventGroup"
    """
    global flt_evt_cnt
    method_count = 0
    event_count = 0
    node.append('void {0}_InitConsumedServices()'.format(node_name))
    node.append('{')
    node.append('  if (gAfterPreStart == 0) return;')
    node.append('  #if (TOOL_MAJOR_VERSION >= 13)')
    node.append('  if (AREthILControlGetStatus() != 4) return;//4->active[Returns current status of AUTOSAR Eth IL]')
    node.append('  #endif')
    node.append('  write("{0} InitConsumedServices called [%.6f].", TimeNowNS()/1e9);'.format(node_name))

    filtered_df = node_df.iloc[1:]
    for srvID, srvID_df in filtered_df.groupby(col_Service_ID):
        #print("Service ID:", srvID)
        for MajVer, MajVer_df in srvID_df.groupby(col_Major_version):
            #print("MajorVersion:", MajVer)
            for MinVer, MinVer_df in MajVer_df.groupby(col_Minor_version):
                #print("MinorVersion:", MinVer)
                for InstID, InstID_df in MinVer_df.groupby(col_Instance_ID):
                    #print("Instance_ID:", InstID)
                    for SDType, SDType_df in InstID_df.groupby(col_SDType):
                        #print("SD Type:", SDType)
                        autosar_type = SDType_df[col_autosar_type].iloc[0]
                        
                        if 'classic' in autosar_type.lower():
                            service_namespace = 'sif_{0}'.format(srvID)
                        else:
                            service_namespace = SDType_df[col_Service_name].unique()[0]
                        if "consume" in SDType.lower():
                            unique_udp_ports = SDType_df[col_udp_port].unique()
                            for udp_port in unique_udp_ports:
                                port_namespace = f'_{udp_port}' if len(unique_udp_ports)>1 else ''
                                node.append('  gCService_sif_{0}_{1}_{2}_{3}{5} = AREthGetConsumedObjectHandle("{4}::{1}::{2}::{3}");// create consume object of Service "sif_{0}_{1}_{2}_{3}"'.format(srvID,MajVer,MinVer,InstID,service_namespace, port_namespace))
                                for MemberType,MemberType_df in SDType_df.groupby(col_MemberType):
                                    if "method" in MemberType.lower():
                                        for Method, Method_df in MemberType_df.groupby(col_Member_name):
                                            if "not" not in str(Method_df[col_Member_ID].values[0]):
                                                Method_id = int(Method_df[col_Member_ID].values[0])
                                                #print("Method_Name:", Method)
                                                if ('true' in list(Method_df[col_fire_and_forget])):
                                                    node.append('  gCm_sif_{0}_{1}_{2}_{3}_{4}{5} = AREthGetConsumedObjectHandle("sif_{0}::{4}::{1}::{2}::{3}");// consumed object handle "sif_{0}_{1}_{2}_{3}" Method "{4}" fire and forget'.format(srvID, MajVer, MinVer, InstID, Method, port_namespace))
                                                else:
                                                    node.append('  gCm_sif_{0}_{1}_{2}_{3}_{4}{6} = AREthCreateMethodCall(gCService_sif_{0}_{1}_{2}_{3}{6},{5}, "OnMethod_sif_{0}_{1}_{2}_{3}_{4}{6}_Response");// Callback for Service "sif_{0}_{1}_{2}_{3}{6}" Method "{4}"'.format(srvID,MajVer,MinVer,InstID,Method,Method_id, port_namespace))
                                    elif "event" in MemberType.lower():
                                        for eventgroup_id,eventgroup_id_df in MemberType_df.groupby(col_EventGroup):
                                            event_count += 1
                                            #print("Event group:", eventgroup)
                                            node.append('  SubscribeEventGroup("{7}::{4}::{1}::{2}::{3}", @sysvar::ETH_{5}::{6}::CONSUMED_SERVICES::sif_{0}_{1}_{2}_{3}{8}::EVENTGROUPS::{4}::CONTROLS::Subscribe);'.format(srvID,MajVer,MinVer,InstID,eventgroup_id,bus_name,node_name,service_namespace, port_namespace))
                                            for event, event_df in eventgroup_id_df.groupby(col_Member_name):
                                                node.append('  gCe_sif_{0}_{1}_{2}_{3}_{4}{6} = AREthGetConsumedObjectHandle("{5}::{4}::{1}::{2}::{3}");// create consume object of Service "sif_{0}_{1}_{2}_{3}" Event "{4}"'.format(srvID,MajVer,MinVer,InstID,event,service_namespace, port_namespace))
                                                node.append('  AREthRegisterCallback(gCe_sif_{0}_{1}_{2}_{3}_{4}{5}, "OnEvent_sif_{0}_{1}_{2}_{3}_{4}{5}");// Callback for Service "sif_{0}_{1}_{2}_{3}{5}" Event "{4}"'.format(srvID,MajVer,MinVer,InstID,event, port_namespace))


                                    elif "field" in MemberType.lower():
                                        for field, field_df in MemberType_df.groupby(col_field_name):
                                            for field_type, field_type_df in field_df.groupby(col_field_type):
                                                field_type_name = field_type_df[col_Member_name].values[0]
                                                field_id = int(field_type_df[col_Member_ID].values[0])
                                                if 'getter' in field_type.lower():
                                                    field_suffix = 'get_' + field
                                                    node.append('  gCg_sif_{0}_{1}_{2}_{3}_{4}{6} = AREthCreateMethodCall(gCService_sif_{0}_{1}_{2}_{3}{6},{5}, "OnGetter_sif_{0}_{1}_{2}_{3}_{4}{6}_Response");// Callback for Service "sif_{0}_{1}_{2}_{3}{6}" field getter "{4}"'.format(srvID, MajVer, MinVer, InstID, field_suffix, field_id, port_namespace))
                                                elif 'setter' in field_type.lower():
                                                    field_suffix = 'set_' + field
                                                    node.append('  gCs_sif_{0}_{1}_{2}_{3}_{4}{6} = AREthCreateMethodCall(gCService_sif_{0}_{1}_{2}_{3}{6},{5}, "OnSetter_sif_{0}_{1}_{2}_{3}_{4}{6}_Response");// Callback for Service "sif_{0}_{1}_{2}_{3}{6}" field setter "{4}"'.format(srvID, MajVer, MinVer, InstID, field_suffix, field_id, port_namespace))
                                                elif 'notifier' in field_type.lower():
                                                    node.append('  gCn_sif_{0}_{1}_{2}_{3}_{4}{6} = AREthGetConsumedObjectHandle("{5}::{4}::{1}::{2}::{3}");// consumed field notifier "{4}"'.format(srvID,MajVer,MinVer,InstID,field, service_namespace, port_namespace))
                                                    node.append('  AREthRegisterCallback(gCn_sif_{0}_{1}_{2}_{3}_{4}{5}, "OnNotifier_sif_{0}_{1}_{2}_{3}_{4}{5}_Response");// Callback for Service "sif_{0}_{1}_{2}_{3}{5}" field notifier "{4}"'.format(srvID, MajVer, MinVer, InstID, field, port_namespace))

                                    else:
                                        node.append('  //AREthSDReleaseService(gCService_sif_{0}_{1}_{2}_{3});// The  Service "sif_{0}_{1}_{2}_{3} Discovery message (Find Service) is then no longer sent by the node "'.format(srvID,MajVer,MinVer,InstID))
    node.append('}')
    node.append('')
    logger.info(f" InitConsumedServices: generated with {str(method_count)} Methods and {str(event_count)} Event groups.")

def generate_FinishConsumedServices():
    """
    Finish Consumed Services:: Removes consume-object for each method and Service 
    """
    global flt_evt_cnt
    method_count = 0
    node.append('void {0}_FinishConsumedServices()'.format(node_name))
    node.append('{')
    node.append('  if (gAfterPreStart == 0) return;')
    node.append('  #if (TOOL_MAJOR_VERSION >= 13)')
    node.append('  if (AREthILControlGetStatus() != 4) return;//4->active[Returns current status of AUTOSAR Eth IL]')
    node.append('  #endif')
    node.append('  write("{0} FinishConsumedServices called [%.6f].", TimeNowNS()/1e9);'.format(node_name))

    filtered_df = node_df.iloc[1:]
    for srvID, srvID_df in filtered_df.groupby(col_Service_ID):
        #print("Service ID:", srvID)
        for MajVer, MajVer_df in srvID_df.groupby(col_Major_version):
            #print("MajorVersion:", MajVer)
            for MinVer, MinVer_df in MajVer_df.groupby(col_Minor_version):
                #print("MinorVersion:", MinVer)
                for InstID, InstID_df in MinVer_df.groupby(col_Instance_ID):
                    #print("Instance_ID:", InstID)
                    for SDType, SDType_df in InstID_df.groupby(col_SDType):
                        #print("SD Type:", SDType)
                        if "consume" in SDType.lower():
                            unique_udp_ports = SDType_df[col_udp_port].unique()
                            for udp_port in unique_udp_ports:
                                port_namespace = f'_{udp_port}' if len(unique_udp_ports)>1 else ''
                                for MemberType,MemberType_df in SDType_df.groupby(col_MemberType):
                                    if "method" in MemberType.lower():
                                        node.append('  gCService_sif_{0}_{1}_{2}_{3}{4} = 0;// remove consume object of Service "sif_{0}_{1}_{2}_{3}{4}"'.format(srvID,MajVer,MinVer,InstID, port_namespace))
                                        for Method, Method_df in MemberType_df.groupby(col_Member_name):
                                            if "not" not in str(Method_df[col_Member_ID].values[0]):
                                                method_count += 1
                                                #print("Method_Name:", Method)
                                                node.append('  gCm_sif_{0}_{1}_{2}_{3}_{4}{5} = 0; // remove consume object of "sif_{0}_{1}_{2}_{3}{5}" Method "{4}"'.format(srvID,MajVer,MinVer,InstID,Method, port_namespace))

                                    elif "field" in MemberType.lower():
                                        for field, field_df in MemberType_df.groupby(col_field_name):
                                            for field_type, field_type_df in field_df.groupby(col_field_type):
                                                field_type_name = field_type_df[col_Member_name].iloc[0]
                                                tem_field_map = {'setter': 'set_', 'getter': 'get_', 'notifier': ''}
                                                field_suffix = tem_field_map.get(field_type.lower()) + field
                                                node.append('  gC{5}_sif_{0}_{1}_{2}_{3}_{4}{6} = 0;// remove provide object of Service "sif_{0}_{1}_{2}_{3}{6}" Field "{4}"'.format(srvID, MajVer, MinVer, InstID, field_suffix, field_type.lower()[0], port_namespace))
                                    elif 'event' in MemberType.lower():
                                        for eventgroup_id,eventgroup_id_df in MemberType_df.groupby(col_EventGroup):
                                            for event, event_df in eventgroup_id_df.groupby(col_Member_name):
                                                node.append('  gCe_sif_{0}_{1}_{2}_{3}_{4}{5} = 0;// remove consume object of Service "sif_{0}_{1}_{2}_{3}{5}" Event "{4}"'.format(srvID,MajVer,MinVer,InstID,event, port_namespace))



    node.append('}')
    node.append('')
    logger.info(f" FinishConsumedServices: generated with {str(method_count)} Methods.")

def generate_CallBacks():
    """
    response :: OnMethod for each PROVIDED_SERVICES::METHODS
                Read/Get  :: AREthGetValueDWord for input parameter and AREthGetRequestId for Method request Id
                Write/Set :: AREthSetValueDWord for return parameter and AREthSetReturnCode for Method return code
                
    response :: OnMethod for each CONSUMED_SERVICES::METHODS
                Read/Get  :: AREthGetValueDWord for input parameter , AREthGetRequestId for Method request Id and AREthGetReturnCode for Method return code 
    """
    method_Cm = 0
    method_Pm = 0
    filtered_df = node_df.iloc[1:]
    for srvID, srvID_df in filtered_df.groupby(col_Service_ID):
        #print("Service ID:", srvID)
        for MajVer, MajVer_df in srvID_df.groupby(col_Major_version):
            #print("MajorVersion:", MajVer)
            for MinVer, MinVer_df in MajVer_df.groupby(col_Minor_version):
                #print("MinorVersion:", MinVer)
                for InstID, InstID_df in MinVer_df.groupby(col_Instance_ID):
                    #print("Instance_ID:", InstID)
                    for MemberType,MemberType_df in InstID_df.groupby(col_MemberType):
                        if "method" in MemberType.lower():
                            for Method, Method_df in MemberType_df.groupby(col_Member_name):
                                #print("Method_Name:", Method)
                                for SDType, SDType_df in Method_df.groupby(col_SDType):
                                    #print("SD Type:", SDType)
                                    if "consume" in SDType.lower():
                                        unique_udp_ports = SDType_df[col_udp_port].unique()
                                        for udp_port in unique_udp_ports:
                                            port_namespace = f'_{udp_port}' if len(unique_udp_ports)>1 else ''
                                            if ("not" not in str(Method_df[col_Member_ID].values[0])) and  not('true' in list(Method_df[col_fire_and_forget])):
                                                method_Cm += 1
                                                node.append('void OnMethod_sif_{0}_{1}_{2}_{3}_{4}{5}_Response(dword methodCallHandle, dword messageResponseHandle)// Callback for Consumed Service "sif_{0}_{1}_{2}_{3}{5}" Method "{4}"'.format(srvID,MajVer,MinVer,InstID,Method, port_namespace))
                                                node.append('{')
                                                node.append('  // Service method response in client/consumer')
                                                for ParameterType, ParameterType_df in SDType_df.groupby(col_ParameterType):
                                                    if "return" in ParameterType.lower():
                                                        for Parameter, Parameter_df in ParameterType_df.groupby(col_Parameter):
                                                            #print("consume :: struct :: Parameter:", Parameter)
                                                            No_of_Parameters = str(Parameter_df[col_No_of_Parameters].values[0])

                                                            parameter_namespaces_list = generate_parameter_namespaces(Parameter_df)
                                                            symbolic_namespace_list = [replace_para_namespace_pattern(nmp.split('#')[0]) for nmp in parameter_namespaces_list]
                                                            parameter_namespaces_list = [nmp.replace('2D_ARRAY_', 'index') for nmp in parameter_namespaces_list]

                                                            if not(is_empty_or_spaces(No_of_Parameters)):
                                                                for Parameter_index in range(int(No_of_Parameters)):
                                                                    for k,para_namespace in enumerate(parameter_namespaces_list):
                                                                        para_split = para_namespace.split('#')
                                                                        para_namespace = para_split[0]
                                                                        datatype = para_split[1]
                                                                        areth_get = AREthGetValueFunction(datatype)
                                                                        if 'string' in datatype:
                                                                            if para_namespace:
                                                                                node.append(f'  From_AREthGetValueString_To_sysSetVariableString(messageResponseHandle, "{Parameter}[{Parameter_index}]{symbolic_namespace_list[k]}", sysvar::ETH_{bus_name}::{node_name}::CONSUMED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}{port_namespace}::METHODS::{Method}::{ParameterType}::{Parameter}_{Parameter_index}::{para_namespace});')
                                                                            else:
                                                                                node.append(f'  From_AREthGetValueString_To_sysSetVariableString(messageResponseHandle, "{Parameter}[{Parameter_index}]", sysvar::ETH_{bus_name}::{node_name}::CONSUMED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}{port_namespace}::METHODS::{Method}::{ParameterType}::{Parameter}_{Parameter_index});')
                                                                        else:
                                                                            if para_namespace:
                                                                                node.append(f'  @sysvar::ETH_{bus_name}::{node_name}::CONSUMED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}{port_namespace}::METHODS::{Method}::{ParameterType}::{Parameter}_{Parameter_index}::{para_namespace} = {areth_get}(messageResponseHandle, "{Parameter}[{Parameter_index}]{symbolic_namespace_list[k]}");')

                                                                            else:
                                                                                node.append(f'  @sysvar::ETH_{bus_name}::{node_name}::CONSUMED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}{port_namespace}::METHODS::{Method}::{ParameterType}::{Parameter}_{Parameter_index} = {areth_get}(messageResponseHandle, "{Parameter}[{Parameter_index}]");')

                                                                    node.append('')
                                                            else:
                                                                for k,para_namespace in enumerate(parameter_namespaces_list):
                                                                    para_split = para_namespace.split('#')
                                                                    para_namespace = para_split[0]
                                                                    datatype = para_split[1]
                                                                    areth_get = AREthGetValueFunction(datatype)
                                                                    if 'string' in datatype:
                                                                        if para_namespace:
                                                                            node.append(f'  From_AREthGetValueString_To_sysSetVariableString(messageResponseHandle, "{Parameter}{symbolic_namespace_list[k]}", sysvar::ETH_{bus_name}::{node_name}::CONSUMED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}{port_namespace}::METHODS::{Method}::{ParameterType}::{Parameter}::{para_namespace});')
                                                                        else:
                                                                            node.append(f'  From_AREthGetValueString_To_sysSetVariableString(messageResponseHandle, "{Parameter}", sysvar::ETH_{bus_name}::{node_name}::CONSUMED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}{port_namespace}::METHODS::{Method}::{ParameterType}::{Parameter});')
                                                                    else:
                                                                        if para_namespace:
                                                                            node.append(f'  @sysvar::ETH_{bus_name}::{node_name}::CONSUMED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}{port_namespace}::METHODS::{Method}::{ParameterType}::{Parameter}::{para_namespace} = {areth_get}(messageResponseHandle, "{Parameter}{symbolic_namespace_list[k]}");')
                                                                        else:
                                                                            node.append(f'  @sysvar::ETH_{bus_name}::{node_name}::CONSUMED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}{port_namespace}::METHODS::{Method}::{ParameterType}::{Parameter} = {areth_get}(messageResponseHandle, "{Parameter}");')

                                                                node.append('')


                                                node.append('  @sysvar::ETH_{5}::{6}::CONSUMED_SERVICES::sif_{0}_{1}_{2}_{3}{7}::METHODS::{4}::CONTROLS::RequestID = AREthGetRequestId(messageResponseHandle);'.format(srvID,MajVer,MinVer,InstID,Method,bus_name,node_name, port_namespace))
                                                node.append('  @sysvar::ETH_{5}::{6}::CONSUMED_SERVICES::sif_{0}_{1}_{2}_{3}{7}::METHODS::{4}::CONTROLS::ReturnCode = AREthGetReturnCode(messageResponseHandle);'.format(srvID,MajVer,MinVer,InstID,Method,bus_name,node_name, port_namespace))
                                                node.append('}')
                                                node.append('')
                                    elif "provide" in SDType.lower():
                                        method_Pm += 1

                                        if 'true' in list(SDType_df[col_fire_and_forget]):
                                            node.append('void OnMethod_sif_{0}_{1}_{2}_{3}_{4}_Request(dword methodHandle, dword messageHandle)// Callback for Provide Service "sif_{0}_{1}_{2}_{3}" Method "{4}" fireAndForget'.format(srvID, MajVer, MinVer, InstID, Method))
                                        else:
                                            node.append('void OnMethod_sif_{0}_{1}_{2}_{3}_{4}_Request(dword methodHandle, dword messageHandle, dword messageResponseHandle)// Callback for Provide Service "sif_{0}_{1}_{2}_{3}" Method "{4}"'.format(srvID,MajVer,MinVer,InstID,Method))

                                        node.append('{')
                                        node.append('  // Service method invocation in server/provider')
                                        node.append('  @sysvar::ETH_{5}::{6}::PROVIDED_SERVICES::sif_{0}_{1}_{2}_{3}::METHODS::{4}::CONTROLS::RequestID = AREthGetRequestId(messageHandle);'.format(srvID,MajVer,MinVer,InstID,Method,bus_name,node_name))
                                        node.append('')
 
                                        for ParameterType, ParameterType_df in SDType_df.groupby(col_ParameterType):
                                            if "in" in ParameterType.lower():
                                                for Parameter, Parameter_df in ParameterType_df.groupby(col_Parameter):
                                                    #print("consume :: struct :: Parameter:", Parameter)
                                                    No_of_Parameters = str(Parameter_df[col_No_of_Parameters].values[0])

                                                    parameter_namespaces_list = generate_parameter_namespaces(Parameter_df)
                                                    symbolic_namespace_list = [replace_para_namespace_pattern(nmp.split('#')[0]) for nmp in parameter_namespaces_list]
                                                    parameter_namespaces_list = [nmp.replace('2D_ARRAY_', 'index') for nmp in parameter_namespaces_list]

                                                    if not(is_empty_or_spaces(No_of_Parameters)):
                                                        for Parameter_index in range(int(No_of_Parameters)):
                                                            for k,para_namespace in enumerate(parameter_namespaces_list):
                                                                para_split = para_namespace.split('#')
                                                                para_namespace = para_split[0]
                                                                datatype = para_split[1]
                                                                areth_get = AREthGetValueFunction(datatype)
                                                                if 'string' in datatype:
                                                                    if para_namespace:
                                                                        node.append(f'  From_AREthGetValueString_To_sysSetVariableString(messageHandle, "{Parameter}[{Parameter_index}]{symbolic_namespace_list[k]}", sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::METHODS::{Method}::{ParameterType}::{Parameter}_{Parameter_index}::{para_namespace});')
                                                                    else:
                                                                        node.append(f'  From_AREthGetValueString_To_sysSetVariableString(messageHandle, "{Parameter}[{Parameter_index}]", sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::METHODS::{Method}::{ParameterType}::{Parameter}_{Parameter_index});')
                                                                else:
                                                                    if para_namespace:
                                                                        node.append(f'  @sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::METHODS::{Method}::{ParameterType}::{Parameter}_{Parameter_index}::{para_namespace} = {areth_get}(messageHandle, "{Parameter}[{Parameter_index}]{symbolic_namespace_list[k]}");')

                                                                    else:
                                                                        node.append(f'  @sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::METHODS::{Method}::{ParameterType}::{Parameter}_{Parameter_index} = {areth_get}(messageHandle, "{Parameter}[{Parameter_index}]");')
                                                        node.append('')
                                                    else:
                                                        for k,para_namespace in enumerate(parameter_namespaces_list):
                                                            para_split = para_namespace.split('#')
                                                            para_namespace = para_split[0]
                                                            datatype = para_split[1]
                                                            areth_get = AREthGetValueFunction(datatype)
                                                            if 'string' in datatype:
                                                                if para_namespace:
                                                                    node.append(f'  From_AREthGetValueString_To_sysSetVariableString(messageHandle, "{Parameter}{symbolic_namespace_list[k]}", sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::METHODS::{Method}::{ParameterType}::{Parameter}::{para_namespace});')

                                                                else:
                                                                    node.append(f'  From_AREthGetValueString_To_sysSetVariableString(messageHandle, "{Parameter}", sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::METHODS::{Method}::{ParameterType}::{Parameter});')
                                                            else:
                                                                if para_namespace:
                                                                    node.append(f'  @sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::METHODS::{Method}::{ParameterType}::{Parameter}::{para_namespace} = {areth_get}(messageHandle, "{Parameter}{symbolic_namespace_list[k]}");')

                                                                else:
                                                                    node.append(f'  @sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::METHODS::{Method}::{ParameterType}::{Parameter} = {areth_get}(messageHandle, "{Parameter}");')

                                                        node.append('')
                                                    
 
                                            if "return" in ParameterType.lower():
                                                for Parameter, Parameter_df in ParameterType_df.groupby(col_Parameter):
                                                    #print("consume :: struct :: Parameter:", Parameter)
                                                    No_of_Parameters = str(Parameter_df[col_No_of_Parameters].values[0])

                                                    parameter_namespaces_list = generate_parameter_namespaces(Parameter_df)
                                                    symbolic_namespace_list = [replace_para_namespace_pattern(nmp.split('#')[0]) for nmp in parameter_namespaces_list]
                                                    parameter_namespaces_list = [nmp.replace('2D_ARRAY_', 'index') for nmp in parameter_namespaces_list]

                                                    if not(is_empty_or_spaces(No_of_Parameters)):
                                                        for Parameter_index in range(int(No_of_Parameters)):
                                                            for k,para_namespace in enumerate(parameter_namespaces_list):
                                                                para_split = para_namespace.split('#')
                                                                para_namespace = para_split[0]
                                                                datatype = para_split[1]
                                                                areth_set = AREthSetValueFunction(datatype)
                                                                if 'string' in datatype:
                                                                    if para_namespace:
                                                                        node.append(f'  From_sysGetVariableString_To_AREthSetValueString(messageResponseHandle, "{Parameter}[{Parameter_index}]{symbolic_namespace_list[k]}", sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::METHODS::{Method}::{ParameterType}::{Parameter}_{Parameter_index}::{para_namespace});')
                                                                    else:
                                                                        node.append(f'  From_sysGetVariableString_To_AREthSetValueString(messageResponseHandle, "{Parameter}[{Parameter_index}]", sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::METHODS::{Method}::{ParameterType}::{Parameter}_{Parameter_index});')
                                                                else:
                                                                    if para_namespace:
                                                                        node.append(f'  {areth_set}(messageResponseHandle, "{Parameter}[{Parameter_index}]{symbolic_namespace_list[k]}",@sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::METHODS::{Method}::{ParameterType}::{Parameter}_{Parameter_index}::{para_namespace});')

                                                                    else:
                                                                        node.append(f'  {areth_set}(messageResponseHandle, "{Parameter}[{Parameter_index}]",@sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::METHODS::{Method}::{ParameterType}::{Parameter}_{Parameter_index});')

                                                            node.append('')
                                                    else:
                                                        for k,para_namespace in enumerate(parameter_namespaces_list):
                                                            para_split = para_namespace.split('#')
                                                            para_namespace = para_split[0]
                                                            datatype = para_split[1]
                                                            areth_set = AREthSetValueFunction(datatype)
                                                            if 'string' in datatype:
                                                                if para_namespace:
                                                                    node.append(f'  From_sysGetVariableString_To_AREthSetValueString(messageResponseHandle, "{Parameter}{symbolic_namespace_list[k]}", sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::METHODS::{Method}::{ParameterType}::{Parameter}::{para_namespace});')
                                                                else:
                                                                    node.append(f'  From_sysGetVariableString_To_AREthSetValueString(messageResponseHandle, "{Parameter}", sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::METHODS::{Method}::{ParameterType}::{Parameter});')
                                                            else:
                                                                if para_namespace:
                                                                    node.append(f'  {areth_set}(messageResponseHandle, "{Parameter}{symbolic_namespace_list[k]}",@sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::METHODS::{Method}::{ParameterType}::{Parameter}::{para_namespace});')

                                                                else:
                                                                    node.append(f'  {areth_set}(messageResponseHandle, "{Parameter}",@sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::METHODS::{Method}::{ParameterType}::{Parameter});')

                                                        node.append('')

                                        if 'true' not in list(SDType_df[col_fire_and_forget]):
                                            node.append('  AREthSetReturnCode(messageResponseHandle, @sysvar::ETH_{5}::{6}::PROVIDED_SERVICES::sif_{0}_{1}_{2}_{3}::METHODS::{4}::CONTROLS::ReturnCode);'.format(srvID,MajVer,MinVer,InstID,Method,bus_name,node_name))
                                        node.append('}')
                                        node.append('')

                        elif 'field' in MemberType.lower():
                            for field, field_df in MemberType_df.groupby(col_field_name):

                                for field_type, field_type_df in field_df.groupby(col_field_type):
                                    field_type_name = field_type_df[col_Member_name].iloc[0]
                                    for SDType, SDType_df in field_type_df.groupby(col_SDType):
                                        No_of_Parameters = str(SDType_df[col_No_of_Parameters].values[0])

                                        check_column = [col_No_of_Parameters, col_para_level1, col_para_level2,col_para_level3, col_para_level4, col_para_level5]
                                        isempty_list = []
                                        for col in check_column:
                                            is_empty_col = []
                                            for cell in SDType_df[col]:
                                                if is_empty_or_spaces(cell):
                                                    is_empty_col.append(True)
                                                else:
                                                    is_empty_col.append(False)
                                            if all(is_empty_col):
                                                isempty_list.append(True)
                                            else:
                                                isempty_list.append(False)

                                        if 'consume' in SDType.lower():
                                            unique_udp_ports = SDType_df[col_udp_port].unique()
                                            for udp_port in unique_udp_ports:
                                                port_namespace = f'_{udp_port}' if len(unique_udp_ports)>1 else ''

                                                response_set_value_list = []
                                                if all(isempty_list):
                                                    for signal_name, signal_name_df in SDType_df.groupby(col_Signal_name):
                                                        No_of_elements = str(signal_name_df[col_No_of_Elements].values[0])
                                                        datatype = str(signal_name_df[col_Signal_DataType].values[0])
                                                        areth_get = AREthGetValueFunction(datatype)
                                                        areth_set = AREthSetValueFunction(datatype)
                                                        if 'string' in datatype:
                                                            if is_empty_or_spaces(No_of_elements):
                                                                if is_empty_or_spaces(signal_name):
                                                                    response_set_value_list.append(f'  From_AREthGetValueString_To_sysSetVariableString(messageResponseHandle, "", sysvar::ETH_{bus_name}::{node_name}::CONSUMED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}{port_namespace}::FIELDS::{field}::Value::{field});')
                                                                else:
                                                                    response_set_value_list.append(f'  From_AREthGetValueString_To_sysSetVariableString(messageResponseHandle, "{signal_name}", sysvar::ETH_{bus_name}::{node_name}::CONSUMED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}{port_namespace}::FIELDS::{field}::Value::{signal_name});')

                                                            else:
                                                                for j in range(int(No_of_elements)):
                                                                    if is_empty_or_spaces(signal_name):
                                                                        response_set_value_list.append(f'  From_AREthGetValueString_To_sysSetVariableString(messageResponseHandle, "[{j}]", sysvar::ETH_{bus_name}::{node_name}::CONSUMED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}{port_namespace}::FIELDS::{field}::Value::{field}_{j});')
                                                                    else:
                                                                        response_set_value_list.append(f'  From_AREthGetValueString_To_sysSetVariableString(messageResponseHandle, "{signal_name}[{j}]", sysvar::ETH_{bus_name}::{node_name}::CONSUMED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}{port_namespace}::FIELDS::{field}::Value::{signal_name}_{j});')

                                                        else:
                                                            if is_empty_or_spaces(No_of_elements):
                                                                if is_empty_or_spaces(signal_name):
                                                                    response_set_value_list.append(f'  @sysvar::ETH_{bus_name}::{node_name}::CONSUMED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}{port_namespace}::FIELDS::{field}::Value::{field} = {areth_get}(messageResponseHandle, "");')
                                                                else:
                                                                    response_set_value_list.append(f'  @sysvar::ETH_{bus_name}::{node_name}::CONSUMED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}{port_namespace}::FIELDS::{field}::Value::{signal_name} = {areth_get}(messageResponseHandle, "{signal_name}");')
                                                            else:
                                                                for j in range(int(No_of_elements)):
                                                                    if is_empty_or_spaces(signal_name):
                                                                        response_set_value_list.append(f'  @sysvar::ETH_{bus_name}::{node_name}::CONSUMED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}{port_namespace}::FIELDS::{field}::Value::{field}_{j} = {areth_get}(messageResponseHandle, "[{j}]");')
                                                                    else:
                                                                        response_set_value_list.append(f'  @sysvar::ETH_{bus_name}::{node_name}::CONSUMED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}{port_namespace}::FIELDS::{field}::Value::{signal_name}_{j} = {areth_get}(messageResponseHandle, "{signal_name}[{j}]");')
                                                else:

                                                    parameter_namespaces_list = generate_parameter_namespaces(SDType_df)
                                                    symbolic_namespace_list = [replace_para_namespace_pattern(nmp.split('#')[0]) for nmp in parameter_namespaces_list]
                                                    parameter_namespaces_list = [nmp.replace('2D_ARRAY_', 'index') for nmp in parameter_namespaces_list]

                                                    if not(is_empty_or_spaces(No_of_Parameters)):
                                                        for Parameter_index in range(int(No_of_Parameters)):
                                                            for k,para_namespace in enumerate(parameter_namespaces_list):
                                                                para_split = para_namespace.split('#')
                                                                para_namespace = para_split[0]
                                                                datatype = para_split[1]
                                                                areth_get = AREthGetValueFunction(datatype)
                                                                if 'string' in datatype:
                                                                    if para_namespace:
                                                                        response_set_value_list.append(f'  From_AREthGetValueString_To_sysSetVariableString(messageResponseHandle, "[{Parameter_index}]{symbolic_namespace_list[k]}", sysvar::ETH_{bus_name}::{node_name}::CONSUMED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}{port_namespace}::FIELDS::{field}::Value::{field}_{Parameter_index}::{para_namespace});')
                                                                    else:
                                                                        response_set_value_list.append(f'  From_AREthGetValueString_To_sysSetVariableString(messageResponseHandle, "[{Parameter_index}]", sysvar::ETH_{bus_name}::{node_name}::CONSUMED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}{port_namespace}::FIELDS::{field}::Value::{field}_{Parameter_index});')
                                                                else:
                                                                    if para_namespace:
                                                                        response_set_value_list.append(f'  @sysvar::ETH_{bus_name}::{node_name}::CONSUMED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}{port_namespace}::FIELDS::{field}::Value::{field}_{Parameter_index}::{para_namespace} = {areth_get}(messageResponseHandle, "[{Parameter_index}]{symbolic_namespace_list[k]}");')

                                                                    else:
                                                                        response_set_value_list.append(f'  @sysvar::ETH_{bus_name}::{node_name}::CONSUMED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}{port_namespace}::FIELDS::{field}::Value::{field}_{Parameter_index} = {areth_get}(messageResponseHandle, "[{Parameter_index}]");')
                                                        response_set_value_list.append('')
                                                    else:
                                                        for k,para_namespace in enumerate(parameter_namespaces_list):
                                                            para_split = para_namespace.split('#')
                                                            para_namespace = para_split[0]
                                                            datatype = para_split[1]
                                                            areth_get = AREthGetValueFunction(datatype)
                                                            if 'string' in datatype:
                                                                if para_namespace:
                                                                    response_set_value_list.append(f'  From_AREthGetValueString_To_sysSetVariableString(messageResponseHandle, "{replace_for_direct_namespace(symbolic_namespace_list[k])}", sysvar::ETH_{bus_name}::{node_name}::CONSUMED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}{port_namespace}::FIELDS::{field}::Value::{field}::{para_namespace});')
                                                                else:
                                                                    response_set_value_list.append(f'  From_AREthGetValueString_To_sysSetVariableString(messageResponseHandle, "", sysvar::ETH_{bus_name}::{node_name}::CONSUMED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}{port_namespace}::FIELDS::{field}::Value::{field});')
                                                            else:
                                                                if para_namespace:
                                                                    response_set_value_list.append(f'  @sysvar::ETH_{bus_name}::{node_name}::CONSUMED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}{port_namespace}::FIELDS::{field}::Value::{field}::{para_namespace} = {areth_get}(messageResponseHandle, "{replace_for_direct_namespace(symbolic_namespace_list[k])}");')

                                                                else:
                                                                    response_set_value_list.append(f'  @sysvar::ETH_{bus_name}::{node_name}::CONSUMED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}{port_namespace}::FIELDS::{field}::Value::{field} = {areth_get}(messageResponseHandle, "");')
                                                        response_set_value_list.append('')
                                                if 'getter' in field_type.lower():
                                                    node.append('void OnGetter_sif_{0}_{1}_{2}_{3}_{4}{5}_Response(dword methodCallHandle, dword messageResponseHandle)// Callback for Consumed Service "sif_{0}_{1}_{2}_{3}{5}" field getter "{4}"'.format(srvID, MajVer, MinVer, InstID, 'get_'+field, port_namespace))
                                                    node.append('{')
                                                    node.append('  // Service field getter response in client/consumer')
                                                    
                                                    node.extend(response_set_value_list)
                                                    
                                                    node.append(f'  @sysvar::ETH_{bus_name}::{node_name}::CONSUMED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}{port_namespace}::FIELDS::{field}::GET::RETURN_PARAMETERS::RequestID = AREthGetRequestId(messageResponseHandle);')
                                                    node.append(f'  @sysvar::ETH_{bus_name}::{node_name}::CONSUMED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}{port_namespace}::FIELDS::{field}::GET::RETURN_PARAMETERS::ReturnCode = AREthGetReturnCode(messageResponseHandle);')
                                                    node.append('}')

                                                elif 'setter' in field_type.lower():
                                                    node.append('void OnSetter_sif_{0}_{1}_{2}_{3}_{4}{5}_Response(dword methodCallHandle, dword messageResponseHandle)// Callback for Consumed Service "sif_{0}_{1}_{2}_{3}{5}" field setter "{4}"'.format(srvID, MajVer, MinVer, InstID, 'set_'+field, port_namespace))
                                                    node.append('{')
                                                    node.append('  // Service field setter response in client/consumer')
                                                    node.extend(response_set_value_list)
                                                    node.append(f'  @sysvar::ETH_{bus_name}::{node_name}::CONSUMED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}{port_namespace}::FIELDS::{field}::SET::RETURN_PARAMETERS::RequestID = AREthGetRequestId(messageResponseHandle);')
                                                    node.append(f'  @sysvar::ETH_{bus_name}::{node_name}::CONSUMED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}{port_namespace}::FIELDS::{field}::SET::RETURN_PARAMETERS::ReturnCode = AREthGetReturnCode(messageResponseHandle);')
                                                
                                                    node.append('}')
                                                elif 'notifier' in field_type.lower():
                                                    node.append('void OnNotifier_sif_{0}_{1}_{2}_{3}_{4}{5}_Response(dword methodCallHandle, dword messageResponseHandle)// Callback for Consumed Service "sif_{0}_{1}_{2}_{3}{5}" field notifier "{4}"'.format(srvID, MajVer, MinVer, InstID, field, port_namespace))
                                                    node.append('{')
                                                    node.append('  // Service field notifier response in client/consumer')
                                                    node.extend(response_set_value_list)
                                                    node.append('}')

                                        elif 'provide' in SDType.lower():
                                            request_set_value_list = []
                                            request_setter_value_list = [] #set values from requestHandle to Value of setter
                                            if all(isempty_list):
                                                for signal_name, signal_name_df in SDType_df.groupby(col_Signal_name):
                                                    No_of_elements = str(signal_name_df[col_No_of_Elements].values[0])
                                                    datatype = str(signal_name_df[col_Signal_DataType].values[0])
                                                    areth_get = AREthGetValueFunction(datatype)
                                                    areth_set = AREthSetValueFunction(datatype)
                                                    if 'string' in datatype:
                                                        if is_empty_or_spaces(No_of_elements):
                                                            if is_empty_or_spaces(signal_name):
                                                                request_set_value_list.append(f'  From_sysGetVariableString_To_AREthSetValueString(messageResponseHandle, "", sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::FIELDS::{field}::Value::{field});')
                                                                request_setter_value_list.append(f'  From_AREthGetValueString_To_sysSetVariableString(messageResponseHandle, "", sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::FIELDS::{field}::Value::{field});')
                                                            else:
                                                                request_set_value_list.append(f'  From_sysGetVariableString_To_AREthSetValueString(messageResponseHandle, "{signal_name}", sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::FIELDS::{field}::Value::{signal_name});')
                                                                request_setter_value_list.append(f'  From_AREthGetValueString_To_sysSetVariableString(messageResponseHandle, "{signal_name}", sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::FIELDS::{field}::Value::{signal_name});')

                                                        else:
                                                            for j in range(int(No_of_elements)):
                                                                if is_empty_or_spaces(signal_name):
                                                                    request_set_value_list.append(f'  From_sysGetVariableString_To_AREthSetValueString(messageResponseHandle, "[{j}]", sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::FIELDS::{field}::Value::{field}_{j});')
                                                                    request_setter_value_list.append(f'  From_AREthGetValueString_To_sysSetVariableString(messageResponseHandle, "[{j}]", sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::FIELDS::{field}::Value::{field}_{j});')
                                                                else:
                                                                    request_set_value_list.append(f'  From_sysGetVariableString_To_AREthSetValueString(messageResponseHandle, "{signal_name}[{j}]", sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::FIELDS::{field}::Value::{signal_name}_{j});')
                                                                    request_setter_value_list.append(f'  From_AREthGetValueString_To_sysSetVariableString(messageResponseHandle, "{signal_name}[{j}]", sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::FIELDS::{field}::Value::{signal_name}_{j});')

                                                    else:
                                                        if is_empty_or_spaces(No_of_elements):
                                                            if is_empty_or_spaces(signal_name):
                                                                request_set_value_list.append(f'  {areth_set}(messageResponseHandle, "", @sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::FIELDS::{field}::Value::{field});')
                                                                request_setter_value_list.append(f'  @sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::FIELDS::{field}::Value::{field} = {areth_get}(messageResponseHandle, "");')
                                                            
                                                            else:
                                                                request_set_value_list.append(f'  {areth_set}(messageResponseHandle, "{signal_name}", @sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::FIELDS::{field}::Value::{signal_name});')
                                                                request_setter_value_list.append(f'  @sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::FIELDS::{field}::Value::{signal_name} = {areth_get}(messageResponseHandle, "{signal_name}");')
                                                        else:
                                                            for j in range(int(No_of_elements)):
                                                                if is_empty_or_spaces(signal_name):
                                                                    request_set_value_list.append(f'  {areth_set}(messageResponseHandle, "[{j}]", @sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::FIELDS::{field}::Value::{field}_{j});')
                                                                    request_setter_value_list.append(f'  @sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::FIELDS::{field}::Value::{field}_{j} = {areth_get}(messageResponseHandle, "[{j}]");')
                                                                else:
                                                                    request_set_value_list.append(f'  {areth_set}(messageResponseHandle, "{signal_name}[{j}]", @sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::FIELDS::{field}::Value::{signal_name}_{j});')
                                                                    request_setter_value_list.append(f'  @sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::FIELDS::{field}::Value::{signal_name}_{j} = {areth_get}(messageResponseHandle, "{signal_name}[{j}]");')
                                             
                                            else:

                                                parameter_namespaces_list = generate_parameter_namespaces(SDType_df)
                                                symbolic_namespace_list = [replace_para_namespace_pattern(nmp.split('#')[0]) for nmp in parameter_namespaces_list]
                                                parameter_namespaces_list = [nmp.replace('2D_ARRAY_', 'index') for nmp in parameter_namespaces_list]

                                                if not(is_empty_or_spaces(No_of_Parameters)):
                                                    for Parameter_index in range(int(No_of_Parameters)):
                                                        for k,para_namespace in enumerate(parameter_namespaces_list):
                                                            para_split = para_namespace.split('#')
                                                            para_namespace = para_split[0]
                                                            datatype = para_split[1]
                                                            areth_get = AREthGetValueFunction(datatype)
                                                            areth_set = AREthSetValueFunction(datatype)
                                                            if 'string' in datatype:
                                                                if para_namespace:
                                                                    request_set_value_list.append(f'  From_sysGetVariableString_To_AREthSetValueString(messageResponseHandle, "[{Parameter_index}]{symbolic_namespace_list[k]}", sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::FIELDS::{field}::Value::{field}_{Parameter_index}::{para_namespace});')
                                                                    request_setter_value_list.append(f'  From_AREthGetValueString_To_sysSetVariableString(messageResponseHandle, "[{Parameter_index}]{symbolic_namespace_list[k]}", sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::FIELDS::{field}::Value::{field}_{Parameter_index}::{para_namespace});')
                                                                else:
                                                                    request_set_value_list.append(f'  From_sysGetVariableString_To_AREthSetValueString(messageResponseHandle, "[{Parameter_index}]", sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::FIELDS::{field}::Value::{field}_{Parameter_index});')
                                                                    request_setter_value_list.append(f'  From_AREthGetValueString_To_sysSetVariableString(messageResponseHandle, "[{Parameter_index}]", sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::FIELDS::{field}::Value::{field}_{Parameter_index});')
                                                            else:
                                                                if para_namespace:
                                                                    request_set_value_list.append(f'  {areth_set}(messageResponseHandle, "[{Parameter_index}]{symbolic_namespace_list[k]}", @sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::FIELDS::{field}::Value::{field}_{Parameter_index}::{para_namespace});')
                                                                    request_setter_value_list.append(f'  @sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::FIELDS::{field}::Value::{field}_{Parameter_index}::{para_namespace} = {areth_get}(messageResponseHandle, "[{Parameter_index}]{symbolic_namespace_list[k]}");')
                                                                else:
                                                                    request_set_value_list.append(f'  {areth_set}(messageResponseHandle, "[{Parameter_index}]", @sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::FIELDS::{field}::Value::{field}_{Parameter_index});')
                                                                    request_setter_value_list.append(f'  @sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::FIELDS::{field}::Value::{field}_{Parameter_index} = {areth_get}(messageResponseHandle, "[{Parameter_index}]");')

                                                else:
                                                    for k,para_namespace in enumerate(parameter_namespaces_list):
                                                        para_split = para_namespace.split('#')
                                                        para_namespace = para_split[0]
                                                        datatype = para_split[1]
                                                        areth_get = AREthGetValueFunction(datatype)
                                                        areth_set = AREthSetValueFunction(datatype)
                                                        if 'string' in datatype:
                                                            if para_namespace:
                                                                request_set_value_list.append(f'  From_sysGetVariableString_To_AREthSetValueString(messageResponseHandle, "{replace_for_direct_namespace(symbolic_namespace_list[k])}", sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::FIELDS::{field}::Value::{field}::{para_namespace});')
                                                                request_setter_value_list.append(f'  From_AREthGetValueString_To_sysSetVariableString(messageResponseHandle, "{replace_for_direct_namespace(symbolic_namespace_list[k])}", sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::FIELDS::{field}::Value::{field}::{para_namespace});')

                                                            else:
                                                                request_set_value_list.append(f'  From_sysGetVariableString_To_AREthSetValueString(messageResponseHandle, "", sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::FIELDS::{field}::Value::{field});')
                                                                request_setter_value_list.append(f'  From_AREthGetValueString_To_sysSetVariableString(messageResponseHandle, "", sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::FIELDS::{field}::Value::{field});')
                                                        else:
                                                            if para_namespace:
                                                                request_set_value_list.append(f'  {areth_set}(messageResponseHandle, "{replace_for_direct_namespace(symbolic_namespace_list[k])}", @sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::FIELDS::{field}::Value::{field}::{para_namespace});')
                                                                request_setter_value_list.append(f'  @sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::FIELDS::{field}::Value::{field}::{para_namespace} = {areth_get}(messageResponseHandle, "{replace_for_direct_namespace(symbolic_namespace_list[k])}");')

                                                            else:
                                                                request_set_value_list.append(f'  {areth_set}(messageResponseHandle, "", @sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::FIELDS::{field}::Value::{field});')
                                                                request_setter_value_list.append(f'  @sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::FIELDS::{field}::Value::{field} = {areth_get}(messageResponseHandle, "");')

                                            if 'getter' in field_type.lower():
                                                node.append('void OnGetter_sif_{0}_{1}_{2}_{3}_{4}_Request(dword methodHandle, dword messageHandle, dword messageResponseHandle)// Callback for Provide Service "sif_{0}_{1}_{2}_{3}" field getter "{4}"'.format(srvID,MajVer,MinVer,InstID,'get_'+field))
                                                node.append('{')
                                                node.append('  // Service field getter invocation in server/provider')
                                                node.append(f'  @sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::FIELDS::{field}::GET::IN_PARAMETERS::RequestID = AREthGetRequestId(messageHandle);')
                                                node.append('  // Get the requested value of the field Value and set it to the response handle')
                                                node.extend(request_set_value_list)
                                                node.append('')
                                                if 'notifier' in field_df[col_field_type].values:
                                                    node.append('  gPn_sif_{0}_{1}_{2}_{3}_{4}_Flag = 1; // set flag to 1'.format(srvID, MajVer, MinVer, InstID, field))
                                                node.append(f'  AREthSetReturnCode(messageResponseHandle, @sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::FIELDS::{field}::GET::RETURN_PARAMETERS::ReturnCode);')
                                                
                                                node.append('}')
                                            elif 'setter' in field_type.lower():
                                                node.append('void OnSetter_sif_{0}_{1}_{2}_{3}_{4}_Request(dword methodHandle, dword messageHandle, dword messageResponseHandle)// Callback for Provide Service "sif_{0}_{1}_{2}_{3}" field setter "{4}"'.format(srvID,MajVer,MinVer,InstID,'set_'+field))
                                                node.append('{')
                                                node.append('  // Service field setter invocation in server/provider')
                                                node.append(f'  @sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::FIELDS::{field}::SET::IN_PARAMETERS::RequestID = AREthGetRequestId(messageHandle);')
                                                node.append('')
                                                node.append(f'  // set the request value to Setter Value')
                                                for line in request_setter_value_list:
                                                    node.append(line.replace(f'FIELDS::{field}::', f'FIELDS::{field}::SET::IN_PARAMETERS::'))
                                                node.append('')
                                                node.append('  // This method does not define a response. Simply handle the request here ...')
                                                node.append('  //copy value from setter_value to field_value')
                                                
                                                for line in request_setter_value_list:
                                                    node.append(line.replace(f'::SET::IN_PARAMETERS::', f'::'))

                                                node.append('')
                                                node.append('  // Set the requested value to the field Value')
                                                node.extend(request_set_value_list)
                                                node.append('')
                                                if 'notifier' in field_df[col_field_type].values:
                                                    node.append('      gPn_sif_{0}_{1}_{2}_{3}_{4}_Flag = 1; // set flag to 1'.format(srvID, MajVer, MinVer, InstID, field))
                                                node.append(f'  AREthSetReturnCode(messageResponseHandle, @sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::FIELDS::{field}::SET::RETURN_PARAMETERS::ReturnCode);')

                                                node.append('}')
                                            elif 'notifier' in field_type.lower():
                                                #handle when signal value is change from symbol explorer
                                                signal_value_change_list = ['  '+ln.replace('(messageResponseHandle', '(messageHandle') for ln in request_setter_value_list]
                                                node.append('void OnPrepareNotification_sif_{0}_{1}_{2}_{3}_{4}(dword eventHandle, dword messageHandle)//  preprocess notify before sending for Provide Service "sif_{0}_{1}_{2}_{3}" field "{4}"'.format(srvID,MajVer,MinVer,InstID,field))
                                                node.append('{')
                                                node.append('  if (gPn_sif_{0}_{1}_{2}_{3}_{4}_Flag)'.format(srvID, MajVer, MinVer, InstID, field))
                                                node.append('  {')
                                                node.append('    gPn_sif_{0}_{1}_{2}_{3}_{4}_Flag = 0;'.format(srvID, MajVer, MinVer, InstID, field))
                                                node.append('  }')
                                                node.append('  else {')
                                                node.extend(signal_value_change_list)
                                                node.append('  }')
                                                
                                                node.append('}')
                        elif 'event' in MemberType.lower():
                            # TEMP_FIX--------
                            temp_ip = MemberType_df[col_ip_address].iloc[0]
                            MemberType_df = MemberType_df[MemberType_df[col_ip_address] == temp_ip]
                            has_duplicates = MemberType_df.groupby(col_Member_name)[col_Member_ID].nunique().gt(1).any()
                            # ------------------
                            for eventgroup_id,eventgroup_id_df in MemberType_df.groupby(col_EventGroup):
                                for event, event_df in eventgroup_id_df.groupby(col_Member_name):
                                    if not (has_duplicates):
                                        for SDType, SDType_df in event_df.groupby(col_SDType):
                                            if 'consume' in SDType.lower():
                                                unique_udp_ports = SDType_df[col_udp_port].unique()
                                                for udp_port in unique_udp_ports:
                                                    port_namespace = f'_{udp_port}' if len(unique_udp_ports)>1 else ''
                                                    node.append('void OnEvent_sif_{0}_{1}_{2}_{3}_{4}{5}(dword eventCallHandle, dword messageResponseHandle)// Callback for Consumed Service "sif_{0}_{1}_{2}_{3}{5}" Event "{4}"'.format(srvID,MajVer,MinVer,InstID,event, port_namespace))
                                                    node.append('{')
                                                    node.append('  // onEvent received callback')

                                                    node.append('}')
                                                    node.append('')
                                            elif 'provide' in SDType.lower():
                                                node.append('void OnPrepareEvent_sif_{0}_{1}_{2}_{3}_{4}(dword eventHandle, dword messageHandle)// preprocess event before sending for  Service "sif_{0}_{1}_{2}_{3}" Event "{4}"'.format(srvID,MajVer,MinVer,InstID,event))
                                                node.append('{')
                                                node.append('  // preprocess event before sending')
                                                
                                                try:
                                                    node.extend(get_prepare_someip_event(network_name,node_name,srvID,MajVer,MinVer,InstID,event))
                                                except:
                                                    pass

                                                node.append('}')
                                                node.append('')
                                


    logger.info(f" generate_OnConsumeMethod: generated with {str(method_Cm)} onMethods.")
    logger.info(f" generate_OnProvideMethod: generated with {str(method_Pm)} onMethods.")
    
def generate_OnProvideService():
    """
    EnableService for each PROVIDED_SERVICES
    """
    method_Ps = 0

    filtered_df = node_df.iloc[1:]
    for srvID, srvID_df in filtered_df.groupby(col_Service_ID):
        #print("Service ID:", srvID)
        for MajVer, MajVer_df in srvID_df.groupby(col_Major_version):
            #print("MajorVersion:", MajVer)
            for MinVer, MinVer_df in MajVer_df.groupby(col_Minor_version):
                #print("MinorVersion:", MinVer)
                for InstID, InstID_df in MinVer_df.groupby(col_Instance_ID):
                    #print("Instance_ID:", InstID)
                    for SDType, SDType_df in InstID_df.groupby(col_SDType):
                        #print("SD Type:", SDType)
                        autosar_type = SDType_df[col_autosar_type].iloc[0]
                        if 'classic' in autosar_type.lower():
                            service_namespace = 'sif_{0}'.format(srvID)
                        else:
                            service_namespace = SDType_df[col_Service_name].unique()[0]

                        if "provide" in SDType.lower():
                            method_Ps += 1
                            node.append('on sysvar sysvar::ETH_{4}::{5}::PROVIDED_SERVICES::sif_{0}_{1}_{2}_{3}::CONTROLS::Provide'.format(srvID,MajVer,MinVer,InstID,bus_name,node_name))
                            node.append('{')
                            node.append('  EnableService("{0}::{1}::{2}::{3}", @this);'.format(service_namespace,MajVer,MinVer,InstID))
                            node.append('}')
    node.append('')
    logger.info(f" generate_OnEnableService: generated with {str(method_Ps)} Service .")

def generate_OnSubscribeService():
    """
    SubscribeEventGroup for each CONSUMED_SERVICES::EVENTGROUPS
    """
    event_count = 0

    filtered_df = node_df.iloc[1:]
    for srvID, srvID_df in filtered_df.groupby(col_Service_ID):
        #print("Service ID:", srvID)
        for MajVer, MajVer_df in srvID_df.groupby(col_Major_version):
            #print("MajorVersion:", MajVer)
            for MinVer, MinVer_df in MajVer_df.groupby(col_Minor_version):
                #print("MinorVersion:", MinVer)
                for InstID, InstID_df in MinVer_df.groupby(col_Instance_ID):
                    #print("Instance_ID:", InstID)
                    for SDType, SDType_df in InstID_df.groupby(col_SDType):
                        #print("SD Type:", SDType)
                        if "consume"  in SDType.lower():
                            unique_udp_ports = SDType_df[col_udp_port].unique()
                            for udp_port in unique_udp_ports:
                                port_namespace = f'_{udp_port}' if len(unique_udp_ports)>1 else ''
                                for MemberType,MemberType_df in SDType_df.groupby(col_MemberType):
                                    if "event" in MemberType.lower():
                                        for eventgroup_id,eventgroup_id_df in MemberType_df.groupby(col_EventGroup):
                                            autosar_type = eventgroup_id_df[col_autosar_type].iloc[0]
                                            if 'classic' in autosar_type.lower():
                                                service_namespace = 'sif_{0}'.format(srvID)
                                            else:
                                                service_namespace = eventgroup_id_df[col_Service_name].unique()[0]
                                            event_count += 1
                                            #print("Event group:", eventgroup)
                                            node.append('on sysvar sysvar::ETH_{5}::{6}::CONSUMED_SERVICES::sif_{0}_{1}_{2}_{3}{7}::EVENTGROUPS::{4}::CONTROLS::Subscribe'.format(srvID,MajVer,MinVer,InstID,eventgroup_id,bus_name,node_name, port_namespace))
                                            node.append('{')
                                            node.append('  SubscribeEventGroup("{0}::{5}::{1}::{2}::{3}", @this);'.format(service_namespace,MajVer,MinVer,InstID,node_name,eventgroup_id))
                                            node.append('}')
    node.append('')
    logger.info(f" generate_OnSubscribeService: generated with {str(event_count)} events .")

def generate_OnSendOnce():
    """
    SendServiceMessageOnce for each PROVIDED_SERVICES::EVENTS
    
    request ::AREthCallMethod for each CONSUMED_SERVICES::METHODS
                Write/Set :: {areth_set} for input methodparameter
    """
    global flt_evt_cnt
    method_Cm = 0
    event_count = 0

    filtered_df = node_df.iloc[1:]
    for srvID, srvID_df in filtered_df.groupby(col_Service_ID):
        #print("Service ID:", srvID)
        for MajVer, MajVer_df in srvID_df.groupby(col_Major_version):
            #print("MajorVersion:", MajVer)
            for MinVer, MinVer_df in MajVer_df.groupby(col_Minor_version):
                #print("MinorVersion:", MinVer)
                for InstID, InstID_df in MinVer_df.groupby(col_Instance_ID):
                    #print("Instance_ID:", InstID)
                    for SDType, SDType_df in InstID_df.groupby(col_SDType):
                        #print("SD Type:", SDType)
                        if "consume"  in SDType.lower():
                            unique_udp_ports = SDType_df[col_udp_port].unique()
                            for udp_port in unique_udp_ports:
                                port_namespace = f'_{udp_port}' if len(unique_udp_ports)>1 else ''
                                for MemberType,MemberType_df in SDType_df.groupby(col_MemberType):
                                    if "method" in MemberType.lower():
                                        for Method, Method_df in MemberType_df.groupby(col_Member_name):
                                            if "not" not in str(Method_df[col_Member_ID].values[0]):
                                                method_Cm += 1
                                                #print("Method_Name:", Method)
                                                node.append('on sysvar sysvar::ETH_{5}::{6}::CONSUMED_SERVICES::sif_{0}_{1}_{2}_{3}{7}::METHODS::{4}::CONTROLS::SendOnce'.format(srvID,MajVer,MinVer,InstID,Method,bus_name,node_name, port_namespace))
                                                node.append('{')
                                                node.append('  // Service method call in client/consumer')
                                                node.append('  if (@this == 1)')
                                                node.append('  {')
                                                node.append('    if (gCm_sif_{0}_{1}_{2}_{3}_{4}{5}!= 0)'.format(srvID,MajVer,MinVer,InstID,Method,port_namespace))
                                                node.append('    {')
                                                node.append("      // All request parameters are taken from the client's dialog:")
                                                for ParameterType, ParameterType_df in Method_df.groupby(col_ParameterType):
                                                    if "in" in ParameterType.lower():
                                                        for Parameter, Parameter_df in ParameterType_df.groupby(col_Parameter):
                                                            #print("consume :: struct :: Parameter:", Parameter)
                                                            No_of_Parameters = str(Parameter_df[col_No_of_Parameters].values[0])

                                                            parameter_namespaces_list = generate_parameter_namespaces(Parameter_df)
                                                            symbolic_namespace_list = [replace_para_namespace_pattern(nmp.split('#')[0]) for nmp in parameter_namespaces_list]
                                                            parameter_namespaces_list = [nmp.replace('2D_ARRAY_', 'index') for nmp in parameter_namespaces_list]

                                                            if not (is_empty_or_spaces(No_of_Parameters)):
                                                                for Parameter_index in range(int(No_of_Parameters)):
                                                                    for k,para_namespace in enumerate(parameter_namespaces_list):
                                                                        para_split = para_namespace.split('#')
                                                                        para_namespace = para_split[0]
                                                                        datatype = para_split[1]
                                                                        areth_set = AREthSetValueFunction(datatype)
                                                                        if 'string' in datatype:
                                                                            if para_namespace:
                                                                                node.append(f'      From_sysGetVariableString_To_AREthSetValueString(gCm_sif_{srvID}_{MajVer}_{MinVer}_{InstID}_{Method}{port_namespace},"{Parameter}[{Parameter_index}]{symbolic_namespace_list[k]}",sysvar::ETH_{bus_name}::{node_name}::CONSUMED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}{port_namespace}::METHODS::{Method}::{ParameterType}::{Parameter}_{Parameter_index}::{para_namespace});')
                                                                            else:
                                                                                node.append(f'      From_sysGetVariableString_To_AREthSetValueString(gCm_sif_{srvID}_{MajVer}_{MinVer}_{InstID}_{Method}{port_namespace},"{Parameter}[{Parameter_index}]",sysvar::ETH_{bus_name}::{node_name}::CONSUMED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}{port_namespace}::METHODS::{Method}::{ParameterType}::{Parameter}_{Parameter_index});')

                                                                        else:
                                                                            if para_namespace:
                                                                                node.append(f'      {areth_set}(gCm_sif_{srvID}_{MajVer}_{MinVer}_{InstID}_{Method}{port_namespace},"{Parameter}[{Parameter_index}]{symbolic_namespace_list[k]}",@sysvar::ETH_{bus_name}::{node_name}::CONSUMED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}{port_namespace}::METHODS::{Method}::{ParameterType}::{Parameter}_{Parameter_index}::{para_namespace});')

                                                                            else:
                                                                                node.append(f'      {areth_set}(gCm_sif_{srvID}_{MajVer}_{MinVer}_{InstID}_{Method}{port_namespace},"{Parameter}[{Parameter_index}]",@sysvar::ETH_{bus_name}::{node_name}::CONSUMED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}{port_namespace}::METHODS::{Method}::{ParameterType}::{Parameter}_{Parameter_index});')
                                                                    node.append('')
                                                            else:
                                                                for k,para_namespace in enumerate(parameter_namespaces_list):
                                                                    para_split = para_namespace.split('#')
                                                                    para_namespace = para_split[0]
                                                                    datatype = para_split[1]
                                                                    areth_set = AREthSetValueFunction(datatype)
                                                                    if 'string' in datatype:
                                                                        if para_namespace:
                                                                            node.append(f'      From_sysGetVariableString_To_AREthSetValueString(gCm_sif_{srvID}_{MajVer}_{MinVer}_{InstID}_{Method}{port_namespace},"{Parameter}{symbolic_namespace_list[k]}",sysvar::ETH_{bus_name}::{node_name}::CONSUMED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}{port_namespace}::METHODS::{Method}::{ParameterType}::{Parameter}::{para_namespace});')
                                                                        else:
                                                                            node.append(f'      From_sysGetVariableString_To_AREthSetValueString(gCm_sif_{srvID}_{MajVer}_{MinVer}_{InstID}_{Method}{port_namespace},"{Parameter}",sysvar::ETH_{bus_name}::{node_name}::CONSUMED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}{port_namespace}::METHODS::{Method}::{ParameterType}::{Parameter});')

                                                                    else:
                                                                        if para_namespace:
                                                                            node.append(f'      {areth_set}(gCm_sif_{srvID}_{MajVer}_{MinVer}_{InstID}_{Method}{port_namespace},"{Parameter}{symbolic_namespace_list[k]}",@sysvar::ETH_{bus_name}::{node_name}::CONSUMED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}{port_namespace}::METHODS::{Method}::{ParameterType}::{Parameter}::{para_namespace});')
                                                                        else:
                                                                            node.append(f'      {areth_set}(gCm_sif_{srvID}_{MajVer}_{MinVer}_{InstID}_{Method}{port_namespace},"{Parameter}",@sysvar::ETH_{bus_name}::{node_name}::CONSUMED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}{port_namespace}::METHODS::{Method}::{ParameterType}::{Parameter});')
                                                                node.append('')

                                                node.append('      AREthCallMethod(gCm_sif_{0}_{1}_{2}_{3}_{4}{5});'.format(srvID,MajVer,MinVer,InstID,Method, port_namespace))
                                                node.append('    }')
                                                node.append('    @sysvar::ETH_{5}::{6}::CONSUMED_SERVICES::sif_{0}_{1}_{2}_{3}{7}::METHODS::{4}::CONTROLS::SendOnce=0;'.format(srvID,MajVer,MinVer,InstID,Method,bus_name,node_name, port_namespace))
                                                node.append('  }')
                                                node.append('}')
                                                node.append('')
                                    elif "field" in MemberType.lower():
                                        for field, field_df in MemberType_df.groupby(col_field_name):
                                            for field_type, field_type_df in field_df.groupby(col_field_type):
                                                field_type_name = field_type_df[col_Member_name].iloc[0]
                                                if "getter" in field_type.lower():
                                                    field_suffix = 'get_' + field
                                                    node.append(f'on sysvar sysvar::ETH_{bus_name}::{node_name}::CONSUMED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}{port_namespace}::FIELDS::{field}::GET::CONTROLS::SendOnce')
                                                    node.append('{')
                                                    node.append('  // Service method call in client/consumer')
                                                    node.append('  if (@this == 1)')
                                                    node.append('  {')
                                                    node.append('    if (gCg_sif_{0}_{1}_{2}_{3}_{4}{5}!= 0)'.format(srvID, MajVer, MinVer, InstID, field_suffix, port_namespace))
                                                    node.append('    {')
                                                    node.append("      // All request parameters are taken from the client's dialog:")

                                                    node.append('      AREthCallMethod(gCg_sif_{0}_{1}_{2}_{3}_{4}{5});'.format(srvID, MajVer, MinVer, InstID, field_suffix, port_namespace))
                                                    node.append('    }')
                                                    node.append(f'    @sysvar::ETH_{bus_name}::{node_name}::CONSUMED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}{port_namespace}::FIELDS::{field}::GET::CONTROLS::SendOnce=0;')
                                                    node.append('  }')
                                                    node.append('}')
                                                    node.append('')
                                                elif "setter" in field_type.lower():
                                                    field_suffix = 'set_' + field
                                                    node.append(f'on sysvar sysvar::ETH_{bus_name}::{node_name}::CONSUMED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}{port_namespace}::FIELDS::{field}::SET::CONTROLS::SendOnce')
                                                    node.append('{')
                                                    node.append('  // Service method call in client/consumer')
                                                    node.append('  if (@this == 1)')
                                                    node.append('  {')
                                                    node.append('    if (gCs_sif_{0}_{1}_{2}_{3}_{4}{5}!= 0)'.format(srvID, MajVer, MinVer, InstID, field_suffix, port_namespace))
                                                    node.append('    {')
                                                    node.append("      // All request parameters are taken from the client's dialog:")
                                                    
                                                    #set user values to obj handle before sending request
                                                    No_of_Parameters = str(field_type_df[col_No_of_Parameters].values[0])

                                                    check_column = [col_No_of_Parameters, col_para_level1, col_para_level2,col_para_level3, col_para_level4, col_para_level5]
                                                    isempty_list = []
                                                    for col in check_column:
                                                        is_empty_col = []
                                                        for cell in field_type_df[col]:
                                                            if is_empty_or_spaces(cell):
                                                                is_empty_col.append(True)
                                                            else:
                                                                is_empty_col.append(False)
                                                        if all(is_empty_col):
                                                            isempty_list.append(True)
                                                        else:
                                                            isempty_list.append(False)
                                                    
                                                    request_set_value_list = []
                                                    
                                                    if all(isempty_list):
                                                        for signal_name, signal_name_df in field_type_df.groupby(col_Signal_name):
                                                            No_of_elements = str(signal_name_df[col_No_of_Elements].values[0])

                                                            datatype = str(signal_name_df[col_Signal_DataType].values[0])
                                                            areth_get = AREthGetValueFunction(datatype)
                                                            areth_set = AREthSetValueFunction(datatype)
                                                            if 'string' in datatype:
                                                                if is_empty_or_spaces(No_of_elements):
                                                                    if is_empty_or_spaces(signal_name):
                                                                        request_set_value_list.append(f'      From_sysGetVariableString_To_AREthSetValueString(gCs_sif_{srvID}_{MajVer}_{MinVer}_{InstID}_{field_suffix}{port_namespace}, "", sysvar::ETH_{bus_name}::{node_name}::CONSUMED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}{port_namespace}::FIELDS::{field}::SET::IN_PARAMETERS::Value::{field});')
                                                                    else:
                                                                        request_set_value_list.append(f'      From_sysGetVariableString_To_AREthSetValueString(gCs_sif_{srvID}_{MajVer}_{MinVer}_{InstID}_{field_suffix}{port_namespace}, "{signal_name}", sysvar::ETH_{bus_name}::{node_name}::CONSUMED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}{port_namespace}::FIELDS::{field}::SET::IN_PARAMETERS::Value::{signal_name});')
                                                                else:
                                                                    for j in range(int(No_of_elements)):
                                                                        if is_empty_or_spaces(signal_name):
                                                                            request_set_value_list.append(f'      From_sysGetVariableString_To_AREthSetValueString(gCs_sif_{srvID}_{MajVer}_{MinVer}_{InstID}_{field_suffix}{port_namespace}, "[{j}]", sysvar::ETH_{bus_name}::{node_name}::CONSUMED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}{port_namespace}::FIELDS::{field}::SET::IN_PARAMETERS::Value::{field}_{j});')
                                                                        else:
                                                                            request_set_value_list.append(f'      From_sysGetVariableString_To_AREthSetValueString(gCs_sif_{srvID}_{MajVer}_{MinVer}_{InstID}_{field_suffix}{port_namespace}, "{signal_name}[{j}]", sysvar::ETH_{bus_name}::{node_name}::CONSUMED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}{port_namespace}::FIELDS::{field}::SET::IN_PARAMETERS::Value::{signal_name}_{j});')

                                                            else:
                                                                if is_empty_or_spaces(No_of_elements):
                                                                    if is_empty_or_spaces(signal_name):
                                                                        request_set_value_list.append(f'      {areth_set}(gCs_sif_{srvID}_{MajVer}_{MinVer}_{InstID}_{field_suffix}{port_namespace}, "", @sysvar::ETH_{bus_name}::{node_name}::CONSUMED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}{port_namespace}::FIELDS::{field}::SET::IN_PARAMETERS::Value::{field});')
                                                                    else:
                                                                        request_set_value_list.append(f'      {areth_set}(gCs_sif_{srvID}_{MajVer}_{MinVer}_{InstID}_{field_suffix}{port_namespace}, "{signal_name}", @sysvar::ETH_{bus_name}::{node_name}::CONSUMED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}{port_namespace}::FIELDS::{field}::SET::IN_PARAMETERS::Value::{signal_name});')
                                                                else:
                                                                    for j in range(int(No_of_elements)):
                                                                        if is_empty_or_spaces(signal_name):
                                                                            request_set_value_list.append(f'      {areth_set}(gCs_sif_{srvID}_{MajVer}_{MinVer}_{InstID}_{field_suffix}{port_namespace}, "[{j}]", @sysvar::ETH_{bus_name}::{node_name}::CONSUMED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}{port_namespace}::FIELDS::{field}::SET::IN_PARAMETERS::Value::{field}_{j});')
                                                                        else:
                                                                            request_set_value_list.append(f'      {areth_set}(gCs_sif_{srvID}_{MajVer}_{MinVer}_{InstID}_{field_suffix}{port_namespace}, "{signal_name}[{j}]", @sysvar::ETH_{bus_name}::{node_name}::CONSUMED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}{port_namespace}::FIELDS::{field}::SET::IN_PARAMETERS::Value::{signal_name}_{j});')

                                                    else:

                                                        parameter_namespaces_list = generate_parameter_namespaces(field_type_df)
                                                        symbolic_namespace_list = [replace_para_namespace_pattern(nmp.split('#')[0]) for nmp in parameter_namespaces_list]
                                                        parameter_namespaces_list = [nmp.replace('2D_ARRAY_', 'index') for nmp in parameter_namespaces_list]

                                                        if not(is_empty_or_spaces(No_of_Parameters)):
                                                            for Parameter_index in range(int(No_of_Parameters)):
                                                                for k,para_namespace in enumerate(parameter_namespaces_list):
                                                                    para_split = para_namespace.split('#')
                                                                    para_namespace = para_split[0]
                                                                    datatype = para_split[1]
                                                                    areth_set = AREthSetValueFunction(datatype)
                                                                    if 'string' in datatype:
                                                                        if para_namespace:
                                                                            request_set_value_list.append(f'      From_sysGetVariableString_To_AREthSetValueString(gCs_sif_{srvID}_{MajVer}_{MinVer}_{InstID}_{field_suffix}{port_namespace}, "[{Parameter_index}]{symbolic_namespace_list[k]}", sysvar::ETH_{bus_name}::{node_name}::CONSUMED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}{port_namespace}::FIELDS::{field}::SET::IN_PARAMETERS::Value::{field}_{Parameter_index}::{para_namespace});')
                                                                        else:
                                                                            request_set_value_list.append(f'      From_sysGetVariableString_To_AREthSetValueString(gCs_sif_{srvID}_{MajVer}_{MinVer}_{InstID}_{field_suffix}{port_namespace}, "[{Parameter_index}]", sysvar::ETH_{bus_name}::{node_name}::CONSUMED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}{port_namespace}::FIELDS::{field}::SET::IN_PARAMETERS::Value::{field}_{Parameter_index});')

                                                                    else:
                                                                        if para_namespace:
                                                                            request_set_value_list.append(f'      {areth_set}(gCs_sif_{srvID}_{MajVer}_{MinVer}_{InstID}_{field_suffix}{port_namespace}, "[{Parameter_index}]{symbolic_namespace_list[k]}", @sysvar::ETH_{bus_name}::{node_name}::CONSUMED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}{port_namespace}::FIELDS::{field}::SET::IN_PARAMETERS::Value::{field}_{Parameter_index}::{para_namespace});')
                                                                        else:
                                                                            request_set_value_list.append(f'      {areth_set}(gCs_sif_{srvID}_{MajVer}_{MinVer}_{InstID}_{field_suffix}{port_namespace}, "[{Parameter_index}]", @sysvar::ETH_{bus_name}::{node_name}::CONSUMED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}{port_namespace}::FIELDS::{field}::SET::IN_PARAMETERS::Value::{field}_{Parameter_index});')
                                                            request_set_value_list.append('')

                                                        else:
                                                            for k,para_namespace in enumerate(parameter_namespaces_list):
                                                                para_split = para_namespace.split('#')
                                                                para_namespace = para_split[0]
                                                                datatype = para_split[1]
                                                                areth_set = AREthSetValueFunction(datatype)
                                                                if 'string' in datatype:
                                                                    if para_namespace:
                                                                        request_set_value_list.append(f'      From_sysGetVariableString_To_AREthSetValueString(gCs_sif_{srvID}_{MajVer}_{MinVer}_{InstID}_{field_suffix}{port_namespace}, "{replace_for_direct_namespace(symbolic_namespace_list[k])}", sysvar::ETH_{bus_name}::{node_name}::CONSUMED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}{port_namespace}::FIELDS::{field}::SET::IN_PARAMETERS::Value::{field}::{para_namespace});')
                                                                    else:
                                                                        request_set_value_list.append(f'      From_sysGetVariableString_To_AREthSetValueString(gCs_sif_{srvID}_{MajVer}_{MinVer}_{InstID}_{field_suffix}{port_namespace}, "", sysvar::ETH_{bus_name}::{node_name}::CONSUMED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}{port_namespace}::FIELDS::{field}::SET::IN_PARAMETERS::Value::{field});')

                                                                else:
                                                                    if para_namespace:
                                                                        request_set_value_list.append(f'      {areth_set}(gCs_sif_{srvID}_{MajVer}_{MinVer}_{InstID}_{field_suffix}{port_namespace}, "{replace_for_direct_namespace(symbolic_namespace_list[k])}", @sysvar::ETH_{bus_name}::{node_name}::CONSUMED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}{port_namespace}::FIELDS::{field}::SET::IN_PARAMETERS::Value::{field}::{para_namespace});')
                                                                    else:
                                                                        request_set_value_list.append(f'      {areth_set}(gCs_sif_{srvID}_{MajVer}_{MinVer}_{InstID}_{field_suffix}{port_namespace}, "", @sysvar::ETH_{bus_name}::{node_name}::CONSUMED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}{port_namespace}::FIELDS::{field}::SET::IN_PARAMETERS::Value::{field});')
                                                            request_set_value_list.append('')        

                                                    node.extend(request_set_value_list)
                                                    node.append('')
                                                    node.append(f'      AREthCallMethod(gCs_sif_{srvID}_{MajVer}_{MinVer}_{InstID}_{field_suffix}{port_namespace});')
                                                    node.append('    }')
                                                    node.append(f'    @sysvar::ETH_{bus_name}::{node_name}::CONSUMED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}{port_namespace}::FIELDS::{field}::SET::CONTROLS::SendOnce=0;')
                                                    node.append('  }')
                                                    node.append('}')
                                                    node.append('')

                                            
                        elif "provide"  in SDType.lower():
                            for MemberType,MemberType_df in SDType_df.groupby(col_MemberType):
                                if "event" in MemberType.lower():
                                    for eventgroup_id,eventgroup_id_df in MemberType_df.groupby(col_EventGroup):
                                        for event,event_df in eventgroup_id_df.groupby(col_Member_name):
                                            autosar_type = event_df[col_autosar_type].iloc[0]
                                            if 'classic' in autosar_type.lower():
                                                service_namespace = 'sif_{0}'.format(srvID)
                                            else:
                                                service_namespace = event_df[col_Service_name].unique()[0]
                                            event_count += 1
                                            #print("Event group:", eventgroup)
                                            node.append('on sysvar sysvar::ETH_{6}::{7}::PROVIDED_SERVICES::sif_{0}_{1}_{2}_{3}::EVENTGROUPS::{4}::EVENTS::{5}::CONTROLS::SendOnce'.format(srvID,MajVer,MinVer,InstID,eventgroup_id,event,bus_name,node_name))
                                            node.append('{')
                                            node.append('  if (@this == 1)')
                                            node.append('  {')
                                            node.append('    SendServiceMessageOnce("{0}::{4}::{1}::{2}::{3}");'.format(service_namespace,MajVer,MinVer,InstID,event))
                                            node.append('    @sysvar::ETH_{6}::{7}::PROVIDED_SERVICES::sif_{0}_{1}_{2}_{3}::EVENTGROUPS::{4}::EVENTS::{5}::CONTROLS::SendOnce=0;'.format(srvID,MajVer,MinVer,InstID,eventgroup_id,event,bus_name,node_name))
                                            node.append('  }')
                                            node.append('}')
                                            node.append('')
                                elif "field" in MemberType.lower():
                                    for field, field_df in MemberType_df.groupby(col_field_name):
                                        for field_type, field_type_df in field_df.groupby(col_field_type):
                                            field_type_name = field_type_df[col_Member_name].iloc[0]
                                            autosar_type = field_type_df[col_autosar_type].iloc[0]
                                            if 'classic' in autosar_type.lower():
                                                service_namespace = 'sif_{0}'.format(srvID)
                                            else:
                                                service_namespace = field_type_df[col_Service_name].unique()[0]

                                            if "notifier" in field_type.lower():
                                                
                                                No_of_Parameters = str(field_type_df[col_No_of_Parameters].values[0])

                                                check_column = [col_No_of_Parameters, col_para_level1, col_para_level2,col_para_level3, col_para_level4, col_para_level5]
                                                isempty_list = []
                                                for col in check_column:
                                                    is_empty_col = []
                                                    for cell in field_type_df[col]:
                                                        if is_empty_or_spaces(cell):
                                                            is_empty_col.append(True)
                                                        else:
                                                            is_empty_col.append(False)
                                                    if all(is_empty_col):
                                                        isempty_list.append(True)
                                                    else:
                                                        isempty_list.append(False)
                                                
                                                request_set_value_list = []
                                                
                                                if all(isempty_list):
                                                    for signal_name, signal_name_df in field_type_df.groupby(col_Signal_name):
                                                        No_of_elements = str(signal_name_df[col_No_of_Elements].values[0])
                                                        datatype = str(signal_name_df[col_Signal_DataType].values[0])
                                                        areth_set = AREthSetValueFunction(datatype)
                                                        if 'string' in datatype:
                                                            if is_empty_or_spaces(No_of_elements):
                                                                if is_empty_or_spaces(signal_name):
                                                                    request_set_value_list.append(f'      From_sysGetVariableString_To_AREthSetValueString(gPn_sif_{srvID}_{MajVer}_{MinVer}_{InstID}_{field}, "", sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::FIELDS::{field}::Value::{field});')

                                                                else:
                                                                    request_set_value_list.append(f'      From_sysGetVariableString_To_AREthSetValueString(gPn_sif_{srvID}_{MajVer}_{MinVer}_{InstID}_{field}, "{signal_name}", sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::FIELDS::{field}::Value::{signal_name});')
                                                            else:
                                                                for j in range(int(No_of_elements)):
                                                                    if is_empty_or_spaces(signal_name):
                                                                        request_set_value_list.append(f'      From_sysGetVariableString_To_AREthSetValueString(gPn_sif_{srvID}_{MajVer}_{MinVer}_{InstID}_{field}, "[{j}]", sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::FIELDS::{field}::Value::{field}_{j});')
                                                                    else:
                                                                        request_set_value_list.append(f'      From_sysGetVariableString_To_AREthSetValueString(gPn_sif_{srvID}_{MajVer}_{MinVer}_{InstID}_{field}, "{signal_name}[{j}]", sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::FIELDS::{field}::Value::{signal_name}_{j});')

                                                        else:
                                                            if is_empty_or_spaces(No_of_elements):
                                                                if is_empty_or_spaces(signal_name):
                                                                    request_set_value_list.append(f'      {areth_set}(gPn_sif_{srvID}_{MajVer}_{MinVer}_{InstID}_{field}, "", @sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::FIELDS::{field}::Value::{field});')

                                                                else:
                                                                    request_set_value_list.append(f'      {areth_set}(gPn_sif_{srvID}_{MajVer}_{MinVer}_{InstID}_{field}, "{signal_name}", @sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::FIELDS::{field}::Value::{signal_name});')
                                                            else:
                                                                for j in range(int(No_of_elements)):
                                                                    if is_empty_or_spaces(signal_name):
                                                                        request_set_value_list.append(f'      {areth_set}(gPn_sif_{srvID}_{MajVer}_{MinVer}_{InstID}_{field}, "[{j}]", @sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::FIELDS::{field}::Value::{field}_{j});')
                                                                    else:
                                                                        request_set_value_list.append(f'      {areth_set}(gPn_sif_{srvID}_{MajVer}_{MinVer}_{InstID}_{field}, "{signal_name}[{j}]", @sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::FIELDS::{field}::Value::{signal_name}_{j});')
    
                                                else:
                                                    parameter_namespaces_list = generate_parameter_namespaces(field_type_df)
                                                    symbolic_namespace_list = [replace_para_namespace_pattern(nmp.split('#')[0]) for nmp in parameter_namespaces_list]
                                                    parameter_namespaces_list = [nmp.replace('2D_ARRAY_', 'index') for nmp in parameter_namespaces_list]

                                                    if not(is_empty_or_spaces(No_of_Parameters)):
                                                        for Parameter_index in range(int(No_of_Parameters)):
                                                            for k,para_namespace in enumerate(parameter_namespaces_list):
                                                                para_split = para_namespace.split('#')
                                                                para_namespace = para_split[0]
                                                                datatype = para_split[1]
                                                                areth_set = AREthSetValueFunction(datatype)
                                                                if 'string' in datatype:
                                                                    if para_namespace:
                                                                        request_set_value_list.append(f'      From_sysGetVariableString_To_AREthSetValueString(gPn_sif_{srvID}_{MajVer}_{MinVer}_{InstID}_{field}, "[{Parameter_index}]{symbolic_namespace_list[k]}", sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::FIELDS::{field}::Value::{field}_{Parameter_index}::{para_namespace});')
                                                                    else:
                                                                        request_set_value_list.append(f'      From_sysGetVariableString_To_AREthSetValueString(gPn_sif_{srvID}_{MajVer}_{MinVer}_{InstID}_{field}, "[{Parameter_index}]", sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::FIELDS::{field}::Value::{field}_{Parameter_index});')

                                                                else:
                                                                    if para_namespace:
                                                                        request_set_value_list.append(f'      {areth_set}(gPn_sif_{srvID}_{MajVer}_{MinVer}_{InstID}_{field}, "[{Parameter_index}]{symbolic_namespace_list[k]}", @sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::FIELDS::{field}::Value::{field}_{Parameter_index}::{para_namespace});')

                                                                    else:
                                                                        request_set_value_list.append(f'      {areth_set}(gPn_sif_{srvID}_{MajVer}_{MinVer}_{InstID}_{field}, "[{Parameter_index}]", @sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::FIELDS::{field}::Value::{field}_{Parameter_index});')

                                                    else:
                                                        for k,para_namespace in enumerate(parameter_namespaces_list):
                                                            para_split = para_namespace.split('#')
                                                            para_namespace = para_split[0]
                                                            datatype = para_split[1]
                                                            areth_set = AREthSetValueFunction(datatype)
                                                            if 'string' in datatype:
                                                                if para_namespace:
                                                                    request_set_value_list.append(f'      From_sysGetVariableString_To_AREthSetValueString(gPn_sif_{srvID}_{MajVer}_{MinVer}_{InstID}_{field}, "{replace_for_direct_namespace(symbolic_namespace_list[k])}", sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::FIELDS::{field}::Value::{field}::{para_namespace});')
                                                                else:
                                                                    request_set_value_list.append(f'      From_sysGetVariableString_To_AREthSetValueString(gPn_sif_{srvID}_{MajVer}_{MinVer}_{InstID}_{field}, "", sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::FIELDS::{field}::Value::{field});')

                                                            else:
                                                                if para_namespace:
                                                                    request_set_value_list.append(f'      {areth_set}(gPn_sif_{srvID}_{MajVer}_{MinVer}_{InstID}_{field}, "{replace_for_direct_namespace(symbolic_namespace_list[k])}", @sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::FIELDS::{field}::Value::{field}::{para_namespace});')
                                                                else:
                                                                    request_set_value_list.append(f'      {areth_set}(gPn_sif_{srvID}_{MajVer}_{MinVer}_{InstID}_{field}, "", @sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::FIELDS::{field}::Value::{field});')

                                                node.append(f'on sysvar sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::FIELDS::{field}::NOTIFY::CONTROLS::SendOnce')
                                                node.append('{')
                                                node.append('  // set value of field content')
                                                node.append('  if (@this == 1)')
                                                node.append('  {')
                                                node.append('    if (gPn_sif_{0}_{1}_{2}_{3}_{4}!= 0)'.format(srvID, MajVer, MinVer, InstID, field))
                                                node.append('    {')
                                                node.extend(request_set_value_list)
                                                node.append('')
                                                node.append('      gPn_sif_{0}_{1}_{2}_{3}_{4}_Flag = 1; // set flag to 1'.format(srvID, MajVer, MinVer, InstID, field))
                                                node.append('      AREthCommitField(gPn_sif_{0}_{1}_{2}_{3}_{4});'.format(srvID, MajVer, MinVer, InstID, field))
                                                node.append(f'      @sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::FIELDS::{field}::NOTIFY::CONTROLS::SendOnce=0;')
                                                node.append('    }')
                                                node.append('  }')
                                                node.append('}')

                                                node.append('')

    logger.info(f" generate_OnSendOnce: generated with {str(method_Cm)} Methods and {str(event_count)} Events.")

def generate_OnTimers():
    """
    
    """
    mstimer_count=0
    filtered_df = node_df.iloc[1:]
    for SDType, SDType_df in filtered_df.groupby(col_SDType):
        if 'provide' in SDType.lower():
            for srvID, srvID_df in SDType_df.groupby(col_Service_ID):#print("Service ID:", srvID)
                for MajVer, MajVer_df in srvID_df.groupby(col_Major_version):
                    for MinVer, MinVer_df in MajVer_df.groupby(col_Minor_version):
                        for InstID, InstID_df in MinVer_df.groupby(col_Instance_ID):
                            for MemberType,MemberType_df in InstID_df.groupby(col_MemberType):
                                if "field" in MemberType.lower():
                                    if 'notifier' in MemberType_df[col_field_type].values:
                                        mstimer_count+=1
                                        node.append('')
                                        node.append(f'on timer Timer_sif_{srvID}_{MajVer}_{MinVer}_{InstID}')
                                        node.append('{')
                                        for field, field_df in MemberType_df.groupby(col_field_name):
                                            for field_type, field_type_df in field_df.groupby(col_field_type):
                                                if 'notifier' in field_type.lower():
                                                    node.append(f'  @sysvar::ETH_{bus_name}::{node_name}::PROVIDED_SERVICES::sif_{srvID}_{MajVer}_{MinVer}_{InstID}::FIELDS::{field}::NOTIFY::CONTROLS::SendOnce=1;')
                                        node.append('}')
    logger.info(f" generate_OnTimers: generated {str(mstimer_count)} timers.")


def generate_onStartSim(): 
    node.append('')
    node.append('void StartSim(dword busContext, int busActive, int mode)')
    node.append('{')
    node.append('  // Called when ... due to ...')
    node.append('  //   mode == 1: ILControlResume')
    node.append('  //   mode == 2: ILControlStart')
    node.append('  //   mode == 3: ILControlSimulationOn')
    node.append('')
    node.append('  //write("[%.6f {0}] StartSim(0x%X, %d, %d)", TimeNowNS()/1e9, busContext, busActive, mode);'.format(node_name))
    node.append('  if (((mode == 1) || (mode == 2)) && (busContext == gILETH1_BusContext)) AREthEstablishTCPConnection(); // Establish TCP Connection')
    node.append('  if (((mode == 1) || (mode == 2)) && (busContext == gILETH1_BusContext)) InitProvidedServices();')
    node.append('  if (((mode == 1) || (mode == 2)) && (busContext == gILETH1_BusContext)) InitConsumedServices();')
    node.append('')
    node.append('}')

def generate_onStopSim(): 
    node.append('')
    node.append('void StopSim(dword busContext, int busActive, int mode)')
    node.append('{')
    node.append('  // Called when ... due to ...')
    node.append('  //   mode == 1: ILControlWait')
    node.append('  //   mode == 2: ILControlStop')
    node.append('  //   mode == 3: ILControlSimulationOff')
    node.append('')
    node.append('  //write("[%.6f {0}] StopSim(0x%X, %d, %d)", TimeNowNS()/1e9, busContext, busActive, mode);'.format(node_name))
    node.append('  if (((mode == 1) || (mode == 2)) && (busContext == gILETH1_BusContext)) FinishProvidedServices();')
    node.append('  if (((mode == 1) || (mode == 2)) && (busContext == gILETH1_BusContext)) FinishConsumedServices();')
    node.append('  if (mode == 3) && (busContext == gILETH1_BusContext)) AREthCloseEstablishedTCPConnection (); // Close TCP Connection')
    node.append('')
    node.append('}')

def generate_onprestart():
    node.append('')
    node.append('on preStart {')
    node.append('  long result;')
    node.append('  gIL_BusContext = GetBusNameContext(gETH1BusName);')
    node.append('  if (gIL_BusContext == 0)')
    node.append('  {')
    node.append('    write("{0}:: ERROR: cannot determine bus context of Ethernet bus %s", gETH1BusName);'.format(node_name))
    node.append('  }')
    node.append('  SetBusContext(gIL_BusContext);')
    node.append('  result = ARILControlInit(); //Initialization of AUTOSAR PDU IL, disables the automatic IL start')
    node.append('  switch(result)')
    node.append('  {')
    node.append('    case  0   : write("ILControlInit :: {0} :: No error."); break;'.format(node_name))
    node.append('    case -1   : write("ILControlInit :: {0} :: IL is in a state that ignores the given request: State graph of the IL (Link)."); break;'.format(node_name))
    node.append('    case -2   : write("ILControlInit :: {0} :: Allowed value range of a signal exceeded. (The exceeded value is transmitted to the bus nevertheless).");  break;'.format(node_name))
    node.append('    case -50  : write("ILControlInit :: {0} :: Node layer is not active. Presumably it is deactivated in the node s configuration dialog.");  break;'.format(node_name))
    node.append('    case -51  : write("ILControlInit :: {0} :: API function not supported.");  break;'.format(node_name))
    node.append('    case -52  : write("ILControlInit :: {0} :: The node is connected to another bus like the CAN Bus.");  break;'.format(node_name))
    node.append('    case -100 : write("ILControlInit :: {0} :: Signal/message not found in DB, or not mapped to the node.");  break;'.format(node_name))
    node.append('    case -101 : write("ILControlInit :: {0} :: Maximum possible value range exceeded. (There is no transmission to the bus).");  break;'.format(node_name))
    node.append('    case -102 : write("ILControlInit :: {0} :: A network management-, diagnostic, or transport protocol signal was set. IL declines such types of signals, because it is not intended to connect these layers to the IL.");  break;'.format(node_name))
    node.append('    case -103 : write("ILControlInit :: {0} :: Invalid value supplied (for all types of requests).");  break;'.format(node_name))
    node.append('    case -104 : write("ILControlInit :: {0} :: General error for invalid requests that are not described furthermore.");  break;'.format(node_name))
    node.append('  }')
    node.append('  AREthILControlInit();//Initialization of AUTOSAR Eth IL,to prevent the IL autostart function.')
    
    node.append('')
    node.append('  {0}_sim_sysvar_DisableEthPDU();// Set sysvar to 0'.format(node_name))
    node.append('  {0}_sim_SetCycleTimeOffset();'.format(node_name))
    node.append('  {0}_sim_sysvar_DisableServices();//Set sysvar "CONTROLS::Provide" to 0'.format(node_name))
    node.append('  AREthSetProperty("MaxUDPMessageLength", 65519);')
    node.append('  gAfterPreStart = 1;')
    try:
        node.extend(get_onprestart_eth(network_name, node_name))
    except:
        pass
    node.append('}')


def add_remove_csi_endpoints():
    main_list = ['']
    remove_canoe_ports_list = []
    aep_variable_list = []
    aep_list = []
    create_consumed_instance_list = []

    filtered_df = node_df.iloc[1:]
    filtered_df = filtered_df[filtered_df[col_SDType] == 'consume']
    for ip_address, ipaddress_df in filtered_df.groupby(col_ip_address):
        remove_canoe_ports_list.append(f'  Remove_All_CANoe_Port_from_IP("{ip_address}");')
        
        for upd_port, upd_port_df in ipaddress_df.groupby(col_udp_port):
            aep_var = 'aep_'+ip_address.replace('.','_') + '_' + str(upd_port)
            aep_variable_list.append(f"  dword {aep_var};")
            aep_list.append(f"  {aep_var} = AREthOpenLocalApplicationEndpoint( ip_endpoint(UDP:{ip_address}:{upd_port}));")
            for srvID, srvID_df in upd_port_df.groupby(col_Service_ID):
                # print("Service ID:", srvID)
                for MajVer, MajVer_df in srvID_df.groupby(col_Major_version):
                    # print("MajorVersion:", MajVer)
                    for MinVer, MinVer_df in MajVer_df.groupby(col_Minor_version):
                        # print("MinorVersion:", MinVer)
                        for InstID, InstID_df in MinVer_df.groupby(col_Instance_ID):
                            unique_udp_ports = ipaddress_df[(ipaddress_df[col_Service_ID] == srvID)&(ipaddress_df[col_Major_version] == MajVer)&(ipaddress_df[col_Minor_version] == MinVer)&(ipaddress_df[col_Instance_ID] == InstID)][col_udp_port].unique()
                            port_namespace = f'_{upd_port}' if len(unique_udp_ports)>1 else ''
                            create_consumed_instance_list.append(f"  gCService_sif_{srvID}_{MajVer}_{MinVer}_{InstID}{port_namespace} = AREthCreateConsumedServiceInstance( {aep_var}, {srvID}, {InstID}, {MajVer}, {MinVer});")

    if aep_variable_list!=[]:
        aep_variable_list.append('')
    if aep_list!=[]:
        aep_list.append('')    
    if create_consumed_instance_list!=[]:
        create_consumed_instance_list.append('')
    if remove_canoe_ports_list!=[]:
        remove_canoe_ports_list.append('')
    main_list.append(f'void {node_name}_add_remove_csi_endpoints()  // for all consumed services consumed instance will be created')
    main_list.append('{')
    main_list.extend(aep_variable_list)
    main_list.extend(aep_list)
    main_list.extend(create_consumed_instance_list)
    main_list.extend(remove_canoe_ports_list)
    main_list.extend(create_consumed_instance_list)
    main_list.append('}')

    #node.extend(main_list)

    logger.info(f"Created consumed instance for all endpoints")
    return main_list


def generate_onstart(): 
    node.append('')
    node.append('on start {')
    node.append('')
    try:
        node.extend(get_customer_onstart_before_IL(network_name, node_name))
    except:
        pass
    node.append('  SetBusContext(gIL_BusContext);')
    node.append('  ARILControlSimulationOn();//Starts the simulation of the IL. The IL is operational and started. The IL will send PDUs.')
    node.append('  AREthEstablishTCPConnection(); // Establishes one or multiple database defined TCP connection(s).')
    node.append('  //{0}_sim_SetEthParameter();'.format(node_name))
    node.append('  //{0}_sim_SetSomeIPParameter();'.format(node_name))
    node.append('  {0}_sim_DisableServices();'.format(node_name))
    node.append('  //AREthILControlStart();//Enabled sending and receiving SOME/IP messages and Service Discovery. After the start of AUTOSAR Eth IL, all Application Endpoints and the Provided Services must be created.')
    node.append('  //ILControlStart();//Cyclical sending starts; setting signals is now possible.')
    node.append('  {0}_sim_Reset();'.format(node_name))
    #node.append('  //Remove_IPs_From_TcpIp_Stack();')
    #node.append('  {0}_add_remove_csi_endpoints();'.format(node_name))
    try:
        node.extend(get_customer_onstart_after_IL(network_name, node_name))
    except:
        pass
    node.append('}')


def generate_onstop():
    node.append('')
    node.append('on stopMeasurement {')
    node.append('')
    node.append('  {0}_sim_DisableEthPdu();'.format(node_name))
    node.append('  {0}_sim_DisableServices();'.format(node_name))
    node.append('  {0}_sim_sysvar_DisableEthPDU();// Set sysvar to 0'.format(node_name))
    node.append('  {0}_sim_sysvar_DisableServices();//Set sysvar "CONTROLS::Provide" to 0'.format(node_name))
    node.append('  ARILControlStop();//Stops sending of PDUs.')
    node.append('  AREthILControlStop();//Disables sending and receiving SOME/IP messages and Service Discovery. Application-Endpoints and Provided Services will be closed')
    node.append('  AREthCloseEstablishedTCPConnection();//Closes on eor multiple database defined TCP connection(s).')
    node.append('  ARILControlSimulationOff();//Stops the simulation of the IL. The IL is not operational. All API function except function ARILControlSimulationOn will not work.')
    node.append('}')


def generate_onsim():
    if node_name in hil_ctrl_node_dict.keys() and hil_ctrl_node_dict[node_name] != "":
        nod_name = hil_ctrl_node_dict[node_name]
    else:
        nod_name = node_name
    node.append('')
    node.append('on sysvar_update hil_ctrl::{0}'.format(nod_name))
    node.append('{')
    node.append('  {0}_sim_Reset();'.format(node_name))
    node.extend(get_eth_rbs_onsim(network_name, node_name))
    node.append('}')


def generate_onbus():
    node.append('')
    node.append('on sysvar_update Cus_bus::bus_{0}_{1}_ON_OFF'.format(bus_type_from_sheet_name, bus_name))
    node.append('{')
    node.append('  {0}_sim_Reset();'.format(node_name))
    node.append('}')


def generate_customer_functions():
    """gets customer specific functions"""
    try:
        node.extend(get_customer_functions(network_name, node_name))
    except:
        return


def generate_onvariant():
    node.append('')
    node.append('on sysvar_update hil_ctrl::variant')
    node.append('{')
    node.append('  {0}_sim_Reset();'.format(node_name))
    node.append('}')


def generate_void_sim_Reset():
    if node_name in hil_ctrl_node_dict.keys() and hil_ctrl_node_dict[node_name] != "":
        nod_name = hil_ctrl_node_dict[node_name]
    else:
        nod_name = node_name

    node.append('')
    node.append('void {0}_sim_Reset()'.format(node_name))
    node.append('{')
    node.append('  {0}_sim_DisableEthPdu();'.format(node_name))
    node.append('  {0}_sim_DisableServices();'.format(node_name))
    node.append('  {0}_sim_sysvar_DisableEthPDU();// Set sysvar to 0'.format(node_name))
    node.append('  {0}_sim_sysvar_DisableServices();//Set sysvar "CONTROLS::Provide" to 0'.format(node_name))
    node.append('  SetBusContext(gIL_BusContext);')
    node.append('')

    #try to get customer specific conditions
    cus_condition = ''
    try:
        cus_condition = get_customer_reset_condition(network_name,node_name)
        if cus_condition!='':
            cus_condition = ' && ' + cus_condition
    except:
        pass

    node.append('  if ((@Cus_bus::bus_{0}_{1}_ON_OFF == 1) && (@hil_ctrl::{2} == 1){3})'.format(bus_type_from_sheet_name, bus_name, nod_name, cus_condition))
    node.append('  {')
    node.append('    {0}_sim_sysvar_EnableEthPdu();// Set sysvar to 1'.format(node_name))
    node.append('    {0}_sim_sysvar_EnableServices();//Set sysvar "CONTROLS::Provide" to 1'.format(node_name))
    node.append('    //AREthILControlResume();')
    node.append('    //ILControlResume(0);')
    node.append('    AREthILControlStart();//Enabled sending and receiving SOME/IP messages and Service Discovery. After the start of AUTOSAR Eth IL, all Application Endpoints and the Provided Services must be created.')
    node.append('    ARILControlStart(); //When the IL is started, then it is fully operating and sending.')
    # node.append('    {0}_sim_EnableEthPdu();'.format(node_name))
    node.append('    {0}_sim_EnableServices();'.format(node_name))
    node.append('  }')
    node.append('  else')
    node.append('  {')
    node.append('    //ILControlWait(0);')
    node.append('    //AREthILControlWait();')
    node.append('    ARILControlStop();//Stops sending of PDUs.')
    node.append('    AREthILControlStop();//Disables sending and receiving SOME/IP messages and Service Discovery. Application-Endpoints and Provided Services will be closed')
    node.append('  }')
    node.append('}')


def generate_Sysvar_EnableDisable_EthPdu():
    pdu_count = 0
    node_enable = []
    node_disable = []
    node_disable.append('')
    node_disable.append('void {0}_sim_sysvar_DisableEthPDU()//Set sysvar to 0'.format(node_name))
    node_disable.append('{')
    node_enable.append('')
    node_enable.append('void {0}_sim_sysvar_EnableEthPdu()//Set sysvar to 1'.format(node_name))
    node_enable.append('{')
    filtered_df = node_df.iloc[1:]
    for network_protocol, network_protocol_df in filtered_df.groupby(col_network_protocol):
        if 'ETH_PDU' in network_protocol.upper():
            for payload_pdu_type, payload_pdu_type_df in network_protocol_df.groupby(col_payload_pdu_type):
                if is_empty_or_spaces(payload_pdu_type):
                    for pdu_type, pdu_type_df in payload_pdu_type_df.groupby(col_pdu_type):
                        if "I-SIGNAL" in pdu_type.upper():
                            for pdu, pdu_df in pdu_type_df.groupby(col_pdu):
                                pdu_count += 1
                                if not (is_empty_or_spaces(pdu_df[col_cycle_time].iloc[0])):
                                    node_enable.append(f'  @{network_name}::{node_name}::PDUs::{pdu}::{pdu}_ON_OFF = 1;')
                                node_disable.append(f'  @{network_name}::{node_name}::PDUs::{pdu}::{pdu}_ON_OFF = 0;')
                elif 'MULTIPLEXED-I-PDU' in payload_pdu_type.upper():
                    for Payload_PDU, Payload_PDU_df in payload_pdu_type_df.groupby(col_payload_pdu):
                        pdu_count += 1
                        node_enable.append(f'  @{network_name}::{node_name}::PDUs::{Payload_PDU}::{Payload_PDU}_ON_OFF = 1;')
                        node_disable.append(f'  @{network_name}::{node_name}::PDUs::{Payload_PDU}::{Payload_PDU}_ON_OFF = 0;')
    node_disable.append('}')
    node_enable.append('}')
    node.extend(node_disable)
    node.extend(node_enable)
    logger.info(f" generate_Sysvar_EnableDisable_EthPdu: generated PDUs {str(pdu_count)}")


def generate_Disable_EthPdu():
    pdu_count = 0
    node.append('')
    node.append('void {0}_sim_DisableEthPdu()'.format(node_name))
    node.append('{')
    DisablePdu = 'ARILFaultInjectionDisablePDU'
    filtered_df = node_df.iloc[1:]
    for network_protocol, network_protocol_df in filtered_df.groupby(col_network_protocol):
        if 'ETH_PDU' in network_protocol.upper():
            for payload_pdu_type, payload_pdu_type_df in network_protocol_df.groupby(col_payload_pdu_type):
                if is_empty_or_spaces(payload_pdu_type):
                    for pdu_type, pdu_type_df in payload_pdu_type_df.groupby(col_pdu_type):
                        if "I-SIGNAL" in pdu_type.upper():
                            for pdu, pdu_df in pdu_type_df.groupby(col_pdu):
                                pdu_count += 1
                                node.append('  {0}({1});'.format(DisablePdu, pdu))
                elif 'MULTIPLEXED-I-PDU' in payload_pdu_type.upper():
                    for Payload_PDU, Payload_PDU_df in payload_pdu_type_df.groupby(col_payload_pdu):
                        pdu_count += 1
                        node.append('  {0}({1});'.format(DisablePdu, Payload_PDU))
    for pdu in blk_pdu_list:
        node.append('   {0}({1});'.format(DisablePdu, pdu))
    node.append('}')
    logger.info(f"generated_void {str(node_name)}_sim_DisableEthPdu()' with {str(pdu_count)} PDUs.")


def generate_dword_applPDUILTxPending():
    """
    applPDUILTxPending : Callback is being called before the IL sends a PDU to the bus
    """
    #node.append('')
    node.append('dword applPDUILTxPending (dword busContext, dword shortID, dword longID, char name[], dword & aPDULength, byte data[])//This callback is being called before the IL sends a PDU to the bus'.format(node_name))
    node.append('{')
    pdu_count = 0
    filtered_df = node_df.iloc[1:]
    for network_protocol, network_protocol_df in filtered_df.groupby(col_network_protocol):
        if 'ETH_PDU' in network_protocol.upper():
            for payload_pdu_type, payload_pdu_type_df in network_protocol_df.groupby(col_payload_pdu_type):
                if is_empty_or_spaces(payload_pdu_type):
                    for pdu_type, pdu_type_df in payload_pdu_type_df.groupby(col_pdu_type):
                        if "I-SIGNAL" in pdu_type.upper():
                            for pdu, pdu_df in pdu_type_df.groupby(col_pdu):
                                node.append('  if(strncmp(name, "{0}", elcount(name)) == 0)'.format(pdu))
                                node.append('  {')
                                node.append('    cfg_Pdu_{0}(aPDULength, data);'.format(pdu))
                                node.append("    return 1; // don't prevent sending of the pdu")
                                node.append('  }')
                                pdu_count += 1
                elif 'MULTIPLEXED-I-PDU' in payload_pdu_type.upper():
                    for Payload_PDU, Payload_PDU_df in payload_pdu_type_df.groupby(col_payload_pdu):
                        node.append('  if(strncmp(name, "{0}", elcount(name)) == 0)'.format(Payload_PDU))
                        node.append('  {')
                        node.append('    cfg_Pdu_{0}(aPDULength, data);'.format(Payload_PDU))
                        node.append("    return 1; // don't prevent sending of the pdu")
                        node.append('  }')
                        pdu_count += 1
    node.append("  return 1; // don't prevent sending of the pdu")
    node.append('}')
    node.append('')
    logger.info(f"void: generate_dword_applPDUILTxPending with {str(pdu_count)} messages.")


def generate_onsysvar():
    tmr_count = 0
    count_normal = 0  # Normal
    count_ac = 0  # Alive Counter
    count_bc = 0  # Block Counter
    count_ts = 0  # Time Stamp
    count_crc = 0  # CRC or CHKSUM
    signal_sum = 0
    filtered_df = node_df.iloc[1:]
    for network_protocol, network_protocol_df in filtered_df.groupby(col_network_protocol):
        if 'ETH_PDU' in network_protocol.upper():
            for payload_pdu_type, payload_pdu_type_df in network_protocol_df.groupby(col_payload_pdu_type):
                if is_empty_or_spaces(payload_pdu_type):
                    for pdu_type, pdu_type_df in payload_pdu_type_df.groupby(col_pdu_type):
                        if "I-SIGNAL" in pdu_type.upper():
                            for pdu, pdu_df in pdu_type_df.groupby(col_pdu):
                                namespace = pdu
                                EnablePdu = 'ARILFaultInjectionEnablePDU'
                                DisablePdu = 'ARILFaultInjectionDisablePDU'
                                PduEvent = 'ARILSetPDUEvent'
                                function_def = 'void cfg_Pdu_{0}(dword & aPDULength, byte data[])'
                                SigGrp_def = 'void cfg_SigGrp_{0}(dword & aPDULength, byte data[])'
                                SigGrp_cal = '   cfg_SigGrp_{0}(aPDULength, data);'

                                cycletime = pdu_df[col_cycle_time].values[0]

                                try:
                                    cycletime = float(cycletime)
                                except:
                                    cycletime = 0

                                node.append(f'on sysvar {network_name}::{node_name}::PDUs::{namespace}::{namespace}_ON_OFF')
                                node.append('{')

                                # try to get customer specific on_off condition conditions
                                cus_condition = ''
                                try:
                                    cus_condition = get_customer_onoff_condition(network_name, node_name, namespace)
                                    if cus_condition != '':
                                        cus_condition = ' && ' + cus_condition
                                except:
                                    pass

                                node.append('  if ((@this == 1) && (@Cus_bus::bus_{0}_ON_OFF == 1){1})'.format(network_name, cus_condition))
                                node.append('  {')
                                if cycletime == 0:  # Event
                                    node.append('     {0}({1});'.format(PduEvent, namespace))
                                    node.append(f'     @{network_name}::{node_name}::PDUs::{namespace}::{namespace}_ON_OFF = 0;')
                                    node.append('  }')
                                else:  # Cyclic
                                    node.append('    {0}({1});'.format(EnablePdu, namespace))
                                    node.append('  }')
                                    node.append('  else \n  {')
                                    node.append('    {0}({1});'.format(DisablePdu, namespace))
                                    node.append('  }')
                                node.append('}')
                                node.append('')
                                node.append(function_def.format(namespace))
                                node.append('{')

                                Sig_grp_fun = []  # Signal belongs to signal group
                                without_Sig_grp_fun = []  # Signal without signal group

                                siggrp_counts = pdu_df[pdu_df[col_signal_group] != ''][col_signal_group].nunique()
                                empty_siggrp_counts = pdu_df[pdu_df[col_signal_group] == ''][col_signal_group].nunique()

                                e2e_def = False
                                if siggrp_counts == 1:
                                    if (pdu_df[pdu_df[col_endtoendprotection] != ''][col_endtoendprotection].nunique() != 0):  # If pdu group has E2E profile
                                        e2e_def = True

                                signal_dict = {}
                                for signal_name, signal_df in pdu_df.groupby(col_Signal_name):
                                    signal_byte_order = signal_df[col_Signal_ByteOrder].values[0]
                                    signal_start_bit = signal_df[col_startbit].values[0]
                                    signal_length = signal_df[col_signal_length_bit].values[0]
                                    signal_para_list = [signal_byte_order, signal_start_bit, signal_length]
                                    if 'OPAQUE' in signal_byte_order.upper():
                                        signal_dict[signal_name] = signal_para_list

                                # signal_para_list = [signal_name, signal_byte_order, signal_start_bit, signal_length]
                                # print(f'pdu_df = \n{pdu_df}')
                                para_dict = {
                                    "bus_type": bus_type,
                                    "input_file": input_file,
                                    "message": '',
                                    "pdu": pdu,
                                    "network_name": network_name,
                                    "e2e_def": e2e_def,
                                    "signal_dict": signal_dict
                                }
                                node.extend(ontmr_variables(para_dict))  # Import from pattern_matching

                                try:
                                    node.extend(get_prepare_pdu_event(network_name, node_name, namespace))
                                except:
                                    pass

                                for siggrp, siggrp_df in pdu_df.groupby(col_signal_group):
                                    # print("Signal group:", siggrp)
                                    if siggrp_counts > 1:
                                        if not is_empty_or_spaces(siggrp):
                                            node.append(SigGrp_cal.format(siggrp))  # call signal group config

                                    num_of_signals = siggrp_df.shape[0]  # print("num_of_signals",num_of_signals)

                                    byte_order = str(siggrp_df[col_Signal_ByteOrder].values[0])

                                    SigGrp_startbit = min(list(siggrp_df[col_startbit].values))  # print("SigGrp_startbit",SigGrp_startbit)
                                    SigGrp_startbit_length = siggrp_df[col_signal_length_bit].values[list(siggrp_df[col_startbit].values).index(SigGrp_startbit)]  # print("SigGrp_startbit_length",SigGrp_startbit_length)

                                    SigGrp_endbit = max(list(siggrp_df[col_startbit].values))  # print("SigGrp_endbit",SigGrp_endbit)
                                    SigGrp_endbit_length = siggrp_df[col_signal_length_bit].values[list(siggrp_df[col_startbit].values).index(
                                        SigGrp_endbit)]  # print("SigGrp_endbit_length",SigGrp_endbit_length)

                                    if not is_empty_or_spaces(siggrp):
                                        DataLength = Get_DataLength(SigGrp_startbit, SigGrp_startbit_length, SigGrp_endbit,
                                                                    SigGrp_endbit_length, byte_order)

                                    siggrp_info = {
                                        "siggrp": siggrp,
                                        "SigGrp_startbit": SigGrp_startbit,
                                        "SigGrp_startbit_length": SigGrp_startbit_length,
                                        "SigGrp_endbit": SigGrp_endbit,
                                        "SigGrp_endbit_length": SigGrp_endbit_length,
                                        "siggrp_counts": siggrp_counts,
                                        "empty_siggrp_counts": empty_siggrp_counts
                                    }

                                    total_used_Pdu_bytes = ceil((SigGrp_endbit + SigGrp_endbit_length) / 8) - 1  # print("total_used_Pdu_bytes",total_used_Pdu_bytes)

                                    part_Pdu_null = []  # Normal
                                    part_Pdu_ac = []  # Alive Counter
                                    part_Pdu_bc = []  # Block Counter
                                    part_Pdu_ts = []  # Time Stamp
                                    part_Pdu_crc = []  # CRC or CHKSUM
                                    tmr_count += 1
                                    signal_sum += num_of_signals
                                    Pdu_length = 0  # in bits
                                    # for j in range(num_of_signals):
                                    # Pdu_length = Pdu_length + byteOrder_to_8byte(int(siggrp_df[col_signal_length_bit].values[j]))

                                    for j in range(num_of_signals):
                                        break_out = 0
                                        signal_name = str(siggrp_df[col_Signal_name].values[j])
                                        byte_order = str(siggrp_df[col_Signal_ByteOrder].values[j])
                                        endtoendprotection = str(siggrp_df[col_endtoendprotection].values[j])
                                        comment = str(siggrp_df[col_comment].values[j])
                                        startbit = int(siggrp_df[col_startbit].values[j]) if (
                                                str(siggrp_df[col_startbit].values[j]) != "") else 0
                                        Pdu_length = byteOrder_to_8byte(int(siggrp_df[col_signal_length_bit].values[j])) if (
                                                str(siggrp_df[col_signal_length_bit].values[j]) != "") else 0

                                        sig_factor = float(siggrp_df[col_Factor].values[j]) if (
                                                str(siggrp_df[col_Factor].values[j]) != "") else float(1)

                                        sig_offset = float(siggrp_df[col_Offset].values[j]) if (
                                                str(siggrp_df[col_Offset].values[j]) != "") else float(0)

                                        dlc = int(siggrp_df[col_dlc].values[j]) if (str(siggrp_df[col_dlc].values[j]) != "") else 0

                                        sig_length = int(siggrp_df[col_signal_length_bit].values[j]) if (
                                                    str(siggrp_df[col_signal_length_bit].values[j]) != "") else 0

                                        para_dict = {
                                            "namespace": pdu,
                                            "signal_name": signal_name,
                                            "siggrp_info": siggrp_info,
                                            "startbit": startbit,
                                            "msg_length": Pdu_length,
                                            "sig_length": sig_length,
                                            "byte_order": byte_order,
                                            "sig_factor": sig_factor,
                                            "sig_offset": sig_offset,
                                            "comment": comment,
                                            "cycletime": cycletime,
                                            "dlc": dlc,
                                            "bus_type": bus_type,
                                            "input_file": input_file,
                                            "network_name": network_name,
                                            "total_used_msg_bytes": total_used_Pdu_bytes,
                                            "node_name": node_name,
                                            "endtoendprotection": endtoendprotection
                                        }

                                        if (break_out != 1):
                                            for sig_pattern, sig_fun in ac_dict.items():
                                                if (None != re.search(sig_pattern, signal_name, re.M | re.I)):  # Pattern :: Alive counter
                                                    # print(siggrp,signal_name)
                                                    if is_empty_or_spaces(endtoendprotection):
                                                        part_Pdu_ac = part_Pdu_ac + sig_fun(para_dict)
                                                    else:
                                                        part_Pdu_ac = part_Pdu_ac + SigGrp_ARILFaultInjectionDisturbSequenceCounter(para_dict)
                                                    count_ac += 1
                                                    break_out = 1
                                                    break
                                        if (break_out != 1):
                                            for sig_pattern, sig_fun in bc_dict.items():
                                                if (None != re.search(sig_pattern, signal_name, re.M | re.I)):  # Pattern :: Block counter
                                                    part_Pdu_bc = part_Pdu_bc + sig_fun(para_dict)
                                                    count_bc += 1
                                                    break_out = 1
                                                    break
                                        if (break_out != 1):
                                            for sig_pattern, sig_fun in ts_dict.items():
                                                if (None != re.search(sig_pattern, signal_name, re.M | re.I)):  # Pattern :: Time Stamp
                                                    part_Pdu_ts = part_Pdu_ts + sig_fun(para_dict)
                                                    count_ts += 1
                                                    break_out = 1
                                                    break
                                        if (break_out != 1):
                                            for sig_pattern, sig_fun in crc_dict.items():
                                                if (None != re.search(sig_pattern, signal_name, re.M | re.I)):  # Pattern ::CRC / CHK_SUM
                                                    if is_empty_or_spaces(endtoendprotection):
                                                        part_Pdu_crc = part_Pdu_crc + sig_fun(para_dict)
                                                    else:
                                                        part_Pdu_crc = part_Pdu_crc + SigGrp_ARILFaultInjectionDisturbChecksum(para_dict)
                                                    count_crc += 1
                                                    break_out = 1
                                                    break
                                        if (break_out != 1):
                                            count_normal += 1
                                            # part_Pdu_null.append('   Pdu_{0}.{1} = f_Phys2Raw(@{0}::{1}_Rv,{2},{3});'.format(pdu, signal_name,sig_factor,sig_offset))#Pdu_{0}.{1}.phys = @{0}::{1}_Rv

                                    if is_empty_or_spaces(siggrp):
                                        without_Sig_grp_fun.extend(part_Pdu_null)
                                        without_Sig_grp_fun.extend(part_Pdu_ac)
                                        without_Sig_grp_fun.extend(part_Pdu_bc)
                                        without_Sig_grp_fun.extend(part_Pdu_ts)
                                        without_Sig_grp_fun.extend(part_Pdu_crc)
                                    else:
                                        if siggrp_counts > 1:
                                            Sig_grp_fun.append(SigGrp_def.format(siggrp))
                                            Sig_grp_fun.append('{')
                                            Sig_grp_fun.append('   int64 startbit,bitlength;')
                                            Sig_grp_fun.append('   int64 i,k;//For looping')
                                            Sig_grp_fun.append('   byte DataLength = {0};//Signal group length in byte'.format(DataLength))
                                            Sig_grp_fun.append('   byte SigGrp_Data[{0}];'.format(DataLength))
                                            Sig_grp_fun.append('   long result;')
                                            if (siggrp_df[siggrp_df[col_endtoendprotection] != ''][
                                                col_endtoendprotection].nunique() != 0):  # If signal group has E2E profile
                                                Sig_grp_fun.append('   long type=0;//reserved (should be set to 0).')
                                                Sig_grp_fun.append(
                                                    '   long disturbanceMode  =0; //0 :: set to disturbanceValue //2 :: set to Random value //3 :: Offset Signal value is increased by disturbanceValue')
                                                Sig_grp_fun.append(
                                                    '   long disturbanceCount =0;//-1 ::Infinite Disturbance is continuously applied //0 ::Stop An active disturbance is stopped and the SequenceCounter/crc will be calculated again appropriately //n > 0 ::Count Do exactly n Repetition/disturbances')
                                                Sig_grp_fun.append(
                                                    '   long disturbanceValue =0;//According to the disturbance mode the SequenceCounter/crc will optionally be set to this value')
                                                Sig_grp_fun.append('   dword crc[1];')

                                            Sig_grp_fun.extend(part_Pdu_null)
                                            Sig_grp_fun.extend(part_Pdu_ac)
                                            Sig_grp_fun.extend(part_Pdu_bc)
                                            Sig_grp_fun.extend(part_Pdu_ts)
                                            Sig_grp_fun.extend(part_Pdu_crc)
                                            Sig_grp_fun.append('')
                                            Sig_grp_fun.append('}')
                                            Sig_grp_fun.append('')
                                        else:
                                            without_Sig_grp_fun.extend(part_Pdu_null)
                                            without_Sig_grp_fun.extend(part_Pdu_ac)
                                            without_Sig_grp_fun.extend(part_Pdu_bc)
                                            without_Sig_grp_fun.extend(part_Pdu_ts)
                                            without_Sig_grp_fun.extend(part_Pdu_crc)

                                node.extend(without_Sig_grp_fun)
                                node.append('}')
                                node.append('')
                                node.extend(Sig_grp_fun)
                elif 'MULTIPLEXED-I-PDU' in payload_pdu_type.upper():
                    EnablePdu = 'ARILFaultInjectionEnablePDU'
                    DisablePdu = 'ARILFaultInjectionDisablePDU'
                    for Payload_PDU, Payload_PDU_df in payload_pdu_type_df.groupby(col_payload_pdu):
                        if not is_empty_or_spaces(Payload_PDU):
                            node.append('')
                            node.append(f'on sysvar {network_name}::{node_name}::{Payload_PDU}::{Payload_PDU}_ON_OFF')
                            node.append('{')
                            node.append('  if (@this == 1)')
                            node.append('  {')
                            node.append('    {0}({1});'.format(EnablePdu, Payload_PDU))
                            node.append('  }')
                            node.append('  else \n  {')
                            node.append('    {0}({1});'.format(DisablePdu, Payload_PDU))
                            node.append('  }')
                            node.append('}')
                            for selector_field_signal, selector_field_signal_df in Payload_PDU_df.groupby(col_selector_field_signal):
                                if 'yes' in selector_field_signal.lower():
                                    node.append('')
                                    node.append(f'on sysvar {network_name}::{node_name}::{Payload_PDU}::{selector_field_signal_df[col_Signal_name].values[0]}_Selection_Field_Code')
                                    node.append('{')
                                    node.append(f'  ${Payload_PDU}::{selector_field_signal_df[col_Signal_name].values[0]} = @this;')
                                    node.append('}')
                            
                            #cfg_pdu function for multiplexing pdu currently keeping empty
                            node.append('')
                            node.append(f'void cfg_Pdu_{Payload_PDU}(dword & aPDULength, byte data[])')
                            node.append('{')
                            node.append('}')

    logger.info("Timer count: {0} on timer functions were generated with {1} signals in total.".format(str(tmr_count), str(signal_sum)))
    logger.info("\tNormal: {0}; Alive Counter: {1}; Time Stamp: {2}; CRC or CHKSUM: {3}".format(count_normal, count_ac, count_ts, count_crc))


def generate_set_pdu_CycleTime():
    """
        Configure cycle time for cyclic pdu
    """
    pdu_count = 0
    node.append('')
    node.append('//  ARILSetPDUTimingCyclic (dbMsg dbMessage, long TrueOrFalse, long offset, long period, long disturbanceCount, long flags);')
    node.append('//  TrueOrFalse ; 0: denotes the False timing; 1: denotes the True timing; 3: denotes the True and the False timing.')
    node.append('//  offset :Defines the delay in [ms] from now when the first cyclic transmission will start.')
    node.append('//  period :Defines the period in [ms] for the cyclic transmission.')
    node.append('//  disturbanceCount :Reserved/unused; should be set to -1 (infinite).')
    node.append('//  flags :Reserved; should be set to 0.')
    filtered_df = node_df.iloc[1:]
    for network_protocol, network_protocol_df in filtered_df.groupby(col_network_protocol):
        if 'ETH_PDU' in network_protocol.upper():
            for payload_pdu_type, payload_pdu_type_df in network_protocol_df.groupby(col_payload_pdu_type):
                if is_empty_or_spaces(payload_pdu_type):
                    # print(f'generate_set_pdu_CycleTime::payload_pdu_type = {payload_pdu_type}')
                    for pdu_type, pdu_type_df in payload_pdu_type_df.groupby(col_pdu_type):
                        if "I-SIGNAL" in pdu_type.upper():
                            for pdu, pdu_df in pdu_type_df.groupby(col_pdu):
                                # print(f'generate_set_pdu_CycleTime::pdu = {pdu}')
                                cycletime = pdu_df[col_cycle_time].values[0]
                                try:
                                    cycletime = float(cycletime)
                                except:
                                    cycletime = 0
                                if cycletime != 0:
                                    node.append('')
                                    node.append(f'on sysvar {network_name}::{node_name}::PDUs::{pdu}::{pdu}_CycleTime')
                                    node.append('{')
                                    node.append('  ARILSetPDUTimingCyclic ({0}, 3, {1}, @this, -1, 0);//TrueOrFalse=3,offset={1};disturbanceCount =-1;flags=0'.format(pdu, pdu_count))
                                    node.append('  //ARILResetPDUTimingCyclic({0},3);//TrueOrFalse=3'.format(pdu))
                                    node.append('}')
                                    if ((pdu_count + 1) % 10 != 0):  # Should not collide with 10ms pdu
                                        pdu_count += 1
                                    else:
                                        pdu_count += 2
    logger.info(f"generate_set_pdu_CycleTime with {str(pdu_count)} messages.")


def generate_void_sim_SetCycleTimeOffset():
    """
        Configure cycle time and offset for cyclic pdu
    """
    pdu_count = 0
    node.append('')
    node.append('void {0}_sim_SetCycleTimeOffset()'.format(node_name))
    node.append('{')
    node.append('')
    node.append('//  ARILSetPDUTimingCyclic (dbMsg dbMessage, long TrueOrFalse, long offset, long period, long disturbanceCount, long flags);')
    node.append('//  TrueOrFalse ; 0: denotes the False timing; 1: denotes the True timing; 3: denotes the True and the False timing.')
    node.append('//  offset :Defines the delay in [ms] from now when the first cyclic transmission will start.')
    node.append('//  period :Defines the period in [ms] for the cyclic transmission.')
    node.append('//  disturbanceCount :Reserved/unused; should be set to -1 (infinite).')
    node.append('//  flags :Reserved; should be set to 0.')
    filtered_df = node_df.iloc[1:]
    for network_protocol, network_protocol_df in filtered_df.groupby(col_network_protocol):
        if 'ETH_PDU' in network_protocol.upper():
            for payload_pdu_type, payload_pdu_type_df in network_protocol_df.groupby(col_payload_pdu_type):
                if is_empty_or_spaces(payload_pdu_type):
                    # print(f'generate_void_sim_SetCycleTimeOffset::payload_pdu_type = {payload_pdu_type}')
                    for pdu_type, pdu_type_df in payload_pdu_type_df.groupby(col_pdu_type):
                        if "I-SIGNAL" in pdu_type.upper():
                            for pdu, pdu_df in pdu_type_df.groupby(col_pdu):
                                # print(f'generate_void_sim_SetCycleTimeOffset::pdu = {pdu}')
                                cycletime = pdu_df[col_cycle_time].values[0]
                                try:
                                    cycletime = float(cycletime)
                                except:
                                    cycletime = 0
                                if cycletime != 0:
                                    node.append('  ARILSetPDUTimingCyclic ({0}, 3, {1}, {2}, -1, 0);//TrueOrFalse=3;offset={1};cycletime={2};disturbanceCount =-1;flags=0'.format(pdu, pdu_count, cycletime))
                                    node.append('  //ARILResetPDUTimingCyclic({0},3);//TrueOrFalse=3'.format(pdu))
                                    if ((pdu_count + 1) % 10 != 0):  # Should not collide with 10ms pdu
                                        pdu_count += 1
                                    else:
                                        pdu_count += 2
    node.append('')
    node.append('}')
    logger.info(f"void: generate_void_sim_SetCycleTimeOffset with {str(pdu_count)} messages.")


def generate_EnableDisableServices():
    node.append('')
    node.append('void {0}_sim_DisableServices()'.format(node_name))
    node.append('{')
    node.append('  {0}_FinishProvidedServices();'.format(node_name))
    node.append('  {0}_FinishConsumedServices();'.format(node_name))
    node.append('}')
    node.append('')
    node.append('void {0}_sim_EnableServices()'.format(node_name))
    node.append('{')
    node.append('  //AREthEstablishTCPConnection(); // Establish TCP Connection')
    node.append('  {0}_InitProvidedServices();'.format(node_name))
    node.append('  {0}_InitConsumedServices();'.format(node_name))
    node.append('}')


def generate_EnableDisable():
    # Initialse Provided Services
    Ps_count = 0
    event_count = 0
    node_enable =[]
    node_disable =[]
    node_disable.append('')
    node_disable.append('void {0}_sim_sysvar_DisableServices()//Set sysvar "CONTROLS::Provide" and "CONTROLS::Subscribe" to 0'.format(node_name))
    node_disable.append('{')
    node_enable.append('')
    node_enable.append('void {0}_sim_sysvar_EnableServices()//Set sysvar "CONTROLS::Provide" and "CONTROLS::Subscribe" to 1'.format(node_name))
    node_enable.append('{')
    filtered_df = node_df.iloc[1:]

    block_services = dict_block_services.get(network_name + '_' + node_name, [])
    for srvID, srvID_df in filtered_df.groupby(col_Service_ID):
        #print("Service ID:", srvID)
        for MajVer, MajVer_df in srvID_df.groupby(col_Major_version):
            #print("MajorVersion:", MajVer)
            for MinVer, MinVer_df in MajVer_df.groupby(col_Minor_version):
                #print("MinorVersion:", MinVer)
                for InstID, InstID_df in MinVer_df.groupby(col_Instance_ID):
                    #print("Instance_ID:", InstID)
                    for SDType, SDType_df in InstID_df.groupby(col_SDType):
                        #print("SD Type:", SDType)
                        if "provide"  in SDType.lower():
                            Ps_count += 1
                            service_namespace = f'sif_{srvID}_{MajVer}_{MinVer}_{InstID}'
                            node_disable.append('  @sysvar::ETH_{1}::{2}::PROVIDED_SERVICES::{0}::CONTROLS::Provide=0;'.format(service_namespace, bus_name,node_name))
                            block_service_exception = False
                            exception_service_blocked_list = []
                            try:
                                exception_service_blocked_list = block_provided_services_exception_dict.get(network_name, {}).get(node_name, [])
                            except:
                                pass

                            if (service_namespace in exception_service_blocked_list) or (str(srvID) in exception_service_blocked_list):
                                block_service_exception = True
                            if service_namespace in block_services or block_service_exception:
                                node_enable.append('  //@sysvar::ETH_{1}::{2}::PROVIDED_SERVICES::{0}::CONTROLS::Provide=1;'.format(service_namespace, bus_name, node_name))
                            else:
                                node_enable.append('  @sysvar::ETH_{1}::{2}::PROVIDED_SERVICES::{0}::CONTROLS::Provide=1;'.format(service_namespace, bus_name,node_name))
                                
                            for MemberType,MemberType_df in SDType_df.groupby(col_MemberType):
                                if 'field' in MemberType.lower():
                                    if 'notifier' in MemberType_df[col_field_type].values:
                                        if service_namespace in block_services or block_service_exception:
                                            node_enable.append(f'  //setTimer(Timer_sif_{srvID}_{MajVer}_{MinVer}_{InstID}, 5);')
                                        else:
                                            node_enable.append(f'  setTimer(Timer_sif_{srvID}_{MajVer}_{MinVer}_{InstID}, 5);')
                                            
                        if "consume"  in SDType.lower():
                            unique_udp_ports = SDType_df[col_udp_port].unique()
                            for udp_port in unique_udp_ports:
                                port_namespace = f'_{udp_port}' if len(unique_udp_ports)>1 else ''
                                for MemberType,MemberType_df in SDType_df.groupby(col_MemberType):
                                    if "event" in MemberType.lower():
                                        for eventgroup_id,eventgroup_id_df in MemberType_df.groupby(col_EventGroup):
                                            event_count += 1
                                            node_disable.append('  @sysvar::ETH_{5}::{6}::CONSUMED_SERVICES::sif_{0}_{1}_{2}_{3}{7}::EVENTGROUPS::{4}::CONTROLS::Subscribe=0;'.format(srvID,MajVer,MinVer,InstID,eventgroup_id,bus_name,node_name, port_namespace))
                                            node_enable.append('  @sysvar::ETH_{5}::{6}::CONSUMED_SERVICES::sif_{0}_{1}_{2}_{3}{7}::EVENTGROUPS::{4}::CONTROLS::Subscribe=1;'.format(srvID,MajVer,MinVer,InstID,eventgroup_id,bus_name,node_name, port_namespace))
                
    node_disable.append('}')
    # node_disable.append('')
    node_enable.append('}')
    node_enable.append('')
    node.extend(node_disable)
    node.extend(node_enable)
    logger.info(f" generate_EnableDisable: generated with {str(Ps_count)} Provide Services and {str(event_count)} Consume event group.")

def generate_RemoveIpFromTcpIpStack():
    node_remove_ip_from_tcpip_stack = []
    removed_ip_address_function_call_list = ['',
                                             'void Remove_IPs_From_TcpIp_Stack()',
                                             '{']
    #if block_nwt_endpoint_list
    if (block_nwt_endpoint_list==[]) or (blocked_df.empty):
        removed_ip_address_function_call_list.append('}')
        node.extend(removed_ip_address_function_call_list)
        return


    for ip_address, adapter_index in block_nwt_endpoint_list:
        ip_string = ip_address.replace('.','_')
        removed_ip_address_function_call_list.append(f'  Remove_{ip_string}_From_TcpIP_Stack();')
        
        unique_ports = []
        ip_df = pd.DataFrame([])
        if not(blocked_df.empty):
            ip_df = blocked_df[blocked_df[col_ip_address]==ip_address]
            if not(ip_df.empty):
                unique_ports = ip_df[col_udp_port].unique()
        
        node_remove_ip_from_tcpip_stack.append('')
        node_remove_ip_from_tcpip_stack.append(f'void Remove_{ip_string}_From_TcpIP_Stack()')
        node_remove_ip_from_tcpip_stack.append('{')
        node_remove_ip_from_tcpip_stack.append(f'  dword application_endpoint;')
        node_remove_ip_from_tcpip_stack.append(f'  dword csi;')
        node_remove_ip_from_tcpip_stack.append('')

        if not(ip_df.empty):
            for udp_port, upd_port_df in ip_df.groupby(col_udp_port):
                node_remove_ip_from_tcpip_stack.append(f'  application_endpoint = AREthOpenLocalApplicationEndpoint(17, {udp_port}, IpGetAddressAsNumber("{ip_address}"));')
                for service_type, service_type_df in ip_df.groupby(col_SDType):
                    for service_id, service_id_df in service_type_df.groupby(col_Service_ID):
                        for MajVer, MajVer_df in service_id_df.groupby(col_Major_version):
                            for MinVer, MinVer_df in MajVer_df.groupby(col_Minor_version):
                                for InstID, InstID_df in MinVer_df.groupby(col_Instance_ID):
                                    if service_type=='provide':
                                        node_remove_ip_from_tcpip_stack.append(f'  EnableService("sif_{service_id}::{MajVer}::{MinVer}::{InstID}", 0);')
                                    elif service_type=='consume':
                                        node_remove_ip_from_tcpip_stack.append(f'  csi = AREthCreateConsumedServiceInstance(application_endpoint, {service_id}, {InstID}, {MajVer}, {MinVer});')
                                        node_remove_ip_from_tcpip_stack.append(f'  AREthSDReleaseService(csi);')
                                        node_remove_ip_from_tcpip_stack.append(f'  AREthReleaseConsumedServiceInstance(csi);')
                
                node_remove_ip_from_tcpip_stack.append('')
                node_remove_ip_from_tcpip_stack.append(f'  AREthCloseLocalApplicationEndpoint(application_endpoint);')
                node_remove_ip_from_tcpip_stack.append('')

        
        # for i,port in enumerate(unique_ports):
        #     node_remove_ip_from_tcpip_stack.append(f'  application_endpoint = AREthOpenLocalApplicationEndpoint(17, {port}, IpGetAddressAsNumber("{ip_address}"));')
        #     node_remove_ip_from_tcpip_stack.append(f'  AREthCloseLocalApplicationEndpoint(application_endpoint);')

        node_remove_ip_from_tcpip_stack.append(f'  Remove_All_CANoe_Port_from_IP("{ip_address}");')
    

        node_remove_ip_from_tcpip_stack.append(f'  IpRemoveAdapterAddress({adapter_index}, IpGetAddressAsNumber("{ip_address}"), 24);')
        node_remove_ip_from_tcpip_stack.append('}')
        node_remove_ip_from_tcpip_stack.append('')

        logger.info(f'Removed {ip_address} and its endpoints from TCP IP stack')


    node.extend(node_remove_ip_from_tcpip_stack)

    removed_ip_address_function_call_list.append('}')
    removed_ip_address_function_call_list.append('')

    node.extend(removed_ip_address_function_call_list)

def save_file(file_path):
    # Open and save node .can file
    try:
        file = ('{3}/{0}_{1}_{2}.can'.format(node_name, bus_type_from_sheet_name, bus_name, file_path))
        with open(file, 'w') as outfile:
            #outfile.write("\n".join(str(item) for item in node))

            #this is temporary fix for CAPL keyword handling need to be revisited-----------------#TEMP_FIX
            outfile.write("\n".join(str(item).replace('::message::', '::message_::') if '::message::' in item else str(item) for item in node))
            #-----------------------------------------------------------

        logger.info(f"{file} updated successfully")
    except:
        logger.warning(f"{node_name} was NOT updated successfully", exc_info=True)

def get_block_services_info(wb):
    """
    If a service is provided by both RBS node from classic and adaptive then block the service from adaptive.
    This function returns the list of services that needs to be blocked in adaptive node.

    Args:
        wb (workbook): complete workbook

    Returns:
        dict: (network_name_node_name: [list of blocked services])
    """
    result_dict = {}
    #start_time = time.time()

    if 'SysVarDatabaseETH' not in wb.sheetnames:
        return {}
    generate_sysvardatabase_dataframe(wb, 'SysVarDatabaseETH')
    if len(sysvar_df[col_autosar_type].unique())<2:
        return result_dict
    unique_node_names = sysvar_df[sysvar_df[col_network_protocol]=='ETH_SOMEIP'][col_sender].unique()
    rbs_node_name = 'RBS'
    for nd_name in unique_node_names:
        if 'rbs' in nd_name.lower():
            rbs_node_name = nd_name
            break
    services_df = sysvar_df[(sysvar_df[col_network_protocol]=='ETH_SOMEIP')&(sysvar_df[col_sender]==rbs_node_name)&(sysvar_df[col_SDType]=='provide')]

    if len(services_df[col_autosar_type].unique()) < 2:
        return result_dict
    classic_services = {}
    for index, row in services_df.iterrows():
        if row[col_autosar_type] == 'Classic':
            ky = f"sif_{row[col_Service_ID]}_{row[col_Major_version]}_{row[col_Minor_version]}_{row[col_Instance_ID]}"
            if ky not in classic_services:
                classic_services[ky] = [row[col_member_type], f"{row[col_network_name]}_{row[col_sender]}"]
    adaptive_services = [f"sif_{row[col_Service_ID]}_{row[col_Major_version]}_{row[col_Minor_version]}_{row[col_Instance_ID]}::{row[col_network_name]}_{row[col_sender]}" for index, row in services_df.iterrows() if row[col_autosar_type] == 'Adaptive']
    adaptive_services = set(adaptive_services)

    for adp_service in adaptive_services:
        if adp_service.split('::')[0] in classic_services.keys():
            if classic_services[adp_service.split('::')[0]][0] == 'not_found': # if classic there are no members defined in service then block in classic
                if classic_services[adp_service.split('::')[0]][1] not in result_dict:
                    result_dict[classic_services[adp_service.split('::')[0]][1]] = []
                result_dict[classic_services[adp_service.split('::')[0]][1]].append(adp_service.split('::')[0])
            else:
                split = adp_service.split('::')
                if split[1] not in result_dict:
                    result_dict[split[1]] = []
                result_dict[split[1]].append(split[0])

    # end_time = time.time()
    # execution_time = end_time - start_time
    # print(f"Execution time: {execution_time} seconds")
    if result_dict!={}:
        logger.info(f'Adaptive services provided by RBS node in classic RBS node as well:\n{result_dict}')
    return result_dict


# main function to create nodes
def create_node_main(workbook, sheet, path):
    logger.info("++++++++++Creating node for: {0} ++++++++++".format(sheet))
    try:
        generate_dataframe(workbook, sheet)
        generate_includes()
        generate_variables()
        generate_onprestart()
        generate_onstart()
        generate_onstop()
        node.extend(generated_app_endpoint_initialize_list)
        generate_onsim()
        generate_onvariant()
        generate_customer_functions()
        generate_onbus()
        generate_void_sim_Reset()
        generate_Sysvar_EnableDisable_EthPdu()
        generate_Disable_EthPdu()
        generate_EnableDisableServices()
        generate_EnableDisable()
        generate_InitProvidedServices()
        generate_FinishProvidedServices()
        generate_InitConsumedServices()
        generate_FinishConsumedServices()
        generate_CallBacks()
        generate_OnProvideService()
        generate_OnSubscribeService()
        generate_OnSendOnce()
        generate_OnTimers()
        
        #generate_RemoveIpFromTcpIpStack()
        generate_dword_applPDUILTxPending()
        generate_onsysvar()
        generate_set_pdu_CycleTime()
        generate_void_sim_SetCycleTimeOffset()
        save_file(path)
        return True
    except Exception as exp:
        logger.error(f"Failed to create node for: {sheet}")
        logger.error(f"Error: {exp}")
        raise exp


def generate_sysvardatabase_dataframe(wb, sht_name):
    """reads sysvarsheet and store as global sheet"""
    global sysvar_df, col_sender, col_network_name, col_SDType, col_Service_ID, col_Major_version, col_Minor_version,col_Instance_ID,col_member_type,col_autosar_type,col_network_protocol

    if sht_name not in wb.sheetnames:
        return
    ws_sysvar = wb[sht_name]

    # create the different dataframes
    sysvar_df = pd.DataFrame(ws_sysvar.values)
    sysvar_df = sysvar_df.replace(np.nan, '', regex=True)
    # get column names
    col_names = list(sysvar_df.iloc[0])
    col_sender = col_names.index("Node Name")
    col_network_name = col_names.index("Network Name")
    col_SDType = col_names.index("SD Type")
    col_Service_ID = col_names.index("Service ID")
    col_Major_version = col_names.index("Major version")
    col_Minor_version = col_names.index("Minor version")
    col_Instance_ID = col_names.index("Instance ID")
    col_member_type = col_names.index("Member Type")
    col_autosar_type = col_names.index("Autosar Type")
    col_network_protocol = col_names.index("Network Protocol")

def initialize_dict(wb):
    global dict_network_info, dict_entity, dict_hilctrl, dict_network_name, dict_network_type, dict_block_messages, dict_pass_messages, dict_block_services
    dict_network_info = getMappingData(wb, "network_category")
    dict_network_name = getMappingData(wb, "network_name")
    dict_blk_msg = getMappingData(wb, "block_messages")
    dict_block_messages = {}
    for st, value in dict_blk_msg.items():
        if (value not in ['na', '', None]):
            value = value.strip().split(";")
            value = [i.strip() for i in value]
            dict_block_messages[st] = value

    dict_pass_msg = getMappingData(wb, "pass_messages")
    dict_pass_messages = {}
    for st, value in dict_pass_msg.items():
        if (value not in ['na', '', None]):
            value = value.strip().split(";")
            value = [i.strip() for i in value]
            dict_pass_messages[st] = value

    dict_network_type = getMappingData(wb, "network_type")
    dict_entity = getMappingData(wb, "node_entity")
    dict_hilctrl = extractHilCtrl(wb)

    dict_block_services = get_block_services_info(wb)


    
def external_call():
    try:
        logger.info(f"###### START 'Create Nodes(.can)' DEBUG INFORMATION ######")
        script_path = os.path.dirname(os.path.abspath(__file__))
        autosar_path = script_path + r'/../../../../CustomerPrj/Restbus/Autosar_Gen_Database.xlsx'
        workbook = load_workbook(autosar_path, data_only=True)
        initialize_dict(workbook)
        sheet_list = [name for name in workbook.sheetnames if len(name.split('_')) > 2]

        eth_sheets_found = False
        for sheet in sheet_list:
            if (len(sheet.split('_')) > 2) and ("sysvardatabase" not in sheet.lower()) and ("ETH_" in sheet.upper()):  # if it is an ECU sheet
                eth_sheets_found = True
                create_node_main(workbook, sheet, script_path + r'/../../../../CustomerPrj/Restbus/Nodes')
        if eth_sheets_found==False:
            logger.warning(f"No ETHERNET sheets found in Autosar_Gen_Database.xlsx")
            return

    except Exception as e:
        logger.error(f"Create Nodes(.can) failed. Exception --> {e}")
        traceback.print_exc()
        raise Exception(e)
    logger.info("###### END 'Create Nodes(.can)' DEBUG INFORMATION ######")
    logger.info('-' * 80)


if __name__ == "__main__":
    external_call()
