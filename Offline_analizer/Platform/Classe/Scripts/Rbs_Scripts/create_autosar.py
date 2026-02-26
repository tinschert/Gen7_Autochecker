# -*- coding: utf-8 -*-
# @file create_autosar.py
# @author ADAS_HIL_TEAM
# @date 08-21-2023

##################################################################
# C O P Y R I G H T S
# ----------------------------------------------------------------
# Copyright (c) 2023 by Robert Bosch GmbH. All rights reserved.

# The reproduction, distribution and utilization of this file as
# well as the communication of its contents to others without express
# authorization is prohibited. Offenders will be held liable for the
# payment of damages. All rights reserved in the event of the grant
# of a patent, utility model or design.

##################################################################
"""extract CAN data from DBC files and write to excel"""

import os
import re
import numpy as np
import pandas as pd
try:
    pd.set_option('future.no_silent_downcasting', True)
except:
    pass
from openpyxl import load_workbook
from win32com.client import Dispatch
import sys

try:
    from Control.logging_config import logger
except ImportError:
    sys.path.append(os.getcwd() + r"\..\Control")
    from logging_config import logger

enum_pattern = re.compile(r'(\d+)\s+"([^"]+)"')
byte_pattern = re.compile(r'0x(\w+)')

#patterns to extract data from dbc
messageBO_pattern = r"^BO_\s*(\d+)\s*(\w+)\s*\:\s*(\d+)\s*(\w+)"
siggroup_pattern = r"^SIG_GROUP_\s(\d*)\s(\w*)\s\d*\s:"
initvalue_pattern = r'"GenSigStartValue"\sSG_\s(\d+)\s(\w+\s?\w*)\s(\d+\.?\d*?)\;$'
cycletime_pattern = r'"GenMsgCycleTime"\s*BO_\s*(\d+)\s*(\d+)\;'
transmission_mode_pattern = r'"GenMsgSendType"\s*BO_\s*(\d+)\s*(\d)\;'
dataid_pattern = r'"DataID"\s*BO_\s*(\d+)\s*(\d+)\;'
e2e_profile_pattern = r'"E2Eprofile"\s*BO_\s*(\d+)\s*"(.+)"\;'
longSignal_pattern = r'"SystemSignalLongSymbol"\s*SG_\s*(\d+)\s*(\w+\s?\w*)\s\"(\w+)"\;'
valuetable_pattern = r'^VAL_\s(\d+)\s(\w+)\s'
valuetype_pattern = r'^SIG_VALTYPE_\s*(\d+)\s*(.*)\s*\:\s*(\d+)\s*;$'
comment_pattern = r"CM_\s\SG_\s(.*)\s(.*)\s\"(.*)\"\;"
tr_type_enum_pattern = r'BA_DEF_\s+BO_\s+"GenMsgSendType"\s+ENUM\s+"(.*?)";'
signal_parameters_pattern = r"SG_\s+(\w+)\s*([Mm\d*]*)\s*:\s*(\d*)\|(\d*)\@(\d+)([-+])\s*\((.*)\,(.*)\)\s*\[(.*)\|(.*)\]\s*\"(.*)\""


dbc_directory = os.path.dirname(os.path.abspath(__file__)) + r"/../../../../CustomerPrj/Restbus/Database/DBC"

# columns = ["Name", "group", "pdu", "PDU Type","Message", "Multiplexing/Group", "Startbit", "Length [Bit]", "Byte Order",
#            "Value Type", "Initial Value", "Factor", "Offset", "Minimum", "Maximum", "Unit",
#            "Value Table", "Comment", "Message ID", "Cycle Time [ms]", "texttable", "texttable values", 'EndToEndProtection',"max_value",
#            "dlc", "variant", "Block_size", "Address_formate","Padding_active", "STMin","MAXFC_wait"]  #,"Block_size", "Address_formate","Padding_active", "STMin","MAXFC_wait","Message ID"
columns = ("Signal",
                    "Signal_Group",
                    "PDU",
                    "PDU_Type",
                    "Payload_PDU_Type",
                    "Payload_PDU",
                    "Selector_Field_Signal",
                    "Message",
                    "Message ID",
                    "Header ID",
                    "Startbit",
                    "PDU_Length [Byte]",
                    "Signal_Length [Bit]",
                    "Payload_PDU_Length [Byte]",
                    "Signal Base Type",
                    "Initial Value",
                    "max_value",
                    "Transmission Mode",
                    "Cycle Time [ms]",
                    "texttable",
                    "texttable values",
                    "Value Type",
                    "Comment",
                    "dlc",
                    "Payload PDU DLC",
                    "variant",
                    "Value Table",
                    "EndToEndProtection",
                    "Byte Order",
                    "Factor",
                    "Offset",
                    "Minimum",
                    "Maximum",
                    "Unit",
                    'Cantp_Pdu_Type',
                    'Block_size',
                    'Address_formate',
                    'Padding_active',
                    'STMin',
                    'MAXFC_wait'
                    )

col_index = {key.lower():columns.index(key) for key in columns}

mess_trans_str = "BO_TX_BU_ "  # Message Transmitter

pd.set_option('display.width', 400)
pd.set_option('display.max_columns', 10)
dict_file_dfs = {}
dict_bo_tx_bu_dbc = {}
previous_file = ""

main_sheets = ["DBCmapping", "ArxmlMapping", "ETHArxmlMapping", "SysVarDatabase", "SysVarDatabaseETH"]  # "ini_out"

def declare_variables():
    """ declare global variables"""
    global dict_comment_list, dict_multiplexer, dict_byte_order, dict_raw_initial, dict_tr_type, dict_cycle_time, \
        list_value_table, dict_value_type, dict_texttable_values, dict_longSig, dict_BO_TX_BU, dict_e2e_information

    dict_comment_list = {}#dict to store comment line from dbc
    dict_multiplexer = {} #dict to store msgid : multiplexer pair
    dict_byte_order = {"Intel":[], "Motorola":[]}
    dict_raw_initial = {}  #for storing initial value as raw from dbc
    #dict_tr_type_enum = {} #enum for send type
    dict_tr_type = {} # send type event or periodic
    dict_cycle_time = {}  #for storing cycletime
    dict_e2e_information = {} 
    list_value_table = []  #for storing value table from dbc
    dict_value_type = {"IEEE Float":[], "IEEE Double":[]} #for storing value type from dbc
    dict_texttable_values = {}
    dict_longSig = {}
    dict_BO_TX_BU = {}

def load_excel_sheet():  # output workbook
    """ load excel into workbook"""
    script_path = os.path.dirname(os.path.abspath(__file__))
    autosar_directory = script_path + r"/../../../../CustomerPrj/Restbus/"

    wb = load_workbook(autosar_directory + r"\Autosar_Gen_Database.xlsx")
    sheet_list = wb.sheetnames
    return wb, sheet_list, autosar_directory


def create_list_from_column_excel(wb, find_column_name, sheet):
    """
    creates value list from given column

    Args:
      wb (workbook): excel workbook
      find_column_name (str): column name to extract
      sheet (sheet): sheet name

    Returns:
        list: list of values
    """
    ws = wb.get_sheet_by_name(sheet)
    list = []
    for col in ws.iter_cols(min_row=1, max_col=ws.max_column, max_row=1):
        for cell in col:
            if cell.value == find_column_name:
                dbc_name_col = cell.column
    for row in ws.iter_rows(min_row=2, min_col=dbc_name_col, max_col=dbc_name_col, max_row=ws.max_row):
        for cell in row:
            if (cell.value==0) or (cell.value):
                list.append(cell.value)
    return list


def delete_excel_sheets(wb, sheet_list, autosar_directory):
    """
    delete old excel sheets when starting new run

    Args:
      wb (workbook): excel workbook
      sheet_list (list): sheet list
      autosar_directory (str): excel path

    """
    for sheet in sheet_list:
        if (sheet.lower().endswith('_dbc')) and (sheet not in main_sheets):
            del wb[sheet]
    wb.save(autosar_directory + "\\" + "Autosar_Gen_Database.xlsx")
    logger.info(f"Deleted old DBC CAN node sheets")


def get_node_name(wb, sheet, direction=False):  # input workbook, output list with all dbc paths
    """
    gets DBC and ARXML mapping info into lists and dicts and returns

    Args:
      wb (workbook): excel workbook
      sheet (str): required mapping sheet
      direction (bool):  (Default value = False) True/False

    Returns:
        list_node, dict_node, dict_ecu_name, list_tx_messages,list_block_messages
    """
    list_autosar_database_network = create_list_from_column_excel(wb, "network_name", sheet)
    list_autosar_database_node = create_list_from_column_excel(wb, "canoe_ecu_node", sheet)
    list_dbc_network_index = create_list_from_column_excel(wb, "canoe_ecu_node_index", sheet)
    list_tx_messages = create_list_from_column_excel(wb, "pass_messages", sheet)
    list_block_messages = create_list_from_column_excel(wb, "block_messages", sheet)
    if direction:
        signal_direction = create_list_from_column_excel(wb, "direction", sheet)

    list_dbc_node = create_list_from_column_excel(wb, "dbc_network_node", sheet)
    if "dbcmapping" in sheet.lower():
        file = "_dbc"
    elif "arxmlmapping" in sheet.lower():
        file = "_arxml"
    else:
        file=""
    list_dbc_node_fixed = []
    for data in list_dbc_node:
        data_fix = data.split(';')
        list_dbc_node_fixed.append(data_fix)

    list_node = []
    for (elem1, elem2) in zip(list_autosar_database_network, list_autosar_database_node):
        list_node.append(elem1 + "_" + elem2 + file)

    dict_node = dict(zip(list_node, list_dbc_network_index))

    dict_ecu_name = dict(zip(list_node, list_dbc_node_fixed))
    if direction:
        return list_node, dict_node, dict_ecu_name, list_tx_messages, list_block_messages, signal_direction
    else:
        return list_node, dict_node, dict_ecu_name, list_tx_messages,list_block_messages


def get_dbc_path(wb, db_map, sheet):
    """
    map dbc mapping info to dbc file paths

    Args:
      wb (workbook): autosar_gen excel workbook
      db_map (list): [(autosar_index, database_index),...]
      sheet (str): mapping sheet name

    Returns:
        autosar_mapping (list): [{index:file_path},...]
    """
    list_dbc_name = create_list_from_column_excel(wb,"db_file_name", sheet)
    #reading dbc file index column into a list
    list_dbc_name_index = create_list_from_column_excel(wb,"db_file_name_index",sheet)

    list_autosar_database_index = db_map
    list_dbc_path = []
    for elem in list_dbc_name:
        list_dbc_path.append(dbc_directory + "\\" + elem)

    dict_dbc_index_file=dict(zip(list_dbc_name_index,list_dbc_path))

    autosar_mapping = []
    for autosar_index, database_index in list_autosar_database_index:
        if type(database_index) is tuple:
            for i in database_index:
                mapped = {autosar_index:dict_dbc_index_file[i]}
                autosar_mapping.append(mapped)
        else:
            mapped = {autosar_index:dict_dbc_index_file[database_index]}
            autosar_mapping.append(mapped)
    return autosar_mapping


def create_excel_sheets(wb, autosar_directory, list_node):
    """
    creates sheet names with data filled in dbc mapping sheet

    Args:
      wb (workbook): autosar_gen excel workbook
      autosar_directory (str): excel path
      list_node (list): list of node name

    Returns:
        workbook with updated sheet names

    """
    logger.info("Creating excel column names based on DBCMapping sheet")
    for elem in reversed(list_node):
        wb.create_sheet(elem, 0)
    wb.save(autosar_directory + "\\" + "Autosar_Gen_Database.xlsx")

def create_df_dbc(dbc_file, ecu_name, dict_ecu_name, tx_messages,block_messages):
    """
    Creates dataframe from dbc files

    Args:
      dbc_file: dbc file names
      ecu_name: 
      dict_ecu_name: 
      tx_messages: pass message list
      block_messages: block message list

    Returns:
            parsed_array : extracted data parsed into an array
    """
    global previous_file
    logger.info(f"++++++++++ Start parsing of the DBC ++++++++++")
    msg_list = []
    list_ecu_name = []
    extracted = False
    for key, value_list in dict_ecu_name.items():
        for value in value_list:
            list_ecu_name.append(value)
    logger.info(f"DBC file --> {dbc_file}")
    if dbc_file not in dict_file_dfs.keys():
        declare_variables()
        try: #try encoding with utf-8
            with open(dbc_file, 'r',encoding="utf-8") as file:
                for line in file.readlines():
                    msg_list.append(line.strip())
                    if line.strip().startswith(mess_trans_str):
                        extract_BO_TX_BU(line.strip())
                df = pd.DataFrame(msg_list)
        except UnicodeDecodeError as e: #if utf-8 throws an error then except part is executed
            with open(dbc_file, 'r') as file:
                for line in file.readlines():
                    msg_list.append(line.strip())
                    if line.strip().startswith(mess_trans_str):
                        extract_BO_TX_BU(line.strip())
                df = pd.DataFrame(msg_list)
        except FileNotFoundError as e:
            logger.error(f"File not found-->{e}\n check typo in dbcmapping sheet")
            raise Exception(f"File not found-->{e}\n check typo in dbcmapping sheet")
        except Exception as e:
            logger.error(f"Error occured while reading the file-->{e}")
            raise Exception(f"Error occured while reading the file-->{e}")
        
        dict_file_dfs[dbc_file]=df
        dict_bo_tx_bu_dbc[dbc_file] = dict_BO_TX_BU.copy()
    else:
        df = dict_file_dfs[dbc_file]
        if previous_file == dbc_file:
            extracted = True
        else:
            declare_variables()
    num_of_rows = df.shape[0]
    parsed_df = pd.DataFrame(columns=columns)
    mess_str = "BO_ "
    mess_block_str_formatted=[]
    mess_str_formatted=[]
    if tx_messages != "na":
        tx_messages_list = tx_messages.split(";")
        mess_str_formatted = [mess_str + str(int(s, 16)) + " " for s in tx_messages_list]
        mess_trans_str_formatted = [mess_trans_str + str(int(s, 16)) + " " for s in tx_messages_list]
        message_search_format = tuple(mess_str_formatted)
        message_trans_search_format = tuple(mess_trans_str_formatted)
    else:
        message_search_format = mess_str
        message_trans_search_format = mess_trans_str
        # if block_messages != "na":
        #     block_messages_list = block_messages.split(";")
        #     mess_block_str_formatted = [str(int(s, 16)) for s in block_messages_list]

    no_signal_messages_default_values = {'Signal':'-', 'Startbit': 0, 'Signal_Length [Bit]': 0, 'Factor': 1, 'Offset':0, 'Minimum':0, 'Maximum':255, 'max_value':255}
    #this defaults are not as per dbc, this is just to avoid warnings in generation since this type of messages details are not needed for generation

    for i in range(num_of_rows):
        current_line = df[0].values[i].strip()
        if current_line.startswith('SG_'):
            continue
        dict_values = {key:"" for key in columns}

        if current_line.startswith(message_search_format):
            sender_ecu_name = current_line.split(" ")[-1]
            msg_id = current_line.split(" ")[1]

            #if msg_id in mess_block_str_formatted:#commented, so that blocked msgs are also extracted, but will be handled in update_sysvartab and create_nodes py files
                #continue

            #ecu_msgID = sender_ecu_name + msg_id

            if ((ecu_name == sender_ecu_name) or (ecu_name in dict_bo_tx_bu_dbc[dbc_file].get(msg_id, []))) and ecu_name != "RBS":
                dict_values = extract_BO(current_line, dict_values)
                if not(df[0].values[i+1].startswith("SG_")):
                    dict_values.update(no_signal_messages_default_values)
                    parsed_df.loc[len(parsed_df)] = dict_values
                else:
                    for j in range(i+1, num_of_rows):
                        if df[0].values[j].startswith('SG_ '):
                            dict_values = extract_SG(df[0].values[j], dict_values)
                            parsed_df.loc[len(parsed_df)] = dict_values
                        elif df[0].values[j].startswith("BO_"):
                            break

            # elif ((msg_id in dict_BO_TX_BU.keys()) and (ecu_name in dict_BO_TX_BU[msg_id])):
            #     #ecu_name = dict_BO_TX_BU[ecu_msgID]
            #     dict_values = extract_BO(current_line, dict_values)
            #     if not(df[0].values[i+1].startswith("SG_")):
            #         print(dict_values)
            #         dict_values.update(no_signal_messages_default_values)
            #         parsed_df.loc[len(parsed_df)] = dict_values
                    
                    
            #     else:    
            #         for j in range(i + 1, num_of_rows):
            #             if df[0].values[j].startswith('SG_ '):
            #                 dict_values = extract_SG(df[0].values[j], dict_values)
            #                 parsed_df.loc[len(parsed_df)] = dict_values
            #             elif df[0].values[j].startswith("BO_"):
            #                 break

            if ecu_name == "RBS":
                if sender_ecu_name not in list_ecu_name:
                    dict_values = {key:"" for key in columns}
                    dict_values = extract_BO(df[0].values[i].strip(), dict_values)
                    if not(df[0].values[i+1].startswith("SG_")):
                        dict_values.update(no_signal_messages_default_values)
                        parsed_df.loc[len(parsed_df)] = dict_values
                    else:
                        for j in range(i + 1, num_of_rows):
                            if df[0].values[j].startswith('SG_ '):
                                dict_values = extract_SG(df[0].values[i + j], dict_values)
                                parsed_df.loc[len(parsed_df)] = dict_values
                            elif df[0].values[j].startswith("BO_"):
                                break

        elif not(extracted) and (current_line.startswith("CM_ SG_")):     #to extract comment from dbc
            if not(current_line.endswith(";")):      #to extract multiline comment
                n=i
                while True:
                    n = n + 1
                    if df[0].values[n].endswith(";"):
                        current_line += ' ' + df[0].values[n].strip()
                        break
                    else:
                        current_line += ' ' + df[0].values[n].strip()
            extract_CM(current_line)
        elif dict_tr_type_enum[dbc_file].get('GenMsgSendType', {})=={} and (current_line.startswith("BA_DEF_ BO_")):
            extract_tr_enum(current_line)

        elif not(extracted) and current_line.startswith("SIG_GROUP_"):      #to extract Signal_Group from dbc
            extract_SIG_GROUP(current_line)

        elif not(extracted) and current_line.startswith("BA_ "):
            extract_BA(current_line)

        elif not(extracted) and current_line.startswith("VAL_ "):  #to extract value table from dbc
            extract_VAL(current_line)

        elif not(extracted) and (current_line.startswith("SIG_VALTYPE_ ")): #to extract value type from dbc
            extract_SIG_VALTYPE(current_line)

    previous_file = dbc_file
    parsed_df = update_columns(parsed_df)
    parsed_array = parsed_df.to_numpy()
    # sort array by column startbit, Message & Message ID
    parsed_array = parsed_array[np.lexsort((parsed_array[:, col_index['startbit']], parsed_array[:, col_index['message']], parsed_array[:, col_index['message id']]))]
    return parsed_array

def raw_to_physicalValue(raw_value, factor, offset):
    """
    Converts raw_value from dbc to initial value

    Args:
      raw_value: raw value as extracted from dbc 
      factor: factor value as per dbc 
      offset: offset value as per dbc

    Returns:
            returns converted physical value
    """
    #converts raw_value from dbc to initial value
    return (raw_value * factor) + offset

def physicalValue_to_raw(physical_value, factor, offset):
    """
    Converts initial value to raw value(currently this function is not used)

    Args:
      physical_value: init value
      factor: factor value as per dbc 
      offset: offset value as per dbc

    Returns:
            returns converted raw value
    """
    #converts initial value to raw value (function not uesd)
    return (physical_value - offset)/factor

def update_columns(df):
    """
    Updates extracted data into excel sheet columns

    Args:
      df: dataframe

    Returns:
            dataframe with updated column data
    """
    for index, row in df.iterrows():
        sig = row["Signal"]
        msgid = row["Message ID"]
        msg_name = row["Message"]
        if sig!='-':
            factor = float(row["Factor"])
            offset = float(row["Offset"])
        sig_msgid = sig + msgid
        df.at[index,"Block_size"] = None    # Added None for 5 CANTP parameter to make architecture similar for Arxml and DBC
        df.at[index,"Address_formate"] = None
        df.at[index,"Padding_active"] = None
        df.at[index,"STMin"] = None
        df.at[index,"MAXFC_wait"] = None

        if msgid.upper() in dict_e2e_information.keys():
            df.at[index,"EndToEndProtection"] = str(dict_e2e_information.get(msgid.upper(), {}).get('dataid','')) + '::' + str(dict_e2e_information.get(msgid.upper(), {}).get('profile',''))

        if sig_msgid in dict_comment_list.keys():
            df.at[index,"Comment"] = df.at[index,"Comment"] + dict_comment_list[sig_msgid]

        if msgid in dict_multiplexer.keys():
            df.at[index, "Signal_Group"] = dict_multiplexer[msgid]
        else:
            df.at[index, "Signal_Group"] = ""

        for key,value in dict_value_type.items():
            if sig_msgid in value:
                df.at[index, "Value Type"] = key

        for key,value in dict_byte_order.items():
            if sig_msgid in value:
                df.at[index, "Byte Order"] = key

        if sig_msgid in dict_raw_initial.keys():
            df.at[index, "Initial Value"] = raw_to_physicalValue(dict_raw_initial[sig_msgid], factor, offset)
        else:
            df.at[index, "Initial Value"] = 0

        #filling transmission type
        if msgid in dict_tr_type.keys():
            df.at[index, "Transmission Mode"] = dict_tr_type[msgid]
        else:
            df.at[index, "Transmission Mode"] = None

        #filling cycle time
        if msgid in dict_cycle_time.keys():
            df.at[index, "Cycle Time [ms]"] = dict_cycle_time[msgid]
        else:
            df.at[index, "Cycle Time [ms]"] = 0

        #filling value table
        if sig_msgid not in dict_texttable_values.keys():
            df.at[index, "Value Table"] = '<none>'
        else:
            df.at[index, "Value Table"] = 'VtSig_' + sig
            df.at[index, "texttable"] = 'texttable'
            df.at[index, "texttable values"] = dict_texttable_values[sig_msgid]

        #filling variants
        if msg_name in dict_variants["a_variant"]:
            df.at[index, "variant"] = "a_variant"
        elif msg_name in dict_variants["b_variant"]:
            df.at[index, "variant"] = "b_variant"
        else:
            df.at[index, "variant"] = "Common"

        #filling long signal names
        if sig_msgid in dict_longSig.keys():
            df.at[index, "Signal"] = dict_longSig[sig_msgid]
    return df


def extract_BO_TX_BU(current_line):
    """
    Extracts transmit message with matched patterns 

    Args:
      current_line: hold the current line where pattern is matched

    Returns:
            returns extracted message in dict
    """
    matchObj = re.search(r"^BO_TX_BU_\s*(\d+)\s*:\s*([\w,]+)\s*;",current_line)
    if matchObj!=None:
        dict_BO_TX_BU[str(matchObj.group(1))] = matchObj.group(2).strip().split(",")


        #dict_BO_TX_BU[str(matchObj.group(2))+str(matchObj.group(1))] = str(matchObj.group(3))

def extract_tr_enum(current_line):
    """
    Extracts transmission type enum with matched pattern

    Args:
      current_line: hold the current line where pattern is matched

    Returns:
            returns extracted in dict
    """
    matchObj = re.search(tr_type_enum_pattern, current_line)
    if (None != matchObj):
        enum_string = matchObj.group(1)
        enum_list = enum_string.split('","')
        enum_dict = {i: value.strip('"') for i, value in enumerate(enum_list)}

        dict_tr_type_enum[dbc_file]['GenMsgSendType'] = enum_dict.copy()


def extract_CM(current_line):
    """
    Extracts comments with matched pattern

    Args:
      current_line: hold the current line where pattern is matched

    Returns:
            returns extracted comments in dict
    """
    matchObj = re.search(comment_pattern, current_line)
    if (None != matchObj):
        msg_id = hex(int(matchObj.group(1))).upper()
        signal_name = matchObj.group(2)
        sig_comment = matchObj.group(3)
        dict_comment_list[signal_name + msg_id] = sig_comment

def extract_SIG_GROUP(current_line):
    """
    Extracts signal group with matched pattern

    Args:
      current_line: hold the current line where pattern is matched

    Returns:
            returns extracted signal group in dict
    """
    matchObj = re.search(siggroup_pattern, current_line)
    if (matchObj != None):
        msgid = hex(int(matchObj.group(1))).upper()
        multiplexer = matchObj.group(2)
        if (msgid not in dict_multiplexer.keys()):
            dict_multiplexer[msgid] = multiplexer

def extract_SIG_VALTYPE(current_line):
    """
    Extracts signal value type with matched pattern

    Args:
      current_line: hold the current line where pattern is matched

    Returns:
            returns extracted value type in dict
    """
    matchObj = re.search(valuetype_pattern, current_line)
    if (matchObj != None):
        msgid = hex(int(matchObj.group(1))).upper()
        sig_name = str(matchObj.group(2).rsplit(' ')[0])
        sigtype = int(matchObj.group(3))
        if sigtype == 1:
            if sig_name + msgid not in dict_value_type['IEEE Float']:
                dict_value_type['IEEE Float'].append(sig_name + msgid)
        elif sigtype == 2:
            if sig_name + msgid not in dict_value_type['IEEE Double']:
                dict_value_type['IEEE Double'].append(sig_name + msgid)

def extract_BA(current_line):
    """
    Extracts initial value, cycle time and long value signal

    Args:
      current_line: hold the current line where pattern is matched

    Returns:
            returns extracted initial value, cycle time and long value signal in dict
    """
    matchObj_ini = re.search(initvalue_pattern, current_line)  # to extract intial value from dbc
    matchObj_transmission_type = re.search(transmission_mode_pattern, current_line) #Event or Cyclic
    matchObj_cytm = re.search(cycletime_pattern, current_line)  # to extract cycle time from dbc
    matchObj_longSig = re.search(longSignal_pattern, current_line)  # to extract long signal names
    matchobj_dataid = re.search(dataid_pattern, current_line)
    matchobj_e2e_profile = re.search(e2e_profile_pattern, current_line)

    if (matchObj_ini != None):
        msgid = hex(int(matchObj_ini.group(1))).upper()
        sig_name = matchObj_ini.group(2)
        raw_value = float(matchObj_ini.group(3))
        if sig_name+msgid not in dict_raw_initial.keys():
            dict_raw_initial[str(sig_name+msgid)] = raw_value
    elif (matchObj_transmission_type != None) and dict_tr_type_enum[dbc_file].get('GenMsgSendType', {})!={}:
        msgid = hex(int(matchObj_transmission_type.group(1))).upper()
        tr_type = int(matchObj_transmission_type.group(2))
        tr_type = dict_tr_type_enum[dbc_file]['GenMsgSendType'].get(tr_type, str(tr_type))
        if 'event' in tr_type.lower():
            tr_type = 'Event'
        elif 'cyclic' in tr_type.lower() or 'periodic' in tr_type.lower():
            tr_type = 'Periodic'
        dict_tr_type[msgid] = tr_type
    elif (matchObj_cytm != None):
        msgid = hex(int(matchObj_cytm.group(1))).upper()
        cycle_time = int(matchObj_cytm.group(2))
        if msgid not in dict_cycle_time.keys():
            dict_cycle_time[msgid] = cycle_time
    elif (matchObj_longSig != None):
        msgid = hex(int(matchObj_longSig.group(1))).upper()
        sig_name = matchObj_longSig.group(2)
        sig_name_long = matchObj_longSig.group(3)
        if sig_name + msgid not in dict_longSig.keys():
            dict_longSig[str(sig_name + msgid)] = sig_name_long

    elif (matchobj_dataid != None):
        msgid = hex(int(matchobj_dataid.group(1))).upper()
        dataid = int(matchobj_dataid.group(2))
        if msgid not in dict_e2e_information.keys():
            dict_e2e_information[msgid] = {}
        
        dict_e2e_information[msgid]['dataid'] = dataid

    elif (matchobj_e2e_profile != None):
        msgid = hex(int(matchobj_e2e_profile.group(1))).upper()
        profile = matchobj_e2e_profile.group(2)
        if msgid not in dict_e2e_information.keys():
            dict_e2e_information[msgid] = {}
        
        dict_e2e_information[msgid]['profile'] = profile


def findStartBit(srt_bit, length):
    """
    Finds the start bit of a signal

    Args:
      srt_bit: start bit of signal
      length: length of the signal

    Returns:
            returns calculated start bit
    """
    while length != 1:
        if srt_bit % 8 == 0:
            srt_bit += 16
        length -= 1
        srt_bit -= 1
    return int(srt_bit)

def extract_VAL(current_line):
    """
    Extracts number & string of logic value

    Args:
      current_line: hold the current line where pattern is matched

    Returns:

    """
    matchObj = re.search(valuetable_pattern, current_line)
    if (matchObj != None):
        msgid = hex(int(matchObj.group(1))).upper()
        sig_name = str(matchObj.group(2))

    try:
        enum_matches = enum_pattern.findall(current_line)
        enum = []
        for value, description in enum_matches:
            enum.append(f"LogicalValue: {value} {description.strip()}")
        enum = '\n'.join(enum)
        if not enum:  # check enum values
            logger.warning(f"\tlogic_value error in line {current_line}")
        if sig_name + msgid not in dict_texttable_values.keys():
            dict_texttable_values[sig_name + msgid] = enum
    except Exception as exp:
        logger.warning(f"Error in extracting enum values from {current_line}: {exp}")

def extract_BO(line, dict_values):
    """
    Extracts message name, id and dlc values

    Args:
      line: hold the current line where pattern is matched
      dict_values: holds the values of message name, id and dlc values

    Returns:
            returns message name, id and dlc values in dictionary
    """
    matchObj = re.search(messageBO_pattern, line)
    if (matchObj!=None):
        dict_values["Message"] = matchObj.group(2)
        dict_values["Message ID"] = hex(int(matchObj.group(1))).upper()
        dict_values["dlc"] = to_canfd_dlc(int(matchObj.group(3)))
    return dict_values


def extract_SG(line, dict_values):
    """
    Extracts signal name and its parameters

    Args:
      line: hold the current line where pattern is matched
      dict_values: holds the values of message name, id and dlc values

    Returns:
            returns signal name and its parameters in dictionary
    """
    #pattern to extract signal name and all parameters associated with signal
    matchObj = re.search(signal_parameters_pattern, line)

    if (None != matchObj):
        dict_values["Signal"] = matchObj.group(1).rsplit(' ')[0].strip()
        dict_values["Unit"] = matchObj.group(11)
        dict_values["Factor"] = float(matchObj.group(7))
        dict_values["Offset"] = float(matchObj.group(8))
        dict_values["Minimum"] = float(matchObj.group(9))
        dict_values["Maximum"] = float(matchObj.group(10))
        dict_values["Signal_Length [Bit]"] = int(matchObj.group(4))
        dict_values["Value Type"] = matchObj.group(6).replace('+', 'Unsigned').replace('-', 'Signed')
        dict_values["Startbit"] = int(matchObj.group(3))

        byte_order = int(matchObj.group(5))     #byte order
        sig_msgid = dict_values["Signal"] + str(dict_values["Message ID"])
        if byte_order == 0:
            dict_values["Startbit"] = findStartBit(dict_values["Startbit"], dict_values["Signal_Length [Bit]"])
            if (sig_msgid not in dict_byte_order["Motorola"]):
                dict_byte_order["Motorola"].append(sig_msgid)
        elif (byte_order == 1) and (sig_msgid not in dict_byte_order["Intel"]):
            dict_byte_order["Intel"].append(sig_msgid)
        dict_values["max_value"] = float(2**int(dict_values["Signal_Length [Bit]"]) - 1)

        #check if multiplexer is present
        if matchObj.group(2).strip() != "":
            multiplexer = matchObj.group(2).strip()
            if 'M' in multiplexer:
                dict_values["Selector_Field_Signal"] = 'yes'
            elif 'm' in multiplexer:
                multiplexer = multiplexer.replace('m','')
                dict_values["Selector_Field_Signal"] = ''
                dict_values['Comment'] = 'multiplex_code = ' + str(multiplexer) + '.\n'
        else:
            dict_values["Selector_Field_Signal"] = ''


    return dict_values

def extractVariant(autosar_path,sheet):
    """
    Extracts variant selection data from mentioned sheets

    Args:
      autosar_path: path of excel sheet
      sheet: sheet names

    Returns:
           returns dict with variant name
    """
    try:
        df = pd.read_excel(autosar_path, sheet_name=sheet, keep_default_na=False)
        variant_col = list(df["variant_selection"].values)
        dict_variants = {}
        for row in variant_col:
            if row.strip():
                temp = row.split("=")
                key = temp[0].strip()
                if key and (key not in dict_variants.keys()):
                    value1 = temp[1].strip().split(";")
                    value = [val.strip() for val in value1 if val.strip()]
                    dict_variants[key] = value
        logger.info(f"Extracted variant data from {sheet} sheet")
        if dict_variants == {}:
            return {"a_variant":[],"b_variant":[]}
        return dict_variants
    except Exception as e:
        logger.error(f"Failed to extract variants data from {sheet} sheet --> {e}")
        raise Exception(f"Failed to extract variants data from {sheet} sheet --> {e}")


def to_canfd_dlc(msg_dlc):
    """
    reference:-  https://elearning.vector.com/mod/page/view.php?id=368

    Args:
      msg_dlc: 

    Returns:

    """
    msg_dlc=int(msg_dlc)
    if msg_dlc<=8:
        return msg_dlc
    elif (msg_dlc>8 and msg_dlc<=12):
        return 9
    elif (msg_dlc>12 and msg_dlc<=16):
        return 10
    elif (msg_dlc>16 and msg_dlc<=20):
        return 11
    elif (msg_dlc>20 and msg_dlc<=24):
        return 12
    elif (msg_dlc>24 and msg_dlc<=32):
        return 13
    elif (msg_dlc>32 and msg_dlc<=48):
        return 14
    elif (msg_dlc>48 and msg_dlc<=64):
        return 15
    

def write_dbc_in_excel(wb, dict_node, dict_ecu_name, autosar_directory, tx_messages,block_messages):
    """
    Writes all the extracted data from dbc into excel sheet

    Args:
      wb: work book
      dict_node: node name
      dict_ecu_name:  ecu name 
      autosar_directory: excel path
      tx_messages: pass message lists
      block_messages: block message lists

    Returns:
            returns updated in excel sheet
    """
    global dict_variants, dbc_file, dict_tr_type_enum
    dict_variants = extractVariant(autosar_directory + "\\" + "Autosar_Gen_Database.xlsx", "DBCmapping")
    dict_tr_type_enum = {}
    dbc_file = '-'
    for pair in dict_node:
        for index, dbc_file in pair.items():
            if dbc_file not in dict_tr_type_enum:
                dict_tr_type_enum[dbc_file] = {"GenMsgSendType":{}}
            ws = wb.worksheets[index]
            tx_message_list = str(tx_messages[index])
            block_messages_list = str( block_messages[index])
            ecu_name_list = dict_ecu_name[ws.title]
            max = ws.max_row
            parsed_array_list = []
            for ecu_name in ecu_name_list:
                parsed_array_list.append(create_df_dbc(dbc_file, ecu_name.strip(), dict_ecu_name, tx_message_list,block_messages_list))
            shape = parsed_array_list[0].shape

            for x in range(shape[1]):
                ws.cell(row=1, column=x + 1).value = columns[x]
            k = 0
            for parsed_array in parsed_array_list:
                shape = parsed_array.shape
                row_cnt = k
                for i in range(shape[0]):
                    k = k + 1
                    for j in range(shape[1]):
                        if max == 1:
                            ws.cell(row=i + row_cnt + 2, column=j + 1).value = parsed_array[i][j]
                        else:
                            ws.cell(row=i + row_cnt + max+1, column=j + 1).value = parsed_array[i][j]
        wb.save(autosar_directory + "\\" + "Autosar_Gen_Database.xlsx")
    wb.save(autosar_directory + "\\" + "Autosar_Gen_Database.xlsx")


def generate_dbc_map(autosar_excel, sheet):
    """
    Creates dbc mapping sheet

    Args:
      autosar_excel: path of the excel
      sheet: sheet name

    Returns:

    """
    # open Autosar database Excel
    logger.info(f"++++++++++ Start creating DBC map from {sheet} sheet ++++++++++")
    try:
        df = pd.read_excel(autosar_excel, sheet_name=sheet)
        autosar_database_index = df["canoe_ecu_node_index"].fillna("").values
        autosar_database_index = [int(i) for i in autosar_database_index if str(i)]
        database_index = df["mapping_canoe_ecu_to_db"].fillna("").values
        database_index = [i for i in database_index if str(i)]
        database_index_fixed = []
        for data in database_index:
            if type(data) == str:
                data_fix = tuple(map(int, data.split(';')))
                database_index_fixed.append(data_fix)
            else:
                database_index_fixed.append(int(data))
        logger.info(f"++++++++++ Created DBC map from {sheet} sheet ++++++++++")

        return tuple(zip(autosar_database_index, database_index_fixed))
    except Exception as e:
        logger.error(f"Failed to create DBC map from {sheet} sheet --> {e}")
        raise Exception(f"Failed to create DBC map from {sheet} sheet --> {e}")


def extract_vehicle_variants(autosar_excel):
    """
    Extract vehicle variants

    Args:
      autosar_excel: path of the excel

    Returns:
        vehicle variants (list)
    """

    # open Autosar database Excel
    logger.info(f"++++++++++ Extract vehicle variants  ++++++++++")
    try:
        df = pd.read_excel(autosar_excel, sheet_name="DBCmapping")
        variant_list_dbc = df["vehicle_selection"].dropna().values.tolist()
        df = pd.read_excel(autosar_excel, sheet_name="ArxmlMapping")
        variant_list_arxml = df["vehicle_selection"].dropna().values.tolist()

        if variant_list_dbc:
            return variant_list_dbc
        else:
            return variant_list_arxml

    except Exception as e:
        logger.error(f"Failed to extract vehicle variants --> {e}")
        raise Exception(f"Failed to extract vehicle variants --> {e}")
# execute main function
def update_autosar_main(db_map):
    """
    Calls all the functions

    Args:
      db_map: dbc mapping sheet 

    Returns:

    """
    wb, sheet_list, autosar_directory = load_excel_sheet()
    delete_excel_sheets(wb, sheet_list, autosar_directory)
    list_node, dict_node, dict_ecu_name, tx_messages, block_messages = get_node_name(wb, "DBCmapping")
    create_excel_sheets(wb, autosar_directory, list_node)
    dict_node = get_dbc_path(wb, db_map, "DBCmapping")
    write_dbc_in_excel(wb, dict_node, dict_ecu_name, autosar_directory,tx_messages,block_messages)

def external_call():
    """ """
    try:
        script_path = os.path.dirname(os.path.abspath(__file__))
        autosar_path = script_path + r'/../../../../CustomerPrj/Restbus/Autosar_Gen_Database.xlsx'
        update_autosar_main(generate_dbc_map(autosar_path, "DBCmapping"))
    except Exception as e:
        logger.error(f"Error occured in create_autosar -> {e}")
        raise Exception(e)


if __name__ == "__main__":
    external_call()

