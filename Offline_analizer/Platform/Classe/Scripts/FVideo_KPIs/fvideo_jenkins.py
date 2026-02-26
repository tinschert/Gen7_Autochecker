# -*- coding: utf-8 -*-
# @file fvideo_jenkins.py
# @author ADAS_HIL_TEAM
# @date 08-21-2023

##################################################################
# C O P Y R I G H T S
# ----------------------------------------------------------------
# Copyright (c) 2023 by Robert Bosch GmbH. All rights reserved.

# The reproduction, distribution and utilization of this file as
# well as the communication of its contents to others without express
# authorization is prohibited. Offenders will be held liable for the
# payment of damages. All rights reserved in the event of the grant
# of a patent, utility model or design.

##################################################################


import argparse
import os,sys

import os

import openpyxl
from openpyxl.styles import Font,Alignment
from openpyxl.drawing.image import Image
import matplotlib.pyplot as plt
import numpy as np


sys.path.append(os.path.dirname(os.path.abspath(__file__)) + r"\..\Jenkins_scripts")

sys.path.append(os.path.dirname(os.path.abspath(__file__)) + r"\..\Control")
from logging_config import logger
import datetime
import shutil
import subprocess
import pythoncom
import time
import canoe_client_1
from fvideo_kpi import external_call, get_scenario_target_type, copyExcelReport
import pandas as pd
try:
    pd.set_option('future.no_silent_downcasting', True)
except:
    pass
import xml.dom.minidom as minidom
import xml.etree.ElementTree as ET





def custom_sort_key(s):
    parts = s.split("_")
    if len(parts) >= 3:
        try:
            return int(parts[-2])
        except ValueError:
            pass
    return 0

def createSummaryReport(excel_files_path, output_file_path, column_suffix, scenario_type):
    destination_paths = {"ODET": r"\\abtvdfs2.de.bosch.com\ismdfs\iad\verification\ODS\ADAS_HIL_concept\Front_Video\FV_KPI\KPI_Results\ODET",
                         'TSR': r"\\abtvdfs2.de.bosch.com\ismdfs\iad\verification\ODS\ADAS_HIL_concept\Front_Video\FV_KPI\KPI_Results\TSR",
                         'LANES': r"\\abtvdfs2.de.bosch.com\ismdfs\iad\verification\ODS\ADAS_HIL_concept\Front_Video\FV_KPI\KPI_Results\LANE"}
    
    try:
        logger.info("-------Start Summary Report Creation----------")
        summary_dict = {}
        EGO_SPEED_LIST = []
        for file in os.listdir(excel_files_path):
            if file.endswith(".xlsx"):
                file_path = excel_files_path+r"\\"+file
                logger.info(f"Processing -> {file}")
                # get scenario name
                split = file.split("_")
                try:
                    if scenario_type.upper()=='ODET':
                        scenario_name = "_".join(split[:-4])
                        ego_speed = split[-4] + " KPH"
                        if ego_speed not in EGO_SPEED_LIST:
                            EGO_SPEED_LIST.append(ego_speed)
                    else:
                        scenario_name = "_".join(split[:-2])
                        ego_speed = split[-2] + " KPH"
                        if ego_speed not in EGO_SPEED_LIST:
                            EGO_SPEED_LIST.append(ego_speed)
                except:
                    continue
                if scenario_name not in summary_dict.keys():
                    summary_dict[scenario_name] = {}
                # summary_dict[scenario_name][ego_speed] = {}

                excel_file = pd.ExcelFile(file_path)
                sheets = excel_file.sheet_names
                for sheet in sheets:
                    df = pd.read_excel(excel_file, sheet_name=sheet)
                    df["Parameter_Name"] = df["Parameter_Name"].ffill()
                    df = df.fillna("")

                    for i, row in df.iterrows():
                        key = row["Parameter_Name"] + "_" + row["Detection_Range"] if row["Detection_Range"] else row[
                            "Parameter_Name"]
                        if key not in summary_dict[scenario_name].keys():
                            summary_dict[scenario_name][key] = {ego_speed: {"PASS/"+column_suffix: 0, "FAIL/"+column_suffix: 0}}
                        else:
                            if ego_speed not in summary_dict[scenario_name][key].keys():
                                summary_dict[scenario_name][key][ego_speed] = {"PASS/"+column_suffix: 0, "FAIL/"+column_suffix: 0}
                        summary_dict[scenario_name][key][ego_speed][row["Verdict"]+"/"+column_suffix] += 1


        #create Report
        EGO_SPEED_LIST.sort()
        df_dict = {}

        for sheet, data in summary_dict.items():
            #plot graph
            params = list(data.keys())
            count_range = np.arange(int(column_suffix)+1)  # iteration y axis
            fontsize = 45
            for speed in EGO_SPEED_LIST:
                try:
                    pass_data = [data[param][speed]['PASS/'+column_suffix] for param in params]
                    fail_data = [data[param][speed]['FAIL/'+column_suffix] for param in params]

                    bar_width = 0.42
                    index = np.arange(len(params))
                    fig, ax = plt.subplots(figsize=(60, 18))
                    pass_bars = plt.bar(index, pass_data, bar_width, label='PASS/'+column_suffix, color='#00cc00')
                    fail_bars = plt.bar(index + bar_width, fail_data, bar_width, label='FAIL/'+column_suffix, color='#ff4d4d')

                    plt.xlabel('Parameters', fontsize=fontsize)
                    plt.ylabel('No_of_Iterations', fontsize=fontsize)
                    plt.title(f'{speed} -> {sheet}', fontsize=fontsize)
                    plt.xticks(index + bar_width / 2, params, fontsize=fontsize - 10)

                    plt.legend(fontsize=25)
                    plt.yticks(count_range, fontsize=fontsize)
                    plt.ylim(0, 11)
                    ax.set_xticklabels(params, rotation=45, ha='right')

                    ax.spines['bottom'].set_color('black')
                    ax.spines['top'].set_color('black')
                    ax.spines['right'].set_color('black')
                    ax.spines['left'].set_color('black')
                    ax.xaxis.set_tick_params(width=5, size=10)
                    ax.yaxis.set_tick_params(width=2.5, size=8)

                    for bar in pass_bars + fail_bars:
                        height = bar.get_height()
                        if height == 0:
                            continue
                        ax.annotate('{}'.format(height),
                                    xy=(bar.get_x() + bar.get_width() / 2, height),
                                    xytext=(0, 3),  # 3 points vertical offset
                                    textcoords="offset points",
                                    ha='center', va='bottom', fontsize=28)
                    im_name = excel_files_path + r"\\" + sheet+"_"+str(speed).replace(" ","_")+ ".jpg"
                    plt.savefig(im_name, bbox_inches='tight')
                except Exception as e:
                    logger.warning(f"Error while generating graph for -> {speed}-> {sheet} --> {e}")
                    continue



            df_dict[sheet] = []
            columns = [""]
            headings = [""]

            first_time = True
            for parameter, kpi_data in data.items():

                temp = [parameter]

                for ego_speed, result in kpi_data.items():
                    if ego_speed not in columns:
                        for i in range(2):
                            columns.append(ego_speed)

                    temp.append(result["PASS/"+column_suffix])
                    temp.append(result["FAIL/"+column_suffix])

                if first_time:
                    first_time = False
                    for i in range(len(kpi_data.keys())):
                        headings.extend(["PASS/"+column_suffix, "FAIL/"+column_suffix])

                df_dict[sheet].append(temp)

            df_dict[sheet].insert(0, headings)
            df_dict[sheet].insert(0, columns)

        workbook = openpyxl.Workbook()
        for sheet_name, data in df_dict.items():
            sht_name = sheet_name
            if len(sheet_name) > 30:
                split = sheet_name.split("_")
                name = ""
                for text in split:
                    name = name + text[:3] + "_"
                sht_name = name[:-1]
            try:
                sheet = workbook.create_sheet(sht_name)
            except:
                continue

            sheet.append([sheet_name] * len(data[0]))
            # Populate the sheet with data
            for row in data:
                sheet.append(row)

            current_value = None
            start_column_index = 1
            end_column_index = 1

            # Iterate through the first row and merge consecutive identical cells
            for column_index, cell in enumerate(sheet[2], start=1):
                if cell.value == current_value:
                    end_column_index = column_index
                else:
                    if end_column_index > start_column_index:
                        sheet.merge_cells(start_row=2, start_column=start_column_index, end_row=2,
                                          end_column=end_column_index)
                    current_value = cell.value
                    start_column_index = column_index
            # Handle the last group of consecutive identical cells
            if end_column_index > start_column_index:
                sheet.merge_cells(start_row=2, start_column=start_column_index, end_row=2, end_column=end_column_index)

            sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(data[0]))

            # fit max width
            max_widths = {}
            for row in sheet.iter_rows(values_only=True):

                for column_index, cell_value in enumerate(row, start=1):
                    if cell_value == sheet_name:
                        continue
                    column_letter = openpyxl.utils.get_column_letter(column_index)
                    max_width = max(max_widths.get(column_letter, 0), len(str(cell_value)) + 1)
                    max_widths[column_letter] = max_width

            # Set the column width based on the maximum width needed
            for column_letter, max_width in max_widths.items():
                sheet.column_dimensions[column_letter].width = max_width

            # Make the first column bold
            for row in sheet.iter_rows(min_col=1, max_col=1):
                for cell in row:
                    cell.font = Font(bold=True)

            # Make the first two rows bold
            for row in sheet.iter_rows(min_row=1, max_row=3):
                for cell in row:
                    cell.font = Font(bold=True)

            for row in sheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            #attach images
            max_row = sheet.max_row+8
            files_list = os.listdir(excel_files_path)
            files_list = [f for f in files_list if f.endswith(".jpg")]
            files_list = sorted(files_list, key=custom_sort_key)
            for img in files_list:
                if sheet_name not in img:
                    continue
                img_path = excel_files_path+r"\\"+img
                imge = Image(img_path)

                imge.width = 1500
                imge.height = 450

                sheet.add_image(imge, 'A' + str(max_row))
                max_row += 30

        default_sheet = workbook['Sheet']
        workbook.remove(default_sheet)
        workbook.save(output_file_path)
        time.sleep(2)
        logger.info(f"-------Finished Summary Report Creation -> {output_file_path}----------")
        copyExcelReport(output_file_path, destination_paths[scenario_type])
    except Exception as e:
        logger.warning(f"Error while creating summary report -> {e}")
        return


def delete_files(directory_path, extension = ""):
    """
    

    Args:
      directory_path: 
      extension:  (Default value = "")

    Returns:

    """
    try:
        for filename in os.listdir(directory_path):
            file_path = os.path.join(directory_path, filename)
            if os.path.isfile(file_path):
                if extension:
                    if filename.lower().endswith(extension):
                        os.remove(file_path)
                else:
                    os.remove(file_path)

        logger.info(f"Trace folder cleaned -> {directory_path}")
    except Exception as e:
        logger.error(f"error occured while deleting trace files ->{e}")
        raise Exception(f"error occured while deleting trace files ->{e}")


def convert_BLF_to_MF4_FVideoKPI(canoeClient, src_folder, dest_folder):
    """
    

    Args:
      canoeClient: 
      src_folder: 
      dest_folder: 

    Returns:

    """
    try:
        converted_files = []
        if not os.path.exists(dest_folder):
            os.mkdir(dest_folder)
        for (dirpath, dirnames, filenames) in os.walk(src_folder):
            for file in filenames:
                if ".blf" in file.lower():
                    scenario_name_folder = "_".join(file.split("_")[:-1])
                    scnario_folder_path = dest_folder + r"\\" + scenario_name_folder
                    if not os.path.exists(scnario_folder_path):
                        os.mkdir(scnario_folder_path)
                    source_file = src_folder+r"\\"+file
                    destination_file = scnario_folder_path+r"\\"+file.replace(".blf",".mf4")
                    converted = canoeClient.convert_logged_file(destination_file=destination_file, source_file=source_file)
                    converted_files.append(destination_file)
        return converted_files
    except Exception as e:
        logger.error(f"ERROR while converting blf to mf4 -> {e}")

def convert_BLF_to_MDF_FVideoKPI(canoeClient, src_folder, dest_folder, required_messages_list):
    """
    converts blf file to signal based mdf file by applying required message filter
    Args:
        canoeClient:
        src_folder:
        dest_folder:
        required_messages_list:

    Returns:

    """
    try:
        converted_files = []
        if not os.path.exists(dest_folder):
            os.mkdir(dest_folder)
        for (dirpath, dirnames, filenames) in os.walk(src_folder):
            for file in filenames:
                if ".blf" in file.lower():
                    scenario_name_folder = "_".join(file.split("_")[:-1])
                    scnario_folder_path = dest_folder + r"\\" + scenario_name_folder
                    if not os.path.exists(scnario_folder_path):
                        os.mkdir(scnario_folder_path)
                    source_file = src_folder+r"\\"+file
                    destination_file = scnario_folder_path+r"\\"+file.replace(".blf",".mdf")
                    converted = canoeClient.convert_logged_file(destination_file=destination_file, source_file=source_file, filter_symbols_patterns=required_messages_list)
                    converted_files.append(destination_file)
        return converted_files
    except Exception as e:
        logger.error(f"ERROR while converting blf to mf4 -> {e}")

def get_required_symbols_pattern(path_platform, scenario_type):
    """
    return a list of required symbols defined in fvideo kpi database for filtering trace
    conversion
    Args:
        path_platform (str): path to main dir
        sheet (str): Database sheet to consider

    Returns (list): return a list of required symbols

    """
    
    if scenario_type=="ODET":
        sheet = "Objects"
    elif scenario_type=="TSR":
        sheet = "TrafficSigns"
    elif scenario_type.upper()=="LANES":
        sheet = "Lanes"
    else:
        logger.error(f"Invalid scenario type -> {scenario_type}")
        return None

    kpi_database_df = pd.read_excel(path_platform + r"\\CustomerPrj\\FVideo_KPIs\\FVideo_KPI_Database.xlsx",
                                    sheet_name=sheet)
    req_pattern_list = ['CarMaker::RB::GroundTruth::Line::nLines']
    for i, row in kpi_database_df.iterrows():
        if "raw_gt" in row["identifier"]:
            if scenario_type=='TSR':
                req_pattern_list.append(row["Message"] + "::" + row["Name"])
            elif scenario_type=='LANES':
                req_pattern_list.append(row["Message"])
            else:
                req_pattern_list.append(row["Message"] + "." + row["Name"])
        else:
            if scenario_type=='LANES':
                req_pattern_list.append(row["Message"].replace('_[]_[]',''))
            else:
                req_pattern_list.append(row["Message"]+"::"+row["Name"])
            

    return req_pattern_list


def run_canoe_test(cfg_file, path_platform, vxt_file_path,test_module, num_iter,ego_speed_list, target_speed,scenario_type):
    """
    

    Args:
      cfg_file: 
      path_platform: 
      vxt_file_path: 
      test_module: 
      num_iter: 
      ego_speed_list: 
      target_speed: 
      scenario_type: 

    Returns:

    """

    required_messages_list = get_required_symbols_pattern(path_platform, scenario_type)


    #logging_path = path_platform + r"\\CustomerPrj\\Traces\\FrameLogged_T{IncTrigger|001}.mf4"
    logging_folder_path = path_platform + r"\\CustomerPrj\\Traces\\FV_KPI_Scenarios\\" + scenario_type
    # if not(test_module_vxt_file.lower().endswith(".vxt")):
    #     test_module_vxt_file_suffix = test_module_vxt_file + ".vxt"

    evaluated_trace_folder = []
    # convert_logging_path = path_platform + r"\\CustomerPrj\\Traces\\MDF_logged.mdf"
    # converted = False  #just to confirm the conversion was successfull

    p = subprocess.Popen(["powershell.exe", path_platform + r"\Platform\Classe\Scripts\Git_sync_tool\Clear_cache.ps1"],
                         stdout=sys.stdout)
    p.communicate()
    pythoncom.CoInitialize()
    start_canoe = time.time()

    logger.info("Starting CANoe via COM interface")
    canoeClient = canoe_client_1.CANoeClient()
    try:
        canoeClient.openConfiguration(cfg_file)
        canoe_started = time.time()
        logger.info(f"CANoe loaded for {canoe_started - start_canoe} seconds")
        

        for ego_speed in ego_speed_list:
            evaluated_trace_file_list = []
            logger.info(f"TC_INFO#:#********* EGO_SPEED = {str(ego_speed)}kmph *********#:#")
            updateTestModule(vxt_file_path, scenario_type, ego_speed, target_speed)

            canoeClient.activate_measurementSetup_block("Logging")
            # logging_path = canoeClient.set_logging_path(logging_path)
            time.sleep(1)
            #canoeClient.enable_logging_filter([1,2,3,4,7])
            canoeClient.enableTestsModules('FVideoKPI_CM_1V1D_delta5')

            canoeClient.startMeasurement()

            #dasy power check
            if "_OD" in str(cfg_file).upper():
                canoeClient.setVariableToValue("hil_ctrl", "hil_mode", 2)
                canoeClient.setVariableToValue("hil_ctrl", "configuration_od", 1)
                time.sleep(5)
                current_consumption = canoeClient.getSysVariable("PS", "current_display_Ch1")
                if current_consumption.Value > 0.12:
                    logger.info(f"DAsy is powered and alive.Proceeding with test case execution.")
                else:
                    logger.warning(f"Dasy is not powered.Current consumption is {current_consumption.Value}, Trying for 2nd time...")
                    canoeClient.stopMeasurement()
                    time.sleep(5)
                    canoeClient.startMeasurement()
                    time.sleep(3)
                    canoeClient.setVariableToValue("hil_ctrl", "hil_mode", 2)
                    canoeClient.setVariableToValue("hil_ctrl", "configuration_od", 1)
                    time.sleep(5)
                    current_consumption = canoeClient.getSysVariable("PS", "current_display_Ch1")
                    if current_consumption.Value < 0.12:
                        raise Exception(f"Dasy is not powered.Current consumption is {current_consumption.Value}")
                

            logger.info(f"Test Module: {test_module}")

            canoeClient.setVariableToValue("hil_ctrl", "jenkins_control", 1)
            for i in range(num_iter):
                separator = "-"*25
                logger.info(f"{separator} Iteration:{i+1} {separator}")
                #canoeClient.setVariableToValue("hil_ctrl", "trace_logging_start_mf4", 1.0)
                time.sleep(1.5)
                logger.info(f"Started Running...")
                verdict = canoeClient.executeTestsInTestModules(test_module, kpi_run=True)
                logger.info(f"Finished Running")
                if verdict:
                    logger.info(f"TC_INFO#:#Iteration {i+1} Verdict: Failed#:#")
                else:
                    logger.info(f"TC_INFO#:#Iteration {i + 1} Verdict: Passed#:#")
                #canoeClient.setVariableToValue("hil_ctrl", "trace_logging_start_mf4", 0.0)
                
                time.sleep(20)

            canoeClient.stopMeasurement()

            logger.info(f"Finished running {test_module} for {str(num_iter)} no of times")

            time.sleep(5)

            mf4_converted_folder = logging_folder_path+r"\\MF4_CONVERTED"
            mf4_converted_files = convert_BLF_to_MDF_FVideoKPI(canoeClient, src_folder = logging_folder_path,
                                                               dest_folder = mf4_converted_folder, required_messages_list=required_messages_list)

            for folder_name in os.listdir(mf4_converted_folder):
                folder_path = os.path.join(mf4_converted_folder, folder_name)
                if os.path.isdir(folder_path):
                    if folder_name not in evaluated_trace_folder:
                        evaluated_trace_folder.append(folder_name)
                        if 'ODET' in scenario_type.upper():
                            output_file_path = path_platform + r"\\" + folder_name.strip() + "_ego" + str(ego_speed) + "_tgt" + str(target_speed) + "_" + "kmph.xlsx"
                        else:
                            output_file_path = path_platform + r"\\" + folder_name.strip() + "_ego" + str(ego_speed) + "kmph.xlsx"

                        try:
                            scenario_target_type = get_scenario_target_type(folder_name, scenario_type,fvideo_kpi_specific=True)
                            logger.info(f"Scenario {folder_name} mapped to verdict sheet -> {scenario_target_type}")
                            if type(scenario_target_type)==tuple:
                                logger.warning(f"Couldn't figure out scenario_target_type from scanario_name")
                                continue
                        except:
                            logger.warning(f"Couldn't figure out scenario_target_type from scanario_name")
                            continue

                        evaluated_trace_file_list.extend(
                            external_call(folder_path, scenario_type.upper(),
                                          scenario_target_type, output_file_path, evaluated_trace_file_list))
                        time.sleep(5)
                    else:
                        continue


            delete_files(logging_folder_path, extension = ".blf")
            logger.info(f"TC_INFO#:#*************************************#:#")
            time.sleep(5)

        createSummaryReport(path_platform, path_platform+r"\\SUMMARY_REPORT.xlsx",str(num_iter), scenario_type.upper())

        return ""
    
    except Exception as e:
        canoeClient.stopMeasurement()
        createSummaryReport(path_platform, path_platform + r"\\SUMMARY_REPORT.xlsx", str(num_iter), scenario_type.upper())
        logger.error(f"Stop at exception --> {e}")
        raise e
    # finally:
    #     canoeClient.stopMeasurement()
    #     canoeClient.quitCanoe(False)
    #     canoeClient = None


def get_canoe_cfg(target_path):
    """
    

    Args:
      target_path: 

    Returns:

    """
    for (dirpath, dirnames, filenames) in os.walk(target_path):
        for file in filenames:
            if file.endswith(".cfg"):
                logger.info(f"CANoe configuration path --> {target_path}\\{file}")
                return file

def updateTestModule(vxt_file_path, req_test_group, ego_speed, target_speed):
    """
    

    Args:
      vxt_file_path: 
      req_test_group: 
      ego_speed: 
      target_speed: 

    Returns:

    """
    try:
        if 'LANE' in req_test_group.upper():
            req_test_group='Lane'

        tree = ET.parse(vxt_file_path)
        root = tree.getroot()

        # Remove all existing test groups
        for testgroup in root.findall(".//testgroup"):
            if testgroup.get("title").lower() != req_test_group.lower():
                root.remove(testgroup)

        for ego_speed_param in root.findall(f".//testgroup[@title='{req_test_group}']//caplparam[@name='ego_speed']"):
            ego_speed_param.text = str(ego_speed)
        
        if req_test_group.upper() == 'ODET':
            for tgt_speed_param in root.findall(f".//testgroup[@title='{req_test_group}']//caplparam[@name='tgt_speed']"):
                tgt_speed_param.text = str(target_speed)

        tree.write(vxt_file_path, encoding='iso-8859-1')

        # #declaration line
        # with open(vxt_file_path, "r") as original_file:
        #     xml_declaration_line = original_file.readline().strip()
        #     original_file.close()

        # # indentation
        # modified_xml_string = minidom.parseString(ET.tostring(root, encoding="iso-8859-1")).toprettyxml(indent="\t")
        # modified_xml_string = modified_xml_string.split("\n")
        # modified_xml_string[0] = xml_declaration_line
        #
        # modified_xml_string = "\n".join(modified_xml_string)

        # Save the modified XML back to the file
        # with open(vxt_file_path, "w") as file:
        #     file.write(modified_xml_string)
        logger.info(f"Successfully updated test module -> {vxt_file_path} -> test group {req_test_group} with ego speed -> {ego_speed}")
    except Exception as e:
        logger.error(f"Error while updating test module -> {e}")


if __name__ == '__main__':
    # parse command line arguments
    commandLineParser = argparse.ArgumentParser(description='Automated RBS scripts and Canoe Integration tests execution.')
    commandLineParser.add_argument('--testmodule', action="store", dest="testmodule", required=True, help="CAPL test module name")
    commandLineParser.add_argument('--scenario_type', action="store", dest="scenario_type", required=True, help="ODET\LANE\TSR")
    # commandLineParser.add_argument('--scenario_target_type', action="store", dest="scenario_target_type", required=True, help="pedestrian\car\truck...")
    # commandLineParser.add_argument('--testmodule_vxt_file', action="store", dest="testmodule_vxt_file", required=True, help="CAPL test module vxt file to edit name")

    commandLineParser.add_argument('--ego_speed', action="store", dest="ego_speed", required=True, help="Ego speed kmph")
    commandLineParser.add_argument('--target_speed', action="store", dest="target_speed", required=False, default=0, help="Target veh speed kmph")
    commandLineParser.add_argument('--num_iter', action="store", dest="num_iter", required=True, help="no. of time the test module to be exeuted")
    

    arguments = commandLineParser.parse_args()
    platform_path = r"D:\\Jenkins\\ADAS_Platform"

    vxt_file_path = platform_path + r"\\Platform\\TestCases\\FVideoKPI_CM_1V1D.vxt"


    try:
        if not(arguments.testmodule.strip()) or not(arguments.num_iter):
            logger.error("INVALID PARAMETERS from jenkins")
            raise Exception("error INVALID PARAMETERS")
        
        ego_speed_range = arguments.ego_speed.split(",")
        ego_speed_range = [int(i) for i in ego_speed_range]
        
        if len(ego_speed_range)!=3:
            logger.error("INVALID ego_speed_range PARAMETERS from jenkins, correct way: start,stop,step in kmph")
            raise Exception("INVALID ego_speed_range PARAMETERS from jenkins")

        ego_speed_list = [int(speed) for speed in range(ego_speed_range[0], ego_speed_range[1]+1, ego_speed_range[2])]

        
        canoe_cfg = get_canoe_cfg(platform_path)
        conf_full_path = platform_path + r"\\" + canoe_cfg
    
        run_canoe_test(conf_full_path, platform_path,vxt_file_path, arguments.testmodule, int(arguments.num_iter),ego_speed_list, arguments.target_speed,arguments.scenario_type.upper())
   
    except Exception as e:
        logger.error(f"{e}")
        raise Exception(e)
