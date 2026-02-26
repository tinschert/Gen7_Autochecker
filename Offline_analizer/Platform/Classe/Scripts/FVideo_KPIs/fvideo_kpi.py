# -*- coding: utf-8 -*-
# @file fvideo_kpi.py
# @author ADAS_HIL_TEAM
# @date 08-21-2023

##################################################################
# C O P Y R I G H T S
# ----------------------------------------------------------------
# Copyright (c) 2023 by Robert Bosch GmbH. All rights reserved.

# The reproduction, distribution and utilization of this file as
# well as the communication of its contents to others without express
# authorization is prohibited. Offenders will be held liable for the
# payment of damages. All rights reserved in the event of the grant
# of a patent, utility model or design.

##################################################################


import os,sys
import shutil
from datetime import datetime
import pandas as pd
try:
    pd.set_option('future.no_silent_downcasting', True)
except:
    pass
import numpy as np
import matplotlib.pyplot as plt
from asammdf import MDF
import Objects, TrafficSigns, Lanes
from openpyxl import Workbook,load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment,Font
from openpyxl.drawing.image import Image

sys.path.append(os.path.dirname(os.path.abspath(__file__)) + r"\..\Control")
from logging_config import logger



from copy import deepcopy
import time


def parse_iterations(trace_folder,target_sig):
    """
    

    Args:
      trace_folder: 
      target_sig: 

    Returns:

    """
    dict_vfc_obj_001 = {}
    for (dirpath, dirnames, filenames) in os.walk(trace_folder):
        for file in filenames:
            if ".mf4" in file.lower():
                try:
                    f_name = trace_folder+r"\\"+file
                    logger.info("-"*25)
                    logger.info(f"Parsing -> {f_name}")
                    file = MDF(f_name)
                    vfc_obj = file.get(target_sig)
                    vfc_obj.timestamps = vfc_obj.timestamps - np.min(vfc_obj.timestamps)
                    vfc_obj.timestamps = vfc_obj.timestamps - np.min(vfc_obj.timestamps)
                    dict_vfc_obj_001[file] = deepcopy(vfc_obj)
                    logger.info("Done")

                except Exception as e:
                    logger.error(f"While parsing -> {e}")
                    continue

    plot_iteratins(dict_vfc_obj_001,target_sig,trace_folder)


def plot_iteratins(dict_sigs, target_sig,trace_folder):
    """
    

    Args:
      dict_sigs: 
      target_sig: 
      trace_folder: 

    Returns:

    """
    logger.info("-"*25)
    logger.info("Started Plotting...")
    markers = [".",""]
    try:
        if dict_sigs=={}:
            raise Exception("Signal NOT Found")
        for i in range(len(markers)):
            file_name = trace_folder + r"\\..\\..\\"+target_sig+str(i)+".png"
            plt.figure(figsize=(26,21))
            
            count = 1
            for file, sig in dict_sigs.items():
                plt.plot(sig.timestamps, sig.samples, color = "green", marker=markers[i])
                #count+=1

            plt.xlabel("Time (s)")
            plt.ylabel("dx (m)")
            plt.title(target_sig)
            plt.savefig(file_name)
            time.sleep(1)
            #plt.legend(fontsize='large')
        
            logger.info(f"Saved -> {file_name}")
        logger.info("Done Plotting")
        
    except Exception as e:
        logger.error(f"error while ploting -> {e}")
        raise Exception(f"{e}")
    

def removeOutliers(sig):
    """
    

    Args:
      sig: 

    Returns:

    """
    required_times = []
    required_values = []
    leng = len(sig.timestamps)
    for index in range(leng):
        if int(sig.samples[index])!=0:
            required_times.append(sig.timestamps[index])
            required_values.append(sig.samples[index])

    required_times= np.array(required_times)
    required_values= np.array(required_values)

    converted_dxN = deepcopy(sig)
    converted_dxN.timestamps = required_times
    converted_dxN.samples = required_values
    return converted_dxN

def shiftByOffset(values_array, offset):
    """
    

    Args:
      values_array: 
      offset: 

    Returns:

    """
    return values_array + offset


def shiftTimestamps(timestamp_array, shift_time):
    """
    

    Args:
      timestamp_array: 
      shift_time: 

    Returns:

    """
    return timestamp_array + shift_time


def parseTraceGUI(trace_file_list, gt_sysvars, signals, radar_signals):
    """
    

    Args:
      trace_file_list: 
      gt_sysvars: 
      signals: 
      radar_signals: 

    Returns:

    """
    plot_dict = {}
    count=0
    # for (dirpath, dirnames, filenames) in os.walk(trace_folder):
    for f_name in trace_file_list:
        #if ".mf4" in file.lower():
        try:
            count+=1
            key = "Measurement "+str(count)
            if key not in plot_dict.keys():
                plot_dict[key] = {}
            #f_name = trace_folder+r"\\"+file
            logger.info("-"*25)
            logger.info(f"Parsing -> {f_name}")
            file = MDF(f_name)

            #signals
            if signals !=['']:
                for signal in signals: 
                    plot_dict[key][signal] = [deepcopy(file.get(signal)).timestamps, deepcopy(file.get(signal)).samples, file.get(signal).unit]
                    logger.info(f"Extracted - {signal}")

            #sysvars
            if "" not in gt_sysvars:
                namespace_locations = file.whereis(gt_sysvars[0])
                for tup in namespace_locations:
                    tup=list(tup)
                    tup.insert(0,gt_sysvars[0])
                    sig = file.get(tup[0],tup[1],tup[2])
                    if sig.source.name == gt_sysvars[1]:
                        if sig.name == "dxN":
                            sig = deepcopy(removeOutliers(sig))
                        gt_key = sig.source.name.split("::")[-1]+"_"+ sig.name
                        plot_dict[key][gt_key] = [(sig.timestamps), sig.samples, ""]
                        
                        plot_dict[key][gt_key][0] = shiftTimestamps(plot_dict[key][gt_key][0], -0.066)
                        if sig.name == "dxN":
                            plot_dict[key][gt_key][1] = shiftByOffset(plot_dict[key][gt_key][1], 2.033)
                        logger.info(f"Extracted - {gt_key}")
                        break
                
                # for sig in file.iter_channels():
                #     if sig.display_names == {}:
                #         if (sig.name, sig.source.name) in gt_sysvars:
                #             if sig.name == "dxN":
                #                 sig = deepcopy(removeOutliers(sig))
                #             gt_key = sig.source.name.split("::")[-1]+"_"+ sig.name
                #             plot_dict[key][gt_key] = [(sig.timestamps), sig.samples, ""]
                            
                #             plot_dict[key][gt_key][0] = shiftTimestamps(plot_dict[key][gt_key][0], -0.066)
                #             if sig.name == "dxN":
                #                 plot_dict[key][gt_key][1] = shiftByOffset(plot_dict[key][gt_key][1], 2.033)
                #             logger.info(f"Extracted - {gt_key}")
                            
                            
            #radar signals
            if radar_signals !=['']:
                for signal in radar_signals:
                    sig = file.get(signal)
                    sig = deepcopy(removeOutliers(sig))
                    plot_dict[key][signal] = [(sig.timestamps), sig.samples, sig.unit]

                    plot_dict[key][signal][0] = shiftTimestamps(plot_dict[key][signal][0], 0.066)
                    #plot_dict[key][signal][1] = shiftByOffset(plot_dict[key][signal][1], 3.4752)
                    logger.info(f"Extracted - {signal}")

            
            if plot_dict[key] == {}:
                del plot_dict[key]

        except Exception as e:
            if plot_dict[key] == {}:
                del plot_dict[key]
            logger.error(f"While parsing -> {e}")
            continue
                
    return plot_dict






#____________________________________________________________________________________________________________________________________________________________
#JENKINS



def processExcel(buffer, array_df, trace_file, TSR=False, scenario_type=''):
    """
    process fvideo kpi database

    Args:
      buffer: 
      array_df: 
      trace_file: 

    Returns:

    """
    temp= []
    columns = array_df.columns
    array_df = array_df.sort_values(by = "identifier")
    array_mea = dict(array_df.iloc[0])
    if len(array_df) == 2:
        array_gt = dict(array_df.iloc[1])

    # mea
    temp_mea = deepcopy(array_mea)

    if TSR:
        temp_mea["Name"] = temp_mea["Name"].replace("[]",buffer[0])
        temp_mea["Message"] = temp_mea["Message"].replace("[]",buffer[0],1).replace("[]",buffer[1],1)
        temp_mea["pair_index"] = temp_mea["pair_index"] + "_" + str(int(buffer[0]))
        if "raw_gt" in temp_mea["identifier"]:
            search_symbol = temp_mea["Message"] + "::" + temp_mea["Name"]  # need to change from :: to . for structure dtype
        else:
            search_symbol = temp_mea["Message"] + "::" + temp_mea["Name"]
    
    elif scenario_type=='LANES':
        temp_mea["Name"] = temp_mea["Name"].replace("[]", buffer[0]).strip()
        temp_mea["Message"] = temp_mea["Message"].replace("[]_[]", buffer[1]).strip()
        temp_mea["pair_index"] = temp_mea["pair_index"] + "_" + str(int(buffer[0]))
        if "raw_gt" in temp_mea["identifier"]:
            search_symbol = temp_mea["Message"] + "." + temp_mea["Name"]
        else:
            search_symbol = temp_mea["Message"] + "::" + temp_mea["Name"]
    else:
        temp_mea["Name"] = temp_mea["Name"].replace("[]", buffer).strip()
        temp_mea["Message"] = temp_mea["Message"].replace("[]", buffer).strip()
        temp_mea["pair_index"] = temp_mea["pair_index"] + "_" + str(int(buffer))
        if "raw_gt" in temp_mea["identifier"]:
            search_symbol = temp_mea["Message"] + "." + temp_mea["Name"]
        else:
            search_symbol = temp_mea["Message"] + "::" + temp_mea["Name"]
    search_sig = trace_file.search(search_symbol)
    if search_sig == [] or not any([match.endswith(temp_mea["Name"]) for match in search_sig]):
        return []
    else:
        match = False
        for search in search_sig:
            if search.endswith(temp_mea["Name"]):
                match = search
                break
        if match==False:
            return []
        sig = trace_file.get(match)
        if len(sig.timestamps) < 5:
            return []
    
    if len(array_df) == 2:
        if TSR:
            #gt
            temp_gt = deepcopy(array_gt)
            temp_gt["Message"] = temp_gt["Message"].replace("[]",str(int(buffer[0])))
            temp_gt["pair_index"] = temp_gt["pair_index"] + "_" + str(int(buffer[0]))
        elif scenario_type=='LANES':
            temp_gt = deepcopy(array_gt)
            temp_gt['Name'] = temp_gt['Name'].replace('[]',f'[{int(buffer[0])}]')
            #temp_gt["Message"] = temp_gt["Message"]
            temp_gt["pair_index"] = temp_gt["pair_index"] + "_" + str(int(buffer[0]))
        else:
            # gt
            temp_gt = deepcopy(array_gt)
            temp_gt["Message"] = temp_gt["Message"].replace("[]", str(int(buffer)))
            temp_gt["pair_index"] = temp_gt["pair_index"] + "_" + str(int(buffer))

    temp.append(temp_mea)
    if len(array_df) == 2:
        temp.append(temp_gt)

    return pd.DataFrame(temp)





def getTraceData(file, signal="", sysvar=[]):
    """
    

    Args:
      file: 
      signal:  (Default value = "")
      sysvar:  (Default value = [])

    Returns:

    """
    return_list = []      #[timestamp, samples]
    if sysvar==[]:
        #signals
        return_list = [deepcopy(file.get(signal)).timestamps, deepcopy(file.get(signal)).samples]
        logger.info(f"Extracted -> {signal}")
    else:
        #sysvars
        for sig in file.iter_channels():
            if sig.display_names == {}:
                if [sig.name, sig.source.name] == sysvar:
                    return_list = [deepcopy(sig.timestamps), deepcopy(sig.samples)]
                    logger.info(f"Extracted -> {sysvar}")

    return deepcopy(return_list)

def find_nearest_index(array, value):
    """
    

    Args:
      array: 
      value: 

    Returns:

    """
    absolute_diff = np.abs(array - value)
    nearest_index = np.argmin(absolute_diff)

    if array[nearest_index]>value:
        return nearest_index-1

    return nearest_index


def objDx(file, df):
    """
    

    Args:
      file: 
      df: 

    Returns:

    """
    logger.info(f" ---Dx---")
    temp={} #to store keys to identify which is gt and which is camera sig
    dict_dx_data = {}
    for idx, row in df.iterrows():
        #print(row["identifier"])
        if row["identifier"] == "dx_raw_gt":
            temp["dx_gt"] = row["Name"]+"_"+row["Message"].split("::")[-1]
            dict_dx_data[temp["dx_gt"]] = getTraceData(file, sysvar=[row["Name"], row["Message"]])

            #GT dx conversion, shifting, etc..
            dict_dx_data[temp["dx_gt"]][0] = shiftTimestamps(dict_dx_data[temp["dx_gt"]][0], -0.066)
            dict_dx_data[temp["dx_gt"]][1] = shiftByOffset(dict_dx_data[temp["dx_gt"]][1], 2.033)


        else:
            temp["dx_cam"] = row["Name"]
            dict_dx_data[temp["dx_cam"]] = getTraceData(file, signal=row["Name"])
    if len(temp)!=2:
        logger.error("objDx error occured")


    cal_gt = {"timestamps":[],"samples":[]}
    for t in dict_dx_data[temp["dx_cam"]][0]:
        gt_index = find_nearest_index(dict_dx_data[temp["dx_gt"]][0],t)
        cal_gt["timestamps"].append(dict_dx_data[temp["dx_gt"]][0][gt_index])
        cal_gt["samples"].append(dict_dx_data[temp["dx_gt"]][1][gt_index])

    df = pd.DataFrame(columns=["calGT_timestamp", "calGT_values", "camera_timestamp", "camera_values"])
    

    # df_gt = pd.DataFrame(columns = ["gt_timestamp","gt_values"])
    df["calGT_timestamp"] =  cal_gt["timestamps"]
    df["calGT_values"] = cal_gt["samples"]

    # df_cam = pd.DataFrame(columns = ["camera_timestamp", "camera_values"])
    df["camera_timestamp"] = dict_dx_data[temp["dx_cam"]][0]
    df["camera_values"] = dict_dx_data[temp["dx_cam"]][1]
    return df
    # df = pd.concat([df_gt,df_cam],axis=1)
    # df.to_excel(r"C:\\GIT WORKSPACE\\KPIs\\Platform\\Classe\\Scripts\\FVideo_KPIs\\" + r"calGT_vs_meaCamera_" + (str(file.original_name).split("\\")[-1].split(".")[0])+".xlsx", index= False)

def getTodaysDate():
    """ """
    #"day_month_year"
    today = datetime.today()
    date = today.strftime("%d_%m_%Y")
    return date


def createGraphs(graph_parameter_name, data, output_path):
    """generates

    Args:
        graph_parameter_name (str): graph parameter name, also save image with this name
        data (dict): can have two keys
                     gt: [timestamps, values]
                     cam: [timestamps, values]
                     key 'gt' is optional, if not present plot only cam
        output_path (str): output folder to save created graphs, create folder path and then save
    """


    # script_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..\\" * 3)

    # output_path = os.path.join(script_dir, output_path)
    try:
        if not os.path.exists(output_path):
            os.makedirs(output_path)

        plt.figure(figsize=(26,21))
        for keys in data:
            if keys == "gt":
                plt.plot(data[keys][0], data[keys][1], color = "green", label=keys)
            elif keys == "cam":
                plt.plot(data[keys][0], data[keys][1], color = "blue", label=keys)

        plt.title(graph_parameter_name, fontsize=36)
        plt.xlabel("Timestamp (s)", fontsize=36)
        plt.ylabel("Samples", fontsize=36)

        plt.legend(fontsize=18)
        plt.savefig(output_path + r"\\" + graph_parameter_name + ".png")
        plt.close()
        logger.info(f"Generated graph for {graph_parameter_name} saved at {output_path}")
        time.sleep(0.6)
    except Exception as e:
        logger.error(f"Error while generating grapg for parameter {graph_parameter_name} -> {e}")


def copyExcelReport(excel_path, destination_path):
    """
    

    Args:
      excel_path: 
      destination_path: 

    Returns:

    """
    # Get the folder name date
    folder_name = getTodaysDate()
    dest_path = os.path.join(destination_path, folder_name)
    
    if not os.path.exists(dest_path):
        os.makedirs(dest_path)

    # Copy
    shutil.copy2(excel_path, dest_path)

    logger.info(f"Excel copied to -> {dest_path}")

def get_scenario_target_type(test_module_name, scenario_type, fvideo_kpi_specific = False):
    """
    return scenario,scenario_target_type based on info in test module name

    Args:
      test_module_name: 
      fvideo_kpi_specific:  (Default value = False)

    Returns:


    """
    if scenario_type == 'LANES':
        return 'Lanes'
    test_module_name=test_module_name.lower()

    scenario_type_sheetname_mapping_dict = {'ODET': {"pedestrian":"Pedestrain", 
                                                     "EgoLong_TargetLong_Car":"Static_Car",
                                                     "oncoming":"On_Coming_Car",
                                                     "EgoLong_TargetLong_Truck":"EgoLong_TargetLong_Truck",
                                                     'EgoLong_TargetLong_BUS':'EgoLong_TargetLong_Bus',
                                                     'Approaching_Long_Cyclist': 'Approaching_Long_Cyclist',
                                                     'Approaching_Long_MotorCyclist':'Approaching_Long_MotorCyclist',
                                                     'CrossingCarIntersection_Left_Right': 'Crossing_Car_Intersection',
                                                     'CrossingCarIntersection_Right_Left': 'Crossing_Car_Intersection',
                                                     'Crossing_MotorCyclist_Intersection_Right_Left': 'Crossing_MotorCyclist_Intersec',
                                                     'Crossing_MotorCyclist_Intersection_Left_Right': 'Crossing_MotorCyclist_Intersec',
                                                     'CrossingCyclistIntersection_Left_Right': 'Crossing_Cyclist_Intersection',
                                                     'CrossingCyclistIntersection_Right_Left': 'Crossing_Cyclist_Intersection'
                                                     },
                                            'TSR': {'_singlepole' :'SLA_50KPH_SinglePole',
                                                    'on_gantry': 'SLA_50KPH_On_Gantry',
                                                    '_stop': 'DE_TSR_STOP',
                                                    '_warnkids': 'DE_TSR_WARNKIDS',
                                                    '_giveway': 'DE_TSR_GiveWay',
                                                    'no_overtaking':'No_OverTaking',
                                                    'nooverTaking_and_endofovertaking': 'NoOverTaking_and_EndofOv',
                                                    '_validTime': 'SLA_50KPH_ValidTime',
                                                    '_whenwet':'SLA_50KPH_WhenWet'}
                                                    }


    #check if objects
    for scenario_target_type, sheet_name in scenario_type_sheetname_mapping_dict[scenario_type].items():
        if (scenario_target_type.lower() in test_module_name.lower()):
            if fvideo_kpi_specific:
                return sheet_name
            if("aeb" in  test_module_name):
                return ("OBJECTS",sheet_name)
    return (None,None)
   
    
def saveFinalOutputExcel(output_file_path, workbook_dict):
    """
    

    Args:
      output_file_path: 
      workbook_dict: 

    Returns:

    """
    try:
        wb = Workbook()
        empty_sheet = wb.active
        wb.remove(empty_sheet)

        for sheet_name, workbook in workbook_dict.items():
            sheet_to_copy = workbook.active  
            result_columns = [cell.value for cell in sheet_to_copy[1]]

            if sheet_name in wb.sheetnames:
                base_sheet_name = sheet_name
                counter = 1
                while sheet_name in wb.sheetnames:
                    sheet_name = f"{base_sheet_name}_{counter}"
                    counter += 1

            ws = wb.create_sheet(title=sheet_name)
            for row in sheet_to_copy.iter_rows(values_only=True):
                ws.append(row)


            current_value = None
            start_merge_row = None
            for row_num, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if row[0] != current_value:
                    if start_merge_row is not None:
                        ws.merge_cells(f'A{start_merge_row}:A{row_num - 1}')
                    current_value = row[0]
                    start_merge_row = row_num

            
            if start_merge_row is not None:
                ws.merge_cells(f'A{start_merge_row}:A{row_num}')

            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                cell = ws.cell(row=row_num, column=result_columns.index("Verdict")+1)  
                if row[result_columns.index("Verdict")] == "PASS":
                    #for cell in ws[row_num]:
                    cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green
                else:
                    #for cell in ws[row_num]:
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red

                cell = ws.cell(row=row_num, column=result_columns.index("Missing/No_Data_Available_Samples")+1)  
                if row[result_columns.index("Missing/No_Data_Available_Samples")]!=0:
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red

            for column_cells in ws.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                adjusted_width = (length + 2)  
                ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width

            border = Border(left=Side(border_style="thin", color="557072"),
                    right=Side(border_style="thin", color="557072"),
                    top=Side(border_style="thin", color="557072"),
                    bottom=Side(border_style="thin", color="557072"))

            for row in ws.iter_rows():
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            
            for cell in ws[1]:  # Assuming the first row is row index 1.
                cell.font = Font(bold=True)

            #Insert graphs to ws
            # script_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..\\" * 3)
            # sheet = wb[sheet_name]
            # num_rows = sheet.max_row 

            # filepath = os.path.join(script_dir, os.path.splitext(output_file_path)[0], sheet_name)

            try:
                graphs_folder_path = output_file_path.replace('.xlsx','') + r"\\" + sheet_name 

                all_files = os.listdir(graphs_folder_path)
                num_rows = ws.max_row
                for image_name in all_files:
                    if not (image_name.endswith('.png')):
                        continue
                    num_rows += 150
                    cell_location = 'A' + str(num_rows)
                    image_file_path = os.path.join(graphs_folder_path, image_name)
                    img = Image(image_file_path)
                    ws.add_image(img, cell_location)
            except Exception as e:
                logger.error(f"Error while attaching graphs to {output_file_path} for {sheet_name} -> {e}")

        wb.save(output_file_path)
        logger.info(f"Verdict Excel successfully generated -> {output_file_path}")
    except Exception as e:
        logger.error("Error occured while save output excel -> {e}")

def get_carmaker_nlines(trace_file):
    """
    

    Args:
      trace_file: 

    Returns:

    """
    try:
        nline = trace_file.get('CarMaker::RB::GroundTruth::Line::nLines')
        for i, val in enumerate(nline.samples):
            if val>0:
                if val>6:
                    return 6, nline.timestamps[i]
                return val, nline.timestamps[i]
    except:
        logger.warning('CarMaker::RB::GroundTruth::Line::nLines not found')
        return 0,0

def Objects_main(trace_file, database_excel_df, evaluation_excel_df, scenario_target_type,iteration, graphs_output_path):
    """
    objects main function

    Args:
      trace_file (str): trace file
      database_excel_df (dataframe):
      evaluation_excel_df (dataframe):
      scenario_target_type (str):
      iteration (int):

    Returns:

    """
    logger.info(f"------------OBJECTS {iteration}--------------")
    #obj_buffers = ["0"+str(i) for i in range(10)]
    #currently only checking for object0
    obj_buffers = ["00"]
    objects_dict = {}

    buffer_found = True
    for buffer in obj_buffers:
        unique_obj_parameters = list(database_excel_df["pair_index"].unique())
        objects_dict[buffer] = Objects.Object(trace_file, database_excel_df, evaluation_excel_df, scenario_target_type, iteration)
        for param in unique_obj_parameters:
            temp_df = database_excel_df[database_excel_df["pair_index"] == param]
            temp_df = processExcel(buffer, temp_df, trace_file)
            if len(temp_df) == 0 and type(temp_df)==list:
                buffer_found = False
                break
            if len(temp_df) == 0:
                logger.error(f"Error occured input dataframe is empty for {param}")
                return False
            
            pair_id_list = list(temp_df["pair_index"].unique())
            for pair_id in pair_id_list:
                graph_data = {}
                pairs = temp_df[temp_df["pair_index"] == pair_id]

                #dx should be processed first and required data is stored which is required for other parameters
                if "dx_" in pair_id:
                    if "dx" not in objects_dict[buffer].evaluation_database.keys():
                        continue
                    calculated_dx =  objects_dict[buffer].cal_DX(pairs)

                    graph_data = calculated_dx.copy()

                    objects_dict[buffer].object_range, calculated_dx = objects_dict[buffer].trimData(calculated_dx)
                    if calculated_dx is None:
                        continue

                    objects_dict[buffer].pair_indices = objects_dict[buffer].findCameraGtPairs(calculated_dx)
                    objects_dict[buffer].calculated_parameters_dict["dx"] = calculated_dx

                    evaluated_result = objects_dict[buffer].evaluate_DX(calculated_dx)
                    objects_dict[buffer].evaluation_results["dx"] = evaluated_result
                elif "dy_" in pair_id:
                    if "dy" not in objects_dict[buffer].evaluation_database.keys():
                        continue
                    calculated_dy =  objects_dict[buffer].cal_DY(pairs)
                    graph_data = calculated_dy.copy()

                    evaluated_result = objects_dict[buffer].evaluate_DY(calculated_dy)
                    objects_dict[buffer].evaluation_results["dy"] = evaluated_result
                elif "vx_" in pair_id:
                    if "vx" not in objects_dict[buffer].evaluation_database.keys():
                        continue
                    calculated_vx =  objects_dict[buffer].cal_VX(pairs)
                    graph_data = calculated_vx.copy()

                    evaluated_result = objects_dict[buffer].evaluate_VX(calculated_vx)
                    objects_dict[buffer].evaluation_results["vx"] = evaluated_result
                elif "vy_" in pair_id:
                    if "vy" not in objects_dict[buffer].evaluation_database.keys():
                        continue
                    calculated_vy =  objects_dict[buffer].cal_VY(pairs)
                    graph_data = calculated_vy.copy()
                    
                    evaluated_result = objects_dict[buffer].evaluate_VY(calculated_vy)
                    objects_dict[buffer].evaluation_results["vy"] = evaluated_result
                elif "ax_" in pair_id:
                    if "ax" not in objects_dict[buffer].evaluation_database.keys():
                        continue
                    calculated_ax =  objects_dict[buffer].cal_AX(pairs)
                    graph_data = calculated_ax.copy()

                    evaluated_result = objects_dict[buffer].evaluate_AX(calculated_ax)
                    objects_dict[buffer].evaluation_results["ax"] = evaluated_result
                elif "ay_" in pair_id:
                    if "ay" not in objects_dict[buffer].evaluation_database.keys():
                        continue
                    calculated_ay =  objects_dict[buffer].cal_AY(pairs)
                    graph_data = calculated_ay.copy()
                    evaluated_result = objects_dict[buffer].evaluate_AY(calculated_ay)
                    objects_dict[buffer].evaluation_results["ay"] = evaluated_result
                elif "length_" in pair_id:
                    if "length" not in objects_dict[buffer].evaluation_database.keys():
                        continue
                    calculated_length =  objects_dict[buffer].cal_length(pairs)
                    graph_data = calculated_length.copy()

                    evaluated_result = objects_dict[buffer].evaluate_length(calculated_length)
                    objects_dict[buffer].evaluation_results["length"] = evaluated_result   
                elif "width_" in pair_id:
                    if "width" not in objects_dict[buffer].evaluation_database.keys():
                        continue
                    calculated_width = objects_dict[buffer].cal_width(pairs)
                    graph_data = calculated_width.copy()

                    evaluated_result = objects_dict[buffer].evaluate_width(calculated_width)
                    objects_dict[buffer].evaluation_results["width"] = evaluated_result   
                elif "headingangle_" in pair_id:
                    if "headingangle" not in objects_dict[buffer].evaluation_database.keys():
                        continue
                    calculated_headingangle = objects_dict[buffer].cal_headingangle(pairs)
                    graph_data = calculated_headingangle.copy()

                    evaluated_result = objects_dict[buffer].evaluate_headingangle(calculated_headingangle)
                    objects_dict[buffer].evaluation_results["headingangle"] = evaluated_result 
                elif "type_" in pair_id:
                    if "type" not in objects_dict[buffer].evaluation_database.keys():
                        continue
                    calculated_type = objects_dict[buffer].cal_type(pairs)
                    graph_data = calculated_type.copy()

                    evaluated_result = objects_dict[buffer].evaluate_type(calculated_type)
                    objects_dict[buffer].evaluation_results["type"] = evaluated_result
                elif "probexist_" in pair_id:
                    if "probexist" not in objects_dict[buffer].evaluation_database.keys():
                        continue
                    calculated_probexist = objects_dict[buffer].cal_probexist(pairs)
                    graph_data = calculated_probexist.copy()

                    evaluated_result = objects_dict[buffer].evaluate_probexist(calculated_probexist)
                    objects_dict[buffer].evaluation_results["probexist"] = evaluated_result
                elif "scalechange_" in pair_id:
                    if "scalechange" not in objects_dict[buffer].evaluation_database.keys():
                        continue
                    calculated_scalechange = objects_dict[buffer].cal_scalechange(pairs)
                    graph_data = calculated_scalechange.copy()
                    
                    evaluated_result = objects_dict[buffer].evaluate_scalechange(calculated_scalechange)
                    objects_dict[buffer].evaluation_results["scalechange"] = evaluated_result
                elif "typeprob_" in pair_id:
                    if "typeprob" not in objects_dict[buffer].evaluation_database.keys():
                        continue
                    calculated_typeprob = objects_dict[buffer].cal_typeprob(pairs)
                    graph_data = calculated_typeprob.copy()

                    evaluated_result = objects_dict[buffer].evaluate_typeprob(calculated_typeprob)
                    objects_dict[buffer].evaluation_results["typeprob"] = evaluated_result
                elif "visibleview_" in pair_id:
                    if "visibleview" not in objects_dict[buffer].evaluation_database.keys():
                        continue
                    calculated_visibleview = objects_dict[buffer].cal_visibleview(pairs)
                    graph_data = calculated_visibleview.copy()
                    evaluated_result = objects_dict[buffer].evaluate_visibleview(calculated_visibleview)
                    objects_dict[buffer].evaluation_results["visibleview"] = evaluated_result

                if graph_data!={}:
                    graph_param = pair_id.rsplit("_", maxsplit=1)[0]
                    createGraphs(graph_param, graph_data, graphs_output_path)
        
        #if buffer_found==False:
        #assuming 1 object
        if buffer_found == False:
            return False
    
    #assuming one object
    return objects_dict["00"].getVerdictExcel()



def Lanes_Kpi(trace_file, database_excel_df, evaluation_excel_df, scenario_target_type,iteration, graphs_output_path):
    """
    

    Args:
      trace_file: 

    Returns:

    """
    logger.info(f"------------ Lanes {iteration}--------------")
    #obj_buffers = ["0"+str(i) for i in range(10)]
    #currently only checking for object0
    lane_buffers_dict = [("00","00_02"), ("01","00_02"), ("02","00_02"),
                         ("03","03_05"), ("04","03_05"), ("05","03_05")]
    
    nline, start_timestamp = get_carmaker_nlines(trace_file)
    if not(nline):
        return None
    
    logger.info(f'Carmaker variable nLine = {nline}')

    detected_messages = []

    verdict_dfs = []

    objects_dict = {}

    buffer_found = True
    for index in range(nline):
        unique_obj_parameters = list(database_excel_df["pair_index"].unique())
        buffer_key = lane_buffers_dict[index][0]

        #keep track of found messages
        if buffer_key not in detected_messages:
            detected_messages.append(buffer_key)

        objects_dict[buffer_key] = Lanes.Lane(trace_file, database_excel_df, evaluation_excel_df, scenario_target_type, iteration, buffer_key)
        objects_dict[buffer_key].start_timestamp = start_timestamp

        for param in unique_obj_parameters:
            temp_df = database_excel_df[database_excel_df["pair_index"] == param]
            temp_df = processExcel(lane_buffers_dict[index], temp_df, trace_file, scenario_type='LANES')
            if len(temp_df) == 0 and type(temp_df)==list:
                logger.error(f'Data not found in trace for {lane_buffers_dict[index]}')
                buffer_found = False
                break
            if len(temp_df) == 0:
                logger.error(f"Error occured input dataframe is empty for {param}")
                return False
            
            pair_id_list = list(temp_df["pair_index"].unique())
            for pair_id in pair_id_list:
                pairs = temp_df[temp_df["pair_index"] == pair_id]

                if "laneassociation_" in pair_id:
                    key = pair_id.split('_')[0]
                    if key not in objects_dict[buffer_key].evaluation_database.keys():
                        continue
                    calculated_laneassociation = objects_dict[buffer_key].cal_laneassociation(pairs)
                    evaluated_result = objects_dict[buffer_key].evaluate_laneassociation(calculated_laneassociation)
                    objects_dict[buffer_key].evaluation_results[key] = evaluated_result.copy()
                else:
                    key = pair_id.split('_')[0]
                    if key not in objects_dict[buffer_key].evaluation_database.keys():
                        continue
                    claculated_parameter_data = objects_dict[buffer_key].cal_lane_parameter(pairs, pair_id)

                    function_name = 'evaluate_' + key
                    evaluate_func = getattr(objects_dict[buffer_key], function_name)
                    parameter_result = evaluate_func(claculated_parameter_data)
                    objects_dict[buffer_key].evaluation_results[key] = parameter_result.copy()

        # if buffer_found==False:
        # assuming 1 object
        if buffer_found == False:
            return False
    for buf_index in detected_messages:
        verdict_dfs.append(objects_dict[buf_index].getVerdictExcel())

    #concat verdict_dfs
    concat_df = pd.concat(verdict_dfs, ignore_index=True)
    concat_df.to_excel(objects_dict['00'].output_file, index=False)
    wb = load_workbook(objects_dict['00'].output_file)
    return wb


def Traffic_Signs(trace_file, database_excel_df, evaluation_excel_df, scenario_target_type,iteration, graphs_output_path):
    """
    

    Args:
      trace_file: 
      database_excel_df: 
      evaluation_excel_df: 
      scenario_target_type: 
      iteration: 

    Returns:

    """
    logger.info(f"------------Traffic Sign {iteration}--------------")
    #obj_buffers = ["0"+str(i) for i in range(10)]
    #currently only checking for object0
    ts_buffers = [("00","01")]
    objects_dict = {}

    buffer_found = True
    for buffer in ts_buffers:
        unique_obj_parameters = list(database_excel_df["pair_index"].unique())
        buffer_key = buffer[0]+"_"+buffer[1]
        objects_dict[buffer_key] = TrafficSigns.TrafficSign(trace_file, database_excel_df, evaluation_excel_df, scenario_target_type, iteration)
        for param in unique_obj_parameters:
            temp_df = database_excel_df[database_excel_df["pair_index"] == param]
            temp_df = processExcel(buffer, temp_df, trace_file, TSR=True)
            if len(temp_df) == 0 and type(temp_df)==list:
                buffer_found = False
                break
            if len(temp_df) == 0:
                logger.error(f"Error occured input dataframe is empty for {param}")
                return False
            
            pair_id_list = list(temp_df["pair_index"].unique())
            for pair_id in pair_id_list:
                pairs = temp_df[temp_df["pair_index"] == pair_id]

                #dx should be processed first and required data is stored which is required for other parameters


                if "dx_" in pair_id:
                    if "dx" not in objects_dict[buffer_key].evaluation_database.keys():
                        continue
                    calculated_dx =  objects_dict[buffer_key].cal_DX(pairs)

                    objects_dict[buffer_key].object_range, calculated_dx = objects_dict[buffer_key].trimData(calculated_dx)

                    objects_dict[buffer_key].pair_indices = objects_dict[buffer_key].findCameraGtPairs(calculated_dx)
                    objects_dict[buffer_key].calculated_parameters_dict["dx"] = calculated_dx

                    evaluated_result = objects_dict[buffer_key].evaluate_DX(calculated_dx)
                    objects_dict[buffer_key].evaluation_results["dx"] = evaluated_result
                elif "dy_" in pair_id:
                    if "dy" not in objects_dict[buffer_key].evaluation_database.keys():
                        continue
                    calculated_dy = objects_dict[buffer_key].cal_DY(pairs)
                    evaluated_result = objects_dict[buffer_key].evaluate_DY(calculated_dy)
                    objects_dict[buffer_key].evaluation_results["dy"] = evaluated_result
                elif "dz_" in pair_id:
                    if "dz" not in objects_dict[buffer_key].evaluation_database.keys():
                        continue
                    calculated_dz = objects_dict[buffer_key].cal_DZ(pairs)
                    evaluated_result = objects_dict[buffer_key].evaluate_DZ(calculated_dz)
                    objects_dict[buffer_key].evaluation_results["dz"] = evaluated_result
                else:
                    if param not in objects_dict[buffer_key].evaluation_database.keys():
                        continue
                    calculated_type = objects_dict[buffer_key].cal_single_parameter(pairs,param)
                    evaluated_result = objects_dict[buffer_key].evaluate_single_parameter(calculated_type,param)
                    objects_dict[buffer_key].evaluation_results[param] = evaluated_result

        # if buffer_found==False:
        # assuming 1 object
        if buffer_found == False:
            return False

    # assuming one object
    return objects_dict["00_01"].getVerdictExcel()

def external_call(trace_folder_path, scenario_type, scenario_target_type, output_file_path, evaluated_trace_file_list, copy_reports=True):
    """
    

    Args:
      trace_folder_path: 
      scenario_type: 
      scenario_target_type: 
      output_file_path: 
      evaluated_trace_file_list: 

    Returns:

    """
    script_path = os.path.dirname(os.path.abspath(__file__))

    destination_paths = {"ODET": r"\\abtvdfs2.de.bosch.com\ismdfs\iad\verification\ODS\ADAS_HIL_concept\Front_Video\FV_KPI\KPI_Results\ODET",
                         'TSR': r"\\abtvdfs2.de.bosch.com\ismdfs\iad\verification\ODS\ADAS_HIL_concept\Front_Video\FV_KPI\KPI_Results\TSR",
                         'LANES': r"\\abtvdfs2.de.bosch.com\ismdfs\iad\verification\ODS\ADAS_HIL_concept\Front_Video\FV_KPI\KPI_Results\LANE"}
    
    input_excel_path = script_path + r"/../../../../CustomerPrj/FVideo_KPIs/FVideo_KPI_Database.xlsx"

    verdict_excels = {'ODET': 'FVideo_KPI_Verdict_Objects.xlsx',
                      'TSR': 'FVideo_KPI_Verdict_TrafficSign.xlsx',
                      'LANES': 'FVideo_KPI_Verdict_Lane.xlsx'}

    evaluation_excel_path = script_path + r"/../../../../CustomerPrj/FVideo_KPIs/" + verdict_excels[scenario_type]
    
    type_func_mapping = {"ODET":Objects_main, "LANES":Lanes_Kpi, "TSR":Traffic_Signs}

    target_func = type_func_mapping[scenario_type]
    logger.info(f"-----------------Start KPI Analysis for {scenario_type}-----------------")
    iteration_wb_dict = {}

    #signal database
    sheet = ''
    if scenario_type=="ODET" or scenario_type=="OBJECTS":
        sheet = "Objects"
    elif scenario_type=="TSR":
        sheet = "TrafficSigns"
    elif scenario_type.upper()=="LANES":
        sheet = "Lanes"
    else:
        logger.error(f"Invalid scenario type -> {scenario_type}")
        return None
    excel_file_df = pd.read_excel(input_excel_path, na_values="NA", sheet_name=sheet)
    excel_file_df = excel_file_df.fillna("NA")

    for (dirpath, dirnames, filenames) in os.walk(trace_folder_path):
        count=0
        for file in filenames:
            if ".mdf" in file.lower():
                if file not in evaluated_trace_file_list:
                    evaluated_trace_file_list.append(file)
                else:
                    continue
                count+=1
                try:
                    logger.info(f"------------------------------------------Reading -> {file}------------------------------------------")
                    graphs_output_path = output_file_path.replace('.xlsx','') + r"\\" + f"Iteration{count}"
                    
                    f_name = trace_folder_path+r"\\"+file
                    file = MDF(f_name)
                    wb = target_func(file, excel_file_df, evaluation_excel_path, scenario_target_type,"Iteration_"+str(count), graphs_output_path)
                    if not(wb):
                        logger.error(f"Data not found in {f_name}")
                        continue
                    iteration_wb_dict["Iteration"+str(count)] = wb
                    time.sleep(2)
                    file.close()
                except Exception as e:
                    try:
                        file.close()
                    except:
                        pass
                    raise Exception(e)
    
    
    if iteration_wb_dict!={}:
        saveFinalOutputExcel(output_file_path, iteration_wb_dict)
        if copy_reports:
            copyExcelReport(output_file_path, destination_paths[scenario_type])
    else:
        logger.error(f"Data empty for to generate Excel")
    
    return evaluated_trace_file_list
    
    # count = 1
    # writer = pd.ExcelWriter(script_path + r"\\..\\..\\..\\..\\FVIDEO_KPI_Temp.xlsx")
    # for df in global_iter_list:
    #     sheet_name = "Iteration"+str(count)
    #     df.to_excel(writer, sheet_name=sheet_name, index = False)
    #     count+=1
    # writer.save()
        

def runFvideoKPI(tracefile_path,output_excel_path,scenario_type, scenario_target_type):
    """
    main function for fvideo kpi

    Args:
      tracefile_path (str): folder pather where log files are saved
      output_excel_path (str): output excel path
      scenario_type (str): TRS or object or lane
      scenario_target_type (str): if object then object tarrget type peesd/atatic car etc
    """
    script_path = os.path.dirname(os.path.abspath(__file__))
    
    input_excel_path = script_path + r"/../../../../CustomerPrj/FVideo_KPIs/FVideo_KPI_Database.xlsx"
    evaluation_excel_path = script_path + r"/../../../../CustomerPrj/FVideo_KPIs/FVideo_KPI_Verdict_Objects.xlsx"
    
    type_func_mapping = {"OBJECTS":Objects_main, "LANES":Lanes_Kpi, "TRAFFIC SIGNS":Traffic_Signs}

    target_func = type_func_mapping[scenario_type]
    logger.info(f"-----------------Start KPI Analysis for {scenario_type}-----------------")
    excel_file_df = pd.read_excel(input_excel_path, na_values="NA")
    iteration_wb_dict={}

    try:
        logger.info(f"------------------------------------------Reading -> {tracefile_path}------------------------------------------")

        file = MDF(tracefile_path)
        wb = target_func(file, excel_file_df, evaluation_excel_path, scenario_target_type,"Iteration_"+str(1))
        if not(wb):
            logger.error(f"Data not found in {tracefile_path}")
            return
        iteration_wb_dict["Iteration"+str(1)] = wb


        if iteration_wb_dict!={}:
            saveFinalOutputExcel(output_excel_path, iteration_wb_dict)

        time.sleep(2)
        file.close()
    except Exception as e:
        try:
            file.close()
        except:
            pass
        raise Exception(e)
    
    

