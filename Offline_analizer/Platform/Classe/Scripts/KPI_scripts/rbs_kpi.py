# -*- coding: utf-8 -*-
# @file rbs_kpi.py
# @author ADAS_HIL_TEAM
# @date 08-21-2023

##################################################################
# C O P Y R I G H T S
# ----------------------------------------------------------------
# Copyright (c) 2023 by Robert Bosch GmbH. All rights reserved.

# The reproduction, distribution and utilization of this file as
# well as the communication of its contents to others without express
# authorization is prohibited. Offenders will be held liable for the
# payment of damages. All rights reserved in the event of the grant
# of a patent, utility model or design.

##################################################################
"""
used by jenkins, takes log files as input and generates KPI report related to rbs and canoe
"""

import argparse
import os,sys

import openpyxl
from openpyxl.styles import Font,Alignment,Border,Side,PatternFill
from openpyxl.drawing.image import Image

import pandas as pd
try:
    pd.set_option('future.no_silent_downcasting', True)
except:
    pass
import numpy as np
import matplotlib.pyplot as plt
from asammdf import MDF
sys.path.append(os.getcwd() + r"\..\..\..\..\CustomerPrj\Restbus\Scripts")
from pattern_matching import filter_column_df, dict_channel_nwtName_mapping

sys.path.append(os.path.dirname(os.path.abspath(__file__)) + r"\..\Jenkins_scripts")
#from rbs_and_canoe_tests import get_canoe_cfg
sys.path.append(os.path.dirname(os.path.abspath(__file__)) + r"\..\Control")
from logging_config import logger
import shutil
import subprocess
import pythoncom
import time
import canoe_client_1




border = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)
bold_font = Font(bold=True)



required_sig_list = ["init_all", "init_done", "CAN_BusStatistic", "CPUUsageTotal", "MainQueueLoad", "CPUUsageProcess", "TimerAccuracy"]
dict_global_kpi = {}





def merge_sheets_into_single_sheet(destination_wb, source_wb, test_scn_name, sheet_name = "new_sheet"):
    """
    merge all kpi results into single sheet

    Args:
      destination_wb (workbook): destination wb
      source_wb (workbook): source wb
      test_scn_name (str): test scenario name
      sheet_name (str):  (Default value = "new_sheet") sheet name after merge

    Returns:
        destination_wb (workbook)
    """
    try:
        bold_row_no = []
        dict_merge_cells = {}
        current_row = 1
        combined_sheet = destination_wb.create_sheet(title=sheet_name)
        
        combined_sheet.cell(row=current_row, column=1, value=f'{test_scn_name}')
        
        bold_row_no.append(current_row-1)
        current_row+=1
        for _ in range(2):
            combined_sheet.append([])
            current_row+=1
        

        for source_sheet_name in source_wb.sheetnames:
            sheet = source_wb[source_sheet_name]
            combined_sheet.cell(row=current_row, column=1, value=f'{source_sheet_name}')
            dict_merge_cells[current_row] = sheet.max_column
            bold_row_no.append(current_row-1)
            current_row += 1
            bold_row_no.append(current_row-1)
            
            # Copy the data from the current sheet
            for row in sheet.iter_rows(values_only=True):
                combined_sheet.append(row)
                if(source_sheet_name == "CAN Statistics") and (row[0]!="Channel"):
                    for i, cell in enumerate(combined_sheet[current_row]):
                        if i == 4 and cell.value > 10:
                            cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                elif(source_sheet_name == "Frame_Statistics") and (row[0]!="Network name"):
                    for i, cell in enumerate(combined_sheet[current_row]):
                        if i == 9 and cell.value != '-' and cell.value > 5:
                            cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                elif(source_sheet_name == "Canoe Statistics") and (row[0]!="Environment"):
                    if row[0] != "Environment":
                        for i, cell in enumerate(combined_sheet[current_row]):
                            if i == 2 and cell.value != None and cell.value > 50:
                                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    if row[0] == "RT_rack":
                        for i, cell in enumerate(combined_sheet[current_row]):
                            if i == 6 and cell.value != None:
                                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                            if i == 7 and cell.value != None and cell.value > 5:
                                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                current_row += 1
            
            # Add a gap of 4 rows
            for _ in range(3):
                combined_sheet.append([])  # Add empty rows
                current_row += 1
                
            

        dict_merge_cells[1] = combined_sheet.max_column
        
        for row_number in bold_row_no:
            row = combined_sheet[row_number + 1]  # Adjust for 0-based index
            for cell in row:
                cell.font = bold_font
                
        

        # Define the left-aligned text alignment style
        left_alignment = Alignment(horizontal='left', vertical='center')

        # Iterate through all cells in the worksheet and set the alignment style
        for i,row in enumerate(combined_sheet.iter_rows()):
            for cell in row:
                cell.alignment = left_alignment
                if cell.value!=None:
                    cell.border = border

        for row_number, merge_cells in dict_merge_cells.items():
            start_column = 1  # Start from the first column
            end_column = start_column + merge_cells - 1

            # Merge cells in the specified range
            combined_sheet.merge_cells(start_row=row_number, start_column=start_column, end_row=row_number, end_column=end_column)

            # Center the text in the merged cells
            for row in combined_sheet.iter_rows(min_row=row_number, max_row=row_number):
                cell = row[0]
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color="0099ff", end_color="0099ff", fill_type="solid")

        max_widths = {}
        for row in combined_sheet.iter_rows(values_only=True):
            for column_index, cell_value in enumerate(row, start=1):
                if cell_value == test_scn_name:
                    continue
                column_letter = openpyxl.utils.get_column_letter(column_index)
                max_width = max(max_widths.get(column_letter, 0), len(str(cell_value))+1)
                max_widths[column_letter] = max_width

        # Set the column width based on the maximum width needed
        for column_letter, max_width in max_widths.items():
            combined_sheet.column_dimensions[column_letter].width = max_width

        return destination_wb
    except Exception as e:
        logger.error(f"Error while merging sheets -> {e}")
        return None



def merge_excel_files_into_sheets(output_folder_path, main_folder=""):
    """
    merges more than 1 excel file into single excel by creating sheets

    Args:
      output_folder_path (str): output folder path
      main_folder (str):  (Default value = "") source folder where excel files are present

    """
    try:
        merged_workbook = openpyxl.Workbook()
        for (dirpath, dirnames, filenames) in os.walk(output_folder_path):
            for file in filenames:
                if ".xlsx" in file and "RBS_KPI" in file:
                    excel_path = output_folder_path+ "\\" + file
                    workbook = openpyxl.load_workbook(excel_path)

                    test_scn_name = file.replace(".xlsx","").replace("RBS_KPI_","")
                    #just to minimize sheet name length----------
                    if "___" in test_scn_name:
                        sheet_name = test_scn_name.split("___")[0].replace("TC_PER_", "")
                        if len(sheet_name) > 31:
                            sheet_name = sheet_name.replace("_","")[:31]
                    else:
                        temp = sorted(test_scn_name.split("_"), key=lambda x: len(x), reverse=True)
                        if "config" in temp[0].lower():
                            sheet_name = temp[1] + "_" + test_scn_name.split("_")[-1]
                        else:
                            sheet_name = temp[0] + "_" + test_scn_name.split("_")[-1]


                    merged_workbook =  merge_sheets_into_single_sheet(destination_wb = merged_workbook, source_wb = workbook, sheet_name=sheet_name, test_scn_name = test_scn_name)
                    
                    current_worksheet = merged_workbook[sheet_name]
                    #attach image
                    image_folder = os.path.dirname(os.path.abspath(__file__)) + "\\..\\..\\..\\..\\CustomerPrj\\Traces"
                    r=current_worksheet.max_row + 4
                    for (dirpath_img, dirnames_img, img_filenames) in os.walk(image_folder):
                        for img_name in img_filenames:
                            if ".jpg" in img_name.lower() and file.replace("RBS_KPI_","").replace(".xlsx","") in img_name:
                                image_path = image_folder+r"\\"+img_name

                                img = Image(image_path)

                                img.width = 850  
                                img.height = 420 

                                current_worksheet.add_image(img, 'A'+str(r))  
                                r+=25
                    workbook.close()
                    logger.info(f"Merged -> {file}")
        default_sheet = merged_workbook['Sheet']
        merged_workbook.remove(default_sheet)
        merged_workbook.save(main_folder+"\\"+"RESTBUS_KPI.xlsx")
        merged_workbook.close()
        logger.info(f"Final restbus KPI stats excel saved --> RESTBUS_KPI.xlsx")
        time.sleep(10)


    except Exception as e:
        logger.error(f"Error while merging Rest Bus KPI excel --> {e}")
        return






def getAutosarData(autosar_path):
    """
    finds the fastest and slowest frame in a network from the autosar_excel

    Args:
      autosar_path (str): path of autosar_excel

    Returns:

    """
    logger.info("+++++Start parsing Autosar Excel++++++")
    dict_data = {}
    df = pd.read_excel(autosar_path, sheet_name="SysVarDatabase")
    df = filter_column_df(df, "Message", "Signal")
    unique_networks = list(df["network_name"].unique())

    for nwt in unique_networks:
        if nwt not in dict_data.keys():
            dict_data[nwt] = {}
        df_nwt = df[df["network_name"] == nwt]
        cycle_time_list = sorted(list(df_nwt["Cycle Time [ms]"].unique()), reverse=True)

        if 0 in cycle_time_list:
            cycle_time_list.remove(0)

        if cycle_time_list == []:
            dict_data[nwt]["fastest"] = None
            dict_data[nwt]["slowest"] = None
            logger.info(f"No Cycle Time found in {nwt} network data")
            continue
        else:
            # find fastest frame
            min_cycle = min(cycle_time_list)  # ***
            df_min_cycle = df_nwt[df_nwt["Cycle Time [ms]"] == min_cycle]
            msg_ids = [int(i, 16) for i in list(df_min_cycle["Message ID"].unique())]
            min_msg_id = min(msg_ids)  # ***
            min_msg_name = list(df_min_cycle[df_min_cycle["Message ID"] == hex(min_msg_id).upper()]["Message"].unique())[0]  # ***
            required_sig_list.append(list(df_min_cycle[df_min_cycle["Message ID"] == hex(min_msg_id).upper()]["Signal"].unique())[0])
            dict_data[nwt]["fastest"] = {"msg_name": min_msg_name, "msg_id": min_msg_id, "cycle_time": min_cycle}
            # print("FASTEST - > ",nwt,"->",min_msg_id,"->",min_msg_name)

            if len(cycle_time_list) == 1:
                if len(msg_ids) == 1:
                    dict_data[nwt]["slowest"] = None
                    continue
                max_msg_id = max(msg_ids)
                max_msg_name = list(df_min_cycle[df_min_cycle["Message ID"] == hex(max_msg_id).upper()]["Message"].unique())[0]  # ***
                required_sig_list.append(list(df_min_cycle[df_min_cycle["Message ID"] == hex(max_msg_id).upper()]["Signal"].unique())[0])
                dict_data[nwt]["slowest"] = {"msg_name": max_msg_name, "msg_id": max_msg_id, "cycle_time": min_cycle}
                # print("SLOWEST - > ",nwt,"->",max_msg_id,"->",max_msg_name)
            else:
                dict_data[nwt]["slowest"] = None
                for max_cycle in cycle_time_list:
                    df_max_cycle = df_nwt[df_nwt["Cycle Time [ms]"] == max_cycle]
                    msg_ids = [int(i, 16) for i in list(df_max_cycle["Message ID"].unique())]
                    max_msg_id = max(msg_ids)  # ***
                    if max_msg_id <= dict_data[nwt]["fastest"]["msg_id"]:
                        continue
                    max_msg_name = list(df_max_cycle[df_max_cycle["Message ID"] == hex(max_msg_id).upper()]["Message"].unique())[0]  # ***
                    required_sig_list.append(list(df_max_cycle[df_max_cycle["Message ID"] == hex(max_msg_id).upper()]["Signal"].unique())[0])
                    dict_data[nwt]["slowest"] = {"msg_name": max_msg_name, "msg_id": max_msg_id,"cycle_time": max_cycle}
                    break
    for nwt, data in dict_data.items():
        logger.info(f"{nwt} -> {data}")
    logger.info("+++++END parsing Autosar Excel++++++")
    return dict_data

def initInfo(dict_init_sigs):
    """
    finds the init signal information and prints the output to console
    Args:
      dict_init_sigs (dict): init variable info
    """
    logger.info("+++++Start find init info+++++")
    try:
        if ("init_all" in dict_init_sigs.keys()) and (1 in dict_init_sigs["init_all"].samples):
            init_start_time = dict_init_sigs["init_all"].timestamps[list(dict_init_sigs["init_all"].samples).index(1)]
            logger.info(f"init_all time = {init_start_time}")
        else:
            logger.warning(f"init_all not triggered or value not logged")
            return

        if ("init_done" in dict_init_sigs.keys()) and (1 in dict_init_sigs["init_done"].samples):
            init_done_time = dict_init_sigs["init_done"].timestamps[list(dict_init_sigs["init_done"].samples).index(1)]
            logger.info(f"init_done time = {init_done_time}")
        else:
            logger.warning(f"init_done not triggered or value not logged")
            return

        if init_done_time < init_start_time:
            logger.warning(f"init_done_time < init_start_time (init_all/init_done logged twice)")
        else:
            pass
            #logger.info(f"KpiINFO_init#:#Time taken for init_all = {round((init_done_time - init_start_time),3)}s#:#")

    except Exception as e:
        logger.warning(f"while finding initInfo -->{e}")

    logger.info("+++++End find init info+++++")

def find_nearest_indexx(value, value_list):
    """Function to find the index of the nearest value in a list."""
    nearest_index = min(range(len(value_list)), key=lambda i: abs(value_list[i] - value))
    return nearest_index

def parseTrace(log_file_path, dict_autosar_data, rest_bus_log_path):
    """
    trace is MF4 frame logged

    Args:
      log_file_path (str): log file path
      dict_autosar_data (dict): extracted autosar_info

    """
    global trace_fl_name
    try:
        trace_fl_name=log_file_path.split("\\")[-1]
        test_case_name  = log_file_path.split("\\")[-1].replace(".mf4","")
        logger.info(f"+++++ Start Parsing Trace Log File --> {trace_fl_name} ++++")
        dict_check = {}
        
        for nwt, data in dict_autosar_data.items():
            dict_check[nwt] = [i["msg_name"] for i in data.values() if i]
        #print(log_file_path)
        mdf = MDF(log_file_path)
        #print("mdf version is : ", mdf.version)
        mdf.save(log_file_path, overwrite=True)
        #print(mdf.version)
        try:
            scenario_start = mdf.get('scenario_start')
            scenario_start_samples = scenario_start.samples
            #print(scenario_start_samples)
            for index, item in enumerate(scenario_start_samples):
                if item == 1:
                    break
            scenario_start_time_stamp_value = scenario_start.timestamps[index]
        except Exception as e:
            logger.warning(f"scenario_start not found in trace file--> {trace_fl_name} ++++")
        required_sig_filter=[]
        for req_sig in required_sig_list:
            indexs = mdf.whereis(req_sig)
            for tup in indexs:
                required_sig_filter.append([req_sig]+list(tup))
                
        file = mdf.filter(required_sig_filter)
        mdf.close()

        temp_list = []
        dict_frame_sigs = {}
        dict_can_bus_stats ={}
        dict_canoe_stats = {}
    
        dict_init_sigs = {}
        for sig in file.iter_channels():
            if scenario_start:
                timestamp_values = sig.timestamps
                start_idx = find_nearest_indexx(scenario_start_time_stamp_value, timestamp_values)
                sig.timestamps = sig.timestamps[start_idx:]
                sig.samples = sig.samples[start_idx:]
            #init_all
            if sig.name in ["init_all","init_done"] and sig.name not in dict_init_sigs.keys():
                dict_init_sigs[sig.name] = sig
                continue
    
            #can bus stats
            if sig.name == "CAN_BusStatistic":
                if sig.source.path not in dict_can_bus_stats:
                    dict_can_bus_stats[sig.source.path] = sig
                continue
    
            #canoe stats
            if sig.name == "CPUUsageTotal":
                if sig.source.name not in dict_canoe_stats:
                    dict_canoe_stats[sig.source.name] = {}
                dict_canoe_stats[sig.source.name]["CPUUsageTotal"] = sig
                continue

            if sig.name == "CPUUsageProcess":
                if sig.source.name not in dict_canoe_stats:
                    dict_canoe_stats[sig.source.name] = {}
                dict_canoe_stats[sig.source.name]["CPUUsageProcess"] = sig
                continue

            elif sig.name == "MainQueueLoad":
                if sig.source.name not in dict_canoe_stats:
                    dict_canoe_stats[sig.source.name] = {}
                dict_canoe_stats[sig.source.name]["MainQueueLoad"] = sig
                continue
                
            elif sig.name == "TimerAccuracy":
                if sig.source.name not in dict_canoe_stats:
                    dict_canoe_stats[sig.source.name] = {}
                dict_canoe_stats[sig.source.name]["TimerAccuracy"] = sig
                continue
            

    
            #frame stats
            if sig.display_names:
                dict_sig_info = {value: key for key, value in sig.display_names.items()}
                nwt,msg_name = "",""
                if "bus_name" in dict_sig_info.keys():
                    nwt = dict_sig_info["bus_name"].split(".")[0]
                else:
                    continue
                if "message_name" in dict_sig_info.keys():
                    msg_name = dict_sig_info["message_name"].split(".")[0]
                else:
                    continue
    
                if nwt in dict_channel_nwtName_mapping.keys():
                    if dict_channel_nwtName_mapping[nwt] not in dict_frame_sigs.keys():
                        dict_frame_sigs[dict_channel_nwtName_mapping[nwt]] = {}
    
                    if msg_name in dict_check[dict_channel_nwtName_mapping[nwt]]:
                        dict_frame_sigs[dict_channel_nwtName_mapping[nwt]][msg_name] = sig
                        logger.info(f"Found {msg_name} in {nwt} -> {dict_channel_nwtName_mapping[nwt]}")
                        dict_check[dict_channel_nwtName_mapping[nwt]].remove(msg_name)
                else:
                    temp_list.append(nwt)
                    temp_list=list(set(temp_list))

        if temp_list:
            logger.info(f"{temp_list} -> not defined in nwt name-trace name mapping")
        logger.info("+++END parse Log+++")

        #find stats
        initInfo(dict_init_sigs)
        findFrameStatistics(dict_autosar_data,dict_frame_sigs,test_case_name, rest_bus_log_path)
        findCanBusStatistics(dict_can_bus_stats, test_case_name, rest_bus_log_path)
        findCanoeStatistics(dict_canoe_stats, test_case_name, rest_bus_log_path)

        #to excel
        if dict_global_kpi=={}:
            logger.warning("********* No KPI info found in Trace file ***********")
            return
        stat_file_name = "RBS_KPI_" + log_file_path.split("\\")[-1].replace(".mf4",".xlsx")
        excel_path = os.getcwd() + r"\\..\\..\\..\\..\\KPI_OUTPUT\\" + stat_file_name
        writer = pd.ExcelWriter(excel_path)
        for sheet, df in dict_global_kpi.items():
            df.to_excel(writer, sheet_name=sheet, index=False)
        writer.close()
        logger.info(f"KPI stats saved -> {excel_path}")
        time.sleep(1)

    except Exception as e:
        logger.warning(f"ERROR in KPI info extract --> {e}")


def findFrameStatistics(dict_autosar_data, dict_frame_sigs, test_case_name, rest_bus_log_path):
    """
    finds frame statistics

    Args:
      dict_autosar_data (dict): extracted autosar info from excel
      dict_frame_sigs (dict): extracted info from trace file

    """
    try:
        rest_bus_log = open(rest_bus_log_path, 'a')
        logger.info("+++++Start Calculate Frame Statistics +++++")
        if dict_frame_sigs=={}:
            logger.warning(f"Frames not found in Trace file (fastest/slowest)")
            return
        columns=['Network name', 'Prio', 'Msg ID', 'Message', 'Expected CycleTime (ms)', 'Min CycleTime (ms)', 'Max CycleTime (ms)', 'Avg CycleTime (ms)', 'Count','Deviation % (expected vs average)']
        #logger.info(f"KpiINFO_FrameStatsColumns#:#" + ('#:#'.join(columns)) + "#:#")
    
        prios = ["fastest","slowest"]

        not_found_trace = []

        Final_Statistics = []
        for network_name, nwt_info in dict_autosar_data.items():
            if network_name not in dict_frame_sigs.keys():
                not_found_trace.append(network_name)
                continue
            for prio in prios:
                temp_dict= {}
                dict_prio = nwt_info[prio]
                if dict_prio:
                    msg_name = dict_prio["msg_name"]
                    if msg_name in dict_frame_sigs[network_name].keys():
                        sig = dict_frame_sigs[network_name][msg_name]
                        diffs = np.diff(sig.timestamps)
                        avg = round(np.mean(diffs) *1000, 4)
                        minn = round(np.min(diffs) *1000, 4)
                        maxx = round(np.max(diffs) *1000, 4)
                        deviation = round((abs(avg - dict_prio['cycle_time']) / dict_prio['cycle_time']) * 100, 3)
                        #rest_bus_log.write(f'----{test_case_name}----\n')
                        if deviation > 5: 
                            logger.info(f'In test_case : {test_case_name} : Deviation% of {network_name} is {deviation}%\n')
                            rest_bus_log.write(f'[ERROR] : {test_case_name} : Deviation% of {network_name} is {deviation}%\n')
                        temp_dict = {'Network name':network_name, 'Prio':prio, 'Msg ID': hex(dict_prio['msg_id']).upper(), "Message": msg_name,
                                     'Expected CycleTime (ms)':dict_prio['cycle_time'],
                                     'Min CycleTime (ms)':minn, 'Max CycleTime (ms)':maxx, 'Avg CycleTime (ms)':avg, 'Count':len(diffs),
                                     'Deviation % (expected vs average)':deviation}
                        Final_Statistics.append(temp_dict)
                        #logger.info(f"KpiINFO_FrameStatsRow#:#" + ('#:#'.join([str(i) for i in temp_dict.values()])) + "#:#")
    
                    else:
                        logger.warning(f"NOT FOUND IN TRACE -> {network_name} -> {prio} -> {msg_name}")
                        temp_dict = {'Network name': network_name, 'Prio': prio, "Message": msg_name,
                                     'Expected CycleTime (ms)': dict_prio['cycle_time'],
                                     'Msg ID': hex(dict_prio['msg_id']).upper(),
                                     'Min CycleTime (ms)': "-", 'Max CycleTime (ms)': "-", 'Avg CycleTime (ms)': "-",
                                     'Count': "-",
                                     'Deviation % (expected vs average)': "-"}
                        Final_Statistics.append(temp_dict)
                        #logger.info(f"KpiINFO_FrameStatsRow#:#" + ('#:#'.join([str(i) for i in temp_dict.values()])) + "#:#")
                else:
                    logger.info(f"{network_name} -> {prio} -> None")
                    temp_dict = {'Network name': network_name, 'Prio': prio, "Message": "-",
                                 'Expected CycleTime (ms)': "-",
                                 'Msg ID': "-",
                                 'Min CycleTime (ms)': "-", 'Max CycleTime (ms)': "-", 'Avg CycleTime (ms)': "-",
                                 'Count': "-",
                                 'Deviation % (expected vs average)': "-"}
                    Final_Statistics.append(temp_dict)
                    #logger.info(f"KpiINFO_FrameStatsRow#:#" + ('#:#'.join([str(i) for i in temp_dict.values()])) + "#:#")
    
    
        df = pd.DataFrame(Final_Statistics, columns=columns)
        dict_global_kpi["Frame_Statistics"] = df

        if not_found_trace:
            logger.warning(f"{not_found_trace} not found in trace")
        logger.info("+++++END Calculate Frame Statistics +++++")
        rest_bus_log.close()
    
        #df.to_excel("Frame_Statistics.xlsx", index=False)
        
    except Exception as e:
        logger.warning(f"ERROR in KPI info extract --> findFrameStats --> {e}")


def findCanBusStatistics(dict_can_stats, test_case_name, rest_bus_log_path):
    """
    finds the can bus stats from extracted info from trace file

    Args:
      dict_can_stats (dict): extracted info from trace

    """
    if dict_can_stats=={}:
        logger.warning("+++++ CanBus Statistics not found+++++")
        return
    cols = ["Channel","Nwt name","Avg Busload (%)","Max Busload (%)","Total ErrorFrames"]
    logger.info("+++++Start Find CAN Statistics +++++")
    # logger.info(f"KpiINFO_CanBusStatsColumns#:#"+('#:#'.join(cols))+"#:#")
    try:
        rest_bus_log = open(rest_bus_log_path, 'a')
        list_can_data = []
        for bus, sig in dict_can_stats.items():
            total_error_frames = 0
            max_bus_load = 0
            samples = list(sig.samples)
            bus_load = []
            for item in samples:
                total_error_frames += item[-2]
                if item[1] > max_bus_load:
                    max_bus_load = item[1]
                bus_load.append(item[1])
            if bus in dict_channel_nwtName_mapping.keys():
                row = [bus, dict_channel_nwtName_mapping[bus], round((sum(bus_load)/len(bus_load))*0.01,2), round((max_bus_load * 0.01), 2),total_error_frames]
                if row[-1] > 10:
                    logger.info(f'In TestCase : {test_case_name} : total_error_frames of {bus} is : {total_error_frames} is > 10\n')
                    rest_bus_log.write(f'[ERROR] : {test_case_name} : total_error_frames of {bus} is {total_error_frames} is > 10\n')
                list_can_data.append(row)
            # logger.info(f"KpiINFO_CanBusStatsRow#:#"+('#:#'.join([str(i) for i in row]))+"#:#")

        df = pd.DataFrame(list_can_data, columns=cols)
        dict_global_kpi["CAN Statistics"] = df
        logger.info("+++++ END Find CAN Statistics +++++")
        rest_bus_log.close()
    except Exception as e:
        logger.warning(f"While finding CAN statistics -->{e}")



def findCanoeStatistics(dict_canoe_stats, test_case_name, rest_bus_log_path):
    """
    find canoe Statistics from extracted info from trace file

    Args:
      dict_canoe_stats (dict): extracted info from trace file

    """
    if dict_canoe_stats=={}:
        logger.warning("+++++ Canoe Statistics not found+++++")
        return
    logger.info("+++++Start Find Canoe Statistics +++++")
    cols = ["Environment","Avg CPU Usage (%)", "Max CPU Usage (%)", "Max CPUUsageProcess (%)", "Avg CPUUsageProcess (%)", "Avg MainQueueLoad (%)", "Max MainQueueLoad (%)", "Timer accuracy (ms)"]
    #logger.info(f"KpiINFO_CanoeStatsColumns#:#" + ('#:#'.join(cols)) + "#:#")
    try:
        rest_bus_log = open(rest_bus_log_path, 'a')
        list_canoe_data = []
        for canoe_runevn, env_data_dict in dict_canoe_stats.items():
            row={}
            row["Environment"] = "RT_rack" if "Simulation" in canoe_runevn else "GUI_PC"
            for parameter, sig in env_data_dict.items():
                samples = list(sig.samples)
                timestamps = list(sig.timestamps)
                if "MainQueueLoad" in parameter:
                    row["Max MainQueueLoad (%)"] = max(samples)
                    row["Avg MainQueueLoad (%)"] = round(sum(samples)/len(samples),2)
                    if row["Max MainQueueLoad (%)"] !=0 and not np.isnan (row["Max MainQueueLoad (%)"]):
                        rest_bus_log.write(f'[ERROR] : {test_case_name} : MainQueueLoad of {row["Environment"]} is !=0 or != None {row["Max MainQueueLoad (%)"]}\n')
                elif "CPUUsageTotal" in parameter:
                    row["Max CPU Usage (%)"] = max(samples)
                    row["Avg CPU Usage (%)"] = round(sum(samples)/len(samples),2)
                    if row["Max CPU Usage (%)"] > 50:
                        rest_bus_log.write(f'[ERROR] : {test_case_name} : CPU Usage of {row["Environment"]} is {row["Max CPU Usage (%)"]}\n')
                elif "CPUUsageProcess" in parameter:
                    row["Max CPUUsageProcess (%)"] = max(samples)
                    row["Avg CPUUsageProcess (%)"] = round(sum(samples)/len(samples),2)
                    if row["Max CPUUsageProcess (%)"] > 50:
                        rest_bus_log.write(f'[ERROR] : {test_case_name} : CPUUsageProcess of {row["Environment"]} is {row["Max CPUUsageProcess (%)"]}\n')
                elif "TimerAccuracy" in parameter:
                    samples = [abs(sample / 1000000) for sample in samples]
                    row["Timer accuracy (ms)"] = max(samples)
                    if row["Timer accuracy (ms)"] > 5:
                        rest_bus_log.write(f'[ERROR] : {test_case_name} : Timer accuracy of {row["Environment"]} is {row["Timer accuracy (ms)"]}\n')
                else:
                    continue

                #max_cpu_timestamps = [round(timestamps[i], 3) for i in range(len(samples)) if samples[i] == max_cpu]
                # if len(max_cpu_timestamps) == 1:
                #     max_cpu_timestamps = max_cpu_timestamps[0]
                # row = [env, round(sum(samples)/len(samples),2) ,max_cpu, max_cpu_timestamps]

                #logger.info(f"KpiINFO_CanoeStatsRow#:#" + ('#:#'.join([str(i) for i in row])) + "#:#")
                
                image_name =  os.path.dirname(os.path.abspath(__file__)) + "\\..\\..\\..\\..\\CustomerPrj\\Traces\\" + trace_fl_name.replace(".mf4","") +"_"+ canoe_runevn.replace(":","")+"_"+parameter+ ".jpg"
                
                max_idx = np.argmax(samples)
                avg_value = np.mean(samples)

                plt.figure(figsize=(10,6))
                plt.ylim(0, 100)
                plt.plot(timestamps,samples,linewidth=1, color='navy')
                plt.scatter(timestamps[max_idx], samples[max_idx],color='red', zorder=10)
                #plt.axvline(x=timestamps[max_idx],color='red', linestyle='--',linewidth=1)
                plt.annotate(f"Max: {samples[max_idx]}% at {round(timestamps[max_idx],3)}s", xy=(timestamps[max_idx],samples[max_idx]), xytext = (timestamps[max_idx], samples[max_idx]+5), arrowprops = dict(facecolor='black', arrowstyle='->',linewidth=0.5))

                plt.axhline(y=avg_value,color='grey', linestyle='--',linewidth=1)
                plt.annotate(f"Avg: {round(avg_value,2)}%", xy=(timestamps[0]-20,avg_value), xytext = (timestamps[0]-60, samples[max_idx]+2))
                plt.xlabel("Timestamp [s]")
                if "TimerAccuracy" in parameter:
                    plt.ylabel(f"{parameter} [ms]")
                else:
                    plt.ylabel(f"{parameter} [%]")
                plt.title(canoe_runevn+"::"+parameter)
                plt.savefig(image_name)
                logger.info(f"Graph saved --> {image_name}")
                plt.close()

            list_canoe_data.append(row)

        df = pd.DataFrame(list_canoe_data,columns=cols)

        dict_global_kpi["Canoe Statistics"] = df
        logger.info("+++++ END Find Canoe Statistics +++++")
        rest_bus_log.close()
    except Exception as e:
        logger.warning(f"While finding Canoe statistics/graph generation -->{e}")


# def external_call():
#     script_path = os.path.dirname(os.path.abspath(__file__))
#     autosar_path = script_path + r'\..\..\..\..\CustomerPrj\Restbus\Autosar_Gen_Database.xlsx'
#     dict_autosar_data = getAutosarData(autosar_path)
#     platform_path = r"D:\\Jenkins\\ADAS_Platform"
#     #canoe_cfg = get_canoe_cfg(platform_path)
#
#     #conf_full_path = platform_path + "\\" + canoe_cfg
#     xcp_full_path = platform_path + r"\\CustomerPrj\\XCP\\XCP_config_gen.xcpcfg"
#     logging_path = platform_path + r"\\CustomerPrj\\Traces\\Frame_logged_T001.mf4"
#     test_unit = "SystemIntegration"
#     #logged_path = logTrace(conf_full_path,platform_path,xcp_full_path,test_unit,logging_path)
#     parseTrace(logging_path, dict_autosar_data)
#     #findStatistics(dict_autosar_data, dict_frame_sigs)
#
# if __name__ == "__main__":
#     external_call()