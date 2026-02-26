# -*- coding: utf-8 -*-
# @file excel_table_summary.py
# @author ADAS_HIL_TEAM
# @date 08-21-2023

##################################################################
# C O P Y R I G H T S
# ----------------------------------------------------------------
# Copyright (c) 2023 by Robert Bosch GmbH. All rights reserved.

# The reproduction, distribution and utilization of this file as
# well as the communication of its contents to others without express
# authorization is prohibited. Offenders will be held liable for the
# payment of damages. All rights reserved in the event of the grant
# of a patent, utility model or design.

##################################################################


import pandas as pd
try:
    pd.set_option('future.no_silent_downcasting', True)
except:
    pass
from openpyxl.styles import PatternFill
import openpyxl


class ExcelTestResultSummary:
    """ """
    columns = {"ID RQM": [],
               "TC name": [],
               "Results SWv": [],
               "Defect PSA": [],
               "Remark": [],
               }

    sub_tc_names_and_status = {}
    helper_list = []

    def __init__(self, test_run_path):
        self.tool_results_folder_path = test_run_path
        self.excel_summary = "testsummary.xlsx"

    def update_table(self, l_dict, oem):
        """
        

        Args:
          l_dict: 
          oem: 

        Returns:

        """
        self.update_data_frame(l_dict, oem)
        self.create_excel_file_pandas()

    def update_data_frame(self, l_dict, oem):
        """
        

        Args:
          l_dict: 
          oem: 

        Returns:

        """
        for k, v in l_dict["tc_sub_cases"].items():
            self.columns["TC name"].append(l_dict["tc_name"])
            self.columns["ID RQM"].append(k)
            self.columns["Results SWv"].append(v)
            self.columns["Defect PSA"].append("")
            if oem:
                self.columns["Remark"].append("")
            else:
                self.columns["Remark"].append("Not run because test case is not applicable for this variant.")

    @staticmethod
    def fill_cells_with_color(excelpath):
        """
        

        Args:
          excelpath: 

        Returns:

        """
        fill_cell_fail = PatternFill(patternType='solid', fgColor='FF0000')
        fill_cell_pass = PatternFill(patternType='solid', fgColor='92D050')
        fill_cell_skip = PatternFill(patternType='solid', fgColor='FFC000')
        try:
            wb = openpyxl.load_workbook(excelpath)
            sh = wb['Sheet1']
            for i in range(1, sh.max_row + 1):
                cellname = "C" + str(i)
                cell_obj = sh.cell(row=i, column=3)
                # print(f"data == {cell_obj.value} and type is {type(cell_obj.value)} ")
                if 'Failed' in cell_obj.value:
                    sh[cellname].fill = fill_cell_fail
                elif 'Passed' in cell_obj.value:
                    sh[cellname].fill = fill_cell_pass
                elif 'Skipped' or 'Not started' in cell_obj.value:
                    sh[cellname].fill = fill_cell_skip
                elif 'Obsolete' in cell_obj.value:
                    sh[cellname].fill = fill_cell_skip
            wb.save(excelpath)
        except PermissionError as err:
            print(f"PLEASE CLOSE THE SUMMARY FILE IN ORDER FOR THE SUMMARY TO BE COMPLETED!!! {err}")

    def create_excel_file_pandas(self):
        """ """
        # create a dataframe with the columns
        df = pd.DataFrame.from_dict(self.columns)
        try:
            writer = pd.ExcelWriter(self.tool_results_folder_path + self.excel_summary, engine='openpyxl', mode='w')
            df.to_excel(writer, sheet_name='Sheet1', index=False)
            writer.save()
        except PermissionError:
            print("PLEASE CLOSE THE SUMMARY FILE IN ORDER FOR THE SUMMARY TO BE COMPLETED!!!")
        finally:
            self.fill_cells_with_color(self.tool_results_folder_path + self.excel_summary)


