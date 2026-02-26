# -*- coding: utf-8 -*-
# @file cm_to_rbs_mapping.py
# @author ADAS_HIL_TEAM
# @date 08-21-2023

##################################################################
# C O P Y R I G H T S
# ----------------------------------------------------------------
# Copyright (c) 2023 by Robert Bosch GmbH. All rights reserved.

# The reproduction, distribution and utilization of this file as
# well as the communication of its contents to others without express
# authorization is prohibited. Offenders will be held liable for the
# payment of damages. All rights reserved in the event of the grant
# of a patent, utility model or design.

##################################################################


from openpyxl import load_workbook
import pandas as pd
try:
    pd.set_option('future.no_silent_downcasting', True)
except:
    pass

import numpy as np
import re
import os, sys
from collections import Counter
try:
    from Control.logging_config import logger
except ImportError:
    sys.path.append(os.path.dirname(os.path.abspath(__file__)) + r"\..\Control")
    from logging_config import logger
sys.path.append(os.path.dirname(os.path.abspath(__file__)) + r"\..\Rbs_Scripts")
from create_sysvar import generate_sysvar, save_to_file

#allow longer names up to 100 characters
pd.options.display.max_colwidth = 100
pd.options.display.colheader_justify = 'left'

column_names = ["Name", "Message", "Value Type", "Initial Value", "Factor", "Offset", "Minimum",  "Maximum", "Unit", "texttable", "texttable values", "max_value", "variant", "sender", "direction", "texttable mapping", "conversion", "array name", "data_type"]

node_head = '''void cm_hil_vm()
{'''

# read worksheet 'sheet_name' from workbook 'wb'
def generate_dataframe(wb, sheet_name):
    """
    

    Args:
      wb: 
      sheet_name: 

    Returns:

    """
    global node, node_df, num_of_rows, sheet
    global signal_cloe
    global col_name, col_namespace, col_initial_value, col_factor, col_offset, col_minimum, col_maximum, col_unit, col_texttable, col_texttable_values, col_max_value, col_variant, col_sender, col_direction, col_texttable_mapping, col_conversion, col_array_name, col_index
    sheet = sheet_name
    
    node = ['/*@!Encoding:1252*/\n', node_head]
    ws1 = wb[sheet_name]
    #print(node_name, bus_name, bus_type)

    # create the different dataframes
    node_df = pd.DataFrame(ws1.values)
    node_df = node_df.replace(np.nan, '', regex=True)

    # get column names
    col_names = list(node_df.iloc[0])
    # get column index
    try:
        col_index = [col_names.index(col) for col in column_names]
    except:
        logger.error("Column names in Excel are different from the Script:")
        logger.error(';'.join(column_names))
        return False
    [col_name, col_namespace, col_value_type, col_initial_value, col_factor, col_offset, col_minimum, col_maximum, col_unit, col_texttable, col_texttable_values, col_max_value, col_variant, col_sender, col_direction, col_texttable_mapping, col_conversion, col_array_name, col_data_type] = col_index

    signal_cloe.extend(node_df[node_df[col_sender] == "carmaker"][col_name].tolist())
    num_of_rows = node_df.shape[0]
    logger.info(f"Signal count: {num_of_rows-1}")


def generate_onsysvarenum():
    """ """
    global sysvar_df
    message_count = 0
    enum_df = node_df
    enum_df = enum_df[(enum_df[col_texttable] != "")]
    num_of_rows_enum = enum_df.shape[0]
    variants_row = set(enum_df[col_variant].values)
    to_delete = {'', "multi", "variant"}
    variants_row.difference_update(to_delete)

    for i in range(num_of_rows_enum):
        if (enum_df[col_sender].values[i] == "carmaker"):
            # generate sysvar
            add_sysvar_df = enum_df[col_index].iloc[[i]]
            add_sysvar_df.columns = [i for i in range(len(col_index))]
            sysvar_df = pd.concat([sysvar_df, add_sysvar_df])
            if (i == num_of_rows_enum-1) or (enum_df[col_sender].values[i+1] == "carmaker") or (enum_df[col_name].values[i+1] == "na"):
                continue

            if (sheet == "cm_in"):
                namespace = enum_df[col_namespace].values[i]
                signal_name =  enum_df[col_name].values[i]
                namespace_b = []
                signal_name_b = []
                initial_value_b = []
                enum_mapping = []
                if enum_df[col_variant].values[i] == "multi":
                    idx = 1
                    while "variant" in enum_df[col_variant].values[i+idx]:
                        namespace_b.append(enum_df[col_namespace].values[i+idx])
                        signal_name_b.append(enum_df[col_name].values[i+idx])
                        init_val = enum_df[col_initial_value].values[i + idx]
                        if init_val is not None:
                            initial_value_b.append(init_val)
                        else:
                            logger.error(f"Missing initial value for {enum_df[col_name].values[i + idx]} signal")
                            raise Exception("Please check the database")
                        mapping = enum_df[col_texttable_mapping].values[i + idx]
                        if mapping != "":
                            enum_mapping.append(mapping)
                        else:
                            logger.error(f"Missing texttable mapping for {enum_df[col_name].values[i + idx]} signal")
                            raise Exception("Please check the database")
                        if i+idx == num_of_rows_enum:
                            break
                        idx+=1
                    combined_data =  list(zip(namespace_b, signal_name_b, initial_value_b, enum_mapping))
                else:
                    idx = 1
                    while "classe" in enum_df[col_sender].values[i+idx] or "Rbs" in enum_df[col_sender].values[i+idx] or "XCP" in enum_df[col_sender].values[i+idx]:
                        namespace_b.append(enum_df[col_namespace].values[i+idx])
                        signal_name_b.append(enum_df[col_name].values[i+idx])
                        init_val = enum_df[col_initial_value].values[i+idx]
                        if init_val != "":
                            initial_value_b.append(init_val)
                        else:
                            logger.error(f"Missing initial value for {enum_df[col_name].values[i+idx]} signal")
                            raise Exception("Please check the database")
                        mapping = enum_df[col_texttable_mapping].values[i+idx]
                        if mapping != "":
                            enum_mapping.append(mapping)
                        else:
                            logger.error(f"Missing texttable mapping for {enum_df[col_name].values[i+idx]} signal")
                            raise Exception("Please check the database")
                        idx+=1
                        if i+idx == num_of_rows_enum:
                            break
                    combined_data_common = list(zip(namespace_b, signal_name_b, initial_value_b, enum_mapping))


            sender = enum_df[col_sender].values[i+1]
            node.append('')  
            if (sheet == "cm_in"):
                node.append(f'  switch (@CarMaker::{namespace}::{signal_name})')
                node.append('  {')

            if enum_df[col_variant].values[i] != "multi":
                enum_mapping = enum_mapping[0].split()
                enum_length = len(enum_mapping)
                for k in range (enum_length):
                    enum_member = str(enum_mapping[k])
                    enum_member = enum_member.split(':')
                    sig_a_map = enum_member[0]
                    node.append(f'    case {sig_a_map}:')
                    for variant in range(0,len(combined_data_common)):
                        val_temp = combined_data_common[variant][3].split("\n")[k]
                        val = val_temp.split(":")[1]
                        if "XCP" in sender:
                            node.append(f'      @{sender}::{combined_data_common[variant][0][:-1]}::{combined_data_common[variant][1]} = {val};')
                        elif (sheet == "fdx_in" and sheet != "fdx_out" and sender != "classe" ):
                            node.append(f'      @{combined_data_common[variant][0]}::{combined_data_common[variant][1]}_Rv = {val};')
                        else:
                            node.append(f'      @{combined_data_common[variant][0]}::{combined_data_common[variant][1]} = {val};')
                    node.append('      break;')

                node.append('    default:'.format(sig_a_map))
                for variant in range(0,len(combined_data_common)):
                    if "XCP" in sender:
                        node.append(f'      @{sender}::{combined_data_common[variant][0][:-1]}::{combined_data_common[variant][1]} = {combined_data_common[variant][2]};')
                    elif (sheet == "fdx_in" and sheet != "fdx_out" and sender != "classe" ):
                        node.append(f'      @{combined_data_common[variant][0]}::{combined_data_common[variant][1]}_Rv = {combined_data_common[variant][2]};')
                    else:
                        node.append(f'      @{combined_data_common[variant][0]}::{combined_data_common[variant][1]} = {combined_data_common[variant][2]};')
                node.append('      break;')
                node.append('  }')
                message_count += 1
            else:
                enum_mapping = enum_mapping[0].split()
                enum_length = len(enum_mapping)
                for k in range (enum_length):
                    enum_member = str(enum_mapping[k])
                    enum_member = enum_member.split(':')
                    sig_a_map = enum_member[0]
                    node.append(f'    case {sig_a_map}:')
                    for variant in range(0,len(combined_data)):
                        val_temp = combined_data[variant][3].split("\n")[k]
                        val = val_temp.split(":")[1]
                        if (sheet == "cm_in"  and sender != "classe" ):
                            node.append(f'      @{combined_data[variant][0]}::{combined_data[variant][1]}_Rv = {val};')
                        else:
                            node.append(f'      @{combined_data[variant][0]}::{combined_data[variant][1]} = {val};')
                    node.append('      break;')

                node.append('    default:'.format(sig_a_map))
                for variant in range(0,len(combined_data)):
                    if "XCP" in sender:
                        node.append(f'      @{sender}::{combined_data[variant][0][:-1]}::{combined_data[variant][1]} = {combined_data[variant][2]};')
                    elif (sheet == "fdx_in" and sheet != "fdx_out" and sender != "classe" ):
                        node.append(f'      @{combined_data[variant][0]}::{combined_data[variant][1]}_Rv = {combined_data[variant][2]};')
                    else:
                        node.append(f'      @{combined_data[variant][0]}::{combined_data[variant][1]} = {combined_data[variant][2]};')
                node.append('      break;')
                node.append('  }')
                message_count += 1
    node.append('}')
    logger.info(f"enum int: generated with {str(message_count)} messages.")



def generate_onsysvardouble():
    """ """
    global sysvar_df
    message_count = 0
    sysvardouble_df = node_df
    sysvardouble_df = sysvardouble_df[(sysvardouble_df[col_texttable] != "texttable")]
    num_of_rows_double = sysvardouble_df.shape[0]
    variants_row = set(sysvardouble_df[col_variant].values)
    to_delete = {'', "multi", "variant"}
    variants_row.difference_update(to_delete)
    number_of_variants = len(variants_row)
    variants_row = list(variants_row)
    variants_row.sort()

    for i in range(num_of_rows_double):
        if sysvardouble_df[col_sender].values[i] == "carmaker":
            # generate sysvar
            add_sysvar_df = sysvardouble_df[col_index].iloc[[i]]
            add_sysvar_df.columns = [i for i in range(len(col_index))]
            sysvar_df = pd.concat([sysvar_df, add_sysvar_df])
            if (i == num_of_rows_double-1) or (sysvardouble_df[col_sender].values[i+1] == "carmaker") or (sysvardouble_df[col_name].values[i+1] == "na"):
                continue
            if sheet == "cm_in":
                if sysvardouble_df[col_variant].values[i] == "multi":
                    namespace = sysvardouble_df[col_namespace].values[i]
                    signal_name = sysvardouble_df[col_name].values[i]
                    # direction = sysvardouble_df[col_direction].values[i]
                    # texttable = sysvardouble_df[col_texttable].values[i]
                    sig_a_offset = sysvardouble_df[col_offset].values[i]
                    sig_a_factor = sysvardouble_df[col_factor].values[i]
                    sig_a_min = sysvardouble_df[col_minimum].values[i]
                    sig_b_conversion = sysvardouble_df[col_conversion].values[i]
                    sig_b_offset = []
                    sig_b_factor = []
                    sig_b_min = []
                    sig_b_name = []
                    sig_b_namespace = []
                    for variant in range(1,number_of_variants+1):
                        sig_b_offset.append(sysvardouble_df[col_offset].values[i+variant])
                        sig_b_factor.append(sysvardouble_df[col_factor].values[i+variant])
                        sig_b_min.append(sysvardouble_df[col_minimum].values[i+variant])
                        sig_b_name.append(sysvardouble_df[col_name].values[i+variant])
                        sig_b_namespace.append(sysvardouble_df[col_namespace].values[i+variant])
                else:
                    namespace = sysvardouble_df[col_namespace].values[i]
                    signal_name = sysvardouble_df[col_name].values[i]
                    sig_b_name = sysvardouble_df[col_name].values[i + 1]
                    sig_b_namespace = sysvardouble_df[col_namespace].values[i + 1]
                    # direction = sysvardouble_df[col_direction].values[i]
                    # texttable = sysvardouble_df[col_texttable].values[i]
                    sig_a_offset = sysvardouble_df[col_offset].values[i]
                    check_int_double_integrity("offset",sig_a_offset,signal_name)
                    sig_a_factor = sysvardouble_df[col_factor].values[i]
                    check_int_double_integrity("factor", sig_a_factor, signal_name)
                    sig_b_offset = sysvardouble_df[col_offset].values[i+1]
                    check_int_double_integrity("offset", sig_b_offset, sig_b_name)
                    sig_b_factor = sysvardouble_df[col_factor].values[i+1]
                    check_int_double_integrity("factor", sig_b_factor, sig_b_name)
                    sig_b_conversion = sysvardouble_df[col_conversion].values[i]
                    check_int_double_integrity("conversion", sig_b_conversion, sig_b_name)
                    sig_a_min = sysvardouble_df[col_minimum].values[i]
                    check_int_double_integrity("minimum", sig_a_min, signal_name)
                    sig_b_min = sysvardouble_df[col_minimum].values[i+1]
                    check_int_double_integrity("minimum", sig_b_min, sig_b_name)

            sender = sysvardouble_df[col_sender].values[i+1]
            if sheet == "cm_in":
                if sig_b_conversion != 1:
                    node.append(f'  @{sig_b_namespace}::{sig_b_name} = @CarMaker::{namespace}::{signal_name} * ({sig_b_conversion});')
                else:
                    node.append(f'  @{sig_b_namespace}::{sig_b_name} = @CarMaker::{namespace}::{signal_name};')
            elif sheet == "fdx_out" and sender != "classe":
                node.append(f'on signal_update {namespace}::{signal_name}')

            message_count += 1

    logger.info(f"double signal : generated with {str(message_count)} messages.")


def check_int_double_integrity(value_type, data, signal_name):
    """
    

    Args:
      value_type: 
      data: 
      signal_name: 

    Returns:

    """
    if data == "":
        logger.error(f"Missing {value_type} value for {signal_name} signal")
        raise Exception("Please check the database")


def save_file(file_path):
    """
    

    Args:
      file_path: 

    Returns:

    """
    # Open and save node .cin file
    file = f'{file_path}/cm_mapping.cin'
    with open(file, 'w') as outfile:
        outfile.write("\n".join(str(item) for item in node))
        outfile.close()
    logger.info(f"{file} updated successfully")


# main function to create nodes
def create_mapping_main_cm(workbook, file_path):
    """
    

    Args:
      workbook: 
      file_path: 

    Returns:

    """
    global sysvar_df, signal_cloe
    sheet_list = workbook.sheetnames
    sysvar_df = pd.DataFrame([column_names])
    signal_cloe = ["Name"]
    for sheet in sheet_list:
        logger.info(f"++++++++++Creating mapping for: {sheet} ++++++++++")
        try:
            if sheet == 'cm_in':
                generate_dataframe(workbook, sheet)
                generate_onsysvardouble()
                generate_onsysvarenum()
                save_file(file_path+"/Nodes")
        except Exception as exp:
            logger.error(f"Failed to create node for: {sheet}")
            logger.error(f"Error: {exp}")
            return False
    return True


def external_call():
    """ """
    try:
        logger.info("###### START 'Update CarMaker Nodes' DEBUG INFORMATION ######")
        script_path = os.path.dirname(os.path.abspath(__file__))
        wb = load_workbook(script_path + '/../../../../Platform/Carmaker/CM_Database.xlsx', data_only=True)
        create_mapping_main_cm(wb, script_path+'/../../../../Platform/Carmaker/')
    except Exception as e:
        logger.error(f"Update Update CarMaker Nodes failed. Exception --> {e}")
    logger.info("###### END 'Update CarMaker Nodes' DEBUG INFORMATION ######")
    logger.info('-' * 80)

if __name__ == "__main__":
    external_call()
