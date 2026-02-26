# -*- coding: utf-8 -*-
# @file report_handler.py
# @author ADAS_HIL_TEAM
# @date 08-21-2023

##################################################################
# C O P Y R I G H T S
# ----------------------------------------------------------------
# Copyright (c) 2023 by Robert Bosch GmbH. All rights reserved.

# The reproduction, distribution and utilization of this file as
# well as the communication of its contents to others without express
# authorization is prohibited. Offenders will be held liable for the
# payment of damages. All rights reserved in the event of the grant
# of a patent, utility model or design.

##################################################################


import codecs
import os.path
import re
import datetime as dt
import time
from fabric import Connection
import ssh_client
import linecache
import pandas as pd
try:
    pd.set_option('future.no_silent_downcasting', True)
except:
    pass
from openpyxl.styles import PatternFill
import openpyxl
from zipfile import ZipFile, ZIP_DEFLATED
import pathlib
from bs4 import BeautifulSoup


class ReportHandler:
    """ """

    current_tc_status = ''

    def __init__(self, **conn_info):
        self.exc_h = ExcelTestResultSummary()
        self.excel_summary = "testsummary.xlsx"

        self.conn_info = dict()
        self.conn_info.update(conn_info)
        self.host = self.conn_info['hostname']
        self.username = self.conn_info['username']
        self.password = self.conn_info['password']
        temp_path = self.conn_info['test_repo_path']
        suffix = '/' + self.conn_info['test_repo_keyword']
        self.path = temp_path.replace(suffix, '')

        # IMPORTANT: sftp_obj bellow is used only to check if a path is a dir or a file. So if we include in execute client a function that checks that,
        # this object init can be removed and its functions replaced
        self.sftp_obj = ssh_client.SSHClient(**self.conn_info)
        self.connection = Connection(host=self.host, user=self.username, connect_kwargs={"password": self.password})

        self.date_now = str(dt.date.today()) + "_"
        self.time_now = str(time.strftime("%Hh%Mm%Ss"))
        self.def_evald_report_name = "index.html"
        self.tool_results_folder_path = self.conn_info['local_report_output_path']
        self.current_test_run_name = "tc_run_" + self.date_now + self.time_now
        self.local_evald_rf_path = self.tool_results_folder_path + "evald_report\\"
        self.linux_evald_rf_path = self.conn_info['linux_evald_reports_path'] + '/'
        self.mf4_path_linux = self.conn_info['linux_mf4_log_path']
        self.mf4_full_path = ''

    @staticmethod
    def compress_mf4_log(dir_path, mf4_name):
        """
        

        Args:
          dir_path: 
          mf4_name: 

        Returns:

        """
        windows_path = pathlib.Path(dir_path)
        mf4_path = dir_path + mf4_name + '.mf4'
        win_mf4_path = pathlib.Path(mf4_path)
        mf4_zip = dir_path + mf4_name + '.zip'
        try:
            with ZipFile(mf4_zip, 'w', ZIP_DEFLATED, compresslevel=9) as zf:
                zf.write(win_mf4_path, arcname=win_mf4_path.relative_to(windows_path))
        except FileNotFoundError:
            print(f"MF4 log not found")
        else:
            os.remove(mf4_path)

    # trqbva da se kopira v tool_results_folder_path + current_test_run_name + test_status + test_case_name
    def copy_mf4_log(self, mf4_path_linux, local_mf4_path, test_case_id):
        """
        

        Args:
          mf4_path_linux: 
          local_mf4_path: 
          test_case_id: 

        Returns:

        """
        try:
            with self.connection.cd(mf4_path_linux):  # enter the linux evald_report folder
                list_of_mf4_linux_logs = self.sftp_obj.list_directory(mf4_path_linux)  # list the items of the report folder
                # print(f"List with items from list_of_mf4_linux_logs = {list_of_mf4_linux_logs}")
                for mf4_log in list_of_mf4_linux_logs:
                    if mf4_log == test_case_id + ".mf4":  # if mf4 log is found
                        # print(f"DEBUG____ mf4 log found = {mf4_log}")
                        self.connection.get(mf4_path_linux + '/' + mf4_log, local=local_mf4_path, preserve_mode=False)
        except Exception as e:
            print(f"copy failed. Reason = {e}")

    def get_test_case_id(self, report_path):
        """
        

        Args:
          report_path: 

        Returns:

        """
        tc_id = ''
        file = codecs.open(report_path + self.def_evald_report_name, "r", "utf-8")
        id_mask = re.compile(r'tc_id==\d+;\s')
        for line in file.readlines():
            if len(id_mask.findall(line)) != 0:
                for match in id_mask.findall(line):
                    tc_id = re.sub(r'tc_id==', '', match)
                    tc_id = re.sub(r';', '', tc_id)
                    return tc_id.strip()

    def get_test_name_if_obsolete(self, report_path):
        """
        

        Args:
          report_path: 

        Returns:

        """
        name = ''
        file = codecs.open(report_path + self.def_evald_report_name, "r", "utf-8")
        name_mask = re.compile(r'-v,\s\w+?.*?.yml')
        for line in file.readlines():
            if len(name_mask.findall(line)) != 0:
                for match in name_mask.findall(line):
                    name = re.sub(r'tc_name==', '', match)
                    name = re.sub(r';', '', name)
                almost_there = name.split('/TC_')[-1]
                # return name.strip()
                tc_name = "TC_" + almost_there.split('.yml')[0]
                return tc_name
            else:
                date_now = str(dt.date.today()) + "_"
                time_now = str(time.strftime("%Hh%Mm%Ss"))
                tc_name = 'obsolete_' + date_now + time_now
                return tc_name

    def get_test_name(self, report_path):
        """
        

        Args:
          report_path: 

        Returns:

        """
        name = ''
        file = codecs.open(report_path + self.def_evald_report_name, "r", "utf-8")
        name_mask = re.compile(r'tc_name==\w+?.*?;\s')
        for line in file.readlines():
            if len(name_mask.findall(line)) != 0:
                for match in name_mask.findall(line):
                    name = re.sub(r'tc_name==', '', match)
                    name = re.sub(r';', '', name)
                    return name.strip()

    def get_report_status(self, report_path):
        """
        

        Args:
          report_path: 

        Returns:

        """
        fail_flag = 0
        pass_flag = 0
        with open(report_path + self.def_evald_report_name, 'r') as html_file:
            soup = BeautifulSoup(html_file, 'html.parser')
            col_results = soup.find_all(name='td', class_='col-result')
            if len(col_results) == 0:
                return "Skipped"
            for result_tag in col_results:
                col_val = result_tag.getText()
                print(f'result tag: {result_tag}')
                print(f'result val: {col_val}')
                if "Skipped" == col_val or "Error" == col_val:
                    print(f'fin result: {col_val}')
                    return "Skipped"
                elif "Failed" == col_val:
                    fail_flag += 1
                elif "Passed" == col_val:
                    pass_flag += 1
            if fail_flag > 0:
                return "Failed"
            elif fail_flag == 0 and pass_flag > 0:
                return "Passed"
            elif fail_flag == 0 and pass_flag == 0:
                return "Skipped"

    # def get_report_status(self, report_path):
    #     file = codecs.open(report_path + self.def_evald_report_name, "r", "utf-8")
    #     tag_exists = re.compile(r'"col-result">')
    #     iserror = re.compile(r'"col-result">Error</td>')
    #     isfailed = re.compile(r'"col-result">Failed</td>')
    #     isskipped = re.compile(r'"col-result">Skipped</td>')
    #     err_flag = 0
    #     fail_flag = 0
    #     skip_flag = 0
    #     for line in file.readlines():
    #         if len(isfailed.findall(line)) != 0:
    #             fail_flag += 1
    #         if len(isskipped.findall(line)) != 0:
    #             skip_flag += 1
    #         if len(iserror.findall(line)) != 0:
    #             err_flag += 1
    #     if err_flag > 0:
    #         return "Skipped"
    #     if fail_flag > 0 and skip_flag == 0:
    #         return "Failed"
    #     elif fail_flag >= 0 and skip_flag > 0:
    #         return "Skipped"
    #     return "Passed"

    @staticmethod
    def create_test_run_folder(results_folder_path, test_run_folder_name):
        """
        

        Args:
          results_folder_path: 
          test_run_folder_name: 

        Returns:

        """
        if not os.path.exists(results_folder_path):  # create the folder 'Reports_from_Test_Execution_Tool' if it does not already exist
            try:
                os.system("mkdir " + results_folder_path)
            except Exception as e:
                raise e
        if not os.path.exists(results_folder_path + test_run_folder_name):  # create the folder 'exp_tc_run' in 'Reports_from_Test_Execution_Tool' if it does not already exist
            try:
                os.system("mkdir " + results_folder_path + test_run_folder_name)
            except Exception as e:
                try:
                    os.system("rm -r " + results_folder_path + test_run_folder_name)
                except:
                    pass
                os.system("mkdir " + results_folder_path + test_run_folder_name)
        try:
            # create passed, failed and skipped folders
            os.system("mkdir " + results_folder_path + test_run_folder_name + "\\" + "passed")
            os.system("mkdir " + results_folder_path + test_run_folder_name + "\\" + "failed")
            os.system("mkdir " + results_folder_path + test_run_folder_name + "\\" + "skipped")
        except:
            print("Some error?")

    def copy_evald_folder_from_linux2local(self, local_results_folder_path, linux_evald_rf_path):
        """
        

        Args:
          local_results_folder_path: 
          linux_evald_rf_path: 

        Returns:

        """
        try:
            with self.connection.cd(linux_evald_rf_path):  # enter the linux evald_report folder
                evald_report_folder_list = self.sftp_obj.list_directory(linux_evald_rf_path)  # list the items of the report folder
                for dir_item in evald_report_folder_list:
                    if self.sftp_obj.check_is_file(linux_evald_rf_path + dir_item):  # if its not a folder, copy it into the local "D:\\evalD_reports_tool_created\\evald_report\\"
                        self.connection.get(linux_evald_rf_path + dir_item, local=local_results_folder_path, preserve_mode=False)
                    elif self.sftp_obj.check_is_dir(linux_evald_rf_path + dir_item):  # if its a folder, copy it into the local "D:\\evalD_reports_tool_created\\evald_report\\plot\\"
                        plot_items_list = self.sftp_obj.list_directory(linux_evald_rf_path + dir_item)
                        for plot_item in plot_items_list:
                            self.connection.get(linux_evald_rf_path + dir_item + '/' + plot_item,
                                                local=local_results_folder_path + dir_item + '\\', preserve_mode=False)
        except Exception as e:
            print(f"copy failed. Reason = {e}")

    @staticmethod
    def copy_report_to_test_run_folder(test_run_folder_name, tool_results_folder_path, test_name, test_status):
        """
        

        Args:
          test_run_folder_name: 
          tool_results_folder_path: 
          test_name: 
          test_status: 

        Returns:

        """
        if test_status == 'Passed':
            move_command = "move" + " " + tool_results_folder_path + test_name + " " + tool_results_folder_path + test_run_folder_name + "\\passed"
            os.system(move_command)
        if test_status == 'Failed':
            move_command = "move" + " " + tool_results_folder_path + test_name + " " + tool_results_folder_path + test_run_folder_name + "\\failed"
            os.system(move_command)
        if test_status == 'Skipped':
            move_command = "move" + " " + tool_results_folder_path + test_name + " " + tool_results_folder_path + test_run_folder_name + "\\skipped"
            os.system(move_command)

    @staticmethod
    def rename_report_folder(local_evald_rf_path, tool_results_folder_path, rename_to_what):
        """
        

        Args:
          local_evald_rf_path: 
          tool_results_folder_path: 
          rename_to_what: 

        Returns:

        """
        if os.path.exists(tool_results_folder_path + rename_to_what):
            date_now = str(dt.date.today()) + "_"
            time_now = str(time.strftime("%Hh%Mm%Ss"))
            os.rename(tool_results_folder_path + rename_to_what, tool_results_folder_path + rename_to_what + "_old_" + date_now + time_now)
            os.rename(local_evald_rf_path, tool_results_folder_path + rename_to_what)
        else:
            os.rename(local_evald_rf_path, tool_results_folder_path + rename_to_what)

    def get_current_test_status(self):
        """ """
        return self.current_tc_status

    def handle_report(self, test_run_folder_name, yaml_path):
        """
        

        Args:
          test_run_folder_name: 
          yaml_path: 

        Returns:

        """
        # copy from linux folder to localfolder
        # print(f"(HANDLE REPORT FUNC)local_evald_rf_path = {self.local_evald_rf_path}\nlinux_evald_rf_path = {self.linux_evald_rf_path}")
        test_run_folder_path = self.tool_results_folder_path + test_run_folder_name + "\\" + self.excel_summary
        self.copy_evald_folder_from_linux2local(self.local_evald_rf_path, self.linux_evald_rf_path)
        # get the report status
        test_status = self.get_report_status(self.local_evald_rf_path)
        # print(f"report status = {test_status}")
        self.current_tc_status = test_status
        # get the test name and ID and rename the report folder with the test name
        test_case_name = self.get_test_name(self.local_evald_rf_path)
        # print(f"report name = {test_case_name}")
        test_case_id = self.get_test_case_id(self.local_evald_rf_path)
        # print(f"test case id = {test_case_id}")
        # rename and move the report to the test_run_folder/(failed, passed, skipped)

        if test_case_name is not None:
            # print(f"DEBUG_____Test is NOT obsolete!!!!")
            test_case_id = "TC_" + test_case_id
            test_case_name = test_case_id + "_" + test_case_name
            self.rename_report_folder(self.local_evald_rf_path, self.tool_results_folder_path, test_case_name)
            # time.sleep(1.5)
            self.copy_report_to_test_run_folder(test_run_folder_name, self.tool_results_folder_path, test_case_name, test_status)
            # time.sleep(1.5)
            local_path_for_mf4log = self.tool_results_folder_path + test_run_folder_name + "\\"
            local_path_for_mf4log += test_status + "\\" + test_case_name + "\\"
            self.mf4_full_path = local_path_for_mf4log + test_case_id + ".mf4"
            # print(f"DEBUG_____path that mf4 log needs to be copied to = {local_path_for_mf4log}")
            self.copy_mf4_log(self.mf4_path_linux, local_path_for_mf4log, test_case_id)
            self.compress_mf4_log(local_path_for_mf4log, test_case_id)
            self.exc_h.get_all_tc_attributes(local_path_for_mf4log, yaml_path, test_run_folder_path)
            self.mf4_full_path = ''
        elif test_case_name is None:
            print(f"TEST IS OBSOLETE...")
            # here we need to fill the sub tc attributes
            self.exc_h.get_all_tc_attributes(self.local_evald_rf_path, yaml_path, test_run_folder_path)
            # trim the '.yml' part from the tc name
            temp_name = re.sub(r'.yml$', '', self.exc_h.tc_attributes_dict["tc_name"][-1])
            obsolete_name = temp_name + "_OBSOLETE"
            print(f"OBSOLETE NAME IS: {obsolete_name}")
            self.rename_report_folder(self.local_evald_rf_path, self.tool_results_folder_path, obsolete_name)
            test_status = "Skipped"
            self.copy_report_to_test_run_folder(test_run_folder_name, self.tool_results_folder_path, obsolete_name, test_status)


class ExcelTestResultSummary:
    """ """

    tc_attributes_dict = {"tc_id": [], "tc_name": [], "tc_folder": [], "sub_tc_attributes": []}
    sub_tc_names_and_status = {}
    helper_list = []

    def __init__(self):
        self.tc_name_mask = re.compile(r'TC_\w*.*yml')
        self.tc_id_mask = re.compile(r'TC_\d*')
        self.tc_folder_mask = re.compile(r'\w*/\w*/\w*/TC')
        self.tc_sub_id_mask = re.compile(r'>\d*<')
        self.ispassed = re.compile(r'"col-result">Passed</td>')
        self.isfailed = re.compile(r'"col-result">Failed</td>')
        self.isskipped = re.compile(r'"col-result">Skipped</td>')
        self.def_evald_report_name = "index.html"
        self.tool_results_folder_path = "D:\\GUI\\"

    def extract_sub_tc_id(self, line):
        """
        

        Args:
          line: 

        Returns:

        """
        # print(f"______extract_sub_tc_id____________line {line}")
        if len(self.tc_sub_id_mask.findall(line)) != 0:
            for match in self.tc_sub_id_mask.findall(line):
                match = re.sub(r'>', '', match)
                match = re.sub(r'<', '', match)
                sub_tc_id = match
                # print(f"sub tc id is: {sub_tc_id}")
                return sub_tc_id
        elif len(self.tc_sub_id_mask.findall(line)) == 0:
            return ""

    def get_tc_sub_attributes(self, report_path):
        """
        

        Args:
          report_path: 

        Returns:

        """
        func_sub_tc_names_and_status = {}
        try:
            with codecs.open(report_path + self.def_evald_report_name, "r", "utf-8") as file:
                fail_flag = 0
                skip_flag = 0
                for index, line in enumerate(file):
                    if len(self.ispassed.findall(line)) != 0:
                        temp_sub_tc_id = linecache.getline(report_path + self.def_evald_report_name, index)
                        sub_tc_id = self.extract_sub_tc_id(temp_sub_tc_id)
                        func_sub_tc_names_and_status[sub_tc_id] = "Passed"
                    elif len(self.isfailed.findall(line)) != 0:
                        fail_flag += 1
                        temp_sub_tc_id = linecache.getline(report_path + self.def_evald_report_name, index)
                        sub_tc_id = self.extract_sub_tc_id(temp_sub_tc_id)
                        func_sub_tc_names_and_status[sub_tc_id] = "Failed"
                    elif len(self.isskipped.findall(line)) != 0:
                        skip_flag += 1
                        temp_sub_tc_id = linecache.getline(report_path + self.def_evald_report_name, index)
                        if re.search("NA", temp_sub_tc_id):
                            # print("Obsolete test step/case, updating subatribute key as 'obsolete'")
                            sub_tc_id = "NA"
                            func_sub_tc_names_and_status[sub_tc_id] = "Obsolete"
                        elif not re.search("NA", temp_sub_tc_id):
                            sub_tc_id = self.extract_sub_tc_id(temp_sub_tc_id)
                            func_sub_tc_names_and_status[sub_tc_id] = "Skipped"
                    else:
                        pass
                self.helper_list.append(func_sub_tc_names_and_status)
                self.tc_attributes_dict["sub_tc_attributes"] = self.helper_list
        except FileNotFoundError as err:
            print(f"Err: {err}")

    def get_tc_name_from_yaml_path(self, yaml_path):
        """
        

        Args:
          yaml_path: 

        Returns:

        """
        if len(self.tc_name_mask.findall(yaml_path)) != 0:
            for name_match in self.tc_name_mask.findall(yaml_path):
                tc_name = name_match
                self.tc_attributes_dict["tc_name"].append(tc_name)
                # print(f"tc_id match is : {tc_name}")
                if len(self.tc_id_mask.findall(tc_name)) != 0:
                    for id_match in self.tc_id_mask.findall(tc_name):
                        tc_id = id_match
                        self.tc_attributes_dict["tc_id"].append(tc_id)
                # print(f"name_and_folder after adding tc_id : {self.tc_attributes_dict}")
                if len(self.tc_folder_mask.findall(yaml_path)) != 0:
                    for folder_match in self.tc_folder_mask.findall(yaml_path):
                        folder_id = re.sub(r'/TC', '', folder_match)
                        self.tc_attributes_dict["tc_folder"].append(folder_id)
                        # print(f"folder_id match is: {folder_id}")
                        # print(f"name_and_folder after adding tc_folder_id : {self.tc_attributes_dict}")
                        out = self.tc_attributes_dict
                        return out
        elif len(self.tc_name_mask.findall(yaml_path)) == 0:
            # print("match NOT found")
            out = self.tc_attributes_dict
            return out

    def get_all_tc_attributes(self, report_path, yaml_path, test_run_folder_path):
        """
        

        Args:
          report_path: 
          yaml_path: 
          test_run_folder_path: 

        Returns:

        """
        self.get_tc_name_from_yaml_path(yaml_path)
        self.get_tc_sub_attributes(report_path)
        # print(f"ALL TEST ATTRIBUTES = {self.tc_attributes_dict}")
        self.create_excel_file_pandas(self.tc_attributes_dict, test_run_folder_path)

    @staticmethod
    def fill_cells_with_color(excelpath):
        """
        

        Args:
          excelpath: 

        Returns:

        """
        fill_cell_fail = PatternFill(patternType='solid', fgColor='FF0000')
        fill_cell_pass = PatternFill(patternType='solid', fgColor='92D050')
        fill_cell_skip = PatternFill(patternType='solid', fgColor='FFC000')
        try:
            wb = openpyxl.load_workbook(excelpath)
            # sh = wb.active
            sh = wb['Sheet1']
            for i in range(1, sh.max_row + 1):
                # print(f"row {i} ")
                for j in range(4, 5):
                    cellname = "D" + str(i)
                    cell_obj = sh.cell(row=i, column=j)
                    # print(f"data == {cell_obj.value} and type is {type(cell_obj.value)} ")
                    if 'Failed' in cell_obj.value:
                        sh[cellname].fill = fill_cell_fail
                    elif 'Passed' in cell_obj.value:
                        sh[cellname].fill = fill_cell_pass
                    elif 'Skipped' in cell_obj.value:
                        sh[cellname].fill = fill_cell_skip
                    elif 'Obsolete' in cell_obj.value:
                        sh[cellname].fill = fill_cell_skip
            wb.save(excelpath)
        except PermissionError as err:
            print(f"PLEASE CLOSE THE SUMMARY FILE IN ORDER FOR THE SUMMARY TO BE COMPLETED!!! {err}")

    def create_excel_file_pandas(self, inputdict, testrunfolderpath):
        """
        

        Args:
          inputdict: 
          testrunfolderpath: 

        Returns:

        """
        # define columns for the excel file
        columns = {"Req Folder": [],
                   "ID RQM": [],
                   "TC name": [],
                   "Results SWv": [],
                   "Defect PSA": [],
                   "Remark": [],
                   "SIL status": []}
        # create a dataframe with the columns
        subattributes_dict_list = inputdict["sub_tc_attributes"]
        # print(f"subatributes dict = {subattributes}")
        index = 0
        for subattribute in subattributes_dict_list:
            for key, val in subattribute.items():
                # print(f"key = {key}, val = {val}")
                columns["Req Folder"].append(inputdict["tc_folder"][index])
                columns["ID RQM"].append(key)
                columns["TC name"].append(inputdict["tc_name"][index])
                columns["Results SWv"].append(val)
                columns["Defect PSA"].append("")
                columns["Remark"].append("")
                columns["SIL status"].append("")
            index += 1
        df = pd.DataFrame.from_dict(columns)
        # print(f"DataFrame?:\n {df}")
        try:
            writer = pd.ExcelWriter(testrunfolderpath, engine='openpyxl', mode='w')
            df.to_excel(writer, sheet_name='Sheet1', index=False)
            writer.save()
        except PermissionError as err:
            print("PLEASE CLOSE THE SUMMARY FILE IN ORDER FOR THE SUMMARY TO BE COMPLETED!!!")
        self.fill_cells_with_color(testrunfolderpath)
