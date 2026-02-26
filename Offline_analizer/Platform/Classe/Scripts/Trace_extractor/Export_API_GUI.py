# -*- coding: utf-8 -*-
# @file Export_API_GUI.py
# @author ADAS_HIL_TEAM
# @date 08-21-2023

##################################################################
# C O P Y R I G H T S
# ----------------------------------------------------------------
# Copyright (c) 2023 by Robert Bosch GmbH. All rights reserved.

# The reproduction, distribution and utilization of this file as
# well as the communication of its contents to others without express
# authorization is prohibited. Offenders will be held liable for the
# payment of damages. All rights reserved in the event of the grant
# of a patent, utility model or design.

##################################################################


import os
from openpyxl import load_workbook
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from INI_parser import delete_if_exist, main_ini
from export_api_main import change_signal_ini_to_xlsx, update_ini_with_database


class ExportAPI():
    """ """
    def __init__(self):
        # determine the main window
        self.window = tk.Tk()
        self.window.title("Export API")
        self.window.geometry("1000x600")

        self.num_input = 4
        self.var_input = [""] * self.num_input
        self.__var_check = [tk.IntVar(), tk.IntVar(), tk.IntVar(), tk.IntVar()]
        self.__label_list = [["CSV File", "INI File", "Input Database", "Output Database"], [None] * self.num_input]
        self.__text_list = [None] * self.num_input
        self.__check_list = [None] * self.num_input
        self.__cbb_column = None


    def main_section(self):
        """ """
        for i in range(self.num_input):
            self.create_section(i)

        label_column = ttk.Label(self.window, text="Select/Enter the column to be updated:", anchor=tk.CENTER)
        label_column.grid(column=2, row=3, columnspan=2, pady=20)
        self.__cbb_column = ttk.Combobox(self.window, width=15)
        self.__cbb_column.grid(column=2, row=4, columnspan=2)

        button_create = ttk.Button(self.window, text="Create INI", command=self.__clicked_create, width=20)
        button_update = ttk.Button(self.window, text="Update Database", command=self.__clicked_update, width=20)
        button_create.grid(column=0, row=5, columnspan=2, padx=30, pady=50, ipady=5)
        button_update.grid(column=2, row=5, columnspan=2, padx=30, pady=50, ipady=5)

    def create_section(self, i):
        """
        

        Args:
          i: 

        Returns:

        """
        self.__label_list[1][i] = ttk.Label(self.window, text=self.__label_list[0][i], font=("Arial Bold", 10), anchor=tk.CENTER, width=25)    
        self.__label_list[1][i].grid(column=i, row=0, padx=30, pady=50)

        self.__text_list[i] = tk.Text(self.window, height=8, width=20)
        self.__text_list[i].grid(column=i, row=1, padx=30)

        check_text = "Load" if i < self.num_input-1 else "Copy from Input"
        self.__check_list[i] = ttk.Checkbutton(self.window, text=check_text, onvalue=(i+1), offvalue=(-i-1), variable=self.__var_check[i], command=lambda: self.__checked(self.__var_check[i].get()))
        self.__check_list[i].grid(column=i, row=2, padx=30, pady=10)

    def __checked(self, index):
        if index == 1:
            file_name = filedialog.askopenfilename(title="Please select a CSV file.", filetypes = (("CSV Files", "*.csv"), ("All Files", "*.*")))
            if not file_name:
                self.__var_check[index-1].set(-index)
            else:
                self.var_input[index-1] = file_name
                self.__text_list[index-1].delete('1.0', tk.END)
                self.__text_list[index-1].insert(tk.END, file_name)              
        elif index == 2:
            file_name = filedialog.askopenfilename(title="Please select an INI file.", filetypes = (("INI Files", "*.ini"), ("All Files", "*.*")))
            if not file_name:
                self.__var_check[index-1].set(-index)
            else:
                self.var_input[index-1] = file_name
                self.__text_list[index-1].delete('1.0', tk.END)
                self.__text_list[index-1].insert(tk.END, file_name)
                self.__var_check[index].set(-index-1)
                self.var_input[index] = ""
                self.__text_list[index].delete('1.0', tk.END)
                self.__text_list[index].insert(tk.END, "")   
        elif index == 3:
            file_name = filedialog.askopenfilename(title="Please select an Excel file.", filetypes = (("Excel Files", ["*.xls", "*.xlsx"]), ("All Files", "*.*")))
            if not file_name:
                self.__var_check[index-1].set(-index)
            else:
                self.var_input[index-1] = file_name
                self.__text_list[index-1].delete('1.0', tk.END)
                self.__text_list[index-1].insert(tk.END, file_name)
                self.__update_combobox() 
        elif index == 4:
            self.var_input[index-1] = self.var_input[index-2]
            self.__text_list[index-1].delete('1.0', tk.END)
            self.__text_list[index-1].insert(tk.END, self.var_input[index-2])
        elif index < 0:
            self.var_input[-index-1] = ""
            self.__text_list[-index-1].delete('1.0', tk.END)
    
    def __update_combobox(self):
        wb = load_workbook(self.var_input[2])
        ws = wb['SysVarDatabase']
        col_names = [cell.value for cell in ws[1]]
        col_names.insert(0, "<New column>")
        self.__cbb_column['values'] = col_names

    def __clicked_create(self):
        if not self.var_input[0]:
            messagebox.showerror(title="Error", message="The CSV file is not found.\nPlease select a CSV file.")
            return False
        if not self.var_input[1]:
            self.var_input[1] = self.var_input[0].replace(".csv", ".ini")
        else:
            selection = messagebox.askyesno(title="Warning", message="The INI file already exists.\nDo you want to replace the existing file?")
            if not selection:
                return selection
        #print(self.var_input[1])
        self.__var_check[1].set(2)
        self.__text_list[1].delete('1.0', tk.END)
        self.__text_list[1].insert(tk.END, self.var_input[1])
        try:
            delete_if_exist(self.var_input[1]) 
            success = main_ini(self.var_input[0], self.var_input[1].replace(".ini", ""))
            if not success:
                raise ValueError
            messagebox.showinfo(title="Info", message="Successfully created the INI file.")
        except Exception as exp:
            print("\nFailed to create the INI file:", exp)
            messagebox.showerror(title="Error", message="Failed to create the INI file.\nPlease check the Console output.")

    def __clicked_update(self):
        self.var_input[3] = self.__text_list[3].get('1.0', tk.END).strip()
        #print(self.var_input[3])
        if not self.var_input[1]:
            messagebox.showerror(title="Error", message="The INI file is not found.\nPlease select or generate the INI file.")
            return False
        if not self.var_input[2]:
            messagebox.showerror(title="Error", message="Please select your input database.")
            return False
        if not self.__cbb_column.get():
            messagebox.showerror(title="Error", message="Please specify the column to be updated.")
            return False
        if not self.var_input[3]:
            messagebox.showerror(title="Error", message="Please specify your output database.")
            return False
        if os.path.exists(self.var_input[3]):
            selection = messagebox.askyesno(title="Warning", message="The output file: " + self.var_input[3] + " already exists. \nDo you want to replace the existing file?")
            if not selection:
                return selection
        #print(self.__cbb_column.get())
        try:
            change_signal_ini_to_xlsx(self.var_input[1])
            update_ini_with_database(self.var_input[2], self.var_input[3], updated_column=self.__cbb_column.get(), CCF=False)
            messagebox.showinfo(title="Info", message="Successfully updated the output database.")
        except Exception as exp:
            print("\nFailed to update the output database:", exp)
            messagebox.showerror(title="Error", message="Failed to update the output database.\nPlease check the Console output.")
        

    def __call__(self):
        self.main_section()
        self.window.mainloop()


if __name__ == "__main__":
    app = ExportAPI()
    app()