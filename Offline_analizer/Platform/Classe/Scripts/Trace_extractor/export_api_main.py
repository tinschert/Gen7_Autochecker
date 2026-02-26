# -*- coding: utf-8 -*-
# @file export_api_main.py
# @author ADAS_HIL_TEAM
# @date 08-21-2023

##################################################################
# C O P Y R I G H T S
# ----------------------------------------------------------------
# Copyright (c) 2023 by Robert Bosch GmbH. All rights reserved.

# The reproduction, distribution and utilization of this file as
# well as the communication of its contents to others without express
# authorization is prohibited. Offenders will be held liable for the
# payment of damages. All rights reserved in the event of the grant
# of a patent, utility model or design.

##################################################################


import pandas as pd
try:
    pd.set_option('future.no_silent_downcasting', True)
except:
    pass
import openpyxl
import xlwings
from INI_parser import execute_ini
from Ethernet_extractor import execute_ethernet
from CAN_extractor import execute_can


# signal formulas
#CcfDD02MainECUSupVolts_Rv = 12.250000
#CcfDD04InCarTemp_Rv = 24
#CcfDD05OutdoorTemp_Rv = 11
#CcfModelYear_Rv = 2020.000000
#CcfOilLevelTopUpValue_Rv = 1.000000 (expected 10)
##CcfMaxVehicleSpeed_Rv = 190, x = x-108

def func1(x):
    """
    

    Args:
      x: 

    Returns:

    """
    return int(x/0.25)
def func2(x):
    """
    

    Args:
      x: 

    Returns:

    """
    return int(x+40)
def func3(x):
    """
    

    Args:
      x: 

    Returns:

    """
    return int((x-2009.75)/0.25)
def func4(x):
    """
    

    Args:
      x: 

    Returns:

    """
    return int(x/0.1)

formula_dict = {355: func1, 356: func2, 357: func2, 423: func3, 427: func4}


def write_text(file_name, data1):          # generate files with the signals selection and their values    
    """
    

    Args:
      file_name: 
      data1: 

    Returns:

    """
    filename = file_name                                        # add the name of the file to generate with the extension
    file = open (filename, "a")                                 # 4 parameters are to be added, in case 4 parameters are too much
    file.write(data1)                                           # keep the data slot empty                                          # the file will be opened and after each data written,  
    file.write("\n")                                            # a jump line is added qt the end of each new element in the file
    file.close()


def read_text(file_in):                                 # load the file to read, in this case, the csv file
    """
    

    Args:
      file_in: 

    Returns:

    """
    file_temp = open(file_in, "r")                      # the file should be plaed in this same directory as the python file
    data = file_temp.read()
    return data


def serach_data(data1, data2):                  # return the index position of data2 in data1
    """
    

    Args:
      data1: 
      data2: 

    Returns:

    """
    return data1.find(data2)                    # if data2 is not to be foundm the result is -1


def generate_sig_values(str_in):                          # extracts the signal values from the header including the CAN + PDU + signal name
    """
    

    Args:
      str_in: 

    Returns:

    """
    #serach_data(str_in, "=")                              # search for the "=" in the file name and crop the last part
    try:
        str_in = str_in[serach_data(str_in, "=")+2 :]
    except:
        print("error in extracting signal name")    
    return (str_in) 


def generate_sig_wo_ch(str_in):                           # extracts the signal name from the header including the CAN + PDU + signal name
    """
    

    Args:
      str_in: 

    Returns:

    """
    #serach_data(str_in, "[")                              # search for the "[" in the file name and crop the first part
    try:
        str_in = str_in[0 :(serach_data(str_in, "["))]
    except:
        print("error in extracting signal name")    
    return (str_in + '_Rv') 


def generate_sig_wo_equal(str_in):                        # extracts the signal name from the header including the CAN + PDU + signal name
    """
    

    Args:
      str_in: 

    Returns:

    """
    #serach_data(str_in, "=")                              # search for the "=" in the file name and crop the first part
    try:
        str_in = str_in[0 :serach_data(str_in, "=")-1]
    except:
        print("error in extracting signal name")    
    return (str_in) 


lines = []
values = []

def change_signal_ini_to_xlsx(Input_ini):
    """
    

    Args:
      Input_ini: 

    Returns:

    """

    with open(Input_ini) as f:
        lines = f.readlines()

    count = 0
    for line in lines:
        count += 1
    
        values = generate_sig_values(line)

        if(serach_data(line, '[') > -1):
            line = generate_sig_wo_ch(line)
        else:
            line= generate_sig_wo_equal(line)
        
        write_text('update.ini', line +'='+ values)
    
    #df = pd.read_csv(Input_ini, delimiter='=', sep=',\s+', skiprows=11) ,error_bad_lines=False , delimiter='=',engine='python'
    df = pd.read_csv('update.ini', delimiter='=') 
    df.to_excel('ini.xlsx', 'ini')
    
    wb_ini = openpyxl.load_workbook('ini.xlsx')
    
    ini = wb_ini['ini']
     
    wb_ini.save('Ini.xlsx')
    

def update_ini_with_database(Input_database, output_database, updated_column="import_variant", CCF=False):
    """
    

    Args:
      Input_database: 
      output_database: 
      updated_column:  (Default value = "import_variant")
      CCF:  (Default value = False)

    Returns:

    """
    
    wb_ini = openpyxl.load_workbook('Ini.xlsx')
    wb_variant_oldini = openpyxl.load_workbook(Input_database)          #'variant_Ini.xlsx'
    
    ini = wb_ini['ini']
    variant_oldini = wb_variant_oldini['SysVarDatabase']
    #variant_oldini = wb_variant_oldini['DatabaseNI7']
    
    row_count_ini = ini.max_row
    row_count_variant_oldini = variant_oldini.max_row
    #row_count_variant_oldeth = variant_oldeth.max_row
    
    col_names_variant_oldini = [cell.value for cell in variant_oldini[1]]

    try:
        col_updated_variant_oldini = col_names_variant_oldini.index(updated_column)+1
        #col_veh_id_variant_oldeth = row_top_variant_oldeth.index(vehicle_id)+1
    except:
        print("The column is not found in the database, so a new column will be created at the end.")
        print(updated_column, col_names_variant_oldini)
        col_updated_variant_oldini = len(col_names_variant_oldini)+1
        variant_oldini.cell(row = 1, column = col_updated_variant_oldini).value = updated_column
    
    for i in range(2, row_count_variant_oldini+1):
        for j in range(1, row_count_ini+1):
            if (variant_oldini.cell(row = i, column = 1).value == ini.cell(row = j, column = 2).value) or \
            (variant_oldini.cell(row = i, column = 1).value+"_Rv" == ini.cell(row = j, column = 2).value):
                variant_oldini.cell(row = i, column = col_updated_variant_oldini).value = ini.cell(row = j, column = 3).value
                break
    if CCF:
        for k in formula_dict.keys():
            tem = variant_oldini.cell(row = k, column = col_updated_variant_oldini).value
            variant_oldini.cell(row = k, column = col_updated_variant_oldini).value = formula_dict[k](float(tem))
  
    wb_variant_oldini.save(output_database)
    # use xlwings to save Excel to maintain formulas
    app = xlwings.App(visible=False, add_book=False)
    wb_variant_newini = app.books.open(output_database)
    wb_variant_newini.save(output_database)
    wb_variant_newini.close()
    app.quit()


def execute_update_main():
    """ """
    print('Please insert the input database file name with the format "file_name.xlsx", like "Variant_Ini.xlsx"')
    Input_database = input(" file = ")
    Input_database = Input_database.replace("'", "").replace('"', '')
    print('Please name the output database file with the format "file_name.xlsx", like "output.xlsx"')
    Output_database = input('output file name = ')
    Output_database = Output_database.replace("'", "").replace('"', '')
    print('Please insert the input signal.ini file name with the format "file_name.ini", like"signal.ini"')
    Input_ini = input(" file = ")
    Input_ini = Input_ini.replace("'", "").replace('"', '')
    """
    print('Please insert the vehicle id with the format "X*** or L***", like"X590"')
    Input_vehicle = input(" vehicle = ")
    Input_vehicle = Input_vehicle.replace("'", "").replace('"', '')
    """
    change_signal_ini_to_xlsx(Input_ini)
    update_ini_with_database(Input_database, Output_database)


def execute_main_api():
    """ """
    for _ in range(100):
        print("###################################################")
        print("          ~~~ARXML parser started~~~"            )
        print("###################################################")
        print("")
        
        print("What action do you want to execute?")
        print("a) Create init file signal.ini")
        print("b) Create an Ethernet extraction")
        print("c) Create a CAN/Flexray extraction")
        print("d) Update the database with ini file")
        print("q) Quit this api extractor")
        print("")
        
        try:
            choice = input(" selection = ")
            if choice == "a":
                print("------Create ini node------")
                Ini_file = execute_ini()
                print(Ini_file, "Execution successful \n")            
            elif choice == "b":
                print("------Export Etehrnet data------")
                execute_ethernet()
                print("Execution successful \n")   
            elif choice == "c":
                print("------Export CAN/Flexray data------")
                execute_can()
                print("Execution successful \n")
            elif choice == "d":
                print("------Update database with signal.ini------")
                execute_update_main()
                print("Execution successful \n")
            elif choice == "q":
                print("\n Quitted from the application")
                # quit()
                break   
            else:    
                print("\n Wrong input selection, please re-enter your choice") 
            
            if choice == "a":
                print('Do you want to update database file with signal.ini? Please answer with [y or n]')
                choice_1 = input(" selection = ")
                if choice_1 == "n":
                    print("\n Thank you for using ini exporter App \n")
                    break
                elif choice_1 == "y":
                    print('Please insert the input database file name with the format "file_name.xlsx", like "Variant_Ini.xlsx"')
                    Input_database = input(" file = ")
                    Input_database = Input_database.replace("'", "").replace('"', '')
                    print('Please name the output database file with the format "file_name.xlsx", like "output.xlsx"')
                    Output_database = input('output file name = ')
                    """
                    Output_database = Output_database.replace("'", "").replace('"', '')
                    print('Please insert the vehicle id with the format "X*** or L***", like"X590"')
                    Input_vehicle = input(" vehicle = ")
                    Input_vehicle = Input_vehicle.replace("'", "").replace('"', '')
                    """
                    change_signal_ini_to_xlsx(Ini_file)
                    update_ini_with_database(Input_database, Output_database)
                    print("Execution successful")
                    print("Thank you for using ini exporter App \n")
                else:    
                    print("\n Wrong input selection, please re-enter your choice") 
            else:
                print("\n Thank you for using ini exporter App \n")
                # quit()
                break   
        except Exception as exp:
            print("\n Input error occurs: ", exp)
            print("Please check your selection or file names \n")
 

if __name__ == '__main__':
    execute_main_api()